from django.shortcuts import render, get_object_or_404, redirect
from .decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from django.core.cache import cache
from django.template.response import TemplateResponse
from django.core.mail import send_mail
from django.conf import settings
from datetime import timedelta
import json
import io
import csv
from datetime import datetime
from urllib.parse import urlencode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4, A3, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from authentication.models import (
    CustomUser, DonorProfile, FoodBankProfile, RecipientProfile, 
    FoodBankRequest, Notification, SupportMessage, SupportMessageReply, FoodBankSubscription, 
    SubscriptionPayment, AdminLoginLog, Donation, AdminCode, SystemSupportDonation,
    RecipientRequest, DonationAllocation, DonationDiscussion, PaymentTransaction,
    Testimonial, DonorTestimonial, FoodbankTestimonial, UnspecifiedDonationManagement,
    RequestManagement
)
from .utils import (
    get_dashboard_stats, get_user_registration_trends, get_donation_trends,
    get_user_type_distribution, get_donation_type_distribution, get_recent_activity
)
from .config import ITEMS_PER_PAGE, CHART_MONTHS
from .decorators import admin_required, superuser_required
from .forms import (
    CustomUserAdminUpdateForm,
    DonorProfileAdminForm,
    FoodBankProfileAdminForm,
    RecipientProfileAdminForm,
)


def admin_login(request):
    """Custom admin login view"""
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('custom_admin:dashboard')

    context = {
        'entered_email': '',
        'remember_checked': False,
        'field_errors': {},
        'next_url': request.GET.get('next', ''),
    }

    if request.method == 'POST':
        email = (request.POST.get('email') or '').strip()
        password = request.POST.get('password') or ''
        remember = bool(request.POST.get('remember'))
        next_url = request.POST.get('next') or request.GET.get('next', 'custom_admin:dashboard')

        context.update({
            'entered_email': email,
            'remember_checked': remember,
            'next_url': request.POST.get('next', ''),
        })

        field_errors = {}
        if not email:
            field_errors['email'] = 'Please enter your email address.'
        if not password:
            field_errors['password'] = 'Please enter your password.'

        if field_errors:
            context['field_errors'] = field_errors
            messages.error(request, 'Please correct the highlighted fields and try again.')
        else:
            client_ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR', 'unknown')
            attempts_key = f'custom_admin_login_attempts:{client_ip}:{email.lower()}'
            lockout_key = f'{attempts_key}:locked'
            max_attempts = 5
            lockout_seconds = 900  # 15 minutes

            if cache.get(lockout_key):
                field_errors['email'] = 'Too many failed attempts. Please wait 15 minutes before trying again.'
                field_errors['password'] = 'Too many failed attempts. Please wait 15 minutes before trying again.'
                context['field_errors'] = field_errors
                messages.error(request, 'Too many failed login attempts. Please try again later.')
                return render(request, 'custom_admin/login.html', context)

            user = authenticate(request, username=email, password=password)

            if user is not None and user.is_staff:
                login(request, user)
                cache.delete(attempts_key)
                cache.delete(lockout_key)

                # Set session expiry based on remember me
                if not remember:
                    request.session.set_expiry(0)  # Browser close
                else:
                    request.session.set_expiry(1209600)  # 2 weeks

                messages.success(request, f'Welcome back, {user.email}!')
                return redirect(next_url)

            failed_attempts = (cache.get(attempts_key) or 0) + 1
            if failed_attempts >= max_attempts:
                cache.set(lockout_key, True, lockout_seconds)
                cache.delete(attempts_key)
                lockout_message = 'Too many failed login attempts. Please try again in 15 minutes.'
                field_errors['email'] = lockout_message
                field_errors['password'] = lockout_message
                messages.error(request, lockout_message)
            else:
                cache.set(attempts_key, failed_attempts, lockout_seconds)
                remaining_attempts = max_attempts - failed_attempts
                invalid_message = 'Invalid email or password.'
                if user is not None and not user.is_staff:
                    invalid_message = 'You do not have permission to access the admin area.'
                field_errors['email'] = invalid_message
                field_errors['password'] = invalid_message
                messages.error(request, invalid_message)
                messages.warning(request, f'{remaining_attempts} login attempt(s) remaining before temporary lockout.')

            context['field_errors'] = field_errors

    return render(request, 'custom_admin/login.html', context)


def admin_logout(request):
    """Custom admin logout view"""
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('/')


@staff_member_required
def dashboard_home(request):
    """Main dashboard with KPIs and charts"""
    current_hour = timezone.localtime().hour
    if 5 <= current_hour < 12:
        dashboard_salutation = "Good morning"
    elif 12 <= current_hour < 17:
        dashboard_salutation = "Good afternoon"
    else:
        dashboard_salutation = "Good evening"

    # KPIs
    total_users = CustomUser.objects.count()
    total_donors = CustomUser.objects.filter(user_type='DONOR').count()
    total_foodbanks = CustomUser.objects.filter(user_type='FOODBANK').count()
    total_recipients = CustomUser.objects.filter(user_type='RECIPIENT').count()
    total_admins = CustomUser.objects.filter(user_type='ADMIN').count()
    total_donations = Donation.objects.count()
    total_donated_amount = Donation.objects.aggregate(
        total=Sum('amount'))['total'] or 0
    
    # Pending foodbank approvals
    pending_foodbanks = FoodBankProfile.objects.filter(is_approved='pending').count()
    
    # Rejected foodbank applications
    rejected_foodbanks = FoodBankProfile.objects.filter(is_approved='rejected').count()
    
    # Support message statistics
    total_support_messages = SupportMessage.objects.count()
    new_support_messages = SupportMessage.objects.filter(status='new').count()
    urgent_support_messages = SupportMessage.objects.filter(priority='urgent').count()
    
    # Subscription statistics
    total_subscriptions = FoodBankSubscription.objects.count()
    now = timezone.now()
    active_subscriptions = FoodBankSubscription.objects.filter(
        (
            Q(status='trial') &
            (
                Q(trial_end_date__isnull=True) |
                Q(trial_end_date__gt=now)
            )
        ) |
        (
            Q(status='active') &
            (
                Q(subscription_end_date__isnull=True) |
                Q(subscription_end_date__gt=now)
            )
        )
    ).count()
    trial_subscriptions = FoodBankSubscription.objects.filter(status='trial').count()
    expired_subscriptions = FoodBankSubscription.objects.filter(status='expired').count()
    pending_payments = SubscriptionPayment.objects.filter(status='pending').count()
    
    # Revenue calculations
    monthly_revenue = SubscriptionPayment.objects.filter(
        status='approved',
        plan_type='monthly'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    yearly_revenue = SubscriptionPayment.objects.filter(
        status='approved',
        plan_type='yearly'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_subscription_revenue = monthly_revenue + yearly_revenue
    
    # Recent activity
    recent_users = CustomUser.objects.order_by('-date_joined')[:5]
    recent_donations = Donation.objects.select_related('donor', 'foodbank').order_by('-donated_at')[:5]
    recent_requests = FoodBankRequest.objects.select_related('foodbank').order_by('-created_at')[:5]
    
    # User counts by type for pie chart
    user_counts = CustomUser.objects.values('user_type').annotate(count=Count('id'))
    user_type_labels = [uc['user_type'] for uc in user_counts]
    user_type_data = [uc['count'] for uc in user_counts]
    
    # Monthly registrations for line chart
    now = timezone.now()
    months = [(now - timedelta(days=30*i)).strftime('%b %Y') for i in reversed(range(12))]
    registration_counts = []
    for i in reversed(range(12)):
        start = (now - timedelta(days=30*(i+1)))
        end = (now - timedelta(days=30*i))
        count = CustomUser.objects.filter(date_joined__gte=start, date_joined__lt=end).count()
        registration_counts.append(count)
    
    # Donations by category (Food vs Non-Food) for bar chart
    donation_category_counts = Donation.objects.values('donation_category').annotate(count=Count('id'))
    food_count = sum(item['count'] for item in donation_category_counts if item['donation_category'] == 'food')
    non_food_count = sum(item['count'] for item in donation_category_counts if item['donation_category'] != 'food')
    donation_type_labels = ['Food', 'Non-Food']
    donation_type_data = [food_count, non_food_count]
    
    context = {
        'title': 'Dashboard',
        'dashboard_salutation': dashboard_salutation,
        'total_users': total_users,
        'total_donors': total_donors,
        'total_foodbanks': total_foodbanks,
        'total_recipients': total_recipients,
        'total_admins': total_admins,
        'total_donations': total_donations,
        'total_donated_amount': total_donated_amount,
        'pending_foodbanks': pending_foodbanks,
        'rejected_foodbanks': rejected_foodbanks,
        'total_support_messages': total_support_messages,
        'new_support_messages': new_support_messages,
        'urgent_support_messages': urgent_support_messages,
        # Subscription statistics
        'total_subscriptions': total_subscriptions,
        'active_subscriptions': active_subscriptions,
        'trial_subscriptions': trial_subscriptions,
        'expired_subscriptions': expired_subscriptions,
        'pending_payments': pending_payments,
        'total_subscription_revenue': total_subscription_revenue,
        'monthly_revenue': monthly_revenue,
        'yearly_revenue': yearly_revenue,
        # Activity and charts
        'recent_users': recent_users,
        'recent_donations': recent_donations,
        'recent_requests': recent_requests,
        'user_type_labels': json.dumps(user_type_labels),
        'user_type_data': json.dumps(user_type_data),
        'registration_months': json.dumps(months),
        'registration_counts': json.dumps(registration_counts),
        'donation_type_labels': json.dumps(donation_type_labels),
        'donation_type_data': json.dumps(donation_type_data),
    }
    return render(request, 'custom_admin/dashboard.html', context)


@staff_member_required
def user_management(request):
    """User management with separate sections for each user type"""
    # Get search parameter
    search = request.GET.get('search')

    active_tab = request.GET.get('tab', 'donors')
    if active_tab not in {'donors', 'foodbanks', 'recipients', 'admins'}:
        active_tab = 'donors'
    
    # Donors with profiles
    donors = CustomUser.objects.filter(user_type='DONOR').select_related('donor_profile').order_by('-date_joined')
    if search:
        donors = donors.filter(
            Q(email__icontains=search) | 
            Q(phone_number__icontains=search) |
            Q(donor_profile__full_name__icontains=search)
        )
    
    # Food Banks with profiles
    foodbanks = CustomUser.objects.filter(user_type='FOODBANK').select_related('foodbank_profile').order_by('-date_joined')
    if search:
        foodbanks = foodbanks.filter(
            Q(email__icontains=search) | 
            Q(phone_number__icontains=search) |
            Q(foodbank_profile__foodbank_name__icontains=search) |
            Q(foodbank_profile__contact_person__icontains=search)
        )
    
    # Recipients with profiles
    recipients = CustomUser.objects.filter(user_type='RECIPIENT').select_related('recipient_profile').order_by('-date_joined')
    if search:
        recipients = recipients.filter(
            Q(email__icontains=search) | 
            Q(phone_number__icontains=search) |
            Q(recipient_profile__full_name__icontains=search)
        )
    
    # Admins
    admins = CustomUser.objects.filter(user_type='ADMIN').order_by('-date_joined')
    if search:
        admins = admins.filter(
            Q(email__icontains=search) | 
            Q(phone_number__icontains=search)
        )
    
    # Pagination for each type
    donor_paginator = Paginator(donors, 10)
    foodbank_paginator = Paginator(foodbanks, 10)
    recipient_paginator = Paginator(recipients, 10)
    admin_paginator = Paginator(admins, 10)
    
    donor_page = donor_paginator.get_page(request.GET.get('donor_page'))
    foodbank_page = foodbank_paginator.get_page(request.GET.get('foodbank_page'))
    recipient_page = recipient_paginator.get_page(request.GET.get('recipient_page'))
    admin_page = admin_paginator.get_page(request.GET.get('admin_page'))
    
    context = {
        'title': 'User Management',
        'donor_page': donor_page,
        'foodbank_page': foodbank_page,
        'recipient_page': recipient_page,
        'admin_page': admin_page,
        'search': search,
        'active_tab': active_tab,
        'total_donors': donors.count(),
        'total_foodbanks': foodbanks.count(),
        'total_recipients': recipients.count(),
        'total_admins': admins.count(),
    }
    return render(request, 'custom_admin/user_management.html', context)


@staff_member_required
def donors_management(request):
    """Manage donors with table view"""
    search = request.GET.get('search')
    is_active = request.GET.get('is_active')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    sort_by = request.GET.get('sort_by', 'newest')  # newest or oldest
    
    donors = CustomUser.objects.filter(user_type='DONOR').select_related('donor_profile').annotate(
        donations_count=Count('donation', distinct=True)
    )
    
    if search:
        donors = donors.filter(
            Q(email__icontains=search) | 
            Q(phone_number__icontains=search) |
            Q(donor_profile__full_name__icontains=search) |
            Q(donor_profile__organization_name__icontains=search)
        )
    
    if is_active:
        donors = donors.filter(is_active=is_active == 'true')
    
    # Date range filter
    if date_from:
        donors = donors.filter(date_joined__gte=date_from)
    if date_to:
        donors = donors.filter(date_joined__lte=date_to)
    
    # Sorting
    if sort_by == 'oldest':
        donors = donors.order_by('date_joined')
    else:  # newest
        donors = donors.order_by('-date_joined')
    
    # Pagination
    paginator = Paginator(donors, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Donors Management',
        'page_obj': page_obj,
        'search': search,
        'is_active': is_active,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'total_count': donors.count(),
    }
    return render(request, 'custom_admin/donors_management.html', context)


@staff_member_required
def foodbanks_management(request):
    """Manage food banks with table view"""
    search = request.GET.get('search')
    is_active = request.GET.get('is_active')
    approval_status = request.GET.get('approval_status')
    food_type = request.GET.get('food_type')  # food, non_food, both
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    sort_by = request.GET.get('sort_by', 'newest')  # newest or oldest
    
    foodbanks = CustomUser.objects.filter(user_type='FOODBANK').select_related('foodbank_profile').annotate(
        donations_count=Count('foodbank_profile__donation', distinct=True)
    )
    
    if search:
        foodbanks = foodbanks.filter(
            Q(email__icontains=search) | 
            Q(phone_number__icontains=search) |
            Q(foodbank_profile__foodbank_name__icontains=search) |
            Q(foodbank_profile__contact_person__icontains=search) |
            Q(foodbank_profile__address__icontains=search)
        )
    
    if is_active:
        foodbanks = foodbanks.filter(is_active=is_active == 'true')
    
    if approval_status:
        foodbanks = foodbanks.filter(foodbank_profile__is_approved=approval_status)
    
    # Food type filter
    if food_type:
        foodbanks = foodbanks.filter(foodbank_profile__service_type=food_type)
    
    # Date range filter
    if date_from:
        foodbanks = foodbanks.filter(date_joined__gte=date_from)
    if date_to:
        foodbanks = foodbanks.filter(date_joined__lte=date_to)
    
    # Sorting
    if sort_by == 'oldest':
        foodbanks = foodbanks.order_by('date_joined')
    else:  # newest
        foodbanks = foodbanks.order_by('-date_joined')
    
    # Pagination
    paginator = Paginator(foodbanks, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Food Banks Management',
        'page_obj': page_obj,
        'search': search,
        'is_active': is_active,
        'approval_status': approval_status,
        'food_type': food_type,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'total_count': foodbanks.count(),
    }
    return render(request, 'custom_admin/foodbanks_management.html', context)


@staff_member_required
def recipients_management(request):
    """Manage recipients with table view"""
    search = request.GET.get('search', '').strip()
    is_active = request.GET.get('is_active', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort_by', 'newest')  # newest or oldest

    recipients = _get_recipients_queryset(search, is_active, date_from, date_to, sort_by)
    
    # Pagination
    paginator = Paginator(recipients, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Recipients Management',
        'page_obj': page_obj,
        'search': search,
        'is_active': is_active,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'total_count': recipients.count(),
    }
    return render(request, 'custom_admin/recipients_management.html', context)


def _get_recipients_queryset(search='', is_active='', date_from='', date_to='', sort_by='newest'):
    """Return filtered recipients queryset used by table and exports."""
    recipients = CustomUser.objects.filter(user_type='RECIPIENT').select_related('recipient_profile')

    if search:
        recipients = recipients.filter(
            Q(email__icontains=search) |
            Q(phone_number__icontains=search) |
            Q(recipient_profile__full_name__icontains=search)
        )

    if is_active:
        recipients = recipients.filter(is_active=is_active == 'true')

    if date_from:
        recipients = recipients.filter(date_joined__date__gte=date_from)
    if date_to:
        recipients = recipients.filter(date_joined__date__lte=date_to)

    if sort_by == 'oldest':
        recipients = recipients.order_by('date_joined')
    else:
        recipients = recipients.order_by('-date_joined')

    return recipients


def _build_recipients_report_data(recipients):
    """Build recipients report headers/rows matching recipients management table."""
    headers = ['S/No', 'Full Name', 'Email', 'Phone Number', 'Location', 'Account Type', 'Organization', 'Status', 'Date Joined', 'Last Login']
    data = []
    for index, user in enumerate(recipients, start=1):
        profile = getattr(user, 'recipient_profile', None)
        account_type = 'Organization' if (profile and profile.is_organization) else 'Individual'
        organization = (
            profile.organization_name if (profile and profile.is_organization and profile.organization_name)
            else 'Not organisation'
        )
        data.append([
            index,
            profile.full_name if profile else user.email,
            user.email,
            user.phone_number or '-',
            profile.location if profile else '-',
            account_type,
            organization,
            'Active' if user.is_active else 'Inactive',
            timezone.localtime(user.date_joined).strftime('%b %d, %Y') if user.date_joined else '-',
            timezone.localtime(user.last_login).strftime('%b %d, %Y %H:%M') if user.last_login else 'Never',
        ])
    return headers, data


def _get_admins_queryset(search='', is_active='', date_from='', date_to='', sort_by='newest'):
    """Return filtered admins queryset used by table and exports."""
    admins = CustomUser.objects.filter(user_type='ADMIN')

    if search:
        admins = admins.filter(
            Q(email__icontains=search) |
            Q(phone_number__icontains=search)
        )

    if is_active:
        admins = admins.filter(is_active=is_active == 'true')

    if date_from:
        admins = admins.filter(date_joined__date__gte=date_from)
    if date_to:
        admins = admins.filter(date_joined__date__lte=date_to)

    if sort_by == 'oldest':
        admins = admins.order_by('date_joined')
    else:
        admins = admins.order_by('-date_joined')

    return admins


@superuser_required
def admins_management(request):
    """Manage admins with table view."""
    search = request.GET.get('search', '').strip()
    is_active = request.GET.get('is_active', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort_by', 'newest')

    admins = _get_admins_queryset(search, is_active, date_from, date_to, sort_by)

    paginator = Paginator(admins, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'title': 'Administrators Management',
        'page_obj': page_obj,
        'search': search,
        'is_active': is_active,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'total_count': admins.count(),
    }
    return render(request, 'custom_admin/admins_management.html', context)


@superuser_required
def export_admins_excel(request):
    """Export administrators report with the same columns/content as admins table."""
    search = request.GET.get('search', '').strip()
    is_active = request.GET.get('is_active', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort_by', 'newest')

    admins = _get_admins_queryset(search, is_active, date_from, date_to, sort_by)

    wb = Workbook()
    ws = wb.active
    ws.title = "Administrators"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'S/No', 'Email', 'Phone Number', 'Staff Status',
        'Superuser', 'Status', 'Date Joined', 'Last Login'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_num, user in enumerate(admins, 2):
        row_values = [
            row_num - 1,
            user.email,
            user.phone_number or '-',
            'Staff' if user.is_staff else 'Not Staff',
            'Superuser' if user.is_superuser else 'Regular',
            'Active' if user.is_active else 'Inactive',
            user.date_joined.strftime('%b %d, %Y') if user.date_joined else '-',
            user.last_login.strftime('%b %d, %Y %H:%M') if user.last_login else 'Never',
        ]
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [1, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center")

    widths = [8, 34, 18, 14, 14, 12, 16, 20]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="administrators_report_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@superuser_required
def export_admins_pdf(request):
    """Export administrators report on A4 with structured table layout."""
    search = request.GET.get('search', '').strip()
    is_active = request.GET.get('is_active', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort_by', 'newest')

    admins = _get_admins_queryset(search, is_active, date_from, date_to, sort_by)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'AdminsReportTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#dc3545'),
        alignment=1,
        spaceAfter=8,
    )
    info_style = ParagraphStyle(
        'AdminsReportInfo',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
    )

    elements.append(Paragraph("Administrators Report", title_style))

    filter_parts = []
    if search:
        filter_parts.append(f"Search: {search}")
    if is_active:
        filter_parts.append(f"Status: {'Active' if is_active == 'true' else 'Inactive'}")
    if date_from:
        filter_parts.append(f"Date From: {date_from}")
    if date_to:
        filter_parts.append(f"Date To: {date_to}")
    filter_text = " | ".join(filter_parts) if filter_parts else "No filters applied"

    elements.append(Paragraph(
        f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        f"Total Records: {admins.count()}<br/>"
        f"{filter_text}",
        info_style
    ))
    elements.append(Spacer(1, 10))

    data = [[
        'S/No', 'Email', 'Phone Number', 'Staff Status',
        'Superuser', 'Status', 'Date Joined', 'Last Login'
    ]]
    for idx, user in enumerate(admins, 1):
        data.append([
            str(idx),
            user.email,
            user.phone_number or '-',
            'Staff' if user.is_staff else 'Not Staff',
            'Superuser' if user.is_superuser else 'Regular',
            'Active' if user.is_active else 'Inactive',
            user.date_joined.strftime('%b %d, %Y') if user.date_joined else '-',
            user.last_login.strftime('%b %d, %Y %H:%M') if user.last_login else 'Never',
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[0.55 * inch, 2.45 * inch, 1.2 * inch, 1.0 * inch, 1.0 * inch, 0.9 * inch, 1.0 * inch, 1.35 * inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="administrators_report_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@staff_member_required
def user_detail(request, user_id):
    """User detail view with edit capabilities"""
    user = get_object_or_404(CustomUser, id=user_id)

    # Get user profile + correct form class based on type
    profile = None
    profile_form_class = None

    if user.user_type == 'DONOR' and hasattr(user, 'donor_profile'):
        profile = user.donor_profile
        profile_form_class = DonorProfileAdminForm
    elif user.user_type == 'FOODBANK' and hasattr(user, 'foodbank_profile'):
        profile = user.foodbank_profile
        profile_form_class = FoodBankProfileAdminForm
    elif user.user_type == 'RECIPIENT' and hasattr(user, 'recipient_profile'):
        profile = user.recipient_profile
        profile_form_class = RecipientProfileAdminForm

    if request.method == 'POST':
        user_form = CustomUserAdminUpdateForm(request.POST, instance=user)
        profile_form = profile_form_class(request.POST, instance=profile) if profile_form_class and profile else None

        forms_valid = user_form.is_valid() and (profile_form.is_valid() if profile_form else True)
        if forms_valid:
            with transaction.atomic():
                user_form.save()
                if profile_form:
                    profile_form.save()

            messages.success(request, f'User {user.email} updated successfully.')
            return redirect('custom_admin:user_detail', user_id=user.id)
    else:
        user_form = CustomUserAdminUpdateForm(instance=user)
        profile_form = profile_form_class(instance=profile) if profile_form_class and profile else None

    # Stats
    user_stats = {}
    if user.user_type == 'DONOR':
        donations = user.donation_set.all()
        user_stats = {
            'total_donations': donations.count(),
            'total_amount_donated': donations.aggregate(total=Sum('amount'))['total'] or 0,
        }
    elif user.user_type == 'FOODBANK' and profile:
        user_stats = {
            'donations_received': Donation.objects.filter(foodbank=profile).count(),
            'active_requests': FoodBankRequest.objects.filter(foodbank=profile, status='active').count(),
        }
    elif user.user_type == 'RECIPIENT' and profile:
        user_stats = {
            'requests_made': profile.requests.count() if hasattr(profile, 'requests') else 0,
            'allocations_received': DonationAllocation.objects.filter(recipient=profile).count(),
        }

    context = {
        'title': f'User: {user.email}',
        'user': user,
        'profile': profile,
        'user_form': user_form,
        'profile_form': profile_form,
        'user_stats': user_stats,
    }
    return render(request, 'custom_admin/user_detail.html', context)


@staff_member_required
def donation_management(request):
    """Donation management with filtering"""
    donations = Donation.objects.select_related('donor', 'foodbank').order_by('-donated_at')
    
    # Filtering
    donation_type = request.GET.get('donation_type')
    delivery_status = request.GET.get('delivery_status')
    search = request.GET.get('search')
    
    if donation_type:
        donations = donations.filter(donation_type=donation_type)
    if delivery_status:
        donations = donations.filter(delivery_status=delivery_status)
    if search:
        donations = donations.filter(
            Q(donor__email__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(item_name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(donations, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Donation Management',
        'page_obj': page_obj,
        'donation_types': Donation.DONATION_TYPES,
        'delivery_statuses': Donation.DELIVERY_STATUS_CHOICES,
        'current_filters': {
            'donation_type': donation_type,
            'delivery_status': delivery_status,
            'search': search,
        }
    }
    return render(request, 'custom_admin/donation_management.html', context)


@staff_member_required
def foodbank_requests(request):
    """Food bank requests management"""
    requests = FoodBankRequest.objects.select_related('foodbank').order_by('-created_at')
    
    # Filtering
    priority = request.GET.get('priority')
    status = request.GET.get('status')
    search = request.GET.get('search')
    
    if priority:
        requests = requests.filter(priority=priority)
    if status:
        requests = requests.filter(status=status)
    if search:
        requests = requests.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Food Bank Requests',
        'page_obj': page_obj,
        'priorities': FoodBankRequest.PRIORITY_CHOICES,
        'statuses': FoodBankRequest.STATUS_CHOICES,
        'current_filters': {
            'priority': priority,
            'status': status,
            'search': search,
        }
    }
    return render(request, 'custom_admin/foodbank_requests.html', context)


@staff_member_required
def bulk_actions(request):
    """Handle bulk actions on users/donations"""
    if request.method == 'POST':
        action = request.POST.get('action')
        model_type = request.POST.get('model_type')
        selected_ids = request.POST.getlist('selected_items')
        
        if not selected_ids:
            messages.error(request, 'No items selected.')
            return redirect(request.META.get('HTTP_REFERER', 'custom_admin:dashboard'))
        
        if model_type == 'users':
            users = CustomUser.objects.filter(id__in=selected_ids)
            if action == 'activate':
                users.update(is_active=True)
                messages.success(request, f'Activated {len(selected_ids)} users.')
            elif action == 'deactivate':
                users.update(is_active=False)
                messages.success(request, f'Deactivated {len(selected_ids)} users.')
        
        elif model_type == 'donations':
            donations = Donation.objects.filter(id__in=selected_ids)
            if action == 'mark_delivered':
                donations.update(delivery_status='delivered')
                messages.success(request, f'Marked {len(selected_ids)} donations as delivered.')
        
        return redirect(request.META.get('HTTP_REFERER', 'custom_admin:dashboard'))


def export_recipients_pdf(request):
    """Export recipients report as structured A4 PDF matching recipients table columns."""
    search = request.GET.get('search', '').strip()
    is_active = request.GET.get('is_active', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort_by', 'newest')

    recipients = _get_recipients_queryset(search, is_active, date_from, date_to, sort_by)
    _, rows = _build_recipients_report_data(recipients)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RecipientsReportTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#f59e0b'),
        alignment=1,
        spaceAfter=8,
    )
    info_style = ParagraphStyle(
        'RecipientsReportInfo',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
    )

    elements.append(Paragraph("Recipients Report", title_style))

    filter_parts = []
    if search:
        filter_parts.append(f"Search: {search}")
    if is_active:
        filter_parts.append(f"Status: {'Active' if is_active == 'true' else 'Inactive'}")
    if date_from:
        filter_parts.append(f"Date From: {date_from}")
    if date_to:
        filter_parts.append(f"Date To: {date_to}")
    filter_text = " | ".join(filter_parts) if filter_parts else "No filters applied"

    elements.append(Paragraph(
        f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        f"Total Records: {recipients.count()}<br/>"
        f"{filter_text}",
        info_style
    ))
    elements.append(Spacer(1, 10))

    data = [[
        'S/No', 'Full Name', 'Email', 'Phone Number', 'Location',
        'Account Type', 'Organization', 'Status', 'Date Joined', 'Last Login'
    ]]
    for row in rows:
        data.append([str(row[0])] + row[1:])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[0.45 * inch, 1.4 * inch, 2.0 * inch, 1.05 * inch, 1.1 * inch, 0.9 * inch, 1.2 * inch, 0.7 * inch, 0.9 * inch, 1.35 * inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (6, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fff8e6')]),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="recipients_report_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@staff_member_required
def export_report(request, report_type):
    """Export various reports in multiple formats (PDF, Excel, CSV)"""
    export_format = request.GET.get('format', 'pdf').lower()

    if report_type == 'recipients_report':
        search = request.GET.get('search', '').strip()
        is_active = request.GET.get('is_active', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        sort_by = request.GET.get('sort_by', 'newest')
        recipients = _get_recipients_queryset(search, is_active, date_from, date_to, sort_by)
        headers, data = _build_recipients_report_data(recipients)
        title = 'Recipients Report'

        if export_format == 'pdf':
            return export_recipients_pdf(request)
        if export_format == 'csv':
            return export_csv(data, headers, title, report_type)
        if export_format == 'excel':
            return export_excel(data, headers, title, report_type)
    
    # Get data based on report type
    data, headers, title = get_report_data(report_type)
    
    if export_format == 'csv':
        return export_csv(data, headers, title, report_type)
    elif export_format == 'excel':
        return export_excel(data, headers, title, report_type)
    else:  # Default to PDF
        return export_pdf(data, headers, title, report_type)


def get_report_data(report_type):
    """Get data for different report types"""
    if report_type == 'user_summary':
        users = CustomUser.objects.all().order_by('-date_joined')
        headers = ['Email', 'User Type', 'Phone Number', 'Status', 'Date Joined', 'Last Login']
        data = []
        for user in users:
            data.append([
                user.email,
                user.get_user_type_display(),
                user.phone_number or '-',
                'Active' if user.is_active else 'Inactive',
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
            ])
        return data, headers, 'User Summary Report'
    
    elif report_type == 'donations_summary':
        donations = Donation.objects.select_related('donor', 'foodbank').order_by('-donated_at')
        headers = ['Donor Email', 'Food Bank', 'Type', 'Item Name', 'Amount (KES)', 'Subsidized Price', 'Date', 'Status']
        data = []
        for d in donations:
            data.append([
                d.donor.email,
                d.foodbank.foodbank_name,
                d.get_donation_type_display(),
                d.item_name or '-',
                f"{d.amount:.2f}" if d.amount else '-',
                f"{d.subsidized_price:.2f}" if d.subsidized_price else '-',
                d.donated_at.strftime('%Y-%m-%d %H:%M'),
                'Delivered' if hasattr(d, 'is_delivered') and d.is_delivered else 'Pending'
            ])
        return data, headers, 'Donations Summary Report'
    
    elif report_type == 'top_donors':
        top_donors = CustomUser.objects.filter(user_type='DONOR').annotate(
            total_donated=Sum('donation__amount'),
            donation_count=Count('donation')
        ).filter(total_donated__isnull=False).order_by('-total_donated')[:50]
        
        headers = ['Rank', 'Donor', 'Donations', 'Total Donated (KES)']
        data = []
        for idx, donor in enumerate(top_donors, start=1):
            display_name = donor.donor_profile.full_name if hasattr(donor, 'donor_profile') and donor.donor_profile and donor.donor_profile.full_name else donor.email
            data.append([
                idx,
                display_name,
                donor.donation_count,
                f"{donor.total_donated:.2f}"
            ])
        return data, headers, 'Top Donors Report'
    
    elif report_type == 'top_foodbanks':
        top_foodbanks = FoodBankProfile.objects.annotate(
            total_received=Sum('donation__amount'),
            donations_received=Count('donation')
        ).filter(total_received__gt=0).order_by('-total_received')[:50]
        
        headers = ['Rank', 'Food Bank', 'Donations', 'Total Received (KES)']
        data = []
        for idx, fb in enumerate(top_foodbanks, start=1):
            data.append([
                idx,
                fb.foodbank_name,
                fb.donations_received,
                f"{(fb.total_received or 0):.2f}"
            ])
        return data, headers, 'Top Food Banks Report'
    
    elif report_type == 'support_messages':
        messages = SupportMessage.objects.select_related('user').order_by('-created_at')
        headers = ['User Email', 'User Type', 'Subject', 'Status', 'Priority', 'Created', 'Resolved']
        data = []
        for msg in messages:
            data.append([
                msg.user.email,
                msg.user.get_user_type_display(),
                msg.get_subject_display(),
                msg.get_status_display(),
                msg.get_priority_display(),
                msg.created_at.strftime('%Y-%m-%d %H:%M'),
                msg.resolved_at.strftime('%Y-%m-%d %H:%M') if msg.resolved_at else '-'
            ])
        return data, headers, 'Support Messages Report'
    
    elif report_type == 'donors_report':
        donors = CustomUser.objects.filter(user_type='DONOR').select_related('donor_profile').order_by('-date_joined')
        headers = ['S/No', 'Full Name', 'Email', 'Phone Number', 'Donor Type', 'Organization', 'Status', 'Date Joined', 'Last Login', 'Total Donations']
        data = []
        for index, user in enumerate(donors, start=1):
            profile = user.donor_profile
            donor_type = 'Organization' if (profile and profile.is_organization) else 'Individual'
            organization = profile.organization_name if (profile and profile.organization_name) else '-'
            
            data.append([
                index,
                profile.full_name if profile else user.email,
                user.email,
                user.phone_number or '-',
                donor_type,
                organization,
                'Active' if user.is_active else 'Inactive',
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never',
                user.donation_set.count()
            ])
        return data, headers, 'Donors Report'
    
    elif report_type == 'foodbanks_report':
        foodbanks = CustomUser.objects.filter(user_type='FOODBANK').select_related('foodbank_profile').order_by('-date_joined')
        headers = ['S/No', 'Food Bank Name', 'Contact Person', 'Email', 'Phone Number', 'Address', 'Website', 'Established Year', 'Status', 'Approval Status', 'Date Joined', 'Last Login', 'Donations Received']
        data = []
        for index, user in enumerate(foodbanks, start=1):
            profile = user.foodbank_profile
            data.append([
                index,
                profile.foodbank_name if profile else user.email,
                profile.contact_person if profile else '-',
                user.email,
                user.phone_number or '-',
                profile.address if profile else '-',
                profile.website_url if profile else '-',
                profile.established_year if profile and profile.established_year else '-',
                'Active' if user.is_active else 'Inactive',
                profile.get_is_approved_display() if profile else 'Pending',
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never',
                profile.donation_set.count() if profile else 0
            ])
        return data, headers, 'Food Banks Report'
    
    elif report_type == 'recipients_report':
        recipients = _get_recipients_queryset()
        headers, data = _build_recipients_report_data(recipients)
        return data, headers, 'Recipients Report'
    
    elif report_type == 'admins_report':
        admins = CustomUser.objects.filter(user_type='ADMIN').order_by('-date_joined')
        headers = ['S/No', 'Email', 'Phone Number', 'Status', 'Superuser', 'Date Joined', 'Last Login']
        data = []
        for index, user in enumerate(admins, start=1):
            data.append([
                index,
                user.email,
                user.phone_number or '-',
                'Active' if user.is_active else 'Inactive',
                'Yes' if user.is_superuser else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
            ])
        return data, headers, 'Administrators Report'
    
    elif report_type == 'recipient_requests':
        from authentication.models import RequestManagement
        requests = RequestManagement.objects.select_related(
            'recipient__user', 'foodbank', 'assigned_foodbank'
        ).all()
        headers = ['Recipient Name', 'Email', 'Phone', 'Request Type', 'Description', 'Quantity', 'Unit', 
                   'Foodbank', 'Status', 'Delivery Method', 'Location', 'Request Date', 'Fulfilled Date', 'Qty Fulfilled']
        data = []
        for req in requests:
            data.append([
                req.recipient.full_name,
                req.recipient.user.email,
                req.recipient.user.phone_number or '-',
                req.get_request_type_display(),
                req.description,
                req.quantity,
                req.unit,
                req.foodbank.foodbank_name if req.foodbank else (req.assigned_foodbank.foodbank_name if req.assigned_foodbank else '-'),
                req.get_status_display(),
                req.get_delivery_method_display(),
                req.location,
                req.time_of_request.strftime('%Y-%m-%d %H:%M'),
                req.fulfilled_at.strftime('%Y-%m-%d %H:%M') if req.fulfilled_at else '-',
                req.quantity_fulfilled
            ])
        return data, headers, 'Recipient Requests Report'
    
    elif report_type == 'foodbank_requests':
        from authentication.models import FoodBankRequest
        requests = FoodBankRequest.objects.select_related(
            'foodbank__user', 'original_request__recipient'
        ).all()
        headers = ['Foodbank Name', 'Contact Person', 'Email', 'Request Title', 'Description', 'Donation Type', 
                   'Priority', 'Status', 'Quantity Needed', 'Unit', 'Linked Recipient', 'Deadline', 'Created Date']
        data = []
        for req in requests:
            data.append([
                req.foodbank.foodbank_name,
                req.foodbank.contact_person,
                req.foodbank.user.email,
                req.title,
                req.description,
                req.get_donation_type_display(),
                req.get_priority_display(),
                req.get_status_display(),
                req.quantity_needed or '-',
                req.quantity_unit or '-',
                req.original_request.recipient.full_name if req.original_request else '-',
                req.deadline.strftime('%Y-%m-%d %H:%M') if req.deadline else '-',
                req.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        return data, headers, 'FoodBank Requests Report'
    
    elif report_type == 'donor_donations':
        donations = Donation.objects.select_related('donor', 'foodbank__user').all()
        headers = ['Donor Name', 'Donor Email', 'Donor Phone', 'Foodbank Name', 'Donation Type', 'Item Name', 
                   'Quantity', 'Unit', 'Amount (KES)', 'Subsidized Price', 'Donation Date', 'Payment Method', 'Status']
        data = []
        for d in donations:
            data.append([
                d.donor.email,
                d.donor.email,
                d.donor.phone_number or '-',
                d.foodbank.foodbank_name,
                d.get_donation_type_display(),
                d.item_name or '-',
                d.quantity or '-',
                d.quantity_unit or '-',
                f"{d.amount:.2f}" if d.amount else '-',
                f"{d.subsidized_price:.2f}" if d.subsidized_price else '-',
                d.donated_at.strftime('%Y-%m-%d %H:%M'),
                d.payment_method or '-',
                'Delivered' if hasattr(d, 'is_delivered') and d.is_delivered else 'Pending'
            ])
        return data, headers, 'Donor Donations Report'
    
    elif report_type == 'complete_donation_flow':
        from authentication.models import DonationAllocation
        allocations = DonationAllocation.objects.select_related(
            'donation__donor',
            'donation__foodbank__user',
            'recipient__user'
        ).all()
        headers = ['Donor Name', 'Donor Email', 'Donation Type', 'Item Name', 'Donated Amount/Qty', 
                   'Foodbank Name', 'Foodbank Email', 'Allocated Amount/Qty', 'Recipient Name', 'Recipient Email', 
                   'Recipient Phone', 'Recipient Location', 'Allocation Date', 'Acknowledged']
        data = []
        for alloc in allocations:
            data.append([
                alloc.donation.donor.email,
                alloc.donation.donor.email,
                alloc.donation.get_donation_type_display(),
                alloc.donation.item_name or '-',
                f"{alloc.donation.quantity} {alloc.donation.quantity_unit}" if alloc.donation.quantity else f"KES {alloc.donation.amount:.2f}",
                alloc.donation.foodbank.foodbank_name,
                alloc.donation.foodbank.user.email,
                f"{alloc.quantity}" if alloc.quantity else f"KES {alloc.amount:.2f}",
                alloc.recipient.full_name,
                alloc.recipient.user.email,
                alloc.recipient.user.phone_number or '-',
                alloc.recipient.location or '-',
                alloc.allocated_at.strftime('%Y-%m-%d %H:%M'),
                'Yes' if alloc.is_acknowledged else 'No'
            ])
        return data, headers, 'Complete Donation Flow Report'
    
    else:
        return [], [], 'Unknown Report'


def export_csv(data, headers, title, report_type):
    """Export data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write title and metadata
    writer.writerow([title])
    writer.writerow([f'Generated on: {timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([f'Total Records: {len(data)}'])
    writer.writerow([])  # Empty row
    
    # Write headers
    writer.writerow(headers)
    
    # Write data
    for row in data:
        writer.writerow(row)
    
    return response


def export_excel(data, headers, title, report_type):
    """Export data as Excel with custom formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Metadata
    ws[f'A3'] = f'Generated on: {timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")}'
    ws[f'A4'] = f'Total Records: {len(data)}'
    
    # Headers
    header_row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.border = border
            
            # Auto-adjust column width
            column_letter = get_column_letter(col_idx)
            if ws.column_dimensions[column_letter].width < len(str(value)):
                ws.column_dimensions[column_letter].width = min(len(str(value)) + 2, 50)
    
    # Set minimum column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        if ws.column_dimensions[column_letter].width < 15:
            ws.column_dimensions[column_letter].width = 15
    
    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


def export_pdf(data, headers, title, report_type):
    """Export data as PDF with optimized A3 landscape formatting."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A3),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch
    )
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=10,
        alignment=1,
        textColor=colors.darkblue
    )
    meta_style = ParagraphStyle(
        'Metadata',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        leading=10
    )
    body_style = ParagraphStyle(
        'BodyCell',
        parent=styles['Normal'],
        fontSize=7,
        leading=8.2,
        wordWrap='CJK',
    )
    header_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.4,
        textColor=colors.whitesmoke,
        alignment=1,
        leading=8.6,
    )
    
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 4))
    
    story.append(Paragraph(f'Generated on: {timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")}', meta_style))
    story.append(Paragraph(f'Total Records: {len(data)}', meta_style))
    story.append(Spacer(1, 8))

    # Convert headers/rows to wrapped Paragraph cells for better A3 usage.
    wrapped_headers = [Paragraph(str(h), header_style) for h in headers]
    wrapped_rows = []
    for row in data:
        wrapped_rows.append([
            Paragraph(str(value) if value is not None else '-', body_style)
            for value in row
        ])
    table_data = [wrapped_headers] + wrapped_rows

    # Build weighted widths using header + sample data lengths.
    usable_width = landscape(A3)[0] - doc.leftMargin - doc.rightMargin
    col_scores = []
    for idx, header in enumerate(headers):
        sample_values = [str(r[idx]) if idx < len(r) and r[idx] is not None else '' for r in data[:80]]
        max_sample = max([len(v) for v in sample_values], default=0)
        score = max(len(str(header)) * 1.15, min(max_sample, 42), 6)
        col_scores.append(score)
    total_score = sum(col_scores) if col_scores else 1
    col_widths = [usable_width * (s / total_score) for s in col_scores]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.Color(0.72, 0.72, 0.72)),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(0.985, 0.985, 0.985), colors.white]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    
    story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


@staff_member_required
def analytics(request):
    """Comprehensive analytics and reports page with enhanced donation management metrics"""
    # Time period filter
    period = request.GET.get('period', '30')  # days
    try:
        days = int(period)
    except ValueError:
        days = 30
    start_date = timezone.now() - timedelta(days=days)
    
    # ===== USER STATISTICS =====
    total_users = CustomUser.objects.count()
    total_donors = CustomUser.objects.filter(user_type='DONOR').count()
    total_foodbanks = CustomUser.objects.filter(user_type='FOODBANK').count()
    total_recipients = CustomUser.objects.filter(user_type='RECIPIENT').count()
    
    # Active users (logged in within period)
    active_users = CustomUser.objects.filter(last_login__gte=start_date).count()
    new_users = CustomUser.objects.filter(date_joined__gte=start_date).count()
    
    # ===== DONATION STATISTICS =====
    total_donations = Donation.objects.count()
    period_donations = Donation.objects.filter(donated_at__gte=start_date).count()
    
    # Donation by type
    item_donations = Donation.objects.filter(donation_type='item').count()
    money_donations = Donation.objects.filter(donation_type='money').count()
    subsidized_donations = Donation.objects.filter(donation_type='subsidized').count()
    csr_donations = Donation.objects.filter(donation_type='csr').count()
    other_donations = Donation.objects.filter(donation_type='other').count()
    
    # Donation by status
    pending_donations = Donation.objects.filter(status='pending').count()
    accepted_donations = Donation.objects.filter(status='accepted').count()
    declined_donations = Donation.objects.filter(status='declined').count()
    
    # Financial metrics
    total_monetary_value = Donation.objects.filter(
        donation_type='money', status='accepted'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_subsidized_value = Donation.objects.filter(
        donation_type='subsidized', status='accepted'
    ).aggregate(total=Sum('subsidized_price'))['total'] or 0
    
    period_monetary_value = Donation.objects.filter(
        donation_type='money', status='accepted', donated_at__gte=start_date
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_donated_amount = Donation.objects.aggregate(total=Sum('amount'))['total'] or 0

    # Posted donations counts
    posted_unspecified_count = UnspecifiedDonationManagement.objects.filter(
        recipient_status='received'
    ).count()
    posted_direct_count = Donation.objects.filter(
        foodbank_request__isnull=False,
        status='accepted'
    ).count()
    posted_subsidized_count = Donation.objects.filter(
        donation_type='subsidized',
        status='accepted',
        delivery_status='delivered'
    ).count()
    posted_specified_count = RequestManagement.objects.filter(
        status__in=['fulfilled', 'acknowledged']
    ).count()

    # ===== REQUEST STATISTICS =====
    # Recipient Requests
    total_recipient_requests = RecipientRequest.objects.count()
    pending_recipient_requests = RecipientRequest.objects.filter(status='pending').count()
    # ... (rest of the code remains the same)
    completed_recipient_requests = RecipientRequest.objects.filter(status='completed').count()
    
    # Foodbank Requests
    total_foodbank_requests = FoodBankRequest.objects.count()
    active_foodbank_requests = FoodBankRequest.objects.filter(status='active').count()
    fulfilled_foodbank_requests = FoodBankRequest.objects.filter(status='fulfilled').count()
    urgent_foodbank_requests = FoodBankRequest.objects.filter(
        status='active', priority='urgent'
    ).count()
    
    # ===== ALLOCATION STATISTICS =====
    total_allocations = DonationAllocation.objects.count()
    period_allocations = DonationAllocation.objects.filter(allocated_at__gte=start_date).count()
    pending_acknowledgments = DonationAllocation.objects.filter(is_acknowledged=False).count()
    
    total_allocated_quantity = DonationAllocation.objects.aggregate(
        total=Sum('quantity')
    )['total'] or 0
    
    total_allocated_amount = DonationAllocation.objects.aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # ===== DISCUSSION STATISTICS =====
    total_discussions = DonationDiscussion.objects.count()
    active_discussions = DonationDiscussion.objects.filter(status='in_progress').count()
    agreed_discussions = DonationDiscussion.objects.filter(status='agreed').count()
    declined_discussions = DonationDiscussion.objects.filter(status='declined').count()
    
    # ===== PAYMENT STATISTICS =====
    total_transactions = PaymentTransaction.objects.count()
    completed_transactions = PaymentTransaction.objects.filter(status='completed').count()
    pending_transactions = PaymentTransaction.objects.filter(status='pending').count()
    failed_transactions = PaymentTransaction.objects.filter(status='failed').count()
    
    total_transaction_amount = PaymentTransaction.objects.filter(
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # ===== FOODBANK APPROVALS =====
    pending_foodbanks = FoodBankProfile.objects.filter(is_approved='pending').count()
    approved_foodbanks = FoodBankProfile.objects.filter(is_approved='approved').count()
    rejected_foodbanks = FoodBankProfile.objects.filter(is_approved='rejected').count()
    
    # ===== SUPPORT MESSAGES =====
    total_support_messages = SupportMessage.objects.count()
    new_support_messages = SupportMessage.objects.filter(status='new').count()
    urgent_support_messages = SupportMessage.objects.filter(priority='urgent').count()
    
    # Recent activity
    recent_users = CustomUser.objects.order_by('-date_joined')[:10]
    recent_donations = Donation.objects.select_related('donor', 'foodbank').order_by('-donated_at')[:10]
    recent_requests = FoodBankRequest.objects.select_related('foodbank').order_by('-created_at')[:10]
    
    # User counts by type for pie chart
    user_counts = CustomUser.objects.values('user_type').annotate(count=Count('id'))
    user_type_labels = [uc['user_type'] for uc in user_counts]
    user_type_data = [uc['count'] for uc in user_counts]
    
    # Monthly registrations for line chart
    now = timezone.now()
    months = [(now - timedelta(days=30*i)).strftime('%b %Y') for i in reversed(range(12))]
    registration_counts = []
    for i in reversed(range(12)):
        start = (now - timedelta(days=30*(i+1)))
        end = (now - timedelta(days=30*i))
        count = CustomUser.objects.filter(date_joined__gte=start, date_joined__lt=end).count()
        registration_counts.append(count)
    
    # Donations by type for bar chart
    donation_counts = Donation.objects.values('donation_type').annotate(count=Count('id'))
    donation_type_labels = [d['donation_type'] for d in donation_counts]
    donation_type_data = [d['count'] for d in donation_counts]
    
    # Monthly donation trends
    donation_months = []
    donation_amounts = []
    for i in reversed(range(12)):
        start = (now - timedelta(days=30*(i+1)))
        end = (now - timedelta(days=30*i))
        month_donations = Donation.objects.filter(donated_at__gte=start, donated_at__lt=end)
        total_amount = month_donations.aggregate(total=Sum('amount'))['total'] or 0
        donation_months.append(start.strftime('%b %Y'))
        donation_amounts.append(float(total_amount))
    
    # Top donors
    top_donors = CustomUser.objects.filter(user_type='DONOR').annotate(
        total_donated=Sum('donation__amount'),
        donation_count=Count('donation')
    ).filter(total_donated__isnull=False).order_by('-total_donated')[:10]
    
    # Top foodbanks by donations received
    top_foodbanks = FoodBankProfile.objects.annotate(
        donations_received=Count('donation'),
        total_received=Sum('donation__amount')
    ).filter(donations_received__gt=0).order_by('-donations_received')[:10]
    
    context = {
        'title': 'Analytics & Reports',
        'period': days,
        
        # User Statistics
        'total_users': total_users,
        'total_donors': total_donors,
        'total_foodbanks': total_foodbanks,
        'total_recipients': total_recipients,
        'active_users': active_users,
        'new_users': new_users,
        
        # Donation Statistics
        'total_donations': total_donations,
        'period_donations': period_donations,
        'item_donations': item_donations,
        'money_donations': money_donations,
        'subsidized_donations': subsidized_donations,
        'csr_donations': csr_donations,
        'other_donations': other_donations,
        'pending_donations': pending_donations,
        'accepted_donations': accepted_donations,
        'declined_donations': declined_donations,
        
        # Financial Metrics
        'total_monetary_value': total_monetary_value,
        'total_subsidized_value': total_subsidized_value,
        'period_monetary_value': period_monetary_value,
        'total_financial_impact': total_monetary_value + total_subsidized_value,
        'total_donated_amount': total_donated_amount,
        'posted_unspecified_count': posted_unspecified_count,
        'posted_direct_count': posted_direct_count,
        'posted_subsidized_count': posted_subsidized_count,
        'posted_specified_count': posted_specified_count,
        
        # Request Statistics
        'total_recipient_requests': total_recipient_requests,
        'pending_recipient_requests': pending_recipient_requests,
        'completed_recipient_requests': completed_recipient_requests,
        'total_foodbank_requests': total_foodbank_requests,
        'active_foodbank_requests': active_foodbank_requests,
        'fulfilled_foodbank_requests': fulfilled_foodbank_requests,
        'urgent_foodbank_requests': urgent_foodbank_requests,
        
        # Allocation Statistics
        'total_allocations': total_allocations,
        'period_allocations': period_allocations,
        'pending_acknowledgments': pending_acknowledgments,
        'total_allocated_quantity': total_allocated_quantity,
        'total_allocated_amount': total_allocated_amount,
        
        # Discussion Statistics
        'total_discussions': total_discussions,
        'active_discussions': active_discussions,
        'agreed_discussions': agreed_discussions,
        'declined_discussions': declined_discussions,
        
        # Payment Statistics
        'total_transactions': total_transactions,
        'completed_transactions': completed_transactions,
        'pending_transactions': pending_transactions,
        'failed_transactions': failed_transactions,
        'total_transaction_amount': total_transaction_amount,
        
        # Foodbank Approvals
        'pending_foodbanks': pending_foodbanks,
        'approved_foodbanks': approved_foodbanks,
        'rejected_foodbanks': rejected_foodbanks,
        
        # Support Messages
        'total_support_messages': total_support_messages,
        'new_support_messages': new_support_messages,
        'urgent_support_messages': urgent_support_messages,
        
        # Recent Activity
        'recent_users': recent_users,
        'recent_donations': recent_donations,
        'recent_requests': recent_requests,
        
        # Top Performers
        'top_donors': top_donors,
        'top_foodbanks': top_foodbanks,
        
        # Chart Data
        'user_type_labels': json.dumps(user_type_labels),
        'user_type_data': json.dumps(user_type_data),
        'registration_months': json.dumps(months),
        'registration_counts': json.dumps(registration_counts),
        'donation_type_labels': json.dumps(donation_type_labels),
        'donation_type_data': json.dumps(donation_type_data),
        'donation_months': json.dumps(donation_months),
        'donation_amounts': json.dumps(donation_amounts),
    }
    return render(request, 'custom_admin/analytics.html', context)


@staff_member_required
def analytics_api(request):
    """API endpoint for dashboard analytics"""
    chart_type = request.GET.get('type')
    
    if chart_type == 'user_growth':
        # Monthly user growth
        now = timezone.now()
        data = []
        for i in range(12):
            start = (now - timedelta(days=30*i)).replace(day=1)
            end = (now - timedelta(days=30*(i-1))).replace(day=1)
            count = CustomUser.objects.filter(date_joined__gte=start, date_joined__lt=end).count()
            data.append({
                'month': start.strftime('%b %Y'),
                'count': count
            })
        return JsonResponse({'data': list(reversed(data))})
    
    elif chart_type == 'donation_trends':
        # Monthly donation trends
        now = timezone.now()
        data = []
        for i in range(12):
            start = (now - timedelta(days=30*i)).replace(day=1)
            end = (now - timedelta(days=30*(i-1))).replace(day=1)
            donations = Donation.objects.filter(donated_at__gte=start, donated_at__lt=end)
            total_amount = donations.aggregate(total=Sum('amount'))['total'] or 0
            data.append({
                'month': start.strftime('%b %Y'),
                'amount': float(total_amount)
            })
        return JsonResponse({'data': list(reversed(data))})
    
    return JsonResponse({'error': 'Invalid chart type'})


@admin_required
def foodbank_approvals(request):
    """Food bank approval management with filtering"""
    # Get filter parameters
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'pending')  # Default to pending
    service_type_filter = request.GET.get('service_type', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    sort_by = request.GET.get('sort', '-application_date')
    has_documents = request.GET.get('has_documents', '')
    
    # Base queryset
    registrations = FoodBankProfile.objects.select_related('user', 'approved_by')
    
    # Apply status filter
    if status_filter == 'all':
        registrations = registrations.all()
    elif status_filter == 'pending':
        registrations = registrations.filter(is_approved='pending')
    elif status_filter == 'approved':
        registrations = registrations.filter(is_approved='approved')
    elif status_filter == 'rejected':
        registrations = registrations.filter(is_approved='rejected')
    
    # Apply search filter
    if search:
        registrations = registrations.filter(
            Q(foodbank_name__icontains=search) | 
            Q(contact_person__icontains=search) |
            Q(user__email__icontains=search) |
            Q(address__icontains=search) |
            Q(contact_email__icontains=search) |
            Q(contact_phone__icontains=search)
        )
    
    # Apply service type filter
    if service_type_filter:
        registrations = registrations.filter(service_type=service_type_filter)
    
    # Apply date range filter
    if date_from:
        registrations = registrations.filter(application_date__date__gte=date_from)
    if date_to:
        registrations = registrations.filter(application_date__date__lte=date_to)
    
    # Apply documents filter
    if has_documents == 'yes':
        registrations = registrations.filter(
            Q(authority_picture__isnull=False) |
            Q(urgent_request_picture__isnull=False) |
            Q(additional_documents__isnull=False)
        )
    elif has_documents == 'no':
        registrations = registrations.filter(
            authority_picture__isnull=True,
            urgent_request_picture__isnull=True,
            additional_documents__isnull=True
        )
    
    # Apply sorting
    valid_sorts = [
        'application_date', '-application_date',
        'foodbank_name', '-foodbank_name',
        'contact_person', '-contact_person',
        'approval_date', '-approval_date'
    ]
    if sort_by in valid_sorts:
        registrations = registrations.order_by(sort_by)
    else:
        registrations = registrations.order_by('-application_date')
    
    # Get counts for different statuses
    all_registrations = FoodBankProfile.objects.all()
    pending_count = all_registrations.filter(is_approved='pending').count()
    approved_count = all_registrations.filter(is_approved='approved').count()
    rejected_count = all_registrations.filter(is_approved='rejected').count()
    total_count = all_registrations.count()
    
    # Pagination
    paginator = Paginator(registrations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get recent decisions for reference (only if showing pending)
    recent_decisions = []
    if status_filter == 'pending':
        recent_decisions = FoodBankProfile.objects.filter(
            is_approved__in=['approved', 'rejected']
        ).select_related('user', 'approved_by').order_by('-approval_date')[:5]
    
    context = {
        'title': 'Food Bank Approvals',
        'page_obj': page_obj,
        'recent_decisions': recent_decisions,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'total_count': total_count,
        # Filter values for form persistence
        'search': search,
        'status_filter': status_filter,
        'service_type_filter': service_type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'has_documents': has_documents,
        # Service type choices for dropdown
        'service_type_choices': [
            ('food', 'Food Only'),
            ('non_food', 'Non-Food Only'),
            ('both', 'Food and Non-Food')
        ],
    }
    return TemplateResponse(request, 'custom_admin/foodbank_approvals.html', context)


def _foodbank_approval_filters_queryset(request):
    """Build filtered queryset for foodbank approvals list/report."""
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'pending')
    service_type_filter = request.GET.get('service_type', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    sort_by = request.GET.get('sort', '-application_date')
    has_documents = request.GET.get('has_documents', '')

    registrations = FoodBankProfile.objects.select_related('user', 'approved_by')

    if status_filter == 'all':
        registrations = registrations.all()
    elif status_filter in ['pending', 'approved', 'rejected']:
        registrations = registrations.filter(is_approved=status_filter)

    if search:
        registrations = registrations.filter(
            Q(foodbank_name__icontains=search) |
            Q(contact_person__icontains=search) |
            Q(user__email__icontains=search) |
            Q(address__icontains=search) |
            Q(contact_email__icontains=search) |
            Q(contact_phone__icontains=search)
        )

    if service_type_filter:
        registrations = registrations.filter(service_type=service_type_filter)

    if date_from:
        registrations = registrations.filter(application_date__date__gte=date_from)
    if date_to:
        registrations = registrations.filter(application_date__date__lte=date_to)

    if has_documents == 'yes':
        registrations = registrations.filter(
            Q(authority_picture__isnull=False) |
            Q(urgent_request_picture__isnull=False) |
            Q(additional_documents__isnull=False)
        )
    elif has_documents == 'no':
        registrations = registrations.filter(
            authority_picture__isnull=True,
            urgent_request_picture__isnull=True,
            additional_documents__isnull=True
        )

    valid_sorts = [
        'application_date', '-application_date',
        'foodbank_name', '-foodbank_name',
        'contact_person', '-contact_person',
        'approval_date', '-approval_date'
    ]
    if sort_by in valid_sorts:
        registrations = registrations.order_by(sort_by)
    else:
        registrations = registrations.order_by('-application_date')

    return registrations


def _foodbank_documents_summary(registration):
    docs = []
    if registration.authority_picture:
        docs.append('Authority Picture')
    if registration.urgent_request_picture:
        docs.append('Urgent Request Picture')
    if registration.additional_documents:
        docs.append('Additional Documents')
    return ', '.join(docs) if docs else 'None'


def _foodbank_rejection_summary(registration):
    if registration.rejection_reason:
        return registration.rejection_reason
    if registration.is_approved == 'rejected':
        return 'No rejection reason provided'
    return 'N/A'


def _foodbank_approvals_export_rows(registrations):
    rows = []
    for idx, registration in enumerate(registrations, 1):
        rows.append([
            idx,
            registration.foodbank_name,
            registration.contact_person,
            registration.user.email,
            registration.contact_phone or 'N/A',
            registration.get_service_type_display(),
            registration.get_is_approved_display(),
            registration.application_date.strftime('%Y-%m-%d'),
            registration.approval_date.strftime('%Y-%m-%d %H:%M') if registration.approval_date else 'N/A',
            registration.approved_by.email if registration.approved_by else 'N/A',
            _foodbank_documents_summary(registration),
            _foodbank_rejection_summary(registration),
        ])
    return rows


def _export_foodbank_approvals_excel(rows, title, filename_prefix):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'S/No', 'Food Bank Name', 'Contact Person', 'Email', 'Phone',
        'Service Type', 'Status', 'Applied On', 'Decision Date',
        'Decided By', 'Documents', 'Rejection Reason'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num == 12:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    column_widths = [8, 26, 22, 30, 16, 18, 16, 15, 20, 28, 26, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename_prefix}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


def _export_foodbank_approvals_csv(rows, filename_prefix):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="{filename_prefix}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        'S/No', 'Food Bank Name', 'Contact Person', 'Email', 'Phone',
        'Service Type', 'Status', 'Applied On', 'Decision Date',
        'Decided By', 'Documents', 'Rejection Reason'
    ])
    for row in rows:
        writer.writerow(row)
    return response


def _export_foodbank_approvals_pdf(rows, title, filename_prefix):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'FoodbankApprovalsTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=12,
        alignment=1
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f"Generated on: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(rows)}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 12))

    data = [[
        'S/No', 'Food Bank', 'Contact', 'Email', 'Phone',
        'Service Type', 'Status', 'Applied', 'Decision',
        'Decided By', 'Documents', 'Rejection Reason'
    ]]
    for row in rows:
        data.append([
            str(row[0]),
            str(row[1])[:24],
            str(row[2])[:18],
            str(row[3])[:30],
            str(row[4])[:14],
            str(row[5])[:16],
            str(row[6])[:13],
            str(row[7])[:10],
            str(row[8])[:16],
            str(row[9])[:22],
            str(row[10])[:24],
            str(row[11])[:42],
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[0.35 * inch, 1.2 * inch, 1.0 * inch, 1.55 * inch, 0.75 * inch, 0.9 * inch, 0.75 * inch, 0.75 * inch, 1.0 * inch, 1.2 * inch, 1.2 * inch, 2.4 * inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7.0),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="{filename_prefix}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@admin_required
def approved_foodbanks(request):
    """View approved foodbanks with ability to reopen for review"""
    # Get all approved foodbank registrations
    approved_registrations = FoodBankProfile.objects.filter(
        is_approved='approved'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        approved_registrations = approved_registrations.filter(
            Q(foodbank_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(contact_person__icontains=search_query)
        )
    
    # Pagination for approved registrations
    paginator = Paginator(approved_registrations, 10)
    page_number = request.GET.get('page')
    approved_page = paginator.get_page(page_number)
    
    context = {
        'title': 'Approved Food Banks',
        'approved_registrations': approved_page,
        'approved_count': approved_registrations.count(),
        'search_query': search_query,
    }
    return TemplateResponse(request, 'custom_admin/approved_foodbanks.html', context)


@admin_required
def rejected_foodbank_applications(request):
    """View rejected foodbank applications"""
    # Get all rejected foodbank registrations
    rejected_registrations = FoodBankProfile.objects.filter(
        is_approved='rejected'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    
    # Pagination for rejected registrations
    paginator = Paginator(rejected_registrations, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    rejected_page = paginator.get_page(page_number)
    
    context = {
        'title': 'Rejected Food Bank Applications',
        'rejected_registrations': rejected_page,
        'rejected_count': rejected_registrations.count(),
    }
    return TemplateResponse(request, 'custom_admin/rejected_foodbank_applications.html', context)


@admin_required
def export_foodbank_approvals_excel(request):
    registrations = _foodbank_approval_filters_queryset(request)
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_excel(rows, 'Foodbank Approvals', 'foodbank_approvals')


@admin_required
def export_foodbank_approvals_pdf(request):
    registrations = _foodbank_approval_filters_queryset(request)
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_pdf(rows, 'Foodbank Approvals Report', 'foodbank_approvals')


@admin_required
def export_foodbank_approvals_csv(request):
    registrations = _foodbank_approval_filters_queryset(request)
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_csv(rows, 'foodbank_approvals')


@admin_required
def export_approved_foodbanks_excel(request):
    search_query = request.GET.get('search', '')
    registrations = FoodBankProfile.objects.filter(
        is_approved='approved'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    if search_query:
        registrations = registrations.filter(
            Q(foodbank_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(contact_person__icontains=search_query)
        )
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_excel(rows, 'Approved Foodbanks', 'approved_foodbanks')


@admin_required
def export_approved_foodbanks_pdf(request):
    search_query = request.GET.get('search', '')
    registrations = FoodBankProfile.objects.filter(
        is_approved='approved'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    if search_query:
        registrations = registrations.filter(
            Q(foodbank_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(contact_person__icontains=search_query)
        )
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_pdf(rows, 'Approved Foodbanks Report', 'approved_foodbanks')


@admin_required
def export_approved_foodbanks_csv(request):
    search_query = request.GET.get('search', '')
    registrations = FoodBankProfile.objects.filter(
        is_approved='approved'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    if search_query:
        registrations = registrations.filter(
            Q(foodbank_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(contact_person__icontains=search_query)
        )
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_csv(rows, 'approved_foodbanks')


@admin_required
def export_rejected_foodbank_applications_excel(request):
    registrations = FoodBankProfile.objects.filter(
        is_approved='rejected'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_excel(rows, 'Rejected Foodbanks', 'rejected_foodbank_applications')


@admin_required
def export_rejected_foodbank_applications_pdf(request):
    registrations = FoodBankProfile.objects.filter(
        is_approved='rejected'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_pdf(rows, 'Rejected Foodbank Applications Report', 'rejected_foodbank_applications')


@admin_required
def export_rejected_foodbank_applications_csv(request):
    registrations = FoodBankProfile.objects.filter(
        is_approved='rejected'
    ).select_related('user', 'approved_by').order_by('-approval_date')
    rows = _foodbank_approvals_export_rows(registrations)
    return _export_foodbank_approvals_csv(rows, 'rejected_foodbank_applications')


@admin_required
def approve_foodbank(request, foodbank_id):
    """Approve a pending foodbank registration"""
    try:
        foodbank_profile = FoodBankProfile.objects.get(id=foodbank_id, is_approved='pending')
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank registration not found or already processed.')
        return redirect('custom_admin:foodbank_approvals')
    
    if request.method == 'POST':
        # Approve the registration
        foodbank_profile.is_approved = 'approved'
        foodbank_profile.approval_date = timezone.now()
        foodbank_profile.approved_by = request.user
        foodbank_profile.save()
        
        # Create notification for the foodbank
        Notification.objects.create(
            user=foodbank_profile.user,
            notification_type='approval',
            message=f'Congratulations! Your food bank "{foodbank_profile.foodbank_name}" has been approved and is now active on FoodBank Hub.'
        )
        
        messages.success(
            request, 
            f'Food bank "{foodbank_profile.foodbank_name}" has been approved successfully!'
        )
        return redirect('custom_admin:foodbank_approvals')
    
    context = {
        'title': 'Approve Food Bank',
        'foodbank_profile': foodbank_profile,
        'action': 'approve'
    }
    return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)


@admin_required
def reject_foodbank(request, foodbank_id):
    """Reject a pending foodbank registration"""
    try:
        foodbank_profile = FoodBankProfile.objects.get(id=foodbank_id, is_approved='pending')
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank registration not found or already processed.')
        return redirect('custom_admin:foodbank_approvals')
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '').strip()
        
        if not rejection_reason:
            messages.error(request, 'Please provide a reason for rejection.')
            context = {
                'title': 'Reject Food Bank',
                'foodbank_profile': foodbank_profile,
                'action': 'reject'
            }
            return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)
        
        # Reject the registration
        foodbank_profile.is_approved = 'rejected'
        foodbank_profile.approval_date = timezone.now()
        foodbank_profile.approved_by = request.user
        foodbank_profile.rejection_reason = rejection_reason
        foodbank_profile.save()
        
        # Create notification for the foodbank
        Notification.objects.create(
            user=foodbank_profile.user,
            notification_type='rejection',
            message=f'Your food bank application has been reviewed. Please check your email for details.'
        )
        
        # Send rejection email
        try:
            subject = 'Food Bank Application Rejected - FoodBankHub'
            message = f"""
Dear {foodbank_profile.contact_person},

We regret to inform you that your food bank application for "{foodbank_profile.foodbank_name}" has been rejected after careful review.

REASON FOR REJECTION:
{rejection_reason}

NEXT STEPS:
To reapply, please create a new account with updated and accurate information. Ensure that:
1. All required documents are properly uploaded and legible
2. Your registration information is complete and accurate
3. You meet all the eligibility requirements for food bank registration

If you have any questions or need clarification about the rejection, please contact our support team.

To create a new account, please visit: {settings.SITE_URL}/register/foodbank/

Thank you for your interest in joining FoodBankHub.

Best regards,
FoodBankHub Admin Team

---
This is an automated email. Please do not reply directly to this message.
            """
            
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[foodbank_profile.user.email],
                fail_silently=False,
            )
            
            messages.success(
                request, 
                f'Food bank "{foodbank_profile.foodbank_name}" application has been rejected and email notification sent.'
            )
        except Exception as e:
            messages.warning(
                request,
                f'Food bank rejected but email notification failed: {str(e)}'
            )
        
        return redirect('custom_admin:foodbank_approvals')
    
    context = {
        'title': 'Reject Food Bank',
        'foodbank_profile': foodbank_profile,
        'action': 'reject'
    }
    return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)


@admin_required
def view_foodbank_application(request, foodbank_id):
    """View detailed foodbank application"""
    try:
        foodbank_profile = FoodBankProfile.objects.select_related('user').get(id=foodbank_id)
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank application not found.')
        return redirect('custom_admin:foodbank_approvals')
    
    context = {
        'title': f'Application - {foodbank_profile.foodbank_name}',
        'foodbank_profile': foodbank_profile,
    }
    return TemplateResponse(request, 'custom_admin/foodbank_application_detail.html', context)


@admin_required
def reopen_foodbank_application(request, foodbank_id):
    """Reopen a rejected foodbank application for review"""
    try:
        foodbank_profile = FoodBankProfile.objects.get(id=foodbank_id, is_approved='rejected')
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Rejected food bank application not found.')
        return redirect('custom_admin:foodbank_approvals')
    
    if request.method == 'POST':
        reopen_reason = request.POST.get('reopen_reason', '').strip()
        
        if not reopen_reason:
            messages.error(request, 'Please provide a reason for reopening this application.')
            context = {
                'title': 'Reopen Food Bank Application',
                'foodbank_profile': foodbank_profile,
                'action': 'reopen'
            }
            return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)
        
        # Reopen the application
        foodbank_profile.is_approved = 'pending'
        foodbank_profile.approval_date = None
        foodbank_profile.approved_by = None
        foodbank_profile.rejection_reason = f"[REOPENED] {foodbank_profile.rejection_reason or ''}\n\nREOPEN REASON: {reopen_reason}\nReopened by: {request.user.email} on {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}"
        foodbank_profile.save()
        
        # Send notification email to the foodbank
        try:
            subject = f"Application Reopened - {foodbank_profile.foodbank_name}"
            message = f"""
Dear {foodbank_profile.contact_person},

Good news! Your food bank application for "{foodbank_profile.foodbank_name}" has been reopened for review.

REOPEN REASON:
{reopen_reason}

WHAT HAPPENS NEXT:
Your application is now back in the pending review queue. Our admin team will review your application again and you will receive an email notification once a decision has been made.

STATUS: Pending Review
APPLICATION ID: {foodbank_profile.id}

If you need to update any information in your application, please contact our support team immediately.

Thank you for your patience and continued interest in joining FoodBankHub.

Best regards,
FoodBankHub Admin Team

---
This is an automated email. Please do not reply directly to this message.
            """
            
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[foodbank_profile.user.email],
                fail_silently=False,
            )
            
            messages.success(
                request, 
                f'Food bank "{foodbank_profile.foodbank_name}" application has been reopened for review and email notification sent.'
            )
        except Exception as e:
            messages.warning(
                request,
                f'Application reopened but email notification failed: {str(e)}'
            )
        
        return redirect('custom_admin:foodbank_approvals')
    
    context = {
        'title': 'Reopen Food Bank Application',
        'foodbank_profile': foodbank_profile,
        'action': 'reopen'
    }
    return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)


@admin_required
def reopen_approved_foodbank(request, foodbank_id):
    """Reopen an approved foodbank for further review"""
    try:
        foodbank_profile = FoodBankProfile.objects.get(id=foodbank_id, is_approved='approved')
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Approved food bank not found.')
        return redirect('custom_admin:foodbank_approvals')
    
    if request.method == 'POST':
        review_reason = request.POST.get('review_reason', '').strip()
        
        if not review_reason:
            messages.error(request, 'Please provide a reason for reopening this approved foodbank for review.')
            context = {
                'title': 'Reopen Approved Food Bank for Review',
                'foodbank_profile': foodbank_profile,
                'action': 'reopen_approved'
            }
            return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)
        
        # Store previous approval info
        previous_approval_info = f"""
PREVIOUS APPROVAL:
- Approved by: {foodbank_profile.approved_by.email if foodbank_profile.approved_by else 'Unknown'}
- Approval date: {foodbank_profile.approval_date.strftime('%Y-%m-%d %H:%M:%S') if foodbank_profile.approval_date else 'Unknown'}

REOPENED FOR REVIEW:
- Reopened by: {request.user.email}
- Reopened on: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}
- Review reason: {review_reason}
"""
        
        # Reopen the foodbank for review
        foodbank_profile.is_approved = 'pending'
        foodbank_profile.approval_date = None
        foodbank_profile.approved_by = None
        foodbank_profile.rejection_reason = previous_approval_info
        foodbank_profile.save()
        
        # Create notification for the foodbank
        Notification.objects.create(
            user=foodbank_profile.user,
            notification_type='info',
            message=f'Your food bank "{foodbank_profile.foodbank_name}" has been reopened for further review by our admin team. You will be notified once the review is complete.'
        )
        
        # Send notification email to the foodbank
        try:
            subject = f"Food Bank Reopened for Review - {foodbank_profile.foodbank_name}"
            message = f"""
Dear {foodbank_profile.contact_person},

Your food bank "{foodbank_profile.foodbank_name}" has been reopened for further administrative review.

REVIEW REASON:
{review_reason}

WHAT THIS MEANS:
Your food bank status has been temporarily changed to "Pending Review" while our admin team conducts additional verification or checks. This is a standard procedure and does not necessarily indicate any issues with your application.

WHAT HAPPENS NEXT:
- Our admin team will review your food bank profile
- You will receive an email notification once the review is complete
- Your food bank services may be temporarily affected during this review period

STATUS: Pending Review
FOOD BANK ID: {foodbank_profile.id}

If you have any questions or concerns, please contact our support team immediately.

Thank you for your understanding and cooperation.

Best regards,
FoodBankHub Admin Team

---
This is an automated email. Please do not reply directly to this message.
            """
            
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[foodbank_profile.user.email],
                fail_silently=False,
            )
            
            messages.success(
                request, 
                f'Food bank "{foodbank_profile.foodbank_name}" has been reopened for review and email notification sent.'
            )
        except Exception as e:
            messages.warning(
                request,
                f'Food bank reopened but email notification failed: {str(e)}'
            )
        
        return redirect('custom_admin:foodbank_approvals')
    
    context = {
        'title': 'Reopen Approved Food Bank for Review',
        'foodbank_profile': foodbank_profile,
        'action': 'reopen_approved'
    }
    return TemplateResponse(request, 'custom_admin/foodbank_action_confirm.html', context)


@staff_member_required
def support_messages(request):
    """Support messages management view"""
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    subject_filter = request.GET.get('subject', '')
    user_type_filter = request.GET.get('user_type', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    messages_queryset = SupportMessage.objects.select_related('user', 'assigned_to').all()

    from django.db.models import OuterRef, Subquery
    from authentication.models import SupportMessageReply
    last_admin_reply_email = SupportMessageReply.objects.filter(
        support_message=OuterRef('pk'),
        is_from_admin=True,
        author__isnull=False,
    ).order_by('-created_at').values('author__email')[:1]
    messages_queryset = messages_queryset.annotate(last_admin_reply_email=Subquery(last_admin_reply_email))
    
    # Apply filters
    if status_filter:
        messages_queryset = messages_queryset.filter(status=status_filter)
    
    if priority_filter:
        messages_queryset = messages_queryset.filter(priority=priority_filter)

    if subject_filter and subject_filter in dict(SupportMessage.SUBJECT_CHOICES):
        messages_queryset = messages_queryset.filter(subject=subject_filter)
    
    if user_type_filter:
        messages_queryset = messages_queryset.filter(user__user_type=user_type_filter)
    
    if search_query:
        # Also support searching by subject display label from the support form dropdown.
        normalized_search = search_query.strip().lower()
        matching_subject_values = [
            value
            for value, label in SupportMessage.SUBJECT_CHOICES
            if normalized_search in label.lower()
        ]

        search_filters = (
            Q(user__email__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(message__icontains=search_query)
        )
        if matching_subject_values:
            search_filters |= Q(subject__in=matching_subject_values)

        messages_queryset = messages_queryset.filter(search_filters)

    if date_from:
        from django.utils.dateparse import parse_date
        parsed_from = parse_date(date_from)
        if parsed_from:
            messages_queryset = messages_queryset.filter(created_at__date__gte=parsed_from)

    if date_to:
        from django.utils.dateparse import parse_date
        parsed_to = parse_date(date_to)
        if parsed_to:
            messages_queryset = messages_queryset.filter(created_at__date__lte=parsed_to)
    
    # Handle bulk actions
    if request.method == 'POST':
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_messages')
        
        if selected_ids:
            selected_messages = SupportMessage.objects.filter(id__in=selected_ids)
            
            if action == 'mark_in_progress':
                selected_messages.update(status='in_progress')
                messages.success(request, f'Marked {len(selected_ids)} messages as in progress.')
            
            elif action == 'mark_resolved':
                from django.utils import timezone
                selected_messages.update(status='resolved', resolved_at=timezone.now())
                messages.success(request, f'Marked {len(selected_ids)} messages as resolved.')
            
            elif action == 'assign_to_me':
                selected_messages.update(assigned_to=request.user)
                messages.success(request, f'Assigned {len(selected_ids)} messages to you.')
            
            elif action == 'set_priority_high':
                selected_messages.update(priority='high')
                messages.success(request, f'Set priority to high for {len(selected_ids)} messages.')
            
            elif action == 'set_priority_urgent':
                selected_messages.update(priority='urgent')
                messages.success(request, f'Set priority to urgent for {len(selected_ids)} messages.')
        
        return redirect('custom_admin:support_messages')
    
    active_queryset = messages_queryset.exclude(status__in=['resolved', 'closed'])
    resolved_queryset = messages_queryset.filter(status__in=['resolved', 'closed'])

    # Pagination
    active_paginator = Paginator(active_queryset, ITEMS_PER_PAGE)
    resolved_paginator = Paginator(resolved_queryset, ITEMS_PER_PAGE)

    active_page_number = request.GET.get('active_page')
    resolved_page_number = request.GET.get('resolved_page')

    active_messages_page = active_paginator.get_page(active_page_number)
    resolved_messages_page = resolved_paginator.get_page(resolved_page_number)
    
    # Statistics
    total_messages = SupportMessage.objects.count()
    new_messages = SupportMessage.objects.filter(status='new').count()
    in_progress_messages = SupportMessage.objects.filter(status='in_progress').count()
    resolved_messages = SupportMessage.objects.filter(status__in=['resolved', 'closed']).count()
    urgent_messages = SupportMessage.objects.filter(priority='urgent').count()
    
    context = {
        'title': 'Support Messages',
        'support_messages': active_messages_page,
        'resolved_support_messages': resolved_messages_page,
        'total_messages': total_messages,
        'new_messages': new_messages,
        'in_progress_messages': in_progress_messages,
        'resolved_messages': resolved_messages,
        'urgent_messages': urgent_messages,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'subject_filter': subject_filter,
        'user_type_filter': user_type_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': SupportMessage.STATUS_CHOICES,
        'priority_choices': SupportMessage.PRIORITY_CHOICES,
        'subject_choices': SupportMessage.SUBJECT_CHOICES,
        'user_type_choices': CustomUser.USER_TYPE_CHOICES,
    }
    
    return TemplateResponse(request, 'custom_admin/support_messages.html', context)


@staff_member_required
def support_message_detail(request, message_id):
    """Support message detail and response view"""
    support_message = get_object_or_404(SupportMessage, id=message_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            new_status = request.POST.get('status')
            if new_status in dict(SupportMessage.STATUS_CHOICES):
                support_message.status = new_status
                if new_status in ['resolved', 'closed']:
                    from django.utils import timezone
                    support_message.resolved_at = timezone.now()
                else:
                    support_message.resolved_at = None
                support_message.save()
                messages.success(request, f'Status updated to {support_message.get_status_display()}.')
        
        elif action == 'update_priority':
            new_priority = request.POST.get('priority')
            if new_priority in dict(SupportMessage.PRIORITY_CHOICES):
                support_message.priority = new_priority
                support_message.save()
                messages.success(request, f'Priority updated to {support_message.get_priority_display()}.')
        
        elif action == 'assign':
            assigned_to_id = request.POST.get('assigned_to')
            if assigned_to_id:
                try:
                    assigned_user = CustomUser.objects.get(id=assigned_to_id, is_staff=True)
                    support_message.assigned_to = assigned_user
                    support_message.save()
                    messages.success(request, f'Assigned to {assigned_user.email}.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Invalid staff member selected.')
        
        elif action == 'add_response':
            admin_response = request.POST.get('admin_response')
            if admin_response:
                if support_message.status in ['resolved', 'closed']:
                    messages.error(request, 'This message is already resolved. Reopen it to add a new response.')
                    return redirect('custom_admin:support_message_detail', message_id=message_id)
                SupportMessageReply.objects.create(
                    support_message=support_message,
                    author=request.user,
                    message=admin_response,
                    is_from_admin=True,
                )
                support_message.admin_response = admin_response
                support_message.status = 'in_progress'
                support_message.save()
                messages.success(request, 'Response added successfully.')
        
        return redirect('custom_admin:support_message_detail', message_id=message_id)
    
    # Get staff members for assignment
    staff_members = CustomUser.objects.filter(is_staff=True)
    
    replies = support_message.replies.select_related('author').all()

    context = {
        'title': f'Support Message #{support_message.id}',
        'support_message': support_message,
        'replies': replies,
        'staff_members': staff_members,
        'status_choices': SupportMessage.STATUS_CHOICES,
        'priority_choices': SupportMessage.PRIORITY_CHOICES,
    }
    
    return TemplateResponse(request, 'custom_admin/support_message_detail.html', context)


@staff_member_required
def subscription_management(request):
    """Subscription management dashboard for admin"""
    subscriptions = FoodBankSubscription.objects.select_related('foodbank', 'foodbank__user').all()
    
    # Filtering
    status_filter = request.GET.get('status')
    plan_filter = request.GET.get('plan')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if status_filter in {'suspended', 'cancelled'}:
        status_filter = None

    if status_filter:
        subscriptions = subscriptions.filter(status=status_filter)
    if plan_filter:
        subscriptions = subscriptions.filter(plan=plan_filter)
    if search:
        subscriptions = subscriptions.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search)
        )
    if date_from:
        subscriptions = subscriptions.filter(created_at__date__gte=date_from)
    if date_to:
        subscriptions = subscriptions.filter(created_at__date__lte=date_to)
    
    # Statistics
    total_subscriptions = FoodBankSubscription.objects.count()
    active_subscriptions = FoodBankSubscription.objects.filter(status='active').count()
    trial_subscriptions = FoodBankSubscription.objects.filter(status='trial').count()
    expired_subscriptions = FoodBankSubscription.objects.filter(status='expired').count()
    
    # Revenue calculations
    monthly_revenue = SubscriptionPayment.objects.filter(
        status='approved',
        plan_type='monthly'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    yearly_revenue = SubscriptionPayment.objects.filter(
        status='approved',
        plan_type='yearly'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_revenue = monthly_revenue + yearly_revenue
    
    # Pagination
    paginator = Paginator(subscriptions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    status_choices = [
        (value, label)
        for value, label in FoodBankSubscription.STATUS_CHOICES
        if value not in {'suspended', 'cancelled'}
    ]
    
    context = {
        'title': 'Subscription Management',
        'page_obj': page_obj,
        'total_subscriptions': total_subscriptions,
        'active_subscriptions': active_subscriptions,
        'trial_subscriptions': trial_subscriptions,
        'expired_subscriptions': expired_subscriptions,
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'yearly_revenue': yearly_revenue,
        'status_choices': status_choices,
        'plan_choices': FoodBankSubscription.PLAN_CHOICES,
        'current_filters': {
            'status': status_filter,
            'plan': plan_filter,
            'search': search,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    return render(request, 'custom_admin/subscription_management.html', context)


@staff_member_required
def export_subscriptions_excel(request):
    """Export subscriptions to Excel."""
    subscriptions = FoodBankSubscription.objects.select_related('foodbank', 'foodbank__user').all()

    status_filter = request.GET.get('status')
    plan_filter = request.GET.get('plan')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status_filter:
        subscriptions = subscriptions.filter(status=status_filter)
    if plan_filter:
        subscriptions = subscriptions.filter(plan=plan_filter)
    if search:
        subscriptions = subscriptions.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search)
        )
    if date_from:
        subscriptions = subscriptions.filter(created_at__date__gte=date_from)
    if date_to:
        subscriptions = subscriptions.filter(created_at__date__lte=date_to)

    subscriptions = subscriptions.order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "All Subscriptions"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'S/No', 'Food Bank Name', 'Email', 'Plan', 'Status',
        'Days Remaining', 'Trial Start Date', 'Trial End Date',
        'Subscription End Date', 'Created At'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_num, subscription in enumerate(subscriptions, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = border
        ws.cell(row=row_num, column=2, value=subscription.foodbank.foodbank_name).border = border
        ws.cell(row=row_num, column=3, value=subscription.foodbank.user.email).border = border
        ws.cell(row=row_num, column=4, value=subscription.get_plan_display()).border = border
        ws.cell(row=row_num, column=5, value=subscription.get_status_display()).border = border
        ws.cell(row=row_num, column=6, value=subscription.days_remaining()).border = border
        ws.cell(
            row=row_num,
            column=7,
            value=subscription.trial_start_date.strftime('%Y-%m-%d') if subscription.trial_start_date else 'N/A'
        ).border = border
        ws.cell(
            row=row_num,
            column=8,
            value=subscription.trial_end_date.strftime('%Y-%m-%d') if subscription.trial_end_date else 'N/A'
        ).border = border
        ws.cell(
            row=row_num,
            column=9,
            value=subscription.subscription_end_date.strftime('%Y-%m-%d') if subscription.subscription_end_date else 'N/A'
        ).border = border
        ws.cell(
            row=row_num,
            column=10,
            value=subscription.created_at.strftime('%Y-%m-%d %H:%M')
        ).border = border

    column_widths = [8, 28, 30, 22, 18, 16, 18, 18, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="all_subscriptions_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@staff_member_required
def export_subscriptions_csv(request):
    """Export subscriptions to CSV."""
    subscriptions = FoodBankSubscription.objects.select_related('foodbank', 'foodbank__user').all()

    status_filter = request.GET.get('status')
    plan_filter = request.GET.get('plan')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status_filter:
        subscriptions = subscriptions.filter(status=status_filter)
    if plan_filter:
        subscriptions = subscriptions.filter(plan=plan_filter)
    if search:
        subscriptions = subscriptions.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search)
        )
    if date_from:
        subscriptions = subscriptions.filter(created_at__date__gte=date_from)
    if date_to:
        subscriptions = subscriptions.filter(created_at__date__lte=date_to)

    subscriptions = subscriptions.order_by('-created_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="all_subscriptions_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        'S/No', 'Food Bank Name', 'Email', 'Plan', 'Status',
        'Days Remaining', 'Trial Start Date', 'Trial End Date',
        'Subscription End Date', 'Created At'
    ])

    for idx, subscription in enumerate(subscriptions, 1):
        writer.writerow([
            idx,
            subscription.foodbank.foodbank_name,
            subscription.foodbank.user.email,
            subscription.get_plan_display(),
            subscription.get_status_display(),
            subscription.days_remaining(),
            subscription.trial_start_date.strftime('%Y-%m-%d') if subscription.trial_start_date else 'N/A',
            subscription.trial_end_date.strftime('%Y-%m-%d') if subscription.trial_end_date else 'N/A',
            subscription.subscription_end_date.strftime('%Y-%m-%d') if subscription.subscription_end_date else 'N/A',
            subscription.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    return response


@staff_member_required
def export_subscriptions_pdf(request):
    """Export subscriptions to PDF."""
    subscriptions = FoodBankSubscription.objects.select_related('foodbank', 'foodbank__user').all()

    status_filter = request.GET.get('status')
    plan_filter = request.GET.get('plan')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status_filter:
        subscriptions = subscriptions.filter(status=status_filter)
    if plan_filter:
        subscriptions = subscriptions.filter(plan=plan_filter)
    if search:
        subscriptions = subscriptions.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search)
        )
    if date_from:
        subscriptions = subscriptions.filter(created_at__date__gte=date_from)
    if date_to:
        subscriptions = subscriptions.filter(created_at__date__lte=date_to)

    subscriptions = subscriptions.order_by('-created_at')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SubscriptionsTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=14,
        alignment=1
    )
    elements.append(Paragraph("All Subscriptions Report", title_style))
    elements.append(Paragraph(
        f"Generated on: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {subscriptions.count()}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 12))

    data = [[
        'S/No', 'Food Bank', 'Email', 'Plan', 'Status',
        'Days Remaining', 'Trial Start', 'Trial End', 'Subscription End', 'Created At'
    ]]

    for idx, subscription in enumerate(subscriptions, 1):
        data.append([
            str(idx),
            subscription.foodbank.foodbank_name[:28],
            subscription.foodbank.user.email[:34],
            subscription.get_plan_display(),
            subscription.get_status_display(),
            str(subscription.days_remaining()),
            subscription.trial_start_date.strftime('%Y-%m-%d') if subscription.trial_start_date else 'N/A',
            subscription.trial_end_date.strftime('%Y-%m-%d') if subscription.trial_end_date else 'N/A',
            subscription.subscription_end_date.strftime('%Y-%m-%d') if subscription.subscription_end_date else 'N/A',
            subscription.created_at.strftime('%Y-%m-%d'),
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[0.45 * inch, 1.7 * inch, 2.1 * inch, 1.4 * inch, 1.1 * inch, 0.9 * inch, 1.1 * inch, 1.1 * inch, 1.15 * inch, 0.95 * inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7.6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F8FB')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="all_subscriptions_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@staff_member_required
def payment_verification(request):
    """Payment verification queue for admin"""
    payments = SubscriptionPayment.objects.select_related(
        'foodbank', 'foodbank__user', 'subscription'
    ).all()
    
    # Filtering
    status_filter = request.GET.get('status', 'pending')  # Default to pending
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if date_from:
        payments = payments.filter(submitted_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(submitted_at__date__lte=date_to)
    
    # Statistics
    pending_count = SubscriptionPayment.objects.filter(status='pending').count()
    approved_count = SubscriptionPayment.objects.filter(status='approved').count()
    rejected_count = SubscriptionPayment.objects.filter(status='rejected').count()
    
    # Pagination
    paginator = Paginator(payments, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    status_choices = [
        (value, label)
        for value, label in SubscriptionPayment.STATUS_CHOICES
        if value != 'requires_info'
    ]

    context = {
        'title': 'Payment Verification',
        'page_obj': page_obj,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'status_choices': status_choices,
        'plan_choices': SubscriptionPayment.PLAN_CHOICES,
        'payment_method_choices': SubscriptionPayment.PAYMENT_METHOD_CHOICES,
        'current_filters': {
            'status': status_filter,
            'plan': plan_filter,
            'payment_method': payment_method_filter,
            'search': search,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    return render(request, 'custom_admin/payment_verification.html', context)


@staff_member_required
def export_payment_verification_excel(request):
    """Export payment verification records to Excel."""
    payments = SubscriptionPayment.objects.select_related(
        'foodbank', 'foodbank__user', 'subscription', 'verified_by'
    ).all()

    status_filter = request.GET.get('status', 'pending')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status_filter:
        payments = payments.filter(status=status_filter)
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if date_from:
        payments = payments.filter(submitted_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(submitted_at__date__lte=date_to)

    payments = payments.order_by('-submitted_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Verification"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'S/No', 'Food Bank Name', 'Email', 'Phone', 'Plan Type', 'Amount (KSH)',
        'Payment Method', 'Transaction Ref', 'Payment Date', 'Submitted At',
        'Message', 'Status', 'Verified By', 'Verified At', 'Rejection Reason'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = border
        ws.cell(row=row_num, column=2, value=payment.foodbank.foodbank_name).border = border
        ws.cell(row=row_num, column=3, value=payment.foodbank.user.email).border = border
        ws.cell(row=row_num, column=4, value=payment.foodbank.user.phone_number or 'N/A').border = border
        ws.cell(row=row_num, column=5, value=payment.get_plan_type_display()).border = border
        ws.cell(row=row_num, column=6, value=float(payment.amount)).border = border
        ws.cell(row=row_num, column=7, value=payment.get_payment_method_display()).border = border
        ws.cell(row=row_num, column=8, value=payment.transaction_reference).border = border
        ws.cell(row=row_num, column=9, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=10, value=payment.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
        ws.cell(row=row_num, column=11, value=payment.notes or '').border = border
        ws.cell(row=row_num, column=12, value=payment.get_status_display()).border = border
        ws.cell(
            row=row_num,
            column=13,
            value=payment.verified_by.email if payment.verified_by else 'N/A'
        ).border = border
        ws.cell(
            row=row_num,
            column=14,
            value=payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else 'N/A'
        ).border = border
        ws.cell(row=row_num, column=15, value=payment.rejection_reason or '').border = border

    column_widths = [8, 26, 30, 16, 20, 15, 18, 24, 14, 20, 30, 18, 26, 20, 36]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="payment_verification_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@staff_member_required
def export_payment_verification_csv(request):
    """Export payment verification records to CSV."""
    payments = SubscriptionPayment.objects.select_related(
        'foodbank', 'foodbank__user', 'subscription', 'verified_by'
    ).all()

    status_filter = request.GET.get('status', 'pending')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status_filter:
        payments = payments.filter(status=status_filter)
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if date_from:
        payments = payments.filter(submitted_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(submitted_at__date__lte=date_to)

    payments = payments.order_by('-submitted_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="payment_verification_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        'S/No', 'Food Bank Name', 'Email', 'Phone', 'Plan Type', 'Amount (KSH)',
        'Payment Method', 'Transaction Ref', 'Payment Date', 'Submitted At',
        'Message', 'Status', 'Verified By', 'Verified At', 'Rejection Reason'
    ])

    for idx, payment in enumerate(payments, 1):
        writer.writerow([
            idx,
            payment.foodbank.foodbank_name,
            payment.foodbank.user.email,
            payment.foodbank.user.phone_number or 'N/A',
            payment.get_plan_type_display(),
            f"{payment.amount:.2f}",
            payment.get_payment_method_display(),
            payment.transaction_reference,
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.submitted_at.strftime('%Y-%m-%d %H:%M'),
            payment.notes or '',
            payment.get_status_display(),
            payment.verified_by.email if payment.verified_by else 'N/A',
            payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else 'N/A',
            payment.rejection_reason or '',
        ])

    return response


@staff_member_required
def export_payment_verification_pdf(request):
    """Export payment verification records to PDF."""
    payments = SubscriptionPayment.objects.select_related(
        'foodbank', 'foodbank__user', 'subscription', 'verified_by'
    ).all()

    status_filter = request.GET.get('status', 'pending')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status_filter:
        payments = payments.filter(status=status_filter)
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if date_from:
        payments = payments.filter(submitted_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(submitted_at__date__lte=date_to)

    payments = payments.order_by('-submitted_at')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PaymentsTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0369A1'),
        spaceAfter=14,
        alignment=1
    )
    cell_wrap_style = ParagraphStyle(
        'PaymentsCellWrap',
        parent=styles['Normal'],
        fontSize=7.2,
        leading=8.2,
        alignment=1,
        wordWrap='CJK',
    )
    elements.append(Paragraph("Payment Verification Report", title_style))
    elements.append(Paragraph(
        f"Generated on: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {payments.count()}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 12))

    data = [[
        'S/No', 'Food Bank Name', 'Email', 'Phone', 'Plan Type', 'Amount (KSH)',
        'Payment Method', 'Transaction Ref', 'Payment Date', 'Submitted At', 'Message',
        'Status', 'Verified By', 'Verified At', 'Rejection Reason'
    ]]

    for idx, payment in enumerate(payments, 1):
        data.append([
            str(idx),
            payment.foodbank.foodbank_name[:24],
            payment.foodbank.user.email[:30],
            (payment.foodbank.user.phone_number or 'N/A')[:16],
            payment.get_plan_type_display()[:18],
            f"{payment.amount:,.2f}",
            payment.get_payment_method_display()[:16],
            payment.transaction_reference[:18],
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.submitted_at.strftime('%Y-%m-%d %H:%M'),
            Paragraph((payment.notes or '')[:45], cell_wrap_style),
            Paragraph(payment.get_status_display(), cell_wrap_style),
            payment.verified_by.email[:22] if payment.verified_by else 'N/A',
            payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else 'N/A',
            (payment.rejection_reason or '')[:34],
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            0.35 * inch, 1.0 * inch, 1.5 * inch, 0.8 * inch, 0.9 * inch,
            0.8 * inch, 0.95 * inch, 0.95 * inch, 0.75 * inch, 0.85 * inch,
            1.3 * inch, 0.9 * inch, 1.0 * inch, 0.85 * inch, 1.2 * inch
        ]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0369A1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7.4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="payment_verification_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@staff_member_required
def payment_detail(request, payment_id):
    """Payment detail and verification view"""
    payment = get_object_or_404(
        SubscriptionPayment.objects.select_related('foodbank', 'subscription'),
        id=payment_id
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            if payment.approve_payment(request.user):
                messages.success(
                    request,
                    f'Payment approved! {payment.foodbank.foodbank_name} subscription has been activated.'
                )
                return redirect('custom_admin:payment_verification')
            else:
                messages.error(
                    request,
                    f'Failed to approve payment. Could not extend subscription. Please check the subscription details.'
                )
        
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason')
            if rejection_reason:
                payment.reject_payment(request.user, rejection_reason)
                messages.success(request, 'Payment rejected and food bank has been notified.')
                return redirect('custom_admin:payment_verification')
            else:
                messages.error(request, 'Please provide a rejection reason.')
        
        elif action == 'add_note':
            admin_notes = request.POST.get('admin_notes')
            if admin_notes:
                payment.admin_notes = admin_notes
                payment.save()
                messages.success(request, 'Admin notes updated.')
    
    context = {
        'title': f'Payment Verification - {payment.foodbank.foodbank_name}',
        'payment': payment,
    }
    
    return render(request, 'custom_admin/payment_detail.html', context)


@staff_member_required
def subscription_detail(request, subscription_id):
    """Subscription detail view with management options"""
    subscription = get_object_or_404(
        FoodBankSubscription.objects.select_related('foodbank', 'foodbank__user'),
        id=subscription_id
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'suspend':
            subscription.status = 'suspended'
            subscription.save()
            messages.success(request, f'{subscription.foodbank.foodbank_name} subscription suspended.')
        
        elif action == 'activate':
            subscription.status = 'active'
            subscription.save()
            messages.success(request, f'{subscription.foodbank.foodbank_name} subscription activated.')
        
        elif action == 'extend_trial':
            try:
                days = int(request.POST.get('days', 30))
            except (TypeError, ValueError):
                days = 30
            days = max(days, 1)

            if not subscription.trial_end_date:
                base_date = subscription.trial_start_date or timezone.now()
                subscription.trial_end_date = base_date + timedelta(days=90)

            subscription.trial_end_date = subscription.trial_end_date + timedelta(days=days)
            subscription.save()
            messages.success(request, f'Trial extended by {days} days.')

        elif action == 'reduce_trial':
            try:
                days = int(request.POST.get('days', 30))
            except (TypeError, ValueError):
                days = 30
            days = max(days, 1)

            if not subscription.trial_end_date:
                base_date = subscription.trial_start_date or timezone.now()
                subscription.trial_end_date = base_date + timedelta(days=90)

            min_trial_end = subscription.trial_start_date or timezone.now()
            proposed_end = subscription.trial_end_date - timedelta(days=days)

            if proposed_end < min_trial_end:
                actual_reduction = max((subscription.trial_end_date - min_trial_end).days, 0)
                subscription.trial_end_date = min_trial_end
                subscription.save()
                messages.warning(
                    request,
                    f'Trial reduced by {actual_reduction} days (minimum end date reached).'
                )
            else:
                subscription.trial_end_date = proposed_end
                subscription.save()
                messages.success(request, f'Trial reduced by {days} days.')
        
        elif action == 'cancel':
            subscription.status = 'cancelled'
            subscription.save()
            messages.success(request, f'{subscription.foodbank.foodbank_name} subscription cancelled.')
        
        return redirect('custom_admin:subscription_detail', subscription_id=subscription_id)
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Get payment history
    payment_history = SubscriptionPayment.objects.filter(
        subscription=subscription
    ).order_by('-submitted_at')

    if date_from:
        payment_history = payment_history.filter(payment_date__gte=date_from)
    if date_to:
        payment_history = payment_history.filter(payment_date__lte=date_to)
    
    context = {
        'title': f'Subscription - {subscription.foodbank.foodbank_name}',
        'subscription': subscription,
        'payment_history': payment_history,
        'current_filters': {
            'date_from': date_from,
            'date_to': date_to,
        },
    }
    
    return render(request, 'custom_admin/subscription_detail.html', context)


@staff_member_required
def expired_accounts(request):
    """List accounts whose trial/subscription has elapsed"""
    now = timezone.now()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    trial_expired = Q(status='trial') & (
        (Q(trial_end_date__isnull=False) & Q(trial_end_date__lt=now)) |
        (Q(trial_end_date__isnull=True) & Q(trial_start_date__isnull=False) & Q(trial_start_date__lt=now - timedelta(days=90)))
    )

    subscription_expired = Q(status__in=['active', 'suspended', 'cancelled']) & (
        (Q(subscription_end_date__isnull=False) & Q(subscription_end_date__lt=now)) |
        (Q(subscription_end_date__isnull=True) & Q(subscription_start_date__isnull=False) & Q(plan='monthly') & Q(subscription_start_date__lt=now - timedelta(days=30))) |
        (Q(subscription_end_date__isnull=True) & Q(subscription_start_date__isnull=False) & Q(plan='yearly') & Q(subscription_start_date__lt=now - timedelta(days=365)))
    )

    expired_qs = FoodBankSubscription.objects.select_related('foodbank', 'foodbank__user').filter(
        Q(status='expired') |
        trial_expired |
        subscription_expired
    ).order_by('-updated_at', '-created_at')

    if date_from:
        expired_qs = expired_qs.filter(updated_at__date__gte=date_from)
    if date_to:
        expired_qs = expired_qs.filter(updated_at__date__lte=date_to)

    paginator = Paginator(expired_qs, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'title': 'Expired Accounts',
        'page_obj': page_obj,
        'current_filters': {
            'date_from': date_from,
            'date_to': date_to,
        },
    }

    return render(request, 'custom_admin/expired_accounts.html', context)


@staff_member_required
def testimonials_overview(request):
    recipient_approved_qs = Testimonial.objects.filter(
        approval_status='approved',
        display_on_public=True,
    ).select_related('recipient', 'recipient__user').order_by('-created_at')
    recipient_displayed = [t for t in recipient_approved_qs if t.is_currently_displayed()]

    donor_approved_qs = DonorTestimonial.objects.filter(
        approval_status='approved',
        display_on_public=True,
    ).select_related('donor', 'donor__user').order_by('-created_at')
    donor_displayed = [t for t in donor_approved_qs if t.is_currently_displayed()]

    foodbank_approved_qs = FoodbankTestimonial.objects.filter(
        approval_status='approved',
        display_on_public=True,
    ).select_related('foodbank', 'foodbank__user').order_by('-created_at')
    foodbank_displayed = [t for t in foodbank_approved_qs if t.is_currently_displayed()]

    context = {
        'title': 'Testimonials',
        'recipient_displayed_count': len(recipient_displayed),
        'donor_displayed_count': len(donor_displayed),
        'foodbank_displayed_count': len(foodbank_displayed),
        'recipient_pending_count': Testimonial.objects.filter(approval_status='pending').count(),
        'donor_pending_count': DonorTestimonial.objects.filter(approval_status='pending').count(),
        'foodbank_pending_count': FoodbankTestimonial.objects.filter(approval_status='pending').count(),
        'recipient_displayed_preview': recipient_displayed[:3],
        'donor_displayed_preview': donor_displayed[:3],
        'foodbank_displayed_preview': foodbank_displayed[:3],
    }
    return render(request, 'custom_admin/testimonials_overview.html', context)


def _testimonial_category_config(category):
    if category == 'recipient':
        return {
            'model': Testimonial,
            'select_related': ('recipient', 'recipient__user', 'reviewed_by'),
        }
    if category == 'donor':
        return {
            'model': DonorTestimonial,
            'select_related': ('donor', 'donor__user', 'reviewed_by'),
        }
    if category == 'foodbank':
        return {
            'model': FoodbankTestimonial,
            'select_related': ('foodbank', 'foodbank__user', 'reviewed_by'),
        }
    return None


def _apply_testimonial_filters(qs, category, search, featured, has_image, date_from, date_to):
    if search:
        if category == 'recipient':
            qs = qs.filter(
                Q(recipient__full_name__icontains=search) |
                Q(recipient__organization_name__icontains=search) |
                Q(recipient__user__email__icontains=search) |
                Q(message__icontains=search)
            )
        elif category == 'donor':
            qs = qs.filter(
                Q(donor__full_name__icontains=search) |
                Q(donor__organization_name__icontains=search) |
                Q(donor__user__email__icontains=search) |
                Q(message__icontains=search)
            )
        else:
            qs = qs.filter(
                Q(foodbank__foodbank_name__icontains=search) |
                Q(foodbank__user__email__icontains=search) |
                Q(message__icontains=search)
            )

    if featured == 'featured':
        qs = qs.filter(is_featured=True)
    elif featured == 'not_featured':
        qs = qs.filter(is_featured=False)

    if has_image == 'yes':
        qs = qs.filter(impact_image__isnull=False)
    elif has_image == 'no':
        qs = qs.filter(impact_image__isnull=True)

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    return qs


def _testimonial_status_template(category, status):
    if category == 'recipient':
        if status == 'displayed':
            return 'custom_admin/testimonials_recipient_displayed.html'
        if status == 'pending':
            return 'custom_admin/testimonials_recipient_pending.html'
        if status == 'rejected':
            return 'custom_admin/testimonials_recipient_rejected.html'
        return 'custom_admin/testimonials_recipient_archived.html'

    if category == 'donor':
        if status == 'displayed':
            return 'custom_admin/testimonials_donor_displayed.html'
        if status == 'pending':
            return 'custom_admin/testimonials_donor_pending.html'
        if status == 'rejected':
            return 'custom_admin/testimonials_donor_rejected.html'
        return 'custom_admin/testimonials_donor_archived.html'

    if status == 'displayed':
        return 'custom_admin/testimonials_foodbank_displayed.html'
    if status == 'pending':
        return 'custom_admin/testimonials_foodbank_pending.html'
    if status == 'rejected':
        return 'custom_admin/testimonials_foodbank_rejected.html'
    return 'custom_admin/testimonials_foodbank_archived.html'


@staff_member_required
def testimonials_category(request, category):
    return redirect('custom_admin:testimonials_category_status', category=category, status='displayed')


@staff_member_required
def testimonials_category_status(request, category, status):
    config = _testimonial_category_config(category)
    if not config:
        return redirect('custom_admin:testimonials_overview')

    if status not in ['pending', 'displayed', 'archived', 'rejected']:
        return redirect('custom_admin:testimonials_category_status', category=category, status='displayed')

    model = config['model']
    template_name = _testimonial_status_template(category, status)

    search = (request.GET.get('search') or '').strip()
    featured = (request.GET.get('featured') or '').strip()
    has_image = (request.GET.get('has_image') or '').strip()
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()

    base_qs = model.objects.select_related(*config['select_related']).order_by('-created_at')
    base_qs = _apply_testimonial_filters(base_qs, category, search, featured, has_image, date_from, date_to)

    displayed_qs = base_qs.filter(approval_status='approved', display_on_public=True)
    displayed_all = [t for t in displayed_qs if t.is_currently_displayed()]
    displayed_ids = [t.id for t in displayed_all]

    pending_qs = base_qs.filter(approval_status='pending')
    rejected_qs = base_qs.filter(approval_status='rejected')
    archived_qs = base_qs.filter(approval_status='approved').exclude(id__in=displayed_ids)

    if status == 'displayed':
        listing = displayed_all
        total_count = len(displayed_all)
    elif status == 'pending':
        listing = pending_qs
        total_count = pending_qs.count()
    elif status == 'rejected':
        listing = rejected_qs
        total_count = rejected_qs.count()
    else:
        listing = archived_qs
        total_count = archived_qs.count()

    paginator = Paginator(listing, ITEMS_PER_PAGE)
    page_obj = paginator.get_page(request.GET.get('page'))

    current_filters = {
        'search': search,
        'featured': featured,
        'has_image': has_image,
        'date_from': date_from,
        'date_to': date_to,
    }
    filter_query = urlencode({k: v for k, v in current_filters.items() if v})

    context = {
        'title': f"{category.title()} Testimonials",
        'category': category,
        'status': status,
        'page_obj': page_obj,
        'displayed_count': len(displayed_all),
        'pending_count': pending_qs.count(),
        'archived_count': archived_qs.count(),
        'rejected_count': rejected_qs.count(),
        'current_filters': current_filters,
        'filter_query': filter_query,
        'total_count': total_count,
    }
    return render(request, template_name, context)


@staff_member_required
def approve_testimonial(request, category, testimonial_id):
    config = _testimonial_category_config(category)
    if not config:
        return redirect('custom_admin:testimonials_overview')
    if request.method != 'POST':
        return redirect('custom_admin:testimonials_category', category=category)

    model = config['model']
    testimonial = get_object_or_404(model, id=testimonial_id)
    testimonial.approval_status = 'approved'
    # Ensure re-approved archived testimonials return to public display.
    testimonial.display_on_public = True
    testimonial.reviewed_by = request.user
    testimonial.reviewed_at = timezone.now()
    testimonial.set_default_display_period()
    testimonial.save()

    messages.success(request, 'Testimonial approved.')
    return redirect('custom_admin:testimonials_category_status', category=category, status='displayed')


@staff_member_required
def reject_testimonial(request, category, testimonial_id):
    config = _testimonial_category_config(category)
    if not config:
        return redirect('custom_admin:testimonials_overview')

    model = config['model']
    testimonial = get_object_or_404(model, id=testimonial_id)

    if request.method == 'POST':
        testimonial.approval_status = 'rejected'
        testimonial.reviewed_by = request.user
        testimonial.reviewed_at = timezone.now()
        testimonial.rejection_reason = request.POST.get('rejection_reason', '').strip()
        testimonial.save()
        messages.success(request, 'Testimonial rejected.')
        return redirect('custom_admin:testimonials_category_status', category=category, status='rejected')

    context = {
        'title': 'Reject Testimonial',
        'category': category,
        'testimonial': testimonial,
    }
    return render(request, 'custom_admin/testimonial_reject.html', context)


@staff_member_required
def hide_testimonial(request, category, testimonial_id):
    config = _testimonial_category_config(category)
    if not config:
        return redirect('custom_admin:testimonials_overview')
    if request.method != 'POST':
        return redirect('custom_admin:testimonials_category', category=category)

    model = config['model']
    testimonial = get_object_or_404(model, id=testimonial_id)

    if getattr(testimonial, 'approval_status', None) != 'approved':
        messages.error(request, 'Only approved testimonials can be removed from display.')
        return redirect('custom_admin:testimonials_category', category=category)

    testimonial.display_on_public = False
    testimonial.save()
    messages.success(request, 'Testimonial removed from public display.')
    return redirect('custom_admin:testimonials_category_status', category=category, status='archived')


@staff_member_required
def restore_testimonial(request, category, testimonial_id):
    config = _testimonial_category_config(category)
    if not config:
        return redirect('custom_admin:testimonials_overview')
    if request.method != 'POST':
        return redirect('custom_admin:testimonials_category', category=category)

    model = config['model']
    testimonial = get_object_or_404(model, id=testimonial_id)

    if getattr(testimonial, 'approval_status', None) != 'approved':
        messages.error(request, 'Only archived approved testimonials can be restored to pending review.')
        return redirect('custom_admin:testimonials_category_status', category=category, status='archived')

    testimonial.approval_status = 'pending'
    testimonial.display_on_public = False
    testimonial.reviewed_by = None
    testimonial.reviewed_at = None
    testimonial.rejection_reason = None
    testimonial.display_start_date = None
    testimonial.display_end_date = None
    testimonial.save()

    messages.success(request, 'Testimonial restored to pending review.')
    return redirect('custom_admin:testimonials_category_status', category=category, status='pending')


@staff_member_required
def download_testimonial_pdf(request, category, testimonial_id):
    config = _testimonial_category_config(category)
    if not config:
        return redirect('custom_admin:testimonials_overview')

    model = config['model']
    testimonial = get_object_or_404(
        model.objects.select_related(*config['select_related']),
        id=testimonial_id,
    )

    if category == 'recipient':
        owner_name = testimonial.recipient.organization_name if getattr(testimonial.recipient, 'is_organization', False) else testimonial.recipient.full_name
        owner_email = testimonial.recipient.user.email
        owner_label = 'Recipient'
    elif category == 'donor':
        owner_name = testimonial.donor.organization_name if getattr(testimonial.donor, 'is_organization', False) else testimonial.donor.full_name
        owner_email = testimonial.donor.user.email
        owner_label = 'Donor'
    else:
        owner_name = getattr(testimonial.foodbank, 'foodbank_name', None) or testimonial.foodbank.user.email
        owner_email = testimonial.foodbank.user.email
        owner_label = 'Foodbank'

    if testimonial.approval_status == 'pending':
        display_state = 'Pending'
    elif testimonial.approval_status == 'approved' and getattr(testimonial, 'display_on_public', False):
        display_state = 'Displayed'
    elif testimonial.approval_status == 'approved' and not getattr(testimonial, 'display_on_public', True):
        display_state = 'Archived'
    else:
        display_state = testimonial.get_approval_status_display() if hasattr(testimonial, 'get_approval_status_display') else str(testimonial.approval_status)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']
    label_style = ParagraphStyle(
        'LabelStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#333333'),
    )

    story = []
    story.append(Paragraph('Testimonial Details', title_style))
    story.append(Spacer(1, 10))

    meta_data = [
        [Paragraph('<b>Category</b>', label_style), Paragraph(category.title(), normal_style)],
        [Paragraph(f'<b>{owner_label}</b>', label_style), Paragraph(owner_name, normal_style)],
        [Paragraph('<b>Email</b>', label_style), Paragraph(owner_email, normal_style)],
        [Paragraph('<b>Status</b>', label_style), Paragraph(display_state, normal_style)],
        [Paragraph('<b>Featured</b>', label_style), Paragraph('Yes' if getattr(testimonial, 'is_featured', False) else 'No', normal_style)],
        [Paragraph('<b>Created</b>', label_style), Paragraph(testimonial.created_at.strftime('%Y-%m-%d %H:%M'), normal_style)],
    ]

    if getattr(testimonial, 'reviewed_by', None):
        meta_data.append([Paragraph('<b>Reviewed By</b>', label_style), Paragraph(getattr(testimonial.reviewed_by, 'email', ''), normal_style)])
    if getattr(testimonial, 'reviewed_at', None):
        meta_data.append([Paragraph('<b>Reviewed At</b>', label_style), Paragraph(testimonial.reviewed_at.strftime('%Y-%m-%d %H:%M'), normal_style)])

    meta_table = Table(meta_data, colWidths=[110, 390])
    meta_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 12))

    impact_image = getattr(testimonial, 'impact_image', None)
    if impact_image:
        image_path = getattr(impact_image, 'path', None)
        if image_path:
            try:
                story.append(Paragraph('<b>Impact Image</b>', label_style))
                story.append(Spacer(1, 6))
                img = Image(image_path)
                img.drawWidth = 5.5 * inch
                img.drawHeight = 3.1 * inch
                story.append(img)
                story.append(Spacer(1, 12))
            except Exception:
                pass

    story.append(Paragraph('<b>Message</b>', label_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph((testimonial.message or '').replace('\n', '<br/>'), normal_style))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="testimonial_{category}_{testimonial_id}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


@staff_member_required
def admin_login_logs(request):
    """Display admin login logs with filtering and search"""
    # Get filter parameters
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    user_filter = request.GET.get('user_filter')
    
    # Base queryset - only admin users
    logs = AdminLoginLog.objects.select_related('user').filter(
        user__user_type='ADMIN'
    ).order_by('-login_time')
    
    # Apply search filter
    if search:
        logs = logs.filter(
            Q(user__email__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(user_agent__icontains=search)
        )
    
    # Apply date filters
    if date_from:
        logs = logs.filter(login_time__date__gte=date_from)
    if date_to:
        logs = logs.filter(login_time__date__lte=date_to)
    
    # Apply user filter
    if user_filter:
        logs = logs.filter(user_id=user_filter)
    
    # Get all admin users for filter dropdown
    admin_users = CustomUser.objects.filter(user_type='ADMIN').order_by('email')
    
    # Pagination
    paginator = Paginator(logs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_logs = logs.count()
    unique_admins = logs.values('user').distinct().count()
    today_logs = logs.filter(login_time__date=timezone.now().date()).count()
    
    # Recent activity (last 7 days)
    week_ago = timezone.now() - timedelta(days=7)
    recent_logs = logs.filter(login_time__gte=week_ago).count()
    
    context = {
        'title': 'Admin Login Logs',
        'page_obj': page_obj,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'user_filter': user_filter,
        'admin_users': admin_users,
        'total_logs': total_logs,
        'unique_admins': unique_admins,
        'today_logs': today_logs,
        'recent_logs': recent_logs,
    }
    
    return render(request, 'custom_admin/admin_login_logs.html', context)


@superuser_required
def admin_codes_management(request):
    """Manage admin registration codes (Superuser only)"""
    # Handle POST requests for adding/updating codes
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_code':
            code = request.POST.get('code', '').strip()
            description = request.POST.get('description', '').strip()
            
            if code:
                # Check if maximum limit reached
                active_codes_count = AdminCode.objects.filter(is_active=True).count()
                if active_codes_count >= 5:
                    messages.error(request, 'Maximum limit of 5 active admin codes reached.')
                else:
                    try:
                        AdminCode.objects.create(
                            code=code,
                            description=description,
                            created_by=request.user
                        )
                        messages.success(request, f'Admin code "{code}" created successfully.')
                    except Exception as e:
                        messages.error(request, f'Error creating code: {str(e)}')
            else:
                messages.error(request, 'Code cannot be empty.')
        
        elif action == 'toggle_status':
            code_id = request.POST.get('code_id')
            try:
                admin_code = AdminCode.objects.get(id=code_id)
                admin_code.is_active = not admin_code.is_active
                admin_code.save()
                status = 'activated' if admin_code.is_active else 'deactivated'
                messages.success(request, f'Admin code "{admin_code.code}" {status} successfully.')
            except AdminCode.DoesNotExist:
                messages.error(request, 'Admin code not found.')
        
        elif action == 'delete_code':
            code_id = request.POST.get('code_id')
            try:
                admin_code = AdminCode.objects.get(id=code_id)
                code_name = admin_code.code
                admin_code.delete()
                messages.success(request, f'Admin code "{code_name}" deleted successfully.')
            except AdminCode.DoesNotExist:
                messages.error(request, 'Admin code not found.')
        
        return redirect('custom_admin:admin_codes_management')
    
    search_query = request.GET.get('search', '').strip()

    # Get all admin codes
    all_admin_codes = AdminCode.objects.all()
    admin_codes = all_admin_codes.select_related('created_by', 'last_used_by')
    if search_query:
        admin_codes = admin_codes.filter(
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(created_by__email__icontains=search_query) |
            Q(last_used_by__email__icontains=search_query)
        )
    admin_codes = admin_codes.order_by('-created_at')
    
    # Statistics
    total_codes = all_admin_codes.count()
    active_codes = all_admin_codes.filter(is_active=True).count()
    used_codes = all_admin_codes.filter(used_count__gt=0).count()
    unused_codes = all_admin_codes.filter(used_count=0).count()
    
    # Check if we can add more codes
    can_add_more = active_codes < 5
    
    context = {
        'title': 'Admin Registration Codes',
        'admin_codes': admin_codes,
        'total_codes': total_codes,
        'active_codes': active_codes,
        'used_codes': used_codes,
        'unused_codes': unused_codes,
        'can_add_more': can_add_more,
        'max_codes': 5,
        'search_query': search_query,
    }
    
    return render(request, 'custom_admin/admin_codes_management.html', context)


def _get_admin_codes_queryset(search_query=''):
    """Return admin codes queryset with optional search filtering."""
    admin_codes = AdminCode.objects.select_related('created_by', 'last_used_by')
    if search_query:
        admin_codes = admin_codes.filter(
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(created_by__email__icontains=search_query) |
            Q(last_used_by__email__icontains=search_query)
        )
    return admin_codes.order_by('-created_at')


@superuser_required
def export_admin_codes_excel(request):
    """Export admin registration codes to Excel."""
    search_query = request.GET.get('search', '').strip()
    admin_codes = _get_admin_codes_queryset(search_query)

    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Codes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4e73df", end_color="4e73df", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'S/No', 'Code', 'Description', 'Status', 'Usage',
        'Created By', 'Created Date', 'Last Used'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_num, code in enumerate(admin_codes, 2):
        last_used_value = 'Never used'
        if code.last_used_at:
            last_used_by = code.last_used_by.email if code.last_used_by else 'Unknown'
            last_used_value = f"{code.last_used_at.strftime('%Y-%m-%d %H:%M:%S')} by {last_used_by}"

        row_values = [
            row_num - 1,
            code.code,
            code.description or 'No description',
            'Active' if code.is_active else 'Inactive',
            f"{code.used_count} time{'s' if code.used_count != 1 else ''}" if code.used_count > 0 else 'Unused',
            code.created_by.email if code.created_by else 'Unknown',
            code.created_at.strftime('%Y-%m-%d %H:%M:%S') if code.created_at else 'N/A',
            last_used_value,
        ]

        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [1, 4, 5]:
                cell.alignment = Alignment(horizontal="center")

    column_widths = [8, 22, 35, 12, 14, 30, 22, 35]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="admin_codes_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@superuser_required
def export_admin_codes_pdf(request):
    """Export admin registration codes to PDF."""
    search_query = request.GET.get('search', '').strip()
    admin_codes = _get_admin_codes_queryset(search_query)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.4 * inch,
        rightMargin=0.4 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'AdminCodesTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#4e73df'),
        alignment=1,
        spaceAfter=10,
    )

    elements.append(Paragraph("Admin Registration Codes Report", title_style))
    report_info = (
        f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        f"Total Records: {admin_codes.count()}"
    )
    if search_query:
        report_info += f"<br/>Search Filter: {search_query}"
    elements.append(Paragraph(report_info, styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [[
        'S/No', 'Code', 'Description', 'Status', 'Usage',
        'Created By', 'Created Date', 'Last Used'
    ]]

    for idx, code in enumerate(admin_codes, 1):
        last_used_value = 'Never used'
        if code.last_used_at:
            last_used_value = code.last_used_at.strftime('%Y-%m-%d %H:%M')

        data.append([
            str(idx),
            code.code,
            (code.description or 'No description')[:45],
            'Active' if code.is_active else 'Inactive',
            f"{code.used_count}x" if code.used_count > 0 else 'Unused',
            (code.created_by.email if code.created_by else 'Unknown')[:30],
            code.created_at.strftime('%Y-%m-%d %H:%M') if code.created_at else 'N/A',
            last_used_value,
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[0.55 * inch, 1.2 * inch, 2.0 * inch, 0.9 * inch, 0.8 * inch, 2.0 * inch, 1.2 * inch, 1.2 * inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4e73df')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),
        ('ALIGN', (5, 1), (5, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fc')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="admin_codes_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


def _filtered_system_support_donations_queryset(request):
    status_filter = (request.GET.get('status') or '').strip()
    search_query = (request.GET.get('search') or '').strip()
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()

    donations = SystemSupportDonation.objects.select_related(
        'donor', 'donor__donor_profile', 'verified_by'
    ).all()

    if status_filter:
        donations = donations.filter(status=status_filter)

    if date_from:
        from django.utils.dateparse import parse_date
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(created_at__date__gte=parsed_from)

    if date_to:
        from django.utils.dateparse import parse_date
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(created_at__date__lte=parsed_to)

    if search_query:
        donations = donations.filter(
            Q(donor__email__icontains=search_query) |
            Q(transaction_reference__icontains=search_query) |
            Q(amount__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    return donations.order_by('-created_at'), status_filter, search_query, date_from, date_to


@staff_member_required
def system_support_donations(request):
    """View and manage system support donations"""
    donations, status_filter, search_query, date_from, date_to = _filtered_system_support_donations_queryset(request)

    paginator = Paginator(donations, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_donations = SystemSupportDonation.objects.count()
    pending_count = SystemSupportDonation.objects.filter(status='pending').count()
    approved_count = SystemSupportDonation.objects.filter(status='approved').count()
    rejected_count = SystemSupportDonation.objects.filter(status='rejected').count()
    total_amount = SystemSupportDonation.objects.filter(status='approved').aggregate(
        total=Sum('amount')
    )['total'] or 0

    context = {
        'title': 'System Support Donations',
        'page_obj': page_obj,
        'donations': page_obj,
        'status_filter': status_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'total_donations': total_donations,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'total_amount': total_amount,
    }
    return render(request, 'custom_admin/system_support_donations.html', context)


@staff_member_required
def export_system_support_donations_excel(request):
    """Export system support donations to Excel."""
    donations, _, _, _, _ = _filtered_system_support_donations_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "System Support Donations"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'ID',
        'Donor',
        'Amount (KES)',
        'Transaction Ref',
        'Donor Note',
        'Status',
        'Submitted At',
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_num, donation in enumerate(donations, 2):
        donor_name = ''
        donor_profile = getattr(donation.donor, 'donor_profile', None)
        if donor_profile and donor_profile.full_name:
            donor_name = f"{donation.donor.email} ({donor_profile.full_name})"
        else:
            donor_name = donation.donor.email

        values = [
            donation.id,
            donor_name,
            float(donation.amount or 0),
            donation.transaction_reference or 'N/A',
            donation.notes or '',
            donation.get_status_display() if hasattr(donation, 'get_status_display') else donation.status,
            timezone.localtime(donation.created_at).strftime('%Y-%m-%d %H:%M') if donation.created_at else 'N/A',
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [2, 4, 5]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    column_widths = [10, 38, 16, 22, 46, 16, 22]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="system_support_donations_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@staff_member_required
def export_system_support_donations_pdf(request):
    """Export system support donations to PDF."""
    import html

    donations, _, _, _, _ = _filtered_system_support_donations_queryset(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.3 * inch,
        rightMargin=0.3 * inch,
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SupportDonationsTitle',
        parent=styles['Heading1'],
        fontSize=15,
        textColor=colors.HexColor('#0369A1'),
        spaceAfter=12,
        alignment=1
    )
    wrap_style = ParagraphStyle(
        'SupportDonationsWrap',
        parent=styles['Normal'],
        fontSize=7.2,
        leading=8.4,
        alignment=0,
        wordWrap='CJK'
    )

    elements.append(Paragraph("System Support Donations Report", title_style))
    elements.append(Paragraph(
        f"Generated on: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {donations.count()}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 10))

    data = [[
        'ID',
        'Donor',
        'Amount (KES)',
        'Transaction Ref',
        'Donor Note',
        'Status',
        'Submitted At',
    ]]

    for donation in donations:
        donor_profile = getattr(donation.donor, 'donor_profile', None)
        if donor_profile and donor_profile.full_name:
            donor_display = f"{donation.donor.email} ({donor_profile.full_name})"
        else:
            donor_display = donation.donor.email

        note_text = html.escape((donation.notes or '').strip())
        if len(note_text) > 220:
            note_text = f"{note_text[:220]}..."

        data.append([
            str(donation.id),
            Paragraph(html.escape(donor_display), wrap_style),
            f"{float(donation.amount or 0):,.2f}",
            Paragraph(html.escape((donation.transaction_reference or 'N/A')[:60]), wrap_style),
            Paragraph(note_text or 'N/A', wrap_style),
            donation.get_status_display() if hasattr(donation, 'get_status_display') else (donation.status or 'N/A'),
            timezone.localtime(donation.created_at).strftime('%Y-%m-%d %H:%M') if donation.created_at else 'N/A',
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[0.55 * inch, 2.05 * inch, 1.2 * inch, 1.35 * inch, 3.4 * inch, 1.0 * inch, 1.4 * inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0369A1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.2),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7.2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="system_support_donations_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@staff_member_required
def system_support_donation_detail(request, donation_id):
    """View and verify a specific system support donation"""
    donation = get_object_or_404(SystemSupportDonation, id=donation_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            notes = request.POST.get('admin_notes', '').strip()
            donation.approve(request.user, notes)
            
            # Notify donor
            Notification.objects.create(
                user=donation.donor,
                notification_type='acknowledgement',
                message=f'Your system support donation of KES {donation.amount} has been approved! Thank you for your support.'
            )
            
            messages.success(request, f'Donation of KES {donation.amount} approved successfully!')
            return redirect('custom_admin:system_support_donations')
            
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '').strip()
            if not reason:
                messages.error(request, 'Please provide a reason for rejection.')
            else:
                donation.reject(request.user, reason)
                
                # Notify donor
                Notification.objects.create(
                    user=donation.donor,
                    notification_type='system',
                    message=f'Your system support donation of KES {donation.amount} could not be verified. Reason: {reason}'
                )
                
                messages.warning(request, f'Donation of KES {donation.amount} rejected.')
                return redirect('custom_admin:system_support_donations')
    
    context = {
        'title': 'System Support Donation Details',
        'donation': donation,
    }
    
    return render(request, 'custom_admin/system_support_donation_detail.html', context)


@staff_member_required
def news_section_list(request):
    """List all news sections"""
    from authentication.models import NewsSection
    
    news_sections = NewsSection.objects.all()
    
    context = {
        'title': 'Manage News Sections',
        'news_sections': news_sections,
    }
    
    return render(request, 'custom_admin/news_section_list.html', context)


@staff_member_required
def news_section_create(request):
    """Create a new news section"""
    from authentication.models import NewsSection
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        image = request.FILES.get('image')
        is_active = request.POST.get('is_active') == 'on'
        display_order = request.POST.get('display_order', 0)
        
        if not title or not content:
            messages.error(request, 'Title and content are required.')
        elif not image:
            messages.error(request, 'Please upload an image.')
        else:
            news_section = NewsSection.objects.create(
                title=title,
                content=content,
                image=image,
                is_active=is_active,
                display_order=display_order,
                created_by=request.user
            )
            messages.success(request, f'News section "{title}" created successfully!')
            return redirect('custom_admin:news_section_list')
    
    context = {
        'title': 'Create News Section',
    }
    
    return render(request, 'custom_admin/news_section_form.html', context)


@staff_member_required
def news_section_edit(request, news_id):
    """Edit an existing news section"""
    from authentication.models import NewsSection
    
    news_section = get_object_or_404(NewsSection, id=news_id)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        image = request.FILES.get('image')
        is_active = request.POST.get('is_active') == 'on'
        display_order = request.POST.get('display_order', 0)
        
        if not title or not content:
            messages.error(request, 'Title and content are required.')
        else:
            news_section.title = title
            news_section.content = content
            if image:
                news_section.image = image
            news_section.is_active = is_active
            news_section.display_order = display_order
            news_section.save()
            
            messages.success(request, f'News section "{title}" updated successfully!')
            return redirect('custom_admin:news_section_list')
    
    context = {
        'title': 'Edit News Section',
        'news_section': news_section,
    }
    
    return render(request, 'custom_admin/news_section_form.html', context)


@staff_member_required
def news_section_delete(request, news_id):
    """Delete a news section"""
    from authentication.models import NewsSection
    
    news_section = get_object_or_404(NewsSection, id=news_id)
    
    if request.method == 'POST':
        title = news_section.title
        news_section.delete()
        messages.success(request, f'News section "{title}" deleted successfully!')
        return redirect('custom_admin:news_section_list')
    
    context = {
        'title': 'Delete News Section',
        'news_section': news_section,
    }
    
    return render(request, 'custom_admin/news_section_delete.html', context)


@staff_member_required
def news_section_toggle(request, news_id):
    """Toggle active status of a news section"""
    from authentication.models import NewsSection
    
    news_section = get_object_or_404(NewsSection, id=news_id)
    news_section.is_active = not news_section.is_active
    news_section.save()
    
    status = 'activated' if news_section.is_active else 'deactivated'
    messages.success(request, f'News section "{news_section.title}" {status}!')
    
    return redirect('custom_admin:news_section_list')


# ============================================================================
# COMPREHENSIVE REPORTS VIEWS
# ============================================================================

@staff_member_required
def recipient_requests_report(request):
    """Display recipient requests in a table with export options"""
    from authentication.models import RequestManagement
    from django.db.models import Q
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    request_type_filter = request.GET.get('request_type', '')
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Base queryset
    requests = RequestManagement.objects.select_related(
        'recipient__user', 'foodbank', 'assigned_foodbank'
    ).all()
    
    # Apply filters
    if status_filter:
        requests = requests.filter(status=status_filter)
    if request_type_filter:
        requests = requests.filter(request_type=request_type_filter)
    if search_query:
        requests = requests.filter(
            Q(recipient__full_name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(recipient__user__email__icontains=search_query)
        )
    if start_date:
        requests = requests.filter(time_of_request__gte=start_date)
    if end_date:
        requests = requests.filter(time_of_request__lte=end_date)
    
    context = {
        'title': 'Recipient Requests Report',
        'requests': requests,
        'status_choices': RequestManagement.STATUS_CHOICES,
        'request_type_choices': RequestManagement.REQUEST_TYPE_CHOICES,
        'status_filter': status_filter,
        'request_type_filter': request_type_filter,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'custom_admin/reports/recipient_requests.html', context)


@staff_member_required
def foodbank_requests_report(request):
    """Display foodbank requests in a table with export options"""
    from authentication.models import FoodBankRequest
    from django.db.models import Q
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    donation_type_filter = request.GET.get('donation_type', '')
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Base queryset
    requests = FoodBankRequest.objects.select_related(
        'foodbank__user', 'original_request__recipient'
    ).all()
    
    # Apply filters
    if status_filter:
        requests = requests.filter(status=status_filter)
    if priority_filter:
        requests = requests.filter(priority=priority_filter)
    if donation_type_filter:
        requests = requests.filter(donation_type=donation_type_filter)
    if search_query:
        requests = requests.filter(
            Q(foodbank__foodbank_name__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    if start_date:
        requests = requests.filter(created_at__gte=start_date)
    if end_date:
        requests = requests.filter(created_at__lte=end_date)
    
    context = {
        'title': 'FoodBank Requests Report',
        'requests': requests,
        'status_choices': FoodBankRequest.STATUS_CHOICES,
        'priority_choices': FoodBankRequest.PRIORITY_CHOICES,
        'donation_type_choices': FoodBankRequest.DONATION_TYPE_CHOICES,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'donation_type_filter': donation_type_filter,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'custom_admin/reports/foodbank_requests.html', context)


@staff_member_required
def donor_donations_report(request):
    """Display donor donations in a table with export options"""
    from authentication.models import Donation
    from django.db.models import Q
    
    # Get filter parameters
    donation_type_filter = request.GET.get('donation_type', '')
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Base queryset
    donations = Donation.objects.select_related(
        'donor', 'foodbank__user'
    ).all()
    
    # Apply filters
    if donation_type_filter:
        donations = donations.filter(donation_type=donation_type_filter)
    if search_query:
        donations = donations.filter(
            Q(donor__email__icontains=search_query) |
            Q(foodbank__foodbank_name__icontains=search_query) |
            Q(item_name__icontains=search_query)
        )
    if start_date:
        donations = donations.filter(donated_at__gte=start_date)
    if end_date:
        donations = donations.filter(donated_at__lte=end_date)
    
    context = {
        'title': 'Donor Donations Report',
        'donations': donations,
        'donation_type_choices': Donation.DONATION_TYPES,
        'donation_type_filter': donation_type_filter,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'custom_admin/reports/donor_donations.html', context)


@staff_member_required
def complete_donation_flow_report(request):
    """Display complete donation flow: Donor → FoodBank → Recipient"""
    from authentication.models import DonationAllocation
    from django.db.models import Q
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    acknowledged_filter = request.GET.get('acknowledged', '')
    
    # Base queryset with all relationships
    allocations = DonationAllocation.objects.select_related(
        'donation__donor',
        'donation__foodbank__user',
        'recipient__user'
    ).prefetch_related(
        'recipient__managed_requests'
    ).all()
    
    # Apply filters
    if search_query:
        allocations = allocations.filter(
            Q(donation__donor__email__icontains=search_query) |
            Q(donation__foodbank__foodbank_name__icontains=search_query) |
            Q(recipient__full_name__icontains=search_query) |
            Q(donation__item_name__icontains=search_query)
        )
    if start_date:
        allocations = allocations.filter(allocated_at__gte=start_date)
    if end_date:
        allocations = allocations.filter(allocated_at__lte=end_date)
    if acknowledged_filter:
        is_acknowledged = acknowledged_filter == 'true'
        allocations = allocations.filter(is_acknowledged=is_acknowledged)
    
    context = {
        'title': 'Complete Donation Flow Report',
        'allocations': allocations,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
        'acknowledged_filter': acknowledged_filter,
    }
    
    return render(request, 'custom_admin/reports/complete_donation_flow.html', context)


@staff_member_required
@require_http_methods(["POST"])
def toggle_user_status(request, user_id):
    """Toggle user active/inactive status via AJAX"""
    try:
        user = CustomUser.objects.get(id=user_id)
        user.is_active = not user.is_active
        user.save()
        
        return JsonResponse({
            'success': True,
            'is_active': user.is_active,
            'message': f'User {"activated" if user.is_active else "deactivated"} successfully'
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@staff_member_required
@require_http_methods(["POST"])
def delete_user(request, user_id):
    """Delete user account via AJAX"""
    try:
        user = CustomUser.objects.get(id=user_id)
        user_email = user.email
        user.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'User {user_email} deleted successfully'
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@staff_member_required
def posted_subscriptions(request):
    """View all approved/posted subscription payments"""
    allowed_subscription_statuses = {
        value for value, _ in FoodBankSubscription.STATUS_CHOICES if value not in {'suspended', 'cancelled'}
    }

    # Get filter parameters
    search = request.GET.get('search', '')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    subscription_status_filter = request.GET.get('subscription_status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    sort_by = request.GET.get('sort', '-verified_at')
    
    # Base queryset - only approved payments
    payments = SubscriptionPayment.objects.filter(
        status='approved'
    ).select_related('foodbank', 'foodbank__user', 'subscription', 'verified_by')
    
    # Apply search filter
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    
    # Apply plan filter
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    
    # Apply payment method filter
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)

    # Apply subscription status filter
    if subscription_status_filter and subscription_status_filter in allowed_subscription_statuses:
        payments = payments.filter(subscription__status=subscription_status_filter)
    elif subscription_status_filter:
        subscription_status_filter = ''
    
    # Apply date range filter
    if date_from:
        payments = payments.filter(verified_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(verified_at__date__lte=date_to)
    
    # Apply sorting
    valid_sorts = [
        'verified_at', '-verified_at',
        'payment_date', '-payment_date',
        'amount', '-amount',
        'foodbank__foodbank_name', '-foodbank__foodbank_name'
    ]
    if sort_by in valid_sorts:
        payments = payments.order_by(sort_by)
    else:
        payments = payments.order_by('-verified_at')
    
    # Statistics
    total_posted = payments.count()
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
    monthly_payments = payments.filter(plan_type='monthly').count()
    yearly_payments = payments.filter(plan_type='yearly').count()
    
    # Recent approvals (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_approvals = payments.filter(verified_at__gte=thirty_days_ago).count()
    
    # Pagination
    paginator = Paginator(payments, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Posted Subscriptions',
        'page_obj': page_obj,
        'total_posted': total_posted,
        'total_revenue': total_revenue,
        'monthly_payments': monthly_payments,
        'yearly_payments': yearly_payments,
        'recent_approvals': recent_approvals,
        'plan_choices': SubscriptionPayment.PLAN_CHOICES,
        'payment_method_choices': SubscriptionPayment.PAYMENT_METHOD_CHOICES,
        'subscription_status_choices': [
            (value, label)
            for value, label in FoodBankSubscription.STATUS_CHOICES
            if value not in {'suspended', 'cancelled'}
        ],
        'current_filters': {
            'search': search,
            'plan': plan_filter,
            'payment_method': payment_method_filter,
            'subscription_status': subscription_status_filter,
            'date_from': date_from,
            'date_to': date_to,
            'sort': sort_by,
        }
    }
    
    return render(request, 'custom_admin/posted_subscriptions.html', context)


@staff_member_required
def export_posted_subscriptions_excel(request):
    """Export posted subscriptions to Excel"""
    allowed_subscription_statuses = {
        value for value, _ in FoodBankSubscription.STATUS_CHOICES if value not in {'suspended', 'cancelled'}
    }

    # Get the same filters as the main view
    search = request.GET.get('search', '')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    subscription_status_filter = request.GET.get('subscription_status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset - only approved payments
    payments = SubscriptionPayment.objects.filter(
        status='approved'
    ).select_related('foodbank', 'foodbank__user', 'subscription', 'verified_by')
    
    # Apply filters
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if subscription_status_filter and subscription_status_filter in allowed_subscription_statuses:
        payments = payments.filter(subscription__status=subscription_status_filter)
    if date_from:
        payments = payments.filter(verified_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(verified_at__date__lte=date_to)
    
    payments = payments.order_by('-verified_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Posted Subscriptions"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'S/No', 'Food Bank Name', 'Email', 'Phone', 'Plan Type', 'Amount (KSH)', 
        'Payment Method', 'Transaction Ref', 'Payment Date', 'Message',
        'Verified By', 'Verified At', 'Subscription Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = border
        ws.cell(row=row_num, column=2, value=payment.foodbank.foodbank_name).border = border
        ws.cell(row=row_num, column=3, value=payment.foodbank.user.email).border = border
        ws.cell(row=row_num, column=4, value=payment.foodbank.user.phone_number or 'N/A').border = border
        ws.cell(row=row_num, column=5, value=payment.get_plan_type_display()).border = border
        ws.cell(row=row_num, column=6, value=float(payment.amount)).border = border
        ws.cell(row=row_num, column=7, value=payment.get_payment_method_display()).border = border
        ws.cell(row=row_num, column=8, value=payment.transaction_reference).border = border
        ws.cell(row=row_num, column=9, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=10, value=payment.notes or '').border = border
        ws.cell(row=row_num, column=11, value=payment.verified_by.email if payment.verified_by else 'N/A').border = border
        ws.cell(row=row_num, column=12, value=payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else 'N/A').border = border
        ws.cell(row=row_num, column=13, value=payment.subscription.get_status_display()).border = border
    
    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    # Add summary row
    summary_row = payments.count() + 3
    ws.cell(row=summary_row, column=1, value="TOTAL REVENUE:").font = Font(bold=True)
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
    ws.cell(row=summary_row, column=6, value=float(total_revenue)).font = Font(bold=True)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="posted_subscriptions_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@staff_member_required
def export_posted_subscriptions_csv(request):
    """Export posted subscriptions to CSV."""
    allowed_subscription_statuses = {
        value for value, _ in FoodBankSubscription.STATUS_CHOICES if value not in {'suspended', 'cancelled'}
    }

    search = request.GET.get('search', '')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    subscription_status_filter = request.GET.get('subscription_status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    payments = SubscriptionPayment.objects.filter(
        status='approved'
    ).select_related('foodbank', 'foodbank__user', 'subscription', 'verified_by')

    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if subscription_status_filter and subscription_status_filter in allowed_subscription_statuses:
        payments = payments.filter(subscription__status=subscription_status_filter)
    if date_from:
        payments = payments.filter(verified_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(verified_at__date__lte=date_to)

    payments = payments.order_by('-verified_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="posted_subscriptions_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        'S/No', 'Food Bank Name', 'Email', 'Phone', 'Plan Type', 'Amount (KSH)',
        'Payment Method', 'Transaction Ref', 'Payment Date', 'Message',
        'Verified By', 'Verified At', 'Subscription Status'
    ])

    for idx, payment in enumerate(payments, 1):
        writer.writerow([
            idx,
            payment.foodbank.foodbank_name,
            payment.foodbank.user.email,
            payment.foodbank.user.phone_number or 'N/A',
            payment.get_plan_type_display(),
            f"{payment.amount:.2f}",
            payment.get_payment_method_display(),
            payment.transaction_reference,
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.notes or '',
            payment.verified_by.email if payment.verified_by else 'N/A',
            payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else 'N/A',
            payment.subscription.get_status_display(),
        ])

    return response


@staff_member_required
def export_posted_subscriptions_pdf(request):
    """Export posted subscriptions to PDF"""
    allowed_subscription_statuses = {
        value for value, _ in FoodBankSubscription.STATUS_CHOICES if value not in {'suspended', 'cancelled'}
    }

    # Get the same filters as the main view
    search = request.GET.get('search', '')
    plan_filter = request.GET.get('plan')
    payment_method_filter = request.GET.get('payment_method')
    subscription_status_filter = request.GET.get('subscription_status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset - only approved payments
    payments = SubscriptionPayment.objects.filter(
        status='approved'
    ).select_related('foodbank', 'foodbank__user', 'subscription', 'verified_by')
    
    # Apply filters
    if search:
        payments = payments.filter(
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__user__phone_number__icontains=search) |
            Q(transaction_reference__icontains=search)
        )
    if plan_filter:
        payments = payments.filter(plan_type=plan_filter)
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    if subscription_status_filter and subscription_status_filter in allowed_subscription_statuses:
        payments = payments.filter(subscription__status=subscription_status_filter)
    if date_from:
        payments = payments.filter(verified_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(verified_at__date__lte=date_to)
    
    payments = payments.order_by('-verified_at')
    
    # Create PDF
    from xml.sax.saxutils import escape

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.3 * inch,
        rightMargin=0.3 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch
    )
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#10b981'),
        spaceAfter=30,
        alignment=1  # Center
    )
    header_wrap_style = ParagraphStyle(
        'PostedSubHeaderWrap',
        parent=styles['Normal'],
        fontSize=8,
        leading=9,
        alignment=1,
        wordWrap='CJK',
    )
    cell_wrap_style = ParagraphStyle(
        'PostedSubCellWrap',
        parent=styles['Normal'],
        fontSize=7,
        leading=8.2,
        alignment=0,
        wordWrap='CJK',
    )
    
    # Title
    title = Paragraph("Posted Subscriptions Report", title_style)
    elements.append(title)
    
    # Report info
    report_info = Paragraph(
        f"Generated on: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        f"Total Records: {payments.count()}",
        styles['Normal']
    )
    elements.append(report_info)
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [[
        Paragraph('S/No', header_wrap_style),
        Paragraph('Food Bank', header_wrap_style),
        Paragraph('Email', header_wrap_style),
        Paragraph('Phone', header_wrap_style),
        Paragraph('Plan', header_wrap_style),
        Paragraph('Amount', header_wrap_style),
        Paragraph('Payment Method', header_wrap_style),
        Paragraph('Transaction Ref', header_wrap_style),
        Paragraph('Payment Date', header_wrap_style),
        Paragraph('Message', header_wrap_style),
        Paragraph('Verified By', header_wrap_style),
        Paragraph('Verified At', header_wrap_style),
        Paragraph('Subscription Status', header_wrap_style),
    ]]
    
    for idx, payment in enumerate(payments, 1):
        plan_text = payment.get_plan_type_display()
        method_text = payment.get_payment_method_display()
        transaction_ref = payment.transaction_reference or 'N/A'
        message_text = payment.notes or ''
        verified_by_text = payment.verified_by.email if payment.verified_by else 'N/A'
        subscription_status_text = payment.subscription.get_status_display() if payment.subscription else 'N/A'

        data.append([
            str(idx),
            Paragraph(escape((payment.foodbank.foodbank_name or '')[:32]), cell_wrap_style),
            Paragraph(escape((payment.foodbank.user.email or '')[:36]), cell_wrap_style),
            Paragraph(escape((payment.foodbank.user.phone_number or 'N/A')[:20]), cell_wrap_style),
            Paragraph(escape(plan_text[:22]), cell_wrap_style),
            f"KSH {payment.amount:,.2f}",
            Paragraph(escape(method_text[:30]), cell_wrap_style),
            Paragraph(escape(transaction_ref[:40]), cell_wrap_style),
            payment.payment_date.strftime('%Y-%m-%d'),
            Paragraph(escape(message_text[:65]), cell_wrap_style),
            Paragraph(escape(verified_by_text[:30]), cell_wrap_style),
            payment.verified_at.strftime('%Y-%m-%d') if payment.verified_at else 'N/A',
            Paragraph(escape(subscription_status_text[:22]), cell_wrap_style),
        ])
    
    # Create table
    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            0.33 * inch, 0.9 * inch, 1.15 * inch, 0.7 * inch, 0.75 * inch,
            0.75 * inch, 0.95 * inch, 1.0 * inch, 0.75 * inch, 1.05 * inch,
            0.8 * inch, 0.8 * inch, 0.85 * inch
        ]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (8, 1), (8, -1), 'CENTER'),
        ('ALIGN', (11, 1), (11, -1), 'CENTER'),
        ('ALIGN', (12, 1), (12, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('TOPPADDING', (0, 0), (-1, 0), 7),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.7, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
    summary = Paragraph(
        f"<b>Total Revenue: KSH {total_revenue:,.2f}</b>",
        styles['Heading2']
    )
    elements.append(summary)
    
    # Build PDF
    doc.build(elements)
    
    # Create response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="posted_subscriptions_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response
