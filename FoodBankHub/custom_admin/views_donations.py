"""
Comprehensive Donations Management Views for Custom Admin Panel
Handles: Recipient Requests, Foodbank Requests, Donor Donations, Allocations, and Discussions
"""

from django.shortcuts import render, get_object_or_404, redirect
from .decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q, F, Avg
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.core.paginator import Paginator
from datetime import timedelta, datetime
from decimal import Decimal
import json
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from authentication.models import (
    RecipientRequest, FoodBankRequest, Donation, DonationAllocation,
    DonationDiscussion, PaymentTransaction, RecipientProfile, FoodBankProfile,
    DonorProfile, CustomUser, UnspecifiedDonationManagement, RequestManagement,
    DonationResponse
)


@staff_member_required
def recipient_requests_management(request):
    """
    Comprehensive management of recipient requests to foodbanks
    Shows all requests with detailed filtering and status tracking
    """
    requests = RecipientRequest.objects.select_related(
        'recipient__user', 
        'foodbank'
    ).prefetch_related('declined_by').order_by('-created_at')
    
    # Statistics
    total_requests = requests.count()
    pending_requests = requests.filter(status='pending').count()
    accepted_requests = requests.filter(status='accepted').count()
    completed_requests = requests.filter(status='completed').count()
    declined_requests = requests.filter(status='declined').count()
    
    # Filtering
    status_filter = request.GET.get('status')
    foodbank_filter = request.GET.get('foodbank')
    anonymous_filter = request.GET.get('anonymous')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        requests = requests.filter(status=status_filter)
    if foodbank_filter:
        requests = requests.filter(foodbank_id=foodbank_filter)
    if anonymous_filter == 'yes':
        requests = requests.filter(is_anonymous=True)
    elif anonymous_filter == 'no':
        requests = requests.filter(is_anonymous=False)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            requests = requests.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            requests = requests.filter(created_at__date__lte=parsed_to)
    if search:
        requests = requests.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(recipient__user__email__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(location__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all active foodbanks for filter dropdown
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    
    context = {
        'title': 'Recipient Requests Management',
        'page_obj': page_obj,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'accepted_requests': accepted_requests,
        'completed_requests': completed_requests,
        'declined_requests': declined_requests,
        'status_choices': RecipientRequest.STATUS_CHOICES,
        'foodbanks': foodbanks,
        'current_filters': {
            'status': status_filter,
            'foodbank': foodbank_filter,
            'anonymous': anonymous_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    return render(request, 'custom_admin/recipient_requests_management.html', context)


@staff_member_required
def foodbank_requests_enhanced(request):
    """
    Enhanced foodbank requests management with comprehensive filters
    Shows requests from foodbanks to donors with detailed tracking
    """
    requests = FoodBankRequest.objects.select_related(
        'foodbank', 
        'original_request'
    ).prefetch_related('donations').order_by('-created_at')
    
    # Statistics
    total_requests = requests.count()
    active_requests = requests.filter(status='active').count()
    fulfilled_requests = requests.filter(status='fulfilled').count()
    expired_requests = requests.filter(status='expired').count()
    urgent_requests = requests.filter(priority='urgent', status='active').count()
    
    # Calculate fulfillment rate
    if total_requests > 0:
        fulfillment_rate = (fulfilled_requests / total_requests) * 100
    else:
        fulfillment_rate = 0
    
    # Filtering
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    donation_type_filter = request.GET.get('donation_type')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    has_deadline = request.GET.get('has_deadline')
    search = request.GET.get('search')
    
    if status_filter:
        requests = requests.filter(status=status_filter)
    if donation_type_filter:
        requests = requests.filter(donation_type=donation_type_filter)
    if foodbank_filter:
        requests = requests.filter(foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            requests = requests.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            requests = requests.filter(created_at__date__lte=parsed_to)
    if has_deadline == 'yes':
        requests = requests.exclude(deadline__isnull=True)
    elif has_deadline == 'no':
        requests = requests.filter(deadline__isnull=True)
    if search:
        requests = requests.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(item_name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all active foodbanks for filter dropdown
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    
    context = {
        'title': 'Foodbank Requests Management',
        'page_obj': page_obj,
        'total_requests': total_requests,
        'active_requests': active_requests,
        'fulfilled_requests': fulfilled_requests,
        'expired_requests': expired_requests,
        'urgent_requests': urgent_requests,
        'fulfillment_rate': round(fulfillment_rate, 1),
        'status_choices': FoodBankRequest.STATUS_CHOICES,
        'priority_choices': FoodBankRequest.PRIORITY_CHOICES,
        'donation_type_choices': FoodBankRequest.DONATION_TYPE_CHOICES,
        'foodbanks': foodbanks,
        'current_filters': {
            'status': status_filter,
            'priority': priority_filter,
            'donation_type': donation_type_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'has_deadline': has_deadline,
            'search': search,
        }
    }
    return render(request, 'custom_admin/foodbank_requests_enhanced.html', context)


@staff_member_required
def donor_donations_management(request):
    """
    Comprehensive donor donations management
    Shows all donations from donors to foodbanks with detailed tracking
    """
    donations = Donation.objects.select_related(
        'donor',
        'foodbank',
        'foodbank_request'
    ).prefetch_related('allocations').order_by('-donated_at')
    
    # Statistics
    total_donations = donations.count()
    pending_donations = donations.filter(status='pending').count()
    accepted_donations = donations.filter(status='accepted').count()
    declined_donations = donations.filter(status='declined').count()
    
    # Financial statistics
    total_monetary_value = donations.filter(
        donation_type='money'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_subsidized_value = donations.filter(
        donation_type='subsidized'
    ).aggregate(total=Sum('subsidized_price'))['total'] or 0
    
    # Delivery statistics
    pending_delivery = donations.filter(delivery_status='pending').count()
    in_transit = donations.filter(delivery_status='in_transit').count()
    delivered = donations.filter(delivery_status='delivered').count()
    
    # Filtering
    donation_type_filter = request.GET.get('donation_type')
    donation_category_filter = request.GET.get('donation_category')
    donation_mode_filter = request.GET.get('donation_mode')
    status_filter = request.GET.get('status')
    delivery_status_filter = request.GET.get('delivery_status')
    foodbank_filter = request.GET.get('foodbank')
    donor_filter = request.GET.get('donor')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    min_amount = request.GET.get('min_amount')
    max_amount = request.GET.get('max_amount')
    search = request.GET.get('search')
    
    if donation_type_filter:
        donations = donations.filter(donation_type=donation_type_filter)
    if donation_category_filter:
        donations = donations.filter(donation_category=donation_category_filter)
    if donation_mode_filter:
        donations = donations.filter(donation_mode=donation_mode_filter)
    if status_filter:
        donations = donations.filter(status=status_filter)
    if delivery_status_filter:
        donations = donations.filter(delivery_status=delivery_status_filter)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if donor_filter:
        donations = donations.filter(donor_id=donor_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if min_amount:
        donations = donations.filter(amount__gte=min_amount)
    if max_amount:
        donations = donations.filter(amount__lte=max_amount)
    if search:
        donations = donations.filter(
            Q(donor__email__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(item_name__icontains=search) |
            Q(message__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    donation_list = list(page_obj.object_list)
    donation_ids = [donation.id for donation in donation_list]
    recipient_notes_by_donation = {}

    if donation_ids:
        donation_responses = DonationResponse.objects.filter(
            donation_id__in=donation_ids
        ).exclude(
            notes__isnull=True
        ).exclude(
            notes__exact=''
        ).order_by('-responded_at')

        for response in donation_responses:
            if response.donation_id not in recipient_notes_by_donation:
                recipient_notes_by_donation[response.donation_id] = response.notes

        for donation in donation_list:
            donation.latest_recipient_note = recipient_notes_by_donation.get(donation.id)
    
    # Get foodbanks and donors for filters
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    donors = CustomUser.objects.filter(user_type='DONOR').order_by('email')
    
    context = {
        'title': 'Donor Donations Management',
        'page_obj': page_obj,
        'total_donations': total_donations,
        'pending_donations': pending_donations,
        'accepted_donations': accepted_donations,
        'declined_donations': declined_donations,
        'total_monetary_value': total_monetary_value,
        'total_subsidized_value': total_subsidized_value,
        'pending_delivery': pending_delivery,
        'in_transit': in_transit,
        'delivered': delivered,
        'donation_types': Donation.DONATION_TYPES,
        'donation_categories': Donation.DONATION_CATEGORIES,
        'donation_modes': Donation.DONATION_MODES,
        'status_choices': Donation.STATUS_CHOICES,
        'delivery_statuses': Donation.DELIVERY_STATUS_CHOICES,
        'foodbanks': foodbanks,
        'donors': donors,
        'current_filters': {
            'donation_type': donation_type_filter,
            'donation_category': donation_category_filter,
            'donation_mode': donation_mode_filter,
            'status': status_filter,
            'delivery_status': delivery_status_filter,
            'foodbank': foodbank_filter,
            'donor': donor_filter,
            'date_from': date_from,
            'date_to': date_to,
            'min_amount': min_amount,
            'max_amount': max_amount,
            'search': search,
        }
    }
    return render(request, 'custom_admin/donor_donations_management.html', context)


@staff_member_required
def received_donations_management(request):
    """
    Admin view to list all donations received by foodbanks (mirrors foodbank donations list).
    """
    from django.http import HttpResponseRedirect
    from urllib.parse import urlencode
    if 'allocation_status' in request.GET:
        qd = request.GET.copy()
        del qd['allocation_status']
        query = qd.urlencode()
        url = request.path + ('?' + query if query else '')
        return HttpResponseRedirect(url)

    donations = Donation.objects.filter(
        foodbank__isnull=False
    ).select_related(
        'donor', 'donor__donor_profile', 'foodbank', 'foodbank__user',
        'foodbank_request', 'accepted_by_recipient', 'declined_by_recipient',
        'accepted_by_recipient__user', 'declined_by_recipient__user',
        'request_management', 'request_management__foodbank', 'request_management__recipient',
        'foodbank_request__original_request', 'foodbank_request__original_request__foodbank',
        'foodbank_request__original_request__recipient'
    ).prefetch_related(
        'allocations__recipient', 'discussion'
    ).order_by('-donated_at')

    donation_type = request.GET.get('type', '')
    category = request.GET.get('category', '')
    delivery_status = request.GET.get('delivery_status', '')
    delivery_method = request.GET.get('delivery_method', '')
    foodbank_filter = request.GET.get('foodbank', '')
    status_filter = request.GET.get('status', '').strip()
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if donation_type:
        donations = donations.filter(donation_type=donation_type)

    if category:
        donations = donations.filter(donation_category=category)

    if delivery_status:
        donations = donations.filter(delivery_status=delivery_status)

    if delivery_method:
        donations = donations.filter(delivery_method=delivery_method)

    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)

    if status_filter:
        donations = donations.filter(status=status_filter)
    else:
        donations = donations.filter(status__in=['accepted', 'fulfilled', 'partial'])

    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)

    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)

    if search:
        donations = donations.filter(
            Q(item_name__icontains=search) |
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(message__icontains=search) |
            Q(csr_description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search)
        )

    total_donations = donations.count()
    total_items = donations.filter(donation_type='item').aggregate(
        total=Sum('quantity')
    )['total'] or 0
    total_money = donations.filter(donation_type='money').aggregate(
        total=Sum('amount')
    )['total'] or 0
    total_subsidized = donations.filter(donation_type='subsidized').aggregate(
        total=Sum('subsidized_price')
    )['total'] or 0
    unallocated_count = donations.filter(is_allocated=False).count()

    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    from authentication.donation_views import get_display_status, STATUS_CLASS_MAP

    for donation in page_obj.object_list:
        donation.recipient_display = donation.get_recipient_name()
        donation.status_display = get_display_status(donation)
        donation.status_class = STATUS_CLASS_MAP.get(donation.status_display, 'pending')

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    context = {
        'title': 'Received Donations',
        'donations': page_obj,
        'page_obj': page_obj,
        'total_donations': total_donations,
        'total_items': total_items,
        'total_money': total_money,
        'total_subsidized': total_subsidized,
        'unallocated_count': unallocated_count,
        'foodbanks': foodbanks,
    }

    return render(request, 'custom_admin/received_donations_management.html', context)


def _received_donations_queryset(request):
    donations = Donation.objects.filter(
        foodbank__isnull=False
    ).select_related(
        'donor', 'donor__donor_profile', 'foodbank', 'foodbank__user',
        'foodbank_request', 'accepted_by_recipient', 'accepted_by_recipient__user'
    ).order_by('-donated_at')

    donation_type = request.GET.get('type', '')
    category = request.GET.get('category', '')
    delivery_status = request.GET.get('delivery_status', '')
    delivery_method = request.GET.get('delivery_method', '')
    foodbank_filter = request.GET.get('foodbank', '')
    status_filter = request.GET.get('status', '').strip()
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if donation_type:
        donations = donations.filter(donation_type=donation_type)
    if category:
        donations = donations.filter(donation_category=category)
    if delivery_status:
        donations = donations.filter(delivery_status=delivery_status)
    if delivery_method:
        donations = donations.filter(delivery_method=delivery_method)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if status_filter:
        donations = donations.filter(status=status_filter)
    else:
        donations = donations.filter(status__in=['accepted', 'fulfilled', 'partial'])
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(item_name__icontains=search) |
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(message__icontains=search) |
            Q(csr_description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search)
        )
    return donations


def _received_donations_rows(donations):
    rows = []
    for donation in donations:
        donor_name = donation.donor.donor_profile.full_name if getattr(donation.donor, 'donor_profile', None) and donation.donor.donor_profile.full_name else donation.donor.email
        foodbank_name = donation.foodbank.foodbank_name if donation.foodbank else '-'
        recipient = donation.get_recipient_name() if hasattr(donation, 'get_recipient_name') else ''
        category = 'Free Goods' if donation.donation_type == 'item' else donation.get_donation_type_display()
        donation_type_display = donation.get_donation_category_display()
        description = donation.item_name or donation.other_description or donation.subsidized_product_type or donation.message or 'Donation'
        quantity_amount = ''
        if donation.donation_type == 'item':
            quantity_amount = f"{donation.quantity or '-'} {donation.quantity_unit or 'units'}"
        elif donation.donation_type == 'money':
            quantity_amount = f"KES {donation.amount or 0}"
        elif donation.donation_type == 'subsidized':
            quantity_amount = f"{donation.subsidized_quantity or '-'} {donation.subsidized_quantity_unit or ''} / KES {donation.subsidized_price or 0}"
        elif donation.donation_type in ['csr', 'other']:
            qty = f"{donation.quantity} {donation.quantity_unit}" if donation.quantity else ''
            amt = f"KES {donation.amount}" if donation.amount else ''
            quantity_amount = qty or amt or '-'
        rows.append([
            donation.donated_at.strftime('%Y-%m-%d %H:%M') if donation.donated_at else '',
            donor_name,
            foodbank_name,
            recipient,
            category,
            donation_type_display,
            description,
            quantity_amount,
            donation.get_status_display(),
            donation.get_delivery_method_display(),
            donation.foodbank.address if donation.foodbank else '',
        ])
    return rows


@staff_member_required
def export_received_donations_csv(request):
    donations = _received_donations_queryset(request)
    rows = _received_donations_rows(donations)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="received_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Donor', 'Foodbank', 'Recipient', 'Category', 'Type',
        'Description', 'Quantity/Amount', 'Status', 'Delivery Method', 'Location'
    ])
    writer.writerows(rows)
    return response


@staff_member_required
def export_received_donations_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    donations = _received_donations_queryset(request)
    rows = _received_donations_rows(donations)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    story = []
    styles = getSampleStyleSheet()
    story.append(Paragraph('Received Donations Report', styles['Title']))
    story.append(Paragraph(datetime.now().strftime('%b %d, %Y %H:%M'), styles['Normal']))
    story.append(Spacer(1, 12))

    data = [[
        'Date', 'Donor', 'Foodbank', 'Recipient', 'Category', 'Type',
        'Description', 'Quantity/Amount', 'Status', 'Delivery Method', 'Location'
    ]] + rows

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    story.append(table)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="received_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)
    return response


@staff_member_required
def donation_allocations_management(request):
    """
    Management of donation allocations to recipients
    Tracks how donations are distributed to individual recipients
    """
    allocations = DonationAllocation.objects.select_related(
        'donation__donor',
        'donation__foodbank',
        'recipient__user'
    ).order_by('-allocated_at')
    
    # Statistics
    total_allocations = allocations.count()
    acknowledged_allocations = allocations.filter(is_acknowledged=True).count()
    pending_acknowledgment = allocations.filter(is_acknowledged=False).count()
    
    # Value statistics
    total_quantity_allocated = allocations.exclude(
        quantity__isnull=True
    ).aggregate(total=Sum('quantity'))['total'] or 0
    
    total_amount_allocated = allocations.exclude(
        amount__isnull=True
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Filtering
    donation_type_filter = request.GET.get('donation_type')
    foodbank_filter = request.GET.get('foodbank')
    recipient_filter = request.GET.get('recipient')
    acknowledged_filter = request.GET.get('acknowledged')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if donation_type_filter:
        allocations = allocations.filter(donation__donation_type=donation_type_filter)
    if foodbank_filter:
        allocations = allocations.filter(donation__foodbank_id=foodbank_filter)
    if recipient_filter:
        allocations = allocations.filter(recipient_id=recipient_filter)
    if acknowledged_filter == 'yes':
        allocations = allocations.filter(is_acknowledged=True)
    elif acknowledged_filter == 'no':
        allocations = allocations.filter(is_acknowledged=False)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            allocations = allocations.filter(allocated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            allocations = allocations.filter(allocated_at__date__lte=parsed_to)
    if search:
        allocations = allocations.filter(
            Q(recipient__user__email__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(donation__foodbank__foodbank_name__icontains=search) |
            Q(donation__donor__email__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(allocations, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get foodbanks and recipients for filters
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    recipients = RecipientProfile.objects.select_related('user').order_by('user__email')
    
    context = {
        'title': 'Donation Allocations Management',
        'page_obj': page_obj,
        'total_allocations': total_allocations,
        'acknowledged_allocations': acknowledged_allocations,
        'pending_acknowledgment': pending_acknowledgment,
        'total_quantity_allocated': total_quantity_allocated,
        'total_amount_allocated': total_amount_allocated,
        'donation_types': Donation.DONATION_TYPES,
        'foodbanks': foodbanks,
        'recipients': recipients,
        'current_filters': {
            'donation_type': donation_type_filter,
            'foodbank': foodbank_filter,
            'recipient': recipient_filter,
            'acknowledged': acknowledged_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    return render(request, 'custom_admin/donation_allocations_management.html', context)


@staff_member_required
def donation_discussions_management(request):
    """
    Monitor donation discussions between donors and foodbanks
    For 'other' type donations that require negotiation
    """
    discussions = DonationDiscussion.objects.select_related(
        'donation__donor',
        'donation__foodbank',
        'donor',
        'foodbank'
    ).prefetch_related('messages').order_by('-created_at')
    
    # Statistics
    total_discussions = discussions.count()
    pending_discussions = discussions.filter(status='pending').count()
    in_progress_discussions = discussions.filter(status='in_progress').count()
    agreed_discussions = discussions.filter(status='agreed').count()
    declined_discussions = discussions.filter(status='declined').count()
    
    # Filtering
    status_filter = request.GET.get('status')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        discussions = discussions.filter(status=status_filter)
    if foodbank_filter:
        discussions = discussions.filter(donation__foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            discussions = discussions.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            discussions = discussions.filter(created_at__date__lte=parsed_to)
    if search:
        discussions = discussions.filter(
            Q(donor__email__icontains=search) |
            Q(foodbank__user__email__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(donation__foodbank__foodbank_name__icontains=search) |
            Q(donation__other_description__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(discussions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get foodbanks for filter
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    
    context = {
        'title': 'Donation Discussions Management',
        'page_obj': page_obj,
        'total_discussions': total_discussions,
        'pending_discussions': pending_discussions,
        'in_progress_discussions': in_progress_discussions,
        'agreed_discussions': agreed_discussions,
        'declined_discussions': declined_discussions,
        'status_choices': DonationDiscussion._meta.get_field('status').choices,
        'foodbanks': foodbanks,
        'current_filters': {
            'status': status_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    return render(request, 'custom_admin/donation_discussions_management.html', context)


@staff_member_required
def payment_transactions_management(request):
    """
    Monitor payment transactions for subsidized and monetary donations
    """
    transactions = PaymentTransaction.objects.select_related(
        'donation__donor',
        'donation__foodbank'
    ).order_by('-created_at')
    
    # Statistics
    total_transactions = transactions.count()
    pending_transactions = transactions.filter(status='pending').count()
    completed_transactions = transactions.filter(status='completed').count()
    failed_transactions = transactions.filter(status='failed').count()
    
    total_transaction_amount = transactions.filter(
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Filtering
    status_filter = request.GET.get('status')
    payment_method_filter = request.GET.get('payment_method')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    min_amount = request.GET.get('min_amount')
    max_amount = request.GET.get('max_amount')
    search = request.GET.get('search')
    
    if status_filter:
        transactions = transactions.filter(status=status_filter)
    if payment_method_filter:
        transactions = transactions.filter(payment_method=payment_method_filter)
    if foodbank_filter:
        transactions = transactions.filter(donation__foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            transactions = transactions.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            transactions = transactions.filter(created_at__date__lte=parsed_to)
    if min_amount:
        transactions = transactions.filter(amount__gte=min_amount)
    if max_amount:
        transactions = transactions.filter(amount__lte=max_amount)
    if search:
        transactions = transactions.filter(
            Q(id__icontains=search) |
            Q(mpesa_receipt_number__icontains=search) |
            Q(donation__donor__email__icontains=search) |
            Q(donation__foodbank__foodbank_name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(transactions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get foodbanks for filter
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    
    context = {
        'title': 'Payment Transactions Management',
        'page_obj': page_obj,
        'total_transactions': total_transactions,
        'pending_transactions': pending_transactions,
        'completed_transactions': completed_transactions,
        'failed_transactions': failed_transactions,
        'total_transaction_amount': total_transaction_amount,
        'status_choices': PaymentTransaction.PAYMENT_STATUS_CHOICES,
        'payment_methods': PaymentTransaction.PAYMENT_METHOD_CHOICES,
        'foodbanks': foodbanks,
        'current_filters': {
            'status': status_filter,
            'payment_method': payment_method_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'min_amount': min_amount,
            'max_amount': max_amount,
            'search': search,
        }
    }
    return render(request, 'custom_admin/payment_transactions_management.html', context)


@staff_member_required
def donations_overview_dashboard(request):
    """
    Comprehensive overview dashboard for all donation-related activities
    Provides high-level metrics and insights
    """
    # Time range filter
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    # Recipient Requests Stats
    recipient_requests_total = RecipientRequest.objects.count()
    recipient_requests_recent = RecipientRequest.objects.filter(created_at__gte=start_date).count()
    recipient_requests_pending = RecipientRequest.objects.filter(status='pending').count()
    
    # Foodbank Requests Stats
    foodbank_requests_total = FoodBankRequest.objects.count()
    foodbank_requests_recent = FoodBankRequest.objects.filter(created_at__gte=start_date).count()
    foodbank_requests_active = FoodBankRequest.objects.filter(status='active').count()
    foodbank_requests_urgent = FoodBankRequest.objects.filter(priority='urgent', status='active').count()
    
    # Donations Stats
    donations_total = Donation.objects.count()
    donations_recent = Donation.objects.filter(donated_at__gte=start_date).count()
    donations_pending = Donation.objects.filter(status='pending').count()
    donations_accepted = Donation.objects.filter(status='accepted').count()
    
    # Financial Stats
    total_monetary_donations = Donation.objects.filter(
        donation_type='money',
        status='accepted'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_subsidized_value = Donation.objects.filter(
        donation_type='subsidized',
        status='accepted'
    ).aggregate(total=Sum('subsidized_price'))['total'] or 0
    
    # Allocations Stats
    allocations_total = DonationAllocation.objects.count()
    allocations_recent = DonationAllocation.objects.filter(allocated_at__gte=start_date).count()
    allocations_pending_ack = DonationAllocation.objects.filter(is_acknowledged=False).count()
    
    # Discussions Stats
    discussions_total = DonationDiscussion.objects.count()
    discussions_pending = DonationDiscussion.objects.filter(status='pending').count()
    discussions_in_progress = DonationDiscussion.objects.filter(status='in_progress').count()
    
    # Payment Transactions Stats
    transactions_total = PaymentTransaction.objects.count()
    transactions_pending = PaymentTransaction.objects.filter(status='pending').count()
    transactions_completed = PaymentTransaction.objects.filter(status='completed').count()
    
    # NEW: Unspecified Donations Stats
    unspecified_total = UnspecifiedDonationManagement.objects.count()
    unspecified_pending_foodbank = UnspecifiedDonationManagement.objects.filter(
        foodbank_status='pending_foodbank'
    ).count()
    unspecified_pending_recipient = UnspecifiedDonationManagement.objects.filter(
        recipient_status='pending_recipient'
    ).count()
    unspecified_received = UnspecifiedDonationManagement.objects.filter(
        recipient_status='received'
    ).count()
    
    # NEW: Subsidized Donations Stats
    subsidized_total = Donation.objects.filter(donation_type='subsidized').count()
    subsidized_pending = Donation.objects.filter(
        donation_type='subsidized', status='pending'
    ).count()
    subsidized_with_responses = Donation.objects.filter(
        donation_type='subsidized', subsidized_responded_by__isnull=False
    ).distinct().count()
    
    # NEW: Direct Donations Stats (donations tied to specific requests)
    direct_donations_total = Donation.objects.filter(foodbank_request__isnull=False).count()
    direct_donations_pending = Donation.objects.filter(
        foodbank_request__isnull=False, status='pending'
    ).count()
    direct_donations_accepted = Donation.objects.filter(
        foodbank_request__isnull=False, status='accepted'
    ).count()
    
    # NEW: Request Management Stats
    request_management_total = RequestManagement.objects.count()
    request_management_pending = RequestManagement.objects.filter(status='pending').count()
    request_management_fulfilled = RequestManagement.objects.filter(status='fulfilled').count()
    request_management_anonymous = RequestManagement.objects.filter(is_anonymous=True).count()
    
    # NEW: Donation Responses Stats
    donation_responses_total = DonationResponse.objects.count()
    donation_responses_accepted = DonationResponse.objects.filter(response_type='accepted').count()
    donation_responses_declined = DonationResponse.objects.filter(response_type='declined').count()
    donation_responses_recent = DonationResponse.objects.filter(
        responded_at__gte=start_date
    ).count()
    
    # Donation Type Breakdown
    donation_types_breakdown = Donation.objects.filter(
        status='accepted'
    ).values('donation_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Top Donors
    top_donors = Donation.objects.filter(
        status='accepted'
    ).values(
        'donor__email'
    ).annotate(
        total_donations=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-total_donations')[:10]
    
    # Top Foodbanks (by donations received)
    top_foodbanks = Donation.objects.filter(
        status='accepted'
    ).values(
        'foodbank__foodbank_name'
    ).annotate(
        total_donations=Count('id')
    ).order_by('-total_donations')[:10]
    
    # Recent Activity
    recent_recipient_requests = RecipientRequest.objects.select_related(
        'recipient__user', 'foodbank'
    ).order_by('-created_at')[:5]
    
    recent_donations = Donation.objects.select_related(
        'donor', 'foodbank'
    ).order_by('-donated_at')[:5]
    
    recent_unspecified = UnspecifiedDonationManagement.objects.select_related(
        'donation__donor', 'donation__foodbank'
    ).order_by('-created_at')[:5]
    
    context = {
        'title': 'Donations Overview Dashboard',
        'days': days,
        'recipient_requests_total': recipient_requests_total,
        'recipient_requests_recent': recipient_requests_recent,
        'recipient_requests_pending': recipient_requests_pending,
        'foodbank_requests_total': foodbank_requests_total,
        'foodbank_requests_recent': foodbank_requests_recent,
        'foodbank_requests_active': foodbank_requests_active,
        'foodbank_requests_urgent': foodbank_requests_urgent,
        'donations_total': donations_total,
        'donations_recent': donations_recent,
        'donations_pending': donations_pending,
        'donations_accepted': donations_accepted,
        'total_monetary_donations': total_monetary_donations,
        'total_subsidized_value': total_subsidized_value,
        'allocations_total': allocations_total,
        'allocations_recent': allocations_recent,
        'allocations_pending_ack': allocations_pending_ack,
        'discussions_total': discussions_total,
        'discussions_pending': discussions_pending,
        'discussions_in_progress': discussions_in_progress,
        'transactions_total': transactions_total,
        'transactions_pending': transactions_pending,
        'transactions_completed': transactions_completed,
        # New donation system stats
        'unspecified_total': unspecified_total,
        'unspecified_pending_foodbank': unspecified_pending_foodbank,
        'unspecified_pending_recipient': unspecified_pending_recipient,
        'unspecified_received': unspecified_received,
        'subsidized_total': subsidized_total,
        'subsidized_pending': subsidized_pending,
        'subsidized_with_responses': subsidized_with_responses,
        'direct_donations_total': direct_donations_total,
        'direct_donations_pending': direct_donations_pending,
        'direct_donations_accepted': direct_donations_accepted,
        'request_management_total': request_management_total,
        'request_management_pending': request_management_pending,
        'request_management_fulfilled': request_management_fulfilled,
        'request_management_anonymous': request_management_anonymous,
        'donation_responses_total': donation_responses_total,
        'donation_responses_accepted': donation_responses_accepted,
        'donation_responses_declined': donation_responses_declined,
        'donation_responses_recent': donation_responses_recent,
        'donation_types_breakdown': donation_types_breakdown,
        'top_donors': top_donors,
        'top_foodbanks': top_foodbanks,
        'recent_recipient_requests': recent_recipient_requests,
        'recent_donations': recent_donations,
        'recent_unspecified': recent_unspecified,
    }
    return render(request, 'custom_admin/donations_overview_dashboard.html', context)


@staff_member_required
def unspecified_donations_management(request):
    """
    Management of unspecified/general donations not tied to specific requests
    """
    donations = UnspecifiedDonationManagement.objects.select_related(
        'donation__donor',
        'donation__foodbank',
        'accepted_by_recipient'
    ).order_by('-created_at')
    
    # Statistics
    total_donations = donations.count()
    pending_foodbank = donations.filter(foodbank_status='pending_foodbank').count()
    accepted_by_foodbank = donations.filter(foodbank_status='accepted_by_foodbank').count()
    declined_by_foodbank = donations.filter(foodbank_status='declined_by_foodbank').count()
    pending_recipient = donations.filter(recipient_status='pending_recipient').count()
    accepted_by_recipient = donations.filter(recipient_status='accepted_by_recipient').count()
    received = donations.filter(recipient_status='received').count()
    
    # Filtering (use .get with default '' so current_filters always has strings for form value persistence)
    foodbank_status_filter = request.GET.get('foodbank_status') or ''
    recipient_status_filter = request.GET.get('recipient_status') or ''
    foodbank_filter = request.GET.get('foodbank') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    search = (request.GET.get('search') or '').strip()
    donation_type_filter = request.GET.get('donation_type') or ''
    category_filter = request.GET.get('category') or ''
    delivery_filter = request.GET.get('delivery') or ''

    if foodbank_status_filter:
        donations = donations.filter(foodbank_status=foodbank_status_filter)
    if recipient_status_filter:
        donations = donations.filter(recipient_status=recipient_status_filter)
    if foodbank_filter:
        donations = donations.filter(donation__foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(created_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(donation__donor__email__icontains=search) |
            Q(donation__donor__donor_profile__full_name__icontains=search) |
            Q(donation__donor__donor_profile__organization_name__icontains=search) |
            Q(donation__foodbank__foodbank_name__icontains=search) |
            Q(donation__item_name__icontains=search) |
            Q(donation__message__icontains=search)
        )
    if donation_type_filter:
        donations = donations.filter(donation__donation_type=donation_type_filter)
    if category_filter:
        donations = donations.filter(donation__donation_category=category_filter)
    if delivery_filter:
        donations = donations.filter(donation__delivery_method=delivery_filter)

    # Pagination
    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    pagination_query_string = get_copy.urlencode()

    context = {
        'title': 'Unspecified Donations Management',
        'page_obj': page_obj,
        'total_donations': total_donations,
        'pending_foodbank': pending_foodbank,
        'accepted_by_foodbank': accepted_by_foodbank,
        'declined_by_foodbank': declined_by_foodbank,
        'pending_recipient': pending_recipient,
        'accepted_by_recipient': accepted_by_recipient,
        'received': received,
        'foodbanks': foodbanks,
        'donation_type_choices': Donation.DONATION_TYPES,
        'category_choices': Donation.DONATION_CATEGORIES,
        'delivery_choices': Donation.DELIVERY_METHODS,
        'pagination_query_string': pagination_query_string,
        'current_filters': {
            'foodbank_status': foodbank_status_filter,
            'recipient_status': recipient_status_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'donation_type': donation_type_filter,
            'category': category_filter,
            'delivery': delivery_filter,
        }
    }
    return render(request, 'custom_admin/unspecified_donations_management.html', context)


def _get_unspecified_filtered_queryset(request):
    """Shared filter pipeline for unspecified donations management exports."""
    donations = UnspecifiedDonationManagement.objects.select_related(
        'donation__donor__donor_profile',
        'donation__foodbank',
        'accepted_by_recipient',
        'accepted_by_recipient__user',
    ).order_by('-created_at')

    foodbank_status_filter = request.GET.get('foodbank_status') or ''
    recipient_status_filter = request.GET.get('recipient_status') or ''
    foodbank_filter = request.GET.get('foodbank') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    search = (request.GET.get('search') or '').strip()
    donation_type_filter = request.GET.get('donation_type') or ''
    category_filter = request.GET.get('category') or ''
    delivery_filter = request.GET.get('delivery') or ''

    if foodbank_status_filter:
        donations = donations.filter(foodbank_status=foodbank_status_filter)
    if recipient_status_filter:
        donations = donations.filter(recipient_status=recipient_status_filter)
    if foodbank_filter:
        donations = donations.filter(donation__foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(created_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(donation__donor__email__icontains=search) |
            Q(donation__donor__donor_profile__full_name__icontains=search) |
            Q(donation__donor__donor_profile__organization_name__icontains=search) |
            Q(donation__foodbank__foodbank_name__icontains=search) |
            Q(donation__item_name__icontains=search) |
            Q(donation__message__icontains=search) |
            Q(donation__other_description__icontains=search) |
            Q(donation__csr_description__icontains=search)
        )
    if donation_type_filter:
        donations = donations.filter(donation__donation_type=donation_type_filter)
    if category_filter:
        donations = donations.filter(donation__donation_category=category_filter)
    if delivery_filter:
        donations = donations.filter(donation__delivery_method=delivery_filter)

    return donations


def _get_unspecified_neutral_status(item):
    from custom_admin.utils import get_neutral_status
    fb = getattr(item, 'foodbank_status', '') or ''
    rec = getattr(item, 'recipient_status', '') or ''
    return get_neutral_status('unspecified', fb, {'foodbank_status': fb, 'recipient_status': rec})


def _build_unspecified_export_row(item, idx):
    """Build one export row that mirrors the admin unspecified table display."""
    donation = item.donation

    date_display = item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else 'N/A'

    profile = getattr(donation.donor, 'donor_profile', None)
    if profile and getattr(profile, 'organization_name', None):
        donor_name = profile.organization_name
    elif profile and getattr(profile, 'full_name', None):
        donor_name = profile.full_name
    else:
        donor_name = donation.donor.email
    donor_email = donation.donor.email
    donor_display = donor_name if donor_name == donor_email else f"{donor_name}\n{donor_email}"

    if donation.donation_type == 'item':
        if donation.donation_mode == 'free':
            type_display = 'Free Goods'
        elif donation.donation_mode == 'subsidized':
            type_display = 'Subsidized'
        else:
            type_display = donation.get_donation_mode_display() or donation.get_donation_type_display()
    else:
        type_display = donation.get_donation_type_display() or 'Unknown'

    category_display = donation.get_donation_category_display() if donation.donation_category else 'Uncategorized'

    if donation.donation_type == 'item':
        if donation.donation_mode == 'free':
            description = donation.item_name or 'General donation'
        else:
            description = donation.subsidized_product_type or donation.item_name or 'Subsidized goods'
    elif donation.donation_type == 'subsidized':
        description = donation.subsidized_product_type or 'Subsidized goods'
    elif donation.donation_type == 'csr':
        description = donation.csr_description or 'CSR initiative'
    elif donation.donation_type == 'other':
        description = donation.other_description or 'Other donation'
    elif donation.donation_type == 'money':
        description = donation.message or 'Monetary donation'
    else:
        description = 'No description'

    foodbank_name = donation.foodbank.foodbank_name if donation.foodbank else 'N/A'

    if donation.donation_type == 'item':
        qty_amount = f"{donation.quantity} {donation.quantity_unit or 'units'}"
    elif donation.donation_type == 'money':
        qty_amount = f"KES {donation.amount:,.0f}" if donation.amount is not None else '-'
    elif donation.donation_type in ('csr', 'other'):
        parts = []
        if donation.quantity:
            parts.append(f"{donation.quantity} {donation.quantity_unit or 'units'}")
        if donation.amount:
            parts.append(f"KES {donation.amount:,.0f}")
        qty_amount = "\n".join(parts) if parts else '-'
    else:
        qty_amount = '-'

    if donation.delivery_method:
        delivery_display = donation.get_delivery_method_display() or donation.delivery_method
    elif donation.donation_category == 'monetary' or donation.donation_type == 'money':
        delivery_display = 'Not applicable'
    else:
        delivery_display = 'Not specified'

    location = donation.foodbank.address if donation.foodbank and donation.foodbank.address else 'Location not provided'

    if item.accepted_by_recipient:
        recipient_name = item.accepted_by_recipient.full_name or item.accepted_by_recipient.user.email
        recipient_display = f"Claimed - {recipient_name}"
    else:
        recipient_display = 'Unclaimed'

    donor_note = donation.message or donation.other_description or 'No donor note'
    recipient_note = item.recipient_notes or 'No recipient note'
    notes_display = f"Donor: {donor_note}\nRecipient: {recipient_note}"
    decline_note = (item.recipient_decline_reason or item.foodbank_decline_reason or 'No decline note').strip()

    status_display = _get_unspecified_neutral_status(item)

    return {
        's_no': idx,
        'date': date_display,
        'donor': donor_display,
        'type': type_display,
        'category': category_display,
        'description': description,
        'foodbank': foodbank_name,
        'qty_amount': qty_amount,
        'delivery': delivery_display,
        'location': location,
        'recipient': recipient_display,
        'notes': notes_display,
        'decline_note': decline_note,
        'status': status_display,
    }


@staff_member_required
def export_unspecified_donations_csv(request):
    """
    Export unspecified donations table to CSV with applied filters.
    Mirrors the same row mapping used by Excel/PDF exports.
    """
    donations = _get_unspecified_filtered_queryset(request)
    data_rows = [_build_unspecified_export_row(item, idx) for idx, item in enumerate(donations, start=1)]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="unspecified_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Unspecified Donations Report'])
    writer.writerow([f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'])
    writer.writerow([])

    headers = ['S/No', 'Date', 'Donor', 'Type', 'Category', 'Description', 'Food Bank',
               'Qty/Amount', 'Delivery', 'Location', 'Recipient', 'Notes', 'Decline Note', 'Status']
    writer.writerow(headers)

    for row in data_rows:
        writer.writerow([
            row['s_no'],
            row['date'],
            (row['donor'] or '').replace('\n', ' | '),
            row['type'],
            row['category'],
            (row['description'] or '').replace('\n', ' '),
            row['foodbank'],
            (row['qty_amount'] or '').replace('\n', ' | '),
            row['delivery'],
            (row['location'] or '').replace('\n', ' '),
            row['recipient'],
            (row['notes'] or '').replace('\n', ' | '),
            (row['decline_note'] or '').replace('\n', ' | '),
            row['status'],
        ])

    writer.writerow([])
    writer.writerow(['Total Records', len(data_rows)])

    return response


@staff_member_required
def export_unspecified_donations_excel(request):
    """
    Export unspecified donations table to Excel with applied filters
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    donations = _get_unspecified_filtered_queryset(request)
    data_rows = [_build_unspecified_export_row(item, idx) for idx, item in enumerate(donations, start=1)]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Unspecified Donations"
    
    # Define styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = "Unspecified Donations Report"
    title_cell.font = Font(size=16, bold=True, color="2C3E50")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtitle with date
    ws.merge_cells('A2:Q2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    subtitle_cell.font = Font(size=10, italic=True, color="7F8C8D")
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Headers - mirror admin table columns (excluding Actions)
    headers = ['S/No', 'Date', 'Donor', 'Type', 'Category', 'Description', 'Food Bank',
               'Qty/Amount', 'Delivery', 'Location', 'Recipient', 'Notes', 'Decline Note', 'Status']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row_num = 5
    for row in data_rows:
        row_data = [
            row['s_no'], row['date'], row['donor'], row['type'], row['category'], row['description'],
            row['foodbank'], row['qty_amount'], row['delivery'], row['location'],
            row['recipient'], row['notes'], row['decline_note'], row['status'],
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        row_num += 1
    
    # Adjust column widths
    column_widths = [6, 16, 28, 14, 14, 30, 24, 18, 14, 24, 24, 32, 32, 26]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Auto-filter and freeze panes
    if data_rows:
        ws.auto_filter.ref = f"A4:N{row_num - 1}"
    ws.freeze_panes = "A5"
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="unspecified_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@staff_member_required
def export_unspecified_donations_pdf(request):
    """
    Export unspecified donations table to PDF with applied filters
    """
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from datetime import datetime
    
    donations = _get_unspecified_filtered_queryset(request)
    data_rows = [_build_unspecified_export_row(item, idx) for idx, item in enumerate(donations, start=1)]

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), 
                           rightMargin=0.5*inch, leftMargin=0.5*inch,
                           topMargin=0.75*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Title
    elements.append(Paragraph("Unspecified Donations Report", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    
    wrap = ParagraphStyle('Wrap', parent=styles['BodyText'], fontSize=7, leading=9, wordWrap='CJK')
    
    # Table data - mirror admin table columns (excluding Actions)
    table_data = [['S/No', 'Date', 'Donor', 'Type', 'Category', 'Description',
                   'Food Bank', 'Qty/Amount', 'Delivery', 'Location', 'Recipient', 'Notes', 'Decline Note', 'Status']]

    for row in data_rows:
        table_data.append([
            str(row['s_no']),
            Paragraph(row['date'], wrap),
            Paragraph((row['donor'] or '-').replace('\n', '<br/>')[:90], wrap),
            Paragraph((row['type'] or '-')[:35], wrap),
            Paragraph((row['category'] or '-')[:35], wrap),
            Paragraph((row['description'] or '-').replace('\n', '<br/>')[:120], wrap),
            Paragraph((row['foodbank'] or '-')[:60], wrap),
            Paragraph((row['qty_amount'] or '-').replace('\n', '<br/>')[:70], wrap),
            Paragraph((row['delivery'] or '-')[:30], wrap),
            Paragraph((row['location'] or '-')[:70], wrap),
            Paragraph((row['recipient'] or '-')[:70], wrap),
            Paragraph((row['notes'] or '-').replace('\n', '<br/>')[:160], wrap),
            Paragraph((row['decline_note'] or '-').replace('\n', '<br/>')[:120], wrap),
            Paragraph((row['status'] or '-')[:60], wrap),
        ])
    
    # Create table and stretch columns to full A3 usable width
    base_col_widths = [
        0.30*inch, 0.55*inch, 1.10*inch, 0.65*inch, 0.65*inch, 1.20*inch,
        0.95*inch, 0.80*inch, 0.65*inch, 0.85*inch, 0.95*inch, 1.10*inch, 1.10*inch, 1.00*inch
    ]
    usable_width = landscape(A3)[0] - doc.leftMargin - doc.rightMargin
    total_base_width = sum(base_col_widths)
    scale_factor = (usable_width / total_base_width) if total_base_width else 1
    col_widths = [w * scale_factor for w in base_col_widths]

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    elements.append(table)
    
    # Add summary
    elements.append(Spacer(1, 0.3*inch))
    summary_text = f"Total Records: {len(data_rows)}"
    elements.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF from buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="unspecified_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)
    
    return response


@staff_member_required
def unspecified_donation_details(request, donation_id):
    """
    API endpoint to fetch detailed information about an unspecified donation
    including tracking history and comments
    """
    try:
        donation_obj = UnspecifiedDonationManagement.objects.select_related(
            'donation__donor__donor_profile',
            'donation__foodbank',
            'accepted_by_recipient'
        ).get(id=donation_id)
        
        donation = donation_obj.donation
        
        # Build tracking history
        tracking_history = []
        
        # Created event
        try:
            donor_message = getattr(donation, 'message', None) or getattr(donation, 'other_description', None) or getattr(donation, 'csr_description', None)
            tracking_history.append({
                'status': 'Created',
                'description': 'Donation created by donor',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'badge_class': 'bg-secondary',
                'comment': donor_message if donor_message else None
            })
        except:
            pass
        
        # Foodbank status changes
        if donation_obj.foodbank_status == 'accepted_by_foodbank':
            try:
                tracking_history.append({
                    'status': 'Accepted by Foodbank',
                    'description': f'Foodbank {donation.foodbank.foodbank_name} accepted the donation',
                    'timestamp': donation_obj.foodbank_reviewed_at.strftime('%b %d, %Y %H:%M') if donation_obj.foodbank_reviewed_at else (donation_obj.updated_at.strftime('%b %d, %Y %H:%M') if donation_obj.updated_at else 'N/A'),
                    'updated_by': 'Foodbank Admin',
                    'badge_class': 'bg-success',
                    'comment': None
                })
            except:
                pass
        elif donation_obj.foodbank_status == 'declined_by_foodbank':
            try:
                tracking_history.append({
                    'status': 'Declined by Foodbank',
                    'description': f'Foodbank {donation.foodbank.foodbank_name} declined the donation',
                    'timestamp': donation_obj.foodbank_reviewed_at.strftime('%b %d, %Y %H:%M') if donation_obj.foodbank_reviewed_at else (donation_obj.updated_at.strftime('%b %d, %Y %H:%M') if donation_obj.updated_at else 'N/A'),
                    'updated_by': 'Foodbank Admin',
                    'badge_class': 'bg-danger',
                    'comment': donation_obj.foodbank_decline_reason if donation_obj.foodbank_decline_reason else None
                })
            except:
                pass
        
        # Recipient status changes
        if donation_obj.accepted_by_recipient:
            try:
                tracking_history.append({
                    'status': 'Assigned to Recipient',
                    'description': f'Assigned to {donation_obj.accepted_by_recipient.full_name}',
                    'timestamp': donation_obj.recipient_accepted_at.strftime('%b %d, %Y %H:%M') if donation_obj.recipient_accepted_at else (donation_obj.updated_at.strftime('%b %d, %Y %H:%M') if donation_obj.updated_at else 'N/A'),
                    'updated_by': donation_obj.accepted_by_recipient.full_name,
                    'badge_class': 'bg-info',
                    'comment': donation_obj.recipient_notes if donation_obj.recipient_notes else None
                })
            except:
                pass
        
        if donation_obj.recipient_status == 'received' and donation_obj.accepted_by_recipient:
            try:
                tracking_history.append({
                    'status': 'Received',
                    'description': f'Recipient {donation_obj.accepted_by_recipient.full_name} confirmed receipt',
                    'timestamp': donation_obj.received_at.strftime('%b %d, %Y %H:%M') if donation_obj.received_at else (donation_obj.updated_at.strftime('%b %d, %Y %H:%M') if donation_obj.updated_at else 'N/A'),
                    'updated_by': donation_obj.accepted_by_recipient.full_name,
                    'badge_class': 'bg-success',
                    'comment': donation_obj.recipient_notes if donation_obj.recipient_notes else None
                })
            except:
                pass
        
        # Safely get donation details
        if donation.donation_type == 'item':
            details = f"{getattr(donation, 'quantity', 0)} {getattr(donation, 'quantity_unit', '')} of {getattr(donation, 'item_name', 'Unknown')}"
        elif donation.donation_type == 'money':
            details = f"KES {getattr(donation, 'amount', 0)}"
        else:
            details = donation.get_donation_type_display()
        
        # Build response data
        data = {
            'success': True,
            'donation': {
                'id': donation_obj.id,
                'donation_type': donation.get_donation_type_display(),
                'details': details,
                'amount': getattr(donation, 'amount', None) if donation.donation_type == 'money' else None,
                'expiry_date': donation.expiry_date.strftime('%b %d, %Y') if hasattr(donation, 'expiry_date') and donation.expiry_date else None,
                'created_at': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_at': donation_obj.updated_at.strftime('%b %d, %Y %H:%M') if donation_obj.updated_at else 'N/A',
                'delivery_method': getattr(donation, 'delivery_method', None),
                'location': getattr(donation.donor.donor_profile, 'location', None) if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else None,
                
                # Donor info
                'donor_name': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'donor_email': donation.donor.email,
                'donor_phone': getattr(donation.donor, 'phone_number', None),
                'donor_organization': getattr(donation.donor.donor_profile, 'organization_name', None) if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile and getattr(donation.donor.donor_profile, 'is_organization', False) else None,
                
                # Foodbank info
                'foodbank_name': donation.foodbank.foodbank_name,
                'foodbank_status': donation_obj.get_foodbank_status_display(),
                'foodbank_status_class': donation_obj.get_foodbank_status_badge_class() if hasattr(donation_obj, 'get_foodbank_status_badge_class') else 'bg-secondary',
                'foodbank_contact': getattr(donation.foodbank, 'contact_person', None),
                
                # Recipient info
                'recipient_name': donation_obj.accepted_by_recipient.full_name if donation_obj.accepted_by_recipient else None,
                'recipient_status': donation_obj.get_recipient_status_display(),
                'recipient_status_class': donation_obj.get_recipient_status_badge_class() if hasattr(donation_obj, 'get_recipient_status_badge_class') else 'bg-secondary',
                'received_date': donation_obj.updated_at.strftime('%b %d, %Y') if donation_obj.recipient_status == 'received' and donation_obj.updated_at else None,
                
                # Tracking history
                'tracking_history': tracking_history
            }
        }
        
        return JsonResponse(data)
        
    except UnspecifiedDonationManagement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Donation not found'
        }, status=404)
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


@staff_member_required
def export_donation_details_pdf(request, donation_id):
    """
    Export donation details to a stylish PDF report
    """
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    from datetime import datetime
    
    try:
        donation_obj = UnspecifiedDonationManagement.objects.select_related(
            'donation__donor__donor_profile',
            'donation__foodbank',
            'accepted_by_recipient'
        ).get(id=donation_id)
        
        donation = donation_obj.donation
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch,
                                topMargin=1*inch, bottomMargin=0.75*inch)
        
        # Container for PDF elements
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=6
        )
        
        # Title
        elements.append(Paragraph("Donation Details Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                                 ParagraphStyle('Subtitle', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))
        elements.append(Spacer(1, 0.3*inch))
        
        # Donation Information Section
        elements.append(Paragraph("Donation Information", heading_style))
        
        donation_data = [
            ['Donation ID:', f"#{donation_obj.id}"],
            ['Type:', donation.get_donation_type_display()],
            ['Details:', f"{donation.quantity} {donation.quantity_unit} of {donation.item_name}" if donation.donation_type == 'item' else f"KES {donation.amount}"],
            ['Created:', donation.donated_at.strftime('%B %d, %Y at %I:%M %p') if donation.donated_at else 'N/A'],
            ['Last Updated:', donation_obj.updated_at.strftime('%B %d, %Y at %I:%M %p') if donation_obj.updated_at else 'N/A'],
        ]
        
        if hasattr(donation, 'expiry_date') and donation.expiry_date:
            donation_data.append(['Expiry Date:', donation.expiry_date.strftime('%B %d, %Y')])
        
        if hasattr(donation, 'delivery_method') and donation.delivery_method:
            donation_data.append(['Delivery Method:', donation.delivery_method.title()])
        
        donation_table = Table(donation_data, colWidths=[2*inch, 4.5*inch])
        donation_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(donation_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Donor Information Section
        elements.append(Paragraph("Donor Information", heading_style))
        
        donor_data = [
            ['Name:', donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email],
            ['Email:', donation.donor.email],
            ['Phone:', getattr(donation.donor, 'phone_number', 'N/A')],
        ]
        
        if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile and getattr(donation.donor.donor_profile, 'is_organization', False):
            donor_data.append(['Organization:', getattr(donation.donor.donor_profile, 'organization_name', 'N/A')])
        
        if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile and hasattr(donation.donor.donor_profile, 'location'):
            donor_data.append(['Location:', getattr(donation.donor.donor_profile, 'location', 'N/A')])
        
        donor_table = Table(donor_data, colWidths=[2*inch, 4.5*inch])
        donor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f5e9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a5d6a7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(donor_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Foodbank Information Section
        elements.append(Paragraph("Foodbank Information", heading_style))
        
        foodbank_data = [
            ['Foodbank:', donation.foodbank.foodbank_name],
            ['Status:', donation_obj.get_foodbank_status_display()],
        ]
        
        if hasattr(donation.foodbank, 'contact_person') and donation.foodbank.contact_person:
            foodbank_data.append(['Contact Person:', donation.foodbank.contact_person])
        
        foodbank_table = Table(foodbank_data, colWidths=[2*inch, 4.5*inch])
        foodbank_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fff3e0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ffcc80')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(foodbank_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Recipient Information Section (if applicable)
        if donation_obj.accepted_by_recipient:
            elements.append(Paragraph("Recipient Information", heading_style))
            
            recipient_data = [
                ['Name:', donation_obj.accepted_by_recipient.full_name],
                ['Status:', donation_obj.get_recipient_status_display()],
            ]
            
            if donation_obj.recipient_status == 'received' and donation_obj.received_at:
                recipient_data.append(['Received Date:', donation_obj.received_at.strftime('%B %d, %Y')])
            
            recipient_table = Table(recipient_data, colWidths=[2*inch, 4.5*inch])
            recipient_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e3f2fd')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#90caf9')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(recipient_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Tracking History Section
        elements.append(Paragraph("Tracking History & Comments", heading_style))
        
        # Build tracking history
        tracking_history = []
        
        # Created event
        try:
            donor_message = getattr(donation, 'message', None) or getattr(donation, 'other_description', None) or getattr(donation, 'csr_description', None)
            tracking_history.append({
                'status': 'Created',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'comment': donor_message if donor_message else None
            })
        except:
            pass
        
        # Foodbank status
        if donation_obj.foodbank_status == 'accepted_by_foodbank':
            try:
                tracking_history.append({
                    'status': 'Accepted by Foodbank',
                    'timestamp': donation_obj.foodbank_reviewed_at.strftime('%b %d, %Y %H:%M') if donation_obj.foodbank_reviewed_at else 'N/A',
                    'updated_by': 'Foodbank Admin',
                    'comment': None
                })
            except:
                pass
        elif donation_obj.foodbank_status == 'declined_by_foodbank':
            try:
                tracking_history.append({
                    'status': 'Declined by Foodbank',
                    'timestamp': donation_obj.foodbank_reviewed_at.strftime('%b %d, %Y %H:%M') if donation_obj.foodbank_reviewed_at else 'N/A',
                    'updated_by': 'Foodbank Admin',
                    'comment': donation_obj.foodbank_decline_reason if donation_obj.foodbank_decline_reason else None
                })
            except:
                pass
        
        # Recipient status
        if donation_obj.accepted_by_recipient:
            try:
                tracking_history.append({
                    'status': 'Assigned to Recipient',
                    'timestamp': donation_obj.recipient_accepted_at.strftime('%b %d, %Y %H:%M') if donation_obj.recipient_accepted_at else 'N/A',
                    'updated_by': donation_obj.accepted_by_recipient.full_name,
                    'comment': donation_obj.recipient_notes if donation_obj.recipient_notes else None
                })
            except:
                pass
        
        if donation_obj.recipient_status == 'received' and donation_obj.accepted_by_recipient:
            try:
                tracking_history.append({
                    'status': 'Received',
                    'timestamp': donation_obj.received_at.strftime('%b %d, %Y %H:%M') if donation_obj.received_at else 'N/A',
                    'updated_by': donation_obj.accepted_by_recipient.full_name,
                    'comment': donation_obj.recipient_notes if donation_obj.recipient_notes else None
                })
            except:
                pass
        
        # Create tracking history table
        if tracking_history:
            for idx, event in enumerate(tracking_history):
                # Event header
                event_data = [
                    [Paragraph(f"<b>{event['status']}</b>", normal_style)],
                    [Paragraph(f"<i>Time:</i> {event['timestamp']}", normal_style)],
                    [Paragraph(f"<i>By:</i> {event['updated_by']}", normal_style)],
                ]
                
                if event['comment']:
                    event_data.append([Paragraph(f"<i>Comment:</i> {event['comment']}", normal_style)])
                
                event_table = Table(event_data, colWidths=[6.5*inch])
                event_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ]))
                elements.append(event_table)
                
                if idx < len(tracking_history) - 1:
                    elements.append(Spacer(1, 0.1*inch))
        else:
            elements.append(Paragraph("No tracking history available", normal_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="donation_{donation_obj.id}_details.pdf"'
        response.write(pdf)
        
        return response
        
    except UnspecifiedDonationManagement.DoesNotExist:
        return HttpResponse("Donation not found", status=404)
    except Exception as e:
        import traceback
        return HttpResponse(f"Error generating PDF: {str(e)}\n{traceback.format_exc()}", status=500)


@staff_member_required
def subsidized_donations_management(request):
    """
    Management of subsidized goods donations
    """
    donations = Donation.objects.filter(
        donation_type='subsidized'
    ).select_related(
        'donor',
        'foodbank',
        'accepted_by_recipient',
        'declined_by_recipient'
    ).prefetch_related('subsidized_responded_by').order_by('-donated_at')
    
    # Statistics
    total_donations = donations.count()
    pending = donations.filter(status='pending').count()
    accepted = donations.filter(status='accepted').count()
    declined = donations.filter(status='declined').count()
    total_value = donations.filter(status='accepted').aggregate(
        total=Sum('subsidized_price')
    )['total'] or 0
    
    # Filtering (normalize to '' so current_filters always has strings for form value persistence)
    status_filter = request.GET.get('status') or ''
    foodbank_filter = request.GET.get('foodbank') or ''
    has_responses = request.GET.get('has_responses') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    search = (request.GET.get('search') or '').strip()
    category_filter = request.GET.get('category') or ''
    delivery_filter = request.GET.get('delivery') or ''
    recipient_claimed_filter = request.GET.get('recipient_claimed') or ''

    if status_filter:
        donations = donations.filter(status=status_filter)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if has_responses == 'yes':
        donations = donations.filter(subsidized_responded_by__isnull=False).distinct()
    elif has_responses == 'no':
        donations = donations.filter(subsidized_responded_by__isnull=True)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(donor__donor_profile__organization_name__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(subsidized_product_type__icontains=search) |
            Q(message__icontains=search)
        )
    if category_filter:
        donations = donations.filter(donation_category=category_filter)
    if delivery_filter:
        donations = donations.filter(delivery_method=delivery_filter)
    if recipient_claimed_filter == 'yes':
        donations = donations.filter(accepted_by_recipient__isnull=False)
    elif recipient_claimed_filter == 'no':
        donations = donations.filter(accepted_by_recipient__isnull=True)

    # Pagination
    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    donation_list = list(page_obj.object_list)
    donation_ids = [donation.id for donation in donation_list]
    recipient_notes_by_donation = {}

    if donation_ids:
        donation_responses = DonationResponse.objects.filter(
            donation_id__in=donation_ids
        ).exclude(
            notes__isnull=True
        ).exclude(
            notes__exact=''
        ).order_by('-responded_at')

        for response in donation_responses:
            if response.donation_id not in recipient_notes_by_donation:
                recipient_notes_by_donation[response.donation_id] = response.notes

        for donation in donation_list:
            donation.latest_recipient_note = recipient_notes_by_donation.get(donation.id)
    
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    pagination_query_string = get_copy.urlencode()

    recipient_claimed_choices = [
        ('', 'All'),
        ('yes', 'Claimed by recipient'),
        ('no', 'Not claimed'),
    ]

    context = {
        'title': 'Subsidized Donations Management',
        'page_obj': page_obj,
        'total_donations': total_donations,
        'pending': pending,
        'accepted': accepted,
        'declined': declined,
        'total_value': total_value,
        'foodbanks': foodbanks,
        'category_choices': Donation.DONATION_CATEGORIES,
        'delivery_choices': Donation.DELIVERY_METHODS,
        'recipient_claimed_choices': recipient_claimed_choices,
        'pagination_query_string': pagination_query_string,
        'current_filters': {
            'status': status_filter,
            'foodbank': foodbank_filter,
            'has_responses': has_responses,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'category': category_filter,
            'delivery': delivery_filter,
            'recipient_claimed': recipient_claimed_filter,
        }
    }
    return render(request, 'custom_admin/subsidized_donations_management.html', context)


@staff_member_required
def subsidized_donation_details(request, donation_id):
    """
    API endpoint to fetch detailed information about a subsidized donation
    """
    try:
        donation = Donation.objects.select_related(
            'donor__donor_profile',
            'foodbank',
            'accepted_by_recipient',
            'declined_by_recipient'
        ).prefetch_related('subsidized_responded_by').get(id=donation_id, donation_type='subsidized')
        
        # Build tracking history
        tracking_history = []
        
        # Created event
        try:
            donor_message = getattr(donation, 'message', None)
            tracking_history.append({
                'status': 'Created',
                'description': 'Subsidized donation created',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'badge_class': 'bg-secondary',
                'comment': donor_message if donor_message else None
            })
        except:
            pass
        
        # Status changes
        if donation.status == 'accepted':
            tracking_history.append({
                'status': 'Accepted',
                'description': 'Donation accepted',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': 'System',
                'badge_class': 'bg-success',
                'comment': None
            })
        elif donation.status == 'declined':
            tracking_history.append({
                'status': 'Declined',
                'description': 'Donation declined',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': 'System',
                'badge_class': 'bg-danger',
                'comment': donation.decline_message if donation.decline_message else None
            })
        
        # Recipients list
        recipients = []
        for recipient_user in donation.subsidized_responded_by.all():
            if hasattr(recipient_user, 'recipient_profile'):
                recipients.append({
                    'name': recipient_user.recipient_profile.full_name,
                    'status': 'Responded',
                    'status_class': 'bg-info'
                })
        
        if donation.accepted_by_recipient:
            recipients.append({
                'name': donation.accepted_by_recipient.full_name,
                'status': 'Accepted',
                'status_class': 'bg-success'
            })
        
        # Calculate discount percentage
        discount_display = '0'
        if donation.subsidized_market_price and donation.subsidized_price and donation.subsidized_market_price > 0:
            discount = ((donation.subsidized_market_price - donation.subsidized_price) / donation.subsidized_market_price) * 100
            discount_display = f"{discount:.1f}"
        
        # Build response data
        data = {
            'success': True,
            'donation': {
                'id': donation.id,
                'donation_type': donation.get_donation_type_display(),
                'product': donation.subsidized_product_type,
                'quantity': f"{donation.subsidized_quantity} {donation.subsidized_quantity_unit}" if donation.subsidized_quantity else None,
                'price': str(donation.subsidized_price) if donation.subsidized_price else '0.00',
                'discount': discount_display,
                'market_price': str(donation.subsidized_market_price) if donation.subsidized_market_price else None,
                'status': donation.get_status_display(),
                'status_class': 'bg-success' if donation.status == 'accepted' else ('bg-danger' if donation.status == 'declined' else 'bg-warning'),
                'created_at': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_at': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                
                # Donor info
                'donor_name': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'donor_email': donation.donor.email,
                'donor_phone': getattr(donation.donor, 'phone_number', None),
                'donor_organization': getattr(donation.donor.donor_profile, 'organization_name', None) if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile and getattr(donation.donor.donor_profile, 'is_organization', False) else None,
                
                # Foodbank info
                'foodbank_name': donation.foodbank.foodbank_name,
                'foodbank_contact': getattr(donation.foodbank, 'contact_person', None),
                
                # Recipients
                'recipients': recipients,
                
                # Tracking history
                'tracking_history': tracking_history
            }
        }
        
        return JsonResponse(data)
        
    except Donation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Donation not found'
        }, status=404)
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


@staff_member_required
def export_subsidized_donation_details_pdf(request, donation_id):
    """
    Export subsidized donation details to PDF
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from datetime import datetime
    
    try:
        donation = Donation.objects.select_related(
            'donor__donor_profile',
            'foodbank'
        ).prefetch_related('subsidized_responded_by').get(id=donation_id, donation_type='subsidized')
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch,
                                topMargin=1*inch, bottomMargin=0.75*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=6
        )
        
        # Title
        elements.append(Paragraph("Subsidized Donation Details Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                                 ParagraphStyle('Subtitle', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))
        elements.append(Spacer(1, 0.3*inch))
        
        # Donation Information
        elements.append(Paragraph("Donation Information", heading_style))
        
        # Calculate discount percentage
        discount_display = 'N/A'
        if donation.subsidized_market_price and donation.subsidized_price and donation.subsidized_market_price > 0:
            discount = ((donation.subsidized_market_price - donation.subsidized_price) / donation.subsidized_market_price) * 100
            discount_display = f"{discount:.1f}%"
        
        donation_data = [
            ['Donation ID:', f"#{donation.id}"],
            ['Type:', donation.get_donation_type_display()],
            ['Product:', donation.subsidized_product_type or 'N/A'],
            ['Quantity:', f"{donation.subsidized_quantity} {donation.subsidized_quantity_unit}" if donation.subsidized_quantity else 'N/A'],
            ['Price:', f"KES {donation.subsidized_price}" if donation.subsidized_price else 'N/A'],
            ['Discount:', discount_display],
            ['Market Price:', f"KES {donation.subsidized_market_price}" if donation.subsidized_market_price else 'N/A'],
            ['Status:', donation.get_status_display()],
            ['Created:', donation.donated_at.strftime('%B %d, %Y at %I:%M %p') if donation.donated_at else 'N/A'],
        ]
        
        donation_table = Table(donation_data, colWidths=[2*inch, 4.5*inch])
        donation_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(donation_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Donor Information
        elements.append(Paragraph("Donor Information", heading_style))
        
        donor_data = [
            ['Name:', donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email],
            ['Email:', donation.donor.email],
            ['Phone:', getattr(donation.donor, 'phone_number', 'N/A')],
        ]
        
        donor_table = Table(donor_data, colWidths=[2*inch, 4.5*inch])
        donor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f5e9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a5d6a7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(donor_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Foodbank Information
        elements.append(Paragraph("Foodbank Information", heading_style))
        
        foodbank_data = [
            ['Foodbank:', donation.foodbank.foodbank_name],
            ['Contact:', getattr(donation.foodbank, 'contact_person', 'N/A')],
        ]
        
        foodbank_table = Table(foodbank_data, colWidths=[2*inch, 4.5*inch])
        foodbank_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fff3e0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ffcc80')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(foodbank_table)
        
        # Build PDF
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="subsidized_donation_{donation.id}_details.pdf"'
        response.write(pdf)
        
        return response
        
    except Donation.DoesNotExist:
        return HttpResponse("Donation not found", status=404)
    except Exception as e:
        import traceback
        return HttpResponse(f"Error generating PDF: {str(e)}\n{traceback.format_exc()}", status=500)


def _get_subsidized_filtered_queryset(request):
    """Return subsidized donations queryset with the same filters used in the admin table."""
    donations = Donation.objects.filter(
        donation_type='subsidized'
    ).select_related(
        'donor__donor_profile',
        'foodbank',
        'accepted_by_recipient'
    ).prefetch_related('subsidized_responded_by').order_by('-donated_at')

    status_filter = request.GET.get('status')
    foodbank_filter = request.GET.get('foodbank')
    has_responses = request.GET.get('has_responses')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    category_filter = request.GET.get('category')
    delivery_filter = request.GET.get('delivery')
    recipient_claimed_filter = request.GET.get('recipient_claimed')

    if status_filter:
        donations = donations.filter(status=status_filter)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if has_responses == 'yes':
        donations = donations.filter(subsidized_responded_by__isnull=False).distinct()
    elif has_responses == 'no':
        donations = donations.filter(subsidized_responded_by__isnull=True)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(donor__donor_profile__organization_name__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(subsidized_product_type__icontains=search) |
            Q(message__icontains=search)
        )
    if category_filter:
        donations = donations.filter(donation_category=category_filter)
    if delivery_filter:
        donations = donations.filter(delivery_method=delivery_filter)
    if recipient_claimed_filter == 'yes':
        donations = donations.filter(accepted_by_recipient__isnull=False)
    elif recipient_claimed_filter == 'no':
        donations = donations.filter(accepted_by_recipient__isnull=True)

    return donations


def _get_latest_subsidized_recipient_notes(donation_ids):
    """Map donation_id -> latest non-empty recipient note."""
    if not donation_ids:
        return {}

    notes_qs = (
        DonationResponse.objects.filter(donation_id__in=donation_ids)
        .exclude(notes__isnull=True)
        .exclude(notes__exact='')
        .order_by('-responded_at')
    )

    latest_notes = {}
    for response in notes_qs:
        if response.donation_id not in latest_notes:
            latest_notes[response.donation_id] = response.notes
    return latest_notes


@staff_member_required
def export_subsidized_donations_csv(request):
    """
    Export subsidized donations table to CSV.
    """
    from datetime import datetime

    donations = list(_get_subsidized_filtered_queryset(request))
    latest_notes = _get_latest_subsidized_recipient_notes([d.id for d in donations])

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="subsidized_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'S/No', 'Date', 'Donor', 'Food Bank', 'Location', 'Type', 'Category', 'Product',
        'Quantity', 'Unit', 'Market Price (KES)', 'Subsidy (KES)', 'New Price (KES)',
        'Status', 'Delivery', 'Recipient', 'Donor Note', 'Recipient Note', 'Decline Note'
    ])

    for idx, donation in enumerate(donations, start=1):
        date_str = donation.donated_at.strftime('%Y-%m-%d %H:%M') if donation.donated_at else 'N/A'

        if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile:
            if donation.donor.donor_profile.is_organization:
                donor_name = donation.donor.donor_profile.organization_name or donation.donor.donor_profile.full_name
            else:
                donor_name = donation.donor.donor_profile.full_name
        else:
            donor_name = donation.donor.email

        fb_name = donation.foodbank.foodbank_name if donation.foodbank else 'N/A'
        location = donation.foodbank.address if donation.foodbank and donation.foodbank.address else 'Not provided'
        type_display = 'Subsidized'
        category_display = donation.get_donation_category_display() or 'Food'
        product = donation.subsidized_product_type or 'Subsidized Goods'
        qty_value = donation.subsidized_quantity if donation.subsidized_quantity is not None else 'N/A'
        qty_unit = donation.subsidized_quantity_unit or ('units' if qty_value != 'N/A' else 'N/A')
        market_price = donation.subsidized_initial_amount or donation.subsidized_market_price or ''
        subsidy = donation.subsidized_subsidy_amount or ''
        new_price = donation.subsidized_price or ''

        if donation.status == 'pending':
            status_display = 'Awaiting foodbank'
        elif donation.status == 'accepted':
            status_display = 'Accepted by food bank'
        elif donation.status == 'fulfilled':
            status_display = 'Fulfilled & allocated'
        elif donation.status == 'declined':
            status_display = 'Declined by food bank'
        else:
            status_display = donation.get_status_display() or 'Status update pending'

        delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
        if donation.accepted_by_recipient:
            recipient_display = f"Claimed - {donation.accepted_by_recipient.full_name or donation.accepted_by_recipient.user.email}"
        else:
            recipient_display = 'Unclaimed'

        donor_note = (donation.message or donation.csr_description or donation.other_description or '').replace('\n', ' ').replace('\r', '')
        recipient_note = (latest_notes.get(donation.id, '') or '').replace('\n', ' ').replace('\r', '')
        if donation.status == 'declined':
            decline_note = (donation.decline_message or 'No decline note').replace('\n', ' ').replace('\r', '')
        else:
            decline_note = 'Not declined'

        writer.writerow([
            idx, date_str, donor_name, fb_name, location, type_display, category_display, product,
            qty_value, qty_unit, market_price, subsidy, new_price, status_display, delivery, recipient_display,
            donor_note, recipient_note, decline_note
        ])

    return response


@staff_member_required
def export_subsidized_donations_excel(request):
    """
    Export subsidized donations table to Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Get filtered donations
    donations = list(_get_subsidized_filtered_queryset(request))
    latest_notes = _get_latest_subsidized_recipient_notes([d.id for d in donations])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Subsidized Donations"
    
    # Define styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_cell.value = "Subsidized Donations Report"
    title_cell.font = Font(size=16, bold=True, color="2C3E50")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtitle
    ws.merge_cells('A2:S2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    subtitle_cell.font = Font(size=10, italic=True, color="7F8C8D")
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Headers - matching template columns exactly
    headers = ['S/No', 'Date', 'Donor', 'Food Bank', 'Location', 'Type', 'Category',
               'Product', 'Quantity', 'Unit', 'Market Price (KES)', 'Subsidy (KES)', 'New Price (KES)',
               'Status', 'Delivery', 'Recipient', 'Donor Note', 'Recipient Note', 'Decline Note']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 5
    s_no = 1
    for donation in donations:
        # Date
        date_str = donation.donated_at.strftime('%Y-%m-%d %H:%M') if donation.donated_at else 'N/A'
        
        # Donor
        if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile:
            if donation.donor.donor_profile.is_organization:
                donor_name = donation.donor.donor_profile.organization_name or donation.donor.donor_profile.full_name
            else:
                donor_name = donation.donor.donor_profile.full_name
        else:
            donor_name = donation.donor.email
        
        # Foodbank and Location
        fb_name = donation.foodbank.foodbank_name if donation.foodbank else 'N/A'
        location = donation.foodbank.address if donation.foodbank and donation.foodbank.address else 'Not provided'
        
        # Type and Category
        type_display = 'Subsidized'
        category_display = donation.get_donation_category_display() or 'Food'
        
        # Product, Quantity
        product = donation.subsidized_product_type or 'Subsidized Goods'
        qty_value = donation.subsidized_quantity if donation.subsidized_quantity is not None else 'N/A'
        qty_unit = donation.subsidized_quantity_unit or ('units' if qty_value != 'N/A' else 'N/A')
        
        # Prices
        market_price = donation.subsidized_initial_amount or donation.subsidized_market_price or ''
        subsidy = donation.subsidized_subsidy_amount or ''
        new_price = donation.subsidized_price or ''
        
        # Status
        if donation.status == 'pending':
            status_display = 'Awaiting foodbank'
        elif donation.status == 'accepted':
            status_display = 'Accepted by food bank'
        elif donation.status == 'fulfilled':
            status_display = 'Fulfilled & allocated'
        elif donation.status == 'declined':
            status_display = 'Declined by food bank'
        else:
            status_display = donation.get_status_display() or 'Status update pending'
        
        # Delivery
        delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
        
        # Recipient
        if donation.accepted_by_recipient:
            recipient_display = f"Claimed - {donation.accepted_by_recipient.full_name or donation.accepted_by_recipient.user.email}"
        else:
            recipient_display = 'Unclaimed'
        
        # Notes
        donor_note = (donation.message or donation.csr_description or donation.other_description or '').replace('\n', ' ').replace('\r', '')
        recipient_note = (latest_notes.get(donation.id, '') or '').replace('\n', ' ').replace('\r', '')
        if donation.status == 'declined':
            decline_note = (donation.decline_message or 'No decline note').replace('\n', ' ').replace('\r', '')
        else:
            decline_note = 'Not declined'
        
        row_data = [
            s_no, date_str, donor_name, fb_name, location, type_display, category_display,
            product, qty_value, qty_unit, market_price, subsidy, new_price,
            status_display, delivery, recipient_display, donor_note, recipient_note, decline_note,
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        row_num += 1
        s_no += 1
    
    # Adjust column widths
    column_widths = [6, 16, 22, 20, 22, 14, 12, 18, 10, 10, 14, 14, 14, 20, 12, 22, 28, 28, 28]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Auto-filter and freeze panes
    if donations:
        ws.auto_filter.ref = f"A4:S{row_num - 1}"
    ws.freeze_panes = "A5"
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="subsidized_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@staff_member_required
def export_subsidized_donations_pdf(request):
    """
    Export subsidized donations table to PDF
    """
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from datetime import datetime
    
    # Get filtered donations
    donations = list(_get_subsidized_filtered_queryset(request))
    latest_notes = _get_latest_subsidized_recipient_notes([d.id for d in donations])

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), 
                           rightMargin=0.5*inch, leftMargin=0.5*inch,
                           topMargin=0.75*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Title
    elements.append(Paragraph("Subsidized Donations Report", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    
    wrap = ParagraphStyle('Wrap', parent=styles['BodyText'], fontSize=7, leading=9, wordWrap='CJK')
    
    # Table data - matching template columns
    table_data = [['S/No', 'Date', 'Donor', 'Food Bank', 'Location', 'Type', 'Category', 'Product',
                   'Quantity', 'Market Price', 'Subsidy', 'New Price', 'Status', 'Delivery',
                   'Recipient', 'Notes', 'Decline Note']]
    
    s_no = 1
    for donation in donations:
        # Date
        date_str = donation.donated_at.strftime('%b %d, %Y') if donation.donated_at else 'N/A'
        
        # Donor
        if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile:
            if donation.donor.donor_profile.is_organization:
                donor_name = donation.donor.donor_profile.organization_name or donation.donor.donor_profile.full_name
            else:
                donor_name = donation.donor.donor_profile.full_name
        else:
            donor_name = donation.donor.email
        
        # Foodbank and Location
        fb_name = donation.foodbank.foodbank_name if donation.foodbank else 'N/A'
        location = (donation.foodbank.address or 'Not provided')[:25] if donation.foodbank else 'N/A'
        
        # Type, Category and Product
        type_display = 'Subsidized'
        category = donation.get_donation_category_display() or 'Food'
        product = (donation.subsidized_product_type or 'Subsidized Goods')[:20]
        
        # Quantity
        qty = f"{donation.subsidized_quantity} {donation.subsidized_quantity_unit or 'units'}" if donation.subsidized_quantity else '-'
        
        # Prices
        market = f"KES {donation.subsidized_initial_amount:,.0f}" if donation.subsidized_initial_amount else ('-')
        if not donation.subsidized_initial_amount and donation.subsidized_market_price:
            market = f"KES {donation.subsidized_market_price:,.0f}"
        subsidy = f"KES {donation.subsidized_subsidy_amount:,.0f}" if donation.subsidized_subsidy_amount else '-'
        new_price = f"<font color='#059669'><b>KES {donation.subsidized_price:,.0f}</b></font>" if donation.subsidized_price else '-'
        
        # Status
        if donation.status == 'pending':
            status_display = '<font color="#d97706">Awaiting foodbank</font>'
        elif donation.status == 'accepted':
            status_display = '<font color="green">Accepted by FB</font>'
        elif donation.status == 'fulfilled':
            status_display = '<font color="green">Fulfilled</font>'
        elif donation.status == 'declined':
            status_display = '<font color="red">Declined by FB</font>'
        else:
            status_display = donation.get_status_display() or 'Pending'
        
        # Delivery
        delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
        
        # Recipient
        if donation.accepted_by_recipient:
            rec_name = (donation.accepted_by_recipient.full_name or donation.accepted_by_recipient.user.email)[:18]
            recipient_display = f"Claimed<br/><font size='6'>{rec_name}</font>"
        else:
            recipient_display = 'Unclaimed'
        
        # Notes
        donor_note = (donation.message or donation.csr_description or donation.other_description or 'No note')[:50]
        rec_note = (latest_notes.get(donation.id, '') or 'No note')[:50]
        notes_display = f"<font size='6' color='#6b7280'>DONOR:</font> {donor_note}<br/><font size='6' color='#6b7280'>RECIP:</font> {rec_note}"
        if donation.status == 'declined':
            decline_note = (donation.decline_message or 'No decline note')[:60]
        else:
            decline_note = 'Not declined'
        
        row = [
            str(s_no),
            Paragraph(date_str, wrap),
            Paragraph(donor_name[:20], wrap),
            Paragraph(fb_name[:20], wrap),
            Paragraph(location, wrap),
            Paragraph(type_display, wrap),
            Paragraph(category, wrap),
            Paragraph(product, wrap),
            Paragraph(qty, wrap),
            Paragraph(market, wrap),
            Paragraph(subsidy, wrap),
            Paragraph(new_price, wrap),
            Paragraph(status_display, wrap),
            Paragraph(delivery, wrap),
            Paragraph(recipient_display, wrap),
            Paragraph(notes_display, wrap),
            Paragraph(decline_note, wrap),
        ]
        
        table_data.append(row)
        s_no += 1
    
    # Create table and stretch columns to full A3 usable width
    base_col_widths = [
        0.30*inch, 0.60*inch, 1.00*inch, 1.00*inch, 1.00*inch, 0.75*inch, 0.75*inch, 1.10*inch,
        0.75*inch, 0.80*inch, 0.80*inch, 0.80*inch, 1.00*inch, 0.75*inch, 1.00*inch, 1.45*inch, 1.35*inch
    ]
    usable_width = landscape(A3)[0] - doc.leftMargin - doc.rightMargin
    total_base_width = sum(base_col_widths)
    scale_factor = (usable_width / total_base_width) if total_base_width else 1
    col_widths = [w * scale_factor for w in base_col_widths]

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    elements.append(table)
    
    # Add summary
    elements.append(Spacer(1, 0.3*inch))
    summary_text = f"Total Records: {len(donations)}"
    elements.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="subsidized_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)
    
    return response


@staff_member_required
def direct_donations_management(request):
    """
    Direct Requests: FoodBankRequest where original_request is None.
    Cloned layout from dashboard_foodbank_direct_requests_table.
    """
    qs = FoodBankRequest.objects.filter(
        original_request__isnull=True
    ).select_related(
        'foodbank',
        'linked_request_management'
    ).prefetch_related(
        'donations__donor__donor_profile',
        'donations__allocations'
    ).order_by('-created_at')

    # Statistics
    total = qs.count()
    active = qs.filter(status='active').count()
    fulfilled = qs.filter(status='fulfilled').count()
    expired = qs.filter(status='expired').count()

    # Filtering (normalize to '' so current_filters always has strings for form value persistence)
    status_filter = request.GET.get('status') or ''
    foodbank_filter = request.GET.get('foodbank') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    search = (request.GET.get('search') or '').strip()
    donation_type_filter = request.GET.get('donation_type') or ''
    priority_filter = request.GET.get('priority') or ''
    delivery_filter = request.GET.get('delivery') or ''
    response_status_filter = request.GET.get('response_status') or ''
    deadline_filter = request.GET.get('deadline_filter') or ''
    category_filter = request.GET.get('category') or ''

    if status_filter:
        qs = qs.filter(status=status_filter)
    if foodbank_filter:
        qs = qs.filter(foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            qs = qs.filter(created_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            qs = qs.filter(created_at__date__lte=parsed_to)
    if search:
        qs = qs.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search)
        )
    if donation_type_filter:
        qs = qs.filter(donation_type=donation_type_filter)
    if priority_filter:
        qs = qs.filter(priority=priority_filter)
    if delivery_filter:
        qs = qs.filter(donations__delivery_method=delivery_filter).distinct()
    if category_filter:
        qs = qs.filter(donations__donation_type=category_filter).distinct()
    if response_status_filter:
        if response_status_filter == 'sent_to_donors':
            qs = qs.filter(donations__isnull=True)
        elif response_status_filter == 'donation_made':
            qs = qs.filter(donations__status='pending').exclude(
                donations__status='accepted'
            ).distinct()
        elif response_status_filter == 'declined':
            qs = qs.filter(donations__status='declined').exclude(
                donations__status='accepted'
            ).exclude(donations__status='pending').distinct()
        elif response_status_filter in ('partially_fulfilled', 'fulfilled'):
            qs = qs.filter(donations__status='accepted').distinct()
            requests_list = list(qs.prefetch_related('donations'))
            target_label = 'Partially Fulfilled' if response_status_filter == 'partially_fulfilled' else 'Fulfilled'
            matching_ids = [r.id for r in requests_list if r.get_foodbank_requests_status_label() == target_label]
            qs = qs.filter(id__in=matching_ids)
    if deadline_filter:
        now = timezone.now()
        if deadline_filter == 'has_deadline':
            qs = qs.filter(deadline__isnull=False)
        elif deadline_filter == 'past':
            qs = qs.filter(deadline__lt=now)
        elif deadline_filter == 'upcoming':
            qs = qs.filter(deadline__gte=now)
        elif deadline_filter == 'no_deadline':
            qs = qs.filter(deadline__isnull=True)

    # Pagination
    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Attach display attributes (matching dashboard logic)
    delivery_label_map = dict(Donation.DELIVERY_METHODS)
    for req in page_obj.object_list:
        req.status_label = req.get_foodbank_requests_status_label()
        req.quantity_display = req.get_requested_quantity_display()
        req.progress_percent = req.get_fulfillment_percentage()

        donations = list(req.donations.all())

        req.total_received = req.get_total_donations_received()
        req.stock_is_monetary = False

        if not (req.total_received and getattr(req, 'quantity_needed', None)):
            money_donations = [d for d in donations if d.donation_type == 'money' and d.status == 'accepted']
            if money_donations:
                total_money = sum(Decimal(str(getattr(d, 'amount', 0) or 0)) for d in money_donations)
                if total_money > 0:
                    req.total_received = float(total_money)
                    req.stock_is_monetary = True

        used_from_allocations = 0
        used_pieces = 0
        for donation in donations:
            allocations = getattr(donation, 'allocations', None)
            allocation_iter = allocations.all() if hasattr(allocations, 'all') else []
            for allocation in allocation_iter:
                if getattr(allocation, 'declined_by_recipient', False):
                    continue
                used_from_allocations += (allocation.quantity or allocation.amount or 0)
                used_pieces += (allocation.quantity or 0)

        req_quantity_needed = getattr(req, 'quantity_needed', None)
        if req_quantity_needed is not None:
            stock_capacity = req_quantity_needed
            if req.total_received:
                stock_capacity = min(req_quantity_needed, req.total_received)
            req.stock_used = min(used_pieces, stock_capacity)
            req.stock_required = stock_capacity
            req.stock_remaining = max(0, stock_capacity - req.stock_used)
            req.stock_is_monetary = False
        else:
            stock_capacity = req.total_received or 0
            req.stock_required = stock_capacity
            req.stock_used = min(used_from_allocations, stock_capacity) if stock_capacity else 0
            req.stock_remaining = max(0, stock_capacity - req.stock_used) if stock_capacity else 0

        linked = getattr(req, 'linked_request_management', None)
        donor_delivery_methods = []
        type_labels = []
        for donation in donations:
            method = getattr(donation, 'delivery_method', None)
            if method and method not in donor_delivery_methods:
                donor_delivery_methods.append(method)
            if donation.donation_type == 'item':
                type_label = 'Free Goods'
            elif donation.donation_type == 'money' and getattr(req, 'original_request', None):
                type_label = req.original_request.get_request_type_display()
            else:
                type_label = donation.get_donation_type_display()
            if type_label and type_label not in type_labels:
                type_labels.append(type_label)
        if donor_delivery_methods:
            req.delivery_display = ", ".join(
                delivery_label_map.get(m, m.replace('_', ' ').title())
                for m in donor_delivery_methods
            )
        else:
            req.delivery_display = getattr(linked, 'delivery_method', None) or '-'
        req.category_display = ", ".join(type_labels) if type_labels else req.get_donation_type_display()
        req.location_display = (
            getattr(linked, 'location', None)
            or getattr(req.foodbank, 'address', None)
            or '-'
        )
        req.notes_display = getattr(linked, 'additional_notes', None) or req.description

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    # Response status choices (display label, GET value)
    response_status_choices = [
        ('sent_to_donors', 'Published to donors'),
        ('donation_made', 'Donation received by foodbank'),
        ('partially_fulfilled', 'Partially fulfilled'),
        ('fulfilled', 'Fulfilled'),
        ('declined', 'Declined by donor'),
    ]
    deadline_filter_choices = [
        ('', 'Any'),
        ('has_deadline', 'Has deadline'),
        ('past', 'Past deadline'),
        ('upcoming', 'Upcoming'),
        ('no_deadline', 'No deadline'),
    ]

    # Query string for pagination (all GET params except page)
    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    pagination_query_string = get_copy.urlencode()

    context = {
        'title': 'Direct Donations (Direct Requests)',
        'page_obj': page_obj,
        'total': total,
        'active': active,
        'fulfilled': fulfilled,
        'expired': expired,
        'foodbanks': foodbanks,
        'status_choices': FoodBankRequest.STATUS_CHOICES,
        'donation_type_choices': FoodBankRequest.DONATION_TYPE_CHOICES,
        'priority_choices': FoodBankRequest.PRIORITY_CHOICES,
        'delivery_choices': Donation.DELIVERY_METHODS,
        'response_status_choices': response_status_choices,
        'deadline_filter_choices': deadline_filter_choices,
        'category_choices': Donation.DONATION_TYPES,
        'current_filters': {
            'status': status_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'donation_type': donation_type_filter,
            'priority': priority_filter,
            'delivery': delivery_filter,
            'category': category_filter,
            'response_status': response_status_filter,
            'deadline_filter': deadline_filter,
        },
        'pagination_query_string': pagination_query_string,
    }
    return render(request, 'custom_admin/direct_donations_management.html', context)


@staff_member_required
def specified_donations_management(request):
    """
    Specified Donations: RequestManagement (recipient requests).
    Cloned layout from dashboard_foodbank Recipient Requests section.
    """
    request_qs = RequestManagement.objects.select_related(
        'recipient',
        'foodbank',
        'assigned_foodbank',
        'updated_by',
    ).prefetch_related(
        'foodbank_request_created__donations',
        'foodbank_request_created__donations__donor__donor_profile',
        'donor_requests__donations',
    ).order_by('-time_of_request')

    # Statistics
    total_requests = request_qs.count()
    request_pending = request_qs.filter(status='pending').count()
    request_partial = request_qs.filter(status='partial').count()
    request_fulfilled = request_qs.filter(status='fulfilled').count()
    request_declined = request_qs.filter(status='declined').count()

    donations = Donation.objects.filter(
        foodbank_request__isnull=False
    ).select_related(
        'donor',
        'foodbank',
        'foodbank_request',
        'foodbank_request__original_request'
    ).prefetch_related(
        'allocations__recipient'
    ).order_by('-donated_at')

    total_donations = donations.count()
    pending = donations.filter(status='pending').count()
    accepted = donations.filter(status='accepted').count()
    declined = donations.filter(status='declined').count()
    total_monetary = donations.filter(
        donation_type='money', status='accepted'
    ).aggregate(total=Sum('amount'))['total'] or 0
    total_items = donations.filter(
        donation_type='item', status='accepted'
    ).count()

    # Filtering (normalize to '' so current_filters always has strings for form value persistence)
    status_filter = request.GET.get('status') or ''
    foodbank_filter = request.GET.get('foodbank') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    search = (request.GET.get('search') or '').strip()
    request_type_filter = request.GET.get('request_type') or ''
    request_category_filter = request.GET.get('request_category') or ''
    delivery_filter = request.GET.get('delivery') or ''
    donation_type_filter = request.GET.get('donation_type') or ''

    if status_filter:
        request_qs = request_qs.filter(status=status_filter)
    if foodbank_filter:
        request_qs = request_qs.filter(
            Q(foodbank_id=foodbank_filter) |
            Q(assigned_foodbank_id=foodbank_filter)
        )
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            request_qs = request_qs.filter(time_of_request__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            request_qs = request_qs.filter(time_of_request__date__lte=parsed_to)
    if search:
        request_qs = request_qs.filter(
            Q(description__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(recipient__user__email__icontains=search) |
            Q(location__icontains=search) |
            Q(additional_notes__icontains=search)
        )
    if request_type_filter:
        request_qs = request_qs.filter(request_type=request_type_filter)
    if request_category_filter:
        request_qs = request_qs.filter(request_category=request_category_filter)
    if delivery_filter:
        request_qs = request_qs.filter(delivery_method=delivery_filter)
    if donation_type_filter:
        request_qs = request_qs.filter(
            foodbank_request_created__donations__donation_type=donation_type_filter
        ).distinct()

    request_paginator = Paginator(request_qs, 15)
    request_page = request.GET.get('requests_page')
    request_page_obj = request_paginator.get_page(request_page)

    def _admin_status_meta(req_obj):
        base_text = req_obj.get_display_status()
        meta = {
            'text': base_text,
            'bg': '#e5e7eb',
            'color': '#374151',
            'icon': 'fas fa-info-circle'
        }
        if req_obj.status == 'pending':
            meta.update(text='Sent to Foodbank', bg='#e5e7eb', color='#374151', icon='fas fa-paper-plane')
        elif req_obj.status == 'submitted':
            meta.update(text='Submitted to Donors', bg='#fff7ed', color='#c2410c', icon='fas fa-bullhorn')
        elif req_obj.status == 'donation_received':
            meta.update(text='Awaiting Foodbank Review', bg='#dbeafe', color='#1d4ed8', icon='fas fa-clipboard-check')
        elif req_obj.status == 'assigned':
            meta.update(text='Assigned to Foodbank', bg='#ede9fe', color='#6d28d9', icon='fas fa-handshake')
        elif req_obj.status == 'awaiting_recipient':
            meta.update(text='Awaiting Recipient Response', bg='#fef3c7', color='#92400e', icon='fas fa-clock')
        elif req_obj.status == 'partial':
            meta.update(text=base_text, bg='#dbeafe', color='#1d4ed8', icon='fas fa-adjust')
        elif req_obj.status == 'fulfilled':
            meta.update(text=base_text, bg='#dcfce7', color='#166534', icon='fas fa-check-circle')
        elif req_obj.status == 'declined':
            meta.update(text='Declined', bg='#fee2e2', color='#b91c1c', icon='fas fa-times-circle')
        elif req_obj.status == 'acknowledged':
            meta.update(text='Recipient Acknowledged', bg='#dcfce7', color='#15803d', icon='fas fa-check-double')
        return meta

    from authentication.views import _get_request_fulfillment_breakdown
    for req in request_page_obj.object_list:
        fb_request, linked_donation = _resolve_specified_links(req)
        req.linked_foodbank_request = fb_request
        req.linked_donation = linked_donation
        req.fulfillment_breakdown = _get_request_fulfillment_breakdown(req)
        status_meta = _admin_status_meta(req)
        req.admin_status_text = status_meta['text']
        req.admin_status_bg = status_meta['bg']
        req.admin_status_color = status_meta['color']
        req.admin_status_icon = status_meta['icon']

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    get_copy = request.GET.copy()
    get_copy.pop('requests_page', None)
    pagination_query_string = get_copy.urlencode()

    context = {
        'title': 'Specified Donations Management',
        'request_page_obj': request_page_obj,
        'total_requests': total_requests,
        'request_pending': request_pending,
        'request_partial': request_partial,
        'request_fulfilled': request_fulfilled,
        'request_declined': request_declined,
        'total_donations': total_donations,
        'pending': pending,
        'accepted': accepted,
        'declined': declined,
        'total_monetary': total_monetary,
        'total_items': total_items,
        'foodbanks': foodbanks,
        'status_choices': RequestManagement.STATUS_CHOICES,
        'request_type_choices': RequestManagement.REQUEST_TYPE_CHOICES,
        'request_category_choices': RequestManagement.REQUEST_CATEGORY_CHOICES,
        'delivery_choices': RequestManagement.DELIVERY_METHOD_CHOICES,
        'donation_type_choices': Donation.DONATION_TYPES,
        'pagination_query_string': pagination_query_string,
        'current_filters': {
            'status': status_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'request_type': request_type_filter,
            'request_category': request_category_filter,
            'delivery': delivery_filter,
            'donation_type': donation_type_filter,
        }
    }
    return render(request, 'custom_admin/specified_donations_management.html', context)


def _get_specified_filtered_queryset(request):
    """Return filtered RequestManagement queryset for specified donations exports."""
    request_qs = RequestManagement.objects.select_related(
        'recipient', 'foodbank', 'assigned_foodbank', 'updated_by',
        'donation', 'donation__donor__donor_profile',
        'foodbank_request',
    ).prefetch_related(
        'foodbank_request_created__donations',
        'foodbank_request_created__donations__donor__donor_profile',
        'donor_requests__donations',
        'donor_requests__donations__donor__donor_profile',
    ).order_by('-time_of_request')

    status_filter = request.GET.get('status')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    request_type_filter = request.GET.get('request_type')
    request_category_filter = request.GET.get('request_category')
    delivery_filter = request.GET.get('delivery')
    donation_type_filter = request.GET.get('donation_type')

    if status_filter:
        request_qs = request_qs.filter(status=status_filter)
    if foodbank_filter:
        request_qs = request_qs.filter(
            Q(foodbank_id=foodbank_filter) | Q(assigned_foodbank_id=foodbank_filter)
        )
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            request_qs = request_qs.filter(time_of_request__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            request_qs = request_qs.filter(time_of_request__date__lte=parsed_to)
    if search:
        request_qs = request_qs.filter(
            Q(description__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(recipient__user__email__icontains=search) |
            Q(location__icontains=search) |
            Q(additional_notes__icontains=search)
        )
    if request_type_filter:
        request_qs = request_qs.filter(request_type=request_type_filter)
    if request_category_filter:
        request_qs = request_qs.filter(request_category=request_category_filter)
    if delivery_filter:
        request_qs = request_qs.filter(delivery_method=delivery_filter)
    if donation_type_filter:
        request_qs = request_qs.filter(
            foodbank_request_created__donations__donation_type=donation_type_filter
        ).distinct()

    return request_qs


def _resolve_specified_links(req):
    """
    Resolve a RequestManagement row to its best-linked FoodBankRequest and Donation.
    This checks all supported relation paths to avoid false 'Request' type labels.
    """
    fb_requests = {}

    def _add_fb_request(fb_req):
        if fb_req and getattr(fb_req, 'id', None):
            fb_requests[fb_req.id] = fb_req

    _add_fb_request(getattr(req, 'foodbank_request', None))

    created_rel = getattr(req, 'foodbank_request_created', None)
    if created_rel is not None:
        if hasattr(created_rel, 'all'):
            for fb_req in created_rel.all():
                _add_fb_request(fb_req)
        else:
            _add_fb_request(created_rel)

    donor_rel = getattr(req, 'donor_requests', None)
    if donor_rel is not None:
        if hasattr(donor_rel, 'all'):
            for fb_req in donor_rel.all():
                _add_fb_request(fb_req)
        else:
            _add_fb_request(donor_rel)

    donation_candidates = {}

    def _add_donation(donation):
        if donation and getattr(donation, 'id', None):
            donation_candidates[donation.id] = donation

    _add_donation(getattr(req, 'donation', None))

    for fb_req in fb_requests.values():
        rel = getattr(fb_req, 'donations', None)
        if rel is None:
            continue
        if hasattr(rel, 'all'):
            for donation in rel.all():
                _add_donation(donation)
        else:
            _add_donation(rel)

    donation = max(donation_candidates.values(), key=lambda d: d.id) if donation_candidates else None

    if donation is None:
        donation_query = Q(foodbank_request__original_request=req) | Q(foodbank_request__linked_request_management=req)
        if getattr(req, 'donation_id', None):
            donation_query |= Q(id=req.donation_id)
        donation = Donation.objects.filter(donation_query).select_related(
            'donor__donor_profile', 'foodbank_request'
        ).order_by('-id').first()

    if donation and getattr(donation, 'foodbank_request', None):
        fb_request = donation.foodbank_request
    elif fb_requests:
        fb_request = max(fb_requests.values(), key=lambda fr: fr.id)
    else:
        fb_request = None

    return fb_request, donation


def _get_specified_donor_display(donation):
    """Get donor display for specified exports with a no-donation placeholder."""
    if not donation:
        return 'Not donated'

    profile = getattr(donation.donor, 'donor_profile', None)
    if profile and getattr(profile, 'organization_name', None):
        return profile.organization_name
    if profile and getattr(profile, 'full_name', None):
        return profile.full_name
    return donation.donor.email or 'Not donated'


@staff_member_required
def export_specified_donations_csv(request):
    """Export specified donations (RequestManagement) to CSV."""
    from datetime import datetime
    from custom_admin.utils import get_neutral_status

    request_qs = _get_specified_filtered_queryset(request)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="specified_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'S/No', 'Code', 'Type', 'Category', 'Description', 'Donor', 'Recipient', 'Foodbank',
        'Quantity', 'Delivery', 'Location', 'Status', 'Requested', 'Notes',
        'Request Rejection Note', 'Donation Rejection Note'
    ])

    for s_no, req in enumerate(request_qs, 1):
        fb_request, donation = _resolve_specified_links(req)

        code = donation.donation_code if donation and donation.donation_code else 'DN-'
        type_display = 'Anonymous' if req.is_anonymous else (donation.get_donation_type_display() if donation else 'Request')
        cat = req.request_category or req.request_type or ''
        category_display = 'Food' if cat == 'food' else ('Non-Food' if cat == 'non_food' else 'Other')
        desc = (req.description or '').replace('\n', ' ').replace('\r', '')
        donor_display = _get_specified_donor_display(donation)
        rec_name = req.recipient.full_name if req.recipient else 'N/A'
        rec_email = req.recipient.user.email if req.recipient else ''
        recipient_display = f"{rec_name} ({rec_email})" if rec_email else rec_name
        fb_name = getattr(req, 'foodbank_name', None) or 'N/A'
        qty_requested = f"{req.quantity} {req.get_unit_display()}"
        qty_fulfilled = ''
        if req.quantity_fulfilled and req.quantity_fulfilled > 0:
            qty_fulfilled = f" / Fulfilled: {req.quantity_fulfilled} {req.get_unit_display()}"
            if req.status == 'partial':
                remaining = req.get_remaining_quantity() if hasattr(req, 'get_remaining_quantity') else max(0, (req.quantity or 0) - (req.quantity_fulfilled or 0))
                qty_fulfilled += f" / Remaining: {remaining} {req.get_unit_display()}"
        quantity_display = f"Requested: {qty_requested}{qty_fulfilled}"
        delivery = req.get_delivery_method_display() if hasattr(req, 'get_delivery_method_display') else 'N/A'
        location = req.location or '-'
        status_display = get_neutral_status('specified', str(req.status).lower() if req.status else '', {'request_obj': req})
        date_str = req.time_of_request.strftime('%Y-%m-%d %H:%M') if req.time_of_request else '-'
        if req.fulfilled_at:
            date_str += f" (Fulfilled: {req.fulfilled_at.strftime('%b %d')})"
        notes = (req.additional_notes or '-').replace('\n', ' ').replace('\r', '')[:100]
        request_rejection_note = 'Not declined'
        if str(req.status).lower() == 'declined':
            request_rejection_note = (req.decline_message or 'No rejection note')
        request_rejection_note = request_rejection_note.replace('\n', ' ').replace('\r', '')

        if donation:
            if str(donation.status).lower() == 'declined':
                donation_rejection_note = (donation.decline_message or 'No rejection note')
            else:
                donation_rejection_note = 'Not declined'
        else:
            donation_rejection_note = 'Not donated'
        donation_rejection_note = donation_rejection_note.replace('\n', ' ').replace('\r', '')

        writer.writerow([
            s_no, code, type_display, category_display, desc, donor_display, recipient_display, fb_name,
            quantity_display, delivery, location, status_display, date_str, notes,
            request_rejection_note, donation_rejection_note
        ])

    return response


@staff_member_required
def export_specified_donations_excel(request):
    """Export specified donations (RequestManagement) to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from custom_admin.utils import get_neutral_status

    request_qs = _get_specified_filtered_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Specified Donations"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:S1')
    ws['A1'].value = "FOODBANKHUB"
    ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A2:S2')
    ws['A2'].value = "Specified Donations Report"
    ws['A2'].font = Font(size=14, bold=True, color="1F4E78")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22
    ws.merge_cells('A3:S3')
    ws['A3'].value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  |  Total records: {request_qs.count()}"
    ws['A3'].font = Font(size=9, color="6B7280")
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 18

    headers = ['S/No', 'Code', 'Type', 'Category', 'Description', 'Donor', 'Recipient', 'Foodbank',
               'Requested Qty', 'Fulfilled Qty', 'Remaining Qty', 'Unit', 'Delivery', 'Location',
               'Status', 'Requested At', 'Notes', 'Request Rejection Note', 'Donation Rejection Note']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[5].height = 24

    row_num = 6
    for s_no, req in enumerate(request_qs, 1):
        fb_request, donation = _resolve_specified_links(req)
        code = donation.donation_code if donation and donation.donation_code else 'DN-'
        type_display = 'Anonymous' if req.is_anonymous else (donation.get_donation_type_display() if donation else 'Request')
        cat = req.request_category or req.request_type or ''
        category_display = 'Food' if cat == 'food' else ('Non-Food' if cat == 'non_food' else 'Other')
        desc = (req.description or '')
        donor_display = _get_specified_donor_display(donation)
        rec_name = req.recipient.full_name if req.recipient else 'N/A'
        rec_email = req.recipient.user.email if req.recipient else ''
        recipient_display = f"{rec_name} ({rec_email})" if rec_email else rec_name
        fb_name = getattr(req, 'foodbank_name', None) or 'N/A'
        requested_qty = req.quantity if req.quantity is not None else ''
        fulfilled_qty = req.quantity_fulfilled if req.quantity_fulfilled is not None else ''
        remaining_qty = (
            req.get_remaining_quantity()
            if hasattr(req, 'get_remaining_quantity')
            else max(0, (req.quantity or 0) - (req.quantity_fulfilled or 0))
        )
        unit_display = req.get_unit_display() if hasattr(req, 'get_unit_display') else '-'
        delivery = req.get_delivery_method_display() if hasattr(req, 'get_delivery_method_display') else 'N/A'
        location = req.location or '-'
        status_display = get_neutral_status('specified', str(req.status).lower() if req.status else '', {'request_obj': req})
        date_str = req.time_of_request.strftime('%Y-%m-%d %H:%M') if req.time_of_request else '-'
        if req.fulfilled_at:
            date_str += f" (Fulfilled: {req.fulfilled_at.strftime('%b %d')})"
        notes = (req.additional_notes or '-')[:100]
        request_rejection_note = 'Not declined'
        if str(req.status).lower() == 'declined':
            request_rejection_note = (req.decline_message or 'No rejection note')

        if donation:
            if str(donation.status).lower() == 'declined':
                donation_rejection_note = (donation.decline_message or 'No rejection note')
            else:
                donation_rejection_note = 'Not declined'
        else:
            donation_rejection_note = 'Not donated'

        row_data = [s_no, code, type_display, category_display, desc[:80], donor_display, recipient_display,
                    fb_name, requested_qty, fulfilled_qty, remaining_qty, unit_display, delivery, location, status_display,
                    date_str, notes, request_rejection_note[:120], donation_rejection_note[:120]]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col_num in (5, 6, 7, 8, 17, 18, 19) else 'center')
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        ws.row_dimensions[row_num].height = max(18, min(45, 24))
        row_num += 1

    column_widths = [6, 12, 14, 12, 24, 20, 24, 20, 12, 12, 12, 12, 12, 18, 18, 18, 28, 24, 24]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = min(width, 50)
    if request_qs.exists():
        ws.auto_filter.ref = f"A5:S{row_num - 1}"
    ws.freeze_panes = 'A6'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="specified_donations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def export_specified_donations_pdf(request):
    """Export specified donations (RequestManagement) to PDF."""
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.platypus import Paragraph
    from authentication.report_utils import (
        get_report_styles,
        build_report_header,
        collect_active_filters,
    )
    from custom_admin.utils import get_neutral_status

    request_qs = _get_specified_filtered_queryset(request)

    requests_list = list(request_qs)
    total_count = len(requests_list)
    styles = get_report_styles()
    active_filters = collect_active_filters(request, [
        ('status', 'Status'),
        ('foodbank', 'Foodbank'),
        ('date_from', 'From'),
        ('date_to', 'To'),
        ('search', 'Search'),
        ('request_type', 'Request type'),
        ('request_category', 'Category'),
        ('delivery', 'Delivery'),
        ('donation_type', 'Donation type'),
    ])

    elements = []
    build_report_header(
        elements,
        report_title="Specified Donations Report",
        generated_for="Admin",
        total_records=total_count,
        active_filters=active_filters or None,
        styles_dict=styles,
    )

    header_row = ['S/No', 'Code', 'Type', 'Category', 'Description', 'Donor', 'Recipient', 'Foodbank',
                 'Quantity', 'Delivery', 'Location', 'Status', 'Requested', 'Notes',
                 'Request Rejection Note', 'Donation Rejection Note']
    table_data = [header_row]
    wrap = styles['wrap']
    wrap_center = styles['wrap_center']
    pdf_pagesize = landscape(A3)

    for s_no, req in enumerate(requests_list, 1):
        fb_request, donation = _resolve_specified_links(req)
        code = (donation.donation_code or 'DN-')[:10] if donation else 'DN-'
        type_display = 'Anonymous' if req.is_anonymous else (donation.get_donation_type_display()[:12] if donation else 'Request')
        cat = req.request_category or req.request_type or ''
        category_display = 'Food' if cat == 'food' else ('Non-Food' if cat == 'non_food' else 'Other')
        desc = (req.description or '')[:50]
        donor_display = (_get_specified_donor_display(donation) or 'Not donated')[:30]
        rec_name = (req.recipient.full_name if req.recipient else 'N/A')[:20]
        fb_name = (getattr(req, 'foodbank_name', None) or 'N/A')[:20]
        qty_str = f"<b>{req.quantity} {req.get_unit_display()}</b><br/><font size='6' color='#6b7280'>Requested</font>"
        if req.quantity_fulfilled and req.quantity_fulfilled > 0:
            qty_str += f"<br/><font color='green'>{req.quantity_fulfilled} {req.get_unit_display()}</font><br/><font size='6' color='#6b7280'>Fulfilled</font>"
            if req.status == 'partial':
                remaining = req.get_remaining_quantity() if hasattr(req, 'get_remaining_quantity') else max(0, (req.quantity or 0) - (req.quantity_fulfilled or 0))
                qty_str += f"<br/><font color='#d97706'>{remaining} {req.get_unit_display()}</font><br/><font size='6' color='#6b7280'>Remaining</font>"
        delivery = (req.get_delivery_method_display() if hasattr(req, 'get_delivery_method_display') else 'N/A')[:12]
        location = (req.location or '-')[:22]
        status_display = get_neutral_status('specified', str(req.status).lower() if req.status else '', {'request_obj': req})
        date_str = req.time_of_request.strftime('%b %d, %Y') if req.time_of_request else '-'
        if req.fulfilled_at:
            date_str += f"<br/><font size='6' color='green'>Done: {req.fulfilled_at.strftime('%b %d')}</font>"
        notes = (req.additional_notes or '-')[:50]
        request_rejection_note = 'Not declined'
        if str(req.status).lower() == 'declined':
            request_rejection_note = (req.decline_message or 'No rejection note')
        request_rejection_note = request_rejection_note.replace('\n', '<br/>')[:55]

        if donation:
            if str(donation.status).lower() == 'declined':
                donation_rejection_note = (donation.decline_message or 'No rejection note')
            else:
                donation_rejection_note = 'Not declined'
        else:
            donation_rejection_note = 'Not donated'
        donation_rejection_note = donation_rejection_note.replace('\n', '<br/>')[:55]
        table_data.append([
            Paragraph(str(s_no), wrap_center),
            Paragraph(code, wrap_center),
            Paragraph(type_display, wrap_center),
            Paragraph(category_display, wrap_center),
            Paragraph(desc, wrap),
            Paragraph(donor_display, wrap),
            Paragraph(rec_name, wrap),
            Paragraph(fb_name, wrap),
            Paragraph(qty_str, wrap),
            Paragraph(delivery, wrap_center),
            Paragraph(location, wrap),
            Paragraph(status_display, wrap_center),
            Paragraph(date_str, wrap),
            Paragraph(notes, wrap),
            Paragraph(request_rejection_note, wrap),
            Paragraph(donation_rejection_note, wrap),
        ])

    from authentication.report_utils import get_branded_table_style, build_pdf_document, make_full_width_table, build_report_summary
    col_weights = [0.35, 0.55, 0.70, 0.55, 1.20, 0.90, 0.90, 0.85, 1.10, 0.55, 0.80, 0.75, 0.85, 1.00, 1.05, 1.05]
    table = make_full_width_table(table_data, repeat_rows=1, col_weights=col_weights, pagesize=pdf_pagesize)
    table.setStyle(get_branded_table_style(len(table_data)))
    elements.append(table)

    summary_items = [("Total records", total_count)]
    pending_count = sum(1 for r in requests_list if r.status == 'pending')
    fulfilled_count = sum(1 for r in requests_list if r.status == 'fulfilled')
    partial_count = sum(1 for r in requests_list if r.status == 'partial')
    summary_items.extend([("Pending", pending_count), ("Fulfilled", fulfilled_count), ("Partial", partial_count)])
    build_report_summary(elements, summary_items, styles_dict=styles)

    return build_pdf_document(
        elements,
        filename_prefix="specified_donations",
        user_display_name="Admin",
        pagesize=pdf_pagesize,
    )


@staff_member_required
def direct_donations_manage(request):
    """
    Direct donations in the pipeline: not yet accepted by foodbank (pending, declined).
    Table shows status as donations move through the flow.
    """
    donations = Donation.objects.filter(
        foodbank_request__isnull=False
    ).exclude(
        status='accepted'
    ).select_related(
        'donor', 'donor__donor_profile', 'foodbank',
        'foodbank_request', 'foodbank_request__original_request',
        'foodbank_request__original_request__recipient'
    ).order_by('-donated_at')

    # Filters
    status_filter = request.GET.get('status', '')
    foodbank_filter = request.GET.get('foodbank', '')
    search = request.GET.get('search', '').strip()
    if status_filter:
        donations = donations.filter(status=status_filter)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if search:
        donations = donations.filter(
            Q(item_name__icontains=search) |
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(message__icontains=search) |
            Q(foodbank_request__title__icontains=search) |
            Q(foodbank_request__original_request__description__icontains=search)
        )

    pending_count = Donation.objects.filter(foodbank_request__isnull=False, status='pending').count()
    declined_count = Donation.objects.filter(foodbank_request__isnull=False, status='declined').count()
    total = donations.count()

    paginator = Paginator(donations, 15)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    pagination_query_string = get_copy.urlencode()

    for d in page_obj.object_list:
        d.admin_status_text = 'Pending' if d.status == 'pending' else 'Declined'
        d.admin_status_bg = '#fef3c7' if d.status == 'pending' else '#fee2e2'
        d.admin_status_color = '#92400e' if d.status == 'pending' else '#b91c1c'

    context = {
        'title': 'Direct Donations – Manage',
        'section': 'manage',
        'donations': page_obj,
        'page_obj': page_obj,
        'total': total,
        'pending_count': pending_count,
        'declined_count': declined_count,
        'foodbanks': foodbanks,
        'pagination_query_string': pagination_query_string,
        'current_filters': {
            'status': status_filter,
            'foodbank': foodbank_filter,
            'search': search,
        }
    }
    return render(request, 'custom_admin/direct_donations_manage_inventory.html', context)


@staff_member_required
def direct_donations_inventory(request):
    """
    Direct donations accepted by foodbank (in stock). status='accepted'.
    """
    donations = Donation.objects.filter(
        foodbank_request__isnull=False,
        status='accepted'
    ).select_related(
        'donor', 'donor__donor_profile', 'foodbank',
        'foodbank_request', 'foodbank_request__original_request',
        'foodbank_request__original_request__recipient'
    ).prefetch_related('allocations__recipient').order_by('-donated_at')

    # Filters (comprehensive, aligned with direct_donations_management)
    foodbank_filter = request.GET.get('foodbank', '')
    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    donation_type_filter = request.GET.get('donation_type')
    priority_filter = request.GET.get('priority')
    delivery_filter = request.GET.get('delivery')
    category_filter = request.GET.get('category')
    allocation_filter = request.GET.get('allocation')
    deadline_filter = request.GET.get('deadline_filter')

    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if search:
        donations = donations.filter(
            Q(item_name__icontains=search) |
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(message__icontains=search) |
            Q(foodbank_request__title__icontains=search) |
            Q(foodbank_request__description__icontains=search) |
            Q(foodbank_request__original_request__description__icontains=search)
        )
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if donation_type_filter:
        donations = donations.filter(donation_type=donation_type_filter)
    if priority_filter:
        donations = donations.filter(foodbank_request__priority=priority_filter)
    if delivery_filter:
        donations = donations.filter(delivery_method=delivery_filter)
    if category_filter:
        donations = donations.filter(donation_category=category_filter)
    if allocation_filter:
        if allocation_filter == 'allocated':
            donations = donations.filter(is_allocated=True)
        elif allocation_filter == 'unallocated':
            donations = donations.filter(is_allocated=False)
    if deadline_filter:
        now = timezone.now()
        if deadline_filter == 'has_deadline':
            donations = donations.filter(foodbank_request__deadline__isnull=False)
        elif deadline_filter == 'past':
            donations = donations.filter(foodbank_request__deadline__lt=now)
        elif deadline_filter == 'upcoming':
            donations = donations.filter(foodbank_request__deadline__gte=now)
        elif deadline_filter == 'no_deadline':
            donations = donations.filter(foodbank_request__deadline__isnull=True)

    total = donations.count()
    allocated_count = donations.filter(is_allocated=True).count()
    unallocated_count = total - allocated_count

    paginator = Paginator(donations, 15)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    deadline_filter_choices = [
        ('', 'Any'),
        ('has_deadline', 'Has deadline'),
        ('past', 'Past deadline'),
        ('upcoming', 'Upcoming'),
        ('no_deadline', 'No deadline'),
    ]
    allocation_choices = [
        ('', 'All'),
        ('allocated', 'Allocated'),
        ('unallocated', 'Unallocated'),
    ]
    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    pagination_query_string = get_copy.urlencode()

    for d in page_obj.object_list:
        d.admin_status_text = 'In stock'
        d.admin_status_bg = '#dcfce7'
        d.admin_status_color = '#166534'
        if d.is_allocated:
            d.admin_status_text = 'Partially / Fully allocated'
            d.admin_status_bg = '#dbeafe'
            d.admin_status_color = '#1d4ed8'

    context = {
        'title': 'Direct Donations – Inventory',
        'section': 'inventory',
        'donations': page_obj,
        'page_obj': page_obj,
        'total': total,
        'allocated_count': allocated_count,
        'unallocated_count': unallocated_count,
        'foodbanks': foodbanks,
        'donation_type_choices': Donation.DONATION_TYPES,
        'priority_choices': FoodBankRequest.PRIORITY_CHOICES,
        'delivery_choices': Donation.DELIVERY_METHODS,
        'category_choices': Donation.DONATION_CATEGORIES,
        'deadline_filter_choices': deadline_filter_choices,
        'allocation_choices': allocation_choices,
        'pagination_query_string': pagination_query_string,
        'current_filters': {
            'foodbank': foodbank_filter,
            'search': search,
            'date_from': date_from,
            'date_to': date_to,
            'donation_type': donation_type_filter,
            'priority': priority_filter,
            'delivery': delivery_filter,
            'category': category_filter,
            'allocation': allocation_filter,
            'deadline_filter': deadline_filter,
        }
    }
    return render(request, 'custom_admin/direct_donations_manage_inventory.html', context)


@staff_member_required
def direct_donation_details(request, donation_id):
    """
    API endpoint to fetch detailed information about a direct donation
    """
    try:
        donation = Donation.objects.select_related(
            'donor__donor_profile',
            'foodbank',
            'foodbank_request__original_request',
            'foodbank_request__original_request__recipient'
        ).get(id=donation_id, foodbank_request__isnull=False)
        
        # Build tracking history
        tracking_history = []
        
        # Created event
        try:
            donor_message = getattr(donation, 'message', None)
            tracking_history.append({
                'status': 'Created',
                'description': 'Direct donation created',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'badge_class': 'bg-secondary',
                'comment': donor_message if donor_message else None
            })
        except:
            pass
        
        # Status changes
        if donation.status == 'accepted':
            tracking_history.append({
                'status': 'Accepted',
                'description': 'Donation accepted',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': 'System',
                'badge_class': 'bg-success',
                'comment': None
            })
        elif donation.status == 'declined':
            tracking_history.append({
                'status': 'Declined',
                'description': 'Donation declined',
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'updated_by': 'System',
                'badge_class': 'bg-danger',
                'comment': donation.decline_message if hasattr(donation, 'decline_message') and donation.decline_message else None
            })
        
        # Build details string
        details = 'N/A'
        if donation.donation_type == 'item':
            details = f"{donation.quantity} {donation.quantity_unit} of {donation.item_name}"
        elif donation.donation_type == 'money':
            details = f"KES {donation.amount}"
        elif donation.donation_type == 'subsidized':
            details = f"KES {donation.subsidized_price}"
        
        # Build response data
        data = {
            'success': True,
            'donation': {
                'id': donation.id,
                'donation_type': donation.get_donation_type_display(),
                'details': details,
                'status': donation.get_status_display(),
                'status_class': 'bg-success' if donation.status == 'accepted' else ('bg-danger' if donation.status == 'declined' else 'bg-warning'),
                'created_at': donation.donated_at.strftime('%b %d, %Y %H:%M') if donation.donated_at else 'N/A',
                'delivery_method': donation.get_delivery_method_display() if donation.delivery_method else None,
                
                # Donor info
                'donor_name': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email,
                'donor_email': donation.donor.email,
                'donor_phone': getattr(donation.donor, 'phone_number', None),
                'donor_organization': getattr(donation.donor.donor_profile, 'organization_name', None) if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile and getattr(donation.donor.donor_profile, 'is_organization', False) else None,
                
                # Foodbank info
                'foodbank_name': donation.foodbank.foodbank_name,
                'foodbank_contact': getattr(donation.foodbank, 'contact_person', None),
                
                # Request info
                'request_title': donation.foodbank_request.title if donation.foodbank_request else None,
                'request_description': donation.foodbank_request.description if donation.foodbank_request else None,
                'recipient_name': donation.foodbank_request.original_request.recipient.full_name if donation.foodbank_request and donation.foodbank_request.original_request and donation.foodbank_request.original_request.recipient else None,
                
                # Tracking history
                'tracking_history': tracking_history
            }
        }
        
        return JsonResponse(data)
        
    except Donation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Donation not found'
        }, status=404)
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


@staff_member_required
def export_direct_donation_details_pdf(request, donation_id):
    """
    Export direct donation details to PDF
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from datetime import datetime
    
    try:
        donation = Donation.objects.select_related(
            'donor__donor_profile',
            'foodbank',
            'foodbank_request__original_request',
            'foodbank_request__original_request__recipient'
        ).get(id=donation_id, foodbank_request__isnull=False)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch,
                                topMargin=1*inch, bottomMargin=0.75*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=6
        )
        
        # Title
        elements.append(Paragraph("Direct Donation Details Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                                 ParagraphStyle('Subtitle', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))
        elements.append(Spacer(1, 0.3*inch))
        
        # Donation Information
        elements.append(Paragraph("Donation Information", heading_style))
        
        details = 'N/A'
        if donation.donation_type == 'item':
            details = f"{donation.quantity} {donation.quantity_unit} of {donation.item_name}"
        elif donation.donation_type == 'money':
            details = f"KES {donation.amount}"
        elif donation.donation_type == 'subsidized':
            details = f"KES {donation.subsidized_price}"
        
        donation_data = [
            ['Donation ID:', f"#{donation.id}"],
            ['Type:', donation.get_donation_type_display()],
            ['Details:', details],
            ['Status:', donation.get_status_display()],
            ['Delivery Method:', donation.get_delivery_method_display() if donation.delivery_method else 'N/A'],
            ['Created:', donation.donated_at.strftime('%B %d, %Y at %I:%M %p') if donation.donated_at else 'N/A'],
        ]
        
        donation_table = Table(donation_data, colWidths=[2*inch, 4.5*inch])
        donation_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(donation_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Donor Information
        elements.append(Paragraph("Donor Information", heading_style))
        
        donor_data = [
            ['Name:', donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile else donation.donor.email],
            ['Email:', donation.donor.email],
            ['Phone:', getattr(donation.donor, 'phone_number', 'N/A')],
        ]
        
        donor_table = Table(donor_data, colWidths=[2*inch, 4.5*inch])
        donor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f5e9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a5d6a7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(donor_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Foodbank Information
        elements.append(Paragraph("Foodbank Information", heading_style))
        
        foodbank_data = [
            ['Foodbank:', donation.foodbank.foodbank_name],
            ['Contact:', getattr(donation.foodbank, 'contact_person', 'N/A')],
        ]
        
        foodbank_table = Table(foodbank_data, colWidths=[2*inch, 4.5*inch])
        foodbank_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fff3e0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ffcc80')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(foodbank_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Request Information
        if donation.foodbank_request:
            elements.append(Paragraph("Request Information", heading_style))
            
            request_data = [
                ['Title:', donation.foodbank_request.title],
                ['Recipient:', donation.foodbank_request.original_request.recipient.full_name if donation.foodbank_request.original_request and donation.foodbank_request.original_request.recipient else 'N/A'],
            ]
            
            request_table = Table(request_data, colWidths=[2*inch, 4.5*inch])
            request_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e3f2fd')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#90caf9')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(request_table)
        
        # Build PDF
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="direct_donation_{donation.id}_details.pdf"'
        response.write(pdf)
        
        return response
        
    except Donation.DoesNotExist:
        return HttpResponse("Donation not found", status=404)
    except Exception as e:
        import traceback
        return HttpResponse(f"Error generating PDF: {str(e)}\n{traceback.format_exc()}", status=500)


@staff_member_required
def export_direct_donations_excel(request):
    """
    Export direct donations table to Excel.
    When no section: export FoodBankRequest (Direct Requests).
    When section=manage/inventory: export RequestManagement (Specified / pipeline).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from custom_admin.utils import get_neutral_status

    section = request.GET.get('section', '')
    status_filter = request.GET.get('status')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    donation_type_filter = request.GET.get('donation_type')
    priority_filter = request.GET.get('priority')
    delivery_filter = request.GET.get('delivery')
    response_status_filter = request.GET.get('response_status')
    deadline_filter = request.GET.get('deadline_filter')
    category_filter = request.GET.get('category')

    # When no section: export Direct Requests (FoodBankRequest)
    if not section:
        qs = FoodBankRequest.objects.filter(
            original_request__isnull=True
        ).select_related('foodbank', 'linked_request_management').prefetch_related(
            'donations__donor__donor_profile'
        ).order_by('-created_at')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if foodbank_filter:
            qs = qs.filter(foodbank_id=foodbank_filter)
        if date_from:
            parsed_from = parse_date(date_from)
            if parsed_from:
                qs = qs.filter(created_at__date__gte=parsed_from)
        if date_to:
            parsed_to = parse_date(date_to)
            if parsed_to:
                qs = qs.filter(created_at__date__lte=parsed_to)
        if search:
            qs = qs.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(foodbank__foodbank_name__icontains=search)
            )
        if donation_type_filter:
            qs = qs.filter(donation_type=donation_type_filter)
        if priority_filter:
            qs = qs.filter(priority=priority_filter)
        if delivery_filter:
            qs = qs.filter(donations__delivery_method=delivery_filter).distinct()
        if category_filter:
            qs = qs.filter(donations__donation_type=category_filter).distinct()
        if response_status_filter:
            if response_status_filter == 'sent_to_donors':
                qs = qs.filter(donations__isnull=True)
            elif response_status_filter == 'donation_made':
                qs = qs.filter(donations__status='pending').exclude(
                    donations__status='accepted'
                ).distinct()
            elif response_status_filter == 'declined':
                qs = qs.filter(donations__status='declined').exclude(
                    donations__status='accepted'
                ).exclude(donations__status='pending').distinct()
            elif response_status_filter in ('partially_fulfilled', 'fulfilled'):
                qs = qs.filter(donations__status='accepted').distinct()
                requests_list = list(qs.prefetch_related('donations'))
                target_label = 'Partially Fulfilled' if response_status_filter == 'partially_fulfilled' else 'Fulfilled'
                matching_ids = [r.id for r in requests_list if r.get_foodbank_requests_status_label() == target_label]
                qs = qs.filter(id__in=matching_ids)
        if deadline_filter:
            now = timezone.now()
            if deadline_filter == 'has_deadline':
                qs = qs.filter(deadline__isnull=False)
            elif deadline_filter == 'past':
                qs = qs.filter(deadline__lt=now)
            elif deadline_filter == 'upcoming':
                qs = qs.filter(deadline__gte=now)
            elif deadline_filter == 'no_deadline':
                qs = qs.filter(deadline__isnull=True)
        delivery_label_map = dict(Donation.DELIVERY_METHODS)
        wb = Workbook()
        ws = wb.active
        ws.title = "Direct Requests"
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.merge_cells('A1:L1')
        ws['A1'].value = "FOODBANKHUB - Direct Requests Report"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        ws.merge_cells('A2:L2')
        ws['A2'].value = f"Generated: {datetime.now().strftime('%B %d, %Y')}  |  Total: {qs.count()}"
        ws['A2'].font = Font(size=9, color="6B7280")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18
        headers = ['S/No', 'Date', 'Category', 'Type', 'Description', 'Food Bank', 'Qty/Amount', 'Delivery', 'Location', 'Notes', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        row_num = 5
        for s_no, req in enumerate(qs, 1):
            linked = getattr(req, 'linked_request_management', None)
            status_label = req.get_foodbank_requests_status_label()
            status_display = get_neutral_status('direct', str(status_label).lower() if status_label else '', {})
            donor_delivery_methods = []
            type_labels = []
            for donation in req.donations.all():
                method = getattr(donation, 'delivery_method', None)
                if method and method not in donor_delivery_methods:
                    donor_delivery_methods.append(method)
                if donation.donation_type == 'item':
                    type_label = 'Free Goods'
                elif donation.donation_type == 'money' and getattr(req, 'original_request', None):
                    type_label = req.original_request.get_request_type_display()
                else:
                    type_label = donation.get_donation_type_display()
                if type_label and type_label not in type_labels:
                    type_labels.append(type_label)
            delivery_display = ", ".join(delivery_label_map.get(m, m.replace('_', ' ').title()) for m in donor_delivery_methods) if donor_delivery_methods else (getattr(linked, 'delivery_method', None) or '-')
            category_display = ", ".join(type_labels) if type_labels else req.get_donation_type_display()
            location_display = getattr(linked, 'location', None) or getattr(req.foodbank, 'address', None) or '-'
            notes_display = getattr(linked, 'additional_notes', None) or req.description or '-'
            row_data = [
                s_no,
                req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else '-',
                category_display,
                req.get_donation_type_display(),
                (req.title or '')[:60] + (('...' if len(req.title or '') > 60 else '')),
                req.foodbank.foodbank_name if req.foodbank else '-',
                req.get_requested_quantity_display() or '-',
                delivery_display,
                location_display,
                (notes_display or '-')[:80],
                status_display,
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            row_num += 1
        column_widths = [6, 14, 12, 12, 28, 20, 16, 14, 18, 24, 16]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = min(width, 50)
        if qs.exists():
            ws.auto_filter.ref = f"A4:L{row_num - 1}"
        ws.freeze_panes = 'A5'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="direct_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response

    # Section provided: export RequestManagement (Specified - used by manage_inventory)
    donations = Donation.objects.filter(
        foodbank_request__isnull=False
    ).select_related(
        'donor__donor_profile',
        'foodbank',
        'foodbank_request__original_request',
        'foodbank_request__original_request__recipient'
    ).order_by('-donated_at')
    if section == 'manage':
        donations = donations.exclude(status='accepted')
    elif section == 'inventory':
        donations = donations.filter(status='accepted')
    
    # Apply filters
    status_filter = request.GET.get('status')
    donation_type_filter = request.GET.get('donation_type')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        donations = donations.filter(status=status_filter)
    if donation_type_filter:
        donations = donations.filter(donation_type=donation_type_filter)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(donor__email__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(item_name__icontains=search) |
            Q(foodbank_request__title__icontains=search)
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Direct Donations"
    
    # Brand colours (FoodBankHub)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Use RequestManagement as the primary data source (matching the template)
    from authentication.models import RequestManagement
    request_qs = RequestManagement.objects.select_related(
        'recipient', 'foodbank', 'assigned_foodbank'
    ).prefetch_related(
        'foodbank_request_created__donations',
        'foodbank_request_created__donations__donor__donor_profile',
    ).order_by('-time_of_request')
    
    # Apply filters to request list (matching the main view)
    if status_filter:
        request_qs = request_qs.filter(status=status_filter)
    if foodbank_filter:
        request_qs = request_qs.filter(
            Q(foodbank_id=foodbank_filter) | Q(assigned_foodbank_id=foodbank_filter)
        )
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            request_qs = request_qs.filter(time_of_request__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            request_qs = request_qs.filter(time_of_request__date__lte=parsed_to)
    if search:
        request_qs = request_qs.filter(
            Q(description__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(location__icontains=search)
        )
    
    # Brand header block
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = "FOODBANKHUB"
    title_cell.font = Font(size=18, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    
    ws.merge_cells('A2:M2')
    tag_cell = ws['A2']
    tag_cell.value = "Connecting Donors to Communities"
    tag_cell.font = Font(size=10, italic=True, color="6B7280")
    tag_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18
    
    report_title = "Specified Donations Report"
    if section == 'manage':
        report_title += " – Manage (Pipeline)"
    elif section == 'inventory':
        report_title += " – Inventory (In Stock)"
    ws.merge_cells('A3:M3')
    report_cell = ws['A3']
    report_cell.value = report_title
    report_cell.font = Font(size=14, bold=True, color="1F4E78")
    report_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 22
    
    ws.merge_cells('A4:M4')
    meta_cell = ws['A4']
    meta_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  |  Total records: {request_qs.count()}"
    meta_cell.font = Font(size=9, color="6B7280")
    meta_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 18
    
    # Table headers (row 6) - matching template columns exactly
    headers = ['S/No', 'Code', 'Type', 'Category', 'Description', 'Recipient', 'Foodbank',
               'Quantity', 'Delivery', 'Location', 'Status', 'Requested', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[6].height = 24
    
    # Data rows
    row_num = 7
    s_no = 1
    for req in request_qs:
        fb_request = req.foodbank_request_created.first() if hasattr(req, 'foodbank_request_created') else None
        donation = fb_request.donations.first() if fb_request else None
        
        # Code
        code = donation.donation_code if donation and donation.donation_code else 'DN-'
        
        # Type
        if req.is_anonymous:
            type_display = 'Anonymous'
        elif donation:
            type_display = donation.get_donation_type_display()
        else:
            type_display = 'Request'
        
        # Category
        cat = req.request_category or req.request_type or ''
        if cat == 'food':
            category_display = 'Food'
        elif cat == 'non_food':
            category_display = 'Non-Food'
        else:
            category_display = 'Other'
        
        # Description
        desc = req.description or ''
        donor_info = ''
        if donation and hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile:
            donor_info = donation.donor.donor_profile.full_name
        elif donation:
            donor_info = donation.donor.email
        if donor_info:
            desc = f"{desc} (Donor: {donor_info})"
        
        # Recipient
        rec_name = req.recipient.full_name if req.recipient else 'N/A'
        rec_email = req.recipient.user.email if req.recipient else ''
        recipient_display = f"{rec_name} ({rec_email})" if rec_email else rec_name
        
        # Foodbank
        fb_name = req.foodbank_name if hasattr(req, 'foodbank_name') and req.foodbank_name else 'N/A'
        
        # Quantity
        qty_requested = f"{req.quantity} {req.get_unit_display()}"
        qty_fulfilled = ''
        if req.quantity_fulfilled and req.quantity_fulfilled > 0:
            qty_fulfilled = f" / Fulfilled: {req.quantity_fulfilled} {req.get_unit_display()}"
            if req.status == 'partial':
                remaining = req.get_remaining_quantity if hasattr(req, 'get_remaining_quantity') else (req.quantity - req.quantity_fulfilled)
                qty_fulfilled += f" / Remaining: {remaining} {req.get_unit_display()}"
        quantity_display = f"Requested: {qty_requested}{qty_fulfilled}"
        
        # Delivery
        delivery = req.get_delivery_method_display() if hasattr(req, 'get_delivery_method_display') else 'N/A'
        
        # Location
        location = req.location or '-'
        
        # Status
        status_display = req.get_display_status() if hasattr(req, 'get_display_status') else req.get_status_display()
        
        # Requested date
        date_str = req.time_of_request.strftime('%Y-%m-%d %H:%M') if req.time_of_request else '-'
        fulfilled_str = ''
        if req.fulfilled_at:
            fulfilled_str = f" (Fulfilled: {req.fulfilled_at.strftime('%b %d')})"
        date_display = f"{date_str}{fulfilled_str}"
        
        # Notes
        notes = req.additional_notes or '-'
        
        row_data = [
            s_no, code, type_display, category_display, desc[:80], recipient_display,
            fb_name, quantity_display, delivery, location, status_display,
            date_display, notes[:100],
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col_num in (5, 6, 8, 13) else 'center')
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        
        ws.row_dimensions[row_num].height = max(18, min(45, 24))
        row_num += 1
        s_no += 1
    
    # Column widths
    column_widths = [6, 12, 14, 12, 28, 24, 20, 24, 12, 18, 18, 18, 28]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = min(width, 50)
    
    # Auto-filter and freeze panes
    if request_qs.exists():
        ws.auto_filter.ref = f"A6:M{row_num - 1}"
    ws.freeze_panes = 'A7'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = "direct_donations"
    if section:
        fname += f"_{section}"
    response['Content-Disposition'] = f'attachment; filename="{fname}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@staff_member_required
def export_direct_donations_pdf(request):
    """
    Export direct donations table to PDF.
    When no section: export FoodBankRequest (Direct Requests).
    When section=manage/inventory: export RequestManagement (Specified).
    """
    from reportlab.platypus import Paragraph
    from authentication.report_utils import (
        get_report_styles,
        build_report_header,
        get_branded_table_style,
        make_full_width_table,
        build_report_summary,
        build_pdf_document,
        collect_active_filters,
    )
    from custom_admin.utils import get_neutral_status

    section = request.GET.get('section', '')
    status_filter = request.GET.get('status')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    donation_type_filter = request.GET.get('donation_type')
    priority_filter = request.GET.get('priority')
    delivery_filter = request.GET.get('delivery')
    response_status_filter = request.GET.get('response_status')
    deadline_filter = request.GET.get('deadline_filter')
    category_filter = request.GET.get('category')

    # When no section: export Direct Requests (FoodBankRequest)
    if not section:
        qs = FoodBankRequest.objects.filter(
            original_request__isnull=True
        ).select_related('foodbank', 'linked_request_management').prefetch_related(
            'donations__donor__donor_profile'
        ).order_by('-created_at')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if foodbank_filter:
            qs = qs.filter(foodbank_id=foodbank_filter)
        if date_from:
            parsed_from = parse_date(date_from)
            if parsed_from:
                qs = qs.filter(created_at__date__gte=parsed_from)
        if date_to:
            parsed_to = parse_date(date_to)
            if parsed_to:
                qs = qs.filter(created_at__date__lte=parsed_to)
        if search:
            qs = qs.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(foodbank__foodbank_name__icontains=search)
            )
        if donation_type_filter:
            qs = qs.filter(donation_type=donation_type_filter)
        if priority_filter:
            qs = qs.filter(priority=priority_filter)
        if delivery_filter:
            qs = qs.filter(donations__delivery_method=delivery_filter).distinct()
        if category_filter:
            qs = qs.filter(donations__donation_type=category_filter).distinct()
        if response_status_filter:
            if response_status_filter == 'sent_to_donors':
                qs = qs.filter(donations__isnull=True)
            elif response_status_filter == 'donation_made':
                qs = qs.filter(donations__status='pending').exclude(
                    donations__status='accepted'
                ).distinct()
            elif response_status_filter == 'declined':
                qs = qs.filter(donations__status='declined').exclude(
                    donations__status='accepted'
                ).exclude(donations__status='pending').distinct()
            elif response_status_filter in ('partially_fulfilled', 'fulfilled'):
                qs = qs.filter(donations__status='accepted').distinct()
                requests_list = list(qs.prefetch_related('donations'))
                target_label = 'Partially Fulfilled' if response_status_filter == 'partially_fulfilled' else 'Fulfilled'
                matching_ids = [r.id for r in requests_list if r.get_foodbank_requests_status_label() == target_label]
                qs = qs.filter(id__in=matching_ids)
        if deadline_filter:
            now = timezone.now()
            if deadline_filter == 'has_deadline':
                qs = qs.filter(deadline__isnull=False)
            elif deadline_filter == 'past':
                qs = qs.filter(deadline__lt=now)
            elif deadline_filter == 'upcoming':
                qs = qs.filter(deadline__gte=now)
            elif deadline_filter == 'no_deadline':
                qs = qs.filter(deadline__isnull=True)
        delivery_label_map = dict(Donation.DELIVERY_METHODS)
        styles = get_report_styles()
        active_filters = collect_active_filters(request, [
            ('status', 'Status'), ('foodbank', 'Foodbank'),
            ('date_from', 'From'), ('date_to', 'To'), ('search', 'Search'),
            ('donation_type', 'Type'), ('priority', 'Priority'), ('delivery', 'Delivery'),
            ('category', 'Category'), ('response_status', 'Response status'), ('deadline_filter', 'Deadline'),
        ])
        elements = []
        build_report_header(
            elements,
            report_title="Direct Requests Report",
            generated_for="Admin",
            total_records=qs.count(),
            active_filters=active_filters or None,
            styles_dict=styles,
        )
        header_row = ['S/No', 'Date', 'Category', 'Type', 'Description', 'Food Bank', 'Qty/Amount', 'Delivery', 'Location', 'Notes', 'Status']
        table_data = [header_row]
        wrap = styles['wrap']
        wrap_center = styles['wrap_center']
        for s_no, req in enumerate(qs, 1):
            linked = getattr(req, 'linked_request_management', None)
            status_label = req.get_foodbank_requests_status_label()
            status_display = get_neutral_status('direct', str(status_label).lower() if status_label else '', {})
            donor_delivery_methods = []
            type_labels = []
            for donation in req.donations.all():
                method = getattr(donation, 'delivery_method', None)
                if method and method not in donor_delivery_methods:
                    donor_delivery_methods.append(method)
                if donation.donation_type == 'item':
                    type_label = 'Free Goods'
                elif donation.donation_type == 'money' and getattr(req, 'original_request', None):
                    type_label = req.original_request.get_request_type_display()
                else:
                    type_label = donation.get_donation_type_display()
                if type_label and type_label not in type_labels:
                    type_labels.append(type_label)
            delivery_display = ", ".join(delivery_label_map.get(m, m.replace('_', ' ').title()) for m in donor_delivery_methods) if donor_delivery_methods else (getattr(linked, 'delivery_method', None) or '-')
            category_display = ", ".join(type_labels) if type_labels else req.get_donation_type_display()
            location_display = getattr(linked, 'location', None) or getattr(req.foodbank, 'address', None) or '-'
            notes_display = (getattr(linked, 'additional_notes', None) or req.description or '-')[:50]
            table_data.append([
                Paragraph(str(s_no), wrap_center),
                Paragraph(req.created_at.strftime('%b %d, %Y') if req.created_at else '-', wrap_center),
                Paragraph(category_display, wrap_center),
                Paragraph(req.get_donation_type_display(), wrap_center),
                Paragraph((req.title or '')[:40], wrap),
                Paragraph(req.foodbank.foodbank_name if req.foodbank else '-', wrap_center),
                Paragraph(req.get_requested_quantity_display() or '-', wrap_center),
                Paragraph(delivery_display[:16], wrap_center),
                Paragraph(location_display[:20], wrap),
                Paragraph(notes_display, wrap),
                Paragraph(status_display, wrap_center),
            ])
        col_weights = [0.4, 0.9, 0.8, 0.6, 1.4, 1.0, 0.9, 0.8, 0.9, 1.0, 0.8]
        table = make_full_width_table(table_data, repeat_rows=1, col_weights=col_weights)
        table.setStyle(get_branded_table_style(len(table_data)))
        elements.append(table)
        summary_items = [("Total records", qs.count())]
        active_count = qs.filter(status='active').count()
        fulfilled_count = qs.filter(status='fulfilled').count()
        expired_count = qs.filter(status='expired').count()
        summary_items.extend([("Active", active_count), ("Fulfilled", fulfilled_count), ("Expired", expired_count)])
        build_report_summary(elements, summary_items, styles_dict=styles)
        return build_pdf_document(elements, filename_prefix="direct_requests", user_display_name="Admin")

    # Section provided: export RequestManagement
    donations = Donation.objects.filter(
        foodbank_request__isnull=False
    ).select_related(
        'donor__donor_profile',
        'foodbank',
        'foodbank_request__original_request',
        'foodbank_request__original_request__recipient'
    ).order_by('-donated_at')
    if section == 'manage':
        donations = donations.exclude(status='accepted')
    elif section == 'inventory':
        donations = donations.filter(status='accepted')
    
    donation_type_filter = request.GET.get('donation_type')
    
    if status_filter:
        donations = donations.filter(status=status_filter)
    if donation_type_filter:
        donations = donations.filter(donation_type=donation_type_filter)
    if foodbank_filter:
        donations = donations.filter(foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations = donations.filter(donated_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations = donations.filter(donated_at__date__lte=parsed_to)
    if search:
        donations = donations.filter(
            Q(donor__email__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(item_name__icontains=search) |
            Q(foodbank_request__title__icontains=search)
        )
    
    # Use RequestManagement as primary data source (matching the template)
    from authentication.models import RequestManagement
    request_qs = RequestManagement.objects.select_related(
        'recipient', 'foodbank', 'assigned_foodbank'
    ).prefetch_related(
        'foodbank_request_created__donations',
        'foodbank_request_created__donations__donor__donor_profile',
    ).order_by('-time_of_request')
    
    # Apply filters to request list
    if status_filter:
        request_qs = request_qs.filter(status=status_filter)
    if foodbank_filter:
        request_qs = request_qs.filter(
            Q(foodbank_id=foodbank_filter) | Q(assigned_foodbank_id=foodbank_filter)
        )
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            request_qs = request_qs.filter(time_of_request__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            request_qs = request_qs.filter(time_of_request__date__lte=parsed_to)
    if search:
        request_qs = request_qs.filter(
            Q(description__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(location__icontains=search)
        )
    
    requests_list = list(request_qs)
    total_count = len(requests_list)
    
    styles = get_report_styles()
    report_title = "Specified Donations Report"
    if section == 'manage':
        report_title += " – Manage (Pipeline)"
    elif section == 'inventory':
        report_title += " – Inventory (In Stock)"
    
    active_filters = collect_active_filters(request, [
        ('section', 'Section'),
        ('status', 'Status'),
        ('foodbank', 'Foodbank'),
        ('date_from', 'From'),
        ('date_to', 'To'),
        ('search', 'Search'),
    ])
    
    elements = []
    build_report_header(
        elements,
        report_title=report_title,
        generated_for="Admin",
        total_records=total_count,
        active_filters=active_filters or None,
        styles_dict=styles,
    )
    
    # Table - matching template columns exactly
    header_row = ['S/No', 'Code', 'Type', 'Category', 'Description', 'Recipient', 'Foodbank',
                  'Quantity', 'Delivery', 'Location', 'Status', 'Requested', 'Notes']
    table_data = [header_row]
    
    wrap = styles['wrap']
    wrap_center = styles['wrap_center']
    
    for s_no, req in enumerate(requests_list, 1):
        fb_request = req.foodbank_request_created.first() if hasattr(req, 'foodbank_request_created') else None
        donation = fb_request.donations.first() if fb_request else None
        
        # Code
        code = (donation.donation_code or 'DN-')[:10] if donation else 'DN-'
        
        # Type
        if req.is_anonymous:
            type_display = 'Anonymous'
        elif donation:
            type_display = donation.get_donation_type_display()[:12]
        else:
            type_display = 'Request'
        
        # Category
        cat = req.request_category or req.request_type or ''
        if cat == 'food':
            category_display = 'Food'
        elif cat == 'non_food':
            category_display = 'Non-Food'
        else:
            category_display = 'Other'
        
        # Description (with donor info)
        desc = (req.description or '')[:50]
        donor_info = ''
        if donation and hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile:
            donor_info = donation.donor.donor_profile.full_name[:20]
        elif donation:
            donor_info = donation.donor.email[:20]
        if donor_info:
            desc_display = f"{desc}<br/><font size='6' color='#6b7280'>Donor: {donor_info}</font>"
        else:
            desc_display = desc
        
        # Recipient
        rec_name = (req.recipient.full_name if req.recipient else 'N/A')[:20]
        
        # Foodbank
        fb_name = (req.foodbank_name if hasattr(req, 'foodbank_name') and req.foodbank_name else 'N/A')[:20]
        
        # Quantity
        qty_str = f"<b>{req.quantity} {req.get_unit_display()}</b><br/><font size='6' color='#6b7280'>Requested</font>"
        if req.quantity_fulfilled and req.quantity_fulfilled > 0:
            qty_str += f"<br/><font color='green'>{req.quantity_fulfilled} {req.get_unit_display()}</font><br/><font size='6' color='#6b7280'>Fulfilled</font>"
            if req.status == 'partial':
                remaining = req.get_remaining_quantity if hasattr(req, 'get_remaining_quantity') else (req.quantity - req.quantity_fulfilled)
                qty_str += f"<br/><font color='#d97706'>{remaining} {req.get_unit_display()}</font><br/><font size='6' color='#6b7280'>Remaining</font>"
        
        # Delivery
        delivery = (req.get_delivery_method_display() if hasattr(req, 'get_delivery_method_display') else 'N/A')[:12]
        
        # Location
        location = (req.location or '-')[:22]
        
        # Status
        status_display = (req.get_display_status() if hasattr(req, 'get_display_status') else req.get_status_display())[:20]
        
        # Date
        date_str = req.time_of_request.strftime('%b %d, %Y') if req.time_of_request else '-'
        if req.fulfilled_at:
            date_str += f"<br/><font size='6' color='green'>Done: {req.fulfilled_at.strftime('%b %d')}</font>"
        
        # Notes
        notes = (req.additional_notes or '-')[:50]
        
        table_data.append([
            Paragraph(str(s_no), wrap_center),
            Paragraph(code, wrap_center),
            Paragraph(type_display, wrap_center),
            Paragraph(category_display, wrap_center),
            Paragraph(desc_display, wrap),
            Paragraph(rec_name, wrap),
            Paragraph(fb_name, wrap),
            Paragraph(qty_str, wrap),
            Paragraph(delivery, wrap_center),
            Paragraph(location, wrap),
            Paragraph(status_display, wrap_center),
            Paragraph(date_str, wrap_center),
            Paragraph(notes, wrap),
        ])
    
    col_weights = [0.4, 0.7, 0.8, 0.6, 1.5, 1.0, 0.9, 1.2, 0.6, 0.9, 0.9, 0.9, 1.2]
    table = make_full_width_table(table_data, repeat_rows=1, col_weights=col_weights)
    table.setStyle(get_branded_table_style(len(table_data)))
    elements.append(table)
    
    summary_items = [("Total records", total_count)]
    pending_count = sum(1 for r in requests_list if r.status == 'pending')
    fulfilled_count = sum(1 for r in requests_list if r.status == 'fulfilled')
    partial_count = sum(1 for r in requests_list if r.status == 'partial')
    summary_items.extend([("Pending", pending_count), ("Fulfilled", fulfilled_count), ("Partial", partial_count)])
    
    build_report_summary(elements, summary_items, styles_dict=styles)
    
    fname = "specified_donations"
    if section:
        fname += f"_{section}"
    return build_pdf_document(elements, filename_prefix=fname, user_display_name="Admin")


@staff_member_required
def request_management_admin(request):
    """
    Unified request management system
    """
    requests = RequestManagement.objects.select_related(
        'recipient',
        'foodbank',
        'assigned_foodbank',
        'donation',
        'foodbank_request'
    ).order_by('-time_of_request')
    
    # Statistics
    total_requests = requests.count()
    pending = requests.filter(status='pending').count()
    assigned = requests.filter(status='assigned').count()
    fulfilled = requests.filter(status='fulfilled').count()
    partial = requests.filter(status='partial').count()
    declined = requests.filter(status='declined').count()
    anonymous_requests = requests.filter(is_anonymous=True).count()
    
    # Filtering
    status_filter = request.GET.get('status')
    request_type_filter = request.GET.get('request_type')
    foodbank_filter = request.GET.get('foodbank')
    anonymous_filter = request.GET.get('anonymous')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        requests = requests.filter(status=status_filter)
    if request_type_filter:
        requests = requests.filter(request_type=request_type_filter)
    if foodbank_filter:
        requests = requests.filter(Q(foodbank_id=foodbank_filter) | Q(assigned_foodbank_id=foodbank_filter))
    if anonymous_filter == 'yes':
        requests = requests.filter(is_anonymous=True)
    elif anonymous_filter == 'no':
        requests = requests.filter(is_anonymous=False)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            requests = requests.filter(time_of_request__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            requests = requests.filter(time_of_request__date__lte=parsed_to)
    if search:
        requests = requests.filter(
            Q(description__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(location__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    
    context = {
        'title': 'Request Management',
        'page_obj': page_obj,
        'total_requests': total_requests,
        'pending': pending,
        'assigned': assigned,
        'fulfilled': fulfilled,
        'partial': partial,
        'declined': declined,
        'anonymous_requests': anonymous_requests,
        'foodbanks': foodbanks,
        'status_choices': RequestManagement.STATUS_CHOICES,
        'request_type_choices': RequestManagement.REQUEST_TYPE_CHOICES,
        'current_filters': {
            'status': status_filter,
            'request_type': request_type_filter,
            'foodbank': foodbank_filter,
            'anonymous': anonymous_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    return render(request, 'custom_admin/request_management_admin.html', context)


@staff_member_required
def donation_responses_management(request):
    """
    Management of recipient responses to subsidized donations
    """
    responses = DonationResponse.objects.select_related(
        'donation__donor',
        'donation__foodbank',
        'recipient'
    ).order_by('-responded_at')
    
    # Statistics
    total_responses = responses.count()
    accepted = responses.filter(response_type='accepted').count()
    declined = responses.filter(response_type='declined').count()
    
    # Filtering
    response_filter = request.GET.get('response')
    foodbank_filter = request.GET.get('foodbank')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if response_filter:
        responses = responses.filter(response_type=response_filter)
    if foodbank_filter:
        responses = responses.filter(donation__foodbank_id=foodbank_filter)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            responses = responses.filter(responded_at__date__gte=parsed_from)
    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            responses = responses.filter(responded_at__date__lte=parsed_to)
    if search:
        responses = responses.filter(
            Q(donation__donor__email__icontains=search) |
            Q(donation__foodbank__foodbank_name__icontains=search) |
            Q(recipient__full_name__icontains=search) |
            Q(donation__subsidized_product_type__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(responses, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    
    context = {
        'title': 'Donation Responses Management',
        'page_obj': page_obj,
        'total_responses': total_responses,
        'accepted': accepted,
        'declined': declined,
        'foodbanks': foodbanks,
        'response_choices': DonationResponse.RESPONSE_CHOICES,
        'current_filters': {
            'response': response_filter,
            'foodbank': foodbank_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    return render(request, 'custom_admin/donation_responses_management.html', context)
