from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from datetime import timedelta
from .models import Donation, DonationResponse, UnspecifiedDonationManagement, FoodBankProfile
from .decorators import donor_required


def _unspecified_export_category(d):
    """Type-column display for unspecified exports: donation category (CSR includes sub-type)."""
    if d.donation_type == 'csr' or getattr(d, 'donation_category', None) == 'csr':
        if getattr(d, 'csr_subcategory', None) == 'other' and getattr(d, 'csr_custom_subcategory', None):
            csr_type = d.csr_custom_subcategory
        elif getattr(d, 'csr_subcategory', None):
            csr_type = d.get_csr_subcategory_display()
        else:
            csr_type = 'CSR Initiative'
        return f"CSR\n{csr_type}"
    return d.get_donation_category_display()


def _unspecified_export_donation_type_label(d):
    """Category-column base label for unspecified exports."""
    return (
        'Free Goods' if (d.donation_type == 'item' and getattr(d, 'donation_mode', None) == 'free') else
        'Subsidized' if d.donation_type == 'subsidized' or (d.donation_type == 'item' and getattr(d, 'donation_mode', None) == 'subsidized') else
        'Monetary' if d.donation_type == 'money' else
        'CSR' if d.donation_type == 'csr' else
        d.get_donation_type_display()
    )


def _unspecified_export_type_category_pair(d):
    """
    Return (type_display, category_display).
    For CSR donations, swap Type and Category values to match table behavior.
    """
    type_display = _unspecified_export_category(d)
    category_display = _unspecified_export_donation_type_label(d)
    if d.donation_type == 'csr' or getattr(d, 'donation_category', None) == 'csr':
        return category_display, type_display
    return type_display, category_display


def _unspecified_export_description(d):
    """Description for unspecified donations export - match table: item name, product type, csr_description, other_description, or message for money."""
    if d.donation_type == 'item':
        if getattr(d, 'donation_mode', None) == 'free':
            return (d.item_name or 'General donation')[:300]
        return (d.subsidized_product_type or d.item_name or 'Subsidized goods')[:300]
    if d.donation_type == 'subsidized':
        return (d.subsidized_product_type or 'Subsidized goods')[:300]
    if d.donation_type == 'csr':
        return (d.csr_description or '')[:300]
    if d.donation_type == 'other':
        return (d.other_description or 'Other donation')[:300]
    if d.donation_type == 'money':
        return (d.message or 'Monetary donation')[:300]
    return ''


def _unspecified_export_qty_amount(d):
    """Qty/Amount text to match donor unspecified table formatting."""
    if d.donation_type == 'item':
        if d.quantity:
            return f"{d.quantity:,} {d.quantity_unit or 'units'}"
        return '-'

    if d.donation_type == 'money':
        if d.amount is not None:
            return f"KES {d.amount:,.0f}"
        return '-'

    if d.donation_type in ('csr', 'other'):
        has_amount = d.amount is not None
        has_quantity = bool(d.quantity)
        unit = d.quantity_unit or 'units'

        if has_amount and has_quantity:
            return f"KES {d.amount:,.0f} for {d.quantity:,} {unit}"
        if has_amount:
            return f"KES {d.amount:,.0f}"
        if has_quantity:
            return f"{d.quantity:,} {unit}"
        return '-'

    return '-'


def _build_unspecified_effective_note_maps(donations_data):
    """
    Build effective recipient note and recipient decline note maps for unspecified exports.
    If donation is accepted/received by any recipient, recipient decline note must be blank.
    """
    item_list = list(donations_data)
    donation_ids = [item.donation_id for item in item_list if getattr(item, 'donation_id', None)]
    if not donation_ids:
        return {}, {}

    latest_non_decline_notes_by_donation = {}
    latest_declines_by_donation = {}
    accepted_notes_by_donation = {}
    accepted_recipient_by_donation = {
        item.donation_id: item.accepted_by_recipient_id
        for item in item_list
        if getattr(item, 'accepted_by_recipient_id', None)
    }

    responses = DonationResponse.objects.filter(
        donation_id__in=donation_ids
    ).exclude(
        notes__isnull=True
    ).exclude(
        notes__exact=''
    ).select_related('recipient').order_by('-responded_at')

    for response in responses:
        donation_id = response.donation_id

        if response.response_type == 'declined':
            if donation_id not in latest_declines_by_donation:
                latest_declines_by_donation[donation_id] = response.notes
        elif donation_id not in latest_non_decline_notes_by_donation:
            latest_non_decline_notes_by_donation[donation_id] = response.notes

        if (
            response.response_type == 'accepted'
            and accepted_recipient_by_donation.get(donation_id) == response.recipient_id
            and donation_id not in accepted_notes_by_donation
        ):
            accepted_notes_by_donation[donation_id] = response.notes

    effective_recipient_notes = {}
    effective_recipient_declines = {}

    for item in item_list:
        donation_id = item.donation_id
        base_recipient_note = (getattr(item, 'recipient_notes', None) or '').strip()
        base_decline_note = (getattr(item, 'recipient_decline_reason', None) or '').strip()
        accepted_or_received = bool(getattr(item, 'accepted_by_recipient_id', None)) or getattr(item, 'recipient_status', None) in (
            'accepted_by_recipient', 'received'
        )

        if accepted_or_received:
            effective_recipient_notes[donation_id] = (
                accepted_notes_by_donation.get(donation_id)
                or base_recipient_note
                or ''
            )
            effective_recipient_declines[donation_id] = ''
        else:
            effective_recipient_notes[donation_id] = (
                latest_non_decline_notes_by_donation.get(donation_id)
                or base_recipient_note
                or ''
            )
            effective_recipient_declines[donation_id] = (
                base_decline_note
                or latest_declines_by_donation.get(donation_id)
                or ''
            )

    return effective_recipient_notes, effective_recipient_declines


@login_required
@donor_required
def donor_unspecified_export(request, format):
    """Export unspecified donations in PDF or CSV format"""
    # Get all unspecified donations for this donor with same filters
    unspecified_donations = UnspecifiedDonationManagement.objects.filter(
        donation__donor=request.user
    ).select_related(
        'donation', 'donation__foodbank', 'accepted_by_recipient'
    ).order_by('-created_at')
    
    # Apply same filters as donor_unspecified_donations_detail (aligned with table columns)
    type_filter = (request.GET.get('type') or 'all').strip().lower()
    if type_filter not in ('all', 'food', 'non_food'):
        type_filter = 'all'
    category_filter = (request.GET.get('category') or 'all').strip().lower()
    foodbank_location_filter = request.GET.get('foodbank_location', '').strip()
    foodbank_status_filter = request.GET.get('foodbank_status', 'all')
    recipient_status_filter = request.GET.get('recipient_status', 'all')
    delivery_filter = request.GET.get('delivery', 'all')
    quantity_range = request.GET.get('quantity_range', 'all').strip()
    amount_range = request.GET.get('amount_range', 'all').strip()
    date_filter = request.GET.get('date_range', 'all')
    search_query = request.GET.get('search', '').strip()

    # Normalize delivery value for UI/backward compatibility.
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    
    if type_filter != 'all':
        # "Type" filter maps to donation_category in this table.
        unspecified_donations = unspecified_donations.filter(donation__donation_category=type_filter)
    
    if category_filter != 'all':
        # "Category" filter maps to donation_type in this table.
        unspecified_donations = unspecified_donations.filter(donation__donation_type=category_filter)

    if foodbank_location_filter:
        unspecified_donations = unspecified_donations.filter(
            donation__foodbank__address__icontains=foodbank_location_filter
        )
    
    if foodbank_status_filter != 'all':
        unspecified_donations = unspecified_donations.filter(foodbank_status=foodbank_status_filter)
    
    if recipient_status_filter != 'all':
        unspecified_donations = unspecified_donations.filter(recipient_status=recipient_status_filter)
    
    if delivery_filter != 'all':
        if delivery_filter == 'delivery':
            unspecified_donations = unspecified_donations.filter(
                donation__delivery_method__in=['dropoff', 'delivery']
            )
        else:
            unspecified_donations = unspecified_donations.filter(donation__delivery_method=delivery_filter)

    quantity_range_map = {
        '1-100': (1, 100),
        '101-500': (101, 500),
        '501-1000': (501, 1000),
        '1001-5000': (1001, 5000),
        '5001_or_more': (5001, None),
    }
    qty_min, qty_max = quantity_range_map.get(quantity_range, (None, None))
    if qty_min is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__quantity__gte=qty_min) |
            Q(donation__subsidized_quantity__gte=qty_min)
        )
    if qty_max is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__quantity__lte=qty_max) |
            Q(donation__subsidized_quantity__lte=qty_max)
        )

    amount_range_map = {
        '1-10000': (1, 10000),
        '10001-50000': (10001, 50000),
        '50001-100000': (50001, 100000),
        '100001-500000': (100001, 500000),
        '500001_or_more': (500001, None),
    }
    amt_min, amt_max = amount_range_map.get(amount_range, (None, None))
    if amt_min is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__amount__gte=amt_min) |
            Q(donation__subsidized_price__gte=amt_min)
        )
    if amt_max is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__amount__lte=amt_max) |
            Q(donation__subsidized_price__lte=amt_max)
        )
    
    if search_query:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__item_name__icontains=search_query) |
            Q(donation__subsidized_product_type__icontains=search_query) |
            Q(donation__csr_description__icontains=search_query) |
            Q(donation__other_description__icontains=search_query) |
            Q(donation__foodbank__foodbank_name__icontains=search_query) |
            Q(donation__message__icontains=search_query) |
            Q(recipient_notes__icontains=search_query) |
            Q(foodbank_decline_reason__icontains=search_query) |
            Q(recipient_decline_reason__icontains=search_query)
        )
    
    # Filter by date range
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    if date_filter == 'custom' and (date_from or date_to):
        if date_from:
            try:
                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
                unspecified_donations = unspecified_donations.filter(created_at__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                unspecified_donations = unspecified_donations.filter(created_at__lte=to_date)
            except ValueError:
                pass
    elif date_filter != 'all':
        now = timezone.now()
        if date_filter == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            unspecified_donations = unspecified_donations.filter(created_at__gte=start_date)
        elif date_filter == 'week':
            start_date = now - timedelta(days=7)
            unspecified_donations = unspecified_donations.filter(created_at__gte=start_date)
        elif date_filter == 'month':
            start_date = now - timedelta(days=30)
            unspecified_donations = unspecified_donations.filter(created_at__gte=start_date)
        elif date_filter == '3months':
            start_date = now - timedelta(days=90)
            unspecified_donations = unspecified_donations.filter(created_at__gte=start_date)
    
    donations_data = list(unspecified_donations)
    effective_recipient_notes, effective_recipient_declines = _build_unspecified_effective_note_maps(donations_data)
    for item in donations_data:
        item.effective_recipient_note = effective_recipient_notes.get(item.donation_id, '') or ''
        item.effective_recipient_decline_note = effective_recipient_declines.get(item.donation_id, '') or ''
    
    if format.lower() == 'pdf':
        return export_unspecified_pdf(request, donations_data)
    elif format.lower() == 'csv':
        return export_unspecified_csv(request, donations_data)
    elif format.lower() == 'excel':
        return export_unspecified_excel(request, donations_data)
    else:
        messages.error(request, 'Invalid export format.')
        return redirect('donor_unspecified_donations_detail')


def export_unspecified_pdf(request, donations_data):
    """Generate PDF report for unspecified donations using shared FoodBankHub branding."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.pagesizes import A3, landscape
    from .report_utils import (
        get_report_styles,
        build_report_header,
        get_branded_table_style,
        build_report_summary,
        build_pdf_document,
        collect_active_filters,
        make_full_width_table,
    )

    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    styles_dict = get_report_styles()
    report_pagesize = landscape(A3)
    filter_keys = [
        ('type', 'Type'), ('foodbank_location', 'Food Bank Location'), ('category', 'Category'),
        ('foodbank_status', 'Status (FB)'), ('recipient_status', 'Recipient'),
        ('delivery', 'Delivery'),
        ('quantity_range', 'Quantity Range'),
        ('amount_range', 'Amount Range'),
        ('search', 'Search'),
    ]
    active_filters = collect_active_filters(request, filter_keys)

    build_report_header(
        elements := [],
        report_title="Unspecified Donations Report",
        generated_for=donor_name,
        total_records=len(donations_data),
        active_filters=active_filters,
        styles_dict=styles_dict,
    )

    if not donations_data:
        elements.append(Paragraph("No donations found matching the current filters.", styles_dict['normal']))
    else:
        wrap = styles_dict['wrap']
        data = [[
            'S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank',
            'Qty/Amount', 'Delivery', 'Location', 'Recipient', 'Donor Note',
            'Recipient Note', 'Recipient Decline Reason', 'Foodbank Decline Note', 'Status'
        ]]

        for idx, item in enumerate(donations_data, start=1):
            date_str = item.created_at.strftime('%b %d, %Y')
            d = item.donation
            details = _unspecified_export_description(d)
            qty_amount = _unspecified_export_qty_amount(d)

            type_display, category_label = _unspecified_export_type_category_pair(d)
            delivery = (
                'Delivery'
                if d.delivery_method in ['dropoff', 'delivery']
                else (d.get_delivery_method_display() if d.delivery_method else '-')
            )
            location = (d.foodbank.address or 'No location specified')[:60]
            recipient_name = (item.accepted_by_recipient.full_name if item.accepted_by_recipient else 'Unclaimed')[:40]
            donor_note = (d.message or d.csr_description or d.other_description or 'No donor note')[:120]
            recipient_note = ((getattr(item, 'effective_recipient_note', None) or item.recipient_notes or 'No recipient note'))[:120]
            recipient_decline_note = ((getattr(item, 'effective_recipient_decline_note', None) or '')[:160])
            foodbank_note = 'no decline note'
            if item.foodbank_status == 'declined_by_foodbank' and item.foodbank_decline_reason:
                foodbank_note = str(item.foodbank_decline_reason)[:160]

            # Status: match table exactly (same labels as donor_unspecified_donations_detail.html)
            if item.foodbank_status == 'accepted_by_foodbank':
                if item.recipient_status == 'received':
                    status = 'Received by recipient'
                elif item.recipient_status == 'accepted_by_recipient':
                    status = 'Claimed by recipient'
                elif item.recipient_status == 'declined_by_recipient':
                    status = 'Declined by recipient -  broadcasted to other recipients'
                else:
                    status = 'Accepted by FB - Sent to Recipients'
            elif item.foodbank_status == 'pending_foodbank':
                status = 'Awaiting foodbank'
            elif item.foodbank_status == 'declined_by_foodbank':
                status = 'Declined by foodbank'
            else:
                status = item.get_foodbank_status_display() or '-'

            type_para = Paragraph(type_display.replace('\n', '<br/>'), wrap)
            category_para = Paragraph(str(category_label).replace('\n', '<br/>'), wrap)
            desc_para = Paragraph((details or '—').replace('\n', '<br/>'), wrap)
            qty_amount_para = Paragraph(str(qty_amount).replace('\n', '<br/>'), wrap)
            data.append([
                str(idx), date_str, type_para, category_para,
                desc_para,
                Paragraph((d.foodbank.foodbank_name or '')[:50], wrap),
                qty_amount_para, delivery, Paragraph(location, wrap),
                Paragraph(recipient_name, wrap), Paragraph(donor_note, wrap), Paragraph(recipient_note, wrap),
                Paragraph(recipient_decline_note.replace('\n', '<br/>'), wrap),
                Paragraph(foodbank_note.replace('\n', '<br/>'), wrap),
                Paragraph(status, wrap),
            ])

        _col_weights = [0.028, 0.05, 0.05, 0.06, 0.085, 0.072, 0.065, 0.045, 0.06, 0.06, 0.072, 0.07, 0.09, 0.085, 0.096]
        table = make_full_width_table(data, repeat_rows=1, col_weights=_col_weights, pagesize=report_pagesize)
        table.setStyle(get_branded_table_style(len(data)))
        elements.append(table)

    build_report_summary(elements, [("Total Unspecified Donations", len(donations_data))], styles_dict=styles_dict)
    return build_pdf_document(elements, "unspecified_donations", donor_name, pagesize=report_pagesize)


def export_unspecified_csv(request, donations_data):
    """Generate CSV report for unspecified donations"""
    import csv
    import datetime
    
    response = HttpResponse(content_type='text/csv')
    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    response['Content-Disposition'] = f'attachment; filename="unspecified_donations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Donor', donor_name])
    writer.writerow(['Report Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    
    writer.writerow([
        'S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity', 'Unit',
        'Amount(ksh)', 'Delivery', 'Location', 'Recipient', 'Donor Note', 'Recipient Note',
        'Recipient Decline Reason', 'Foodbank Decline Note', 'Status'
    ])
    
    for idx, item in enumerate(donations_data, start=1):
        date_str = item.created_at.strftime('%Y-%m-%d %H:%M')
        d = item.donation
        details = _unspecified_export_description(d)
        type_display, category_label = _unspecified_export_type_category_pair(d)
        quantity_value = ''
        unit_value = ''
        amount_value = ''
        if d.donation_type == 'item':
            quantity_value = d.quantity
            unit_value = d.quantity_unit or 'units'
        elif d.donation_type == 'money':
            amount_value = f"{d.amount:.0f}" if d.amount is not None else ''
        elif d.donation_type in ('csr', 'other'):
            if d.quantity:
                quantity_value = d.quantity
                unit_value = d.quantity_unit or 'units'
            if d.amount is not None:
                amount_value = f"{d.amount:.0f}"
        delivery = (
            'Delivery'
            if d.delivery_method in ['dropoff', 'delivery']
            else (d.get_delivery_method_display() if d.delivery_method else '-')
        )
        location = (d.foodbank.address or 'No location specified').replace('\n', ' ')
        recipient_name = item.accepted_by_recipient.full_name if item.accepted_by_recipient else 'Unclaimed'
        donor_note = (d.message or d.csr_description or d.other_description or 'No donor note').replace('\n', ' | ').replace('\r', '')
        recipient_note = (getattr(item, 'effective_recipient_note', None) or item.recipient_notes or 'No recipient note').replace('\n', ' | ').replace('\r', '')
        recipient_decline_note = (getattr(item, 'effective_recipient_decline_note', None) or '').replace('\n', ' | ').replace('\r', '')
        foodbank_note = 'no decline note'
        if item.foodbank_status == 'declined_by_foodbank' and item.foodbank_decline_reason:
            foodbank_note = str(item.foodbank_decline_reason).replace('\n', ' | ').replace('\r', '')
        if item.foodbank_status == 'accepted_by_foodbank':
            if item.recipient_status == 'received':
                status = 'Received by recipient'
            elif item.recipient_status == 'accepted_by_recipient':
                status = 'Claimed by recipient'
            elif item.recipient_status == 'declined_by_recipient':
                status = 'Rejected by recipient - broadcasted to other recipients'
            else:
                status = 'Accepted by FB - Sent to Recipients'
        elif item.foodbank_status == 'pending_foodbank':
            status = 'Awaiting foodbank'
        elif item.foodbank_status == 'declined_by_foodbank':
            status = 'Declined by foodbank'
        else:
            status = item.get_foodbank_status_display() or '-'
        
        writer.writerow([
            idx,
            date_str,
            type_display,
            category_label,
            details,
            d.foodbank.foodbank_name,
            quantity_value,
            unit_value,
            amount_value,
            delivery,
            location,
            recipient_name,
            donor_note,
            recipient_note,
            recipient_decline_note,
            foodbank_note,
            status
        ])
    
    return response

def export_unspecified_excel(request, donations_data):
    """Generate Excel report for unspecified donations"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Unspecified Donations"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title and info
    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = f"{donor_name} - Unspecified Donations Report"
    title_cell.font = Font(bold=True, size=14, color="10b981")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:Q2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total: {len(donations_data)}"
    info_cell.alignment = Alignment(horizontal="center")

    headers = [
        'S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity', 'Unit',
        'Amount(ksh)', 'Delivery', 'Location', 'Recipient', 'Donor Note', 'Recipient Note',
        'Recipient Decline Reason', 'Foodbank Decline Note', 'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for idx, item in enumerate(donations_data, 1):
        row_num = idx + 4
        d = item.donation
        date_str = item.created_at.strftime('%Y-%m-%d %H:%M')
        details = _unspecified_export_description(d)
        type_display, category_label = _unspecified_export_type_category_pair(d)
        quantity_value = ''
        unit_value = ''
        amount_value = ''
        if d.donation_type == 'item':
            quantity_value = d.quantity
            unit_value = d.quantity_unit or 'units'
        elif d.donation_type == 'money':
            amount_value = float(d.amount) if d.amount is not None else ''
        elif d.donation_type in ('csr', 'other'):
            if d.quantity:
                quantity_value = d.quantity
                unit_value = d.quantity_unit or 'units'
            if d.amount is not None:
                amount_value = float(d.amount)
        delivery = (
            'Delivery'
            if d.delivery_method in ['dropoff', 'delivery']
            else (d.get_delivery_method_display() if d.delivery_method else '-')
        )
        location = d.foodbank.address or 'No location specified'
        recipient_name = item.accepted_by_recipient.full_name if item.accepted_by_recipient else 'Unclaimed'
        donor_note = d.message or d.csr_description or d.other_description or 'No donor note'
        recipient_note = getattr(item, 'effective_recipient_note', None) or item.recipient_notes or 'No recipient note'
        recipient_decline_note = getattr(item, 'effective_recipient_decline_note', None) or ''
        foodbank_note = 'no decline note'
        if item.foodbank_status == 'declined_by_foodbank' and item.foodbank_decline_reason:
            foodbank_note = str(item.foodbank_decline_reason)
        if item.foodbank_status == 'accepted_by_foodbank':
            if item.recipient_status == 'received':
                status = 'Received by recipient'
            elif item.recipient_status == 'accepted_by_recipient':
                status = 'Claimed by recipient'
            elif item.recipient_status == 'declined_by_recipient':
                status = 'Rejected by recipient - broadcasted to other recipients'
            else:
                status = 'Accepted by FB - Sent to Recipients'
        elif item.foodbank_status == 'pending_foodbank':
            status = 'Awaiting foodbank'
        elif item.foodbank_status == 'declined_by_foodbank':
            status = 'Declined by foodbank'
        else:
            status = item.get_foodbank_status_display() or '-'

        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=date_str).border = border
        type_cell = ws.cell(row=row_num, column=3, value=type_display)
        type_cell.border = border
        type_cell.alignment = Alignment(wrap_text=True)
        cat_cell = ws.cell(row=row_num, column=4, value=category_label)
        cat_cell.border = border
        ws.cell(row=row_num, column=5, value=details).border = border
        ws.cell(row=row_num, column=6, value=d.foodbank.foodbank_name).border = border
        ws.cell(row=row_num, column=7, value=quantity_value).border = border
        ws.cell(row=row_num, column=8, value=unit_value).border = border
        ws.cell(row=row_num, column=9, value=amount_value).border = border
        ws.cell(row=row_num, column=10, value=delivery).border = border
        ws.cell(row=row_num, column=11, value=location).border = border
        ws.cell(row=row_num, column=12, value=recipient_name).border = border
        ws.cell(row=row_num, column=13, value=donor_note).border = border
        ws.cell(row=row_num, column=14, value=recipient_note).border = border
        ws.cell(row=row_num, column=15, value=recipient_decline_note).border = border
        ws.cell(row=row_num, column=16, value=foodbank_note).border = border
        ws.cell(row=row_num, column=17, value=status).border = border

    column_widths = [6, 18, 12, 12, 30, 22, 10, 10, 12, 12, 24, 18, 22, 22, 24, 22, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"unspecified_donations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@donor_required
def donor_subsidized_export(request, format):
    """Export subsidized donations in PDF or CSV format"""
    subsidized_donations = Donation.objects.filter(
        donor=request.user,
        donation_type='subsidized',
        foodbank_request__isnull=True  # Only donor-initiated (unspecified) subsidized donations
    ).select_related(
        'foodbank', 'accepted_by_recipient', 'unspecified_management', 'declined_by_recipient'
    ).order_by('-donated_at')
    
    # Apply same filters as donor_subsidized_donations_detail (aligned with table columns)
    foodbank_filter = request.GET.get('foodbank', 'all')
    type_filter = (request.GET.get('type') or request.GET.get('category') or 'all').strip().lower()
    if type_filter not in ('all', 'food', 'non_food'):
        type_filter = 'all'
    status_filter = request.GET.get('status', 'all')
    delivery_filter = (request.GET.get('delivery', request.GET.get('delivery_status', 'all')) or 'all').strip().lower()
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    if delivery_filter not in ('all', 'pickup', 'delivery'):
        delivery_filter = 'all'
    date_filter = request.GET.get('date_range', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', request.GET.get('price', 'all'))
    recipient_filter = request.GET.get('recipient', 'all')
    search_query = request.GET.get('search', '').strip()
    
    if foodbank_filter != 'all':
        subsidized_donations = subsidized_donations.filter(foodbank__id=foodbank_filter)
    
    if type_filter != 'all':
        subsidized_donations = subsidized_donations.filter(donation_category=type_filter)
    
    subsidized_donations = _apply_donor_subsidized_status_filter(subsidized_donations, status_filter)
    
    if delivery_filter != 'all':
        if delivery_filter == 'delivery':
            subsidized_donations = subsidized_donations.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            subsidized_donations = subsidized_donations.filter(delivery_method='pickup')
    
    if recipient_filter == 'claimed':
        subsidized_donations = subsidized_donations.filter(accepted_by_recipient__isnull=False)
    elif recipient_filter == 'unclaimed':
        subsidized_donations = subsidized_donations.filter(accepted_by_recipient__isnull=True)
    
    if search_query:
        subsidized_donations = subsidized_donations.filter(
            Q(subsidized_product_type__icontains=search_query) |
            Q(foodbank__foodbank_name__icontains=search_query) |
            Q(foodbank__address__icontains=search_query) |
            Q(message__icontains=search_query) |
            Q(csr_description__icontains=search_query) |
            Q(other_description__icontains=search_query)
        )
    
    # Filter by date range
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    if date_filter == 'custom' and (date_from or date_to):
        if date_from:
            try:
                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
                subsidized_donations = subsidized_donations.filter(donated_at__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                subsidized_donations = subsidized_donations.filter(donated_at__lte=to_date)
            except ValueError:
                pass
    elif date_filter != 'all':
        now = timezone.now()
        if date_filter == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            subsidized_donations = subsidized_donations.filter(donated_at__gte=start_date)
        elif date_filter == 'week':
            start_date = now - timedelta(days=7)
            subsidized_donations = subsidized_donations.filter(donated_at__gte=start_date)
        elif date_filter == 'month':
            start_date = now - timedelta(days=30)
            subsidized_donations = subsidized_donations.filter(donated_at__gte=start_date)
        elif date_filter == '3months':
            start_date = now - timedelta(days=90)
            subsidized_donations = subsidized_donations.filter(donated_at__gte=start_date)
    
    # Filter by quantity
    if quantity_filter == 'small':
        subsidized_donations = subsidized_donations.filter(subsidized_quantity__lte=100)
    elif quantity_filter == 'medium':
        subsidized_donations = subsidized_donations.filter(subsidized_quantity__gt=100, subsidized_quantity__lte=500)
    elif quantity_filter == 'large':
        subsidized_donations = subsidized_donations.filter(subsidized_quantity__gt=500)

    # Filter by amount (subsidized new price)
    if amount_filter == 'small':
        subsidized_donations = subsidized_donations.filter(subsidized_price__lte=100)
    elif amount_filter == 'medium':
        subsidized_donations = subsidized_donations.filter(subsidized_price__gt=100, subsidized_price__lte=500)
    elif amount_filter == 'large':
        subsidized_donations = subsidized_donations.filter(subsidized_price__gt=500)
    
    donations_data = list(subsidized_donations)
    latest_recipient_notes, latest_recipient_declines = _build_subsidized_effective_note_maps(donations_data)
    
    if format.lower() == 'pdf':
        return export_subsidized_pdf(request, donations_data, latest_recipient_notes, latest_recipient_declines)
    elif format.lower() == 'csv':
        return export_subsidized_csv(request, donations_data, latest_recipient_notes, latest_recipient_declines)
    elif format.lower() == 'excel':
        return export_subsidized_excel(request, donations_data, latest_recipient_notes, latest_recipient_declines)
    else:
        messages.error(request, 'Invalid export format.')
        return redirect('donor_subsidized_donations_detail')


def _subsidized_export_status(donation):
    if donation.accepted_by_recipient:
        return 'Accepted by recipient'
    if getattr(donation, 'declined_by_recipient', None) and donation.status == 'accepted':
        return 'Rejected by recipient - broadcasted to other recipients'
    if donation.status == 'pending':
        return 'Awaiting foodbank'
    if donation.status == 'accepted':
        return 'Accepted by fb sent to recipients'
    if donation.status == 'fulfilled':
        return 'Fulfilled & allocated'
    if donation.status == 'declined':
        return 'Declined by food bank'
    return donation.get_status_display() or 'Status update pending'


def _apply_donor_subsidized_status_filter(queryset, status_filter):
    """Apply donor subsidized status filtering using table-visible status values."""
    status_key = (status_filter or '').strip().lower()
    if not status_key or status_key == 'all':
        return queryset

    if status_key in ('pending', 'awaiting_foodbank'):
        return queryset.filter(status='pending')
    if status_key in ('declined', 'declined_by_foodbank'):
        return queryset.filter(status='declined')
    if status_key in ('fulfilled', 'fulfilled_allocated'):
        return queryset.filter(status='fulfilled')
    if status_key == 'accepted_by_recipient':
        return queryset.filter(accepted_by_recipient__isnull=False)
    if status_key == 'rejected_by_recipient_broadcasted':
        return queryset.filter(
            status='accepted',
            declined_by_recipient__isnull=False,
            accepted_by_recipient__isnull=True,
        )
    if status_key in ('accepted', 'accepted_by_foodbank'):
        return queryset.filter(
            status='accepted',
            accepted_by_recipient__isnull=True,
            declined_by_recipient__isnull=True,
        )

    # Unknown status key should not hide records unexpectedly.
    return queryset


def _build_subsidized_effective_note_maps(donations_data):
    """
    Build recipient note and recipient decline note maps aligned with donor subsidized table behavior.
    If a donation has been accepted by a recipient, recipient decline note must be blank.
    """
    donation_ids = [d.id for d in donations_data]
    if not donation_ids:
        return {}, {}

    latest_notes_by_donation = {}
    latest_declines_by_donation = {}
    accepted_notes_by_donation = {}
    accepted_recipient_by_donation = {
        donation.id: donation.accepted_by_recipient_id
        for donation in donations_data
        if donation.accepted_by_recipient_id
    }

    responses = DonationResponse.objects.filter(
        donation_id__in=donation_ids
    ).exclude(
        notes__isnull=True
    ).exclude(
        notes__exact=''
    ).select_related('recipient').order_by('-responded_at')

    for response in responses:
        donation_id = response.donation_id

        if donation_id not in latest_notes_by_donation:
            latest_notes_by_donation[donation_id] = response.notes

        if response.response_type == 'declined' and donation_id not in latest_declines_by_donation:
            latest_declines_by_donation[donation_id] = response.notes

        if (
            response.response_type == 'accepted'
            and accepted_recipient_by_donation.get(donation_id) == response.recipient_id
            and donation_id not in accepted_notes_by_donation
        ):
            accepted_notes_by_donation[donation_id] = response.notes

    effective_recipient_notes = {}
    effective_recipient_declines = {}

    for donation in donations_data:
        unspecified = getattr(donation, 'unspecified_management', None)
        unspecified_recipient_note = (getattr(unspecified, 'recipient_notes', None) or '').strip()
        unspecified_decline_note = (getattr(unspecified, 'recipient_decline_reason', None) or '').strip()

        if donation.accepted_by_recipient_id:
            effective_recipient_notes[donation.id] = (
                accepted_notes_by_donation.get(donation.id)
                or unspecified_recipient_note
                or ''
            )
            effective_recipient_declines[donation.id] = ''
        else:
            effective_recipient_notes[donation.id] = (
                latest_notes_by_donation.get(donation.id)
                or unspecified_recipient_note
                or ''
            )
            effective_recipient_declines[donation.id] = (
                unspecified_decline_note
                or latest_declines_by_donation.get(donation.id)
                or ''
            )

    return effective_recipient_notes, effective_recipient_declines


def export_subsidized_pdf(request, donations_data, latest_recipient_notes=None, latest_recipient_declines=None):
    """Generate PDF report for subsidized donations using shared FoodBankHub branding."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.pagesizes import A3, landscape
    from .report_utils import (
        get_report_styles,
        build_report_header,
        get_branded_table_style,
        build_report_summary,
        build_pdf_document,
        collect_active_filters,
        make_full_width_table,
    )

    if latest_recipient_notes is None:
        latest_recipient_notes = {}
    if latest_recipient_declines is None:
        latest_recipient_declines = {}

    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    styles_dict = get_report_styles()
    filter_keys = [
        ('foodbank', 'Food Bank'), ('type', 'Type'), ('status', 'Status'),
        ('delivery', 'Delivery'), ('date_range', 'Date Range'), ('date_from', 'Date From'), ('date_to', 'Date To'),
        ('price', 'New Price'), ('recipient', 'Recipient'), ('search', 'Search'),
    ]
    active_filters = collect_active_filters(request, filter_keys)

    build_report_header(
        elements := [],
        report_title="Subsidized Donations Report",
        generated_for=donor_name,
        total_records=len(donations_data),
        active_filters=active_filters,
        styles_dict=styles_dict,
    )

    if not donations_data:
        elements.append(Paragraph("No donations found matching the current filters.", styles_dict['normal']))
    else:
        wrap = styles_dict['wrap']
        report_pagesize = landscape(A3)
        data = [[
            'S/No', 'Date', 'Food Bank', 'Category', 'Type', 'Product', 'Quantity',
            'Market Price', 'Subsidy', 'New Price', 'Status', 'Delivery',
            'Recipient', 'Foodbank Decline Notes', 'Recipient Decline Note', 'Donor Note', 'Recipient Note'
        ]]

        for idx, donation in enumerate(donations_data, start=1):
            date_str = donation.donated_at.strftime('%b %d, %Y')
            product = (donation.subsidized_product_type or 'Subsidized Goods')[:50]
            category = (donation.get_donation_category_display() or 'Food')[:12]
            type_label = 'Subsidized'
            quantity = f"{donation.subsidized_quantity} {donation.subsidized_quantity_unit or 'units'}"[:18]
            initial_amount = donation.subsidized_initial_amount
            subsidy_amount = donation.subsidized_subsidy_amount
            initial_str = f"{float(initial_amount):.0f}" if initial_amount is not None else "-"
            subsidy_str = f"{float(subsidy_amount):.0f}" if subsidy_amount is not None else "-"
            price_str = f"{float(donation.subsidized_price):.0f}" if donation.subsidized_price is not None else "-"

            status = _subsidized_export_status(donation)

            delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
            recipient = (donation.accepted_by_recipient.full_name if donation.accepted_by_recipient else 'Not claimed')[:25]
            decline_note = '<i>Not declined</i>'
            if donation.status == 'declined' and getattr(donation, 'decline_message', None):
                decline_note = str(donation.decline_message)[:140]
            recipient_decline_note = (latest_recipient_declines.get(donation.id) or '')[:140]
            donor_note = (donation.message or donation.csr_description or donation.other_description or 'No donor note')[:80]
            recipient_note = (latest_recipient_notes.get(donation.id) or 'No recipient note')[:80]

            data.append([
                str(idx), date_str,
                Paragraph((donation.foodbank.foodbank_name or '')[:30], wrap),
                category, type_label, Paragraph(product, wrap), quantity,
                initial_str, subsidy_str, price_str, Paragraph(status, wrap), delivery,
                Paragraph(recipient, wrap), Paragraph(decline_note, wrap), Paragraph(recipient_decline_note, wrap), Paragraph(donor_note, wrap),
                Paragraph(recipient_note, wrap),
            ])

        # S/No, Date, Food Bank, Category, Type, Product, Quantity, Market Price, Subsidy, New Price, Status, Delivery, Recipient, Foodbank Decline Notes, Recipient Decline Note, Donor Note, Recipient Note
        _col_weights = [
            0.030, 0.054, 0.080, 0.046, 0.054, 0.094, 0.044, 0.064, 0.046, 0.046,
            0.056, 0.046, 0.064, 0.090, 0.090, 0.060, 0.060,
        ]
        table = make_full_width_table(data, repeat_rows=1, col_weights=_col_weights, pagesize=report_pagesize)
        table.setStyle(get_branded_table_style(len(data)))
        elements.append(table)

    build_report_summary(elements, [("Total Subsidized Donations", len(donations_data))], styles_dict=styles_dict)
    return build_pdf_document(elements, "subsidized_donations", donor_name, pagesize=landscape(A3))


def export_subsidized_csv(request, donations_data, latest_recipient_notes=None, latest_recipient_declines=None):
    """Generate CSV report for subsidized donations"""
    import csv
    import datetime

    if latest_recipient_notes is None:
        latest_recipient_notes = {}
    if latest_recipient_declines is None:
        latest_recipient_declines = {}
    
    response = HttpResponse(content_type='text/csv')
    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    response['Content-Disposition'] = f'attachment; filename="subsidized_donations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Donor', donor_name])
    writer.writerow(['Report Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    
    writer.writerow([
        'S/No', 'Date', 'Food Bank', 'Category', 'Type', 'Product', 'Quantity', 'Unit',
        'Market Price (KSH)', 'Subsidy (KSH)', 'New Price (KSH)', 'Status', 'Delivery',
        'Recipient', 'Foodbank Decline', 'Recipient Decline Note', 'Donor Note', 'Recipient Note'
    ])
    
    for idx, donation in enumerate(donations_data, start=1):
        date_str = donation.donated_at.strftime('%Y-%m-%d %H:%M')
        product = donation.subsidized_product_type or 'Subsidized Goods'
        quantity_value = donation.subsidized_quantity or '-'
        unit_value = donation.subsidized_quantity_unit or ''
        price = f"{donation.subsidized_price:,.2f}"
        initial_amount = donation.subsidized_initial_amount
        subsidy_amount = donation.subsidized_subsidy_amount
        
        status = _subsidized_export_status(donation)
        delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
        
        recipient = donation.accepted_by_recipient.full_name if donation.accepted_by_recipient else 'Not claimed'
        decline_note = 'Not declined'
        if donation.status == 'declined' and getattr(donation, 'decline_message', None):
            decline_note = str(donation.decline_message)
        recipient_decline_note = latest_recipient_declines.get(donation.id) or ''
        donor_note = donation.message or donation.csr_description or donation.other_description or 'No donor note'
        recipient_note = latest_recipient_notes.get(donation.id) or 'No recipient note'
        
        writer.writerow([
            idx,
            date_str,
            donation.foodbank.foodbank_name,
            donation.get_donation_category_display() or 'Food',
            'Subsidized',
            product,
            quantity_value,
            unit_value,
            f"{initial_amount:,.2f}" if initial_amount else '-',
            f"{subsidy_amount:,.2f}" if subsidy_amount else '-',
            price,
            status,
            delivery,
            recipient,
            decline_note,
            recipient_decline_note,
            donor_note,
            recipient_note
        ])
    
    return response


def export_subsidized_excel(request, donations_data, latest_recipient_notes=None, latest_recipient_declines=None):
    """Generate Excel report for subsidized donations"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    if latest_recipient_notes is None:
        latest_recipient_notes = {}
    if latest_recipient_declines is None:
        latest_recipient_declines = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Subsidized Donations"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title and info
    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    ws.merge_cells('A1:R1')
    title_cell = ws['A1']
    title_cell.value = f"{donor_name} - Subsidized Donations Report"
    title_cell.font = Font(bold=True, size=14, color="3b82f6")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:R2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total: {len(donations_data)}"
    info_cell.alignment = Alignment(horizontal="center")

    # Headers - match table: S/No, Date, Food Bank, Category, Type, Product, Quantity, Market Price, Subsidy, New Price, ...
    headers = [
        'S/No', 'Date', 'Food Bank', 'Category', 'Type', 'Product', 'Quantity', 'Unit',
        'Market Price (KSH)', 'Subsidy (KSH)', 'New Price (KSH)', 'Status', 'Delivery',
        'Recipient', 'Foodbank Decline', 'Recipient Decline Note', 'Donor Note', 'Recipient Note'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for idx, donation in enumerate(donations_data, 1):
        row_num = idx + 4

        date_str = donation.donated_at.strftime('%Y-%m-%d %H:%M')
        product = donation.subsidized_product_type or 'Subsidized Goods'
        quantity = donation.subsidized_quantity or '-'
        unit = donation.subsidized_quantity_unit or '-'
        price = donation.subsidized_price or 0
        category = donation.get_donation_category_display() or 'Food'
        type_label = 'Subsidized'
        initial_amount = donation.subsidized_initial_amount
        subsidy_amount = donation.subsidized_subsidy_amount

        status = _subsidized_export_status(donation)
        delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'

        recipient = donation.accepted_by_recipient.full_name if donation.accepted_by_recipient else 'Not claimed'
        decline_note = 'Not declined'
        if donation.status == 'declined' and getattr(donation, 'decline_message', None):
            decline_note = str(donation.decline_message)
        recipient_decline_note = latest_recipient_declines.get(donation.id) or ''
        donor_note = donation.message or donation.csr_description or donation.other_description or 'No donor note'
        recipient_note = latest_recipient_notes.get(donation.id) or 'No recipient note'

        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=date_str).border = border
        ws.cell(row=row_num, column=3, value=donation.foodbank.foodbank_name).border = border
        ws.cell(row=row_num, column=4, value=category).border = border
        ws.cell(row=row_num, column=5, value=type_label).border = border
        ws.cell(row=row_num, column=6, value=product).border = border
        ws.cell(row=row_num, column=7, value=quantity).border = border
        ws.cell(row=row_num, column=8, value=unit).border = border
        ws.cell(row=row_num, column=9, value=initial_amount if initial_amount else '-').border = border
        ws.cell(row=row_num, column=10, value=subsidy_amount if subsidy_amount else '-').border = border
        ws.cell(row=row_num, column=11, value=price).border = border
        ws.cell(row=row_num, column=12, value=status).border = border
        ws.cell(row=row_num, column=13, value=delivery).border = border
        ws.cell(row=row_num, column=14, value=recipient).border = border
        ws.cell(row=row_num, column=15, value=decline_note).border = border
        ws.cell(row=row_num, column=16, value=recipient_decline_note).border = border
        ws.cell(row=row_num, column=17, value=donor_note).border = border
        ws.cell(row=row_num, column=18, value=recipient_note).border = border

    # Adjust column widths
    column_widths = [6, 18, 20, 15, 15, 20, 10, 10, 15, 12, 12, 18, 16, 20, 24, 24, 30, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"subsidized_donations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
