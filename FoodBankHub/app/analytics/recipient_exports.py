"""
Recipient Export Views
Contains export functions for recipient donation reports
"""
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
from authentication.models import DonationAllocation
import csv
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required
def recipient_regular_donations_export(request):
    """Export regular donations received by recipient"""
    if request.user.user_type != 'RECIPIENT':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    recipient_profile = request.user.recipient_profile
    
    # Get regular item donations allocated to this recipient
    allocations = DonationAllocation.objects.filter(
        recipient=recipient_profile,
        donation__donation_type='item'
    ).select_related('donation', 'donation__foodbank', 'donation__donor').order_by('-allocated_at')
    
    # Generate Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="regular_donations_received_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Regular Donations'
    
    # Header styling
    header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Date Allocated', 'Item Name', 'Category', 'Quantity', 'Foodbank', 'Donor', 'Status']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for allocation in allocations:
        donor_name = 'Anonymous'
        if allocation.donation.donor:
            if hasattr(allocation.donation.donor, 'donor_profile'):
                donor_name = allocation.donation.donor.donor_profile.full_name or allocation.donation.donor.email
        
        ws.append([
            allocation.allocated_at.strftime('%Y-%m-%d %H:%M'),
            allocation.donation.item_name or 'N/A',
            allocation.donation.get_donation_category_display() if allocation.donation.donation_category else 'N/A',
            allocation.quantity,
            allocation.donation.foodbank.foodbank_name if allocation.donation.foodbank else 'N/A',
            donor_name,
            'Received'
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@login_required
def recipient_subsidized_donations_export(request):
    """Export subsidized donations claimed by recipient"""
    if request.user.user_type != 'RECIPIENT':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    recipient_profile = request.user.recipient_profile
    
    # Get subsidized donations allocated to this recipient
    allocations = DonationAllocation.objects.filter(
        recipient=recipient_profile,
        donation__donation_type='subsidized'
    ).select_related('donation', 'donation__foodbank', 'donation__donor').order_by('-allocated_at')
    
    # Generate Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="subsidized_donations_claimed_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subsidized Donations'
    
    # Header styling
    header_fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Date Claimed', 'Item Name', 'Category', 'Quantity', 'Price Paid (KSH)', 'Foodbank', 'Status']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for allocation in allocations:
        quantity = allocation.quantity or 0
        price_paid = quantity * (allocation.donation.subsidized_price or 0)
        
        ws.append([
            allocation.allocated_at.strftime('%Y-%m-%d %H:%M'),
            allocation.donation.item_name or 'N/A',
            allocation.donation.get_donation_category_display() if allocation.donation.donation_category else 'N/A',
            quantity,
            f"{price_paid:.2f}",
            allocation.donation.foodbank.foodbank_name if allocation.donation.foodbank else 'N/A',
            'Claimed'
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@login_required
def recipient_all_allocations_export(request):
    """Export all allocations for recipient"""
    if request.user.user_type != 'RECIPIENT':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    recipient_profile = request.user.recipient_profile
    
    # Get all allocations for this recipient
    allocations = DonationAllocation.objects.filter(
        recipient=recipient_profile
    ).select_related('donation', 'donation__foodbank', 'donation__donor').order_by('-allocated_at')
    
    # Generate Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="complete_allocation_history_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Allocations'
    
    # Header styling
    header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Date', 'Type', 'Item Name', 'Category', 'Quantity', 'Price (KSH)', 'Foodbank', 'Donor', 'Status']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for allocation in allocations:
        donor_name = 'Anonymous'
        if allocation.donation.donor:
            if hasattr(allocation.donation.donor, 'donor_profile'):
                donor_name = allocation.donation.donor.donor_profile.full_name or allocation.donation.donor.email
        
        donation_type = allocation.donation.get_donation_type_display() if allocation.donation.donation_type else 'N/A'
        quantity = allocation.quantity or 0
        price = ''
        if allocation.donation.donation_type == 'subsidized':
            price = f"{quantity * (allocation.donation.subsidized_price or 0):.2f}"
        
        ws.append([
            allocation.allocated_at.strftime('%Y-%m-%d %H:%M'),
            donation_type,
            allocation.donation.item_name or 'N/A',
            allocation.donation.get_donation_category_display() if allocation.donation.donation_category else 'N/A',
            quantity,
            price,
            allocation.donation.foodbank.foodbank_name if allocation.donation.foodbank else 'N/A',
            donor_name,
            'Received'
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response
