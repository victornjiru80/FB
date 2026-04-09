"""
Environmental & Social Impact Analytics for Custom Admin Panel
Calculates sustainability metrics based on donation data
"""

from django.shortcuts import render, redirect
from django.http import HttpResponse
from .decorators import staff_member_required
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.urls import reverse
from datetime import timedelta
import json

from authentication.models import (
    Donation, FoodBankProfile, DonationAllocation,
    CustomUser, PaymentTransaction, UnspecifiedDonationManagement,
    RequestManagement, DonationResponse
)
from custom_admin.impact_calculations import (
    calculate_food_waste_prevented,
    calculate_co2_saved,
    calculate_non_food_impact,
    get_donation_impact,
)


def calculate_economic_impact(monetary_donations, csr_donations):
    """
    Calculate economic and social impact value
    
    Impact Values (KES per unit):
    - Monetary: 100 KES per unit
    - CSR & Others: 250 KES per unit
    
    Returns: Total economic impact in KES
    """
    monetary_impact = monetary_donations * 100
    csr_impact = csr_donations * 250
    
    return monetary_impact + csr_impact


def estimate_meals_provided(food_kg):
    """
    Estimate meals provided from food donations
    
    Assumption: 1 meal = 0.5 kg of food
    
    Returns: Number of meals
    """
    meals = food_kg / 0.5
    return int(meals)


def estimate_lives_impacted(total_allocations, avg_family_size=4):
    """
    Estimate lives impacted based on allocations
    
    Assumption: Each allocation helps an average family
    
    Returns: Estimated number of people helped
    """
    return total_allocations * avg_family_size


@staff_member_required
def environmental_impact_dashboard(request):
    """
    Wingu Impact Report 2025 - Comprehensive Environmental & Social Impact Dashboard
    Shows sustainability metrics, donation impact, and beneficiary analytics
    """
    # Get time period (default: last 12 months)
    days = int(request.GET.get('days', 365))
    start_date = timezone.now() - timedelta(days=days)
    
    # === COMPREHENSIVE DONATION BREAKDOWN ===
    all_donations = Donation.objects.filter(
        donated_at__gte=start_date,
        status='accepted'
    )
    
    # A. DONATION IMPACT REPORTS - By Category
    donation_breakdown = {
        'free_food': all_donations.filter(
            donation_category='food', 
            donation_mode='free'
        ).aggregate(
            count=Count('id'),
            total_kg=Sum('quantity')
        ),
        'subsidized_food': all_donations.filter(
            donation_category='food', 
            donation_mode='subsidized'
        ).aggregate(
            count=Count('id'),
            total_kg=Sum('subsidized_quantity')
        ),
        'free_non_food': all_donations.filter(
            donation_category='non_food', 
            donation_mode='free'
        ).aggregate(
            count=Count('id'),
            total_units=Sum('quantity')
        ),
        'subsidized_non_food': all_donations.filter(
            donation_category='non_food', 
            donation_mode='subsidized'
        ).aggregate(
            count=Count('id'),
            total_units=Sum('subsidized_quantity')
        ),
        'monetary': all_donations.filter(
            donation_category='monetary'
        ).aggregate(
            count=Count('id'),
            total_amount=Sum('amount')
        ),
        'csr_others': all_donations.filter(
            donation_category__in=['csr', 'other']
        ).aggregate(
            count=Count('id'),
            total_units=Sum('quantity')
        )
    }
    
    # Clean up None values
    for category in donation_breakdown:
        for key, value in donation_breakdown[category].items():
            if value is None:
                donation_breakdown[category][key] = 0
    
    # === FOOD WASTE PREVENTED CALCULATION ===
    free_food_kg = donation_breakdown['free_food']['total_kg']
    subsidized_food_kg = donation_breakdown['subsidized_food']['total_kg']
    
    # Calculate food waste prevented using UNEP factors
    food_waste_prevented = calculate_food_waste_prevented(free_food_kg, subsidized_food_kg)
    
    # === COâ‚‚ SAVED CALCULATION ===
    co2_saved_food = calculate_co2_saved(food_waste_prevented)
    
    # Non-food COâ‚‚ impact
    free_non_food_units = donation_breakdown['free_non_food']['total_units']
    subsidized_non_food_units = donation_breakdown['subsidized_non_food']['total_units']
    non_food_impact = calculate_non_food_impact(free_non_food_units, subsidized_non_food_units)
    
    # Total COâ‚‚ saved
    total_co2_saved = co2_saved_food + non_food_impact['co2_saved_tons']
    
    # === BENEFICIARIES SUPPORTED ===
    # Get unique recipients who received allocations
    unique_recipients = DonationAllocation.objects.filter(
        allocated_at__gte=start_date
    ).values('recipient').distinct().count()
    
    # Estimate total beneficiaries (including family members)
    estimated_beneficiaries = estimate_lives_impacted(unique_recipients, avg_family_size=4)
    
    # Households assisted (direct recipients)
    households_assisted = unique_recipients
    
    # === DISTRIBUTION BY FOODBANK ===
    foodbank_distribution = []
    total_donations_count = all_donations.count()
    
    foodbank_stats = all_donations.values('foodbank__foodbank_name').annotate(
        donation_count=Count('id'),
        total_kg=Sum('quantity'),
        total_amount=Sum('amount')
    ).order_by('-donation_count')
    
    for fb_stat in foodbank_stats:
        if total_donations_count > 0:
            percentage = round((fb_stat['donation_count'] / total_donations_count) * 100, 1)
            foodbank_distribution.append({
                'name': fb_stat['foodbank__foodbank_name'],
                'percentage': percentage,
                'donation_count': fb_stat['donation_count'],
                'total_kg': fb_stat['total_kg'] or 0,
                'total_amount': fb_stat['total_amount'] or 0
            })

    foodbank_impact = []
    foodbank_impact_stats = all_donations.values('foodbank__foodbank_name').annotate(
        donations_count=Count('id'),
        free_food_kg=Sum('quantity', filter=Q(donation_category='food', donation_mode='free')),
        subsidized_food_kg=Sum('subsidized_quantity', filter=Q(donation_category='food', donation_mode='subsidized')),
        free_non_food_units=Sum('quantity', filter=Q(donation_category='non_food', donation_mode='free')),
        subsidized_non_food_units=Sum('subsidized_quantity', filter=Q(donation_category='non_food', donation_mode='subsidized')),
    )

    for fb_stat in foodbank_impact_stats:
        name = fb_stat.get('foodbank__foodbank_name') or 'Unknown'
        free_food_fb = fb_stat.get('free_food_kg') or 0
        subsidized_food_fb = fb_stat.get('subsidized_food_kg') or 0
        free_non_food_fb = fb_stat.get('free_non_food_units') or 0
        subsidized_non_food_fb = fb_stat.get('subsidized_non_food_units') or 0

        food_waste_prevented_fb = calculate_food_waste_prevented(free_food_fb, subsidized_food_fb)
        co2_saved_food_fb = calculate_co2_saved(food_waste_prevented_fb)
        non_food_impact_fb = calculate_non_food_impact(free_non_food_fb, subsidized_non_food_fb)

        waste_prevented_total_fb = float(food_waste_prevented_fb) + float(non_food_impact_fb['waste_prevented_kg'])
        co2_saved_total_fb = float(co2_saved_food_fb) + float(non_food_impact_fb['co2_saved_tons'])

        foodbank_impact.append({
            'name': name,
            'waste_prevented': waste_prevented_total_fb,
            'co2_saved': co2_saved_total_fb,
            'donations_count': fb_stat.get('donations_count') or 0,
        })

    foodbank_impact.sort(key=lambda row: row.get('waste_prevented', 0), reverse=True)
    max_waste_prevented = max((row.get('waste_prevented', 0) for row in foodbank_impact), default=0)
    for row in foodbank_impact:
        row['impact_score_percent'] = (row.get('waste_prevented', 0) / max_waste_prevented * 100) if max_waste_prevented else 0
    
    # === MEALS PROVIDED ESTIMATION ===
    meals_provided = estimate_meals_provided(food_waste_prevented)
    
    # === ECONOMIC IMPACT ===
    monetary_donations_count = donation_breakdown['monetary']['count']
    csr_donations_count = donation_breakdown['csr_others']['count']
    total_monetary_value = donation_breakdown['monetary']['total_amount']
    
    economic_impact = calculate_economic_impact(monetary_donations_count, csr_donations_count)
    
    # === DONOR & CSR ANALYTICS ===
    # Contributions by donor type
    donor_type_breakdown = all_donations.values('donor__user_type').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    )
    
    # CSR subcategory breakdown
    csr_breakdown = all_donations.filter(
        donation_category='csr'
    ).values('csr_subcategory').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # === OPERATIONAL METRICS ===
    # Matching success rate (donations linked to requests vs general donations)
    request_linked_donations = all_donations.filter(foodbank_request__isnull=False).count()
    general_donations = all_donations.filter(foodbank_request__isnull=True).count()
    matching_success_rate = (request_linked_donations / total_donations_count * 100) if total_donations_count > 0 else 0
    
    # Distribution efficiency (accepted vs total donations)
    total_all_donations = Donation.objects.filter(donated_at__gte=start_date).count()
    distribution_efficiency = (total_donations_count / total_all_donations * 100) if total_all_donations > 0 else 0
    
    # === ENVIRONMENTAL EQUIVALENTS ===
    # Trees planted equivalent (1 tree absorbs ~21 kg COâ‚‚ per year)
    trees_equivalent = int(total_co2_saved * 1000 / 21)
    
    # Cars off road equivalent (1 car = ~4.6 tons COâ‚‚ per year)
    cars_equivalent = round(total_co2_saved / 4.6, 1)
    
    # Homes powered equivalent (1 home = ~7.5 tons COâ‚‚ per year)
    homes_equivalent = round(total_co2_saved / 7.5, 1)
    
    # === SDG ALIGNMENT ===
    sdg_metrics = {
        'sdg_2': meals_provided,  # Zero Hunger
        'sdg_12': food_waste_prevented,  # Responsible Consumption
        'sdg_13': total_co2_saved,  # Climate Action
        'sdg_17': len(foodbank_distribution)  # Partnerships
    }

    donation_chart_labels = ['Free Food', 'Subsidized Food', 'Monetary', 'CSR & Others']
    donation_chart_data = [
        donation_breakdown['free_food']['count'],
        donation_breakdown['subsidized_food']['count'],
        donation_breakdown['monetary']['count'],
        donation_breakdown['csr_others']['count'],
    ]

    environmental_chart_labels = ['Food Waste Prevented', 'Non-Food Waste Prevented', 'COâ‚‚ Saved (tons)']
    environmental_chart_data = [
        float(food_waste_prevented),
        float(non_food_impact['waste_prevented_kg']),
        float(total_co2_saved),
    ]

    sdg_chart_labels = ['SDG 2: Zero Hunger', 'SDG 12: Responsible Consumption', 'SDG 13: Climate Action', 'SDG 17: Partnerships']
    sdg_chart_data = [
        sdg_metrics['sdg_2'],
        float(sdg_metrics['sdg_12']),
        float(sdg_metrics['sdg_13']),
        sdg_metrics['sdg_17'],
    ]

    top_foodbanks = foodbank_distribution[:8]
    def _truncate_label(value, max_len=20):
        if not value:
            return ''
        value = str(value)
        if len(value) <= max_len:
            return value
        return value[: max_len - 1] + 'â€¦'

    foodbank_chart_labels = [_truncate_label(fb.get('name', ''), 20) for fb in top_foodbanks]
    foodbank_chart_data = [fb.get('percentage', 0) for fb in top_foodbanks]

    climate_champion_progress = min((float(total_co2_saved) / 10.0) * 100.0, 100.0) if total_co2_saved else 0
    zero_waste_progress = min((float(food_waste_prevented) / 1000.0) * 100.0, 100.0) if food_waste_prevented else 0
    community_builder_progress = min((float(estimated_beneficiaries) / 500.0) * 100.0, 100.0) if estimated_beneficiaries else 0
    partnership_master_progress = min((float(len(foodbank_distribution)) / 10.0) * 100.0, 100.0) if foodbank_distribution else 0

    # === POSTED DONATIONS ENVIRONMENTAL TABLES ===
    posted_unspecified_page_param = 'unspecified_page'
    posted_subsidized_page_param = 'subsidized_page'
    posted_specified_page_param = 'specified_page'

    def _table_query_string(page_param):
        query = request.GET.copy()
        query.pop(page_param, None)
        return query.urlencode()

    # Posted Unspecified (completed = received by recipient)
    posted_unspecified_qs = UnspecifiedDonationManagement.objects.select_related(
        'donation__donor',
        'donation__donor__donor_profile',
        'donation__foodbank',
        'accepted_by_recipient',
    ).filter(
        recipient_status='received',
        created_at__gte=start_date,
    ).order_by('-created_at')
    posted_unspecified_total = posted_unspecified_qs.count()
    posted_unspecified_page_obj = Paginator(
        posted_unspecified_qs, 10
    ).get_page(request.GET.get(posted_unspecified_page_param))
    for item in posted_unspecified_page_obj.object_list:
        impact = get_donation_impact(item.donation)
        item.food_waste_prevented_kg = impact['waste_prevented_kg']
        item.co2_saved_tons = impact['co2_saved_tons']

    # Posted Subsidized (completed = accepted + delivered)
    posted_subsidized_qs = Donation.objects.filter(
        donation_type='subsidized',
        status='accepted',
        delivery_status='delivered',
        donated_at__gte=start_date,
    ).select_related(
        'donor',
        'donor__donor_profile',
        'foodbank',
        'accepted_by_recipient',
    ).prefetch_related(
        'subsidized_responded_by'
    ).order_by('-donated_at')
    posted_subsidized_total = posted_subsidized_qs.count()
    posted_subsidized_page_obj = Paginator(
        posted_subsidized_qs, 10
    ).get_page(request.GET.get(posted_subsidized_page_param))
    posted_subsidized_rows = list(posted_subsidized_page_obj.object_list)
    if posted_subsidized_rows:
        donation_ids = [donation.id for donation in posted_subsidized_rows]
        recipient_notes = (
            DonationResponse.objects.filter(donation_id__in=donation_ids)
            .exclude(notes__isnull=True)
            .exclude(notes__exact='')
            .order_by('-responded_at')
        )
        latest_notes = {}
        for response in recipient_notes:
            if response.donation_id not in latest_notes:
                latest_notes[response.donation_id] = response.notes

        for donation in posted_subsidized_rows:
            donation.latest_recipient_note = latest_notes.get(donation.id)
            impact = get_donation_impact(donation)
            donation.food_waste_prevented_kg = impact['waste_prevented_kg']
            donation.co2_saved_tons = impact['co2_saved_tons']

    # Posted Specified (completed = fulfilled/acknowledged and donation-backed accepted+delivered)
    from .views_donations import _resolve_specified_links
    posted_specified_base_qs = RequestManagement.objects.filter(
        status__in=['fulfilled', 'acknowledged'],
        time_of_request__gte=start_date,
    ).select_related(
        'recipient',
        'foodbank',
        'assigned_foodbank',
        'donation',
        'donation__donor__donor_profile',
        'foodbank_request',
    ).prefetch_related(
        'foodbank_request_created__donations',
        'foodbank_request_created__donations__donor__donor_profile',
        'donor_requests__donations',
        'donor_requests__donations__donor__donor_profile',
    ).order_by('-time_of_request')
    posted_specified_rows = []
    for req in posted_specified_base_qs:
        fb_request, linked_donation = _resolve_specified_links(req)
        if not linked_donation:
            continue
        if linked_donation.status != 'accepted' or linked_donation.delivery_status != 'delivered':
            continue
        req.linked_foodbank_request = fb_request
        req.linked_donation = linked_donation
        impact = get_donation_impact(linked_donation)
        req.food_waste_prevented_kg = impact['waste_prevented_kg']
        req.co2_saved_tons = impact['co2_saved_tons']
        posted_specified_rows.append(req)

    posted_specified_total = len(posted_specified_rows)
    posted_specified_page_obj = Paginator(
        posted_specified_rows, 15
    ).get_page(request.GET.get(posted_specified_page_param))

    # Top summary on this page should reflect only completed donations represented
    # by the three posted tables above.
    completed_donation_ids = set(
        posted_unspecified_qs.filter(donation__isnull=False).values_list('donation_id', flat=True)
    )
    completed_donation_ids.update(posted_subsidized_qs.values_list('id', flat=True))
    completed_donation_ids.update(
        req.linked_donation.id for req in posted_specified_rows if getattr(req, 'linked_donation', None)
    )

    completed_food_waste_prevented = 0.0
    completed_total_co2_saved = 0.0
    if completed_donation_ids:
        completed_donations = Donation.objects.filter(id__in=completed_donation_ids)
        for donation in completed_donations:
            impact = get_donation_impact(donation)
            completed_food_waste_prevented += float(impact.get('waste_prevented_kg') or 0)
            completed_total_co2_saved += float(impact.get('co2_saved_tons') or 0)
    
    context = {
        'title': 'Wingu Impact Report 2025',
        'period_days': days,
        
        # === A. DONATION IMPACT REPORTS ===
        'donation_breakdown': donation_breakdown,
        'total_donations': total_donations_count,
        'foodbank_distribution': foodbank_distribution,
        'foodbank_impact': foodbank_impact,
        
        # Environmental Impact
        'food_waste_prevented': food_waste_prevented,
        'co2_saved_food': co2_saved_food,
        'non_food_waste_prevented': non_food_impact['waste_prevented_kg'],
        'co2_saved_non_food': non_food_impact['co2_saved_tons'],
        'total_co2_saved': total_co2_saved,
        'total_waste_prevented': food_waste_prevented + non_food_impact['waste_prevented_kg'],
        
        # === B. BENEFICIARIES SUPPORTED ===
        'estimated_beneficiaries': estimated_beneficiaries,
        'households_assisted': households_assisted,
        'unique_recipients': unique_recipients,
        'meals_provided': meals_provided,
        
        # === C. DONOR & CSR REPORTS ===
        'donor_type_breakdown': donor_type_breakdown,
        'csr_breakdown': csr_breakdown,
        'total_monetary_value': total_monetary_value,
        
        # === D. OPERATIONAL REPORTS ===
        'matching_success_rate': round(matching_success_rate, 1),
        'distribution_efficiency': round(distribution_efficiency, 1),
        'request_linked_donations': request_linked_donations,
        'general_donations': general_donations,
        
        # === E. ENVIRONMENTAL & SUSTAINABILITY ===
        'trees_equivalent': trees_equivalent,
        'cars_equivalent': cars_equivalent,
        'homes_equivalent': homes_equivalent,
        'sdg_metrics': sdg_metrics,

        'donation_chart_labels': json.dumps(donation_chart_labels),
        'donation_chart_data': json.dumps(donation_chart_data),
        'environmental_chart_labels': json.dumps(environmental_chart_labels),
        'environmental_chart_data': json.dumps(environmental_chart_data),
        'sdg_chart_labels': json.dumps(sdg_chart_labels),
        'sdg_chart_data': json.dumps(sdg_chart_data),
        'foodbank_chart_labels': json.dumps(foodbank_chart_labels),
        'foodbank_chart_data': json.dumps(foodbank_chart_data),

        'climate_champion_progress': climate_champion_progress,
        'zero_waste_progress': zero_waste_progress,
        'community_builder_progress': community_builder_progress,
        'partnership_master_progress': partnership_master_progress,
        
        # === F. FINANCIAL REPORTS ===
        'economic_impact': economic_impact,
        'monetary_donations_count': monetary_donations_count,
        'csr_donations_count': csr_donations_count,
        
        # Raw metrics for detailed breakdown
        'free_food_kg': free_food_kg,
        'subsidized_food_kg': subsidized_food_kg,
        'free_non_food_units': free_non_food_units,
        'subsidized_non_food_units': subsidized_non_food_units,
        
        # Legacy compatibility
        'unique_foodbanks': len(foodbank_distribution),

        # Posted tables with per-donation impact
        'posted_unspecified_page_obj': posted_unspecified_page_obj,
        'posted_unspecified_total': posted_unspecified_total,
        'posted_unspecified_page_param': posted_unspecified_page_param,
        'posted_unspecified_pagination_query_string': _table_query_string(posted_unspecified_page_param),

        'posted_subsidized_page_obj': posted_subsidized_page_obj,
        'posted_subsidized_total': posted_subsidized_total,
        'posted_subsidized_page_param': posted_subsidized_page_param,
        'posted_subsidized_pagination_query_string': _table_query_string(posted_subsidized_page_param),

        'posted_specified_page_obj': posted_specified_page_obj,
        'posted_specified_total': posted_specified_total,
        'posted_specified_page_param': posted_specified_page_param,
        'posted_specified_pagination_query_string': _table_query_string(posted_specified_page_param),

        # Top-of-page summary sourced from the completed posted tables only
        'completed_donations_total': len(completed_donation_ids),
        'completed_food_waste_prevented': completed_food_waste_prevented,
        'completed_total_co2_saved': completed_total_co2_saved,
    }
    
    return render(request, 'custom_admin/environmental_impact.html', context)


@staff_member_required
def export_impact_report_pdf(request):
    """
    Export Wingu Impact Report 2025 as PDF
    """
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus.frames import Frame
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
    from io import BytesIO
    import datetime
    
    # Get the same data as the dashboard
    days = int(request.GET.get('days', 365))
    start_date = timezone.now() - timedelta(days=days)
    
    # Get all the metrics (reusing the same logic from environmental_impact_dashboard)
    all_donations = Donation.objects.filter(
        donated_at__gte=start_date,
        status='accepted'
    )
    
    # Donation breakdown
    donation_breakdown = {
        'free_food': all_donations.filter(
            donation_category='food', 
            donation_mode='free'
        ).aggregate(
            count=Count('id'),
            total_kg=Sum('quantity')
        ),
        'subsidized_food': all_donations.filter(
            donation_category='food', 
            donation_mode='subsidized'
        ).aggregate(
            count=Count('id'),
            total_kg=Sum('subsidized_quantity')
        ),
        'monetary': all_donations.filter(
            donation_category='monetary'
        ).aggregate(
            count=Count('id'),
            total_amount=Sum('amount')
        ),
        'csr_others': all_donations.filter(
            donation_category__in=['csr', 'other']
        ).aggregate(
            count=Count('id'),
            total_units=Sum('quantity')
        )
    }
    
    # Clean up None values
    for category in donation_breakdown:
        for key, value in donation_breakdown[category].items():
            if value is None:
                donation_breakdown[category][key] = 0
    
    # Calculate impact metrics
    free_food_kg = donation_breakdown['free_food']['total_kg']
    subsidized_food_kg = donation_breakdown['subsidized_food']['total_kg']
    food_waste_prevented = calculate_food_waste_prevented(free_food_kg, subsidized_food_kg)
    total_co2_saved = calculate_co2_saved(food_waste_prevented)
    meals_provided = estimate_meals_provided(food_waste_prevented)
    
    # Beneficiaries
    unique_recipients = DonationAllocation.objects.filter(
        allocated_at__gte=start_date
    ).values('recipient').distinct().count()
    estimated_beneficiaries = estimate_lives_impacted(unique_recipients, avg_family_size=4)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center alignment
        textColor=colors.HexColor('#10b981')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor('#059669')
    )
    
    normal_style = styles['Normal']
    
    # Title
    title = Paragraph("Wingu Impact Report 2025", title_style)
    elements.append(title)
    
    # Report period
    period_text = f"Report Period: Last {days} days ({start_date.strftime('%B %d, %Y')} - {timezone.now().strftime('%B %d, %Y')})"
    elements.append(Paragraph(period_text, normal_style))
    elements.append(Spacer(1, 20))
    
    # Executive Summary
    elements.append(Paragraph("Executive Summary", heading_style))
    summary_data = [
        ['Metric', 'Value', 'Impact'],
        ['Total COâ‚‚ Saved', f'{total_co2_saved} tons', 'Climate Action'],
        ['Food Waste Prevented', f'{food_waste_prevented:,.0f} kg', 'Sustainability'],
        ['Meals Provided', f'{meals_provided:,.0f}', 'Zero Hunger'],
        ['Beneficiaries Supported', f'{estimated_beneficiaries:,.0f}', 'Social Impact'],
        ['Total Donations', f'{all_donations.count():,.0f}', 'Community Engagement']
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # A. Donation Impact Reports
    elements.append(Paragraph("A. Donation Impact Reports", heading_style))
    donation_data = [
        ['Category', 'Count', 'Quantity/Amount', 'Impact'],
        ['Free Food', f"{donation_breakdown['free_food']['count']}", f"{free_food_kg:,.0f} kg", "Direct waste prevention"],
        ['Subsidized Food', f"{donation_breakdown['subsidized_food']['count']}", f"{subsidized_food_kg:,.0f} kg", "Affordable nutrition"],
        ['Monetary', f"{donation_breakdown['monetary']['count']}", f"KES {donation_breakdown['monetary']['total_amount']:,.0f}", "Operational support"],
        ['CSR & Others', f"{donation_breakdown['csr_others']['count']}", f"{donation_breakdown['csr_others']['total_units']:,.0f} units", "Community programs"]
    ]
    
    donation_table = Table(donation_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 2*inch])
    donation_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(donation_table)
    elements.append(Spacer(1, 20))
    
    # Environmental Impact
    elements.append(Paragraph("Environmental & Sustainability Impact", heading_style))
    trees_equivalent = int(total_co2_saved * 1000 / 21)
    cars_equivalent = round(total_co2_saved / 4.6, 1)
    
    env_data = [
        ['Environmental Metric', 'Value', 'Equivalent'],
        ['Food Waste Prevented', f'{food_waste_prevented:,.0f} kg', 'Reduced landfill burden'],
        ['COâ‚‚ Emissions Saved', f'{total_co2_saved} tons', f'{trees_equivalent:,} trees planted'],
        ['Carbon Footprint Reduction', f'{total_co2_saved} tons COâ‚‚', f'{cars_equivalent} cars off road/year'],
    ]
    
    env_table = Table(env_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
    env_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(env_table)
    elements.append(Spacer(1, 20))
    
    # SDG Alignment
    elements.append(Paragraph("SDG Alignment", heading_style))
    sdg_text = f"""
    <b>SDG 2 (Zero Hunger):</b> {meals_provided:,.0f} meals provided<br/>
    <b>SDG 12 (Responsible Consumption):</b> {food_waste_prevented:,.0f} kg waste prevented<br/>
    <b>SDG 13 (Climate Action):</b> {total_co2_saved} tons COâ‚‚ saved<br/>
    <b>SDG 17 (Partnerships):</b> Multi-stakeholder collaboration
    """
    elements.append(Paragraph(sdg_text, normal_style))
    elements.append(Spacer(1, 20))
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_text = f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')} | Wingu Impact Report 2025"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        textColor=colors.grey
    )
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Wingu_Impact_Report_{days}days_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


@staff_member_required
def environmental_impact_reports(request):
    """
    Hidden endpoint: redirect the legacy environmental reports page to the
    main environmental impact dashboard.
    """
    target = reverse('custom_admin:environmental_impact')
    query = request.GET.urlencode()
    if query:
        return redirect(f'{target}?{query}')
    return redirect(target)


def _environmental_days_and_start(request):
    """Parse dashboard days filter and return (days, start_date)."""
    try:
        days = int(request.GET.get('days', 365))
    except (TypeError, ValueError):
        days = 365
    return days, timezone.now() - timedelta(days=days)


def _environmental_posted_unspecified_queryset(start_date):
    return (
        UnspecifiedDonationManagement.objects.select_related(
            'donation__donor',
            'donation__donor__donor_profile',
            'donation__foodbank',
            'accepted_by_recipient',
        )
        .filter(recipient_status='received', created_at__gte=start_date)
        .order_by('-created_at')
    )


def _environmental_posted_subsidized_queryset(start_date):
    return (
        Donation.objects.filter(
            donation_type='subsidized',
            status='accepted',
            delivery_status='delivered',
            donated_at__gte=start_date,
        )
        .select_related(
            'donor',
            'donor__donor_profile',
            'foodbank',
            'accepted_by_recipient',
        )
        .prefetch_related('subsidized_responded_by')
        .order_by('-donated_at')
    )


def _environmental_posted_specified_rows(start_date):
    from .views_donations import _resolve_specified_links

    base_qs = (
        RequestManagement.objects.filter(
            status__in=['fulfilled', 'acknowledged'],
            time_of_request__gte=start_date,
        )
        .select_related(
            'recipient',
            'foodbank',
            'assigned_foodbank',
            'donation',
            'donation__donor__donor_profile',
            'foodbank_request',
        )
        .prefetch_related(
            'foodbank_request_created__donations',
            'foodbank_request_created__donations__donor__donor_profile',
            'donor_requests__donations',
            'donor_requests__donations__donor__donor_profile',
        )
        .order_by('-time_of_request')
    )

    rows = []
    for req in base_qs:
        fb_request, linked_donation = _resolve_specified_links(req)
        if not linked_donation:
            continue
        if linked_donation.status != 'accepted' or linked_donation.delivery_status != 'delivered':
            continue
        req.linked_foodbank_request = fb_request
        req.linked_donation = linked_donation
        rows.append(req)
    return rows


def _donor_display_name(donor):
    if not donor:
        return 'Unknown Donor'
    profile = getattr(donor, 'donor_profile', None)
    if profile and getattr(profile, 'organization_name', None):
        return profile.organization_name
    if profile and getattr(profile, 'full_name', None):
        return profile.full_name
    return donor.email or 'Unknown Donor'


def _recipient_display_name(recipient_profile):
    if not recipient_profile:
        return 'Unclaimed'
    full_name = getattr(recipient_profile, 'full_name', None)
    if full_name:
        return full_name
    user = getattr(recipient_profile, 'user', None)
    if user and getattr(user, 'email', None):
        return user.email
    return 'Claimed'


def _unspecified_type_label(donation):
    if donation.donation_type == 'item':
        if donation.donation_mode == 'free':
            return 'Free Goods'
        if donation.donation_mode == 'subsidized':
            return 'Subsidized'
        return donation.get_donation_mode_display()
    return donation.get_donation_type_display()


def _unspecified_description_label(donation):
    if donation.donation_type == 'item':
        return donation.item_name or 'General donation'
    if donation.donation_type == 'subsidized':
        return donation.subsidized_product_type or 'Subsidized goods'
    if donation.donation_type == 'csr':
        return donation.csr_description or 'CSR donation'
    if donation.donation_type == 'other':
        return donation.other_description or 'Other donation'
    if donation.donation_type == 'money':
        return donation.message or 'Monetary donation'
    return 'No description'


def _unspecified_quantity_label(donation):
    if donation.donation_type == 'item':
        return f"{donation.quantity or 0} {donation.quantity_unit or 'units'}"
    if donation.donation_type == 'money':
        return f"KES {float(donation.amount or 0):,.0f}"
    if donation.donation_type in {'csr', 'other'}:
        qty = f"{donation.quantity} {donation.quantity_unit or 'units'}" if donation.quantity else ''
        amt = f"KES {float(donation.amount):,.0f}" if donation.amount else ''
        if qty and amt:
            return f"{qty} | {amt}"
        return qty or amt or '-'
    return '-'


def _build_environmental_posted_export_payload(request, table_type):
    days, start_date = _environmental_days_and_start(request)

    if table_type == 'unspecified':
        headers = [
            'S/No', 'Date', 'Donor', 'Type', 'Category', 'Description', 'Food Bank',
            'Qty/Amount', 'Delivery', 'Location', 'Recipient', 'Notes', 'Status',
            'Food Waste Prevented (kg)', 'CO2 Saved (tons)',
        ]
        rows = []
        for idx, item in enumerate(_environmental_posted_unspecified_queryset(start_date), start=1):
            donation = item.donation
            impact = get_donation_impact(donation)
            notes = []
            if donation.message:
                notes.append(f"Donor: {donation.message}")
            if item.recipient_notes:
                notes.append(f"Recipient: {item.recipient_notes}")
            delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
            status_label = (item.recipient_status or 'received').replace('_', ' ').title()
            rows.append([
                idx,
                item.created_at.strftime('%Y-%m-%d %H:%M'),
                _donor_display_name(donation.donor),
                _unspecified_type_label(donation),
                donation.get_donation_category_display(),
                _unspecified_description_label(donation),
                donation.foodbank.foodbank_name if donation.foodbank else '-',
                _unspecified_quantity_label(donation),
                delivery,
                donation.foodbank.address if donation.foodbank and donation.foodbank.address else 'Location not provided',
                _recipient_display_name(item.accepted_by_recipient),
                ' | '.join(notes) if notes else 'No notes',
                status_label,
                impact['waste_prevented_kg'],
                impact['co2_saved_tons'],
            ])
        return {
            'filename_prefix': 'posted_unspecified_environmental',
            'title': f'Posted Unspecified Donations Environmental Impact (Last {days} days)',
            'headers': headers,
            'rows': rows,
        }

    if table_type == 'subsidized':
        headers = [
            'S/No', 'Date', 'Donor', 'Food Bank', 'Location', 'Type', 'Category', 'Product',
            'Quantity', 'Market Price', 'Subsidy', 'New Price', 'Status', 'Delivery',
            'Recipient', 'Notes', 'Decline Note', 'Food Waste Prevented (kg)', 'CO2 Saved (tons)',
        ]
        donations = list(_environmental_posted_subsidized_queryset(start_date))
        latest_notes = {}
        if donations:
            donation_ids = [donation.id for donation in donations]
            recipient_notes = (
                DonationResponse.objects.filter(donation_id__in=donation_ids)
                .exclude(notes__isnull=True)
                .exclude(notes__exact='')
                .order_by('-responded_at')
            )
            for response in recipient_notes:
                if response.donation_id not in latest_notes:
                    latest_notes[response.donation_id] = response.notes

        rows = []
        for idx, donation in enumerate(donations, start=1):
            impact = get_donation_impact(donation)
            notes = []
            if donation.message:
                notes.append(f"Donor: {donation.message}")
            if latest_notes.get(donation.id):
                notes.append(f"Recipient: {latest_notes[donation.id]}")
            delivery_label = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'
            rows.append([
                idx,
                donation.donated_at.strftime('%Y-%m-%d %H:%M') if donation.donated_at else '-',
                _donor_display_name(donation.donor),
                donation.foodbank.foodbank_name if donation.foodbank else '-',
                donation.foodbank.address if donation.foodbank and donation.foodbank.address else 'Not provided',
                donation.get_donation_category_display() if hasattr(donation, 'get_donation_category_display') else 'Food',
                'Subsidized',
                donation.subsidized_product_type or 'Subsidized Goods',
                f"{donation.subsidized_quantity or 0} {donation.subsidized_quantity_unit or 'units'}",
                float(donation.subsidized_initial_amount or 0),
                float(donation.subsidized_subsidy_amount or 0),
                float(donation.subsidized_price or 0),
                (donation.status or '').replace('_', ' ').title() or 'Accepted',
                (donation.delivery_status or '').replace('_', ' ').title() or delivery_label,
                _recipient_display_name(donation.accepted_by_recipient),
                ' | '.join(notes) if notes else 'No notes',
                donation.decline_message or 'Not declined',
                impact['waste_prevented_kg'],
                impact['co2_saved_tons'],
            ])
        return {
            'filename_prefix': 'posted_subsidized_environmental',
            'title': f'Posted Subsidized Donations Environmental Impact (Last {days} days)',
            'headers': headers,
            'rows': rows,
        }

    if table_type == 'specified':
        headers = [
            'S/No', 'Category', 'Type', 'Description', 'Donor', 'Recipient', 'Foodbank',
            'Quantity', 'Delivery', 'Location', 'Status', 'Requested', 'Notes',
            'Food Waste Prevented (kg)', 'CO2 Saved (tons)',
        ]
        rows = []
        for idx, req in enumerate(_environmental_posted_specified_rows(start_date), start=1):
            donation = req.linked_donation
            impact = get_donation_impact(donation)
            if getattr(req, 'foodbank_name', None):
                foodbank_name = req.foodbank_name
            elif req.foodbank:
                foodbank_name = req.foodbank.foodbank_name
            elif req.assigned_foodbank:
                foodbank_name = req.assigned_foodbank.foodbank_name
            else:
                foodbank_name = '-'
            rows.append([
                idx,
                donation.get_donation_category_display() if hasattr(donation, 'get_donation_category_display') else (req.request_category or req.request_type),
                donation.get_donation_type_display() if hasattr(donation, 'get_donation_type_display') else 'Request',
                req.description or '-',
                _donor_display_name(donation.donor),
                _recipient_display_name(req.recipient),
                foodbank_name,
                f"{req.quantity or 0} {req.get_unit_display() if hasattr(req, 'get_unit_display') else ''}".strip(),
                req.get_delivery_method_display() if hasattr(req, 'get_delivery_method_display') else (req.delivery_method or '-'),
                req.location or '-',
                (req.status or '').replace('_', ' ').title(),
                req.time_of_request.strftime('%Y-%m-%d %H:%M') if req.time_of_request else '-',
                req.additional_notes or '-',
                impact['waste_prevented_kg'],
                impact['co2_saved_tons'],
            ])
        return {
            'filename_prefix': 'posted_specified_environmental',
            'title': f'Posted Specified Donations Environmental Impact (Last {days} days)',
            'headers': headers,
            'rows': rows,
        }

    raise ValueError(f'Unsupported table_type: {table_type}')


def _export_environmental_posted_excel(payload):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{payload["filename_prefix"]}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Posted Impact'

    ws['A1'] = payload['title']
    ws['A1'].font = Font(size=14, bold=True, color='065F46')
    ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=9, color='6B7280')

    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)

    header_row = 4
    for col, header in enumerate(payload['headers'], 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    data_start = header_row + 1
    for row_idx, row in enumerate(payload['rows'], start=data_start):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for idx, header in enumerate(payload['headers'], start=1):
        ws.column_dimensions[chr(64 + idx)].width = min(max(len(str(header)) + 4, 12), 36) if idx <= 26 else 20

    wb.save(response)
    return response


def _export_environmental_posted_pdf(payload):
    import re
    from xml.sax.saxutils import escape
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="{payload["filename_prefix"]}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    )

    # Prefer A4 for portability; fall back to A3 for very wide tables.
    headers = payload['headers']
    page_size = landscape(A4) if len(headers) <= 16 else landscape(A3)
    doc = SimpleDocTemplate(response, pagesize=page_size, leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PostedImpactPdfTitle',
        parent=styles['Heading2'],
        fontSize=12,
        leading=14,
        textColor=colors.HexColor('#065F46'),
        spaceAfter=6,
    )
    header_style = ParagraphStyle(
        'PostedImpactPdfHeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6.5,
        leading=7.5,
        textColor=colors.whitesmoke,
        alignment=1,
        wordWrap='CJK',
    )
    cell_style = ParagraphStyle(
        'PostedImpactPdfCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6,
        leading=7,
        textColor=colors.HexColor('#111827'),
        wordWrap='CJK',
    )
    number_style = ParagraphStyle(
        'PostedImpactPdfNumberCell',
        parent=cell_style,
        alignment=2,
    )
    elements = [Paragraph(escape(payload['title']), title_style), Spacer(1, 8)]

    def _header_weight(header):
        header_lower = header.lower()
        if header_lower in {'s/no', 'id'}:
            return 0.6
        if 'date' in header_lower or 'requested' in header_lower:
            return 1.1
        if header_lower in {'status', 'type', 'category', 'delivery'}:
            return 0.95
        if 'co2' in header_lower or 'food waste' in header_lower:
            return 1.0
        if any(token in header_lower for token in ['price', 'amount', 'subsidy', 'quantity', 'qty']):
            return 1.0
        if any(token in header_lower for token in ['description', 'notes', 'location', 'product']):
            return 1.8
        if any(token in header_lower for token in ['donor', 'recipient', 'food bank', 'foodbank']):
            return 1.35
        return 1.0

    def _text_limit(header):
        header_lower = header.lower()
        if any(token in header_lower for token in ['description', 'notes', 'location']):
            return 60
        if any(token in header_lower for token in ['product', 'donor', 'recipient', 'food bank', 'foodbank']):
            return 40
        return 24

    def _as_paragraph(value, style, max_len):
        if value is None:
            text = '-'
        elif isinstance(value, float):
            text = f"{value:.3f}".rstrip('0').rstrip('.')
        else:
            text = str(value)
        text = re.sub(r'\s+', ' ', text).strip() or '-'
        if len(text) > max_len:
            text = text[: max_len - 1] + '…'
        return Paragraph(escape(text), style)

    numeric_headers = {'food waste prevented (kg)', 'co2 saved (tons)', 'quantity', 'qty/amount', 'market price', 'subsidy', 'new price'}
    processed_rows = [[Paragraph(escape(str(h)), header_style) for h in headers]]
    for row in payload['rows']:
        rendered = []
        for idx, value in enumerate(row):
            header = headers[idx]
            style = number_style if header.lower() in numeric_headers else cell_style
            rendered.append(_as_paragraph(value, style, _text_limit(header)))
        processed_rows.append(rendered)
    if len(processed_rows) == 1:
        processed_rows.append([Paragraph('No records found', cell_style)] + [''] * (len(headers) - 1))

    available_width = page_size[0] - doc.leftMargin - doc.rightMargin
    weights = [_header_weight(h) for h in headers]
    total_weight = sum(weights) or 1
    col_widths = [available_width * (w / total_weight) for w in weights]

    table = Table(processed_rows, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


@staff_member_required
def export_environmental_posted_unspecified_excel(request):
    payload = _build_environmental_posted_export_payload(request, 'unspecified')
    return _export_environmental_posted_excel(payload)


@staff_member_required
def export_environmental_posted_unspecified_pdf(request):
    payload = _build_environmental_posted_export_payload(request, 'unspecified')
    return _export_environmental_posted_pdf(payload)


@staff_member_required
def export_environmental_posted_subsidized_excel(request):
    payload = _build_environmental_posted_export_payload(request, 'subsidized')
    return _export_environmental_posted_excel(payload)


@staff_member_required
def export_environmental_posted_subsidized_pdf(request):
    payload = _build_environmental_posted_export_payload(request, 'subsidized')
    return _export_environmental_posted_pdf(payload)


@staff_member_required
def export_environmental_posted_specified_excel(request):
    payload = _build_environmental_posted_export_payload(request, 'specified')
    return _export_environmental_posted_excel(payload)


@staff_member_required
def export_environmental_posted_specified_pdf(request):
    payload = _build_environmental_posted_export_payload(request, 'specified')
    return _export_environmental_posted_pdf(payload)


def _impact_base_queryset(days=None, date_from=None, date_to=None):
    """Base accepted donations queryset for impact reports with optional date range."""
    qs = Donation.objects.filter(status='accepted').select_related('donor', 'foodbank')
    if days is not None:
        start = timezone.now() - timedelta(days=int(days))
        qs = qs.filter(donated_at__gte=start)
    if date_from:
        qs = qs.filter(donated_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(donated_at__date__lte=date_to)
    return qs.order_by('-donated_at')


@staff_member_required
def impact_donations_list(request):
    """View all donations contributing to environmental impact, with filters and pagination."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        all_donations = _impact_base_queryset(days=days_int)
    else:
        all_donations = _impact_base_queryset(date_from=date_from, date_to=date_to)

    category_filter = request.GET.get('donation_category')
    mode_filter = request.GET.get('donation_mode')
    foodbank_filter = request.GET.get('foodbank')
    donor_filter = request.GET.get('donor')
    if category_filter:
        all_donations = all_donations.filter(donation_category=category_filter)
    if mode_filter:
        all_donations = all_donations.filter(donation_mode=mode_filter)
    if foodbank_filter:
        all_donations = all_donations.filter(foodbank_id=foodbank_filter)
    if donor_filter:
        all_donations = all_donations.filter(donor_id=donor_filter)

    paginator = Paginator(all_donations, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    donors = CustomUser.objects.filter(user_type='DONOR').order_by('email')

    context = {
        'title': 'Impact Donations',
        'page_obj': page_obj,
        'foodbanks': foodbanks,
        'donors': donors,
        'days': days_int if not (date_from or date_to) else None,
        'current_filters': {
            'days': days,
            'date_from': date_from_str,
            'date_to': date_to_str,
            'donation_category': category_filter,
            'donation_mode': mode_filter,
            'foodbank': foodbank_filter,
            'donor': donor_filter,
        },
        'donation_categories': Donation.DONATION_CATEGORIES,
        'donation_modes': Donation.DONATION_MODES,
    }
    return render(request, 'custom_admin/impact_donations_list.html', context)


@staff_member_required
def received_donations_impact_list(request):
    """Full list of received donations (accepted, with foodbank) with Waste Prevented and COâ‚‚ Saved columns."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        qs = _impact_base_queryset(days=days_int)
    else:
        qs = _impact_base_queryset(date_from=date_from, date_to=date_to)
    qs = qs.filter(foodbank__isnull=False).select_related('donor', 'donor__donor_profile', 'foodbank')

    category_filter = request.GET.get('donation_category')
    mode_filter = request.GET.get('donation_mode')
    foodbank_filter = request.GET.get('foodbank')
    if category_filter:
        qs = qs.filter(donation_category=category_filter)
    if mode_filter:
        qs = qs.filter(donation_mode=mode_filter)
    if foodbank_filter:
        qs = qs.filter(foodbank_id=foodbank_filter)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    for donation in page_obj.object_list:
        impact = get_donation_impact(donation)
        donation.waste_prevented_kg = impact['waste_prevented_kg']
        donation.co2_saved_tons = impact['co2_saved_tons']

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')
    context = {
        'title': 'Received Donations with Environmental Impact',
        'page_obj': page_obj,
        'foodbanks': foodbanks,
        'days': days_int if not (date_from or date_to) else None,
        'current_filters': {
            'days': days,
            'date_from': date_from_str or '',
            'date_to': date_to_str or '',
            'donation_category': category_filter or '',
            'donation_mode': mode_filter or '',
            'foodbank': foodbank_filter or '',
        },
        'donation_categories': Donation.DONATION_CATEGORIES,
        'donation_modes': Donation.DONATION_MODES,
    }
    return render(request, 'custom_admin/received_donations_impact_list.html', context)


@staff_member_required
def impact_monthly_breakdown(request):
    """Full monthly impact breakdown with filters: year or months (6/12/24)."""
    months_param = request.GET.get('months', '12')
    year_param = request.GET.get('year')
    try:
        n_months = int(months_param)
        n_months = min(max(n_months, 6), 24)
    except (TypeError, ValueError):
        n_months = 12

    from datetime import datetime
    if year_param:
        try:
            year = int(year_param)
            start_date = timezone.make_aware(datetime(year, 1, 1))
            end_date = timezone.make_aware(datetime(year, 12, 31, 23, 59, 59))
            month_starts = [timezone.make_aware(datetime(year, m, 1)) for m in range(1, 13)]
            n_months = 12
        except (TypeError, ValueError):
            start_date = timezone.now() - timedelta(days=30 * n_months)
            month_starts = [timezone.now() - timedelta(days=30 * (n_months - 1 - i)) for i in range(n_months)]
    else:
        start_date = timezone.now() - timedelta(days=30 * n_months)
        month_starts = [timezone.now() - timedelta(days=30 * (n_months - 1 - i)) for i in range(n_months)]

    all_donations = Donation.objects.filter(
        donated_at__gte=start_date,
        status='accepted'
    ).select_related('donor', 'foodbank')
    if year_param and len(month_starts) == 12:
        end_date = timezone.make_aware(datetime(year, 12, 31, 23, 59, 59))
        all_donations = all_donations.filter(donated_at__lte=end_date)

    monthly_data = []
    for i, month_start in enumerate(month_starts):
        if i + 1 < len(month_starts):
            month_end = month_starts[i + 1]
        else:
            month_end = month_start + timedelta(days=31)
        month_donations = all_donations.filter(
            donated_at__gte=month_start,
            donated_at__lt=month_end
        )
        food_donations = month_donations.filter(donation_category='food')
        free_food = food_donations.filter(donation_mode='free').aggregate(total=Sum('quantity'))['total'] or 0
        subsidized_food = food_donations.filter(donation_mode='subsidized').aggregate(total=Sum('subsidized_quantity'))['total'] or 0
        waste_prevented = (free_food * 1.0) + (subsidized_food * 0.8)
        co2_saved = (waste_prevented * 2.5) / 1000
        monthly_data.append({
            'month': month_start.strftime('%B %Y'),
            'food_donated': round(free_food, 2),
            'subsidized_food': round(subsidized_food, 2),
            'waste_prevented': round(waste_prevented, 2),
            'co2_saved': round(co2_saved, 3),
            'donation_count': month_donations.count(),
        })

    context = {
        'title': 'Monthly Impact Breakdown',
        'monthly_data': monthly_data,
        'months': n_months,
        'current_filters': {'months': months_param, 'year': year_param},
    }
    return render(request, 'custom_admin/impact_monthly_breakdown.html', context)


@staff_member_required
def impact_donors_list(request):
    """View all donors with impact metrics (count, quantity, amount), filters and pagination."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        all_donations = _impact_base_queryset(days=days_int)
    else:
        all_donations = _impact_base_queryset(date_from=date_from, date_to=date_to)

    search = request.GET.get('search')
    if search:
        all_donations = all_donations.filter(
            Q(donor__email__icontains=search) |
            Q(donor__first_name__icontains=search) |
            Q(donor__last_name__icontains=search)
        )

    donors_qs = all_donations.values('donor_id', 'donor__email', 'donor__first_name', 'donor__last_name').annotate(
        total_donations=Count('id'),
        total_quantity=Sum('quantity'),
        total_amount=Sum('amount'),
        total_subsidized_value=Sum('subsidized_price'),
    ).order_by('-total_donations')
    paginator = Paginator(donors_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'title': 'Impact Donors',
        'page_obj': page_obj,
        'days': days_int if not (date_from or date_to) else None,
        'current_filters': {
            'days': days,
            'date_from': date_from_str,
            'date_to': date_to_str,
            'search': search,
        },
    }
    return render(request, 'custom_admin/impact_donors_list.html', context)


@staff_member_required
def impact_foodbanks_list(request):
    """View all food banks by received volume with filters and pagination."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        all_donations = _impact_base_queryset(days=days_int)
    else:
        all_donations = _impact_base_queryset(date_from=date_from, date_to=date_to)

    foodbanks_qs = all_donations.values('foodbank_id', 'foodbank__foodbank_name').annotate(
        total_received=Count('id'),
        total_quantity=Sum('quantity'),
    ).order_by('-total_received')
    paginator = Paginator(foodbanks_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'title': 'Impact Food Banks',
        'page_obj': page_obj,
        'days': days_int if not (date_from or date_to) else None,
        'current_filters': {
            'days': days,
            'date_from': date_from_str,
            'date_to': date_to_str,
        },
    }
    return render(request, 'custom_admin/impact_foodbanks_list.html', context)


# --- Export helpers (PDF/Excel) for impact list views ---

def _impact_donations_queryset_from_request(request):
    """Build filtered impact donations queryset from request GET (same logic as impact_donations_list)."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        qs = _impact_base_queryset(days=days_int)
    else:
        qs = _impact_base_queryset(date_from=date_from, date_to=date_to)
    if request.GET.get('donation_category'):
        qs = qs.filter(donation_category=request.GET.get('donation_category'))
    if request.GET.get('donation_mode'):
        qs = qs.filter(donation_mode=request.GET.get('donation_mode'))
    if request.GET.get('foodbank'):
        qs = qs.filter(foodbank_id=request.GET.get('foodbank'))
    if request.GET.get('donor'):
        qs = qs.filter(donor_id=request.GET.get('donor'))
    return qs


@staff_member_required
def export_impact_donations_pdf(request):
    """Export impact donations list to PDF (same filters as list view)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    donations = _impact_donations_queryset_from_request(request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="impact_donations_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#10b981'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph('Impact Donations Report', title_style))
    elements.append(Paragraph(f'Total: {donations.count()} donations', styles['Normal']))
    elements.append(Spacer(1, 15))
    table_data = [['Donor', 'Food Bank', 'Category', 'Mode', 'Quantity / Amount', 'Donated at']]
    for d in donations[:500]:
        qty_amt = []
        if d.quantity:
            qty_amt.append(str(d.quantity))
        if d.subsidized_quantity:
            qty_amt.append(f"{d.subsidized_quantity} (subsidized)")
        if d.amount:
            qty_amt.append(f"KES {d.amount:,.0f}")
        table_data.append([
            d.donor.email if d.donor else 'â€”',
            d.foodbank.foodbank_name if d.foodbank else 'â€”',
            d.get_donation_category_display() if hasattr(d, 'get_donation_category_display') else str(d.donation_category),
            d.get_donation_mode_display() if hasattr(d, 'get_donation_mode_display') else str(d.donation_mode),
            ' / '.join(qty_amt) if qty_amt else 'â€”',
            d.donated_at.strftime('%Y-%m-%d') if d.donated_at else 'â€”',
        ])
    t = Table(table_data, colWidths=[1.5*inch, 1.8*inch, 1*inch, 1*inch, 1.5*inch, 1*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(t)
    doc.build(elements)
    return response


@staff_member_required
def export_impact_donations_excel(request):
    """Export impact donations list to Excel (same filters as list view)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    donations = _impact_donations_queryset_from_request(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="impact_donations_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Impact Donations"
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ws['A1'] = "Impact Donations Report"
    ws['A1'].font = Font(size=14, bold=True, color="059669")
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total: {donations.count()}"
    ws['A2'].font = Font(size=9, color="6B7280")
    headers = ['Donor', 'Food Bank', 'Category', 'Mode', 'Quantity / Amount', 'Donated at']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col)
        c.value = h
        c.fill = header_fill
        c.font = header_font
    for row_idx, d in enumerate(donations, start=5):
        qty_amt = []
        if d.quantity:
            qty_amt.append(str(d.quantity))
        if d.subsidized_quantity:
            qty_amt.append(f"{d.subsidized_quantity} (subsidized)")
        if d.amount:
            qty_amt.append(f"KES {d.amount:,.0f}")
        ws.cell(row=row_idx, column=1, value=d.donor.email if d.donor else 'â€”')
        ws.cell(row=row_idx, column=2, value=d.foodbank.foodbank_name if d.foodbank else 'â€”')
        ws.cell(row=row_idx, column=3, value=d.get_donation_category_display() if hasattr(d, 'get_donation_category_display') else str(d.donation_category))
        ws.cell(row=row_idx, column=4, value=d.get_donation_mode_display() if hasattr(d, 'get_donation_mode_display') else str(d.donation_mode))
        ws.cell(row=row_idx, column=5, value=' / '.join(qty_amt) if qty_amt else 'â€”')
        ws.cell(row=row_idx, column=6, value=d.donated_at.strftime('%Y-%m-%d') if d.donated_at else 'â€”')
    wb.save(response)
    return response


def _impact_monthly_data_from_request(request):
    """Build monthly breakdown data from request GET (same logic as impact_monthly_breakdown)."""
    months_param = request.GET.get('months', '12')
    year_param = request.GET.get('year')
    try:
        n_months = int(months_param)
        n_months = min(max(n_months, 6), 24)
    except (TypeError, ValueError):
        n_months = 12
    from datetime import datetime
    if year_param:
        try:
            year = int(year_param)
            start_date = timezone.make_aware(datetime(year, 1, 1))
            month_starts = [timezone.make_aware(datetime(year, m, 1)) for m in range(1, 13)]
            n_months = 12
        except (TypeError, ValueError):
            start_date = timezone.now() - timedelta(days=30 * n_months)
            month_starts = [timezone.now() - timedelta(days=30 * (n_months - 1 - i)) for i in range(n_months)]
    else:
        start_date = timezone.now() - timedelta(days=30 * n_months)
        month_starts = [timezone.now() - timedelta(days=30 * (n_months - 1 - i)) for i in range(n_months)]
    all_donations = Donation.objects.filter(donated_at__gte=start_date, status='accepted')
    if year_param and len(month_starts) == 12:
        end_date = timezone.make_aware(datetime(int(year_param), 12, 31, 23, 59, 59))
        all_donations = all_donations.filter(donated_at__lte=end_date)
    monthly_data = []
    for i, month_start in enumerate(month_starts):
        if i + 1 < len(month_starts):
            month_end = month_starts[i + 1]
        else:
            month_end = month_start + timedelta(days=31)
        month_donations = all_donations.filter(donated_at__gte=month_start, donated_at__lt=month_end)
        food_donations = month_donations.filter(donation_category='food')
        free_food = food_donations.filter(donation_mode='free').aggregate(total=Sum('quantity'))['total'] or 0
        subsidized_food = food_donations.filter(donation_mode='subsidized').aggregate(total=Sum('subsidized_quantity'))['total'] or 0
        waste_prevented = (free_food * 1.0) + (subsidized_food * 0.8)
        co2_saved = (waste_prevented * 2.5) / 1000
        monthly_data.append({
            'month': month_start.strftime('%B %Y'),
            'food_donated': round(free_food, 2),
            'subsidized_food': round(subsidized_food, 2),
            'waste_prevented': round(waste_prevented, 2),
            'co2_saved': round(co2_saved, 3),
            'donation_count': month_donations.count(),
        })
    return monthly_data


@staff_member_required
def export_impact_monthly_pdf(request):
    """Export monthly impact breakdown to PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    monthly_data = _impact_monthly_data_from_request(request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="impact_monthly_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#3b82f6'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph('Monthly Impact Breakdown', title_style))
    elements.append(Spacer(1, 10))
    table_data = [['Month', 'Donations', 'Food Donated (kg)', 'Subsidized Food (kg)', 'Waste Prevented (kg)', 'COâ‚‚ Saved (tons)']]
    for row in monthly_data:
        table_data.append([row['month'], str(row['donation_count']), str(row['food_donated']), str(row['subsidized_food']), str(row['waste_prevented']), str(row['co2_saved'])])
    t = Table(table_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(t)
    doc.build(elements)
    return response


@staff_member_required
def export_impact_monthly_excel(request):
    """Export monthly impact breakdown to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    monthly_data = _impact_monthly_data_from_request(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="impact_monthly_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Impact"
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ws['A1'] = "Monthly Impact Breakdown"
    ws['A1'].font = Font(size=14, bold=True, color="3b82f6")
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=9, color="6B7280")
    headers = ['Month', 'Donations', 'Food Donated (kg)', 'Subsidized Food (kg)', 'Waste Prevented (kg)', 'COâ‚‚ Saved (tons)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col)
        c.value = h
        c.fill = header_fill
        c.font = header_font
    for row_idx, row in enumerate(monthly_data, start=5):
        ws.cell(row=row_idx, column=1, value=row['month'])
        ws.cell(row=row_idx, column=2, value=row['donation_count'])
        ws.cell(row=row_idx, column=3, value=row['food_donated'])
        ws.cell(row=row_idx, column=4, value=row['subsidized_food'])
        ws.cell(row=row_idx, column=5, value=row['waste_prevented'])
        ws.cell(row=row_idx, column=6, value=row['co2_saved'])
    wb.save(response)
    return response


def _impact_donors_data_from_request(request):
    """Build donors aggregate list from request GET (same filters as impact_donors_list)."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        all_donations = _impact_base_queryset(days=days_int)
    else:
        all_donations = _impact_base_queryset(date_from=date_from, date_to=date_to)
    search = request.GET.get('search')
    if search:
        all_donations = all_donations.filter(
            Q(donor__email__icontains=search) |
            Q(donor__first_name__icontains=search) |
            Q(donor__last_name__icontains=search)
        )
    return list(all_donations.values('donor_id', 'donor__email', 'donor__first_name', 'donor__last_name').annotate(
        total_donations=Count('id'),
        total_quantity=Sum('quantity'),
        total_amount=Sum('amount'),
        total_subsidized_value=Sum('subsidized_price'),
    ).order_by('-total_donations'))


@staff_member_required
def export_impact_donors_pdf(request):
    """Export impact donors list to PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    donors = _impact_donors_data_from_request(request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="impact_donors_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#f59e0b'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph('Impact Donors Report', title_style))
    elements.append(Paragraph(f'Total: {len(donors)} donors', styles['Normal']))
    elements.append(Spacer(1, 15))
    table_data = [['#', 'Donor', 'Donations', 'Quantity (kg)', 'Amount (KES)']]
    for idx, d in enumerate(donors, 1):
        name = f"{d.get('donor__first_name') or ''} {d.get('donor__last_name') or ''}".strip() or d.get('donor__email') or 'â€”'
        amt = d.get('total_amount')
        sub_val = d.get('total_subsidized_value')
        amt_str = f"{amt:,.0f}" if amt else "â€”"
        if sub_val:
            amt_str += f" (Subsidized: {sub_val:,.0f})"
        table_data.append([
            str(idx),
            name[:40],
            str(d.get('total_donations', 0)),
            str(d.get('total_quantity') or 0),
            amt_str,
        ])
    t = Table(table_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.2*inch, 1.8*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(t)
    doc.build(elements)
    return response


@staff_member_required
def export_impact_donors_excel(request):
    """Export impact donors list to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    donors = _impact_donors_data_from_request(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="impact_donors_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Impact Donors"
    header_fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ws['A1'] = "Impact Donors Report"
    ws['A1'].font = Font(size=14, bold=True, color="f59e0b")
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total: {len(donors)} donors"
    ws['A2'].font = Font(size=9, color="6B7280")
    headers = ['#', 'Donor', 'Donations', 'Quantity (kg)', 'Amount (KES)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col)
        c.value = h
        c.fill = header_fill
        c.font = header_font
    for row_idx, d in enumerate(donors, start=5):
        name = f"{d.get('donor__first_name') or ''} {d.get('donor__last_name') or ''}".strip() or d.get('donor__email') or 'â€”'
        amt = d.get('total_amount')
        sub_val = d.get('total_subsidized_value')
        amt_str = f"{amt:,.0f}" if amt else "â€”"
        if sub_val:
            amt_str += f" (Subsidized: {sub_val:,.0f})"
        ws.cell(row=row_idx, column=1, value=row_idx - 4)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=d.get('total_donations', 0))
        ws.cell(row=row_idx, column=4, value=d.get('total_quantity') or 0)
        ws.cell(row=row_idx, column=5, value=amt_str)
    wb.save(response)
    return response


def _impact_foodbanks_data_from_request(request):
    """Build food banks aggregate list from request GET (same filters as impact_foodbanks_list)."""
    days = request.GET.get('days', '365')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    date_from = parse_date(date_from_str) if date_from_str else None
    date_to = parse_date(date_to_str) if date_to_str else None
    try:
        days_int = int(days) if days else 365
    except (TypeError, ValueError):
        days_int = 365
    if not date_from and not date_to:
        all_donations = _impact_base_queryset(days=days_int)
    else:
        all_donations = _impact_base_queryset(date_from=date_from, date_to=date_to)
    return list(all_donations.values('foodbank_id', 'foodbank__foodbank_name').annotate(
        total_received=Count('id'),
        total_quantity=Sum('quantity'),
    ).order_by('-total_received'))


@staff_member_required
def export_impact_foodbanks_pdf(request):
    """Export impact food banks list to PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    foodbanks = _impact_foodbanks_data_from_request(request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="impact_foodbanks_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#3b82f6'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph('Impact Food Banks Report', title_style))
    elements.append(Paragraph(f'Total: {len(foodbanks)} food banks', styles['Normal']))
    elements.append(Spacer(1, 15))
    table_data = [['#', 'Food Bank', 'Donations Received', 'Quantity (kg)']]
    for idx, fb in enumerate(foodbanks, 1):
        table_data.append([str(idx), (fb.get('foodbank__foodbank_name') or 'â€”')[:50], str(fb.get('total_received', 0)), str(fb.get('total_quantity') or 0)])
    t = Table(table_data, colWidths=[0.5*inch, 3*inch, 1.5*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(t)
    doc.build(elements)
    return response


@staff_member_required
def export_impact_foodbanks_excel(request):
    """Export impact food banks list to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    foodbanks = _impact_foodbanks_data_from_request(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="impact_foodbanks_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Impact Food Banks"
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ws['A1'] = "Impact Food Banks Report"
    ws['A1'].font = Font(size=14, bold=True, color="3b82f6")
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total: {len(foodbanks)} food banks"
    ws['A2'].font = Font(size=9, color="6B7280")
    headers = ['#', 'Food Bank', 'Donations Received', 'Quantity (kg)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col)
        c.value = h
        c.fill = header_fill
        c.font = header_font
    for row_idx, fb in enumerate(foodbanks, start=5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4)
        ws.cell(row=row_idx, column=2, value=fb.get('foodbank__foodbank_name') or 'â€”')
        ws.cell(row=row_idx, column=3, value=fb.get('total_received', 0))
        ws.cell(row=row_idx, column=4, value=fb.get('total_quantity') or 0)
    wb.save(response)
    return response
