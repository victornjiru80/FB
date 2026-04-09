from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
import json
import html
from django.utils import timezone
from datetime import timedelta
from .models import Donation, FoodBankProfile, DonorProfile, DonationResponse, DonationAllocation
from .decorators import donor_required, foodbank_required
STATUS_CLASS_MAP = {
    'Acknowledged by Recipient': 'fulfilled',
    'Received By Recipient': 'fulfilled',
    'Fulfilled': 'fulfilled',
    'Fulfilled-Acknowledged': 'fulfilled',
    'Fulfilled-Received': 'fulfilled',
    'Awaiting Recipient': 'accepted',
    'Accepted → Awaiting Pickup': 'accepted',
    'Delivery In Transit': 'accepted',
    'Delivery Scheduled': 'accepted',
    'Accepted by Foodbank': 'accepted',
    'Delivered': 'fulfilled',
    'Accepted': 'accepted',
    'Awaiting Foodbank Review': 'pending',
    'Awaiting FB Review': 'pending',
    'Pending': 'pending',
    'Partially Fulfilled': 'partial',
    'Partially Fulfilled – Acknowledged': 'partial',
    'Partially Fulfilled – Received': 'partial',
    'Declined by FB': 'declined',
    'Declined by Recipient': 'declined',
    'Declined': 'declined',
}


def _resolve_unit_label(obj, unit_field='quantity_unit', custom_field='custom_unit', fallback='units'):
    """Resolve unit label, preferring custom text when unit is 'other'."""
    if not obj:
        return fallback

    unit_value = getattr(obj, unit_field, None)
    custom_value = (getattr(obj, custom_field, None) or '').strip()

    if unit_value == 'other' and custom_value:
        return custom_value

    display_method_name = f'get_{unit_field}_display'
    display_method = getattr(obj, display_method_name, None)
    if callable(display_method):
        try:
            display_value = display_method()
            if display_value:
                return display_value
        except Exception:
            pass

    return unit_value or fallback


def _get_related_request(donation):
    """Return the RequestManagement instance linked to this donation, if any."""
    if hasattr(donation, 'request_management') and donation.request_management:
        return donation.request_management
    if hasattr(donation, 'foodbank_request') and donation.foodbank_request:
        fb_req = donation.foodbank_request
        # FoodBankRequest can link to recipient request via original_request or linked_request_management
        req = getattr(fb_req, 'original_request', None) or getattr(fb_req, 'linked_request_management', None)
        if req:
            return req
        # Fallback: RequestManagement has foodbank_request FK (related_name='recipient_requests')
        if hasattr(fb_req, 'recipient_requests'):
            first_req = fb_req.recipient_requests.order_by('-time_of_request').first()
            if first_req:
                return first_req
    return None


def _get_foodbank_note_for_export(donation):
    """Collect foodbank note for PDF/CSV/Excel export. Excludes recipient content (e.g. additional_notes is used as recipient note in UI)."""
    note = None
    if getattr(donation, 'foodbank_request', None):
        note = getattr(donation.foodbank_request, 'foodbank_note', None)
    if not note and getattr(donation, 'request_management', None):
        note = getattr(donation.request_management, 'foodbank_response_note', None)
    # Decline messages from foodbank (when they decline donation or request)
    if not note and getattr(donation, 'decline_message', None) and donation.decline_message:
        note = donation.decline_message
    if not note and getattr(donation, 'request_management', None) and getattr(donation.request_management, 'decline_message', None) and donation.request_management.decline_message:
        note = donation.request_management.decline_message
    if not note and getattr(donation, 'foodbank_request', None) and getattr(donation.foodbank_request, 'original_request', None) and getattr(donation.foodbank_request.original_request, 'decline_message', None) and donation.foodbank_request.original_request.decline_message:
        note = donation.foodbank_request.original_request.decline_message
    return note or ''


def _get_donor_foodbank_note_for_export(donation):
    """Foodbank note for donor exports: acceptance note, or decline reason if foodbank declined."""
    status_display = get_display_status(donation)
    if status_display in {"Declined by Foodbank", "Declined by FB"}:
        decline_reason = (_get_donor_foodbank_decline_reason(donation) or '').strip()
        return decline_reason or 'No note'

    note = ''
    if getattr(donation, 'foodbank_request', None):
        note = getattr(donation.foodbank_request, 'foodbank_note', None) or ''
    if not note and getattr(donation, 'request_management', None):
        note = getattr(donation.request_management, 'foodbank_response_note', None) or ''
    # Some direct-request acceptance notes are stored on Donation.decline_message.
    if (
        not note
        and getattr(donation, 'decline_message', None)
        and not getattr(donation, 'declined_by_recipient_id', None)
        and getattr(donation, 'status', None) != 'declined'
    ):
        note = donation.decline_message
    note = (note or '').strip()
    return note or 'No note'


def _get_description_for_export(donation):
    """Get description for export; for direct donations use title + description; for monetary use linked request or other_description/message; for CSR use csr_description or subcategory."""
    # Direct donations (foodbank-created request, no recipient request): same as table - title and description
    fb_req = getattr(donation, 'foodbank_request', None)
    if fb_req and getattr(fb_req, 'original_request', None) is None:
        title = (getattr(fb_req, 'title', None) or '').strip() or 'Untitled Request'
        desc = (getattr(fb_req, 'description', None) or '').strip() or 'No description provided'
        return f"{title}\n{desc}"
    if donation.donation_type == 'item':
        if fb_req:
            if getattr(fb_req, 'original_request', None) and getattr(fb_req.original_request, 'description', None):
                return fb_req.original_request.description
            if getattr(fb_req, 'linked_request_management', None) and getattr(fb_req.linked_request_management, 'description', None):
                return fb_req.linked_request_management.description
            if getattr(fb_req, 'description', None) and fb_req.description:
                return fb_req.description
        return donation.item_name or 'Unnamed Item'
    if donation.donation_type == 'money':
        fb_req = getattr(donation, 'foodbank_request', None)
        if fb_req:
            if getattr(fb_req, 'original_request', None) and fb_req.original_request.description:
                return fb_req.original_request.description
            if getattr(fb_req, 'linked_request_management', None) and fb_req.linked_request_management.description:
                return fb_req.linked_request_management.description
            if getattr(fb_req, 'title', None) and fb_req.title:
                return fb_req.title
            if getattr(fb_req, 'description', None) and fb_req.description:
                return fb_req.description
        # Unspecified monetary: use other_description or message
        return (getattr(donation, 'other_description', None) or '').strip() or (getattr(donation, 'message', None) or '').strip() or 'Monetary Donation'
    if donation.donation_type == 'subsidized':
        return donation.subsidized_product_type or 'Subsidized Goods'
    if donation.donation_type == 'csr':
        csr_desc = (getattr(donation, 'csr_description', None) or '').strip()
        if csr_desc:
            return csr_desc
        if getattr(donation, 'csr_subcategory', None) == 'other' and getattr(donation, 'csr_custom_subcategory', None):
            return donation.csr_custom_subcategory
        return f"CSR - {donation.get_csr_subcategory_display()}"
    return donation.other_description or 'Other donation'


def _get_donor_requested_qty_unit(donation):
    """Resolve requested quantity/unit for donor exports."""
    fb_req = getattr(donation, 'foodbank_request', None)
    if fb_req and getattr(fb_req, 'quantity_needed', None) is not None:
        return fb_req.quantity_needed, _resolve_unit_label(fb_req, 'quantity_unit', 'custom_unit', 'units')
    if fb_req and getattr(fb_req, 'original_request', None):
        orig = fb_req.original_request
        if getattr(orig, 'quantity', None) is not None:
            return orig.quantity, _resolve_unit_label(orig, 'unit', 'custom_unit', 'units')
    return None, ''


def _get_donor_qty_amount_display(donation):
    """Qty/Amount text matching donor table display."""
    requested_qty, requested_unit = _get_donor_requested_qty_unit(donation)
    req_suffix = f"(of {requested_qty} {requested_unit} requested)" if requested_qty is not None else ""
    for_suffix = f"(for {requested_qty} {requested_unit})" if requested_qty is not None else ""

    if donation.donation_type == 'item':
        text = f"{donation.quantity} {donation.quantity_unit or 'units'}"
        if req_suffix:
            text += f"\n{req_suffix}"
        return text

    if donation.donation_type == 'money':
        text = f"KES {donation.amount:.0f}" if donation.amount is not None else '-'
        if for_suffix and donation.amount is not None:
            text += f"\n{for_suffix}"
        if req_suffix and donation.amount is not None:
            text += f"\n{req_suffix}"
        return text

    if donation.donation_type == 'subsidized':
        if donation.subsidized_price is None:
            return '-'
        text = f"KES {donation.subsidized_price:.0f}"
        if donation.subsidized_market_price:
            text += f"\n(Market KES {donation.subsidized_market_price:.0f})"
        if for_suffix:
            text += f"\n{for_suffix}"
        if req_suffix:
            text += f"\n{req_suffix}"
        return text

    if donation.donation_type in ('csr', 'other'):
        parts = []
        if donation.quantity:
            parts.append(f"{donation.quantity} {donation.quantity_unit or 'units'}")
        if donation.amount:
            parts.append(f"KES {donation.amount:.0f}")
        return '\n'.join(parts) if parts else '-'

    if donation.quantity:
        return f"{donation.quantity} {donation.quantity_unit or 'units'}"

    return '-'


def _get_donor_foodbank_decline_reason(donation):
    """Return foodbank decline reason for donor exports."""

    def _clean(value):
        return (value or '').strip()

    def _is_foodbank_declined_request(req):
        if not req:
            return False
        if bool(getattr(req, 'foodbank_declined_request', False)) or bool(getattr(req, 'foodbank_declined_donation', False)):
            return True
        method = getattr(req, 'was_declined_by_foodbank', None)
        if callable(method):
            try:
                if bool(method()):
                    return True
            except Exception:
                pass
        if bool(getattr(req, 'recipient_declined_request', False)):
            return False
        req_status = getattr(req, 'status', None)
        if req_status != 'declined':
            return False
        updater = getattr(req, 'updated_by', None)
        updater_type = _clean(getattr(updater, 'user_type', '')).upper()
        if updater_type == 'RECIPIENT':
            return False
        return True

    def _declined_donation_reason(d):
        if not d:
            return ''
        if getattr(d, 'status', None) != 'declined':
            return ''
        if getattr(d, 'declined_by_recipient_id', None):
            return ''
        return _clean(getattr(d, 'decline_message', None))

    # 1) This donation.
    reason = _declined_donation_reason(donation)
    if reason:
        return reason

    # 2) Unspecified/general donation flow.
    unspecified = getattr(donation, 'unspecified_management', None)
    if unspecified and getattr(unspecified, 'foodbank_status', None) == 'declined_by_foodbank':
        reason = _clean(getattr(unspecified, 'foodbank_decline_reason', None))
        if reason:
            return reason
        reason = _clean(getattr(donation, 'decline_message', None))
        if reason:
            return reason

    # 3) Related request decline reason.
    related_request = _get_related_request(donation)
    if _is_foodbank_declined_request(related_request):
        reason = _clean(getattr(related_request, 'decline_message', None))
        if reason:
            return reason

    # 4) Collect linked FoodBankRequest records and scan their request messages/donations.
    linked_foodbank_requests = []
    seen_fb_req_ids = set()

    def _push_fb_req(fb_req):
        if not fb_req:
            return
        fb_req_id = getattr(fb_req, 'id', None)
        if fb_req_id and fb_req_id in seen_fb_req_ids:
            return
        if fb_req_id:
            seen_fb_req_ids.add(fb_req_id)
        linked_foodbank_requests.append(fb_req)

    _push_fb_req(getattr(donation, 'foodbank_request', None))
    if related_request:
        _push_fb_req(getattr(related_request, 'foodbank_request', None))
        try:
            for fb_req in related_request.foodbank_request_created.all():
                _push_fb_req(fb_req)
        except Exception:
            pass
        try:
            for fb_req in related_request.donor_requests.all():
                _push_fb_req(fb_req)
        except Exception:
            pass

    for fb_req in linked_foodbank_requests:
        orig_req = getattr(fb_req, 'original_request', None)
        if _is_foodbank_declined_request(orig_req):
            reason = _clean(getattr(orig_req, 'decline_message', None))
            if reason:
                return reason

        linked_req = getattr(fb_req, 'linked_request_management', None)
        if _is_foodbank_declined_request(linked_req):
            reason = _clean(getattr(linked_req, 'decline_message', None))
            if reason:
                return reason

        try:
            for linked_donation in fb_req.donations.all().order_by('-donated_at', '-id'):
                reason = _declined_donation_reason(linked_donation)
                if reason:
                    return reason
        except Exception:
            pass

    # 5) Scan donations directly linked to request_management relations.
    if related_request:
        try:
            for linked_donation in related_request.donations.all().order_by('-donated_at', '-id'):
                reason = _declined_donation_reason(linked_donation)
                if reason:
                    return reason
        except Exception:
            pass
        try:
            for alloc in related_request.donation_allocations.select_related('donation').order_by('-allocated_at', '-id'):
                reason = _declined_donation_reason(getattr(alloc, 'donation', None))
                if reason:
                    return reason
        except Exception:
            pass

    # 6) Discussion-driven declines (other/csr) can be marked declined without donation.decline_message.
    discussion = getattr(donation, 'discussion', None)
    if discussion and getattr(discussion, 'status', None) == 'declined':
        try:
            reason = _clean(
                discussion.messages
                .filter(sender__user_type='FOODBANK')
                .order_by('-sent_at')
                .values_list('message', flat=True)
                .first()
            )
            if reason:
                return reason
        except Exception:
            pass

    # 7) If status still reads foodbank-declined, avoid "No note" and return explicit fallback.
    if get_display_status(donation) == "Declined by Foodbank":
        return "Declined by Foodbank (reason not recorded)"

    return 'No note'


def _get_donor_recipient_decline_reason(donation, latest_recipient_declines=None):
    """Return recipient decline reason for donor exports, if available."""
    latest_recipient_declines = latest_recipient_declines or {}

    def _clean(value):
        return (value or '').strip()

    reason = _clean(latest_recipient_declines.get(getattr(donation, 'id', None)))
    if reason:
        return reason

    if getattr(donation, 'declined_by_recipient_id', None):
        reason = _clean(getattr(donation, 'decline_message', None))
        if reason:
            return reason

    unspecified = getattr(donation, 'unspecified_management', None)
    if unspecified:
        reason = _clean(getattr(unspecified, 'recipient_decline_reason', None))
        if reason:
            return reason

    related_request = _get_related_request(donation)
    if related_request and bool(getattr(related_request, 'recipient_declined_request', False)):
        reason = _clean(getattr(related_request, 'decline_message', None))
        if reason:
            return reason

    fb_req = getattr(donation, 'foodbank_request', None)
    orig_req = getattr(fb_req, 'original_request', None) if fb_req else None
    if orig_req and bool(getattr(orig_req, 'recipient_declined_request', False)):
        reason = _clean(getattr(orig_req, 'decline_message', None))
        if reason:
            return reason

    return ''


def _get_donor_decline_reason_for_export(donation, latest_recipient_declines=None):
    """Return decline reason (recipient or foodbank) matching donor table status semantics."""
    recipient_reason = _get_donor_recipient_decline_reason(donation, latest_recipient_declines)
    status_display = get_display_status(donation)

    if status_display == "Declined by Recipient" or recipient_reason:
        return recipient_reason or 'Declined by Recipient (no note provided)'

    return _get_donor_foodbank_decline_reason(donation) or 'No note'


def _normalize_donor_type_category_filters(type_filter, category_filter):
    """Normalize donor table filters where Type=Food/Non-Food and Category=donation kind."""
    request_type_choices = {'food', 'non_food'}
    donation_kind_choices = {'item', 'subsidized', 'money', 'csr', 'other', 'monetary'}

    type_filter = (type_filter or '').strip()
    category_filter = (category_filter or '').strip()

    # Backward compatibility with older swapped filter semantics.
    if type_filter in donation_kind_choices and category_filter in request_type_choices:
        type_filter, category_filter = category_filter, type_filter
    elif type_filter in donation_kind_choices and not category_filter:
        category_filter = type_filter
        type_filter = ''
    elif category_filter in request_type_choices and not type_filter:
        type_filter = category_filter
        category_filter = ''

    if category_filter == 'monetary':
        category_filter = 'money'

    if type_filter and type_filter not in request_type_choices:
        type_filter = ''
    if category_filter and category_filter not in {'item', 'subsidized', 'money', 'csr', 'other'}:
        category_filter = ''

    return type_filter, category_filter


def _get_requested_quantity_value_for_export(donation):
    """Return only the requested quantity number (for CSV/Excel Requested column), or ''."""
    fb_req = getattr(donation, 'foodbank_request', None)
    if fb_req and getattr(fb_req, 'quantity_needed', None) is not None:
        return str(fb_req.quantity_needed)
    if fb_req and getattr(fb_req, 'original_request', None):
        orig = fb_req.original_request
        q = getattr(orig, 'quantity', None)
        if q is not None:
            return str(q)
    rm = getattr(donation, 'request_management', None)
    if rm and getattr(rm, 'quantity', None) is not None:
        return str(rm.quantity)
    return ''


def _get_requested_unit_for_export(donation):
    fb_req = getattr(donation, 'foodbank_request', None)
    if fb_req and getattr(fb_req, 'quantity_needed', None) is not None:
        return _resolve_unit_label(fb_req, 'quantity_unit', 'custom_unit', 'units')
    if fb_req and getattr(fb_req, 'original_request', None):
        orig = fb_req.original_request
        if getattr(orig, 'quantity', None) is not None:
            return _resolve_unit_label(orig, 'unit', 'custom_unit', 'units')
    rm = getattr(donation, 'request_management', None)
    if rm and getattr(rm, 'quantity', None) is not None:
        return _resolve_unit_label(rm, 'unit', 'custom_unit', 'units')
    return ''


def _get_requested_for_export(donation):
    """Return 'for X unit requested' when donation is linked to a request with quantity, else ''."""
    fb_req = getattr(donation, 'foodbank_request', None)
    if fb_req and getattr(fb_req, 'quantity_needed', None) is not None:
        qu = _resolve_unit_label(fb_req, 'quantity_unit', 'custom_unit', 'units')
        return f"for {fb_req.quantity_needed} {qu} requested"
    if fb_req and getattr(fb_req, 'original_request', None):
        orig = fb_req.original_request
        q = getattr(orig, 'quantity', None)
        if q is not None:
            u = _resolve_unit_label(orig, 'unit', 'custom_unit', 'units')
            return f"for {q} {u} requested"
    rm = getattr(donation, 'request_management', None)
    if rm and getattr(rm, 'quantity', None) is not None:
        u = _resolve_unit_label(rm, 'unit', 'custom_unit', 'units')
        return f"for {rm.quantity} {u} requested"
    return ''


def _get_foodbank_export_type_display(donation):
    """Type column for foodbank donations table/export: Food/Non-Food from target request."""
    if donation.donation_type in ('money', 'item', 'subsidized'):
        req_type = None

        # Direct donations may not have a recipient-side request, so prefer the
        # originating foodbank request type first.
        fb_req = getattr(donation, 'foodbank_request', None)
        if fb_req:
            for attr in ('donation_type', 'request_type', 'request_category'):
                value = getattr(fb_req, attr, None)
                if value in ('food', 'non_food'):
                    req_type = value
                    break

        # Then check related recipient/request-management objects.
        if not req_type:
            related_request = _get_related_request(donation)
            if related_request:
                for attr in ('request_type', 'donation_type', 'request_category'):
                    value = getattr(related_request, attr, None)
                    if value in ('food', 'non_food'):
                        req_type = value
                        break

        if not req_type:
            fallback = getattr(donation, 'donation_category', None)
            if fallback in ('food', 'non_food'):
                req_type = fallback
        if req_type == 'food':
            return 'Food'
        if req_type == 'non_food':
            return 'Non-Food'
        return '-'
    return donation.get_donation_type_display() or '-'


def _get_foodbank_export_category_display(donation):
    """Category column for foodbank donations table/export: donation kind."""
    if donation.donation_type == 'money' or donation.donation_category == 'monetary':
        return 'Monetary'
    if donation.donation_type == 'item':
        return 'Free Goods'
    if donation.donation_type == 'subsidized':
        return 'Subsidized Goods'
    if donation.donation_type == 'csr' or donation.donation_category == 'csr':
        if donation.csr_subcategory == 'other' and donation.csr_custom_subcategory:
            return f"CSR {donation.csr_custom_subcategory}"
        return f"CSR {donation.get_csr_subcategory_display()}"
    return donation.get_donation_category_display() or '-'


def _get_foodbank_export_quantity_details(donation):
    """Quantity details text matching table quantity column semantics."""
    requested_qty = _get_requested_quantity_value_for_export(donation) or '-'
    requested_unit = _get_requested_unit_for_export(donation) or '-'

    def _suffix(preposition):
        if requested_qty == '-' or requested_unit == '-':
            return ''
        return f"({preposition} {requested_qty} {requested_unit} requested)"

    if donation.donation_type == 'item':
        text = f"{donation.quantity or '-'} {donation.quantity_unit or 'u'}"
        requested_suffix = _suffix('of')
    elif donation.donation_type == 'money':
        amount_str = f"{float(donation.amount):.0f}" if donation.amount is not None else "-"
        text = f"KES {amount_str}" if amount_str != "-" else "-"
        requested_suffix = _suffix('for')
    elif donation.donation_type == 'subsidized':
        if donation.subsidized_price is not None:
            text = f"KES {float(donation.subsidized_price):.0f}"
        elif donation.amount is not None:
            text = f"KES {float(donation.amount):.0f}"
        elif donation.subsidized_quantity is not None:
            text = f"{donation.subsidized_quantity} {donation.subsidized_quantity_unit or 'units'}"
        else:
            text = '-'
        requested_suffix = _suffix('for')
    elif donation.donation_type in ('csr', 'other'):
        parts = []
        if donation.quantity:
            parts.append(f"{donation.quantity} {donation.quantity_unit or 'units'}")
        if donation.amount is not None:
            parts.append(f"KES {float(donation.amount):.0f}")
        text = '\n'.join(parts) if parts else '-'
        requested_suffix = _suffix('for')
    else:
        text = '-'
        requested_suffix = ''

    if requested_suffix:
        text += f"\n{requested_suffix}"
    return text


def _get_foodbank_export_status_display(donation):
    """Status text for foodbank exports (match donation row status logic used by table)."""
    return get_display_status(donation)


def _get_foodbank_export_decline_reason(donation):
    """Decline reason for foodbank exports (donation-level recipient/foodbank decline), else '-'."""
    related_request = _get_related_request(donation)

    # Donation-level recipient decline
    if getattr(donation, 'declined_by_recipient_id', None):
        direct_reason = (getattr(donation, 'decline_message', None) or '').strip()
        if direct_reason:
            return direct_reason
        if related_request:
            req_reason = (getattr(related_request, 'decline_message', None) or '').strip()
            if req_reason:
                return req_reason

    # Request-level recipient decline linked to this donation
    if related_request and bool(getattr(related_request, 'recipient_declined_request', False)):
        req_reason = (getattr(related_request, 'decline_message', None) or '').strip()
        if req_reason:
            return req_reason

    # Foodbank decline reason on donation/request
    if getattr(donation, 'status', None) == 'declined':
        fb_reason = (getattr(donation, 'decline_message', None) or '').strip()
        if fb_reason:
            return fb_reason
        if related_request:
            req_reason = (getattr(related_request, 'decline_message', None) or '').strip()
            if req_reason:
                return req_reason

    return '-'


def _build_foodbank_table_rows(donations, type_filter='', status_filter=''):
    """Build row objects using the same row semantics as the foodbank donations table."""
    from types import SimpleNamespace

    def is_direct_stock(d):
        if d.foodbank_request_id is not None or (hasattr(d, 'request_management') and d.request_management_id is not None):
            return False
        return True

    def has_allocations(d):
        return len(list(d.allocations.all())) > 0

    def _resolve_related_request(donation, allocation=None):
        if allocation and getattr(allocation, 'request_management', None):
            return allocation.request_management
        rm = getattr(donation, 'request_management', None)
        if rm:
            return rm
        foodbank_request = getattr(donation, 'foodbank_request', None)
        if foodbank_request:
            linked_rm = getattr(foodbank_request, 'linked_request_management', None)
            if linked_rm:
                return linked_rm
            original_request = getattr(foodbank_request, 'original_request', None)
            if original_request:
                return original_request
            return foodbank_request
        return None

    def _extract_request_type_info(request_obj):
        if not request_obj:
            return None, None
        request_type_value = None
        for attr in ('request_type', 'donation_type', 'request_category'):
            value = getattr(request_obj, attr, None)
            if value:
                request_type_value = value
                break
        if not request_type_value:
            return None, None
        label = None
        if hasattr(request_obj, 'get_request_type_display'):
            try:
                label = request_obj.get_request_type_display()
            except Exception:
                label = None
        if not label and hasattr(request_obj, 'get_donation_type_display'):
            try:
                label = request_obj.get_donation_type_display()
            except Exception:
                label = None
        if not label and hasattr(request_obj, 'get_request_category_display'):
            try:
                label = request_obj.get_request_category_display()
            except Exception:
                label = None
        if not label:
            label = str(request_type_value).replace('_', ' ').title()
        return request_type_value, label

    def _attach_row_request_type(row_obj, donation, allocation=None):
        related_request = _resolve_related_request(donation, allocation)
        req_type, req_label = _extract_request_type_info(related_request)
        row_obj.target_request_type = req_type
        row_obj.target_request_type_label = req_label
        if allocation and getattr(allocation, 'request_management', None):
            row_obj.requested_unit_label = _resolve_unit_label(
                allocation.request_management,
                'unit',
                'custom_unit',
                'units'
            )
        else:
            row_obj.requested_unit_label = _get_requested_unit_for_export(donation) or 'units'
        return row_obj

    rows = []
    for donation in list(donations):
        if has_allocations(donation):
            allocs = sorted(
                donation.allocations.all(),
                key=lambda a: a.allocated_at or timezone.now() - timedelta(days=99999),
                reverse=True
            )
            for alloc in allocs:
                req = getattr(alloc, 'request_management', None)
                status_display = _get_allocation_row_status(req, allocation=alloc)
                status_class = STATUS_CLASS_MAP.get(status_display, 'accepted')
                row_entry = SimpleNamespace(
                    donation=donation,
                    is_allocation=True,
                    allocation=alloc,
                    recipient_display=alloc.recipient.full_name if alloc.recipient else '-',
                    status_display=status_display,
                    status_class=status_class,
                )
                rows.append(_attach_row_request_type(row_entry, donation, alloc))

            if is_direct_stock(donation):
                row_entry = SimpleNamespace(
                    donation=donation,
                    is_allocation=False,
                    allocation=None,
                    recipient_display='Not allocated',
                    status_display='Accepted by Foodbank',
                    status_class='accepted',
                )
                rows.append(_attach_row_request_type(row_entry, donation))
            else:
                display = get_display_status(donation)
                row_entry = SimpleNamespace(
                    donation=donation,
                    is_allocation=False,
                    allocation=None,
                    recipient_display='Not allocated',
                    status_display=display,
                    status_class=STATUS_CLASS_MAP.get(display, 'pending'),
                )
                rows.append(_attach_row_request_type(row_entry, donation))
        else:
            recipient_display = 'Not allocated' if is_direct_stock(donation) else donation.get_recipient_name()
            display = get_display_status(donation)
            row_entry = SimpleNamespace(
                donation=donation,
                is_allocation=False,
                allocation=None,
                recipient_display=recipient_display,
                status_display=display,
                status_class=STATUS_CLASS_MAP.get(display, 'pending'),
            )
            rows.append(_attach_row_request_type(row_entry, donation))

    if type_filter:
        def _row_type_value(row):
            value = getattr(row, 'target_request_type', None)
            if value:
                return str(value).lower()
            fallback = getattr(row.donation, 'donation_category', None)
            if fallback in ('food', 'non_food'):
                return fallback
            return ''
        rows = [row for row in rows if _row_type_value(row) == type_filter]

    declined_labels = {'Declined by Recipient', 'Declined by Foodbank', 'Declined by FB', 'Declined'}
    pending_labels = {'Awaiting Foodbank Review', 'Awaiting FB Review'}
    status_filter_map = {
        'awaiting_foodbank_review': pending_labels,
        'accepted_by_foodbank': {'Accepted by Foodbank'},
        'awaiting_recipient': {
            'Awaiting Recipient',
            'Accepted -> Awaiting Pickup',
            'Accepted â†’ Awaiting Pickup',
            'Accepted Ã¢â€ â€™ Awaiting Pickup',
            'Delivery In Transit',
            'Delivery Scheduled',
            'Accepted',
        },
        'received_by_recipient': {
            'Received By Recipient',
            'Acknowledged by Recipient',
            'Delivered',
            'Fulfilled',
            'Fulfilled-Acknowledged',
            'Fulfilled-Received',
        },
        'declined_by_recipient': {'Declined by Recipient'},
        'declined_by_foodbank': {'Declined by Foodbank', 'Declined by FB'},
        'declined': declined_labels,
        'pending': pending_labels,
    }
    if status_filter:
        if status_filter == 'accepted':
            rows = [
                row for row in rows
                if row.status_display not in declined_labels and row.status_display not in pending_labels
            ]
        elif status_filter in status_filter_map:
            allowed_statuses = status_filter_map[status_filter]
            rows = [row for row in rows if row.status_display in allowed_statuses]
        else:
            normalized_filter = status_filter.replace('_', ' ').lower()
            rows = [row for row in rows if normalized_filter in (row.status_display or '').lower()]

    def row_date(r):
        if r.is_allocation and r.allocation and getattr(r.allocation, 'allocated_at', None):
            return r.allocation.allocated_at
        return r.donation.donated_at or timezone.now() - timedelta(days=99999)

    rows.sort(key=row_date, reverse=True)
    return rows


def _get_export_type_display_for_row(row):
    donation = row.donation
    if donation.donation_type in ('money', 'item', 'subsidized'):
        req_label = getattr(row, 'target_request_type_label', None)
        if req_label:
            return req_label
    return _get_foodbank_export_type_display(donation)


def _get_export_description_for_row(row):
    donation = row.donation
    if row.is_allocation and row.allocation and getattr(row.allocation, 'request_management', None):
        return row.allocation.request_management.description or 'Request'
    if donation.donation_type == 'item':
        fb_req = getattr(donation, 'foodbank_request', None)
        if fb_req:
            if getattr(fb_req, 'original_request', None) and getattr(fb_req.original_request, 'description', None):
                return fb_req.original_request.description
            if getattr(fb_req, 'linked_request_management', None) and getattr(fb_req.linked_request_management, 'description', None):
                return fb_req.linked_request_management.description
            if getattr(fb_req, 'description', None):
                return fb_req.description
        rm = getattr(donation, 'request_management', None)
        if rm and getattr(rm, 'description', None):
            return rm.description
    return _get_description_for_export(donation)


def _get_export_requested_for_row(row):
    donation = row.donation
    if row.is_allocation and row.allocation:
        rm = getattr(row.allocation, 'request_management', None)

        # For money/subsidized allocation rows, requested quantity should reflect
        # the specific allocation used for this donation row (not the full request).
        if donation.donation_type in ('money', 'subsidized') and getattr(row.allocation, 'quantity', None) is not None:
            unit = (
                getattr(row, 'requested_unit_label', None)
                or (_resolve_unit_label(rm, 'unit', 'custom_unit', 'units') if rm else '')
                or _get_requested_unit_for_export(donation)
                or 'units'
            )
            return str(row.allocation.quantity), unit

        if rm:
            qty = rm.quantity if getattr(rm, 'quantity', None) is not None else '-'
            unit = _resolve_unit_label(rm, 'unit', 'custom_unit', 'units')
            return str(qty), unit
    return _get_requested_quantity_value_for_export(donation) or '-', _get_requested_unit_for_export(donation) or '-'


def _get_export_quantity_details_for_row(row):
    donation = row.donation
    if not row.is_allocation or not row.allocation:
        return _get_foodbank_export_quantity_details(donation)

    requested_qty, requested_unit = _get_export_requested_for_row(row)

    if donation.donation_type == 'money':
        alloc_amount = row.allocation.amount if row.allocation.amount is not None else donation.amount
        text = f"KES {float(alloc_amount):.0f}" if alloc_amount is not None else '-'
        alloc_requested_qty = None
        if row.allocation.quantity is not None:
            alloc_requested_qty = row.allocation.quantity
        elif getattr(row.allocation, 'request_management', None) and getattr(row.allocation.request_management, 'quantity', None) is not None:
            alloc_requested_qty = row.allocation.request_management.quantity
        elif requested_qty != '-':
            alloc_requested_qty = requested_qty
        unit_label = getattr(row, 'requested_unit_label', None) or requested_unit or 'units'
        if text != '-' and alloc_requested_qty is not None:
            text = f"{text}\n(for {alloc_requested_qty} {unit_label} requested)"
        return text

    if donation.donation_type == 'subsidized':
        alloc_amount = row.allocation.amount
        if alloc_amount is None:
            alloc_amount = donation.subsidized_price if donation.subsidized_price is not None else donation.amount
        text = f"KES {float(alloc_amount):.0f}" if alloc_amount is not None else '-'
        alloc_requested_qty = None
        if row.allocation.quantity is not None:
            alloc_requested_qty = row.allocation.quantity
        elif getattr(row.allocation, 'request_management', None) and getattr(row.allocation.request_management, 'quantity', None) is not None:
            alloc_requested_qty = row.allocation.request_management.quantity
        elif requested_qty != '-':
            alloc_requested_qty = requested_qty
        unit_label = (
            getattr(row, 'requested_unit_label', None)
            or donation.subsidized_quantity_unit
            or donation.quantity_unit
            or requested_unit
            or 'units'
        )
        if text != '-' and alloc_requested_qty is not None:
            text = f"{text}\n(for {alloc_requested_qty} {unit_label} requested)"
        return text

    alloc_qty = row.allocation.quantity if row.allocation.quantity is not None else '-'
    unit = donation.quantity_unit or 'u'
    text = f"{alloc_qty} {unit}"
    if requested_qty != '-' and requested_unit != '-':
        text = f"{text}\n(of {requested_qty} {requested_unit} requested)"
    return text


def _get_export_decline_reason_for_row(row):
    donation = row.donation
    if row.is_allocation and row.allocation and row.status_display == 'Declined by Recipient':
        req = getattr(row.allocation, 'request_management', None)
        if req:
            reason = (getattr(req, 'decline_message', None) or '').strip()
            if reason:
                return reason
    return _get_foodbank_export_decline_reason(donation)


def _get_category_for_export(donation):
    """Get category for export; for CSR show CSR subcategory e.g. 'CSR Volunteerism'."""
    if donation.donation_type == 'csr' or getattr(donation, 'donation_category', None) == 'csr':
        sub = getattr(donation, 'csr_subcategory', None)
        if sub == 'other' and getattr(donation, 'csr_custom_subcategory', None):
            return f"CSR {donation.csr_custom_subcategory}"
        return f"CSR {donation.get_csr_subcategory_display()}"
    return donation.get_donation_category_display()


def _get_related_request_status(donation):
    related_request = _get_related_request(donation)
    return related_request.status if related_request else None


def _get_allocation_row_status(request_management, allocation=None):
    """Status for a row that represents stock allocated to a recipient (from request).
    Per-allocation declined/acknowledged takes precedence over request-level status."""
    if allocation:
        if getattr(allocation, 'declined_by_recipient', False):
            return "Declined by Recipient"
        if getattr(allocation, 'is_acknowledged', False):
            return "Received By Recipient"
    if not request_management:
        return "Accepted by Foodbank"
    if request_management.status == 'declined':
        return "Declined by Recipient"
    if request_management.status == 'acknowledged' or getattr(request_management, 'acknowledged_by_recipient', False):
        return "Received By Recipient"
    return "Awaiting Recipient"


def get_display_status(donation):
    """Derive a human-readable status for donor/foodbank tables."""
    dtype = getattr(donation, 'donation_type', None)
    is_money_or_subsidized = dtype in ('money', 'subsidized')
    related_request = _get_related_request(donation)

    def _request_declined_by_foodbank(req):
        method = getattr(req, 'was_declined_by_foodbank', None)
        return bool(method()) if callable(method) else False

    request_declined = related_request and getattr(related_request, 'status', None) == 'declined'
    request_declined_by_fb = related_request and _request_declined_by_foodbank(related_request)
    request_declined_by_recipient = bool(getattr(related_request, 'recipient_declined_request', False))

    # Monetary and subsidized: use per-donation state only (this donation, not the whole request).
    # So "Add more" donations show Awaiting Recipient until this donation is acknowledged/declined.
    if is_money_or_subsidized:
        # Per-donation: did the recipient acknowledge or decline THIS donation?
        this_accepted = bool(getattr(donation, 'accepted_by_recipient_id', None))
        this_declined = bool(getattr(donation, 'declined_by_recipient_id', None))

        if this_declined:
            # Guard against stale bad data where a foodbank-declined donation was
            # accidentally tagged with declined_by_recipient.
            if donation.status == 'declined' and (getattr(donation, 'decline_message', '') or '').strip():
                return "Declined by Foodbank"
            return "Declined by Recipient"

        if donation.status == 'declined':
            return "Declined by Foodbank"

        if request_declined and not this_accepted:
            if request_declined_by_fb:
                return "Declined by Foodbank"
            if request_declined_by_recipient:
                return "Declined by Recipient"
            return "Declined"


        if donation.status == 'pending':
            return "Awaiting Foodbank Review"

        if this_accepted:
            return "Received By Recipient"
        # Direct donations (response to a request): show Accepted by Foodbank once FB has accepted
        if getattr(donation, 'foodbank_request_id', None):
            return "Accepted by Foodbank"
        return "Awaiting Recipient"

    # Non–monetary/subsidized (e.g. free goods): use request-level but also per-donation accepted.
    # If THIS donation was already accepted by recipient, keep "Received" even if request later declined or reset.
    linked_request_status = related_request.status if related_request else None
    request_acknowledged = bool(getattr(related_request, 'acknowledged_by_recipient', False))
    request_notes = getattr(related_request, 'additional_notes', '') or ''
    request_confirmed = 'Receipt Confirmed' in request_notes
    request_status_acknowledged = linked_request_status == 'acknowledged'
    # Per-donation accepted = this donation was acknowledged by recipient (e.g. free goods already received)
    this_donation_accepted = bool(getattr(donation, 'accepted_by_recipient_id', None))
    this_donation_declined = bool(getattr(donation, 'declined_by_recipient_id', None))
    received = (
        request_acknowledged or request_confirmed or request_status_acknowledged
        or this_donation_accepted
    )

    if this_donation_declined:
        # Guard against stale bad data where a foodbank-declined donation was
        # accidentally tagged with declined_by_recipient.
        if donation.status == 'declined' and (getattr(donation, 'decline_message', '') or '').strip():
            return "Declined by Foodbank"
        return "Declined by Recipient"

    if donation.status == 'declined':
        return "Declined by Foodbank"

    # Request declined: distinguish who declined
    if linked_request_status == 'declined' and not received:
        if request_declined_by_fb:
            return "Declined by Foodbank"
        if request_declined_by_recipient:
            return "Declined by Recipient"
        return "Declined"

    if linked_request_status == 'declined' and received:
        return "Received By Recipient"

    if donation.status == 'pending':
        return "Awaiting Foodbank Review"

    if donation.status == 'accepted':
        awaiting_request_statuses = {'awaiting_recipient', 'partial', 'assigned'}
        fulfilled_not_acknowledged = linked_request_status in ('fulfilled', 'partial') and not received
        if linked_request_status in awaiting_request_statuses or fulfilled_not_acknowledged:
            if received:
                return "Received By Recipient"
            return "Awaiting Recipient"

        if received:
            return "Received By Recipient"

        if (
            donation.donation_type in ['subsidized', 'csr', 'other']
            and getattr(donation, 'foodbank_request_id', None) is None
            and donation.accepted_by_recipient_id is None
        ):
            return "Awaiting Recipient"

        if hasattr(donation, 'foodbank_request') and donation.foodbank_request:
            return "Accepted by Foodbank"
        if donation.delivery_status == 'delivered':
            return "Delivered"
        if donation.delivery_status == 'in_transit':
            return "Delivery In Transit"
        if donation.delivery_status == 'scheduled':
            return "Delivery Scheduled"
        return "Accepted → Awaiting Pickup"

    if donation.status == 'fulfilled':
        if received:
            return "Received By Recipient"
        return "Awaiting Recipient"

    if donation.status == 'partial':
        if received:
            return "Received By Recipient"
        return "Awaiting Recipient"

    return donation.status.replace('_', ' ').title()


@login_required
@donor_required
def donor_donations_list(request):
    """View for donors to see all their donations with filters"""
    donations = Donation.objects.filter(
        donor=request.user
    ).filter(
        unspecified_management__isnull=True
    ).select_related(
        'foodbank', 'foodbank__user', 'foodbank_request', 'accepted_by_recipient',
        'foodbank_request__original_request__recipient',
        'foodbank_request__linked_request_management',
        'request_management', 'request_management__recipient'
    ).prefetch_related('allocations__recipient').order_by('-donated_at')

    direct_only = request.GET.get('direct', '').strip()
    if direct_only:
        donations = donations.filter(
            foodbank_request__isnull=False,
            foodbank_request__original_request__isnull=True
        )
    else:
        donations = donations.exclude(
            Q(foodbank_request__isnull=False) & Q(foodbank_request__original_request__isnull=True)
        ).filter(
            Q(request_management__isnull=False) |
            Q(foodbank_request__original_request__isnull=False) |
            Q(foodbank_request__linked_request_management__isnull=False)
        ).distinct()

    is_direct_only = bool(direct_only)
    include_recipient_columns = not is_direct_only
    
    # Get filter parameters (aligned with table: Type, Category, Food Bank, Status, Delivery, Date, Qty/Amount, Search)
    type_filter = request.GET.get('type', '')
    category_filter = request.GET.get('category', '')
    type_filter, category_filter = _normalize_donor_type_category_filters(type_filter, category_filter)
    foodbank_filter = request.GET.get('foodbank', '')
    status_filter = request.GET.get('status', '')
    delivery_status_raw = request.GET.get('delivery_status', '').strip()
    delivery_filter = delivery_status_raw.lower()
    date_filter = request.GET.get('date_range', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')
    search = request.GET.get('search', '').strip()
    
    # Apply filters
    if category_filter:
        donations = donations.filter(donation_type=category_filter)
    
    if foodbank_filter:
        donations = donations.filter(foodbank__id=foodbank_filter)
    
    if status_filter:
        if status_filter == 'declined':
            if is_direct_only:
                # Direct donor view: "Declined" should only show foodbank declines.
                recipient_decline_q = (
                    Q(declined_by_recipient__isnull=False) |
                    Q(allocations__declined_by_recipient=True) |
                    Q(request_management__status='declined', request_management__updated_by__user_type='RECIPIENT') |
                    Q(
                        foodbank_request__linked_request_management__status='declined',
                        foodbank_request__linked_request_management__updated_by__user_type='RECIPIENT'
                    ) |
                    Q(
                        foodbank_request__original_request__status='declined',
                        foodbank_request__original_request__updated_by__user_type='RECIPIENT'
                    )
                )
                donations = donations.filter(status='declined').exclude(
                    recipient_decline_q
                ).distinct()
            else:
                donations = donations.filter(
                    Q(status='declined') |
                    Q(declined_by_recipient__isnull=False) |
                    Q(allocations__declined_by_recipient=True)
                ).distinct()
        elif status_filter == 'accepted':
            donations = donations.filter(status='accepted').exclude(
                Q(declined_by_recipient__isnull=False) |
                Q(allocations__declined_by_recipient=True)
            ).distinct()
        else:
            donations = donations.filter(status=status_filter)
    
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'

    if delivery_filter:
        if delivery_filter == 'pickup':
            donations = donations.filter(delivery_method='pickup')
        elif delivery_filter == 'delivery':
            donations = donations.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            donations = donations.filter(delivery_status=delivery_filter)
    
    if search:
        donations = donations.filter(
            Q(item_name__icontains=search) |
            Q(subsidized_product_type__icontains=search) |
            Q(other_description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(message__icontains=search) |
            Q(csr_description__icontains=search) |
            Q(foodbank_request__title__icontains=search) |
            Q(foodbank_request__description__icontains=search) |
            Q(foodbank_request__original_request__description__icontains=search) |
            Q(foodbank_request__original_request__additional_notes__icontains=search) |
            Q(request_management__description__icontains=search) |
            Q(request_management__additional_notes__icontains=search)
        )

    # Type filter (Food/Non-Food) is derived from linked request semantics.
    if type_filter:
        filtered_ids = []
        for donation in donations:
            type_display = (_get_foodbank_export_type_display(donation) or '').strip().lower()
            normalized_type = type_display.replace('-', '_').replace(' ', '_')
            if normalized_type == type_filter:
                filtered_ids.append(donation.id)
        donations = donations.filter(id__in=filtered_ids) if filtered_ids else donations.none()
    
    # Filter by date range
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    if date_filter == 'custom' and (date_from or date_to):
        # Custom date range
        if date_from:
            try:
                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
                donations = donations.filter(donated_at__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                donations = donations.filter(donated_at__lte=to_date)
            except ValueError:
                pass
    elif date_filter != 'all':
        now = timezone.now()
        if date_filter == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter == 'week':
            start_date = now - timedelta(days=7)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter == 'month':
            start_date = now - timedelta(days=30)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter == '3months':
            start_date = now - timedelta(days=90)
            donations = donations.filter(donated_at__gte=start_date)
    
    # Filter by quantity range (for item donations)
    if quantity_filter != 'all':
        if quantity_filter == 'small':
            donations = donations.filter(donation_type='item', quantity__lte=50)
        elif quantity_filter == 'medium':
            donations = donations.filter(donation_type='item', quantity__gt=50, quantity__lte=200)
        elif quantity_filter == 'large':
            donations = donations.filter(donation_type='item', quantity__gt=200)
    
    # Filter by amount range (for monetary donations)
    if amount_filter != 'all':
        if amount_filter == 'small':
            donations = donations.filter(donation_type='money', amount__lte=5000)
        elif amount_filter == 'medium':
            donations = donations.filter(donation_type='money', amount__gt=5000, amount__lte=20000)
        elif amount_filter == 'large':
            donations = donations.filter(donation_type='money', amount__gt=20000)
    
    # Get statistics
    total_donations = donations.count()
    total_items = donations.filter(donation_type='item').aggregate(
        total=Sum('quantity')
    )['total'] or 0
    total_money = donations.filter(donation_type='money').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # Get unique foodbanks for filter dropdown (only active ones)
    foodbanks = FoodBankProfile.objects.filter(
        id__in=donations.values_list('foodbank_id', flat=True).distinct(),
        user__is_active=True
    )

    status_class_map = {
        'Accepted': 'accepted',
        'Pending': 'pending',
        'Received By Recipient': 'fulfilled',
        'Partially Fulfilled': 'partial',
        'Partially Fulfilled Acknowledged': 'partial',
        'Declined': 'declined',
        'Declined by FB': 'declined',
        'Declined by Recipient': 'declined',
    }

    # Pagination
    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    donation_list = list(page_obj.object_list)
    latest_non_decline_notes = {}
    latest_decline_notes = {}
    accepted_notes = {}

    if include_recipient_columns and donation_list:
        donation_ids = [donation.id for donation in donation_list]
        accepted_recipient_by_donation = {
            donation.id: donation.accepted_by_recipient_id
            for donation in donation_list
            if getattr(donation, 'accepted_by_recipient_id', None)
        }

        responses = (
            DonationResponse.objects
            .filter(donation_id__in=donation_ids)
            .exclude(notes__isnull=True)
            .exclude(notes__exact='')
            .order_by('-responded_at')
        )

        for response in responses:
            response_note = (response.notes or '').strip()
            if not response_note:
                continue

            response_type = (response.response_type or '').strip().lower()
            donation_id = response.donation_id

            if response_type == 'declined':
                if donation_id not in latest_decline_notes:
                    latest_decline_notes[donation_id] = response_note
                continue

            if donation_id not in latest_non_decline_notes:
                latest_non_decline_notes[donation_id] = response_note

            if (
                response_type == 'accepted'
                and accepted_recipient_by_donation.get(donation_id) == response.recipient_id
                and donation_id not in accepted_notes
            ):
                accepted_notes[donation_id] = response_note

    for donation in donation_list:
        donation.recipient_display = donation.get_recipient_name()
        donation.status_display = get_display_status(donation)
        donation.status_class = status_class_map.get(donation.status_display, 'pending')
        donation.donor_type_display = _get_foodbank_export_type_display(donation)
        donation.donor_category_display = _get_foodbank_export_category_display(donation)
        donation.requested_unit_label = _get_requested_unit_for_export(donation) or 'units'
        _type_norm = (donation.donor_type_display or '').strip().lower()
        donation.donor_type_value = 'food' if _type_norm == 'food' else 'non_food' if _type_norm in ('non-food', 'non food', 'non_food') else ''
        donation.latest_note = donation.message
        rm = getattr(donation, 'request_management', None)
        fb_req = getattr(donation, 'foodbank_request', None)
        orig_req = getattr(fb_req, 'original_request', None) if fb_req else None

        fallback_recipient_note = (
            (getattr(rm, 'additional_notes', None) or '').strip()
            or (getattr(orig_req, 'additional_notes', None) or '').strip()
            or ''
        )
        fallback_recipient_decline_note = (
            (getattr(rm, 'decline_message', None) or '').strip()
            if rm and getattr(rm, 'recipient_declined_request', False)
            else ''
        ) or (
            (getattr(orig_req, 'decline_message', None) or '').strip()
            if orig_req and getattr(orig_req, 'recipient_declined_request', False)
            else ''
        ) or (
            (getattr(donation, 'decline_message', None) or '').strip()
            if getattr(donation, 'declined_by_recipient_id', None)
            else ''
        )

        if include_recipient_columns:
            if getattr(donation, 'accepted_by_recipient_id', None):
                effective_recipient_note = (
                    accepted_notes.get(donation.id)
                    or latest_non_decline_notes.get(donation.id)
                    or fallback_recipient_note
                )
                effective_recipient_decline_note = ''
            else:
                effective_recipient_note = (
                    latest_non_decline_notes.get(donation.id)
                    or fallback_recipient_note
                )
                effective_recipient_decline_note = (
                    latest_decline_notes.get(donation.id)
                    or fallback_recipient_decline_note
                )
        else:
            effective_recipient_note = ''
            effective_recipient_decline_note = ''

        donation.latest_note_recipient = effective_recipient_note
        donation.latest_recipient_note = effective_recipient_note
        donation.effective_recipient_note = effective_recipient_note
        donation.effective_recipient_decline_note = effective_recipient_decline_note

        has_any_note_text = any([
            bool((getattr(donation, 'message', None) or '').strip()),
            bool((getattr(donation, 'decline_message', None) or '').strip()),
            bool((effective_recipient_note or '').strip()),
            bool((effective_recipient_decline_note or '').strip()),
            bool((getattr(rm, 'additional_notes', None) or '').strip()) if rm else False,
            bool((getattr(rm, 'foodbank_response_note', None) or '').strip()) if rm else False,
            bool((getattr(rm, 'decline_message', None) or '').strip()) if rm else False,
            bool((getattr(fb_req, 'foodbank_note', None) or '').strip()) if fb_req else False,
            bool((getattr(orig_req, 'additional_notes', None) or '').strip()) if orig_req else False,
            bool((getattr(orig_req, 'decline_message', None) or '').strip()) if orig_req else False,
        ])
        donation.has_notes_for_modal = bool(rm or fb_req or has_any_note_text)

    context = {
        'donations': page_obj,
        'total_donations': total_donations,
        'total_items': total_items,
        'total_money': total_money,
        'foodbanks': foodbanks,
        'selected_type': type_filter,
        'selected_category': category_filter,
        'selected_foodbank': foodbank_filter,
        'selected_status': status_filter,
        'selected_delivery_status': delivery_filter,
        'date_filter': date_filter,
        'date_from': date_from,
        'date_to': date_to,
        'quantity_filter': quantity_filter,
        'amount_filter': amount_filter,
        'search_query': search,
        'direct_filter': request.GET.get('direct', ''),
    }
    
    return render(request, 'authentication/donor_donations_list.html', context)


@login_required
@foodbank_required
def foodbank_donations_list(request):
    """Foodbank view for received donations using legacy layout"""
    foodbank_profile = request.user.foodbank_profile
    donations = Donation.objects.filter(foodbank=foodbank_profile).order_by('-donated_at')

    donation_type = request.GET.get('type', '').strip()
    if donation_type:
        donations = donations.filter(donation_type=donation_type)

    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if date_from:
        donations = donations.filter(donated_at__date__gte=date_from)
    if date_to:
        donations = donations.filter(donated_at__date__lte=date_to)

    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    donation_list = list(page_obj.object_list)
    donation_ids = [donation.id for donation in donation_list]
    latest_notes = {}

    if donation_ids:
        responses = (
            DonationResponse.objects
            .filter(donation_id__in=donation_ids)
            .exclude(notes__isnull=True)
            .exclude(notes__exact='')
            .order_by('-responded_at')
        )

        for response in responses:
            if response.donation_id not in latest_notes:
                latest_notes[response.donation_id] = response.notes

    for donation in donation_list:
        donation.latest_recipient_note = latest_notes.get(donation.id)
        donation.status_display = get_display_status(donation)
        donation.status_class = STATUS_CLASS_MAP.get(donation.status_display, 'pending')

    context = {
        'donations': page_obj,
        'page_obj': page_obj,
        'total_donations': paginator.count,
        'total_money': Donation.objects.filter(
            foodbank=foodbank_profile,
            donation_type='money'
        ).aggregate(total=Sum('amount'))['total'] or 0,
        'total_items': Donation.objects.filter(
            foodbank=foodbank_profile,
            donation_type='item'
        ).count(),
        'total_subsidized': Donation.objects.filter(
            foodbank=foodbank_profile,
            donation_type='subsidized'
        ).aggregate(total=Sum('subsidized_price'))['total'] or 0,
    }

    return render(request, 'authentication/foodbank_donations.html', context)


@login_required
@donor_required
def donor_donations_export(request, format):
    """Export donor's donations in PDF, CSV, or Excel. Mirrors all filters from donor_donations_list."""
    donations = Donation.objects.filter(donor=request.user).filter(
        unspecified_management__isnull=True
    ).select_related(
        'foodbank', 'foodbank__user', 'foodbank_request', 'accepted_by_recipient',
        'foodbank_request__original_request__recipient',
        'foodbank_request__linked_request_management',
        'request_management', 'request_management__recipient',
        'unspecified_management',
    ).prefetch_related('allocations__recipient').order_by('-donated_at')

    # Direct vs specified: same logic as donor_donations_list
    direct_only = request.GET.get('direct', '').strip()
    if direct_only:
        donations = donations.filter(
            foodbank_request__isnull=False,
            foodbank_request__original_request__isnull=True
        )
    else:
        donations = donations.exclude(
            Q(foodbank_request__isnull=False) & Q(foodbank_request__original_request__isnull=True)
        ).filter(
            Q(request_management__isnull=False) |
            Q(foodbank_request__original_request__isnull=False) |
            Q(foodbank_request__linked_request_management__isnull=False)
        ).distinct()

    is_direct_only = bool(direct_only)
    include_recipient_columns = not is_direct_only

    type_filter = request.GET.get('type', '')
    category_filter = request.GET.get('category', '')
    type_filter, category_filter = _normalize_donor_type_category_filters(type_filter, category_filter)
    foodbank_filter = request.GET.get('foodbank', '')
    status_filter = request.GET.get('status', '')
    delivery_status_raw = request.GET.get('delivery_status', '').strip()
    delivery_filter = delivery_status_raw.lower()
    date_filter = request.GET.get('date_range', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')
    search = request.GET.get('search', '')

    if category_filter:
        donations = donations.filter(donation_type=category_filter)

    if foodbank_filter:
        donations = donations.filter(foodbank__id=foodbank_filter)

    if status_filter:
        if status_filter == 'declined':
            if is_direct_only:
                recipient_decline_q = (
                    Q(declined_by_recipient__isnull=False) |
                    Q(allocations__declined_by_recipient=True) |
                    Q(request_management__status='declined', request_management__updated_by__user_type='RECIPIENT') |
                    Q(
                        foodbank_request__linked_request_management__status='declined',
                        foodbank_request__linked_request_management__updated_by__user_type='RECIPIENT'
                    ) |
                    Q(
                        foodbank_request__original_request__status='declined',
                        foodbank_request__original_request__updated_by__user_type='RECIPIENT'
                    )
                )
                donations = donations.filter(status='declined').exclude(
                    recipient_decline_q
                ).distinct()
            else:
                donations = donations.filter(
                    Q(status='declined') |
                    Q(declined_by_recipient__isnull=False) |
                    Q(allocations__declined_by_recipient=True)
                ).distinct()
        else:
            donations = donations.filter(status=status_filter)

    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'

    if delivery_filter:
        if delivery_filter == 'pickup':
            donations = donations.filter(delivery_method='pickup')
        elif delivery_filter == 'delivery':
            donations = donations.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            donations = donations.filter(delivery_status=delivery_filter)

    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if date_filter == 'custom' and (date_from or date_to):
        if date_from:
            try:
                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
                donations = donations.filter(donated_at__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                donations = donations.filter(donated_at__lte=to_date)
            except ValueError:
                pass
    elif date_filter != 'all':
        now = timezone.now()
        if date_filter == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter == 'week':
            start_date = now - timedelta(days=7)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter == 'month':
            start_date = now - timedelta(days=30)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter == '3months':
            start_date = now - timedelta(days=90)
            donations = donations.filter(donated_at__gte=start_date)

    if quantity_filter != 'all':
        if quantity_filter == 'small':
            donations = donations.filter(donation_type='item', quantity__lte=50)
        elif quantity_filter == 'medium':
            donations = donations.filter(donation_type='item', quantity__gt=50, quantity__lte=200)
        elif quantity_filter == 'large':
            donations = donations.filter(donation_type='item', quantity__gt=200)

    if amount_filter != 'all':
        if amount_filter == 'small':
            donations = donations.filter(donation_type='money', amount__lte=5000)
        elif amount_filter == 'medium':
            donations = donations.filter(donation_type='money', amount__gt=5000, amount__lte=20000)
        elif amount_filter == 'large':
            donations = donations.filter(donation_type='money', amount__gt=20000)

    if search:
        donations = donations.filter(
            Q(item_name__icontains=search) |
            Q(subsidized_product_type__icontains=search) |
            Q(other_description__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(message__icontains=search) |
            Q(csr_description__icontains=search) |
            Q(foodbank_request__title__icontains=search) |
            Q(foodbank_request__description__icontains=search) |
            Q(foodbank_request__original_request__description__icontains=search) |
            Q(foodbank_request__original_request__additional_notes__icontains=search) |
            Q(request_management__description__icontains=search) |
            Q(request_management__additional_notes__icontains=search)
        )

    # Type filter (Food/Non-Food) derived from request semantics.
    if type_filter:
        filtered_ids = []
        for donation in donations:
            type_display = (_get_foodbank_export_type_display(donation) or '').strip().lower()
            normalized_type = type_display.replace('-', '_').replace(' ', '_')
            if normalized_type == type_filter:
                filtered_ids.append(donation.id)
        donations = donations.filter(id__in=filtered_ids) if filtered_ids else donations.none()

    donations_data = list(donations)

    latest_recipient_notes = {}
    latest_recipient_declines = {}
    if include_recipient_columns:
        # Fetch latest recipient notes for all donations from all possible sources
        donation_ids = [d.id for d in donations_data]

        # Source 1: DonationResponse.notes (subsidized donations)
        if donation_ids:
            responses = (
                DonationResponse.objects
                .filter(donation_id__in=donation_ids)
                .exclude(notes__isnull=True)
                .exclude(notes__exact='')
                .order_by('-responded_at')
            )
            for resp in responses:
                if resp.donation_id not in latest_recipient_notes:
                    latest_recipient_notes[resp.donation_id] = resp.notes

            decline_responses = (
                DonationResponse.objects
                .filter(donation_id__in=donation_ids, response_type='declined')
                .exclude(notes__isnull=True)
                .exclude(notes__exact='')
                .order_by('-responded_at')
            )
            for resp in decline_responses:
                if resp.donation_id not in latest_recipient_declines:
                    latest_recipient_declines[resp.donation_id] = resp.notes

        # Source 2: request_management / foodbank_request / unspecified_management
        for d in donations_data:
            if d.id in latest_recipient_notes:
                continue
            try:
                # Direct/request-based: request_management.additional_notes
                rm = d.request_management
                if rm and rm.additional_notes:
                    latest_recipient_notes[d.id] = rm.additional_notes
                    continue
            except Exception:
                pass
            try:
                # Direct/request-based: foodbank_request.original_request.additional_notes
                fbr = d.foodbank_request
                if fbr and fbr.original_request and fbr.original_request.additional_notes:
                    latest_recipient_notes[d.id] = fbr.original_request.additional_notes
                    continue
            except Exception:
                pass
            try:
                # Unspecified: unspecified_management.recipient_notes
                um = d.unspecified_management
                if um and um.recipient_notes:
                    latest_recipient_notes[d.id] = um.recipient_notes
            except Exception:
                pass

        for d in donations_data:
            if d.id in latest_recipient_declines:
                continue
            try:
                um = d.unspecified_management
                if um and um.recipient_decline_reason:
                    latest_recipient_declines[d.id] = um.recipient_decline_reason
                    continue
            except Exception:
                pass
            try:
                rm = d.request_management
                if rm and getattr(rm, 'recipient_declined_request', False) and rm.decline_message:
                    latest_recipient_declines[d.id] = rm.decline_message
                    continue
            except Exception:
                pass
            try:
                fbr = d.foodbank_request
                if fbr and fbr.original_request and getattr(fbr.original_request, 'recipient_declined_request', False) and fbr.original_request.decline_message:
                    latest_recipient_declines[d.id] = fbr.original_request.decline_message
            except Exception:
                pass

    if format.lower() == 'pdf':
        return donor_donations_export_pdf(request, donations_data, latest_recipient_notes, latest_recipient_declines, include_recipient_columns=include_recipient_columns)
    elif format.lower() == 'csv':
        return donor_donations_export_csv(request, donations_data, latest_recipient_notes, latest_recipient_declines, include_recipient_columns=include_recipient_columns)
    elif format.lower() == 'excel':
        return donor_donations_export_excel(request, donations_data, latest_recipient_notes, latest_recipient_declines, include_recipient_columns=include_recipient_columns)
    else:
        messages.error(request, 'Invalid export format.')
        return redirect('donor_donations_list')


def donor_donations_export_pdf(request, donations_data, latest_recipient_notes=None, latest_recipient_declines=None, include_recipient_columns=True):
    """Generate PDF report for donor's donations using shared FoodBankHub branding."""
    from reportlab.platypus import Paragraph, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from .report_utils import (
        get_report_styles,
        build_report_header,
        get_branded_table_style,
        build_report_summary,
        build_pdf_document,
        collect_active_filters,
        make_full_width_table,
    )

    if latest_recipient_notes is None:
        latest_recipient_notes = {}
    if latest_recipient_declines is None:
        latest_recipient_declines = {}

    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    styles_dict = get_report_styles()
    report_pagesize = landscape(A3)
    filter_keys = [
        ('type', 'Type'), ('category', 'Category'), ('foodbank', 'Food Bank'),
        ('status', 'Status'), ('delivery_status', 'Delivery'),
        ('date_range', 'Date Range'), ('date_from', 'Date From'), ('date_to', 'Date To'),
        ('quantity', 'Quantity'), ('amount', 'Amount'), ('search', 'Search'),
    ]
    active_filters = collect_active_filters(request, filter_keys)
    # Show human-readable filter labels in report header.
    type_display_map = {'food': 'Food', 'non_food': 'Non-Food'}
    category_display_map = {'item': 'Free Goods', 'money': 'Monetary', 'subsidized': 'Subsidized', 'csr': 'CSR', 'other': 'Other'}
    if active_filters.get('Type') and active_filters['Type'] in type_display_map:
        active_filters['Type'] = type_display_map[active_filters['Type']]
    if active_filters.get('Category') and active_filters['Category'] in category_display_map:
        active_filters['Category'] = category_display_map[active_filters['Category']]

    build_report_header(
        elements := [],
        report_title="My Donations Report",
        generated_for=donor_name,
        total_records=len(donations_data),
        active_filters=active_filters,
        styles_dict=styles_dict,
    )

    if not donations_data:
        elements.append(Paragraph("No donations found matching the current filters.", styles_dict['normal']))
    else:
        wrap = styles_dict['wrap']
        header_wrap = ParagraphStyle(
            'donor_report_header_wrap',
            parent=wrap,
            textColor=colors.white,
        )
        if include_recipient_columns:
            data = [[
                'S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity Details', 'Delivery', 'Location',
                'Donor Note',
                Paragraph('Foodbank<br/>Note', header_wrap),
                Paragraph('Recipient<br/>Note', header_wrap),
                Paragraph('Decline<br/>Reason', header_wrap),
                'Status'
            ]]
        else:
            data = [[
                'S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity Details', 'Delivery', 'Location',
                'Donor Note', 'Foodbank Note', 'Status'
            ]]

        for idx, donation in enumerate(donations_data, start=1):
            date_str = donation.donated_at.strftime('%b %d, %Y')
            type_display = _get_foodbank_export_type_display(donation)
            category_display = _get_foodbank_export_category_display(donation)

            description = _get_description_for_export(donation)

            # Qty/Amount: shared formatter (includes "(for ...)" and "(of ... requested)" for money/subsidized)
            quantity_amount = _get_donor_qty_amount_display(donation)

            delivery = donation.get_delivery_method_display() if donation.delivery_method else '-'
            location = donation.foodbank.address or 'No location specified'

            donor_note_text = (donation.message or '')[:200]
            donor_note = Paragraph(donor_note_text, wrap)
            foodbank_note_text = (_get_donor_foodbank_note_for_export(donation) or 'No note')[:200]
            foodbank_note = Paragraph(foodbank_note_text, wrap)

            recipient_note_text = (latest_recipient_notes.get(donation.id) or '')[:200]
            recipient_note = Paragraph(recipient_note_text, wrap)
            decline_reason_text = (_get_donor_decline_reason_for_export(donation, latest_recipient_declines) or 'No note')[:200]
            decline_reason = Paragraph(decline_reason_text, wrap)

            status = get_display_status(donation)
            # Description: title on top of description (newlines as <br/> for PDF)
            description_para = Paragraph(description.replace('\n', '<br/>'), wrap)

            if include_recipient_columns:
                data.append([
                    str(idx), date_str, Paragraph(type_display, wrap), Paragraph(category_display, wrap),
                    description_para,
                    Paragraph((donation.foodbank.foodbank_name or '')[:50], wrap),
                    Paragraph(quantity_amount.replace('\n', '<br/>'), wrap), delivery, Paragraph(location[:60], wrap), donor_note, foodbank_note, recipient_note, decline_reason, Paragraph(status, wrap),
                ])
            else:
                data.append([
                    str(idx), date_str, Paragraph(type_display, wrap), Paragraph(category_display, wrap),
                    description_para,
                    Paragraph((donation.foodbank.foodbank_name or '')[:50], wrap),
                    Paragraph(quantity_amount.replace('\n', '<br/>'), wrap), delivery, Paragraph(location[:60], wrap), donor_note, foodbank_note, Paragraph(status, wrap),
                ])

        if include_recipient_columns:
            _col_weights = [0.031, 0.062, 0.055, 0.069, 0.091, 0.078, 0.102, 0.052, 0.069, 0.074, 0.074, 0.089, 0.089, 0.045]
        else:
            _col_weights = [0.032, 0.067, 0.062, 0.080, 0.110, 0.094, 0.138, 0.073, 0.100, 0.094, 0.094, 0.066]
        table = make_full_width_table(data, repeat_rows=1, col_weights=_col_weights, pagesize=report_pagesize)
        table.setStyle(get_branded_table_style(len(data)))
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ]))
        elements.append(table)

    build_report_summary(elements, [("Total Donations", len(donations_data))], styles_dict=styles_dict)
    return build_pdf_document(elements, "donations", donor_name, pagesize=report_pagesize)


def donor_donations_export_csv(request, donations_data, latest_recipient_notes=None, latest_recipient_declines=None, include_recipient_columns=True):
    """Generate CSV report for donor's donations"""
    import csv
    from django.http import HttpResponse
    import datetime

    if latest_recipient_notes is None:
        latest_recipient_notes = {}
    if latest_recipient_declines is None:
        latest_recipient_declines = {}

    response = HttpResponse(content_type='text/csv')
    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    response['Content-Disposition'] = f'attachment; filename="{donor_name}_donations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Donor', donor_name])
    writer.writerow(['Report Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    # Headers - add Quantity Details to mirror table Qty/Amount content
    if include_recipient_columns:
        writer.writerow(['S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity Details', 'Quantity', 'Unit', 'Amount (KSH)', 'Delivery', 'Location', 'Donor Note', 'Foodbank Note', 'Recipient Note', 'Decline Reason', 'Status'])
    else:
        writer.writerow(['S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity Details', 'Quantity', 'Unit', 'Amount (KSH)', 'Delivery', 'Location', 'Donor Note', 'Foodbank Note', 'Status'])

    # Data
    for idx, donation in enumerate(donations_data, start=1):
        # Format date
        date_str = donation.donated_at.strftime('%Y-%m-%d %H:%M')

        # Format type/category to match donor table
        type_display = _get_foodbank_export_type_display(donation)
        category_display = _get_foodbank_export_category_display(donation)

        # Format description (monetary: use linked request description)
        description = _get_description_for_export(donation)
        # CSV importers can mis-handle embedded newlines and shift columns.
        # Keep full details but flatten to one line for stable CSV rendering.
        quantity_details = _get_donor_qty_amount_display(donation).replace('\r\n', '\n').replace('\n', ' | ')

        # Split quantity/unit/amount (monetary: use quantity_needed/unit from foodbank_request when present, like dashboard "KES 500 (for 200 kg)")
        if donation.donation_type == 'item':
            quantity_value = donation.quantity or ''
            unit_value = donation.quantity_unit or 'units'
            amount_value = ''
        elif donation.donation_type == 'money':
            fb_req = getattr(donation, 'foodbank_request', None)
            quantity_value = (fb_req.quantity_needed if fb_req and getattr(fb_req, 'quantity_needed', None) is not None else '') or ''
            unit_value = (fb_req.quantity_unit if fb_req and getattr(fb_req, 'quantity_unit', None) else '') or ''
            amount_value = int(round(float(donation.amount), 0)) if donation.amount is not None else ''
        elif donation.donation_type == 'subsidized':
            quantity_value = donation.subsidized_quantity or ''
            unit_value = donation.subsidized_quantity_unit or ''
            amount_value = int(round(float(donation.subsidized_price), 0)) if donation.subsidized_price is not None else ''
        else:
            quantity_value = ''
            unit_value = ''
            amount_value = ''

        # Delivery method
        delivery = donation.get_delivery_method_display() if donation.delivery_method else '-'
        
        # Location
        location = donation.foodbank.address or 'No location specified'

        recipient_note = ''
        decline_reason = 'No note'
        if include_recipient_columns:
            recipient_note = (latest_recipient_notes.get(donation.id) or '').replace('\n', ' | ').replace('\r', '')
            decline_reason = (_get_donor_decline_reason_for_export(donation, latest_recipient_declines) or 'No note').replace('\n', ' | ').replace('\r', '')

        donor_note = (donation.message or '').replace('\n', ' | ').replace('\r', '')
        foodbank_note = (_get_donor_foodbank_note_for_export(donation) or 'No note').replace('\n', ' | ').replace('\r', '')
        # Status: same as table (get_display_status)
        status = get_display_status(donation)

        if include_recipient_columns:
            writer.writerow([
                idx,
                date_str,
                type_display,
                category_display,
                description,
                donation.foodbank.foodbank_name,
                quantity_details,
                quantity_value,
                unit_value,
                amount_value,
                delivery,
                location,
                donor_note,
                foodbank_note,
                recipient_note,
                decline_reason,
                status
            ])
        else:
            writer.writerow([
                idx,
                date_str,
                type_display,
                category_display,
                description,
                donation.foodbank.foodbank_name,
                quantity_details,
                quantity_value,
                unit_value,
                amount_value,
                delivery,
                location,
                donor_note,
                foodbank_note,
                status
            ])

    return response


def donor_donations_export_excel(request, donations_data, latest_recipient_notes=None, latest_recipient_declines=None, include_recipient_columns=True):
    """Generate Excel report for donor's donations"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    if latest_recipient_notes is None:
        latest_recipient_notes = {}
    if latest_recipient_declines is None:
        latest_recipient_declines = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "My Donations"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title and info
    donor_name = request.user.donor_profile.full_name if request.user.donor_profile else request.user.email
    headers = ['S/No', 'Date', 'Type', 'Category', 'Description', 'Food Bank', 'Quantity Details', 'Quantity', 'Unit', 'Amount (KSH)', 'Delivery Method', 'Location', 'Donor Note', 'Foodbank Note']
    if include_recipient_columns:
        headers.extend(['Recipient Note', 'Decline Reason'])
    headers.append('Donation Status')

    max_col_letter = get_column_letter(len(headers))

    ws.merge_cells(f'A1:{max_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = f"{donor_name} - My Donations Report"
    title_cell.font = Font(bold=True, size=14, color="10b981")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f'A2:{max_col_letter}2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Donations: {len(donations_data)}"
    info_cell.alignment = Alignment(horizontal="center")

    # Headers - matching table columns with separated Quantity, Unit, Amount
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for idx, donation in enumerate(donations_data, 1):
        row_num = idx + 4

        # Format date
        date_str = donation.donated_at.strftime('%Y-%m-%d %H:%M')

        # Format type/category to match donor table
        type_display = _get_foodbank_export_type_display(donation)
        category_display = _get_foodbank_export_category_display(donation)

        # Format description (monetary: use linked request description)
        description = _get_description_for_export(donation)
        quantity_details = _get_donor_qty_amount_display(donation)

        # Format quantity, unit, and amount separately (monetary: use quantity_needed/unit from foodbank_request when present, like dashboard "KES 500 (for 200 kg)")
        if donation.donation_type == 'item':
            quantity = donation.quantity or '-'
            unit = donation.quantity_unit or 'units'
            amount = '-'
        elif donation.donation_type == 'money':
            fb_req = getattr(donation, 'foodbank_request', None)
            quantity = (fb_req.quantity_needed if fb_req and getattr(fb_req, 'quantity_needed', None) is not None else None) or '-'
            unit = (fb_req.quantity_unit if fb_req and getattr(fb_req, 'quantity_unit', None) else None) or '-'
            amount = int(round(float(donation.amount), 0)) if donation.amount is not None else '-'
        elif donation.donation_type == 'subsidized':
            quantity = donation.subsidized_quantity or '-'
            unit = donation.subsidized_quantity_unit or '-'
            amount = int(round(float(donation.subsidized_price), 0)) if donation.subsidized_price is not None else '-'
        else:
            quantity = '-'
            unit = '-'
            amount = '-'

        # Delivery method
        delivery = donation.get_delivery_method_display() if donation.delivery_method else '-'

        # Location
        location = donation.foodbank.address or 'No location specified'

        # Donor note
        donor_note = donation.message or '-'
        foodbank_note = _get_donor_foodbank_note_for_export(donation) or 'No note'

        # Status: same as table (get_display_status)
        status = get_display_status(donation)

        # Write row data; description keeps newlines so title appears on top of description in cell
        desc_cell = ws.cell(row=row_num, column=5, value=description)
        desc_cell.border = border
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=date_str).border = border
        ws.cell(row=row_num, column=3, value=type_display).border = border
        ws.cell(row=row_num, column=4, value=category_display).border = border
        ws.cell(row=row_num, column=6, value=donation.foodbank.foodbank_name).border = border
        qty_details_cell = ws.cell(row=row_num, column=7, value=quantity_details)
        qty_details_cell.border = border
        qty_details_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=8, value=quantity).border = border
        ws.cell(row=row_num, column=9, value=unit).border = border
        ws.cell(row=row_num, column=10, value=amount).border = border
        ws.cell(row=row_num, column=11, value=delivery).border = border
        ws.cell(row=row_num, column=12, value=location).border = border
        ws.cell(row=row_num, column=13, value=donor_note).border = border
        fb_note_cell = ws.cell(row=row_num, column=14, value=foodbank_note)
        fb_note_cell.border = border
        fb_note_cell.alignment = Alignment(wrap_text=True, vertical="top")

        next_col = 15
        if include_recipient_columns:
            recipient_note = latest_recipient_notes.get(donation.id) or '-'
            decline_reason = _get_donor_decline_reason_for_export(donation, latest_recipient_declines) or 'No note'
            rec_note_cell = ws.cell(row=row_num, column=next_col, value=recipient_note)
            rec_note_cell.border = border
            rec_note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            next_col += 1
            rec_decl_cell = ws.cell(row=row_num, column=next_col, value=decline_reason)
            rec_decl_cell.border = border
            rec_decl_cell.alignment = Alignment(wrap_text=True, vertical="top")
            next_col += 1

        ws.cell(row=row_num, column=next_col, value=status).border = border

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_num, column=col)
            if c.alignment is None or not getattr(c.alignment, 'wrap_text', False):
                c.alignment = Alignment(vertical="top")

    # Adjust column widths
    if include_recipient_columns:
        column_widths = [6, 18, 12, 12, 30, 25, 24, 10, 10, 15, 12, 30, 25, 25, 25, 25, 20]
    else:
        column_widths = [6, 18, 12, 12, 30, 25, 24, 10, 10, 15, 12, 30, 25, 25, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{donor_name.replace(' ', '_')}_donations_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@foodbank_required
def foodbank_donations_list(request):
    """View for foodbanks to see all donations received with filters"""
    foodbank_profile = request.user.foodbank_profile
    donations = Donation.objects.filter(
        foodbank=foodbank_profile
    ).select_related(
        'donor', 'donor__donor_profile', 'foodbank_request', 'accepted_by_recipient',
        'declined_by_recipient', 'accepted_by_recipient__user', 'declined_by_recipient__user',
        'request_management', 'request_management__foodbank', 'request_management__recipient',
        'foodbank_request__original_request', 'foodbank_request__original_request__foodbank',
        'foodbank_request__original_request__recipient',
        'foodbank_request__linked_request_management'
    ).prefetch_related(
        'allocations__recipient', 'allocations__request_management', 'discussion',
        'foodbank_request__recipient_requests'
    ).order_by('-donated_at')

    # Show only donations that were made against an explicit request (direct or linked)
    donations = donations.filter(
        Q(foodbank_request__isnull=False) | Q(request_management__isnull=False)
    )
    
    # Get filter parameters
    type_filter = request.GET.get('type', '').strip()
    category_filter = request.GET.get('category', '').strip()
    donor_filter = request.GET.get('donor', '')
    delivery_status_raw = request.GET.get('delivery_status', '').strip()
    delivery_filter = delivery_status_raw.lower()
    quantity_filter = request.GET.get('quantity', '').strip().lower()
    amount_filter = request.GET.get('amount', '').strip().lower()
    status_filter = request.GET.get('status', '').strip()
    date_filter = request.GET.get('date_range', 'all').strip() or 'all'
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    # Backward compatibility: older UI had type/category swapped.
    request_type_choices = {'food', 'non_food'}
    donation_kind_choices = {'item', 'subsidized', 'money', 'monetary'}

    if type_filter in donation_kind_choices and category_filter in request_type_choices:
        type_filter, category_filter = category_filter, type_filter
    elif type_filter in donation_kind_choices and not category_filter:
        category_filter = type_filter
        type_filter = ''
    elif category_filter in request_type_choices and not type_filter:
        type_filter = category_filter
        category_filter = ''

    if category_filter == 'monetary':
        category_filter = 'money'

    if type_filter and type_filter not in request_type_choices:
        type_filter = ''
    if category_filter and category_filter not in {'item', 'subsidized', 'money'}:
        category_filter = ''

    # Apply category filter (donation kind: item/subsidized/money/csr/other)
    if category_filter:
        donations = donations.filter(donation_type=category_filter)
    
    if donor_filter:
        donations = donations.filter(donor__id=donor_filter)
    
    if delivery_filter:
        if delivery_filter == 'dropoff':
            delivery_filter = 'delivery'
        if delivery_filter == 'pickup':
            donations = donations.filter(delivery_method='pickup')
        elif delivery_filter == 'delivery':
            donations = donations.filter(
                Q(delivery_method='delivery') | Q(delivery_method='dropoff')
            ).distinct()
        else:
            donations = donations.filter(delivery_status=delivery_filter)

    if quantity_filter:
        if quantity_filter == 'small':
            donations = donations.filter(
                Q(donation_type='item', quantity__lte=50) |
                Q(donation_type='subsidized', subsidized_quantity__lte=50) |
                Q(allocations__quantity__lte=50)
            ).distinct()
        elif quantity_filter == 'medium':
            donations = donations.filter(
                Q(donation_type='item', quantity__gt=50, quantity__lte=200) |
                Q(donation_type='subsidized', subsidized_quantity__gt=50, subsidized_quantity__lte=200) |
                Q(allocations__quantity__gt=50, allocations__quantity__lte=200)
            ).distinct()
        elif quantity_filter == 'large':
            donations = donations.filter(
                Q(donation_type='item', quantity__gt=200) |
                Q(donation_type='subsidized', subsidized_quantity__gt=200) |
                Q(allocations__quantity__gt=200)
            ).distinct()

    if amount_filter:
        if amount_filter == 'small':
            donations = donations.filter(
                Q(donation_type='money', amount__lte=5000) |
                Q(donation_type='subsidized', subsidized_price__lte=5000) |
                Q(allocations__amount__lte=5000)
            ).distinct()
        elif amount_filter == 'medium':
            donations = donations.filter(
                Q(donation_type='money', amount__gt=5000, amount__lte=20000) |
                Q(donation_type='subsidized', subsidized_price__gt=5000, subsidized_price__lte=20000) |
                Q(allocations__amount__gt=5000, allocations__amount__lte=20000)
            ).distinct()
        elif amount_filter == 'large':
            donations = donations.filter(
                Q(donation_type='money', amount__gt=20000) |
                Q(donation_type='subsidized', subsidized_price__gt=20000) |
                Q(allocations__amount__gt=20000)
            ).distinct()
    # Keep list scope aligned with statuses rendered in the table, but always
    # include recipient-declined rows even if status value is atypical.
    donations = donations.filter(
        Q(status__in=['pending', 'accepted', 'fulfilled', 'partial', 'declined']) |
        Q(declined_by_recipient__isnull=False) |
        Q(allocations__declined_by_recipient=True)
    ).distinct()

    # Date filter (supports presets + custom range)
    if (date_filter == 'custom') or (date_filter in ('', 'all') and (date_from or date_to)):
        if date_from:
            try:
                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
                donations = donations.filter(donated_at__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                donations = donations.filter(donated_at__lte=to_date)
            except ValueError:
                pass
    elif date_filter != 'all':
        now = timezone.now()
        if date_filter == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            donations = donations.filter(donated_at__gte=start_date)
        elif date_filter in ('7days', 'week'):
            donations = donations.filter(donated_at__gte=now - timedelta(days=7))
        elif date_filter == 'month':
            donations = donations.filter(donated_at__gte=now - timedelta(days=30))
        elif date_filter == '3months':
            donations = donations.filter(donated_at__gte=now - timedelta(days=90))
    
    if search:
        search_q = (
            Q(item_name__icontains=search) |
            Q(message__icontains=search) |
            Q(other_description__icontains=search) |
            Q(csr_description__icontains=search) |
            Q(subsidized_product_type__icontains=search) |
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank_request__title__icontains=search) |
            Q(foodbank_request__description__icontains=search) |
            Q(foodbank_request__original_request__description__icontains=search) |
            Q(foodbank_request__original_request__additional_notes__icontains=search) |
            Q(foodbank_request__linked_request_management__description__icontains=search) |
            Q(request_management__description__icontains=search) |
            Q(request_management__additional_notes__icontains=search) |
            Q(request_management__location__icontains=search) |
            Q(request_management__recipient__full_name__icontains=search) |
            Q(request_management__recipient__user__email__icontains=search) |
            Q(foodbank_request__original_request__recipient__full_name__icontains=search) |
            Q(foodbank_request__original_request__recipient__user__email__icontains=search) |
            Q(foodbank_request__original_request__location__icontains=search) |
            Q(allocations__recipient__full_name__icontains=search) |
            Q(allocations__recipient__user__email__icontains=search) |
            Q(allocations__request_management__description__icontains=search) |
            Q(allocations__request_management__additional_notes__icontains=search) |
            Q(allocations__request_management__location__icontains=search) |
            Q(accepted_by_recipient__full_name__icontains=search) |
            Q(accepted_by_recipient__user__email__icontains=search) |
            Q(declined_by_recipient__full_name__icontains=search) |
            Q(declined_by_recipient__user__email__icontains=search) |
            Q(foodbank__address__icontains=search)
        )
        donations = donations.filter(search_q).distinct()
    
    # Get statistics
    total_donations = donations.count()
    total_items = donations.filter(donation_type='item').aggregate(
        total=Sum('quantity')
    )['total'] or 0
    total_money = donations.filter(donation_type='money').aggregate(
        total=Sum('amount')
    )['total'] or 0
    unallocated_count = donations.filter(is_allocated=False).count()
    total_subsidized = donations.filter(donation_type='subsidized').aggregate(
        total=Sum('subsidized_price')
    )['total'] or 0
    
    # Get unique donors for filter dropdown
    donors = request.user.__class__.objects.filter(
        id__in=donations.values_list('donor_id', flat=True).distinct(),
        user_type='DONOR'
    ).select_related('donor_profile')
    
    # Get donation responses for all donations (for recipient notes)
    donation_ids = list(donations.values_list('id', flat=True))
    donation_responses = DonationResponse.objects.filter(
        donation_id__in=donation_ids
    ).select_related('recipient').order_by('-responded_at')

    # Create a dictionary of responses by donation_id
    responses_by_donation = {}
    latest_recipient_notes = {}
    for response in donation_responses:
        donation_id = response.donation_id
        if donation_id not in responses_by_donation:
            responses_by_donation[donation_id] = []
        responses_by_donation[donation_id].append(response)

        if response.notes and donation_id not in latest_recipient_notes:
            latest_recipient_notes[donation_id] = {
                'recipient_name': response.recipient.full_name,
                'notes': response.notes
            }

    # Build table rows: for "stock" donations (direct, no request link) with allocations,
    # show one row "Not allocated" / Accepted by FB and one row per allocation with recipient/status.
    from types import SimpleNamespace

    def is_direct_stock(d):
        """Direct donation to foodbank (no request link)."""
        if d.foodbank_request_id is not None or (hasattr(d, 'request_management') and d.request_management_id is not None):
            return False
        return True

    def has_allocations(d):
        return len(list(d.allocations.all())) > 0

    def _resolve_related_request(donation, allocation=None):
        if allocation and getattr(allocation, 'request_management', None):
            return allocation.request_management
        rm = getattr(donation, 'request_management', None)
        if rm:
            return rm
        foodbank_request = getattr(donation, 'foodbank_request', None)
        if foodbank_request:
            linked_rm = getattr(foodbank_request, 'linked_request_management', None)
            if linked_rm:
                return linked_rm
            original_request = getattr(foodbank_request, 'original_request', None)
            if original_request:
                return original_request
            return foodbank_request
        return None

    def _extract_request_type_info(request_obj):
        if not request_obj:
            return None, None
        request_type_value = None
        for attr in ('request_type', 'donation_type', 'request_category'):
            value = getattr(request_obj, attr, None)
            if value:
                request_type_value = value
                break
        if not request_type_value:
            return None, None
        label = None
        if hasattr(request_obj, 'get_request_type_display'):
            try:
                label = request_obj.get_request_type_display()
            except Exception:
                label = None
        if not label and hasattr(request_obj, 'get_donation_type_display'):
            try:
                label = request_obj.get_donation_type_display()
            except Exception:
                label = None
        if not label and hasattr(request_obj, 'get_request_category_display'):
            try:
                label = request_obj.get_request_category_display()
            except Exception:
                label = None
        if not label:
            label = str(request_type_value).replace('_', ' ').title()
        return request_type_value, label

    def _attach_row_request_type(row_obj, donation, allocation=None):
        related_request = _resolve_related_request(donation, allocation)
        req_type, req_label = _extract_request_type_info(related_request)
        row_obj.target_request_type = req_type
        row_obj.target_request_type_label = req_label
        if allocation and getattr(allocation, 'request_management', None):
            row_obj.requested_unit_label = _resolve_unit_label(
                allocation.request_management,
                'unit',
                'custom_unit',
                'units'
            )
        else:
            row_obj.requested_unit_label = _get_requested_unit_for_export(donation) or 'units'
        return row_obj

    donations_list = list(donations)
    table_rows = []
    for donation in donations_list:
        if has_allocations(donation):
            # Allocation rows first (newest first) — show like new donations on top
            allocs = sorted(
                donation.allocations.all(),
                key=lambda a: a.allocated_at or timezone.now() - timedelta(days=99999),
                reverse=True
            )
            for alloc in allocs:
                req = getattr(alloc, 'request_management', None)
                status_display = _get_allocation_row_status(req, allocation=alloc)
                status_class = STATUS_CLASS_MAP.get(status_display, 'accepted')
                row_entry = SimpleNamespace(
                    donation=donation,
                    is_allocation=True,
                    allocation=alloc,
                    recipient_display=alloc.recipient.full_name if alloc.recipient else '—',
                    status_display=status_display,
                    status_class=status_class,
                )
                table_rows.append(_attach_row_request_type(row_entry, donation, alloc))
            # One summary row: always "Not allocated" (stock row never shows recipient name or "X recipients")
            if is_direct_stock(donation):
                row_entry = SimpleNamespace(
                    donation=donation,
                    is_allocation=False,
                    allocation=None,
                    recipient_display='Not allocated',
                    status_display='Accepted by Foodbank',
                    status_class='accepted',
                )
                table_rows.append(_attach_row_request_type(row_entry, donation))
            else:
                row_entry = SimpleNamespace(
                    donation=donation,
                    is_allocation=False,
                    allocation=None,
                    recipient_display='Not allocated',
                    status_display=get_display_status(donation),
                    status_class=STATUS_CLASS_MAP.get(get_display_status(donation), 'pending'),
                )
                table_rows.append(_attach_row_request_type(row_entry, donation))
        else:
            # Single row: direct/stock donations always show Not allocated (no recipient name)
            recipient_display = 'Not allocated' if is_direct_stock(donation) else donation.get_recipient_name()
            row_entry = SimpleNamespace(
                donation=donation,
                is_allocation=False,
                allocation=None,
                recipient_display=recipient_display,
                status_display=get_display_status(donation),
                status_class=STATUS_CLASS_MAP.get(get_display_status(donation), 'pending'),
            )
            table_rows.append(_attach_row_request_type(row_entry, donation))

    # Apply type filter on row target request type (food/non_food), matching Type column.
    if type_filter:
        def _row_type_value(row):
            value = getattr(row, 'target_request_type', None)
            if value:
                return str(value).lower()
            fallback = getattr(row.donation, 'donation_category', None)
            if fallback in ('food', 'non_food'):
                return fallback
            return ''
        table_rows = [row for row in table_rows if _row_type_value(row) == type_filter]

    # Enforce quantity filters at row-level so only visible rows matching the
    # selected range are shown (prevents non-matching summary rows from leaking in).
    if quantity_filter in {'small', 'medium', 'large'}:
        def _to_number(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        def _row_quantity_value(row):
            donation = row.donation
            allocation = getattr(row, 'allocation', None)

            if row.is_allocation and allocation:
                if getattr(allocation, 'quantity', None) is not None:
                    return _to_number(allocation.quantity)
                req = getattr(allocation, 'request_management', None)
                if req and getattr(req, 'quantity', None) is not None:
                    return _to_number(req.quantity)

            if donation.donation_type == 'item':
                return _to_number(getattr(donation, 'quantity', None))

            if donation.donation_type == 'subsidized':
                qty = getattr(donation, 'subsidized_quantity', None)
                if qty is None:
                    qty = getattr(donation, 'quantity', None)
                return _to_number(qty)

            if donation.donation_type == 'money':
                fb_req = getattr(donation, 'foodbank_request', None)
                if fb_req and getattr(fb_req, 'quantity_needed', None) is not None:
                    return _to_number(fb_req.quantity_needed)
                if fb_req and getattr(fb_req, 'original_request', None) and getattr(fb_req.original_request, 'quantity', None) is not None:
                    return _to_number(fb_req.original_request.quantity)
                req = getattr(donation, 'request_management', None)
                if req and getattr(req, 'quantity', None) is not None:
                    return _to_number(req.quantity)
                return None

            return _to_number(getattr(donation, 'quantity', None))

        def _matches_quantity_band(row):
            qty = _row_quantity_value(row)
            if qty is None:
                return False
            if quantity_filter == 'small':
                return qty <= 50
            if quantity_filter == 'medium':
                return 50 < qty <= 200
            if quantity_filter == 'large':
                return qty > 200
            return True

        table_rows = [row for row in table_rows if _matches_quantity_band(row)]

    declined_labels = {'Declined by Recipient', 'Declined by Foodbank', 'Declined by FB', 'Declined'}
    pending_labels = {'Awaiting Foodbank Review', 'Awaiting FB Review'}
    status_filter_map = {
        'awaiting_foodbank_review': pending_labels,
        'accepted_by_foodbank': {'Accepted by Foodbank'},
        'awaiting_recipient': {
            'Awaiting Recipient',
            'Accepted → Awaiting Pickup',
            'Accepted â†’ Awaiting Pickup',
            'Delivery In Transit',
            'Delivery Scheduled',
            'Accepted',
        },
        'received_by_recipient': {
            'Received By Recipient',
            'Acknowledged by Recipient',
            'Delivered',
            'Fulfilled',
            'Fulfilled-Acknowledged',
            'Fulfilled-Received',
        },
        'declined_by_recipient': {'Declined by Recipient'},
        'declined_by_foodbank': {'Declined by Foodbank', 'Declined by FB'},
        'declined': declined_labels,
        # Backward compatibility with existing links/bookmarks
        'pending': pending_labels,
    }

    if status_filter:
        if status_filter == 'accepted':
            table_rows = [
                row for row in table_rows
                if row.status_display not in declined_labels and row.status_display not in pending_labels
            ]
        elif status_filter in status_filter_map:
            allowed_statuses = status_filter_map[status_filter]
            table_rows = [row for row in table_rows if row.status_display in allowed_statuses]
        else:
            normalized_filter = status_filter.replace('_', ' ').lower()
            table_rows = [
                row for row in table_rows
                if normalized_filter in (row.status_display or '').lower()
            ]

    # Sort entire table by activity date (newest first) so allocation rows appear on top when recent
    def row_date(r):
        if r.is_allocation and r.allocation and getattr(r.allocation, 'allocated_at', None):
            return r.allocation.allocated_at
        return r.donation.donated_at or timezone.now() - timedelta(days=99999)

    table_rows.sort(key=row_date, reverse=True)

    paginator = Paginator(table_rows, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    for row in page_obj.object_list:
        donation = row.donation
        note_entry = latest_recipient_notes.get(donation.id)
        if note_entry:
            donation.latest_note = note_entry.get('notes')
            donation.latest_note_recipient = note_entry.get('recipient_name')
        else:
            donation.latest_note = donation.message
            donation.latest_note_recipient = None
        note_history = []
        if donation.message:
            note_history.append({
                'author': donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile.full_name else donation.donor.email,
                'note': donation.message,
                'timestamp': donation.donated_at.strftime('%b %d, %Y %H:%M')
            })
        for response in responses_by_donation.get(donation.id, []):
            note_history.append({
                'author': response.recipient.full_name,
                'note': response.notes,
                'timestamp': response.responded_at.strftime('%b %d, %Y %H:%M') if hasattr(response, 'responded_at') and response.responded_at else ''
            })
        related_request = _get_related_request(donation)
        if row.is_allocation and row.allocation and getattr(row.allocation, 'request_management', None):
            related_request = row.allocation.request_management
        if related_request and related_request.additional_notes:
            latest_request_timestamp = related_request.updated_at or related_request.time_of_request
            note_history.append({
                'author': (related_request.foodbank.foodbank_name if getattr(related_request, 'foodbank', None) else 'Request Notes'),
                'note': related_request.additional_notes,
                'timestamp': latest_request_timestamp.strftime('%b %d, %Y %H:%M') if latest_request_timestamp else ''
            })
        if not donation.latest_note and note_history:
            donation.latest_note = note_history[0]['note']
        donation.has_note_history = bool(note_history)
        donation.note_history_json = json.dumps(note_history)

    context = {
        'donations': page_obj,
        'page_obj': page_obj,
        'responses_by_donation': responses_by_donation,
        'total_donations': total_donations,
        'total_items': total_items,
        'total_money': total_money,
        'total_subsidized': total_subsidized,
        'unallocated_count': unallocated_count,
        'donors': donors,
        'selected_type': type_filter,
        'selected_category': category_filter,
        'selected_donor': donor_filter,
        'selected_delivery_status': delivery_filter,
        'selected_quantity': quantity_filter,
        'selected_amount': amount_filter,
        'selected_status': status_filter or '',
        'date_filter': date_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search,
        'latest_recipient_notes': latest_recipient_notes,
    }
    
    return render(request, 'authentication/foodbank_donations.html', context)


@login_required
def donation_detail(request, donation_id):
    """View donation details"""
    donation = get_object_or_404(Donation, id=donation_id)
    
    # Check permissions
    if request.user.user_type == 'DONOR' and donation.donor != request.user:
        messages.error(request, 'You do not have permission to view this donation.')
        return redirect('donor_donations_list')
    
    if request.user.user_type == 'FOODBANK' and donation.foodbank != request.user.foodbank_profile:
        messages.error(request, 'You do not have permission to view this donation.')
        return redirect('foodbank_donations_list')
    
    # Get allocations if any
    allocations = donation.allocations.all().select_related('recipient', 'recipient__user')

    status_display = get_display_status(donation)
    status_class = STATUS_CLASS_MAP.get(status_display, 'pending')
    
    context = {
        'donation': donation,
        'allocations': allocations,
        'status_display': status_display,
        'status_class': status_class,
    }
    
    return render(request, 'authentication/donation_detail.html', context)


@login_required
@foodbank_required
def foodbank_donations_export(request):
    """Export donations data as CSV or PDF"""
    import csv
    from django.http import HttpResponse
    from django.template.loader import get_template
    from datetime import datetime
    
    foodbank_profile = request.user.foodbank_profile
    export_format = request.GET.get('format', 'csv')
    
    # Get the same filtered donations as the list view (prefetch for get_display_status and table-style Qty/Amount)
    donations = Donation.objects.filter(
        foodbank=foodbank_profile
    ).select_related(
        'donor', 'donor__donor_profile', 'foodbank_request', 'accepted_by_recipient',
        'request_management', 'request_management__recipient', 'unspecified_management',
        'foodbank_request__original_request', 'foodbank_request__original_request__recipient',
    ).prefetch_related(
        'allocations__recipient',
        'allocations__request_management',
    ).order_by('-donated_at')

    # Keep export scope aligned with /donations/foodbank/ table:
    # include only specified + direct request-linked donations, exclude unspecified.
    donations = donations.filter(
        Q(foodbank_request__isnull=False) | Q(request_management__isnull=False)
    )
    
    # Apply the same filters as the list view
    type_filter = request.GET.get('type', '').strip()
    category_filter = request.GET.get('category', '').strip()
    donor_filter = request.GET.get('donor', '')
    delivery_status_raw = request.GET.get('delivery_status', '').strip()
    delivery_status = delivery_status_raw.lower()
    quantity_filter = request.GET.get('quantity', '').strip().lower()
    amount_filter = request.GET.get('amount', '').strip().lower()
    status_filter = request.GET.get('status', '').strip()
    date_filter = request.GET.get('date_range', 'all').strip() or 'all'
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    # Backward compatibility: older UI had type/category swapped.
    request_type_choices = {'food', 'non_food'}
    donation_kind_choices = {'item', 'subsidized', 'money', 'monetary'}

    if type_filter in donation_kind_choices and category_filter in request_type_choices:
        type_filter, category_filter = category_filter, type_filter
    elif type_filter in donation_kind_choices and not category_filter:
        category_filter = type_filter
        type_filter = ''
    elif category_filter in request_type_choices and not type_filter:
        type_filter = category_filter
        category_filter = ''

    if category_filter == 'monetary':
        category_filter = 'money'

    if type_filter and type_filter not in request_type_choices:
        type_filter = ''
    if category_filter and category_filter not in {'item', 'subsidized', 'money'}:
        category_filter = ''

    # Apply category filter (donation kind)
    if category_filter:
        donations = donations.filter(donation_type=category_filter)
    
    if donor_filter:
        donations = donations.filter(donor__id=donor_filter)
    
    if delivery_status:
        if delivery_status == 'dropoff':
            delivery_status = 'delivery'
        if delivery_status == 'pickup':
            donations = donations.filter(delivery_method='pickup')
        elif delivery_status == 'delivery':
            donations = donations.filter(
                Q(delivery_method='delivery') | Q(delivery_method='dropoff')
            ).distinct()
        else:
            donations = donations.filter(delivery_status=delivery_status)

    if quantity_filter:
        if quantity_filter == 'small':
            donations = donations.filter(
                Q(donation_type='item', quantity__lte=50) |
                Q(donation_type='subsidized', subsidized_quantity__lte=50) |
                Q(allocations__quantity__lte=50)
            ).distinct()
        elif quantity_filter == 'medium':
            donations = donations.filter(
                Q(donation_type='item', quantity__gt=50, quantity__lte=200) |
                Q(donation_type='subsidized', subsidized_quantity__gt=50, subsidized_quantity__lte=200) |
                Q(allocations__quantity__gt=50, allocations__quantity__lte=200)
            ).distinct()
        elif quantity_filter == 'large':
            donations = donations.filter(
                Q(donation_type='item', quantity__gt=200) |
                Q(donation_type='subsidized', subsidized_quantity__gt=200) |
                Q(allocations__quantity__gt=200)
            ).distinct()

    if amount_filter:
        if amount_filter == 'small':
            donations = donations.filter(
                Q(donation_type='money', amount__lte=5000) |
                Q(donation_type='subsidized', subsidized_price__lte=5000) |
                Q(allocations__amount__lte=5000)
            ).distinct()
        elif amount_filter == 'medium':
            donations = donations.filter(
                Q(donation_type='money', amount__gt=5000, amount__lte=20000) |
                Q(donation_type='subsidized', subsidized_price__gt=5000, subsidized_price__lte=20000) |
                Q(allocations__amount__gt=5000, allocations__amount__lte=20000)
            ).distinct()
        elif amount_filter == 'large':
            donations = donations.filter(
                Q(donation_type='money', amount__gt=20000) |
                Q(donation_type='subsidized', subsidized_price__gt=20000) |
                Q(allocations__amount__gt=20000)
            ).distinct()

    # Base status scope aligned with table view; keep recipient-declined rows even
    # when donation.status has an atypical value.
    donations = donations.filter(
        Q(status__in=['pending', 'accepted', 'fulfilled', 'partial', 'declined']) |
        Q(declined_by_recipient__isnull=False) |
        Q(allocations__declined_by_recipient=True)
    ).distinct()

    # Date filter (supports presets + custom range)
    if (date_filter == 'custom') or (date_filter in ('', 'all') and (date_from or date_to)):
        if date_from:
            try:
                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
                donations = donations.filter(donated_at__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                donations = donations.filter(donated_at__lte=to_date)
            except ValueError:
                pass
    elif date_filter != 'all':
        now = timezone.now()
        if date_filter == 'today':
            donations = donations.filter(donated_at__gte=now.replace(hour=0, minute=0, second=0, microsecond=0))
        elif date_filter in ('7days', 'week'):
            donations = donations.filter(donated_at__gte=now - timedelta(days=7))
        elif date_filter == 'month':
            donations = donations.filter(donated_at__gte=now - timedelta(days=30))
        elif date_filter == '3months':
            donations = donations.filter(donated_at__gte=now - timedelta(days=90))
    
    if search:
        search_q = (
            Q(item_name__icontains=search) |
            Q(message__icontains=search) |
            Q(other_description__icontains=search) |
            Q(csr_description__icontains=search) |
            Q(subsidized_product_type__icontains=search) |
            Q(donor__email__icontains=search) |
            Q(donor__donor_profile__full_name__icontains=search) |
            Q(foodbank__foodbank_name__icontains=search) |
            Q(foodbank_request__title__icontains=search) |
            Q(foodbank_request__description__icontains=search) |
            Q(foodbank_request__original_request__description__icontains=search) |
            Q(foodbank_request__original_request__additional_notes__icontains=search) |
            Q(foodbank_request__linked_request_management__description__icontains=search) |
            Q(request_management__description__icontains=search) |
            Q(request_management__additional_notes__icontains=search) |
            Q(request_management__location__icontains=search) |
            Q(request_management__recipient__full_name__icontains=search) |
            Q(request_management__recipient__user__email__icontains=search) |
            Q(foodbank_request__original_request__recipient__full_name__icontains=search) |
            Q(foodbank_request__original_request__recipient__user__email__icontains=search) |
            Q(foodbank_request__original_request__location__icontains=search) |
            Q(allocations__recipient__full_name__icontains=search) |
            Q(allocations__recipient__user__email__icontains=search) |
            Q(allocations__request_management__description__icontains=search) |
            Q(allocations__request_management__additional_notes__icontains=search) |
            Q(allocations__request_management__location__icontains=search) |
            Q(accepted_by_recipient__full_name__icontains=search) |
            Q(accepted_by_recipient__user__email__icontains=search) |
            Q(declined_by_recipient__full_name__icontains=search) |
            Q(declined_by_recipient__user__email__icontains=search) |
            Q(foodbank__address__icontains=search)
        )
        donations = donations.filter(search_q).distinct()
    
    # Build export rows exactly like the table rows (allocation + summary rows).
    report_rows = _build_foodbank_table_rows(
        donations,
        type_filter=type_filter,
        status_filter=status_filter,
    )

    if export_format == 'csv':
        return export_donations_csv(report_rows, foodbank_profile)
    elif export_format == 'pdf':
        return export_donations_pdf(request, report_rows, foodbank_profile)
    elif export_format == 'excel':
        return export_donations_excel(report_rows, foodbank_profile)
    else:
        return HttpResponse("Invalid format", status=400)


def export_donations_csv(rows, foodbank_profile):
    """Export donations as CSV"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{foodbank_profile.foodbank_name}_donations_{timestamp}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header with foodbank info
    writer.writerow([f'Receive Donations Report - {foodbank_profile.foodbank_name}'])
    writer.writerow([f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'])
    writer.writerow([])  # Empty row
    
    # Write column headers - matching table with Recipient Note and Donor Note in separate columns
    writer.writerow([
        'S/No',
        'Date',
        'Donor',
        'Recipient',
        'Type',
        'Category',
        'Description',
        'Quantity Details',
        'Quantity Requested',
        'Quantity Fulfilled',
        'Unit',
        'Amount (KSH)',
        'Status',
        'Decline Reason',
        'Delivery Method',
        'Location',
        'Recipient Note',
        'Donor Note'
    ])
    
    # Write data rows (row-by-row exactly like table: includes allocation rows).
    for idx, row in enumerate(rows, 1):
        donation = row.donation
        description = _get_export_description_for_row(row)
        quantity_details = _get_export_quantity_details_for_row(row).replace('\r\n', '\n').replace('\n', ' ')
        requested_qty, requested_unit = _get_export_requested_for_row(row)
        if row.is_allocation and row.allocation:
            fulfilled_qty = row.allocation.quantity if row.allocation.quantity is not None else '-'
        elif donation.donation_type == 'item':
            fulfilled_qty = donation.quantity or '-'
        elif donation.donation_type == 'money':
            fulfilled_qty = requested_qty if requested_qty != '-' else '-'
        elif donation.donation_type == 'subsidized':
            fulfilled_qty = donation.subsidized_quantity or '-'
        elif donation.donation_type == 'csr':
            fulfilled_qty = '-'
        else:
            fulfilled_qty = donation.quantity or '-'

        unit = requested_unit if row.is_allocation and row.allocation else (
            donation.quantity_unit or '-'
            if donation.donation_type not in ('money', 'subsidized', 'csr')
            else (
                requested_unit if donation.donation_type == 'money'
                else (donation.subsidized_quantity_unit or '-')
                if donation.donation_type == 'subsidized'
                else '-'
            )
        )

        if row.is_allocation and row.allocation and row.allocation.amount is not None:
            amount = f"{row.allocation.amount:,.2f}"
        elif donation.donation_type == 'money':
            amount = f"{donation.amount:,.2f}" if donation.amount else '-'
        elif donation.donation_type == 'subsidized':
            amount = f"{donation.subsidized_price:,.2f}" if donation.subsidized_price else '-'
        elif donation.donation_type == 'csr':
            amount = f"{donation.amount:,.2f}" if donation.amount else '-'
        else:
            amount = f"{donation.amount:,.2f}" if donation.amount else '-'

        recipient = getattr(row, 'recipient_display', None) or donation.get_recipient_name() or "Not allocated"

        recipient_note = ''
        if row.is_allocation and row.allocation and getattr(row.allocation, 'request_management', None) and getattr(row.allocation.request_management, 'additional_notes', None):
            recipient_note = (row.allocation.request_management.additional_notes or '').strip()
        elif getattr(donation, 'request_management', None) and getattr(donation.request_management, 'additional_notes', None):
            recipient_note = (donation.request_management.additional_notes or '').strip()
        elif getattr(donation, 'foodbank_request', None):
            orig = getattr(donation.foodbank_request, 'original_request', None)
            if orig and getattr(orig, 'additional_notes', None):
                recipient_note = (orig.additional_notes or '').strip()
        donor_note = (
            (donation.message or '').strip()
            or (getattr(donation, 'other_description', None) or '').strip()
            or (getattr(donation, 'csr_description', None) or '').strip()
        )

        status_display = getattr(row, 'status_display', None) or _get_foodbank_export_status_display(donation)
        decline_reason = _get_export_decline_reason_for_row(row)

        export_date = donation.donated_at
        if row.is_allocation and row.allocation and getattr(row.allocation, 'allocated_at', None):
            export_date = row.allocation.allocated_at

        writer.writerow([
            idx,
            export_date.strftime('%Y-%m-%d') if export_date else '-',
            donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile.full_name else donation.donor.email,
            recipient,
            _get_export_type_display_for_row(row),
            _get_foodbank_export_category_display(donation),
            description,
            quantity_details,
            requested_qty,
            fulfilled_qty,
            unit,
            amount,
            status_display,
            decline_reason,
            donation.delivery_method or '-',
            donation.foodbank.address if hasattr(donation.foodbank, 'address') and donation.foodbank.address else '-',
            recipient_note or '-',
            donor_note or '-',
        ])
    
    return response


def export_donations_pdf(request, rows, foodbank_profile):
    """Export donations as PDF using FoodBankHub branded report_utils."""
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph
    from .report_utils import (
        get_report_styles, build_report_header, get_branded_table_style,
        build_report_summary, build_pdf_document, collect_active_filters,
        make_full_width_table,
    )

    rows_list = list(rows)
    styles = get_report_styles()
    wrap = styles['wrap']
    # Break long unspaced words so description text never gets clipped.
    wrap_desc = ParagraphStyle('FoodbankDonationDescWrap', parent=wrap, wordWrap='CJK')
    elements = []

    name = foodbank_profile.foodbank_name or foodbank_profile.user.email
    active_filters = collect_active_filters(request, [
        ('type', 'Type'), ('category', 'Category'), ('donor', 'Donor'),
        ('delivery_status', 'Delivery Status'), ('status', 'Status'), ('search', 'Search'),
    ])

    build_report_header(
        elements, "Receive Donations Report", name,
        len(rows_list), active_filters, styles,
    )

    if not rows_list:
        elements.append(Paragraph("No donations found matching the current filters.", styles['normal']))
    else:
        data = [['S/No', 'Date', 'Donor', 'Recipient', 'Type', 'Category', 'Description',
                 'Quantity Details', 'Status', 'Decline Reason', 'Delivery', 'Location', 'Recipient Note', 'Donor Note']]

        for idx, row in enumerate(rows_list, 1):
            donation = row.donation
            description = _get_export_description_for_row(row)
            qty_amt = _get_export_quantity_details_for_row(row)
            type_display = _get_export_type_display_for_row(row)
            category_display = _get_foodbank_export_category_display(donation)

            recipient = getattr(row, 'recipient_display', None) or donation.get_recipient_name() or "Not allocated"

            delivery = donation.get_delivery_method_display() if donation.delivery_method else '-'
            location = (donation.foodbank.address or '-') if getattr(donation.foodbank, 'address', None) else '-'

            # Recipient note: from linked request, foodbank request's original request, or first allocation's request
            recipient_note = ''
            if row.is_allocation and row.allocation and getattr(row.allocation, 'request_management', None) and getattr(row.allocation.request_management, 'additional_notes', None):
                recipient_note = (row.allocation.request_management.additional_notes or '').strip()
            elif getattr(donation, 'request_management', None) and getattr(donation.request_management, 'additional_notes', None):
                recipient_note = (donation.request_management.additional_notes or '').strip()
            elif getattr(donation, 'foodbank_request', None):
                orig = getattr(donation.foodbank_request, 'original_request', None)
                if orig and getattr(orig, 'additional_notes', None):
                    recipient_note = (orig.additional_notes or '').strip()
            if not recipient_note and getattr(donation, 'allocations', None):
                first_alloc = next(iter(donation.allocations.all()), None)
                if first_alloc and getattr(first_alloc, 'request_management', None) and getattr(first_alloc.request_management, 'additional_notes', None):
                    recipient_note = (first_alloc.request_management.additional_notes or '').strip()
            # Donor note: message, other_description, or csr_description
            donor_note = (
                (donation.message or '').strip()
                or (getattr(donation, 'other_description', None) or '').strip()
                or (getattr(donation, 'csr_description', None) or '').strip()
            )
            recipient_note_text = (recipient_note[:200] if recipient_note else '-').replace('\n', '<br/>')
            donor_note_text = (donor_note[:200] if donor_note else '-').replace('\n', '<br/>')
            status_display = getattr(row, 'status_display', None) or _get_foodbank_export_status_display(donation)
            decline_reason = _get_export_decline_reason_for_row(row).replace('\n', '<br/>')
            export_date = donation.donated_at
            if row.is_allocation and row.allocation and getattr(row.allocation, 'allocated_at', None):
                export_date = row.allocation.allocated_at
            description_text = html.escape(description or '-').replace('\n', '<br/>')

            data.append([
                str(idx),
                export_date.strftime('%b %d, %Y') if export_date else '-',
                Paragraph((donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and getattr(donation.donor.donor_profile, 'full_name', None) else donation.donor.email) or '-', wrap),
                Paragraph(recipient, wrap),
                Paragraph(type_display, wrap),
                Paragraph(category_display or '-', wrap),
                Paragraph(description_text, wrap_desc),
                Paragraph(qty_amt, wrap),
                Paragraph(status_display, wrap),
                Paragraph(decline_reason, wrap),
                Paragraph(delivery, wrap),
                Paragraph(location, wrap),
                Paragraph(recipient_note_text, wrap),
                Paragraph(donor_note_text, wrap),
            ])

        col_weights = [
            0.35, 0.70, 1.00, 0.95, 0.65, 0.65,
            1.45, 0.85, 0.80, 0.95, 0.60, 0.85, 0.85, 0.85,
        ]
        table = make_full_width_table(data, repeat_rows=1, col_weights=col_weights)
        table.setStyle(get_branded_table_style(len(data)))
        elements.append(table)

        build_report_summary(elements, [
            ("Total Donations", len(rows_list)),
        ], styles)

    safe_name = (foodbank_profile.foodbank_name or "foodbank").replace(" ", "_")[:30]
    return build_pdf_document(elements, "donations_received", safe_name)


def export_donations_excel(rows, foodbank_profile):
    """Export donations as Excel with formatting"""
    from django.http import HttpResponse
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Donations Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title and info
    ws.merge_cells('A1:R1')
    title_cell = ws['A1']
    title_cell.value = f"Receive Donations Report - {foodbank_profile.foodbank_name}"
    title_cell.font = Font(bold=True, size=14, color="10b981")
    title_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:R2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Donations: {len(rows)}"
    info_cell.alignment = Alignment(horizontal="center")
    
    # Headers - matching table with Requested and Recipient/Donor Note columns
    headers = [
        'S/No', 'Date', 'Donor', 'Recipient', 'Type', 'Category', 
        'Description', 'Quantity Details', 'Quantity Requested', 'Quantity Fulfilled', 'Unit', 'Amount (KSH)', 'Status',
        'Decline Reason', 'Delivery Method', 'Location', 'Recipient Note', 'Donor Note'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for idx, row in enumerate(rows, 1):
        donation = row.donation
        row_num = idx + 4
        description = _get_export_description_for_row(row)
        quantity_details = _get_export_quantity_details_for_row(row)
        requested_qty, requested_unit = _get_export_requested_for_row(row)
        if row.is_allocation and row.allocation:
            fulfilled_qty = row.allocation.quantity if row.allocation.quantity is not None else '-'
        elif donation.donation_type == 'item':
            fulfilled_qty = donation.quantity or '-'
        elif donation.donation_type == 'money':
            fulfilled_qty = requested_qty if requested_qty != '-' else '-'
        elif donation.donation_type == 'subsidized':
            fulfilled_qty = donation.subsidized_quantity or '-'
        elif donation.donation_type == 'csr':
            fulfilled_qty = '-'
        else:
            fulfilled_qty = donation.quantity or '-'

        unit = requested_unit if row.is_allocation and row.allocation else (
            donation.quantity_unit or '-'
            if donation.donation_type not in ('money', 'subsidized', 'csr')
            else (
                requested_unit if donation.donation_type == 'money'
                else (donation.subsidized_quantity_unit or '-')
                if donation.donation_type == 'subsidized'
                else '-'
            )
        )

        if row.is_allocation and row.allocation and row.allocation.amount is not None:
            amount = f"{row.allocation.amount:,.2f}"
        elif donation.donation_type == 'money':
            amount = f"{donation.amount:,.2f}" if donation.amount else '-'
        elif donation.donation_type == 'subsidized':
            amount = f"{donation.subsidized_price:,.2f}" if donation.subsidized_price else '-'
        elif donation.donation_type == 'csr':
            amount = f"{donation.amount:,.2f}" if donation.amount else '-'
        else:
            amount = f"{donation.amount:,.2f}" if donation.amount else '-'

        recipient = getattr(row, 'recipient_display', None) or donation.get_recipient_name() or "Not allocated"
        status_display = getattr(row, 'status_display', None) or _get_foodbank_export_status_display(donation)
        decline_reason = _get_export_decline_reason_for_row(row)

        recipient_note = ''
        if row.is_allocation and row.allocation and getattr(row.allocation, 'request_management', None) and getattr(row.allocation.request_management, 'additional_notes', None):
            recipient_note = (row.allocation.request_management.additional_notes or '').strip()
        elif getattr(donation, 'request_management', None) and getattr(donation.request_management, 'additional_notes', None):
            recipient_note = (donation.request_management.additional_notes or '').strip()
        elif getattr(donation, 'foodbank_request', None):
            orig = getattr(donation.foodbank_request, 'original_request', None)
            if orig and getattr(orig, 'additional_notes', None):
                recipient_note = (orig.additional_notes or '').strip()
        if not recipient_note and getattr(donation, 'allocations', None):
            first_alloc = next(iter(donation.allocations.all()), None)
            if first_alloc and getattr(first_alloc, 'request_management', None) and getattr(first_alloc.request_management, 'additional_notes', None):
                recipient_note = (first_alloc.request_management.additional_notes or '').strip()
        donor_note = (
            (donation.message or '').strip()
            or (getattr(donation, 'other_description', None) or '').strip()
            or (getattr(donation, 'csr_description', None) or '').strip()
        )
        
        ws.cell(row=row_num, column=1, value=idx).border = border
        export_date = donation.donated_at
        if row.is_allocation and row.allocation and getattr(row.allocation, 'allocated_at', None):
            export_date = row.allocation.allocated_at

        ws.cell(row=row_num, column=2, value=export_date.strftime('%Y-%m-%d') if export_date else '-').border = border
        ws.cell(row=row_num, column=3, value=donation.donor.donor_profile.full_name if hasattr(donation.donor, 'donor_profile') and donation.donor.donor_profile.full_name else donation.donor.email).border = border
        ws.cell(row=row_num, column=4, value=recipient).border = border
        ws.cell(row=row_num, column=5, value=_get_export_type_display_for_row(row)).border = border
        ws.cell(row=row_num, column=6, value=_get_foodbank_export_category_display(donation)).border = border
        ws.cell(row=row_num, column=7, value=description).border = border
        qty_details_cell = ws.cell(row=row_num, column=8, value=quantity_details)
        qty_details_cell.border = border
        qty_details_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=9, value=requested_qty).border = border
        ws.cell(row=row_num, column=10, value=fulfilled_qty).border = border
        ws.cell(row=row_num, column=11, value=unit).border = border
        ws.cell(row=row_num, column=12, value=amount).border = border
        ws.cell(row=row_num, column=13, value=status_display).border = border
        decline_cell = ws.cell(row=row_num, column=14, value=decline_reason)
        decline_cell.border = border
        decline_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=15, value=donation.delivery_method or '-').border = border
        ws.cell(row=row_num, column=16, value=donation.foodbank.address if hasattr(donation.foodbank, 'address') and donation.foodbank.address else '-').border = border
        ws.cell(row=row_num, column=17, value=recipient_note or '-').border = border
        ws.cell(row=row_num, column=18, value=donor_note or '-').border = border

    # Adjust column widths (18 columns)
    column_widths = [8, 12, 20, 20, 15, 14, 25, 24, 12, 10, 12, 18, 20, 24, 18, 25, 25, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{foodbank_profile.foodbank_name}_donations_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
