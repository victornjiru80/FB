"""Exports for recipient unspecified donations (available donations table).

All formats share the same columns: S/No, Type, Category, Description,
Foodbank, Quantity, Delivery, Location, Status, Date, Donor Note,
My Note, Decline Reason.
"""

from decimal import Decimal, InvalidOperation


def _format_currency(value):
    """Return a human-readable currency string for numeric values."""
    if value in (None, ''):
        return ''
    try:
        quantized = Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return str(value)
    return f"KES {quantized:,.2f}"


def _format_kes_whole(value):
    """Return whole-number KES format used in qty/amount display strings."""
    if value in (None, ''):
        return ''
    try:
        amount = Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return f"KES {value}"
    return f"KES {amount:.0f}"


def _normalize_note(value):
    """Return a standardized note placeholder when note text is missing."""
    text = (value or '').strip()
    return text if text else 'No note'


def _clean_text(value):
    """Normalize optional text fields to stripped strings."""
    return (value or '').strip()


def _resolve_linked_request_description(donation):
    """Best-effort request description fallback for donations tied to a request."""
    fb_req = getattr(donation, 'foodbank_request', None)
    if not fb_req:
        return ''
    original_req = getattr(fb_req, 'original_request', None)
    if original_req and getattr(original_req, 'description', None):
        return _clean_text(original_req.description)
    linked_req = getattr(fb_req, 'linked_request_management', None)
    if linked_req and getattr(linked_req, 'description', None):
        return _clean_text(linked_req.description)
    return _clean_text(getattr(fb_req, 'description', None)) or _clean_text(getattr(fb_req, 'title', None))


def resolve_available_donation_description(donation):
    """Resolve table/report description text for recipient available donations."""
    if donation is None:
        return 'No description'

    linked_request_desc = _resolve_linked_request_description(donation)
    message = _clean_text(getattr(donation, 'message', None))
    other_desc = _clean_text(getattr(donation, 'other_description', None))
    csr_desc = _clean_text(getattr(donation, 'csr_description', None))
    item_name = _clean_text(getattr(donation, 'item_name', None))

    if donation.donation_type == 'item':
        return item_name or linked_request_desc or 'General items'
    if donation.donation_type == 'money':
        return message or other_desc or csr_desc or linked_request_desc or 'No description'
    if donation.donation_type == 'other':
        return other_desc or message or linked_request_desc or 'No description'
    if donation.donation_type == 'csr':
        return csr_desc or other_desc or message or linked_request_desc or 'No description'
    return other_desc or csr_desc or message or item_name or linked_request_desc or 'No description'


def _resolve_table_description(donation):
    """Match the recipient available-donations table description precedence."""
    return resolve_available_donation_description(donation)


def _get_donation_details(item):
    """Extract common fields from an UnspecifiedDonationManagement item."""
    donation = item.donation

    # Baseline table semantics.
    type_display = donation.get_donation_category_display() if hasattr(donation, 'get_donation_category_display') else ''
    csr_sub = ''
    if donation.donation_type == 'csr' or donation.donation_category == 'csr':
        if donation.csr_subcategory == 'other' and donation.csr_custom_subcategory:
            csr_sub = donation.csr_custom_subcategory
        elif donation.csr_subcategory:
            csr_sub = donation.get_csr_subcategory_display() if hasattr(donation, 'get_csr_subcategory_display') else ''

    # Category baseline: donation mode/type.
    if donation.donation_type == 'item':
        if getattr(donation, 'donation_mode', '') == 'subsidized':
            category = 'Subsidized'
        elif getattr(donation, 'donation_mode', '') == 'free':
            category = 'Free'
        else:
            category = donation.get_donation_type_display() if hasattr(donation, 'get_donation_type_display') else 'Item'
    elif donation.donation_type == 'money':
        category = 'Monetary'
    else:
        category = donation.get_donation_type_display() if hasattr(donation, 'get_donation_type_display') else donation.donation_type

    # For CSR/Monetary rows, Type and Category are intentionally swapped.
    csr_sub_target = 'type'
    if donation.donation_type in ('csr', 'money'):
        type_display, category = category, type_display
        csr_sub_target = 'category'

    # Normalized labels
    if donation.donation_type == 'money':
        type_display = 'Monetary'
    if donation.donation_type == 'csr':
        category = 'CSR'

    # Description follows the same precedence as the table.
    description = _resolve_table_description(donation)

    # Quantity display (single column for all formats)
    if donation.donation_type == 'item':
        quantity = donation.quantity or 0
        unit = donation.quantity_unit or 'units'
        amount = ''
        if quantity:
            quantity_display = f"{quantity} {unit}".strip()
        else:
            quantity_display = unit or '—'
    elif donation.donation_type == 'money':
        quantity = ''
        unit = ''
        amount = donation.amount
        quantity_display = _format_currency(amount) or '—'
    elif donation.donation_type in ('csr', 'other'):
        quantity = donation.quantity or ''
        unit = donation.quantity_unit or ('units' if donation.quantity else '')
        amount = donation.amount
        if amount not in (None, '') and quantity:
            quantity_display = f"{_format_kes_whole(amount)} for {quantity} {unit}".strip()
        elif amount not in (None, ''):
            quantity_display = _format_kes_whole(amount)
        elif quantity:
            quantity_display = f"{quantity} {unit}".strip()
        else:
            quantity_display = '—'
    else:
        quantity = ''
        unit = ''
        amount = ''
        quantity_display = '—'

    # Delivery
    if donation.delivery_method in ('dropoff', 'delivery'):
        delivery = 'Delivery'
    elif donation.delivery_method == 'pickup':
        delivery = 'Pickup'
    else:
        delivery = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'

    # Location
    location = donation.foodbank.address if donation.foodbank else ''

    # Status (recipient-specific effective status when provided by views)
    status = (
        getattr(item, 'effective_recipient_status_display', None)
        or (item.get_recipient_status_display() if hasattr(item, 'get_recipient_status_display') else '')
    )

    # Date
    date_available = item.foodbank_reviewed_at.strftime('%b %d, %Y') if item.foodbank_reviewed_at else '-'
    date_iso = item.foodbank_reviewed_at.strftime('%Y-%m-%d') if item.foodbank_reviewed_at else ''

    # Notes - split into donor, recipient, and foodbank
    donor_note = _normalize_note(donation.message or donation.other_description or donation.csr_description)
    effective_recipient_note = (getattr(item, 'effective_recipient_note', None) or '').strip()
    recipient_note = _normalize_note(effective_recipient_note) if effective_recipient_note else ''
    foodbank_note = item.foodbank_decline_reason or ''
    effective_status = getattr(item, 'effective_recipient_status', getattr(item, 'recipient_status', ''))
    if effective_status == 'declined_by_recipient':
        decline_reason = (getattr(item, 'effective_decline_note', None) or getattr(item, 'recipient_decline_reason', None) or '').strip()
    else:
        decline_reason = ''

    # Foodbank name
    foodbank_name = donation.foodbank.foodbank_name if donation.foodbank else '-'

    return {
        'type': type_display,
        'category': category,
        'csr_sub': csr_sub,
        'csr_sub_target': csr_sub_target,
        'description': description,
        'foodbank': foodbank_name,
        'quantity': quantity,
        'unit': unit,
        'amount': amount,
        'quantity_display': quantity_display,
        'delivery': delivery,
        'location': location,
        'status': status,
        'date': date_available,
        'date_iso': date_iso,
        'donor_note': donor_note,
        'recipient_note': recipient_note,
        'foodbank_note': foodbank_note,
        'decline_reason': decline_reason,
    }


# ═══════════════════════════════════════════════════════════════════
#  PDF
# ═══════════════════════════════════════════════════════════════════

def export_available_donations_pdf(request, donations_data, recipient_profile):
    """Generate branded PDF report for available (unspecified) donations."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.pagesizes import A3, landscape
    from .report_utils import (
        get_report_styles, build_report_header, get_branded_table_style,
        build_report_summary, build_pdf_document, collect_active_filters,
        make_full_width_table,
    )

    styles = get_report_styles()
    wrap = styles['wrap']
    elements = []
    report_pagesize = landscape(A3)

    recipient_name = recipient_profile.full_name or recipient_profile.user.email

    active_filters = collect_active_filters(request, [
        ('status', 'Status'),
        ('type', 'Type'),
        ('category', 'Category'),
        ('delivery', 'Delivery'),
        ('quantity_range', 'Quantity Range'),
        ('amount_range', 'Amount Range'),
        ('sort', 'Sort'),
        ('search', 'Search'),
    ])

    build_report_header(
        elements, "Unspecified Donations Report", recipient_name,
        len(donations_data), active_filters, styles,
    )

    if not donations_data:
        elements.append(Paragraph("No donations found matching the current filters.", styles['normal']))
    else:
        # PDF: Quantity combined (number + unit), Amount shown where relevant
        data = [['S/No', 'Type', 'Category', 'Description', 'Foodbank',
                 'Quantity', 'Delivery', 'Location', 'Status', 'Date',
                 'Donor Note', 'My Note', 'Decline Reason']]

        for idx, item in enumerate(donations_data, 1):
            d = _get_donation_details(item)

            type_display = d['type']
            category_display = d['category']
            if d['csr_sub']:
                if d.get('csr_sub_target') == 'category':
                    category_display += f" ({d['csr_sub']})"
                else:
                    type_display += f" ({d['csr_sub']})"

            data.append([
                str(idx),
                Paragraph(type_display, wrap),
                Paragraph(category_display, wrap),
                Paragraph(d['description'][:120], wrap),
                Paragraph(d['foodbank'], wrap),
                Paragraph(d['quantity_display'], wrap),
                Paragraph(d['delivery'], wrap),
                Paragraph(d['location'] or '-', wrap),
                Paragraph(d['status'], wrap),
                Paragraph(d['date'], wrap),
                Paragraph((d['donor_note'] or '-')[:100], wrap),
                Paragraph((d['recipient_note'] or '-')[:100], wrap),
                Paragraph((d['decline_reason'] or '')[:100], wrap),
            ])

        _col_weights = [
            0.03, 0.06, 0.07, 0.12, 0.09,
            0.07, 0.06, 0.09, 0.07, 0.06,
            0.11, 0.11, 0.10,
        ]
        table = make_full_width_table(data, repeat_rows=1, col_weights=_col_weights, pagesize=report_pagesize)
        table.setStyle(get_branded_table_style(len(data)))
        elements.append(table)

        # Summary
        build_report_summary(elements, [
            ("Total Donations", len(donations_data)),
            ("Available", sum(1 for d in donations_data if d.recipient_status == 'pending_recipient')),
            ("Accepted", sum(1 for d in donations_data if d.recipient_status == 'accepted_by_recipient')),
            ("Received", sum(1 for d in donations_data if d.recipient_status == 'received')),
        ], styles)

    return build_pdf_document(elements, "available_donations", recipient_name, pagesize=report_pagesize)


# ═══════════════════════════════════════════════════════════════════
#  CSV
# ═══════════════════════════════════════════════════════════════════

def export_available_donations_csv(request, donations_data, recipient_profile):
    """Generate CSV report for unspecified donations (same order as table)."""
    import csv
    from django.http import HttpResponse
    import datetime

    response = HttpResponse(content_type='text/csv')
    recipient_name = recipient_profile.full_name or recipient_profile.user.email
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_unspecified_donations_{timestamp}.csv"'

    writer = csv.writer(response, quoting=csv.QUOTE_ALL)

    # Metadata
    writer.writerow(['Recipient', recipient_name])
    writer.writerow(['Report Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Total Records', len(donations_data)])
    writer.writerow([])

    # Headers – split quantity/unit/amount for spreadsheet calculations
    writer.writerow([
        'S/No', 'Type', 'Category', 'Description', 'Foodbank',
        'Quantity', 'Unit', 'Amount (KES)', 'Delivery', 'Location',
        'Status', 'Date', 'Donor Note', 'My Note', 'Decline Reason',
    ])

    for idx, item in enumerate(donations_data, 1):
        d = _get_donation_details(item)

        type_display = d['type']
        category_display = d['category']
        if d['csr_sub']:
            if d.get('csr_sub_target') == 'category':
                category_display += f" ({d['csr_sub']})"
            else:
                type_display += f" ({d['csr_sub']})"

        writer.writerow([
            idx,
            type_display,
            category_display,
            d['description'],
            d['foodbank'],
            d['quantity'],
            d['unit'],
            d['amount'],
            d['delivery'],
            d['location'] or '',
            d['status'],
            d['date'],
            d['donor_note'],
            d['recipient_note'],
            d['decline_reason'],
        ])

    # Summary
    writer.writerow([])
    writer.writerow(['Summary'])
    writer.writerow(['Total Donations', len(donations_data)])
    writer.writerow(['Available', sum(1 for d in donations_data if d.recipient_status == 'pending_recipient')])
    writer.writerow(['Accepted', sum(1 for d in donations_data if d.recipient_status == 'accepted_by_recipient')])
    writer.writerow(['Received', sum(1 for d in donations_data if d.recipient_status == 'received')])

    return response


# ═══════════════════════════════════════════════════════════════════
#  Excel
# ═══════════════════════════════════════════════════════════════════

def export_available_donations_excel(request, donations_data, recipient_profile):
    """Generate Excel report for unspecified donations (table-aligned headers)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse("Excel export requires openpyxl library. Please install it.")

    from django.http import HttpResponse
    import datetime
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Available Donations"

    recipient_name = recipient_profile.full_name or recipient_profile.user.email

    # ── Branded header ──
    ws['A1'] = "FOODBANKHUB"
    ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
    ws['A2'] = f"{recipient_name} - Unspecified Donations Report"
    ws['A2'].font = Font(size=13, bold=True, color="1F4E78")
    ws['A3'] = f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Total Records: {len(donations_data)}"
    ws['A3'].font = Font(size=10, italic=True)

    # ── Headers (row 5) ──
    headers = [
        'S/No', 'Type', 'Category', 'Description', 'Foodbank',
        'Quantity', 'Unit', 'Amount (KES)', 'Delivery', 'Location',
        'Status', 'Date', 'Donor Note', 'My Note', 'Decline Reason',
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data rows ──
    row_num = 6
    for idx, item in enumerate(donations_data, 1):
        d = _get_donation_details(item)

        type_display = d['type']
        category_display = d['category']
        if d['csr_sub']:
            if d.get('csr_sub_target') == 'category':
                category_display += f" ({d['csr_sub']})"
            else:
                type_display += f" ({d['csr_sub']})"

        amount_value = d['amount']
        if isinstance(amount_value, Decimal):
            amount_value = float(amount_value)

        row_data = [
            idx,
            type_display,
            category_display,
            d['description'],
            d['foodbank'],
            d['quantity'],
            d['unit'],
            amount_value,
            d['delivery'],
            d['location'] or '',
            d['status'],
            d['date'],
            d['donor_note'],
            d['recipient_note'],
            d['decline_reason'],
        ]

        is_alt = idx % 2 == 0
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_alt:
                cell.fill = alt_fill

            if col_num in [6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_num == 8 and isinstance(amount_value, (int, float)):
                cell.number_format = '#,##0.00'

        row_num += 1

    # ── Summary ──
    summary_row = row_num + 2
    ws.cell(row=summary_row, column=1).value = "Summary"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="1F4E78")

    summary_data = [
        ('Total Donations', len(donations_data)),
        ('Available', sum(1 for d in donations_data if d.recipient_status == 'pending_recipient')),
        ('Accepted', sum(1 for d in donations_data if d.recipient_status == 'accepted_by_recipient')),
        ('Received', sum(1 for d in donations_data if d.recipient_status == 'received')),
    ]
    for i, (label, value) in enumerate(summary_data):
        ws.cell(row=summary_row + 1 + i, column=1).value = label
        ws.cell(row=summary_row + 1 + i, column=2).value = value
        ws.cell(row=summary_row + 1 + i, column=1).font = Font(bold=True)

    # ── Column widths ──
    column_widths = {
        'A': 6,   # S/No
        'B': 12,  # Type
        'C': 16,  # Category
        'D': 28,  # Description
        'E': 20,  # Foodbank
        'F': 12,  # Quantity
        'G': 10,  # Unit
        'H': 14,  # Amount
        'I': 16,  # Delivery
        'J': 18,  # Location
        'K': 16,  # Status
        'L': 15,  # Date
        'M': 24,  # Donor Note
        'N': 22,  # My Note
        'O': 22,  # Decline Reason
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[5].height = 30

    # ── Response ──
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_available_donations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response
