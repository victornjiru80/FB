from email.policy import default
import re
import html
from urllib.parse import quote

from django.shortcuts import render, redirect, get_object_or_404

from django.http import HttpResponse, JsonResponse

import csv

from reportlab.lib.pagesizes import letter, A4

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from reportlab.lib.units import inch

from io import BytesIO

from django.core.paginator import Paginator

from django.contrib.auth import login, logout, update_session_auth_hash

from django.contrib.auth.decorators import login_required

from django.contrib.auth.forms import PasswordChangeForm

from django.views.generic import CreateView, TemplateView

from django.urls import reverse_lazy, reverse

from django.contrib import messages

from django.db.models import Sum, Count, Q, F, Value, Case, When, DecimalField, CharField, Exists, OuterRef, BooleanField

from django.db.models.functions import TruncMonth, Coalesce, Cast

from django.db import models, transaction

from datetime import datetime, timedelta, date

import uuid

from django.utils import timezone
from django.utils.text import slugify
from django.utils.http import url_has_allowed_host_and_scheme
from .utils import (
    send_welcome_email,
    send_application_received_email,
    notify_admins_new_application,
    send_approval_email,
    send_rejection_email,
    send_donation_confirmation_email,
    send_donation_confirmation_email_async,
    send_foodbank_request_notification_email,
    send_urgent_request_notification_email,
)
from django.conf import settings

from django.views.decorators.csrf import csrf_exempt

from django.core.cache import cache
import logging
import stripe

logger = logging.getLogger(__name__)
from django.utils.timezone import now

from .models import (

    CustomUser, DonorProfile, FoodBankProfile, RecipientProfile, 

    Donation, FoodBankRequest, DonationAllocation, RecipientRequest, 

    Notification, Testimonial, PaymentTransaction, FoodBankGalleryPhoto,

    DonationDiscussion, DonationDiscussionMessage, QUANTITY_UNITS, QuantityUnit,

    RequestManagement, Subscription, AccountDeletionRequest, SystemSupportDonation, DonationResponse,

    UnspecifiedDonationManagement

)

from .donation_views import (
    get_display_status,
    STATUS_CLASS_MAP,
    _get_foodbank_export_type_display,
    _get_foodbank_export_category_display,
)

from .forms import (

    CustomLoginForm, DonorRegistrationForm,

    FoodBankRegistrationForm, RecipientRegistrationForm, AdminRegistrationForm,

    DonationForm, FoodBankRequestForm, FoodBankProfileForm, FoodBankPasswordChangeForm,

    DonationAllocationForm, DonorProfileForm, DonorPasswordChangeForm,

    RecipientRequestForm, TestimonialForm, RecipientProfileForm, RecipientPasswordChangeForm,

    SystemSupportDonationForm

)

from .available_donations_exports import (
    export_available_donations_pdf as available_donations_pdf_report,
    export_available_donations_csv as available_donations_csv_report,
    export_available_donations_excel as available_donations_excel_report,
    resolve_available_donation_description,
)


def _get_quantity_units():
    default_units = list(QUANTITY_UNITS)
    db_units = list(QuantityUnit.objects.values_list('code', 'label'))

    # Start with the full default list, then append any custom units from DB.
    merged_units = []
    default_codes = {code for code, _ in default_units}
    db_by_code = {code: label for code, label in db_units}

    for code, label in default_units:
        merged_units.append((code, db_by_code.get(code, label)))

    for code, label in db_units:
        if code not in default_codes:
            merged_units.append((code, label))

    return merged_units


def _redirect_back_or_default(request, fallback='foodbank_requests_view'):
    """Redirect to a safe caller-provided next URL or a fallback view name."""
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
    if next_url and url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect(fallback)



class HomeView(TemplateView):

    template_name = 'authentication/home.html'

    

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        from .models import DonorTestimonial, FoodbankTestimonial

        def _safe_image_url(image_field):
            if not image_field:
                return ''
            try:
                return image_field.url
            except (ValueError, AttributeError):
                return ''

        real_story_testimonials = []

        recipient_testimonials = Testimonial.objects.filter(
            approval_status='approved',
            display_on_public=True,
        ).select_related('recipient__user')
        recipient_testimonials = [t for t in recipient_testimonials if t.is_currently_displayed()]
        for testimonial in recipient_testimonials:
            recipient = testimonial.recipient
            author_name = ''
            if recipient:
                if getattr(recipient, 'is_organization', False):
                    author_name = (recipient.organization_name or '').strip()
                else:
                    author_name = (recipient.full_name or '').strip()
                if not author_name:
                    user = getattr(recipient, 'user', None)
                    author_name = (getattr(user, 'email', '') or '').strip()

            real_story_testimonials.append({
                'source_type': 'recipient',
                'is_featured': bool(testimonial.is_featured),
                'created_at': testimonial.created_at,
                'message': testimonial.message or '',
                'author_name': author_name or 'Recipient',
                'author_role_label': 'Recipient',
                'category_label': 'Community',
                'impact_image_url': _safe_image_url(getattr(testimonial, 'impact_image', None)),
                'public_website_url': '',
            })

        donor_testimonials = DonorTestimonial.objects.filter(
            approval_status='approved',
            display_on_public=True,
        ).select_related('donor__user')
        donor_testimonials = [t for t in donor_testimonials if t.is_currently_displayed()]
        for testimonial in donor_testimonials:
            donor = testimonial.donor
            author_name = ''
            if donor:
                author_name = (getattr(donor, 'full_name', None) or '').strip()
                if not author_name and getattr(donor, 'user', None):
                    author_name = (donor.user.email or '').strip()

            real_story_testimonials.append({
                'source_type': 'donor',
                'is_featured': bool(testimonial.is_featured),
                'created_at': testimonial.created_at,
                'message': testimonial.message or '',
                'author_name': author_name or 'Donor',
                'author_role_label': 'Donor',
                'category_label': 'Donor',
                'impact_image_url': _safe_image_url(getattr(testimonial, 'impact_image', None)),
                'public_website_url': (testimonial.public_website_url or '').strip(),
            })

        foodbank_testimonials = FoodbankTestimonial.objects.filter(
            approval_status='approved',
            display_on_public=True,
        ).select_related('foodbank__user')
        foodbank_testimonials = [t for t in foodbank_testimonials if t.is_currently_displayed()]
        for testimonial in foodbank_testimonials:
            foodbank = testimonial.foodbank
            author_name = ''
            if foodbank:
                author_name = (getattr(foodbank, 'foodbank_name', None) or '').strip()

            real_story_testimonials.append({
                'source_type': 'foodbank',
                'is_featured': bool(testimonial.is_featured),
                'created_at': testimonial.created_at,
                'message': testimonial.message or '',
                'author_name': author_name or 'Food Bank',
                'author_role_label': 'Food Bank',
                'category_label': 'Food Bank',
                'impact_image_url': _safe_image_url(getattr(testimonial, 'impact_image', None)),
                'public_website_url': '',
            })

        # Featured first, then newest.
        real_story_testimonials.sort(
            key=lambda row: (row['is_featured'], row['created_at']),
            reverse=True,
        )
        context['real_story_testimonials'] = real_story_testimonials[:3]

        

        # Get active news sections for landing page

        from .models import NewsSection

        active_news = NewsSection.objects.filter(is_active=True)[:3]  # Show up to 3 active news items

        context['news_sections'] = active_news

        

        # Platform statistics - pass directly like admin dashboard

        from django.db.models import Sum, Count

        context['total_donors'] = CustomUser.objects.filter(user_type='DONOR', is_active=True).count()

        context['total_foodbanks'] = CustomUser.objects.filter(user_type='FOODBANK', is_active=True).count()

        context['total_recipients'] = CustomUser.objects.filter(user_type='RECIPIENT', is_active=True).count()

        

        # Count total donations (accepted donations)

        total_donations = Donation.objects.filter(status='accepted').count()

        

        # Calculate total meals donated (estimate: 1 donation = ~10 meals)

        context['estimated_meals'] = total_donations * 10

        

        return context



def csrf_debug(request):

    """Debug endpoint to check CSRF token status"""

    from django.middleware.csrf import get_token

    from django.http import JsonResponse

    

    token = get_token(request)

    return JsonResponse({

        'csrf_token': token,

        'session_key': request.session.session_key,

        'cookies': dict(request.COOKIES),

        'method': request.method,

    })


def news_detail(request, news_id: int):
    """
    Public news article/detail page for a single NewsSection.
    """
    from .models import NewsSection

    news = get_object_or_404(NewsSection, id=news_id, is_active=True)

    return render(request, 'authentication/news_detail.html', {
        'news': news,
        'related_news': NewsSection.objects.filter(is_active=True).exclude(id=news.id)[:3],
    })



def login_view(request):

    if request.user.is_authenticated:

        # Redirect admin users to custom admin dashboard

        if request.user.is_staff:

            return redirect('custom_admin:dashboard')

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = CustomLoginForm(data=request.POST)

        if form.is_valid():

            user = form.get_user()

            

            # Check if user is a foodbank and if they're approved

            if user.user_type == 'FOODBANK':

                try:

                    foodbank_profile = user.foodbank_profile

                    if foodbank_profile.is_approved == 'pending':

                        messages.error(

                            request, 

                            'Your food bank application is still under review. '

                            'You will receive an email notification once approved.'

                        )

                        return render(request, 'authentication/login.html', {'form': form})

                    elif foodbank_profile.is_approved == 'rejected':

                        messages.error(

                            request, 

                            'Your food bank application has been rejected. '

                            'Please contact support for more information.'

                        )

                        return render(request, 'authentication/login.html', {'form': form})

                except FoodBankProfile.DoesNotExist:

                    messages.error(request, 'Food bank profile not found. Please contact support.')

                    return render(request, 'authentication/login.html', {'form': form})

            

            login(request, user)

            

            # Redirect admin users to custom admin dashboard

            if user.is_staff:

                messages.success(request, f'Welcome back, {user.email}!')

                return redirect('custom_admin:dashboard')

            

            # Use role-aware dashboard routing for all non-staff users.
            return redirect('dashboard')

    else:

        form = CustomLoginForm()

    return render(request, 'authentication/login.html', {'form': form})



@login_required

def logout_view(request):

    logout(request)

    return redirect('home')



class DonorRegistrationView(CreateView):

    form_class = DonorRegistrationForm

    template_name = 'authentication/register_donor.html'

    success_url = reverse_lazy('login')



    def form_valid(self, form):

        response = super().form_valid(form)

        # Send welcome email to new donor

        try:

            send_welcome_email(self.object, self.object.user_type)

        except Exception as e:

            # Log error but don't break registration

            print(f"Failed to send welcome email: {e}")

        messages.success(self.request, 'Registration successful! You can now log in.')

        return response



class FoodBankRegistrationView(CreateView):

    form_class = FoodBankRegistrationForm

    template_name = 'authentication/register_foodbank.html'

    success_url = reverse_lazy('registration_pending')



    def form_valid(self, form):

        response = super().form_valid(form)

        # Send application received email to new foodbank

        try:

            send_application_received_email(self.object)

        except Exception as e:

            # Log error but don't break registration

            print(f"Failed to send application received email: {e}")

        

        # Notify admins about new application

        try:

            notify_admins_new_application(self.object.foodbank_profile)

        except Exception as e:

            print(f"Failed to notify admins: {e}")

            

        messages.success(

            self.request, 

            'Registration submitted successfully! Your application is now under review. '

            'You will receive an email notification once your application has been approved.'

        )

        return response



class RecipientRegistrationView(CreateView):

    form_class = RecipientRegistrationForm

    template_name = 'authentication/register_recipient.html'

    success_url = reverse_lazy('login')



    def form_valid(self, form):

        response = super().form_valid(form)

        # Send welcome email to new recipient

        try:

            send_welcome_email(self.object, self.object.user_type)

        except Exception as e:

            # Log error but don't break registration

            print(f"Failed to send welcome email: {e}")

        messages.success(self.request, 'Registration successful! You can now log in.')

        return response



class AdminRegistrationView(CreateView):

    form_class = AdminRegistrationForm

    template_name = 'authentication/register_admin.html'

    success_url = reverse_lazy('login')



    def form_valid(self, form):

        response = super().form_valid(form)

        # Send welcome email to new admin

        try:

            send_welcome_email(self.object, self.object.user_type)

        except Exception as e:

            # Log error but don't break registration

            print(f"Failed to send welcome email: {e}")

        messages.success(self.request, 'Admin registration successful! You can now log in with administrative privileges.')

        return response



def public_dashboard(request):

    donation_type = request.GET.get('donation_type', 'food')  # default to 'food'



    foodbank_requests = FoodBankRequest.objects.filter(

        donation_type=donation_type,

        status='active'

    ).order_by('-created_at')



    context = {

        'foodbank_requests': foodbank_requests,

        'selected_type': donation_type,

    }

    return render(request, 'authentication/public_dashboard.html', context)





@login_required

def dashboard(request):

    # Redirect admin users to custom admin dashboard

    if request.user.is_staff:

        return redirect('custom_admin:dashboard')

    

    user_type = request.user.user_type

    template_name = f'authentication/dashboard_{user_type.lower()}.html'

    

    if user_type == 'DONOR':

        # Get all donations by this donor

        donations = Donation.objects.filter(donor=request.user).order_by('-donated_at')



        # Total donated (money + subsidized)

        total_money = donations.filter(donation_type='money').aggregate(total=Sum('amount'))['total'] or 0

        total_subsidized = donations.filter(donation_type='subsidized').aggregate(total=Sum('subsidized_price'))['total'] or 0

        total_donated = float(total_money or 0) + float(total_subsidized or 0)

        item_count = donations.filter(donation_type='item').count()

        money_count = donations.filter(donation_type='money').count()

        subsidized_count = donations.filter(donation_type='subsidized').count()



        # Unique foodbanks supported

        foodbanks_supported = donations.values('foodbank').distinct().count()



         # Dummy data 

        recipients_impacted = donations.count() * 3  



        # Last donation date

        last_donation = donations.first().donated_at if donations.exists() else None



        # Get urgent food bank requests (priority: urgent, high)

        # Only show active requests where deadline hasn't passed (or no deadline set)

        urgent_requests = FoodBankRequest.objects.select_related('foodbank').filter(

            status='active',

            priority__in=['urgent', 'high'],

            foodbank__isnull=False

        ).filter(

            Q(deadline__isnull=True) | Q(deadline__gte=timezone.now())

        ).order_by('-priority', '-created_at')[:5]

        

        # Get all active requests for donors to see (ordered by newest first)

        # Only show active requests where deadline hasn't passed (or no deadline set)

        all_active_requests = FoodBankRequest.objects.select_related('foodbank').filter(

            status='active',

            foodbank__isnull=False

        ).filter(

            Q(deadline__isnull=True) | Q(deadline__gte=timezone.now())

        ).order_by('-created_at')

        

        # Filter out requests that are 100% fulfilled

        active_requests = []

        for req in all_active_requests:

            if req.get_fulfillment_percentage() < 100:

                active_requests.append(req)

            if len(active_requests) >= 15:  # Limit to 15 requests to ensure 6+ are visible

                break

        

        # Get latest requests for sidebar (same as active but limited to 5)

        # Only show active requests where deadline hasn't passed (or no deadline set)

        latest_requests = FoodBankRequest.objects.select_related('foodbank').filter(

            status='active',

            foodbank__isnull=False

        ).filter(

            Q(deadline__isnull=True) | Q(deadline__gte=timezone.now())

        ).order_by('-created_at')[:5]

        

        # Get recent activities (donations, notifications)

        recent_activities = []

        

        # Add recent donations as activities

        for donation in donations[:5]:

            activity = {

                'icon': 'gift',

                'color': 'success',

                'title': f'Donated {donation.donation_type.title()}',

                'description': f'To {donation.foodbank.foodbank_name}',

                'time': donation.donated_at.strftime('%b %d, %H:%M')

            }

            recent_activities.append(activity)

        

        # Add recent notifications as activities

        recent_notifications = Notification.objects.filter(

            user=request.user

        ).order_by('-created_at')[:3]

        

        for notification in recent_notifications:

            activity = {

                'icon': 'bell',

                'color': 'info',

                'title': notification.notification_type.replace('_', ' ').title(),

                'description': notification.message[:50] + '...' if len(notification.message) > 50 else notification.message,

                'time': notification.created_at.strftime('%b %d, %H:%M')

            }

            recent_activities.append(activity)

        

        # Sort activities by time (most recent first)

        recent_activities.sort(key=lambda x: x['time'], reverse=True)

        

        # Calculate impact metrics

        total_donations = donations.count()

        recent_donations = donations.filter(

            donated_at__gte=timezone.now() - timedelta(days=30)

        ).count()

        

        # Calculate impact score based on various factors

        impact_score = min(100, max(0, (

            total_donations * 10 +  # Base score from donations

            recent_donations * 5 +   # Bonus for recent activity

            (urgent_requests.count() * 2) +  # Bonus for responding to urgent needs

            (foodbanks_supported * 3)        # Bonus for supporting multiple food banks

        )))

        

        # Quick Donation Summary

        current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        this_month_donations = donations.filter(donated_at__gte=current_month_start).count()

        this_month_amount = donations.filter(

            donation_type='money',

            donated_at__gte=current_month_start

        ).aggregate(total=Sum('amount'))['total'] or 0

        

        # Most recent donation

        most_recent_donation = donations.first() if donations.exists() else None

        

        # Donation streak (consecutive months with donations)

        donation_streak = 0

        if donations.exists():

            current_date = timezone.now().replace(day=1)

            streak_month = current_date

            while True:

                month_start = streak_month

                month_end = (streak_month + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)

                if donations.filter(donated_at__gte=month_start, donated_at__lte=month_end).exists():

                    donation_streak += 1

                    streak_month = (streak_month - timedelta(days=32)).replace(day=1)

                else:

                    break

        

        # Achievement Badges

        achievements = []

        

        # First Donation

        if total_donations >= 1:

            achievements.append({

                'id': 'first_donation',

                'name': 'First Donation',

                'description': 'Made your first donation',

                'icon': 'fas fa-star',

                'earned': True,

                'earned_date': donations.order_by('donated_at').first().donated_at if donations.exists() else None

            })

        

        # Regular Donor (3+ donations)

        if total_donations >= 3:

            achievements.append({

                'id': 'regular_donor',

                'name': 'Regular Donor',

                'description': 'Made 3+ donations',

                'icon': 'fas fa-heart',

                'earned': True,

                'earned_date': donations.order_by('donated_at')[2].donated_at if donations.count() >= 3 else None

            })

        

        # Community Hero (helped 5+ food banks)

        if foodbanks_supported >= 5:

            achievements.append({

                'id': 'community_hero',

                'name': 'Community Hero',

                'description': 'Helped 5+ food banks',

                'icon': 'fas fa-users',

                'earned': True,

                'earned_date': None  # Could calculate when this was achieved

            })

        

        # Urgent Responder (responded to urgent requests)

        urgent_donations = donations.filter(foodbank_request__priority__in=['urgent', 'high']).count()

        if urgent_donations >= 1:

            achievements.append({

                'id': 'urgent_responder',

                'name': 'Urgent Responder',

                'description': 'Responded to urgent requests',

                'icon': 'fas fa-bolt',

                'earned': True,

                'earned_date': donations.filter(foodbank_request__priority__in=['urgent', 'high']).first().donated_at

            })

        

        # Monthly Champion (donated every month for 3+ months)

        if donation_streak >= 3:

            achievements.append({

                'id': 'monthly_champion',

                'name': 'Monthly Champion',

                'description': f'Donated for {donation_streak} consecutive months',

                'icon': 'fas fa-trophy',

                'earned': True,

                'earned_date': None

            })

        

        # Add unearned achievements for motivation

        if total_donations < 1:

            achievements.append({

                'id': 'first_donation',

                'name': 'First Donation',

                'description': 'Make your first donation',

                'icon': 'fas fa-star',

                'earned': False,

                'earned_date': None

            })

        

        if total_donations < 3:

            achievements.append({

                'id': 'regular_donor',

                'name': 'Regular Donor',

                'description': 'Make 3+ donations',

                'icon': 'fas fa-heart',

                'earned': False,

                'earned_date': None

            })

        

        if foodbanks_supported < 5:

            achievements.append({

                'id': 'community_hero',

                'name': 'Community Hero',

                'description': 'Help 5+ food banks',

                'icon': 'fas fa-users',

                'earned': False,

                'earned_date': None

            })

        

        if urgent_donations < 1:

            achievements.append({

                'id': 'urgent_responder',

                'name': 'Urgent Responder',

                'description': 'Respond to urgent requests',

                'icon': 'fas fa-bolt',

                'earned': False,

                'earned_date': None

            })

        

        if donation_streak < 3:

            achievements.append({

                'id': 'monthly_champion',

                'name': 'Monthly Champion',

                'description': 'Donate for 3+ consecutive months',

                'icon': 'fas fa-trophy',

                'earned': False,

                'earned_date': None

            })

        

        # Personal Impact Tracker

        lives_impacted = total_donations * 3  # Estimate 3 people per donation

        meals_provided = donations.filter(donation_type='item').aggregate(

            total=Sum('quantity')

        )['total'] or 0

        emergency_responses = urgent_donations

        

        # Smart Notifications

        smart_notifications = []

        

        # Recent urgent requests

        recent_urgent = FoodBankRequest.objects.filter(

            status='active',

            priority__in=['urgent', 'high'],

            created_at__gte=timezone.now() - timedelta(days=14)

        ).order_by('-created_at')[:3]

        

        for urgent_request in recent_urgent:

            smart_notifications.append({

                'type': 'urgent_request',

                'title': 'New Urgent Request',

                'message': f'"{urgent_request.title}" from {urgent_request.foodbank.foodbank_name}',

                'time': urgent_request.created_at,

                'icon': 'fas fa-exclamation-triangle',

                'color': 'danger',

                'action_url': f'/donate/{urgent_request.id}/'

            })

        

        # Recent thank you messages (from notifications)

        thank_you_notifications = Notification.objects.filter(

            user=request.user,

            notification_type='acknowledgement',

            created_at__gte=timezone.now() - timedelta(days=7)

        ).order_by('-created_at')[:3]

        

        for notification in thank_you_notifications:

            smart_notifications.append({

                'type': 'thank_you',

                'title': 'Thank You!',

                'message': notification.message,

                'time': notification.created_at,

                'icon': 'fas fa-heart',

                'color': 'success',

                'action_url': '/donation-history/'

            })

        

        # Donation confirmations

        recent_donations_notifications = Notification.objects.filter(

            user=request.user,

            notification_type__in=['donation_received', 'acknowledgement'],

            created_at__gte=timezone.now() - timedelta(days=3)

        ).order_by('-created_at')[:2]

        

        for notification in recent_donations_notifications:

            smart_notifications.append({

                'type': 'donation_confirmation',

                'title': 'Donation Confirmed',

                'message': notification.message,

                'time': notification.created_at,

                'icon': 'fas fa-check-circle',

                'color': 'info',

                'action_url': '/donation-history/'

            })

        

        

        # Phase 2 Chart Data

        import json

        

        # 6-month donation timeline data

        months_back = 6

        timeline_data = []

        for i in range(months_back):

            month_start = (timezone.now().replace(day=1) - timedelta(days=32*i)).replace(day=1)

            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)

            

            month_donations = donations.filter(

                donated_at__gte=month_start,

                donated_at__lte=month_end

            )

            

            timeline_data.append({

                'month': month_start.strftime('%b %Y'),

                'month_short': month_start.strftime('%b'),

                'donations': month_donations.count(),

                'amount': float(month_donations.filter(donation_type='money').aggregate(total=Sum('amount'))['total'] or 0),

                'items': month_donations.filter(donation_type='item').count()

            })

        

        timeline_data.reverse()  # Oldest to newest

        

        # Donation type breakdown for donut chart

        donation_breakdown = {

            'item': item_count,

            'money': money_count,

            'subsidized': subsidized_count

        }

        

        # Geographic impact (food banks helped by location)

        geographic_data = []

        foodbank_locations = donations.values('foodbank__address').annotate(

            count=Count('id'),

            location=F('foodbank__address')

        ).order_by('-count')[:5]

        

        for loc_data in foodbank_locations:

            if loc_data['location']:

                city = loc_data['location'].split(',')[0].strip()

                geographic_data.append({

                    'city': city,

                    'donations': loc_data['count']

                })

        

        # Response time trend (last 10 urgent donations)

        response_trend = []

        recent_urgent = donations.filter(

            foodbank_request__priority__in=['urgent', 'high'],

            foodbank_request__isnull=False

        ).order_by('-donated_at')[:10]

        

        for i, donation in enumerate(recent_urgent):

            if donation.foodbank_request:

                response_hours = (donation.donated_at - donation.foodbank_request.created_at).total_seconds() / 3600

                response_trend.append({

                    'donation': i + 1,

                    'hours': round(response_hours, 1)

                })

        

        response_trend.reverse()  # Chronological order

        

        # Top donors leaderboard (simplified)

        leaderboard = []

        # Get top 5 donors by donation count

        top_donors = CustomUser.objects.filter(user_type='DONOR').annotate(

            donation_count=Count('donation')

        ).order_by('-donation_count')[:5]

        

        for i, donor in enumerate(top_donors):

            is_current_user = donor.id == request.user.id

            leaderboard.append({

                'rank': i + 1,

                'name': 'You' if is_current_user else f'Donor {i + 1}',

                'donations': donor.donation_count,

                'is_current': is_current_user

            })

        

        # Unread notifications

        unread_notifications_count = Notification.objects.filter(user=request.user, is_read=False).count()

        

        # Foodbank support badge:
        # - Do not count newly-created tickets from the foodbank itself.
        # - Count only threads where admin has responded (reply/admin_response),
        #   and hide again when the foodbank follows up (status goes back to 'new').
        from .models import SupportMessage

        unread_support_count = SupportMessage.objects.filter(
            user=request.user
        ).filter(
            Q(admin_response__isnull=False) & ~Q(admin_response='') |
            Q(replies__is_from_admin=True)
        ).exclude(
            status='new'
        ).distinct().count()

        

        # Dummy recipient feedback

        recipient_feedback = [

        {'message': "Thank you for the food! It helped my family so much.", 'recipient_name': "Mary A.", 'date': datetime(2025, 6, 15)},

        {'message': "God bless you for your kindness.", 'recipient_name': "James O.", 'date': datetime(2025, 6, 12)},

        {'message': "Your donation reached us just in time. Forever grateful!", 'recipient_name': "Grace M.", 'date': datetime(2025, 6, 5)},

        ]



        # Get recent donations for table (limit to 10)

        recent_donations_table = list(

            donations.filter(

                foodbank_request__isnull=False,

                foodbank_request__original_request__isnull=False

            ).select_related(

                'foodbank',
                'foodbank__user',
                'foodbank_request',
                'foodbank_request__original_request',
                'foodbank_request__linked_request_management',
                'accepted_by_recipient'

            ).order_by('-donated_at')[:10]

        )

        direct_donations_table = list(

            donations.filter(

                foodbank_request__isnull=False,

                foodbank_request__original_request__isnull=True

            ).select_related(

                'foodbank',
                'foodbank__user',
                'foodbank_request',
                'foodbank_request__original_request',
                'foodbank_request__linked_request_management',
                'accepted_by_recipient'

            ).order_by('-donated_at')[:10]

        )



        for donation in recent_donations_table:

            status_display = get_display_status(donation)

            donation.status_display = status_display

            donation.status_class = STATUS_CLASS_MAP.get(status_display, 'pending')
            donation.donor_type_display = _get_foodbank_export_type_display(donation)
            donation.donor_category_display = _get_foodbank_export_category_display(donation)
            donation.requested_unit_label = _get_request_unit_label(getattr(donation, 'foodbank_request', None)) or 'units'
            _type_norm = (donation.donor_type_display or '').strip().lower()
            donation.donor_type_value = 'food' if _type_norm == 'food' else 'non_food' if _type_norm in ('non-food', 'non food', 'non_food') else ''



        for donation in direct_donations_table:

            status_display = get_display_status(donation)

            donation.status_display = status_display

            donation.status_class = STATUS_CLASS_MAP.get(status_display, 'pending')
            donation.donor_type_display = _get_foodbank_export_type_display(donation)
            donation.donor_category_display = _get_foodbank_export_category_display(donation)
            donation.requested_unit_label = _get_request_unit_label(getattr(donation, 'foodbank_request', None)) or 'units'
            _type_norm = (donation.donor_type_display or '').strip().lower()
            donation.donor_type_value = 'food' if _type_norm == 'food' else 'non_food' if _type_norm in ('non-food', 'non food', 'non_food') else ''

        

        # Get donor's unspecified donations with management status

        donor_unspecified_donations = UnspecifiedDonationManagement.objects.filter(

            donation__donor=request.user

        ).select_related(

            'donation', 'donation__foodbank', 'accepted_by_recipient'

        ).order_by('-created_at')[:5]

        

        donor_unspecified_counts = {

            'pending_foodbank': UnspecifiedDonationManagement.objects.filter(

                donation__donor=request.user, foodbank_status='pending_foodbank'

            ).count(),

            'accepted_by_foodbank': UnspecifiedDonationManagement.objects.filter(

                donation__donor=request.user, foodbank_status='accepted_by_foodbank'

            ).count(),

            'accepted_by_recipient': UnspecifiedDonationManagement.objects.filter(

                donation__donor=request.user, recipient_status='accepted_by_recipient'

            ).count(),

            'received': UnspecifiedDonationManagement.objects.filter(

                donation__donor=request.user, recipient_status='received'

            ).count(),

        }

        

        # Get donor's subsidized donations with status

        donor_subsidized_donations = list(Donation.objects.filter(

            donor=request.user,

            donation_type='subsidized',

            foodbank_request__isnull=True  # Only donor-initiated subsidized donations

        ).select_related('foodbank', 'accepted_by_recipient', 'unspecified_management').order_by('-donated_at')[:5])



        subsidized_ids = [donation.id for donation in donor_subsidized_donations]
        if subsidized_ids:
            responses = DonationResponse.objects.filter(
                donation_id__in=subsidized_ids
            ).exclude(
                notes__isnull=True
            ).exclude(
                notes__exact=''
            ).select_related('recipient').order_by('-responded_at')

            latest_notes = {}
            accepted_notes = {}
            decline_by_donation = {}
            accepted_recipient_by_donation = {
                donation.id: donation.accepted_by_recipient_id
                for donation in donor_subsidized_donations
                if donation.accepted_by_recipient_id
            }

            for response in responses:
                donation_id = response.donation_id
                if donation_id not in latest_notes:
                    latest_notes[donation_id] = response.notes

                if response.response_type == 'declined':
                    decline_by_donation.setdefault(donation_id, []).append(response)

                if (
                    response.response_type == 'accepted'
                    and accepted_recipient_by_donation.get(donation_id) == response.recipient_id
                    and donation_id not in accepted_notes
                ):
                    accepted_notes[donation_id] = response.notes

            for donation in donor_subsidized_donations:
                unspecified = getattr(donation, 'unspecified_management', None)
                unspecified_recipient_note = (getattr(unspecified, 'recipient_notes', None) or '').strip()
                unspecified_decline_note = (getattr(unspecified, 'recipient_decline_reason', None) or '').strip()

                if donation.accepted_by_recipient_id:
                    effective_recipient_note = (
                        accepted_notes.get(donation.id)
                        or unspecified_recipient_note
                        or ''
                    )
                    effective_recipient_decline_note = ''
                    donation.decline_responses = []
                else:
                    effective_recipient_note = (
                        latest_notes.get(donation.id)
                        or unspecified_recipient_note
                        or ''
                    )

                    decline_responses = decline_by_donation.get(donation.id, [])
                    if decline_responses:
                        decline_lines = []
                        for response in decline_responses:
                            recipient = getattr(response, 'recipient', None)
                            recipient_name = (
                                getattr(recipient, 'full_name', None)
                                or getattr(getattr(recipient, 'user', None), 'email', None)
                                or 'Recipient'
                            )
                            decline_lines.append(
                                f"{recipient_name}: {(response.notes or '').strip() or 'No note provided.'}"
                            )
                        effective_recipient_decline_note = "\n".join(decline_lines)
                    else:
                        effective_recipient_decline_note = unspecified_decline_note

                    donation.decline_responses = decline_responses

                donation.latest_recipient_note = effective_recipient_note
                donation.effective_recipient_note = effective_recipient_note
                donation.effective_recipient_decline_note = effective_recipient_decline_note
        else:
            for donation in donor_subsidized_donations:
                donation.latest_recipient_note = ''
                donation.effective_recipient_note = ''
                donation.effective_recipient_decline_note = ''
                donation.decline_responses = []
        

        donor_subsidized_counts = {

            'pending': Donation.objects.filter(

                donor=request.user,

                donation_type='subsidized',

                status='pending',

                foodbank_request__isnull=True

            ).count(),

            'accepted': Donation.objects.filter(

                donor=request.user,

                donation_type='subsidized',

                status='accepted',

                foodbank_request__isnull=True

            ).count(),

            'claimed': Donation.objects.filter(

                donor=request.user,

                donation_type='subsidized',

                status='accepted',

                accepted_by_recipient__isnull=False,

                foodbank_request__isnull=True

            ).count(),

            'delivered': Donation.objects.filter(

                donor=request.user,

                donation_type='subsidized',

                delivery_status='delivered',

                foodbank_request__isnull=True

            ).count(),

        }

        

        # Get donor's request-based donations (donations made to foodbank requests)

        donor_request_donations = Donation.objects.filter(

            donor=request.user,

            foodbank_request__isnull=False

        ).select_related('foodbank', 'foodbank_request').order_by('-donated_at')[:5]

        

        donor_request_counts = {

            'pending': Donation.objects.filter(donor=request.user, foodbank_request__isnull=False, status='pending').count(),

            'accepted': Donation.objects.filter(donor=request.user, foodbank_request__isnull=False, status='accepted').count(),

            'allocated': Donation.objects.filter(donor=request.user, foodbank_request__isnull=False, allocations__isnull=False).distinct().count(),

        }



        # Direct requests (from foodbanks) this donor can respond to
        direct_requests_available_count = FoodBankRequest.objects.filter(
            original_request__isnull=True,
            status='active',
        ).filter(
            Q(deadline__isnull=True) | Q(deadline__gte=timezone.now())
        ).exclude(
            donations__donor=request.user
        ).count()

        # Specified requests (recipient request sent to donors by foodbank) this donor can respond to
        specified_requests_available_count = FoodBankRequest.objects.filter(
            original_request__isnull=False,
            status='active',
        ).filter(
            Q(deadline__isnull=True) | Q(deadline__gte=timezone.now())
        ).exclude(
            donations__donor=request.user
        ).count()

        donor_tab_alerts = {

            'my_donations': 0,

            'direct': 0,  # Notifications for direct requests show on foodbank dashboard only

            'unspecified': 0,

            'subsidized': 0,

        }

        

        context = {

            'donations': donations,

            'total_donated': total_donated,

            'foodbanks_supported': foodbanks_supported,

            'recipients_impacted': recipients_impacted,

            'last_donation_date': last_donation,

            'active_requests': active_requests,

            'latest_requests': latest_requests,

            'urgent_requests': urgent_requests,

            'recipient_feedback': recipient_feedback,

            'item_count': item_count,

            'money_count': money_count,

            'subsidized_count': subsidized_count,

            'unread_notifications_count': unread_notifications_count,

            'unread_support_count': unread_support_count,

            'total_donations': total_donations,

            'recent_donations': recent_donations,

            'recent_donations_table': recent_donations_table,

            'direct_donations_table': direct_donations_table,

            'recent_activities': recent_activities,

            'impact_score': impact_score,

            'lives_impacted': recipients_impacted,

            'communities_served': foodbanks_supported,

            'volunteer_hours': total_donations * 2,  # Estimate 2 hours per donation

            

            # New Dashboard Sections

            'this_month_donations': this_month_donations,

            'this_month_amount': float(this_month_amount or 0),

            'most_recent_donation': most_recent_donation,

            'donation_streak': donation_streak,

            'achievements': achievements,

            'lives_impacted_count': lives_impacted,

            'meals_provided': meals_provided,

            'emergency_responses': emergency_responses,

            'smart_notifications': smart_notifications,

            

            

            # Phase 2 Chart Data

            'timeline_data': json.dumps(timeline_data),

            'donation_breakdown': json.dumps(donation_breakdown),

            'geographic_data': json.dumps(geographic_data),

            'response_trend': json.dumps(response_trend),

            'leaderboard': leaderboard,

            

            # Donation Tracking Tables

            'donor_unspecified_donations': donor_unspecified_donations,

            'donor_unspecified_counts': donor_unspecified_counts,

            'donor_subsidized_donations': donor_subsidized_donations,

            'donor_subsidized_counts': donor_subsidized_counts,

            'donor_request_donations': donor_request_donations,

            'donor_request_counts': donor_request_counts,

            'donor_tab_alerts': donor_tab_alerts,

        }

        return render(request, 'authentication/dashboard_donor.html', context)

    

    elif user_type == 'FOODBANK':

        # Get foodbank profile

        foodbank_profile = request.user.foodbank_profile

        

        # Check profile completion for notifications

        profile_completion = foodbank_profile.get_profile_completion_percentage()

        profile_incomplete = not foodbank_profile.is_profile_complete()

        

        # Get all donations to this foodbank

        donations = Donation.objects.filter(foodbank=foodbank_profile).order_by('-donated_at')

        

        # Calculate KPIs

        total_donations = donations.count()

        total_money = donations.filter(donation_type='money').aggregate(total=Sum('amount'))['total'] or 0

        total_subsidized = donations.filter(donation_type='subsidized').aggregate(total=Sum('subsidized_price'))['total'] or 0

        total_items = donations.filter(donation_type='item').count()

        

        # Unique donors

        unique_donors = donations.values('donor').distinct().count()

        

        # Active requests

        active_requests = FoodBankRequest.objects.filter(

            foodbank=foodbank_profile, 

            status='active'

        ).order_by('-priority', '-created_at')

        

        in_progress_requests = RecipientRequest.objects.filter(

        foodbank=foodbank_profile,

        status='in_progress'

    )



        # Fetch all recipient requests linked to this foodbank

        # Direct requests assigned to this foodbank

        direct_pending = RecipientRequest.objects.filter(

        foodbank=foodbank_profile,

        is_anonymous=False,

        status="pending"

        )

        

        # Anonymous requests that haven't been declined by this foodbank

        anonymous_pending = RecipientRequest.objects.filter(

            foodbank__isnull=True,

            is_anonymous=True,

            status="pending"

        ).exclude(

            declined_by=foodbank_profile  # Exclude already declined ones

        )

        

        # Combine pending requests

        pending_rec_requests = (direct_pending | anonymous_pending).order_by('-created_at')

        

        



        # Anonymous requests available to all foodbanks

        anonymous_requests = RecipientRequest.objects.filter(

            foodbank__isnull=True,

            is_anonymous=True,

            status="pending"

        ).order_by('-created_at')



        # Combine them for display

        linked_requests = (

            RecipientRequest.objects.filter(foodbank=foodbank_profile) |

            RecipientRequest.objects.filter(foodbank__isnull=True, is_anonymous=True)

        ).order_by('-created_at')



        # For stats only

        total_rec_requests = linked_requests.count()



        # Pending requests

        pending_rec_requests = (

            RecipientRequest.objects.filter(foodbank=foodbank_profile, status="pending") |

            RecipientRequest.objects.filter(foodbank__isnull=True, is_anonymous=True, status="pending")

        ).order_by('-created_at')





        #approved requests

        approved_requests = RecipientRequest.objects.filter(

            foodbank=foodbank_profile,

            status="accepted"

        ).order_by('-created_at')



        # Total requests (pending + approved)

        





        

        # Request statistics

        total_requests = FoodBankRequest.objects.filter(foodbank=foodbank_profile).count()

        fulfilled_requests = FoodBankRequest.objects.filter(foodbank=foodbank_profile, status='fulfilled').count()

        urgent_requests = active_requests.filter(priority='urgent').count()

        

        # Recent donations (last 7 days)

        recent_donations = donations.filter(

            donated_at__gte=timezone.now() - timedelta(days=7)

        ).count()

        

        # Donation type breakdown for charts

        donation_types = {

            'item': donations.filter(donation_type='item').count(),

            'money': donations.filter(donation_type='money').count(),

            'subsidized': donations.filter(donation_type='subsidized').count(),

        }

        

        # Request priority breakdown

        request_priorities = {

            'urgent': active_requests.filter(priority='urgent').count(),

            'high': active_requests.filter(priority='high').count(),

            'medium': active_requests.filter(priority='medium').count(),

            'low': active_requests.filter(priority='low').count(),

        }

        

        # Monthly trends data (last 6 months)

        monthly_donations = donations.annotate(

            month=TruncMonth('donated_at')

        ).values('month').annotate(

            count=Count('id')

        ).order_by('month')

        

        # Get last 6 months safely

        if monthly_donations.count() > 6:

            monthly_donations = list(monthly_donations)[-6:]

        else:

            monthly_donations = list(monthly_donations)

        

        # Prepare monthly data for chart

        monthly_labels = []

        monthly_data = []

        for item in monthly_donations:

            monthly_labels.append(item['month'].strftime('%b %Y'))

            monthly_data.append(item['count'])

        

        # Recent activity (last 5 donations)

        recent_activity = donations[:5]

        

        # Unread notifications

        unread_notifications_count = Notification.objects.filter(user=request.user, is_read=False).count()

        

        # Foodbank support badge:
        # - Do not count newly-created tickets from the foodbank itself.
        # - Count only threads where admin has responded (reply/admin_response),
        #   and hide again when the foodbank follows up (status goes back to 'new').
        from .models import SupportMessage

        unread_support_count = SupportMessage.objects.filter(
            user=request.user
        ).filter(
            (Q(admin_response__isnull=False) & ~Q(admin_response='')) |
            Q(replies__is_from_admin=True)
        ).exclude(
            status='new'
        ).distinct().count()

        

        # Estimated recipients served (based on donations)

        estimated_recipients = total_donations * 3  # Rough estimate

        

        # Get recent requests for the dashboard table

        recent_requests = FoodBankRequest.objects.filter(

            foodbank=foodbank_profile

        ).order_by('-created_at')[:10]



        direct_requests_queryset = FoodBankRequest.objects.filter(

            foodbank=foodbank_profile,

            original_request__isnull=True

        ).select_related('foodbank', 'linked_request_management').prefetch_related(

            'donations__donor__donor_profile',

            'donations__allocations'

        )

        recent_direct_requests = list(

            direct_requests_queryset.order_by('-created_at')[:5]

        )

        direct_requests_count = direct_requests_queryset.count()



        delivery_label_map = dict(Donation.DELIVERY_METHODS)

        for direct_req in recent_direct_requests:

            linked = getattr(direct_req, 'linked_request_management', None)

            direct_req.status_label = direct_req.get_foodbank_requests_status_label()

            direct_req.quantity_display = direct_req.get_requested_quantity_display()

            direct_req.progress_percent = direct_req.get_fulfillment_percentage()

            direct_req.total_received = direct_req.get_total_donations_received()

            direct_req.stock_is_monetary = False

            if not (direct_req.total_received and getattr(direct_req, 'quantity_needed', None)):

                money_donations = direct_req.donations.filter(donation_type='money', status='accepted')

                if money_donations.exists():

                    from decimal import Decimal

                    total_money = sum(Decimal(str(d.amount or 0)) for d in money_donations)

                    if total_money > 0:

                        direct_req.total_received = float(total_money)

                        direct_req.stock_is_monetary = True

            used_from_allocations = 0
            used_pieces = 0
            for donation in direct_req.donations.all():
                for allocation in donation.allocations.filter(declined_by_recipient=False):
                    used_from_allocations += (allocation.quantity or allocation.amount or 0)
                    used_pieces += (allocation.quantity or 0)

            direct_req_quantity_needed = getattr(direct_req, 'quantity_needed', None)
            if direct_req_quantity_needed is not None:
                stock_capacity = direct_req_quantity_needed
                if direct_req.total_received:
                    stock_capacity = min(direct_req_quantity_needed, direct_req.total_received)
                direct_req.stock_used = min(used_pieces, stock_capacity)
                direct_req.stock_required = stock_capacity
                direct_req.stock_remaining = max(0, stock_capacity - direct_req.stock_used)
                direct_req.stock_is_monetary = False
            else:
                stock_capacity = direct_req.total_received or 0
                direct_req.stock_required = stock_capacity
                direct_req.stock_used = min(used_from_allocations, stock_capacity) if stock_capacity else 0
                direct_req.stock_remaining = max(0, stock_capacity - direct_req.stock_used) if stock_capacity else 0



            donor_delivery_methods = []

            type_labels = []

            for donation in direct_req.donations.all():

                method = getattr(donation, 'delivery_method', None)

                if method and method not in donor_delivery_methods:

                    donor_delivery_methods.append(method)



                if donation.donation_type == 'item':

                    type_label = 'Free Goods'

                elif donation.donation_type == 'money' and getattr(donation.foodbank_request, 'original_request', None):

                    type_label = donation.foodbank_request.original_request.get_request_type_display()

                else:

                    type_label = donation.get_donation_type_display()



                if type_label and type_label not in type_labels:

                    type_labels.append(type_label)



            if donor_delivery_methods:

                direct_req.delivery_display = ", ".join(

                    delivery_label_map.get(method, method.replace('_', ' ').title())

                    for method in donor_delivery_methods

                )

            else:

                linked_delivery = linked.get_delivery_method_display() if linked and getattr(linked, 'delivery_method', None) else None
                request_delivery = direct_req.get_delivery_method_display() if getattr(direct_req, 'delivery_method', None) else None
                direct_req.delivery_display = linked_delivery or request_delivery or '-'



            direct_req.category_display = ", ".join(type_labels) if type_labels else direct_req.get_donation_type_display()

            direct_req.location_display = (

                getattr(linked, 'location', None)

                or getattr(direct_req.foodbank, 'address', None)

                or '-'

            )

            direct_req.notes_display = getattr(linked, 'additional_notes', None) or direct_req.description



        

        # Get pending requests count

        pending_requests = FoodBankRequest.objects.filter(

            foodbank=foodbank_profile,

            status='PENDING'

        ).count()



        # Get recent requests from RequestManagement model for dashboard display

        # Apply status filter if provided

        status_filter = request.GET.get('status', 'all')



        # Include: Direct requests + Assigned anonymous + Unassigned anonymous

        recent_manage_requests = RequestManagement.objects.filter(

            Q(foodbank=foodbank_profile) |  # Direct requests

            Q(is_anonymous=True, assigned_foodbank=foodbank_profile) |  # Assigned anonymous

            Q(is_anonymous=True, assigned_foodbank__isnull=True)  # Unassigned anonymous

        ).prefetch_related('foodbank_request_created__donations', 'donor_requests__donations')



        if status_filter and status_filter != 'all':

            recent_manage_requests = recent_manage_requests.filter(status=status_filter)



        recent_manage_requests = recent_manage_requests.order_by('-time_of_request')[:5]

        for req in recent_manage_requests:
            req.fulfillment_breakdown = _get_request_fulfillment_breakdown(req)
            req.quantity_timeline = _build_request_quantity_timeline(req)
            req.quantity_timeline_has_declines = any(entry.get('is_declined') for entry in req.quantity_timeline)



        # Get request counts for dashboard stats (include anonymous requests)

        base_queryset = RequestManagement.objects.filter(

            Q(foodbank=foodbank_profile) |

            Q(is_anonymous=True, assigned_foodbank=foodbank_profile) |

            Q(is_anonymous=True, assigned_foodbank__isnull=True)

        )



        manage_request_counts = {

            'total': base_queryset.count(),

            'pending': base_queryset.filter(status='pending').count(),

            'fulfilled': base_queryset.filter(status='fulfilled').count(),

            'partial': base_queryset.filter(status='partial').count(),

            'denied': base_queryset.filter(status='declined').count(),  # Note: use 'declined'

        }



        unspecified_pending_count = UnspecifiedDonationManagement.objects.filter(

            donation__foodbank=foodbank_profile,

            foodbank_status='pending_foodbank'

        ).count()



        subsidized_pending_count = Donation.objects.filter(

            foodbank=foodbank_profile,

            donation_type='subsidized',

            status='pending',

            foodbank_request__isnull=True

        ).count()



        # Direct tab: count donations to this foodbank's direct requests awaiting foodbank response (accept/decline)

        direct_pending_donations_count = Donation.objects.filter(

            foodbank_request__foodbank=foodbank_profile,

            foodbank_request__original_request__isnull=True,

            status='pending'

        ).count()



        foodbank_tab_alerts = {

            'requests': manage_request_counts['pending'],

            'direct': direct_pending_donations_count,

            'unspecified': unspecified_pending_count,

            'subsidized': subsidized_pending_count,

        }
        # Subsidized preview table: attach decline_responses for "Show all rejections"
        subsidized_donations = list(Donation.objects.filter(
            foodbank=foodbank_profile,
            donation_type='subsidized',
            foodbank_request__isnull=True
        ).select_related('donor', 'donor__donor_profile', 'accepted_by_recipient').order_by('-donated_at')[:5])
        if subsidized_donations:
            sub_ids = [d.id for d in subsidized_donations]
            decline_by_donation = {}
            for r in DonationResponse.objects.filter(
                donation_id__in=sub_ids, response_type='declined'
            ).select_related('recipient').order_by('-responded_at'):
                decline_by_donation.setdefault(r.donation_id, []).append(r)
            for d in subsidized_donations:
                d.decline_responses = decline_by_donation.get(d.id, [])

            latest_notes = {}
            for response in DonationResponse.objects.filter(
                donation_id__in=sub_ids
            ).exclude(notes__isnull=True).exclude(notes__exact='').order_by('-responded_at'):
                if response.donation_id not in latest_notes:
                    latest_notes[response.donation_id] = response.notes
            for d in subsidized_donations:
                d.latest_recipient_note = latest_notes.get(d.id)
                donor_note = (
                    (getattr(d, 'message', None) or '').strip()
                    or (getattr(d, 'csr_description', None) or '').strip()
                    or (getattr(d, 'other_description', None) or '').strip()
                )
                recipient_note = (d.latest_recipient_note or '').strip() if d.latest_recipient_note else ''
                foodbank_decline_note = (getattr(d, 'decline_message', None) or '').strip()

                recipient_rejection_entries = []
                for resp in getattr(d, 'decline_responses', []):
                    who = (
                        getattr(getattr(resp, 'recipient', None), 'full_name', None)
                        or getattr(getattr(getattr(resp, 'recipient', None), 'user', None), 'email', None)
                        or 'Recipient'
                    )
                    when = resp.responded_at.strftime('%b %d, %Y %H:%M') if getattr(resp, 'responded_at', None) else ''
                    note = (getattr(resp, 'notes', None) or '').strip() or 'No note provided.'
                    if when:
                        recipient_rejection_entries.append(f"{who} ({when}): {note}")
                    else:
                        recipient_rejection_entries.append(f"{who}: {note}")

                d.donor_note_display = donor_note or 'No donor note'
                d.recipient_note_display = recipient_note or 'No recipient note'
                d.foodbank_decline_note_display = foodbank_decline_note or 'No foodbank rejection note'
                d.recipient_rejection_notes_display = "\n".join(recipient_rejection_entries) if recipient_rejection_entries else 'No recipient rejection notes'
                d.has_foodbank_decline_note = bool(foodbank_decline_note)
                d.has_recipient_rejection_notes = bool(recipient_rejection_entries)

        unspecified_preview_donations = list(
            UnspecifiedDonationManagement.objects.filter(
                donation__foodbank=foodbank_profile
            ).select_related(
                'donation',
                'donation__donor',
                'donation__donor__donor_profile',
                'accepted_by_recipient',
            ).order_by('-created_at')[:5]
        )
        unspecified_preview_note_maps = _build_unspecified_response_note_maps(unspecified_preview_donations)
        for item in unspecified_preview_donations:
            donation = item.donation
            donor = getattr(donation, 'donor', None)
            donor_profile = getattr(donor, 'donor_profile', None) if donor else None

            organization_name = (getattr(donor_profile, 'organization_name', None) or '').strip() if donor_profile else ''
            profile_full_name = (getattr(donor_profile, 'full_name', None) or '').strip() if donor_profile else ''
            donor_full_name = ''
            if donor:
                try:
                    donor_full_name = (donor.get_full_name() or '').strip()
                except Exception:
                    donor_full_name = ''
            item.donor_display = organization_name or profile_full_name or donor_full_name or getattr(donor, 'email', 'Unknown donor')

            donor_note = (
                (getattr(donation, 'message', None) or '').strip()
                or (getattr(donation, 'csr_description', None) or '').strip()
                or (getattr(donation, 'other_description', None) or '').strip()
            )
            recipient_note, recipient_decline_note = _resolve_unspecified_recipient_notes(item, unspecified_preview_note_maps)
            foodbank_decline_note = (getattr(item, 'foodbank_decline_reason', None) or '').strip()

            item.effective_recipient_note = recipient_note
            item.effective_recipient_decline_note = recipient_decline_note
            item.donor_note_display = donor_note or 'No donor note'
            item.recipient_note_display = recipient_note or 'No recipient note'
            item.foodbank_decline_note_display = foodbank_decline_note or 'No foodbank decline note'
            item.recipient_decline_note_display = recipient_decline_note or 'No recipient decline note'
            item.has_foodbank_decline_note = bool(foodbank_decline_note)
            item.has_recipient_decline_note = bool(recipient_decline_note)
        context = {

            'foodbank_profile': foodbank_profile,

            'donations': donations,

            'active_requests': active_requests,

            'recent_requests': recent_requests,

            'recent_direct_requests': recent_direct_requests,

            'direct_foodbank_requests': recent_direct_requests,

            'direct_foodbank_requests_count': direct_requests_count,

            'recent_activity': recent_activity,

            'linked_requests': linked_requests,

            "approved_requests": approved_requests,

            "total_rec_requests": total_rec_requests,

            'pending_rec_requests': pending_rec_requests,

            'in_progress_requests': in_progress_requests,

            

            # KPIs for new dashboard

            'total_donations': total_donations,

            'total_requests': total_requests,

            'urgent_requests': urgent_requests,

            'pending_requests': pending_requests,

            

            # Legacy data (keeping for compatibility)

            'total_money': total_money,

            'total_subsidized': total_subsidized,

            'total_items': total_items,

            'unique_donors': unique_donors,

            'fulfilled_requests': fulfilled_requests,

            'recent_donations': recent_donations,

            'estimated_recipients': estimated_recipients,

            

            # Chart data

            'donation_types': donation_types,

            'request_priorities': request_priorities,

            'monthly_labels': monthly_labels,

            'monthly_data': monthly_data,

            

            # Notifications

            'unread_notifications_count': unread_notifications_count,

            'unread_support_count': unread_support_count,

            

            # Profile completion

            'profile_completion': profile_completion,

            'profile_incomplete': profile_incomplete,

            

            # Request Management data

            'recent_manage_requests': recent_manage_requests,

            'manage_request_counts': manage_request_counts,

            'current_full_path': request.get_full_path(),

            

            # Unspecified Donations Management data - show ALL donations (not just pending)

            'unspecified_donations': unspecified_preview_donations,

            'unspecified_donations_count': UnspecifiedDonationManagement.objects.filter(

                donation__foodbank=foodbank_profile

            ).count(),

            'unspecified_pending_count': unspecified_pending_count,

            
            # Subsidized Donations Management data (preview table with decline_responses for "Show all rejections")
            'subsidized_donations': subsidized_donations,
            'subsidized_donations_count': Donation.objects.filter(

                foodbank=foodbank_profile,

                donation_type='subsidized',

                foodbank_request__isnull=True

            ).count(),

            'subsidized_pending_count': subsidized_pending_count,

            'foodbank_tab_alerts': foodbank_tab_alerts,

        }

        return render(request, template_name, context)

    

    

    elif user_type == 'RECIPIENT':

        # Get recipient profile with proper error handling

        try:

            recipient_profile = request.user.recipient_profile

        except RecipientProfile.DoesNotExist:

            messages.error(request, 'Recipient profile not found. Please contact support.')

            return redirect('home')

        except AttributeError:

            messages.error(request, 'Invalid user account. Please contact support.')

            return redirect('home')

        

        # Get recipient's requests

        recipient_requests = RecipientRequest.objects.filter(recipient=recipient_profile).order_by('-created_at')

        total_recipient_requests = recipient_requests.count()



        # Get donations allocated to this recipient

        try:

            donations_allocated = DonationAllocation.objects.filter(recipient=recipient_profile)

            total_donations = donations_allocated.count()

        except Exception:

            total_donations = 0

        

        # Get available food banks (only approved and active ones)

        foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')[:4]

        

        # Get recent notifications

        recent_notifications = Notification.objects.filter(

            user=request.user, is_read=False

        ).order_by('-created_at')[:5]



        # Get notifications count

        unread_count = Notification.objects.filter(user=request.user, is_read=False).count()



        # Get unacknowledged donations

        try:

            unacknowledged_allocations = DonationAllocation.objects.filter(

                recipient=recipient_profile,

                is_acknowledged=False

            ).order_by('-allocated_at')[:4]

            # Force evaluation to catch DB errors early

            list(unacknowledged_allocations)

        except Exception:

            unacknowledged_allocations = []



        # Get recent requests from RequestManagement model for dashboard display

        recent_requests = RequestManagement.objects.filter(

            recipient=recipient_profile

        ).select_related('foodbank', 'updated_by', 'foodbank_request', 'donation').prefetch_related(
            'foodbank_request__donations',
            'foodbank_request_created__donations',
            'donor_requests__donations',
            'donations',
            'donation_allocations__donation'
        ).order_by('-time_of_request')[:5]

        for req in recent_requests:
            req.fulfillment_breakdown = _get_request_fulfillment_breakdown(req)
            raw_timeline = _build_request_quantity_timeline(req)
            req.quantity_timeline, req.quantity_timeline_has_declines = _build_recipient_timeline(raw_timeline)
            req.primary_donation = _get_request_primary_donation(req)
            if req.primary_donation:
                req.primary_donor_note = (
                    req.primary_donation.message
                    or req.primary_donation.csr_description
                    or req.primary_donation.other_description
                    or ""
                )
            else:
                req.primary_donor_note = ""

        

        # Get request counts for dashboard stats

        request_counts = {

            'total': RequestManagement.objects.filter(recipient=recipient_profile).count(),

            'pending': RequestManagement.objects.filter(recipient=recipient_profile, status='pending').count(),

            'fulfilled': RequestManagement.objects.filter(recipient=recipient_profile, status='fulfilled').count(),

            'partial': RequestManagement.objects.filter(recipient=recipient_profile, status='partial').count(),

            'denied': RequestManagement.objects.filter(recipient=recipient_profile, status='denied').count(),

        }

        

        # Recipient support badge:
        # - Do not count brand-new tickets created by the recipient.
        # - Count tickets once the recipient has followed up (replied) or when the case is resolved/closed.
        from .models import SupportMessage

        unread_support_count = SupportMessage.objects.filter(
            user=request.user
        ).filter(
            Q(replies__is_from_admin=False) |
            Q(status__in=['resolved', 'closed'])
        ).distinct().count()

        

        # Get unspecified donations for organization recipients - show ALL with status

        available_unspecified_donations = []

        available_unspecified_count = 0

        my_accepted_donations = []

        my_accepted_count = 0

        pending_available_count = 0

        

        if recipient_profile.is_organization:

            # Available donations for this recipient.
            # Rebroadcasted entries remain available unless this same recipient declined them.
            available_unspecified_qs = _annotate_recipient_unspecified_decline_state(
                UnspecifiedDonationManagement.objects.filter(
                    foodbank_status='accepted_by_foodbank',
                    recipient_status__in=['pending_recipient', 'declined_by_recipient'],
                ),
                recipient_profile,
            ).filter(
                Q(recipient_status='pending_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
                Q(recipient_status='declined_by_recipient', _declined_by_me=False, _declined_by_me_legacy=False)
            ).select_related('donation', 'donation__donor', 'donation__foodbank').order_by('-created_at')

            # Preview list is a single newest-first stream across all relevant states.
            # This keeps table order strictly latest-to-oldest.
            preview_unspecified_qs = _annotate_recipient_unspecified_decline_state(
                UnspecifiedDonationManagement.objects.filter(
                    foodbank_status='accepted_by_foodbank',
                ),
                recipient_profile,
            ).filter(
                Q(recipient_status='pending_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
                Q(recipient_status='declined_by_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
                (Q(_declined_by_me=True) | Q(_declined_by_me_legacy=True)) |
                Q(accepted_by_recipient=recipient_profile, recipient_status='accepted_by_recipient') |
                Q(accepted_by_recipient=recipient_profile, recipient_status='received')
            ).select_related('donation', 'donation__donor', 'donation__foodbank').order_by(
                '-foodbank_reviewed_at', '-created_at', '-id'
            )



            available_unspecified_donations = _attach_recipient_unspecified_effective_status(
                list(preview_unspecified_qs[:5]),
                recipient_profile,
            )

            

            available_unspecified_count = available_unspecified_qs.count()

            pending_available_count = available_unspecified_count

            my_accepted_qs = UnspecifiedDonationManagement.objects.filter(

                accepted_by_recipient=recipient_profile

            ).select_related('donation', 'donation__donor', 'donation__foodbank').order_by('-created_at')

            # Preview table is now fully driven by available_unspecified_donations (already newest-first).
            my_accepted_donations = []

            

            my_accepted_count = my_accepted_qs.count()

        

        # Subsidized donations for all recipients

        # Available subsidized donations (accepted by foodbank, not yet claimed)

        # Only show unspecified subsidized donations (donor-initiated, not linked to a request)

        available_subsidized_qs = Donation.objects.filter(

            donation_type='subsidized',

            status='accepted',

            accepted_by_recipient__isnull=True,

            foodbank_request__isnull=True  # Exclude specified donations

        ).select_related('donor', 'donor__donor_profile', 'foodbank').order_by('-donated_at')



        available_subsidized_donations = list(available_subsidized_qs[:5])

        subsidized_remaining_slots = max(0, 5 - len(available_subsidized_donations))
        

        available_subsidized_count = available_subsidized_qs.count()

        

        # My accepted subsidized donations (only unspecified - donor-initiated)

        my_subsidized_qs = Donation.objects.filter(

            donation_type='subsidized',

            accepted_by_recipient=recipient_profile,

            foodbank_request__isnull=True  # Exclude specified donations

        ).select_related('donor', 'donor__donor_profile', 'foodbank').order_by('-donated_at')



        if subsidized_remaining_slots:

            my_subsidized_donations = list(my_subsidized_qs[:subsidized_remaining_slots])

        else:

            my_subsidized_donations = []

        # Enrich subsidized preview rows with latest notes so preview modal mirrors "View All" page.
        subsidized_preview_donations = available_subsidized_donations + my_subsidized_donations
        if subsidized_preview_donations:
            preview_ids = [d.id for d in subsidized_preview_donations]
            latest_recipient_notes = {}
            latest_responses = (
                DonationResponse.objects
                .filter(donation_id__in=preview_ids)
                .exclude(notes__isnull=True)
                .exclude(notes__exact='')
                .order_by('-responded_at')
            )
            for response in latest_responses:
                if response.donation_id not in latest_recipient_notes:
                    latest_recipient_notes[response.donation_id] = response.notes

            decline_reasons = {}
            declined_responses = (
                DonationResponse.objects
                .filter(
                    donation_id__in=preview_ids,
                    recipient=recipient_profile,
                    response_type='declined',
                )
                .exclude(notes__isnull=True)
                .exclude(notes__exact='')
                .order_by('-responded_at')
            )
            for response in declined_responses:
                if response.donation_id not in decline_reasons:
                    decline_reasons[response.donation_id] = response.notes

            for donation in subsidized_preview_donations:
                donation.latest_recipient_note = latest_recipient_notes.get(donation.id) or ''
                if not getattr(donation, 'recipient_decline_reason', None):
                    donation.recipient_decline_reason = decline_reasons.get(donation.id) or ''

        

        my_subsidized_count = my_subsidized_qs.count()

        

        current_hour = timezone.localtime().hour

        if current_hour < 12:

            greeting_text = "Good morning"

        elif current_hour < 18:

            greeting_text = "Good afternoon"

        else:

            greeting_text = "Good evening"



        recipient_tab_alerts = {

            'requests': 0,

            'unspecified': available_unspecified_count,

            'subsidized': available_subsidized_count,

        }



        context = {

            'recipient_profile': recipient_profile,

            'total_recipient_requests': total_recipient_requests,

            'total_donations': total_donations,  # FIX: Added missing context variable

            'foodbanks': foodbanks,

            'recent_notifications': recent_notifications,

            'unread_count': unread_count,

            'unacknowledged_allocations': unacknowledged_allocations,

            'recent_requests': recent_requests,

            'request_counts': request_counts,

            'unread_support_count': unread_support_count,  # FIX: Added missing context variable

            

            # Unspecified donations for organization recipients

            'available_unspecified_donations': available_unspecified_donations,

            'available_unspecified_count': available_unspecified_count,

            'my_accepted_donations': my_accepted_donations,

            'my_accepted_count': my_accepted_count,

            

            # Subsidized donations for all recipients

            'available_subsidized_donations': available_subsidized_donations,

            'available_subsidized_count': available_subsidized_count,

            'my_subsidized_donations': my_subsidized_donations,

            'my_subsidized_count': my_subsidized_count,

            'greeting_text': greeting_text,

            'recipient_tab_alerts': recipient_tab_alerts,

        }

        return render(request, template_name, context)

    

    elif user_type == 'ADMIN':

        # Get system statistics for admin dashboard

        total_users = CustomUser.objects.count()

        total_donations = Donation.objects.count()

        total_foodbanks = FoodBankProfile.objects.count()

        total_recipients = RecipientProfile.objects.count()

        

        # Get pending foodbank registrations count

        pending_foodbanks_count = FoodBankProfile.objects.filter(is_approved='pending').count()

        

        context = {

            'total_users': total_users,

            'total_donations': total_donations,

            'total_foodbanks': total_foodbanks,

            'total_recipients': total_recipients,

            'pending_foodbanks_count': pending_foodbanks_count,

        }

        return render(request, template_name, context)

    

    return render(request, template_name)





@login_required

def view_approved_requests(request):

    """View all approved recipient requests for a foodbank"""

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    

    # Get all approved requests

    approved_requests = RecipientRequest.objects.filter(

        foodbank=foodbank_profile,

        status="accepted"

    ).select_related('recipient').order_by('-created_at')



    

    context = {

        'approved_requests': approved_requests,

        'foodbank_profile': foodbank_profile,

    }

    return render(request, 'authentication/approved_requests.html', context)



@login_required

def fulfill_recipient_request(request, pk):

    """Fulfill an approved recipient request by allocating donations"""

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    recipient_request = get_object_or_404(

        RecipientRequest, 

        pk=pk, 

        foodbank=foodbank_profile,

        status="accepted"

    )

    

    if request.method == 'POST':

        allocation_type = request.POST.get('allocation_type')

        notes = request.POST.get('notes', '')

        

        if allocation_type == 'existing_donation':

            # Allocate from an existing donation

            donation_id = request.POST.get('donation_id')

            quantity = request.POST.get('quantity')

            

            try:

                donation = Donation.objects.get(id=donation_id, foodbank=foodbank_profile)

                

                # Create allocation

                allocation = DonationAllocation.objects.create(

                    donation=donation,

                    recipient=recipient_request.recipient,

                    quantity=quantity if donation.donation_type == 'item' else None,

                    amount=quantity if donation.donation_type in ['money', 'subsidized'] else None,

                    notes=notes

                )

                

                # Mark request as completed

                recipient_request.status = 'completed'

                recipient_request.save()

                

                # Notify recipient

                Notification.objects.create(

                    user=recipient_request.recipient.user,

                    notification_type='request_fulfilled',

                    message=f'Your request "{recipient_request.title}" has been fulfilled by {foodbank_profile.foodbank_name}. Please acknowledge receipt when you receive the donation.'

                )

                

                messages.success(request, 'Request fulfilled successfully! The recipient has been notified.')

                return redirect('dashboard')

                

            except Donation.DoesNotExist:

                messages.error(request, 'Selected donation not found.')

        

        elif allocation_type == 'manual':

            # Manual fulfillment without linking to specific donation

            item_description = request.POST.get('item_description')

            quantity_fulfilled = request.POST.get('quantity_fulfilled')

            

            # Update request with fulfillment details

            recipient_request.fulfillment_notes = f"{item_description} - Quantity: {quantity_fulfilled}. {notes}"

            recipient_request.status = 'completed'

            recipient_request.save()

            

            # Notify recipient

            Notification.objects.create(

                user=recipient_request.recipient.user,

                notification_type='request_fulfilled',

                message=f'Your request "{recipient_request.title}" has been fulfilled by {foodbank_profile.foodbank_name}. Details: {item_description}'

            )

            

            messages.success(request, 'Request marked as fulfilled! The recipient has been notified.')

            return redirect('dashboard')

    

    # Get available donations for allocation

    available_donations = Donation.objects.filter(

        foodbank=foodbank_profile,

        donated_at__gte=timezone.now() - timedelta(days=30)  # Recent donations

    ).order_by('-donated_at')

    

    context = {

        'recipient_request': recipient_request,

        'available_donations': available_donations,
        'available_donations_page': available_donations_page,

        'foodbank_profile': foodbank_profile,

    }

    return render(request, 'authentication/fulfill_recipient_request.html', context)



@login_required

def mark_request_completed(request, pk):

    """Mark a request as completed"""

    if request.user.user_type != 'FOODBANK':

        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)

    

    foodbank_profile = request.user.foodbank_profile

    

    try:

        recipient_request = RecipientRequest.objects.get(

            pk=pk, 

            foodbank=foodbank_profile,

            status="accepted"

        )

        

        recipient_request.status = 'completed'

        recipient_request.save()

        

        # Notify recipient

        Notification.objects.create(

            user=recipient_request.recipient.user,

            notification_type='request_fulfilled',

            message=f'Your request "{recipient_request.title}" has been marked as completed by {foodbank_profile.foodbank_name}.'

        )

        

        return JsonResponse({'success': True})

    except RecipientRequest.DoesNotExist:

        return JsonResponse({'success': False, 'error': 'Request not found'}, status=404)

    

@login_required

def create_recipient_request(request):

    if request.method == 'POST':

        form = RecipientRequestForm(request.POST)

        if form.is_valid():

            req = form.save(commit=False)

            req.recipient = request.user.recipient_profile

            req.save()

            messages.success(request, "Request submitted successfully!")

            return redirect('recipient_requests_view')

    else:

        form = RecipientRequestForm()

    return render(request, 'authentication/create_recipient_request.html', {'form': form})



# REMOVED: recipient_requests_list - Use recipient_requests_view instead





@login_required

def view_recipient_request_detail(request, pk):

    """View details of a specific recipient request"""

    req = get_object_or_404(RecipientRequest, pk=pk, recipient=request.user.recipient_profile)

    

    context = {

        'request_obj': req,

    }

    return render(request, 'recipient/request_detail.html', context)



@login_required

def edit_recipient_request(request, pk):

    """Edit an existing recipient request"""

    req = get_object_or_404(RecipientRequest, pk=pk, recipient=request.user.recipient_profile)

    

    if request.method == 'POST':

        form = RecipientRequestForm(request.POST, instance=req)

        if form.is_valid():

            form.save()

            messages.success(request, 'Request updated successfully!')

            return redirect('recipient_requests_view')

        else:

            messages.error(request, 'Please correct the errors below.')

    else:

        form = RecipientRequestForm(instance=req)

    

    context = {

        'form': form,

        'request_obj': req,

    }

    return render(request, 'recipient/edit_request.html', context)



@login_required

def delete_recipient_request(request, pk):

    """Delete a recipient request"""

    req = get_object_or_404(RecipientRequest, pk=pk, recipient=request.user.recipient_profile)

    

    if request.method == 'POST':

        req.delete()

        messages.success(request, 'Request deleted successfully!')

        return redirect('recipient_requests_view')

    return render(request, 'recipient/delete_confirm.html', context)

SUPPORT_MESSAGE_MIN_LENGTH = 15


@login_required

def recipient_contact_support(request):

    """Contact support page for recipients"""

    if request.user.user_type != 'RECIPIENT':

        return redirect('dashboard')

    

    if request.method == 'POST':

        subject = request.POST.get('subject')

        message_text = request.POST.get('message')

        

        if subject and message_text and len(message_text.strip()) >= SUPPORT_MESSAGE_MIN_LENGTH:

            # Create support message in database

            from .models import SupportMessage

            SupportMessage.objects.create(

                user=request.user,

                subject=subject,

                message=message_text.strip(),

            )

            messages.success(request, 'Your message has been sent to our support team. We will get back to you soon!')

            return redirect('recipient_contact_support')

        else:

            messages.error(
                request,
                f'Please fill in all fields and ensure your message is at least {SUPPORT_MESSAGE_MIN_LENGTH} characters long.'
            )

    

    context = {

        'recipient_profile': request.user.recipient_profile

    }

    return render(request, 'recipient/contact_support.html', context)



@login_required

def recipient_privacy_settings(request):

    """Privacy settings page for recipients"""

    if request.user.user_type != 'RECIPIENT':

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if request.method == 'POST':

        # Update privacy settings

        recipient_profile.allow_public_profile = request.POST.get('allow_public_profile') == 'on'

        recipient_profile.show_donation_history = request.POST.get('show_donation_history') == 'on'

        recipient_profile.receive_email_notifications = request.POST.get('receive_email_notifications') == 'on'

        recipient_profile.save()

        

        messages.success(request, 'Privacy settings updated successfully!')

        return redirect('recipient_privacy_settings')

    

    context = {

        'recipient_profile': recipient_profile

    }

    return render(request, 'recipient/privacy_settings.html', context)

@login_required

def available_random_donations(request):

    donations = Donation.objects.filter(foodbank_request__isnull=True)

    print("Available Donations:", donations)

    return render(request, 'authentication/available_random_donations.html', {'donations': donations})



@login_required

def available_foodbanks(request):

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True).order_by('foodbank_name')

    

    print("DEBUG â†’ Foodbanks count:", foodbanks.count())

    for fb in foodbanks:

        print("Foodbank:", fb.foodbank_name, "| Contact:", fb.contact_person)



    return render(request, "authentication/available_foodbanks.html", {"foodbanks": foodbanks})



@login_required

def accept_random_donation(request, donation_id):

    donation = get_object_or_404(Donation, id=donation_id, foodbank_request__isnull=True)



    if request.method == "POST":

        DonationAllocation.objects.create(

            donation=donation,

            recipient=request.user.recipient_profile,  # if you use RecipientProfile FK

            quantity=donation.quantity,

            is_acknowledged=False

        )

        # Remove from available list (by deleting or marking as allocated)

        donation.foodbank_request = None  # keep it as random donation

        donation.is_allocated = True  # add this field in your Donation model

        donation.save()



        messages.success(request, "You have successfully accepted this random donation.")

        return redirect("available_random_donations")





@login_required

def acknowledge_donation(request, allocation_id):

    allocation = get_object_or_404(

        DonationAllocation,

        id=allocation_id,

        recipient=request.user.recipient_profile

    )



    if request.method == 'POST' and not allocation.is_acknowledged:

        allocation.is_acknowledged = True

        allocation.save()



        # Recipient's display name (fall back to email if full name missing)

        recipient_name = allocation.recipient.full_name or request.user.email



        # Donation display name (fall back to donation type)

        donation_name = allocation.donation.item_name or allocation.donation.get_donation_type_display()



        # Create personalized notification for the foodbank

        Notification.objects.create(

            user=allocation.donation.foodbank.user,

            notification_type='acknowledgement',

            message=f"{recipient_name} has acknowledged receipt of '{donation_name}' from you. Thank you!"

        )

        

        # Create notification for the donor

        Notification.objects.create(

            user=allocation.donation.donor,

            notification_type='acknowledgement',

            message=f"{recipient_name} has received and acknowledged your donation of '{donation_name}'. Thank you for making a difference!"

        )



        # Success message for the recipient

        messages.success(request, "Donation acknowledged successfully! Donor and food bank have been notified.")



    return redirect('dashboard')



@login_required

def recipient_notifications(request):

    print("Logged-in user:", request.user.email)

    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')

    print("Notifications:", notifications)

    notifications.filter(is_read=False).update(is_read=True)

    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()

    return render(

        request, 'authentication/recipient_notifications.html',

        {'notifications': notifications, 'unread_count': unread_count})

@login_required

def mark_recipient_notifications_read(request, pk):

    notification = get_object_or_404(Notification, pk=pk, user=request.user)

    notification.is_read = True

    notification.save()

    return redirect('recipient_notifications')



# ussd view

from django.views.decorators.csrf import csrf_exempt



from django.views.decorators.csrf import csrf_exempt

from django.http import HttpResponse



@csrf_exempt  # USSD requests wonâ€™t have CSRF token

def ussd_callback(request):

    session_id   = request.POST.get("sessionId", "")

    service_code = request.POST.get("serviceCode", "")

    phone_number = request.POST.get("phoneNumber", "")

    text         = request.POST.get("text", "")



    if text == "":

        response = "CON Welcome to FoodBank Hub\n1. View Requests\n2. My Status"

    elif text == "1":

        response = "END You have 2 active requests"

    elif text == "2":

        response = "END Your last donation was approved"

    else:

        response = "END Invalid choice"



    return HttpResponse(response, content_type="text/plain")





@login_required

def create_testimonial(request):

    """Recipients can create testimonials with impact photos"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Only recipients can submit testimonials.')

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = TestimonialForm(request.POST, request.FILES)

        if form.is_valid():

            testimonial = form.save(commit=False)

            testimonial.recipient = request.user.recipient_profile

            testimonial.approval_status = 'pending'

            testimonial.save()

            messages.success(

                request, 

                "Testimonial submitted successfully! It will be reviewed by an admin before being displayed publicly."

            )

            return redirect('recipient_testimonials_list')

    else:

        form = TestimonialForm()

    return render(request, 'authentication/create_testimonial.html', {'form': form})



@login_required

def recipient_testimonials_list(request):

    """List all testimonials for the logged-in recipient"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')



    pending_testimonials = Testimonial.objects.filter(

        recipient=request.user.recipient_profile,

        approval_status='pending'

    ).order_by('-created_at')



    approved_qs = Testimonial.objects.filter(

        recipient=request.user.recipient_profile,

        approval_status='approved'

    ).order_by('-created_at')



    displayed_testimonials = [t for t in approved_qs if t.is_currently_displayed()]

    archived_approved = [t for t in approved_qs if not t.is_currently_displayed()]



    rejected_qs = Testimonial.objects.filter(

        recipient=request.user.recipient_profile,

        approval_status='rejected'

    ).order_by('-created_at')



    archived_testimonials = archived_approved

    rejected_testimonials = rejected_qs



    context = {

        'pending_testimonials': pending_testimonials,

        'displayed_testimonials': displayed_testimonials,

        'archived_testimonials': archived_testimonials,

        'rejected_testimonials': rejected_testimonials,

    }

    return render(request, 'authentication/recipient_testimonials_list.html', context)



@login_required

def edit_testimonial(request, testimonial_id):

    """Edit a testimonial (only if pending or rejected)"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    testimonial = get_object_or_404(

        Testimonial, 

        id=testimonial_id, 

        recipient=request.user.recipient_profile

    )

    

    # Only allow editing if pending or rejected

    if testimonial.approval_status == 'approved':

        messages.error(request, 'Cannot edit an approved testimonial.')

        return redirect('recipient_testimonials_list')

    

    if request.method == 'POST':

        form = TestimonialForm(request.POST, request.FILES, instance=testimonial)

        if form.is_valid():

            testimonial = form.save(commit=False)

            testimonial.approval_status = 'pending'  # Reset to pending after edit

            testimonial.save()

            messages.success(request, 'Testimonial updated and resubmitted for review.')

            return redirect('recipient_testimonials_list')

    else:

        form = TestimonialForm(instance=testimonial)

    

    context = {

        'form': form,

        'testimonial': testimonial,

        'is_edit': True,

    }

    return render(request, 'authentication/create_testimonial.html', context)



@login_required

def delete_testimonial(request, testimonial_id):

    """Delete a testimonial"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    testimonial = get_object_or_404(

        Testimonial, 

        id=testimonial_id, 

        recipient=request.user.recipient_profile

    )

    

    if request.method == 'POST':

        testimonial.delete()

        messages.success(request, 'Testimonial deleted successfully.')

        return redirect('recipient_testimonials_list')

    

    context = {'testimonial': testimonial}

    return render(request, 'authentication/delete_testimonial.html', context)



@login_required

def toggle_testimonial_display(request, testimonial_id):

    """Toggle whether a testimonial should be displayed publicly"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    testimonial = get_object_or_404(

        Testimonial, 

        id=testimonial_id, 

        recipient=request.user.recipient_profile

    )

    

    if testimonial.approval_status != 'approved':

        messages.error(request, 'Only approved testimonials can be toggled.')

        return redirect('recipient_testimonials_list')

    

    testimonial.display_on_public = not testimonial.display_on_public

    testimonial.save()

    

    status = "enabled" if testimonial.display_on_public else "disabled"

    messages.success(request, f'Public display {status} for this testimonial.')

    return redirect('recipient_testimonials_list')



# Admin testimonial management views

@login_required

def admin_testimonials_pending(request):

    """Admin view to see all pending testimonials"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    pending_testimonials = Testimonial.objects.filter(

        approval_status='pending'

    ).select_related('recipient__user').order_by('-created_at')

    

    context = {

        'title': 'Pending Testimonials',

        'pending_testimonials': pending_testimonials,

        'pending_count': pending_testimonials.count(),

    }

    return render(request, 'authentication/admin_testimonials_pending.html', context)



@login_required

def admin_approve_testimonial(request, testimonial_id):

    """Admin approves a testimonial"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    testimonial = get_object_or_404(Testimonial, id=testimonial_id)

    

    if request.method == 'POST':

        testimonial.approval_status = 'approved'

        testimonial.reviewed_by = request.user

        testimonial.reviewed_at = timezone.now()

        testimonial.set_default_display_period()  # Set 1-week display period

        testimonial.save()

        

        messages.success(request, f'Testimonial approved and will be displayed for 1 week.')

        return redirect('admin_testimonials_pending')

    

    context = {

        'title': 'Approve Testimonial',

        'testimonial': testimonial

    }

    return render(request, 'authentication/admin_approve_testimonial.html', context)



@login_required

def admin_reject_testimonial(request, testimonial_id):

    """Admin rejects a testimonial"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    testimonial = get_object_or_404(Testimonial, id=testimonial_id)

    

    if request.method == 'POST':

        rejection_reason = request.POST.get('rejection_reason', '')

        testimonial.approval_status = 'rejected'

        testimonial.reviewed_by = request.user

        testimonial.reviewed_at = timezone.now()

        testimonial.rejection_reason = rejection_reason

        testimonial.save()

        

        messages.success(request, 'Testimonial rejected.')

        return redirect('admin_testimonials_pending')

    

    context = {

        'title': 'Reject Testimonial',

        'testimonial': testimonial

    }

    return render(request, 'authentication/admin_reject_testimonial.html', context)



@login_required

def admin_all_testimonials(request):

    """Admin view to see all testimonials with filters"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    status_filter = request.GET.get('status', 'all')

    

    testimonials = Testimonial.objects.select_related('recipient__user', 'reviewed_by')

    

    if status_filter != 'all':

        testimonials = testimonials.filter(approval_status=status_filter)

    

    testimonials = testimonials.order_by('-created_at')

    

    context = {

        'title': 'All Testimonials',

        'testimonials': testimonials,

        'status_filter': status_filter,

        'total_count': testimonials.count(),

    }

    return render(request, 'authentication/admin_all_testimonials.html', context)



@login_required

def select_foodbank_for_donation(request):

    """Step 1: Select a food bank for general donation"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can make donations.')

        return redirect('dashboard')

    

    # Get all approved food banks with their public profile info

    foodbanks = FoodBankProfile.objects.select_related('user').filter(

        user__is_active=True,

        is_approved='approved'

    ).order_by('foodbank_name')

    

    context = {

        'foodbanks': foodbanks,

    }

    return render(request, 'donor/select_foodbank.html', context)



@login_required

def donate_to_foodbank_general(request, foodbank_id):

    """Step 2: Make donation to selected food bank"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can make donations.')

        return redirect('dashboard')

    

    try:

        foodbank = FoodBankProfile.objects.get(id=foodbank_id)

    except FoodBankProfile.DoesNotExist:

        messages.error(request, 'Food bank not found.')

        return redirect('donate')

    

    if request.method == 'POST':

        submission_token = (request.POST.get('submission_token') or '').strip()

        expected_token = request.session.get('donate_general_submission_token')

        if not expected_token or not submission_token or submission_token != expected_token:

            messages.error(request, 'Please refresh the page and try submitting again.')

            return redirect(request.path)

        

        post_data = request.POST.copy()
        custom_quantity_unit = (post_data.get('custom_quantity_unit') or '').strip()
        custom_subsidized_unit = (post_data.get('custom_subsidized_quantity_unit') or '').strip()
        if (post_data.get('quantity_unit') or '').strip().lower() == 'other' and custom_quantity_unit:
            post_data['quantity_unit'] = custom_quantity_unit
        if (post_data.get('subsidized_quantity_unit') or '').strip().lower() == 'other' and custom_subsidized_unit:
            post_data['subsidized_quantity_unit'] = custom_subsidized_unit

        form = DonationForm(post_data)

        # Remove foodbank from form validation since it's pre-selected

        if 'foodbank' in form.fields:

            del form.fields['foodbank']

        if form.is_valid():

            cache_key = f"donate_general_submit:{request.user.id}:{submission_token}"

            if not cache.add(cache_key, True, timeout=120):

                messages.info(request, 'This donation has already been submitted.')

                return redirect('donor_unspecified_donations_detail')



            try:

                donation = form.save(commit=False)

                donation.donor = request.user

                donation.foodbank = foodbank  # Set the selected foodbank

                

                # Handle 'other' and 'csr' type donations - enable discussion

                if donation.donation_type in ['other', 'csr']:

                    donation.requires_discussion = True

                    donation.discussion_status = 'pending'

                

                # Create notification for food bank

                Notification.objects.create(

                    user=donation.foodbank.user,

                    notification_type='donation_received',

                    message=f'New donation received from {request.user.email}: {donation.get_donation_display()}'

                )

                

                donation.save()



                request.session.pop('donate_general_submission_token', None)

                

                # Create UnspecifiedDonationManagement for general donations (not subsidized)

                # This enables the foodbank -> recipient approval workflow

                if donation.foodbank_request is None and donation.donation_mode != 'subsidized':

                    from .models import UnspecifiedDonationManagement

                    UnspecifiedDonationManagement.objects.create(donation=donation)

                

                # For 'other' and 'csr' type donations, notify foodbank about new discussion opportunity

                if donation.donation_type in ['other', 'csr']:

                    # Notify foodbank about the new donation requiring discussion

                    Notification.objects.create(

                        user=donation.foodbank.user,

                        notification_type='new_donor',

                        message=f'New {donation.get_donation_type_display()} donation requires discussion from {request.user.email}'

                    )

                # Send confirmation email in background so response returns quickly
                try:
                    send_donation_confirmation_email_async(donation)
                except Exception as e:
                    logger.exception("Failed to queue donation confirmation email")
                

                messages.success(request, f'Your donation was submitted successfully! ðŸŽ‰ Thank you for helping {donation.foodbank.foodbank_name}.')

                if donation.foodbank_request is None:

                    if donation.donation_mode == 'subsidized' or donation.donation_type == 'subsidized':

                        return redirect('donor_subsidized_donations_detail')

                    return redirect('donor_unspecified_donations_detail')

                return redirect('donor_donations_list')

            except ValueError as e:

                messages.error(request, str(e))

        else:

            messages.error(request, 'Please correct the errors below.')

    else:

        # Create form without foodbank field since it's pre-selected

        form = DonationForm()

        if 'foodbank' in form.fields:

            del form.fields['foodbank']



    submission_token = uuid.uuid4().hex

    request.session['donate_general_submission_token'] = submission_token

    delivery_methods = [
        (value, 'Delivery' if value == 'dropoff' else label)
        for value, label in Donation.DELIVERY_METHODS
    ]

    

    context = {

        'form': form,

        'foodbank': foodbank,

        'submission_token': submission_token,

        'donation_types': Donation.DONATION_TYPES,

        'quantity_units': _get_quantity_units(),

        'delivery_methods': delivery_methods,

        'STRIPE_PUBLISHABLE_KEY': settings.STRIPE_PUBLISHABLE_KEY,

    }

    return render(request, 'donor/donate_to_foodbank_enhanced.html', context)



@login_required

def donate_to_foodbank(request, request_id):

    """Handle donations to specific food bank requests"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can make donations.')

        return redirect('dashboard')



    try:

        foodbank_request = FoodBankRequest.objects.get(pk=request_id, status='active')

    except FoodBankRequest.DoesNotExist:

        messages.error(request, 'Request not found or no longer active.')

        return redirect('dashboard')



    if request.method == 'POST':
        post_data = request.POST.copy()

        # Guard against overlong request titles flowing into CharField(max_length=255)
        # via hidden/default form values in request-response flow.
        for field_name in ('item_name', 'subsidized_product_type'):
            field_value = (post_data.get(field_name) or '').strip()
            if field_value:
                post_data[field_name] = field_value[:255]

        response_type = (post_data.get('response_type') or '').strip().lower()
        title_fallback = (foodbank_request.title or 'Request Item').strip()[:255]

        if response_type == 'free' and not (post_data.get('item_name') or '').strip():
            post_data['item_name'] = title_fallback

        if response_type == 'subsidized' and not (post_data.get('subsidized_product_type') or '').strip():
            post_data['subsidized_product_type'] = title_fallback

        form = DonationForm(post_data, request_context=foodbank_request)

        if form.is_valid():

            donation = form.save(commit=False)

            donation.donor = request.user

            donation.foodbank = foodbank_request.foodbank

            donation.foodbank_request = foodbank_request

            donation.status = 'pending'  # Set initial status to pending

            # If donor contributes more than the request needs, keep the request portion
            # linked to the request and push the remainder into the same foodbank's stock
            # as a general donation (no specific request).
            remainder_donation = None
            try:
                qty_needed = getattr(foodbank_request, 'quantity_needed', None)
                qty_fulfilled = int(getattr(foodbank_request, 'quantity_fulfilled', 0) or 0)
                remaining_needed = None
                if qty_needed is not None:
                    remaining_needed = max(int(qty_needed) - qty_fulfilled, 0)

                if (
                    remaining_needed is not None
                    and donation.donation_type == 'item'
                    and getattr(donation, 'quantity', None)
                ):
                    donated_qty = int(donation.quantity)
                    if donated_qty > remaining_needed:
                        remainder_qty = donated_qty - remaining_needed
                        donation.quantity = remaining_needed if remaining_needed > 0 else donated_qty

                        if remaining_needed > 0 and remainder_qty > 0:
                            remainder_donation = Donation(
                                donor=donation.donor,
                                foodbank=donation.foodbank,
                                foodbank_request=None,
                                donation_type=donation.donation_type,
                                donation_mode=getattr(donation, 'donation_mode', None),
                                donation_category=getattr(donation, 'donation_category', None),
                                item_name=getattr(donation, 'item_name', None),
                                other_description=getattr(donation, 'other_description', None),
                                quantity=remainder_qty,
                                quantity_unit=getattr(donation, 'quantity_unit', None),
                                delivery_method=getattr(donation, 'delivery_method', None),
                                pickup_time=getattr(donation, 'pickup_time', None),
                                message=(getattr(donation, 'message', None) or '').strip() or None,
                                status=donation.status,
                            )

                        if remaining_needed == 0:
                            # Request already satisfied by accepted donations; treat entire donation as stock.
                            donation.foodbank_request = None
            except Exception:
                logger.exception("Failed to split excess donation into stock remainder")

            donation.save()

            if remainder_donation is not None and getattr(remainder_donation, 'quantity', 0):
                remainder_donation.save()

            

            # Update original request metadata so foodbank can review/accept

            original_req = foodbank_request.original_request

            if original_req:

                fields_to_update = []



                # Mark as awaiting foodbank action so Accept button shows up

                if getattr(original_req, 'status', None) not in ['awaiting_recipient', 'fulfilled']:

                    if original_req.status != 'donation_received':

                        original_req.status = 'donation_received'

                        fields_to_update.append('status')

                    if hasattr(original_req, 'awaiting_donors') and original_req.awaiting_donors:

                        original_req.awaiting_donors = False

                        fields_to_update.append('awaiting_donors')



                if fields_to_update:

                    original_req.save()



                # Don't append donor note to original_req

# Keep recipient notes in RequestManagement

# Donor note stays with the donation itself



            

            # Notifications (your existing logic)

            Notification.objects.create(

                user=request.user,

                notification_type='acknowledgement',

                message=f'Thank you for responding to the urgent request from {foodbank_request.foodbank.foodbank_name}.'

            )

            

            donation_display = donation.get_donation_display()

            Notification.objects.create(

                user=foodbank_request.foodbank.user,

                notification_type='donation_received',

                message=f'New donation pending review for request "{foodbank_request.title}": {donation_display} from {request.user.email}'

            )

            

            # Send email (your existing logic)

            try:
                send_donation_confirmation_email_async(donation)
            except Exception:
                logger.exception("Failed to queue donation confirmation email")
            

            messages.success(request, f'Donation submitted successfully! Your donation is pending review by {foodbank_request.foodbank.foodbank_name}. You will be notified once it is reviewed.')

            redirect_url = reverse('donor_donations_list')

            if foodbank_request.original_request is None:

                redirect_url = f"{redirect_url}?direct=1"

            return redirect(redirect_url)

        else:

            messages.error(request, 'Please correct the errors below.')

    else:

        form = DonationForm(request_context=foodbank_request)

    request_unit_value = (getattr(foodbank_request, 'quantity_unit', None) or '').strip()
    if request_unit_value == 'other' and getattr(foodbank_request, 'custom_unit', None):
        request_unit_value = (foodbank_request.custom_unit or '').strip()
    request_unit_label = request_unit_value or (_get_request_unit_label(foodbank_request) or '')

    request_card_description = (getattr(foodbank_request, 'description', '') or '').strip()
    linked_request = getattr(foodbank_request, 'linked_request_management', None) or getattr(foodbank_request, 'original_request', None)
    if linked_request and getattr(linked_request, 'description', None):
        request_card_description = (linked_request.description or '').strip()
    else:
        for marker in ('--- Recipient Note ---', '--- Acknowledgment Note', '--- Receipt Confirmed', '--- Donor Note'):
            if marker in request_card_description:
                request_card_description = request_card_description.split(marker)[0].strip()
    if not request_card_description:
        request_card_description = (foodbank_request.title or 'Request details').strip()

    request_title = (getattr(foodbank_request, 'title', '') or '').strip()
    normalized_title = ' '.join(request_title.split()).lower()
    normalized_description = ' '.join(request_card_description.split()).lower()
    show_request_card_description = bool(request_card_description and normalized_description != normalized_title)



    delivery_methods = [
        (value, 'Delivery' if value == 'dropoff' else label)
        for value, label in Donation.DELIVERY_METHODS
    ]

    context = {

        'form': form,

        'foodbank_request': foodbank_request,
        'request_card_description': request_card_description,
        'show_request_card_description': show_request_card_description,
        'request_unit_label': request_unit_label,
        'request_unit_value': request_unit_value,

        'donation_types': Donation.DONATION_TYPES,

        'quantity_units': _get_quantity_units(),

        'delivery_methods': delivery_methods,

    }

    return render(request, 'donor/donate.html', context)



@login_required

def add_quantity_unit(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    raw_label = (request.POST.get('label') or '').strip()
    if not raw_label:
        return JsonResponse({'success': False, 'message': 'Please enter a unit label.'}, status=400)

    if raw_label.lower() == 'other':
        return JsonResponse({'success': False, 'message': '"Other" is reserved. Please enter a specific unit.'}, status=400)

    existing = QuantityUnit.objects.filter(label__iexact=raw_label).first()
    if existing:
        return JsonResponse({
            'success': True,
            'created': False,
            'unit': {'code': existing.code, 'label': existing.label}
        })

    base_code = slugify(raw_label)
    if not base_code:
        return JsonResponse({'success': False, 'message': 'Invalid unit label.'}, status=400)

    code = base_code
    suffix = 1
    while QuantityUnit.objects.filter(code=code).exists():
        code = f"{base_code}-{suffix}"
        suffix += 1

    unit = QuantityUnit.objects.create(code=code, label=raw_label)
    return JsonResponse({
        'success': True,
        'created': True,
        'unit': {'code': unit.code, 'label': unit.label}
    })



@login_required

def donor_settings(request):

    if request.user.user_type != 'DONOR':

        return redirect('dashboard')

    

    try:

        donor_profile = request.user.donor_profile

    except:

        messages.error(request, 'Donor profile not found.')

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = DonorProfileForm(request.POST, instance=donor_profile, user=request.user)

        if form.is_valid():

            form.save()

            messages.success(request, 'Profile updated successfully!')

            return redirect('donor_settings')

    else:

        form = DonorProfileForm(instance=donor_profile, user=request.user)

    

    context = {

        'form': form,

        'donor_profile': donor_profile

    }

    return render(request, 'donor/donor_settings.html', context)



@login_required

def donation_history(request):

    donations = Donation.objects.filter(donor=request.user).order_by('-donated_at')

    return render(request, 'donor/history.html', {'donations': donations})



@login_required

def notifications_acknowledgements(request):

    notifications = Notification.objects.filter(user=request.user, notification_type='acknowledgement')

    return render(request, 'donor/acknowledgements.html', {'notifications': notifications})



@login_required

def notifications_requests(request):

    notes = Notification.objects.filter(user=request.user, notification_type='request').order_by('-created_at')

    notes.filter(is_read=False).update(is_read=True)

    return render(request, 'donor/requests.html', {'notifications': notes})



@login_required

def notifications_system_updates(request):

    notes = Notification.objects.filter(user=request.user, notification_type='system_update').order_by('-created_at')

    notes.filter(is_read=False).update(is_read=True)

    return render(request, 'donor/system_updates.html', {'notifications': notes})



@login_required

def notifications_foodbank(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')

    notifications.filter(is_read=False).update(is_read=True)

    return render(request, 'authentication/notifications_foodbank.html', {'notifications': notifications})



@login_required

def notifications_foodbank_system(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    notifications = Notification.objects.filter(user=request.user, notification_type__in=['system', 'system_maintenance']).order_by('-created_at')

    notifications.filter(is_read=False).update(is_read=True)

    return render(request, 'authentication/notifications_foodbank_system.html', {'notifications': notifications})



@login_required

def notifications_foodbank_requests(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    notifications = Notification.objects.filter(user=request.user, notification_type__in=['request', 'urgent_request', 'request_fulfilled', 'request_expiring']).order_by('-created_at')

    notifications.filter(is_read=False).update(is_read=True)

    return render(request, 'authentication/notifications_foodbank_requests.html', {'notifications': notifications})



@login_required

def mark_notification_read(request, pk):

    try:

        notification = Notification.objects.get(pk=pk, user=request.user)

        notification.is_read = True

        notification.save()

        messages.success(request, 'Notification marked as read.')

    except Notification.DoesNotExist:

        messages.error(request, 'Notification not found.')

    return redirect(request.META.get('HTTP_REFERER', 'notifications_foodbank'))



@login_required

def change_password(request):

    if request.user.user_type != 'DONOR':

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = DonorPasswordChangeForm(request.user, request.POST)

        if form.is_valid():

            form.save()

            # Update session to prevent logout

            update_session_auth_hash(request, form.user)

            messages.success(request, 'Password changed successfully!')

            return redirect('donor_settings')

    else:

        form = DonorPasswordChangeForm(request.user)

    

    context = {

        'form': form

    }

    return render(request, 'donor/change_password.html', context)



@login_required

def contact_support(request):

    return render(request, 'donor/contact_support.html')



@login_required

def privacy_settings(request):

    return render(request, 'donor/privacy_settings.html')



@login_required

def create_foodbank_request(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    foodbank_profile = request.user.foodbank_profile

    if request.method == 'POST':

        form = FoodBankRequestForm(request.POST)
        form.fields['delivery_method'].choices = [
            choice for choice in form.fields['delivery_method'].choices
            if choice[0] != 'both'
        ]

        if form.is_valid():

            req = form.save(commit=False)

            req.foodbank = foodbank_profile

            req.status = 'active'

            

            # Ensure deadline is set - if not provided, set default to 30 days from now

            if not req.deadline:

                req.deadline = timezone.now() + timedelta(days=30)

            

            # Validate deadline is in the future

            if req.deadline <= timezone.now():

                messages.error(request, 'Deadline must be in the future. Please select a future date and time.')

                return render(request, 'authentication/create_foodbank_request.html', {'form': form})

            

            # Handle custom unit logic

            if req.quantity_unit == 'other' and form.cleaned_data.get('custom_unit'):

                req.custom_unit = form.cleaned_data['custom_unit']

            else:

                req.custom_unit = None

            

            req.save()

            

            # Email notifications to donors temporarily disabled for performance

            # TODO: Implement async email notifications later

            

            messages.success(request, 'Request created successfully!')

            return redirect('foodbank_requests')

    else:

        form = FoodBankRequestForm()
        form.fields['delivery_method'].choices = [
            choice for choice in form.fields['delivery_method'].choices
            if choice[0] != 'both'
        ]

        # Set default deadline to 30 days from now for new requests

        default_deadline = timezone.now() + timedelta(days=30)

        form.fields['deadline'].initial = default_deadline.strftime('%Y-%m-%dT%H:%M')

    return render(request, 'authentication/create_foodbank_request.html', {'form': form})

@login_required

def edit_foodbank_request(request, pk):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    try:

        req = FoodBankRequest.objects.get(pk=pk, foodbank=request.user.foodbank_profile)

    except FoodBankRequest.DoesNotExist:

        messages.error(request, 'Request not found.')

        return redirect('dashboard')

    if request.method == 'POST':

        form = FoodBankRequestForm(request.POST, instance=req)

        if form.is_valid():

            updated_req = form.save(commit=False)

            

            # Ensure deadline is set - if not provided, set default to 30 days from now

            if not updated_req.deadline:

                updated_req.deadline = timezone.now() + timedelta(days=30)

            

            # Validate deadline is in the future

            if updated_req.deadline <= timezone.now():

                messages.error(request, 'Deadline must be in the future. Please select a future date and time.')

                return render(request, 'authentication/edit_foodbank_request.html', {'form': form, 'request_obj': req})

            

            # Handle custom unit logic

            if updated_req.quantity_unit == 'other' and form.cleaned_data.get('custom_unit'):

                updated_req.custom_unit = form.cleaned_data['custom_unit']

            else:

                updated_req.custom_unit = None

                

            updated_req.save()

            messages.success(request, 'Request updated successfully!')

            return redirect('dashboard')

    else:

        form = FoodBankRequestForm(instance=req)

    return render(request, 'authentication/edit_foodbank_request.html', {'form': form, 'request_obj': req})



@login_required

def delete_foodbank_request(request, pk):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    try:

        req = FoodBankRequest.objects.get(pk=pk, foodbank=request.user.foodbank_profile)

    except FoodBankRequest.DoesNotExist:

        messages.error(request, 'Request not found.')

        return redirect('dashboard')

    if request.method == 'POST':

        req.delete()

        messages.success(request, 'Request deleted successfully!')

        return redirect('dashboard')

    return render(request, 'authentication/delete_foodbank_request.html', {'request_obj': req})



@login_required

def fulfill_foodbank_request(request, pk):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    try:

        req = FoodBankRequest.objects.get(pk=pk, foodbank=request.user.foodbank_profile)

    except FoodBankRequest.DoesNotExist:

        messages.error(request, 'Request not found.')

        return redirect('dashboard')

    req.status = 'fulfilled'

    req.save()

    messages.success(request, 'Request marked as fulfilled!')

    return redirect('dashboard')



@login_required

def foodbank_donations(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    donations = Donation.objects.filter(foodbank=foodbank_profile).order_by('-donated_at')

    

    # Filter by donation type

    donation_type = request.GET.get('type')

    if donation_type:

        donations = donations.filter(donation_type=donation_type)

    

    # Filter by date range

    date_from = (request.GET.get('date_from') or '').strip()

    date_to = (request.GET.get('date_to') or '').strip()

    if date_from:

        donations = donations.filter(donated_at__gte=date_from)

    if date_to:

        donations = donations.filter(donated_at__lte=date_to)

    

    # Pagination

    paginator = Paginator(donations, 10)  # 10 items per page

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)



    # Attach latest recipient notes for displayed donations

    donation_list = list(page_obj.object_list)

    donation_ids = [donation.id for donation in donation_list]

    latest_notes = {}



    if donation_ids:

        responses = DonationResponse.objects.filter(

            donation_id__in=donation_ids

        ).exclude(

            notes__isnull=True

        ).exclude(

            notes__exact=''

        ).order_by('-responded_at')



        for response in responses:

            if response.donation_id not in latest_notes:

                latest_notes[response.donation_id] = response.notes



    for donation in donation_list:

        donation.latest_recipient_note = latest_notes.get(donation.id)

    

    context = {

        'donations': page_obj,

        'page_obj': page_obj,

        'total_donations': paginator.count,

        'total_money': Donation.objects.filter(foodbank=foodbank_profile, donation_type='money').aggregate(total=Sum('amount'))['total'] or 0,

        'total_items': Donation.objects.filter(foodbank=foodbank_profile, donation_type='item').count(),

        'total_subsidized': Donation.objects.filter(foodbank=foodbank_profile, donation_type='subsidized').aggregate(total=Sum('subsidized_price'))['total'] or 0,

    }

    

    return render(request, 'authentication/foodbank_donations.html', context)



@login_required

def foodbank_requests(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    requests = FoodBankRequest.objects.filter(

        foodbank=foodbank_profile,

        original_request__isnull=True

    ).select_related(

        'linked_request_management',

        'foodbank'

    ).prefetch_related(

        'donations__donor__donor_profile',

        'donations__allocations'

    ).order_by('-created_at')

    

    # Filter by date range

    date_from = (request.GET.get('date_from') or '').strip()

    date_to = (request.GET.get('date_to') or '').strip()

    if date_from:

        requests = requests.filter(created_at__date__gte=date_from)

    if date_to:

        requests = requests.filter(created_at__date__lte=date_to)



    category_filter = (request.GET.get('category') or '').strip().lower().replace('-', '_')
    type_filter = (request.GET.get('type') or '').strip().lower().replace('-', '_')

    if not type_filter and category_filter in ('food', 'non_food'):
        type_filter = category_filter
        category_filter = ''

    if type_filter in ('food', 'non_food'):
        requests = requests.filter(donation_type=type_filter)

    category_type_map = {
        'free_goods': 'item',
        'item': 'item',
        'monetary': 'money',
        'money': 'money',
        'subsidized': 'subsidized',
        'csr': 'csr',
        'other': 'other',
    }
    selected_category_type = category_type_map.get(category_filter)
    if selected_category_type:
        requests = requests.filter(donations__donation_type=selected_category_type).distinct()



    donation_status_filter = request.GET.get('donation_status', '').strip()

    if donation_status_filter == 'pending':
        requests = requests.filter(Q(donations__status='pending') | Q(donations__isnull=True)).distinct()

    elif donation_status_filter == 'declined':
        declined_request_ids = []
        for req in requests.prefetch_related('donations'):
            donations = list(req.donations.all())
            if donations and all(d.status == 'declined' for d in donations):
                declined_request_ids.append(req.id)
        requests = requests.filter(id__in=declined_request_ids)

    elif donation_status_filter in ('partially_fulfilled', 'fulfilled'):

        requests = requests.filter(donations__status='accepted').distinct()

        requests_list = list(requests.prefetch_related('donations'))

        target_label = 'Partially Fulfilled' if donation_status_filter == 'partially_fulfilled' else 'Fulfilled'

        matching_ids = [r.id for r in requests_list if r.get_foodbank_requests_status_label() == target_label]

        requests = requests.filter(id__in=matching_ids)



    quantity_filter = (request.GET.get('quantity') or '').strip().lower()
    if quantity_filter == 'small':
        requests = requests.filter(quantity_needed__lte=50)
    elif quantity_filter == 'medium':
        requests = requests.filter(quantity_needed__gt=50, quantity_needed__lte=200)
    elif quantity_filter == 'large':
        requests = requests.filter(quantity_needed__gt=200)

    amount_filter = (request.GET.get('amount') or '').strip().lower()
    if amount_filter == 'small':
        requests = requests.filter(
            Q(donations__amount__lte=5000) | Q(donations__subsidized_price__lte=5000)
        ).distinct()
    elif amount_filter == 'medium':
        requests = requests.filter(
            Q(donations__amount__gt=5000, donations__amount__lte=20000) |
            Q(donations__subsidized_price__gt=5000, donations__subsidized_price__lte=20000)
        ).distinct()
    elif amount_filter == 'large':
        requests = requests.filter(
            Q(donations__amount__gt=20000) | Q(donations__subsidized_price__gt=20000)
        ).distinct()

    delivery_filter = (request.GET.get('delivery') or '').strip().lower()
    if delivery_filter in ('all', 'both'):
        delivery_filter = ''
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'

    if delivery_filter in ('pickup', 'delivery'):
        if delivery_filter == 'delivery':
            requests = requests.filter(
                Q(delivery_method='delivery') |
                Q(donations__delivery_method__in=['delivery', 'dropoff']) |
                Q(linked_request_management__delivery_method='delivery')
            ).distinct()
        else:
            requests = requests.filter(
                Q(delivery_method='pickup') |
                Q(donations__delivery_method='pickup') |
                Q(linked_request_management__delivery_method='pickup')
            ).distinct()



    # Search: title, description, location, notes, donor

    search_query = (request.GET.get('search') or '').strip()

    if search_query:

        requests = requests.filter(

            Q(title__icontains=search_query) |

            Q(description__icontains=search_query) |

            Q(linked_request_management__additional_notes__icontains=search_query) |

            Q(linked_request_management__location__icontains=search_query) |

            Q(donations__donor__donor_profile__full_name__icontains=search_query) |

            Q(donations__donor__email__icontains=search_query)

        ).distinct()



    # Sort

    sort_filter = request.GET.get('sort', 'newest')

    if sort_filter == 'oldest':

        requests = requests.order_by('created_at')

    else:

        requests = requests.order_by('-created_at')



    # Pagination - 10 items per page

    paginator = Paginator(requests, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)



    delivery_label_map = {
        'pickup': 'Pickup',
        'dropoff': 'Delivery',
        'delivery': 'Delivery',
    }

    for req in page_obj:

        req.status_label = req.get_foodbank_requests_status_label()

        req.quantity_display = req.get_requested_quantity_display()

        req.progress_percent = req.get_fulfillment_percentage()

        req.total_received = req.get_total_donations_received()

        req.stock_is_monetary = False

        if not (req.total_received and getattr(req, 'quantity_needed', None)):

            money_donations = req.donations.filter(donation_type='money', status='accepted')

            if money_donations.exists():

                from decimal import Decimal

                total_money = sum(Decimal(str(d.amount or 0)) for d in money_donations)

                if total_money > 0:

                    req.total_received = float(total_money)

                    req.stock_is_monetary = True

        used_from_allocations = 0
        used_pieces = 0
        for donation in req.donations.all():
            for allocation in donation.allocations.filter(declined_by_recipient=False):
                used_from_allocations += (allocation.quantity or allocation.amount or 0)
                used_pieces += (allocation.quantity or 0)

        quantity_needed = getattr(req, 'quantity_needed', None)
        if quantity_needed is not None:
            # Show stock in pieces; cap by donated total so used/remaining match received stock
            stock_capacity = quantity_needed
            if req.total_received:
                stock_capacity = min(quantity_needed, req.total_received)
            req.stock_used = min(used_pieces, stock_capacity)
            req.stock_required = stock_capacity
            req.stock_remaining = max(0, stock_capacity - req.stock_used)
            req.stock_is_monetary = False
            req.stock_capacity = stock_capacity
        else:
            stock_capacity = req.total_received or 0
            req.stock_capacity = stock_capacity
            req.stock_required = stock_capacity
            req.stock_used = min(used_from_allocations, stock_capacity) if stock_capacity else 0
            req.stock_remaining = max(0, stock_capacity - req.stock_used) if stock_capacity else 0

        linked = getattr(req, 'linked_request_management', None)

        donor_delivery_methods = []

        type_labels = []

        for donation in req.donations.all():

            method = getattr(donation, 'delivery_method', None)

            if method and method not in donor_delivery_methods:

                donor_delivery_methods.append(method)



            if donation.donation_type == 'item':

                type_label = 'Free Goods'

            elif donation.donation_type == 'money' and getattr(donation.foodbank_request, 'original_request', None):

                type_label = donation.foodbank_request.original_request.get_request_type_display()

            else:
                if donation.donation_type == 'money':
                    type_label = 'Monetary'
                else:
                    type_label = donation.get_donation_type_display()



            if type_label and type_label not in type_labels:

                type_labels.append(type_label)



        if donor_delivery_methods:

            req.delivery_display = ", ".join(

                delivery_label_map.get(method, method.replace('_', ' ').title())

                for method in donor_delivery_methods

            )

        else:

            linked_delivery = linked.get_delivery_method_display() if linked and getattr(linked, 'delivery_method', None) else None
            request_delivery = req.get_delivery_method_display() if getattr(req, 'delivery_method', None) else None
            req.delivery_display = linked_delivery or request_delivery or '-'



        req.category_display = ", ".join(type_labels) if type_labels else req.get_donation_type_display()

        req.location_display = (

            getattr(linked, 'location', None)

            or getattr(req.foodbank, 'address', None)

            or '-'

        )

        req.notes_display = getattr(linked, 'additional_notes', None) or req.description



    # Get counts from original queryset before pagination

    total_requests = requests.count()

    active_requests = requests.filter(status='active').count()

    fulfilled_requests = requests.filter(status='fulfilled').count()

    urgent_requests = requests.filter(priority='urgent').count()

    

    # Get total pending donations count for all requests

    total_pending_donations = Donation.objects.filter(

        foodbank=foodbank_profile,

        foodbank_request__isnull=False,

        foodbank_request__original_request__isnull=True,

        status='pending'

    ).count()

    

    context = {

        'requests': page_obj,

        'page_obj': page_obj,

        'total_requests': total_requests,

        'active_requests': active_requests,

        'fulfilled_requests': fulfilled_requests,

        'urgent_requests': urgent_requests,

        'total_pending_donations': total_pending_donations,

        'type_choices': FoodBankRequest.DONATION_TYPE_CHOICES,
        'type_filter': type_filter,
        'category_choices': [
            ('free_goods', 'Free Goods'),
            ('monetary', 'Monetary'),
            ('subsidized', 'Subsidized'),
        ],
        'category_filter': category_filter,

        'donation_status_choices': [
            ('pending', 'Pending'),
            ('partially_fulfilled', 'Partially Fulfilled'),
            ('fulfilled', 'Fulfilled'),
            ('declined', 'Declined'),
        ],

        'donation_status_filter': donation_status_filter,

        'delivery_choices': [
            ('pickup', 'Pickup'),
            ('delivery', 'Delivery'),
        ],
        'delivery_filter': delivery_filter,
        'quantity_choices': [
            ('small', 'Small (1-50)'),
            ('medium', 'Medium (51-200)'),
            ('large', 'Large (201+)'),
        ],
        'quantity_filter': quantity_filter,
        'amount_choices': [
            ('small', 'Small (<= KES 5,000)'),
            ('medium', 'Medium (KES 5,001 - 20,000)'),
            ('large', 'Large (> KES 20,000)'),
        ],
        'amount_filter': amount_filter,

        'sort_filter': sort_filter,

        'search_query': search_query,

    }

    return render(request, 'authentication/foodbank_requests.html', context)


def _foodbank_requests_list_queryset(request, foodbank_profile):
    """Same filtered FoodBankRequest queryset as foodbank_requests view (no pagination)."""
    requests = FoodBankRequest.objects.filter(
        foodbank=foodbank_profile,
        original_request__isnull=True
    ).select_related('linked_request_management', 'foodbank').prefetch_related(
        'donations__donor__donor_profile', 'donations__allocations'
    ).order_by('-created_at')

    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()
    if date_from:
        requests = requests.filter(created_at__date__gte=date_from)
    if date_to:
        requests = requests.filter(created_at__date__lte=date_to)

    category_filter = (request.GET.get('category') or '').strip().lower().replace('-', '_')
    type_filter = (request.GET.get('type') or '').strip().lower().replace('-', '_')
    if not type_filter and category_filter in ('food', 'non_food'):
        type_filter = category_filter
        category_filter = ''
    if type_filter in ('food', 'non_food'):
        requests = requests.filter(donation_type=type_filter)

    category_type_map = {
        'free_goods': 'item',
        'item': 'item',
        'monetary': 'money',
        'money': 'money',
        'subsidized': 'subsidized',
        'csr': 'csr',
        'other': 'other',
    }
    selected_category_type = category_type_map.get(category_filter)
    if selected_category_type:
        requests = requests.filter(donations__donation_type=selected_category_type).distinct()

    donation_status_filter = request.GET.get('donation_status', '').strip()
    if donation_status_filter == 'pending':
        requests = requests.filter(donations__isnull=True)
    elif donation_status_filter == 'declined':
        requests = requests.filter(donations__status='declined').exclude(
            donations__status='accepted'
        ).exclude(donations__status='pending').distinct()
    elif donation_status_filter in ('partially_fulfilled', 'fulfilled'):
        requests = requests.filter(donations__status='accepted').distinct()
        requests_list = list(requests.prefetch_related('donations'))
        target_label = 'Partially Fulfilled' if donation_status_filter == 'partially_fulfilled' else 'Fulfilled'
        matching_ids = [r.id for r in requests_list if r.get_foodbank_requests_status_label() == target_label]
        requests = requests.filter(id__in=matching_ids)

    quantity_filter = (request.GET.get('quantity') or '').strip().lower()
    if quantity_filter == 'small':
        requests = requests.filter(quantity_needed__lte=50)
    elif quantity_filter == 'medium':
        requests = requests.filter(quantity_needed__gt=50, quantity_needed__lte=200)
    elif quantity_filter == 'large':
        requests = requests.filter(quantity_needed__gt=200)

    amount_filter = (request.GET.get('amount') or '').strip().lower()
    if amount_filter == 'small':
        requests = requests.filter(
            Q(donations__amount__lte=5000) | Q(donations__subsidized_price__lte=5000)
        ).distinct()
    elif amount_filter == 'medium':
        requests = requests.filter(
            Q(donations__amount__gt=5000, donations__amount__lte=20000) |
            Q(donations__subsidized_price__gt=5000, donations__subsidized_price__lte=20000)
        ).distinct()
    elif amount_filter == 'large':
        requests = requests.filter(
            Q(donations__amount__gt=20000) | Q(donations__subsidized_price__gt=20000)
        ).distinct()

    delivery_filter = (request.GET.get('delivery') or '').strip().lower()
    if delivery_filter in ('all', 'both'):
        delivery_filter = ''
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    if delivery_filter in ('pickup', 'delivery'):
        if delivery_filter == 'delivery':
            requests = requests.filter(
                Q(delivery_method='delivery') |
                Q(donations__delivery_method__in=['delivery', 'dropoff']) |
                Q(linked_request_management__delivery_method='delivery')
            ).distinct()
        else:
            requests = requests.filter(
                Q(delivery_method='pickup') |
                Q(donations__delivery_method='pickup') |
                Q(linked_request_management__delivery_method='pickup')
            ).distinct()

    search_query = (request.GET.get('search') or '').strip()
    if search_query:
        requests = requests.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(linked_request_management__additional_notes__icontains=search_query) |
            Q(linked_request_management__location__icontains=search_query) |
            Q(donations__donor__donor_profile__full_name__icontains=search_query) |
            Q(donations__donor__email__icontains=search_query)
        ).distinct()

    sort_filter = request.GET.get('sort', 'newest')
    if sort_filter == 'oldest':
        requests = requests.order_by('created_at')
    else:
        requests = requests.order_by('-created_at')
    return requests


def _foodbank_request_row_data(req, delivery_label_map, idx):
    """Build one row dict for export: date, category_display, type, description, qty_amount, donor, delivery, status, notes, created."""
    status_label = req.get_foodbank_requests_status_label()
    linked = getattr(req, 'linked_request_management', None)
    donor_delivery_methods = []
    type_labels = []
    for donation in req.donations.all():
        method = getattr(donation, 'delivery_method', None)
        if method and method not in donor_delivery_methods:
            donor_delivery_methods.append(method)
        if donation.donation_type == 'item':
            type_label = 'Free Goods'
        elif donation.donation_type == 'money' and getattr(getattr(donation, 'foodbank_request', None), 'original_request', None):
            type_label = donation.foodbank_request.original_request.get_request_type_display()
        else:
            if donation.donation_type == 'money':
                type_label = 'Monetary'
            else:
                type_label = donation.get_donation_type_display()
        if type_label and type_label not in type_labels:
            type_labels.append(type_label)

    if donor_delivery_methods:
        delivery_display = ", ".join(
            delivery_label_map.get(m, m.replace('_', ' ').title()) for m in donor_delivery_methods
        )
    else:
        linked_delivery = linked.get_delivery_method_display() if linked and getattr(linked, 'delivery_method', None) else None
        request_delivery = req.get_delivery_method_display() if getattr(req, 'delivery_method', None) else None
        delivery_display = linked_delivery or request_delivery or '-'
    category_display = ", ".join(type_labels) if type_labels else req.get_donation_type_display()

    requested_unit_label = getattr(req, 'custom_unit', None) or (getattr(req, 'get_quantity_unit_display', lambda: None)() or getattr(req, 'quantity_unit', None) or '')

    if status_label == 'Donation Made':
        lines = []
        for donation in req.donations.all():
            if donation.status == 'pending':
                if donation.donation_type == 'item':
                    lines.append(f"{donation.quantity or 0} (of {req.quantity_needed or 0} {requested_unit_label} requested)")
                elif donation.donation_type == 'money':
                    lines.append(f"{donation.amount:.0f} (for {req.quantity_needed or 0} {requested_unit_label} requested)")
                elif donation.donation_type == 'subsidized':
                    lines.append(f"{donation.subsidized_price:.0f} for {donation.subsidized_quantity or donation.quantity or 0} {donation.subsidized_quantity_unit or donation.quantity_unit or requested_unit_label} (of {req.quantity_needed or 0} {requested_unit_label} requested)")
                else:
                    lines.append(f"{donation.quantity or donation.amount or 0} (of {req.quantity_needed or 0} {requested_unit_label} requested)")
        qty_amount = "\n".join(lines) if lines else "-"
    elif status_label == 'Declined':
        lines = []
        for donation in req.donations.all():
            if donation.status == 'declined':
                if donation.donation_type == 'item':
                    lines.append(f"{donation.quantity or 0} (of {req.quantity_needed or 0} {requested_unit_label} requested)")
                elif donation.donation_type == 'money':
                    lines.append(f"{donation.amount:.0f} (for {req.quantity_needed or 0} {requested_unit_label} requested)")
                elif donation.donation_type == 'subsidized':
                    lines.append(f"{donation.subsidized_price:.0f} for {donation.subsidized_quantity or donation.quantity or 0} {donation.subsidized_quantity_unit or donation.quantity_unit or requested_unit_label} (of {req.quantity_needed or 0} {requested_unit_label} requested)")
                else:
                    lines.append(f"{donation.quantity or donation.amount or 0} (of {req.quantity_needed or 0} {requested_unit_label} requested)")
        qty_amount = "\n".join(lines) if lines else "-"
    else:
        qty_amount = req.get_fulfillment_qty_amount_display() if req.get_fulfillment_qty_amount_display() != '-' else (
            f"{req.quantity_needed} {requested_unit_label}" if req.quantity_needed else "-"
        )

    donor_names = []
    notes_entries = []
    donor_note_entries = []
    acceptance_note_entries = []
    from decimal import Decimal
    amount_total = Decimal('0')
    for donation in req.donations.all():
        d = donation.donor
        donor_name = (
            getattr(getattr(d, 'donor_profile', None), 'full_name', None)
            or getattr(d, 'email', '')
            or 'Not donated'
        )
        if donor_name and donor_name not in donor_names:
            donor_names.append(donor_name)

        donor_note = (
            getattr(donation, 'message', None)
            or getattr(donation, 'other_description', None)
            or getattr(donation, 'csr_description', None)
            or ''
        ).strip()
        decision_note = (getattr(donation, 'decline_message', None) or '').strip()

        note_parts = []
        if donor_note:
            donor_note_clean = re.sub(r'^\s*donor\s*note\s*:\s*', '', donor_note, flags=re.IGNORECASE).strip()
            note_parts.append(f"Donor Note: {donor_note_clean}")
            if donor_note_clean and donor_note_clean not in donor_note_entries:
                donor_note_entries.append(donor_note_clean)
        if decision_note:
            decision_note_clean = re.sub(
                r'^\s*(acceptance|rejection|decision)\s*note\s*:\s*',
                '',
                decision_note,
                flags=re.IGNORECASE
            ).strip()
            if donation.status == 'accepted':
                decision_label = 'Acceptance Note'
                if decision_note_clean and decision_note_clean not in acceptance_note_entries:
                    acceptance_note_entries.append(decision_note_clean)
            elif donation.status == 'declined':
                decision_label = 'Rejection Note'
            else:
                decision_label = 'Decision Note'
            note_parts.append(f"{decision_label}: {decision_note_clean}")
        if note_parts:
            notes_entries.append(" | ".join(note_parts))

        if donation.donation_type == 'money' and getattr(donation, 'amount', None) is not None:
            amount_total += Decimal(str(donation.amount))
        elif donation.donation_type == 'subsidized' and getattr(donation, 'subsidized_price', None) is not None:
            amount_total += Decimal(str(donation.subsidized_price))

    donor_str = ", ".join(donor_names) if donor_names else "Not donated"
    notes_str = "\n".join(notes_entries) if notes_entries else "No notes"
    donor_note_str = "\n".join(donor_note_entries) if donor_note_entries else "-"
    acceptance_note_str = "\n".join(acceptance_note_entries) if acceptance_note_entries else "-"
    amount_kes = f"{amount_total:,.2f}" if amount_total > 0 else "-"

    total_received = req.get_total_donations_received() if getattr(req, 'get_total_donations_received', None) else 0
    stock_is_monetary = False
    if not (total_received and getattr(req, 'quantity_needed', None)):
        money_donations = req.donations.filter(donation_type='money', status='accepted')
        if money_donations.exists():
            from decimal import Decimal
            total_money = sum(Decimal(str(d.amount or 0)) for d in money_donations)
            if total_money > 0:
                total_received = float(total_money)
                stock_is_monetary = True

    used_from_allocations = 0
    used_pieces = 0
    for donation in req.donations.all():
        for allocation in (donation.allocations.all() if hasattr(donation.allocations, 'all') else []):
            if getattr(allocation, 'declined_by_recipient', None):
                continue
            used_from_allocations += (getattr(allocation, 'quantity', None) or getattr(allocation, 'amount', 0) or 0)
            used_pieces += (getattr(allocation, 'quantity', None) or 0)

    quantity_needed = getattr(req, 'quantity_needed', None)
    if quantity_needed is not None:
        stock_capacity = quantity_needed
        if total_received:
            stock_capacity = min(quantity_needed, total_received)
        stock_used = min(used_pieces, stock_capacity)
        stock_required = stock_capacity
        stock_remaining = max(0, stock_capacity - stock_used)
        stock_is_monetary = False
    else:
        stock_capacity = total_received or 0
        stock_required = stock_capacity
        stock_used = min(used_from_allocations, stock_capacity) if stock_capacity else 0
        stock_remaining = max(0, stock_capacity - stock_used) if stock_capacity else 0

    stock_unit_label = (requested_unit_label or '').strip() or 'units'

    if not ((total_received and quantity_needed) or (total_received and stock_is_monetary)):
        stock_display = "-"
    elif stock_required and stock_used >= stock_required:
        stock_display = "Out of Stock"
    elif stock_used > 0:
        prefix = "KES " if stock_is_monetary else ""
        if stock_is_monetary:
            stock_display = f"Used Stock: {prefix}{stock_used:.0f} / {prefix}{stock_required:.0f}, Remaining: {prefix}{stock_remaining:.0f}"
        else:
            stock_display = (
                f"Used Stock: {stock_used:.0f} {stock_unit_label} / {stock_required:.0f} {stock_unit_label}, "
                f"Remaining: {stock_remaining:.0f} {stock_unit_label}"
            )
    else:
        stock_display = "In Stock"

    quantity_detail = (qty_amount or '-').strip() or '-'
    if stock_display and stock_display != '-':
        stock_lines = stock_display
        if stock_display.startswith("Used Stock:") and ", Remaining:" in stock_display:
            used_part, remaining_part = stock_display.split(", Remaining:", 1)
            stock_lines = f"{used_part}\nRemaining Stock:{remaining_part}"
        quantity_detail = f"{quantity_detail}\n\n{stock_lines}" if quantity_detail else stock_lines

    date_str = req.deadline.strftime('%b %d, %Y') if req.deadline else '-'
    created_str = req.created_at.strftime('%b %d, %Y') if req.created_at else '-'
    desc_full = f"{req.title}\n{req.description}" if req.title else (req.description or '-')

    try:
        total_received_num = int(total_received) if total_received is not None else None
    except (TypeError, ValueError):
        total_received_num = float(total_received) if total_received is not None else None

    stock_in_value = total_received_num
    if quantity_needed is not None:
        accepted_qty = 0
        for donation in req.donations.all():
            if getattr(donation, 'status', None) != 'accepted':
                continue
            if donation.donation_type == 'item':
                accepted_qty += int(getattr(donation, 'quantity', 0) or 0)
            elif donation.donation_type == 'subsidized':
                accepted_qty += int((getattr(donation, 'subsidized_quantity', None) or getattr(donation, 'quantity', 0) or 0))
            elif donation.donation_type == 'money':
                # If money donations carry a quantity target, count it as fulfilled quantity.
                money_qty = getattr(donation, 'quantity', 0) or 0
                if not money_qty:
                    money_qty = _get_request_quantity_value(getattr(donation, 'foodbank_request', None)) or 0
                if not money_qty:
                    try:
                        money_qty = sum(
                            int(getattr(alloc, 'quantity', 0) or 0)
                            for alloc in donation.allocations.filter(declined_by_recipient=False)
                        )
                    except Exception:
                        money_qty = 0
                if not money_qty:
                    money_qty = quantity_needed or 0
                accepted_qty += int(money_qty or 0)
        if accepted_qty and quantity_needed:
            accepted_qty = min(accepted_qty, int(quantity_needed))
        stock_in_value = accepted_qty
    try:
        stock_used_num = int(stock_used) if stock_used is not None else None
    except (TypeError, ValueError):
        stock_used_num = float(stock_used) if stock_used is not None else None
    try:
        stock_remaining_num = int(stock_remaining) if stock_remaining is not None else None
    except (TypeError, ValueError):
        stock_remaining_num = float(stock_remaining) if stock_remaining is not None else None

    return {
        'idx': idx,
        'date': date_str,
        'category': category_display,
        'type': req.get_donation_type_display(),
        'description': desc_full,
        'qty_amount': qty_amount,
        'quantity': quantity_needed if quantity_needed is not None else '',
        'unit': requested_unit_label or '',
        'requested': quantity_needed if quantity_needed is not None else '',
        'fulfilled': stock_in_value if stock_in_value is not None else '',
        'used_stock': stock_used_num if stock_used_num is not None else '',
        'remaining_stock': stock_remaining_num if stock_remaining_num is not None else '',
        'donor': donor_str,
        'delivery': delivery_display,
        'status': status_label,
        'stock': stock_display,
        'quantity_detail': quantity_detail,
        'amount_kes': amount_kes,
        'notes': notes_str,
        'donor_note': donor_note_str,
        'acceptance_note': acceptance_note_str,
        'created': created_str,
    }


def _wrap_export_text(value, width=30, max_lines=3):
    import textwrap
    text = (value or '').strip()
    if not text:
        return '-'
    lines = textwrap.wrap(text, width=width)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        if lines[-1]:
            lines[-1] = (lines[-1][:-3] + '...') if len(lines[-1]) > 3 else (lines[-1] + '...')
    return "\n".join(lines)


@login_required
def export_foodbank_requests_list(request, format):
    """Export All Requests (FoodBankRequest list) as PDF, CSV, or Excel - matches foodbank_requests table."""
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    foodbank_profile = get_object_or_404(FoodBankProfile, user=request.user)
    requests_qs = _foodbank_requests_list_queryset(request, foodbank_profile)
    requests_list = list(requests_qs)
    delivery_label_map = {
        'pickup': 'Pickup',
        'dropoff': 'Delivery',
        'delivery': 'Delivery',
    }
    rows = [_foodbank_request_row_data(req, delivery_label_map, i) for i, req in enumerate(requests_list, 1)]

    fmt = (format or '').strip().lower()
    if fmt == 'pdf':
        return _export_foodbank_requests_list_pdf(request, rows, foodbank_profile)
    if fmt == 'csv':
        return _export_foodbank_requests_list_csv(rows, foodbank_profile)
    if fmt == 'excel':
        return _export_foodbank_requests_list_excel(rows, foodbank_profile)
    messages.error(request, 'Invalid format.')
    return redirect('foodbank_requests')


def _export_foodbank_requests_list_pdf(request, rows, foodbank_profile):
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.platypus import Table, Paragraph, TableStyle
    from .report_utils import (
        get_report_styles, build_report_header, get_branded_table_style,
        build_report_summary, build_pdf_document, collect_active_filters,
    )
    styles = get_report_styles()
    report_pagesize = landscape(A3)
    wrap = styles['wrap']
    wrap_center = styles.get('wrap_center', wrap)
    elements = []
    name = foodbank_profile.foodbank_name or foodbank_profile.user.email
    active_filters = collect_active_filters(request, [
        ('type', 'Type'),
        ('category', 'Category'),
        ('donation_status', 'Donation Status'),
        ('delivery', 'Delivery'),
        ('quantity', 'Quantity'),
        ('amount', 'Amount'),
        ('date_from', 'From Date'),
        ('date_to', 'To Date'),
        ('sort', 'Sort'),
        ('search', 'Search'),
    ])
    build_report_header(elements, "Direct Donations Report", name, len(rows), active_filters, styles)

    if not rows:
        elements.append(Paragraph("No requests found matching the current filters.", styles['normal']))
    else:
        data = [[
            'S/No', 'Date', 'Category', 'Type', 'Description', 'Qty/Amount',
            'Donor', 'Delivery', 'Status', 'Stock', 'Notes', 'Created'
        ]]
        for r in rows:
            notes_pdf_text = html.escape((r.get('notes') or '-')[:240]).replace('\n', '<br/>')
            for label in ('Donor Note:', 'Acceptance Note:', 'Rejection Note:', 'Decision Note:'):
                notes_pdf_text = notes_pdf_text.replace(label, f"<b>{label}</b>")
            data.append([
                str(r['idx']),
                Paragraph(r['date'], wrap),
                Paragraph(r['category'], wrap),
                Paragraph(r['type'], wrap),
                Paragraph((r['description'] or '-')[:150].replace('\n', '<br/>'), wrap),
                Paragraph((r['qty_amount'] or '-').replace('\n', '<br/>'), wrap),
                Paragraph((r['donor'] or '-')[:80], wrap),
                Paragraph(r['delivery'], wrap),
                Paragraph(r['status'], wrap),
                Paragraph((r.get('stock') or '-').replace('\n', '<br/>'), wrap),
                Paragraph(notes_pdf_text, wrap),
                Paragraph(r['created'], wrap),
            ])
        col_widths = [0.35*inch, 0.65*inch, 0.75*inch, 0.6*inch, 1.45*inch, 1.05*inch, 1.0*inch, 0.75*inch, 0.9*inch, 0.9*inch, 1.45*inch, 0.75*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(get_branded_table_style(len(data)))
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ]))
        elements.append(table)
        build_report_summary(elements, [("Total Requests", len(rows))], styles)
    safe_name = (foodbank_profile.foodbank_name or "foodbank").replace(" ", "_")[:30]
    return build_pdf_document(elements, "all_requests", safe_name, pagesize=report_pagesize)


def _export_foodbank_requests_list_csv(rows, foodbank_profile):
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{foodbank_profile.foodbank_name.replace(" ", "_")}_all_requests_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow([f'Direct Donations Report - {foodbank_profile.foodbank_name}'])
    writer.writerow([f'Generated: {timezone.localtime().strftime("%Y-%m-%d %H:%M")}'])
    writer.writerow([])
    writer.writerow([
        'S/No', 'Date', 'Category', 'Type', 'Description', 'Quantity Requested', 'Stock In',
        'Used Stock', 'Remaining Stock', 'Unit', 'Quantity Detail', 'Amount (KES)', 'Donor', 'Delivery', 'Status', 'Donor Note', 'Acceptance Note', 'Created'
    ])
    for r in rows:
        writer.writerow([
            r['idx'], r['date'], r['category'], r['type'], (r['description'] or '').replace('\n', ' '),
            r.get('requested', ''), r.get('fulfilled', ''),
            r.get('used_stock', ''), r.get('remaining_stock', ''),
            r.get('unit', ''),
            (r.get('quantity_detail') or '-'),
            (r.get('amount_kes') or '-'),
            r['donor'], r['delivery'], r['status'],
            (r.get('donor_note') or '').replace('\n', ' '),
            (r.get('acceptance_note') or '').replace('\n', ' '),
            r['created'],
        ])
    return response


def _export_foodbank_requests_list_excel(rows, foodbank_profile):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    wb = Workbook()
    ws = wb.active
    ws.title = "All Requests"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = [
        'S/No', 'Date', 'Category', 'Type', 'Description', 'Quantity Requested', 'Stock In',
        'Used Stock', 'Remaining Stock', 'Unit', 'Quantity Detail', 'Amount (KES)', 'Donor', 'Delivery', 'Status', 'Donor Note', 'Acceptance Note', 'Created'
    ]
    num_cols = len(headers)
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'].value = f"{foodbank_profile.foodbank_name} - Direct Donations Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    ws['A2'].value = f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M')} | Total: {len(rows)}"
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    ws.row_dimensions[4].height = 22
    wrap_align = Alignment(wrap_text=True, vertical='top')
    for i, r in enumerate(rows, 1):
        row_num = i + 4
        ws.cell(row=row_num, column=1, value=r['idx']).border = border
        ws.cell(row=row_num, column=2, value=r['date']).border = border
        ws.cell(row=row_num, column=3, value=r['category']).border = border
        ws.cell(row=row_num, column=4, value=r['type']).border = border
        cell_desc = ws.cell(row=row_num, column=5, value=(r['description'] or '').replace('\n', ' '))
        cell_desc.border = border
        cell_desc.alignment = wrap_align
        ws.cell(row=row_num, column=6, value=r.get('requested', '')).border = border
        ws.cell(row=row_num, column=7, value=r.get('fulfilled', '')).border = border
        ws.cell(row=row_num, column=8, value=r.get('used_stock', '')).border = border
        ws.cell(row=row_num, column=9, value=r.get('remaining_stock', '')).border = border
        ws.cell(row=row_num, column=10, value=r.get('unit', '')).border = border
        cell_qty_detail = ws.cell(row=row_num, column=11, value=(r.get('quantity_detail') or '-'))
        cell_qty_detail.border = border
        cell_qty_detail.alignment = wrap_align
        ws.cell(row=row_num, column=12, value=r.get('amount_kes', '-')).border = border
        ws.cell(row=row_num, column=13, value=r['donor']).border = border
        ws.cell(row=row_num, column=14, value=r['delivery']).border = border
        ws.cell(row=row_num, column=15, value=r['status']).border = border
        donor_note_wrapped = _wrap_export_text(r.get('donor_note') or '', width=36, max_lines=4)
        acceptance_note_wrapped = _wrap_export_text(r.get('acceptance_note') or '', width=36, max_lines=4)
        cell_donor_note = ws.cell(row=row_num, column=16, value=donor_note_wrapped)
        cell_donor_note.border = border
        cell_donor_note.alignment = wrap_align
        cell_acceptance_note = ws.cell(row=row_num, column=17, value=acceptance_note_wrapped)
        cell_acceptance_note.border = border
        cell_acceptance_note.alignment = wrap_align
        ws.cell(row=row_num, column=18, value=r['created']).border = border
        if ('\n' in (r.get('quantity_detail') or '')) or ('\n' in donor_note_wrapped) or ('\n' in acceptance_note_wrapped):
            ws.row_dimensions[row_num].height = 45
    col_widths = [8, 12, 14, 12, 28, 14, 12, 10, 12, 10, 36, 14, 22, 14, 16, 28, 28, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{foodbank_profile.foodbank_name.replace(" ", "_")}_all_requests_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required

def export_foodbank_requests(request, format):

    """Export foodbank requests in PDF or CSV format"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')



    foodbank_profile = get_object_or_404(FoodBankProfile, user=request.user)



    # Get filter parameters (same as foodbank_requests_view)
    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')
    anonymous_filter = request.GET.get('anonymous', 'all')
    type_filter = (request.GET.get('type') or request.GET.get('category') or 'all').strip().lower()
    if type_filter not in ('all', 'food', 'non_food'):
        type_filter = 'all'
    category_filter = request.GET.get('category', 'all')
    delivery_filter = request.GET.get('delivery', 'all')
    if delivery_filter == 'both':
        delivery_filter = 'all'
    date_filter = request.GET.get('date_range', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')
    search_query = request.GET.get('search', '').strip()
    sort_filter = request.GET.get('sort', 'newest')

    

    # Get requests for this foodbank (same query as foodbank_requests_view)

    requests = RequestManagement.objects.filter(

        Q(foodbank=foodbank_profile) |

        Q(is_anonymous=True, assigned_foodbank=foodbank_profile) |

        Q(is_anonymous=True, assigned_foodbank__isnull=True)

    ).select_related('recipient', 'updated_by', 'donation__donor', 'donation__donor__donor_profile').prefetch_related('foodbank_request_created__donations')

    

    # Apply filters

    if status_filter != 'all':

        requests = requests.filter(status=status_filter)

    

    if type_filter != 'all':

        requests = requests.filter(request_type=type_filter)

    if anonymous_filter == 'anonymous':

        requests = requests.filter(is_anonymous=True)

    elif anonymous_filter == 'not_anonymous':

        requests = requests.filter(is_anonymous=False)

    

    if delivery_filter != 'all':

        requests = requests.filter(delivery_method=delivery_filter)

    

    # Date filtering

    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()

    

    if (date_filter == 'custom') or (date_filter in ('', 'all') and (date_from or date_to)):

        if date_from:

            try:

                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')

                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))

                requests = requests.filter(time_of_request__gte=from_date)

            except ValueError:

                pass

        if date_to:

            try:

                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')

                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))

                requests = requests.filter(time_of_request__lte=to_date)

            except ValueError:

                pass

    elif date_filter != 'all':

        now = timezone.now()

        if date_filter == 'today':

            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)

            requests = requests.filter(time_of_request__gte=start_date)

        elif date_filter == 'week':

            start_date = now - timedelta(days=7)

            requests = requests.filter(time_of_request__gte=start_date)

        elif date_filter == 'month':

            start_date = now - timedelta(days=30)

            requests = requests.filter(time_of_request__gte=start_date)

        elif date_filter == '3months':

            start_date = now - timedelta(days=90)

            requests = requests.filter(time_of_request__gte=start_date)

    

    # Quantity filtering

    if quantity_filter != 'all':

        if quantity_filter == 'small':

            requests = requests.filter(quantity__lte=10)

        elif quantity_filter == 'medium':

            requests = requests.filter(quantity__gt=10, quantity__lte=50)

        elif quantity_filter == 'large':

            requests = requests.filter(quantity__gt=50)

    

    # Search filtering

    if search_query:

        requests = requests.filter(

            Q(description__icontains=search_query) |

            Q(location__icontains=search_query) |

            Q(recipient__full_name__icontains=search_query) |

            Q(additional_notes__icontains=search_query)

        )

    

    # Sorting

    if sort_filter == 'oldest':

        requests = requests.order_by('time_of_request')

    else:

        requests = requests.order_by('-time_of_request')



    # Get all records (no pagination for export)

    requests_data = list(requests)



    if format.lower() == 'pdf':

        return export_foodbank_requests_pdf(request, requests_data, foodbank_profile)

    elif format.lower() == 'csv':

        return export_foodbank_requests_csv(request, requests_data, foodbank_profile)

    elif format.lower() == 'excel':

        return export_foodbank_requests_excel(request, requests_data, foodbank_profile)

    else:

        messages.error(request, 'Invalid export format.')

        return redirect('foodbank_requests_view')





def get_donation_info_for_request(req):

    """Helper function to get donation info for a request"""

    try:

        foodbank_request = req.foodbank_request_created.first()

        if foodbank_request:

            donation = foodbank_request.donations.first()

            if donation:

                if donation.donation_type == 'money':

                    return f"Monetary: KES {donation.amount}"

                elif donation.donation_type == 'subsidized':

                    return f"Subsidized: KES {donation.subsidized_price}"

                else:

                    return f"Free: {donation.quantity} {donation.quantity_unit}"

    except:

        pass

    return "-"


def _manage_requests_quantity_values(req):
    """Return quantity values exactly as rendered in the manage-requests Quantity cell."""
    unit = (req.get_unit_display() or '').strip() or "-"
    requested = getattr(req, 'quantity', 0) or 0
    fulfilled = getattr(req, 'quantity_fulfilled', 0) or 0
    try:
        remaining = req.get_remaining_quantity() if hasattr(req, 'get_remaining_quantity') else max(0, requested - fulfilled)
    except Exception:
        remaining = max(0, requested - fulfilled)
    return requested, fulfilled, remaining, unit


def _manage_requests_quantity_text(req, line_break="\n"):
    """Build Quantity text exactly like the table cell: requested + fulfilled/remaining."""
    requested, fulfilled, remaining, unit = _manage_requests_quantity_values(req)
    requested_text = f"{requested} {unit}".strip()
    fulfilled_text = f"{fulfilled} {unit}".strip()
    remaining_text = f"{remaining} {unit}".strip()
    return (
        f"{requested_text} Requested"
        f"{line_break}{fulfilled_text} Fulfilled | {remaining_text} Remaining"
    )


def _request_used_amount(req):
    """Return amount actually used for this request (prefer allocation amounts)."""
    try:
        allocations = req.donation_allocations.filter(declined_by_recipient=False)
    except Exception:
        allocations = []

    allocated_total = 0.0
    has_allocated_amount = False
    for allocation in allocations:
        amount = getattr(allocation, 'amount', None)
        if amount in (None, ''):
            continue
        try:
            allocated_total += float(amount)
            has_allocated_amount = True
        except (TypeError, ValueError):
            continue

    if has_allocated_amount:
        return round(allocated_total, 2)

    # Fallback to request-targeted donations only (exclude stock-source donations).
    donation_map = {}

    def add_donation(donation):
        if not donation:
            return
        donation_id = getattr(donation, 'id', None)
        if donation_id:
            donation_map[donation_id] = donation

    add_donation(getattr(req, 'donation', None))

    fb_request = getattr(req, 'foodbank_request', None)
    if fb_request:
        for donation in fb_request.donations.all():
            add_donation(donation)

    related_requests = list(req.foodbank_request_created.all()) + list(req.donor_requests.all())
    for linked_request in related_requests:
        for donation in linked_request.donations.all():
            add_donation(donation)

    for donation in req.donations.all():
        add_donation(donation)

    total_amount = 0.0
    has_amount = False

    for donation in donation_map.values():
        donation_type = getattr(donation, 'donation_type', None)
        if donation_type not in ('money', 'subsidized'):
            continue

        if getattr(donation, 'status', None) == 'declined' or getattr(donation, 'declined_by_recipient_id', None):
            continue

        if donation_type == 'money':
            amount = getattr(donation, 'amount', None)
        else:
            amount = getattr(donation, 'subsidized_price', None)
            if amount in (None, ''):
                amount = getattr(donation, 'amount', None)
        if amount in (None, ''):
            continue

        try:
            total_amount += float(amount)
            has_amount = True
        except (TypeError, ValueError):
            continue

    return round(total_amount, 2) if has_amount else ''


def _manage_requests_money_amount(req):
    """Return report amount for manage-requests exports (amount used for request)."""
    return _request_used_amount(req)


def _manage_requests_qty_details_breakdown(req, line_break="\n", include_timestamps=True):
    """Return manage-view quantity breakdown lines like the View Breakdown modal."""
    timeline = _build_request_quantity_timeline(req)
    if not timeline:
        return "-"

    lines = []
    for entry in timeline:
        display = (entry.get('display') or '').strip()
        if not display:
            continue

        if entry.get('label'):
            display = f"{display} ({entry['label']})"
        if entry.get('is_declined'):
            display = f"{display} (Declined)"

        lines.append(display)

        if include_timestamps:
            timestamp = entry.get('timestamp')
            if timestamp:
                try:
                    ts_text = timezone.localtime(timestamp).strftime('%b %d, %Y %H:%M')
                except Exception:
                    ts_text = str(timestamp)
                lines.append(ts_text)

    return line_break.join(lines) if lines else "-"


def _manage_requests_requested_display(req, multiline=False):
    """Return requested datetime in the same shape as table cells."""
    dt = getattr(req, 'time_of_request', None)
    if not dt:
        return "-"
    local_dt = timezone.localtime(dt)
    if multiline:
        return f"{local_dt.strftime('%b %d, %Y')}\n{local_dt.strftime('%H:%M')}"
    return f"{local_dt.strftime('%b %d, %Y')} {local_dt.strftime('%H:%M')}"


def _manage_requests_status_display(req):
    """Return (status_label, decline_reason) matching foodbank_manage_requests.html table exactly."""
    status = getattr(req, 'status', None)
    decline_reason = (getattr(req, 'decline_message', None) or '').strip() or None
    if status == 'partial':
        return "Partially fulfilled sent to recipient", decline_reason
    if status == 'fulfilled':
        label = "Recipient Recieved" if getattr(req, 'acknowledged_by_recipient', False) else "Fulfilled"
        return label, decline_reason
    if status == 'declined':
        if getattr(req, 'recipient_declined_request', False):
            return "Declined by Recipient", decline_reason
        if getattr(req, 'foodbank_declined_request', False) or getattr(req, 'foodbank_declined_donation', False):
            return "Declined by Foodbank", decline_reason
        return "Declined", decline_reason
    if status == 'donation_received':
        return "Donated awaiting approval", decline_reason
    if status == 'submitted':
        return "Submitted to Donors - Awaiting Donation", decline_reason
    return (req.get_status_display(), decline_reason)


def _manage_requests_decline_reason(req):
    """Return decline reason for exports (foodbank or recipient), otherwise '-'."""
    status_label, decline_reason = _manage_requests_status_display(req)
    if getattr(req, 'status', None) != 'declined':
        return "-"
    return decline_reason or "-"


def _manage_requests_notes_columns(req):
    """Return (recipient_note, donor_note) matching the manage-requests notes modal data."""
    recipient_note = (getattr(req, 'additional_notes', None) or '').strip() or "-"

    donor_bits = []
    fb_req_rel = getattr(req, 'foodbank_request_created', None)
    if fb_req_rel:
        fb_req = fb_req_rel.first()
        if fb_req and getattr(fb_req, 'donations', None):
            for d in fb_req.donations.all():
                t = (
                    (getattr(d, 'message', None) or '')
                    or (getattr(d, 'other_description', None) or '')
                    or (getattr(d, 'csr_description', None) or '')
                )
                if isinstance(t, str):
                    t = t.strip()
                    if t:
                        donor_bits.append(t)

    donor_note = " | ".join(donor_bits).strip() if donor_bits else "-"
    return recipient_note, donor_note


def _manage_request_recipient_display(req):
    """Recipient label for foodbank manage-requests table/preview."""
    if getattr(req, 'is_anonymous', False):
        return "Anonymous"
    # Defensive guard: assigned requests without a direct foodbank link are anonymous in this flow.
    if getattr(req, 'foodbank_id', None) is None and getattr(req, 'assigned_foodbank_id', None) is not None:
        return "Anonymous"
    name = (getattr(req, 'recipient_name', None) or '').strip()
    if not name:
        recipient = getattr(req, 'recipient', None)
        name = (getattr(recipient, 'full_name', None) or '').strip() if recipient else ''
    return name or "Anonymous"


def export_foodbank_requests_pdf(request, requests_data, foodbank_profile):

    """Generate branded PDF report for foodbank manage requests."""

    from reportlab.lib.pagesizes import A3, landscape

    from reportlab.platypus import Paragraph, TableStyle

    from .report_utils import (

        get_report_styles, build_report_header, get_branded_table_style,

        build_report_summary, build_pdf_document, collect_active_filters, make_full_width_table,

    )



    styles = get_report_styles()
    report_pagesize = landscape(A3)

    wrap = styles['wrap']

    elements = []



    name = foodbank_profile.foodbank_name or foodbank_profile.user.email

    active_filters = collect_active_filters(request, [

        ('status', 'Status'), ('type', 'Type'), ('delivery', 'Delivery'),

        ('anonymous', 'Anonymous'),

        ('date_range', 'Date Range'), ('quantity', 'Quantity'),

        ('search', 'Search'), ('sort', 'Sort'),

    ])



    build_report_header(

        elements, "Manage Requests Report", name,

        len(requests_data), active_filters, styles,

    )



    def _build_quantity_breakdown_paragraph(req_obj):
        quantity_breakdown = _manage_requests_qty_details_breakdown(req_obj, line_break='\n', include_timestamps=True)
        quantity_breakdown = html.escape(quantity_breakdown).replace('\n', '<br/>')
        return Paragraph(quantity_breakdown, wrap)

    def _build_quantity_text_paragraph(req_obj):
        quantity_text = _manage_requests_quantity_text(req_obj, line_break='\n')
        quantity_text = html.escape(quantity_text).replace('\n', '<br/>')
        return Paragraph(quantity_text, wrap)

    if not requests_data:

        elements.append(Paragraph("No requests found matching the current filters.", styles['normal']))

    else:

        data = [[
            'S/No', 'Type', 'Description', 'Recipient',
            'Quantity Breakdown', 'Quantity Details',
            'Amount (KES)', 'Delivery', 'Location', 'Notes', 'Status', 'Decline Reason', 'Requested'
        ]]



        for idx, req in enumerate(requests_data, 1):

            req_type = req.get_request_type_display()

            description = req.description or "-"

            recipient_name = _manage_request_recipient_display(req)

            quantity_para = _build_quantity_breakdown_paragraph(req)
            quantity_text_para = _build_quantity_text_paragraph(req)
            amount_value = _manage_requests_money_amount(req)
            amount_text = "-" if amount_value == '' else f"{amount_value:,.2f}"

            delivery = req.get_delivery_method_display()

            location = req.location or "-"

            recipient_note, donor_note = _manage_requests_notes_columns(req)

            notes_text = (
                f"<b>Recipient Note:</b><br/>{html.escape(recipient_note or '-')}<br/><br/>"
                f"<b>Donor Note:</b><br/>{html.escape(donor_note or '-')}"
            )

            status_label, _ = _manage_requests_status_display(req)
            decline_reason = _manage_requests_decline_reason(req)

            status_text = html.escape(status_label)
            decline_reason_text = html.escape(decline_reason)
            requested_date = _manage_requests_requested_display(req, multiline=True)
            requested_date = html.escape(requested_date).replace('\n', '<br/>')
            description_text = html.escape(description).replace('\n', '<br/>')



            data.append([

                str(idx),

                Paragraph(req_type, wrap),

                Paragraph(description_text, wrap),

                Paragraph(recipient_name, wrap),

                quantity_para,
                quantity_text_para,
                Paragraph(html.escape(amount_text), wrap),

                Paragraph(delivery, wrap),

                Paragraph(location[:60], wrap),

                Paragraph(notes_text, wrap),

                Paragraph(status_text, wrap),
                Paragraph(decline_reason_text, wrap),

                Paragraph(requested_date, wrap),

            ])



        col_weights = [0.34, 0.66, 1.45, 0.95, 1.8, 1.1, 0.75, 0.7, 0.85, 1.35, 0.95, 1.2, 0.85]
        table = make_full_width_table(
            data,
            repeat_rows=1,
            col_weights=col_weights,
            pagesize=report_pagesize,
            left_margin=8,
            right_margin=8,
        )

        table.setStyle(get_branded_table_style(len(data)))
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ]))

        elements.append(table)



        build_report_summary(elements, [

            ("Total Requests", len(requests_data)),

            ("Pending", sum(1 for r in requests_data if r.status == 'pending')),

            ("Fulfilled", sum(1 for r in requests_data if r.status == 'fulfilled')),

            ("Partial", sum(1 for r in requests_data if r.status == 'partial')),

            ("Declined", sum(1 for r in requests_data if r.status == 'declined')),

        ], styles)



    safe_name = (foodbank_profile.foodbank_name or "foodbank").replace(" ", "_")[:30]

    return build_pdf_document(elements, "manage_requests", safe_name, pagesize=report_pagesize)





def export_foodbank_requests_csv(request, requests_data, foodbank_profile):

    """Generate CSV report for foodbank manage requests"""

    import csv

    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')

    filename = f"{foodbank_profile.foodbank_name.replace(' ', '_')}_manage_requests_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.csv"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'



    writer = csv.writer(response)

    writer.writerow(['Food Bank', foodbank_profile.foodbank_name])

    writer.writerow(['Report Generated', timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')])

    writer.writerow([])



    # Quantity is split into explicit columns for spreadsheet exports.
    writer.writerow([
        'S/No', 'Type', 'Description', 'Recipient',
        'Quantity Requested', 'Quantity Fulfilled', 'Quantity Remaining', 'Unit',
        'Quantity Details', 'Amount (KES)',
        'Delivery', 'Location', 'Recipient Note', 'Donor Note', 'Status', 'Decline Reason', 'Requested Date/Time'
    ])



    # Data

    for idx, req in enumerate(requests_data, 1):

        req_type = req.get_request_type_display()

        description = req.description if req.description else "-"

        recipient_name = _manage_request_recipient_display(req)

        qty_requested, qty_fulfilled, qty_remaining, qty_unit = _manage_requests_quantity_values(req)
        quantity_details = _manage_requests_qty_details_breakdown(req, line_break=' | ', include_timestamps=True)
        amount_value = _manage_requests_money_amount(req)

        delivery = req.get_delivery_method_display()

        location = req.location if req.location else "-"

        recipient_note, donor_note = _manage_requests_notes_columns(req)

        status_label, _ = _manage_requests_status_display(req)
        decline_reason = _manage_requests_decline_reason(req)

        requested_date = _manage_requests_requested_display(req, multiline=True)

        writer.writerow([

            idx,

            req_type,

            description,

            recipient_name,

            qty_requested,

            qty_fulfilled,

            qty_remaining,

            qty_unit,

            quantity_details,

            amount_value if amount_value != '' else '',

            delivery,

            location,

            recipient_note,
            donor_note,

            status_label,
            decline_reason,

            requested_date,

        ])



    # Summary

    writer.writerow([])

    writer.writerow(['Summary'])

    writer.writerow(['Total Requests', len(requests_data)])

    writer.writerow(['Pending', sum(1 for r in requests_data if r.status == 'pending')])

    writer.writerow(['Fulfilled', sum(1 for r in requests_data if r.status == 'fulfilled')])

    writer.writerow(['Declined', sum(1 for r in requests_data if r.status == 'declined')])

    writer.writerow(['Partial', sum(1 for r in requests_data if r.status == 'partial')])



    return response





def export_foodbank_requests_excel(request, requests_data, foodbank_profile):

    """Generate Excel report for foodbank manage requests - columns match table exactly"""

    from django.http import HttpResponse

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws = wb.active

    ws.title = "Manage Requests"



    # Define styles

    header_font = Font(bold=True, color="FFFFFF", size=10)

    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )



    # Title and info
    ws.merge_cells('A1:Q1')

    title_cell = ws['A1']

    title_cell.value = f"{foodbank_profile.foodbank_name} - Manage Requests Report"

    title_cell.font = Font(bold=True, size=14, color="10b981")

    title_cell.alignment = Alignment(horizontal="center")



    ws.merge_cells('A2:Q2')

    info_cell = ws['A2']

    info_cell.value = f"Generated on: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')} | Total Requests: {len(requests_data)}"

    info_cell.alignment = Alignment(horizontal="center")



    # Quantity is split into explicit columns for spreadsheet exports.
    headers = [
        'S/No', 'Type', 'Description', 'Recipient',
        'Quantity Requested', 'Quantity Fulfilled', 'Quantity Remaining', 'Unit',
        'Quantity Details', 'Amount (KES)',
        'Delivery', 'Location', 'Recipient Note', 'Donor Note', 'Status', 'Decline Reason', 'Requested Date/Time'
    ]



    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=4, column=col_num)

        cell.value = header

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = header_alignment

        cell.border = border



    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    # Data rows

    for idx, req in enumerate(requests_data, 1):

        row_num = idx + 4

        req_type = req.get_request_type_display()

        description = req.description if req.description else "-"

        recipient_name = _manage_request_recipient_display(req)

        qty_requested, qty_fulfilled, qty_remaining, qty_unit = _manage_requests_quantity_values(req)
        quantity_details = _manage_requests_qty_details_breakdown(req, line_break='\n', include_timestamps=True)
        amount_value = _manage_requests_money_amount(req)

        delivery = req.get_delivery_method_display()

        location = req.location if req.location else "-"

        recipient_note, donor_note = _manage_requests_notes_columns(req)

        status_label, _ = _manage_requests_status_display(req)
        decline_reason = _manage_requests_decline_reason(req)
        requested_date = _manage_requests_requested_display(req, multiline=True)

        ws.cell(row=row_num, column=1, value=idx).border = border

        ws.cell(row=row_num, column=2, value=req_type).border = border

        ws.cell(row=row_num, column=3, value=description).border = border

        ws.cell(row=row_num, column=4, value=recipient_name).border = border

        ws.cell(row=row_num, column=5, value=qty_requested).border = border
        ws.cell(row=row_num, column=6, value=qty_fulfilled).border = border
        ws.cell(row=row_num, column=7, value=qty_remaining).border = border
        ws.cell(row=row_num, column=8, value=qty_unit).border = border

        quantity_cell = ws.cell(row=row_num, column=9, value=quantity_details)
        quantity_cell.border = border
        quantity_cell.alignment = wrap_align

        amount_cell = ws.cell(row=row_num, column=10, value=amount_value if amount_value != '' else None)
        amount_cell.border = border
        if amount_value != '':
            amount_cell.number_format = '#,##0.00'

        ws.cell(row=row_num, column=11, value=delivery).border = border
        ws.cell(row=row_num, column=12, value=location).border = border

        recipient_note_cell = ws.cell(row=row_num, column=13, value=recipient_note)
        recipient_note_cell.border = border
        recipient_note_cell.alignment = wrap_align

        donor_note_cell = ws.cell(row=row_num, column=14, value=donor_note)
        donor_note_cell.border = border
        donor_note_cell.alignment = wrap_align

        ws.cell(row=row_num, column=15, value=status_label).border = border
        decline_reason_cell = ws.cell(row=row_num, column=16, value=decline_reason)
        decline_reason_cell.border = border
        decline_reason_cell.alignment = wrap_align

        requested_cell = ws.cell(row=row_num, column=17, value=requested_date)
        requested_cell.border = border
        requested_cell.alignment = wrap_align

        for col in range(1, 18):
            c = ws.cell(row=row_num, column=col)
            if not getattr(c.alignment, 'wrap_text', False):
                c.alignment = top_align



    # Summary section

    summary_row = len(requests_data) + 6

    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)

    ws.cell(row=summary_row + 1, column=1, value="Total Requests")

    ws.cell(row=summary_row + 1, column=2, value=len(requests_data))

    ws.cell(row=summary_row + 2, column=1, value="Pending")

    ws.cell(row=summary_row + 2, column=2, value=sum(1 for r in requests_data if r.status == 'pending'))

    ws.cell(row=summary_row + 3, column=1, value="Fulfilled")

    ws.cell(row=summary_row + 3, column=2, value=sum(1 for r in requests_data if r.status == 'fulfilled'))

    ws.cell(row=summary_row + 4, column=1, value="Declined")

    ws.cell(row=summary_row + 4, column=2, value=sum(1 for r in requests_data if r.status == 'declined'))

    ws.cell(row=summary_row + 5, column=1, value="Partial")

    ws.cell(row=summary_row + 5, column=2, value=sum(1 for r in requests_data if r.status == 'partial'))



    # Adjust column widths (17 columns including separate Recipient/Donor Note columns).
    column_widths = [6, 12, 28, 16, 13, 13, 13, 10, 28, 14, 12, 18, 24, 24, 20, 28, 18]

    for col_num, width in enumerate(column_widths, 1):

        ws.column_dimensions[get_column_letter(col_num)].width = width



    # Create response

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    filename = f"{foodbank_profile.foodbank_name.replace(' ', '_')}_manage_requests_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'



    wb.save(response)

    return response





@login_required

def export_recipient_requests(request, format):

    """Export recipient requests in PDF, CSV, or Excel format.

    Mirrors all filters from recipient_requests_view so the export

    matches exactly what the user sees on screen."""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient privileges required.')

        return redirect('dashboard')



    recipient_profile = get_object_or_404(RecipientProfile, user=request.user)

    # Base queryset â€“ same as recipient_requests_view
    requests_qs = RequestManagement.objects.filter(
        recipient=recipient_profile
    ).select_related(
        'foodbank', 'updated_by', 'foodbank_request', 'donation'
    ).prefetch_related(
        'foodbank_request__donations',
        'foodbank_request_created__donations',
        'donor_requests__donations',
        'donations',
        'donation_allocations__donation'
    ).order_by('-time_of_request')
    # â”€â”€ Apply ALL filters identical to the main view â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    status_filter = request.GET.get('status', 'all')

    if status_filter != 'all':

        requests_qs = requests_qs.filter(status=status_filter)



    type_filter = request.GET.get('type', 'all')
    anonymous_filter = request.GET.get('anonymous', 'all')

    category_filter = request.GET.get('category', 'all')

    request_type_choices = {'all', 'food', 'non_food'}
    donation_type_choices = {'all', 'free_goods', 'subsidized', 'money'}

    # Backward compatibility: older UI had type/category semantics swapped.
    if type_filter in donation_type_choices and category_filter in request_type_choices:
        type_filter, category_filter = category_filter, type_filter
    elif type_filter in donation_type_choices and category_filter == 'all':
        category_filter = type_filter
        type_filter = 'all'
    elif category_filter in request_type_choices and type_filter == 'all':
        type_filter = category_filter
        category_filter = 'all'

    if type_filter not in request_type_choices:
        type_filter = 'all'
    if category_filter not in donation_type_choices:
        category_filter = 'all'

    quantity_filter = (request.GET.get('quantity', 'all') or 'all').strip().lower()
    amount_filter = (request.GET.get('amount', 'all') or 'all').strip().lower()
    valid_range_filters = {'all', 'small', 'medium', 'large'}
    if quantity_filter not in valid_range_filters:
        quantity_filter = 'all'
    if amount_filter not in valid_range_filters:
        amount_filter = 'all'

    if type_filter != 'all':
        requests_qs = requests_qs.filter(
            Q(request_category=type_filter) | Q(request_type=type_filter)
        )

    donation_filter = _build_recipient_type_filter(category_filter)

    if donation_filter is not None:

        requests_qs = requests_qs.filter(donation_filter).distinct()

    if quantity_filter == 'small':
        requests_qs = requests_qs.filter(quantity__gte=1, quantity__lte=50)
    elif quantity_filter == 'medium':
        requests_qs = requests_qs.filter(quantity__gt=50, quantity__lte=200)
    elif quantity_filter == 'large':
        requests_qs = requests_qs.filter(quantity__gt=200)

    amount_q = _build_recipient_amount_filter(amount_filter)
    if amount_q is not None:
        requests_qs = requests_qs.filter(amount_q).distinct()

    if anonymous_filter == 'anonymous':

        requests_qs = requests_qs.filter(is_anonymous=True)

    elif anonymous_filter == 'not_anonymous':

        requests_qs = requests_qs.filter(is_anonymous=False)



    delivery_filter = request.GET.get('delivery', 'all')
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'

    if delivery_filter != 'all':

        requests_qs = requests_qs.filter(delivery_method=delivery_filter)



    acknowledgment_filter = request.GET.get('acknowledgment', 'all')

    if acknowledgment_filter != 'all':

        if acknowledgment_filter == 'received':

            requests_qs = requests_qs.filter(

                Q(additional_notes__icontains='Receipt Confirmed') |

                Q(status='donation_received')

            ).exclude(status='declined')

        elif acknowledgment_filter == 'acknowledged':

            requests_qs = requests_qs.filter(

                Q(acknowledged_by_recipient=True) &

                ~Q(additional_notes__icontains='Receipt Confirmed')

            )

        elif acknowledgment_filter == 'not_acknowledged':

            requests_qs = requests_qs.filter(

                Q(status__in=['fulfilled', 'partial']) &

                Q(acknowledged_by_recipient=False)

            )



    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()



    if date_from:

        try:

            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')

            from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))

            requests_qs = requests_qs.filter(time_of_request__gte=from_date)

        except ValueError:

            pass

    if date_to:

        try:

            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')

            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))

            requests_qs = requests_qs.filter(time_of_request__lte=to_date)

        except ValueError:

            pass



    search_query = request.GET.get('search', '').strip()

    if search_query:

        requests_qs = requests_qs.filter(

            Q(description__icontains=search_query) |

            Q(location__icontains=search_query) |

            Q(foodbank__foodbank_name__icontains=search_query) |

            Q(additional_notes__icontains=search_query)

        )



    sort_filter = request.GET.get('sort', 'newest')

    if sort_filter == 'oldest':

        requests_qs = requests_qs.order_by('time_of_request')

    else:

        requests_qs = requests_qs.order_by('-time_of_request')



    # Get all records (no pagination for export)

    requests_data = list(requests_qs)
    for req in requests_data:
        req.primary_donation = _get_request_primary_donation(req)
        if req.primary_donation:
            req.primary_donor_note = (
                req.primary_donation.message
                or req.primary_donation.csr_description
                or req.primary_donation.other_description
                or ""
            )
        else:
            req.primary_donor_note = ""



    if format.lower() == 'pdf':

        return export_recipient_requests_pdf(request, requests_data, recipient_profile)

    elif format.lower() == 'csv':

        return export_recipient_requests_csv(request, requests_data, recipient_profile)

    elif format.lower() == 'excel':

        return export_recipient_requests_excel(request, requests_data, recipient_profile)

    else:

        messages.error(request, 'Invalid export format.')

        return redirect('recipient_requests_view')



def _recipient_requests_type_display(req):
    donation = getattr(req, 'primary_donation', None) or _get_request_primary_donation(req)
    if donation:
        type_map = {'item': 'Free Goods', 'money': 'Money', 'subsidized': 'Subsidized', 'csr': 'CSR'}
        return type_map.get(donation.donation_type, donation.get_donation_type_display())
    if getattr(req, 'is_anonymous', False):
        return 'Anonymous'
    return 'Request'


def _recipient_requests_category_display(req):
    category = getattr(req, 'request_category', None) or getattr(req, 'request_type', None)
    return {'food': 'Food', 'non_food': 'Non-Food'}.get(category, 'Other')


def _recipient_requests_foodbank_display(req):
    name = getattr(req, 'foodbank_name', None)
    if name:
        return name
    foodbank = getattr(req, 'foodbank', None)
    if foodbank and getattr(foodbank, 'foodbank_name', None):
        return foodbank.foodbank_name
    return "-"


def _recipient_requests_notes_columns(req):
    recipient_note = (getattr(req, 'additional_notes', None) or '').strip() or "-"
    donor_note = (
        getattr(req, 'primary_donor_note', None)
        if hasattr(req, 'primary_donor_note')
        else None
    )
    if donor_note is None:
        donation = getattr(req, 'primary_donation', None) or _get_request_primary_donation(req)
        donor_note = (
            (getattr(donation, 'message', None) or '')
            or (getattr(donation, 'csr_description', None) or '')
            or (getattr(donation, 'other_description', None) or '')
        ) if donation else ''
    donor_note = donor_note.strip() if isinstance(donor_note, str) else ''
    return recipient_note, (donor_note or "-")


def _recipient_requests_quantity_values(req):
    unit = (req.get_unit_display() or '').strip() or "-"
    requested = getattr(req, 'quantity', 0) or 0
    fulfilled = getattr(req, 'quantity_fulfilled', 0) or 0
    try:
        remaining = req.get_remaining_quantity() if hasattr(req, 'get_remaining_quantity') else max(0, requested - fulfilled)
    except Exception:
        remaining = max(0, requested - fulfilled)
    return requested, fulfilled, remaining, unit


def _recipient_requests_quantity_text(req, line_break="\n"):
    requested, fulfilled, remaining, unit = _recipient_requests_quantity_values(req)
    return (
        f"{requested} {unit} Requested"
        f"{line_break}{fulfilled} {unit} Fulfilled | {remaining} {unit} Remaining"
    )


def _recipient_requests_qty_details_breakdown(req, line_break="\n"):
    """Return recipient-visible quantity breakdown lines like the View Breakdown modal."""
    raw_timeline = _build_request_quantity_timeline(req)
    timeline, _ = _build_recipient_timeline(raw_timeline)
    if not timeline:
        return "-"

    lines = []
    for entry in timeline:
        display = (entry.get('display') or '').strip()
        if not display:
            continue
        is_declined = bool(entry.get('is_declined') or entry.get('state') == 'declined')
        if is_declined:
            display = f"{display} (Declined)"
        lines.append(display)
        timestamp = entry.get('timestamp')
        if timestamp:
            try:
                ts_text = timezone.localtime(timestamp).strftime('%b %d, %Y %H:%M')
            except Exception:
                ts_text = str(timestamp)
            lines.append(ts_text)

    return line_break.join(lines) if lines else "-"


def _recipient_requests_money_amount(req):
    """Return report amount for recipient exports (amount used for request)."""
    return _request_used_amount(req)


def _recipient_requests_status_display(req):
    notes = getattr(req, 'additional_notes', None) or ""
    status = getattr(req, 'status', None)
    if status == 'partial':
        label = 'Partially Fulfilled'
        if "Receipt Confirmed" in notes:
            label += " (Received)"
        elif getattr(req, 'acknowledged_by_recipient', False):
            label += " (Acknowledged)"
        return label
    if status == 'awaiting_recipient':
        return "Awaiting Your Response"
    if status == 'fulfilled':
        if "Receipt Confirmed" in notes:
            return "Received"
        if getattr(req, 'acknowledged_by_recipient', False):
            return "Acknowledged"
        return "Fulfilled"
    if status == 'donation_received':
        return "Donated awaiting approval"
    if status == 'submitted':
        return "Sent to Donors"
    if status == 'declined':
        if getattr(req, 'foodbank_declined_request', False) or getattr(req, 'foodbank_declined_donation', False):
            return "Declined by Foodbank"
        if getattr(req, 'recipient_declined_request', False):
            return "Declined by Recipient"
        return "Declined"
    return "Sent to FB"


def _recipient_requests_decline_reason(req):
    """Return decline reason for recipient exports, or '-'."""
    if getattr(req, 'status', None) != 'declined':
        return "-"

    reason = (getattr(req, 'decline_message', None) or '').strip()
    if reason:
        return reason

    # Fallback to linked declined donation message (if present).
    for donation in _iter_request_linked_donations(req):
        if getattr(donation, 'status', None) == 'declined':
            donation_reason = (getattr(donation, 'decline_message', None) or '').strip()
            if donation_reason:
                return donation_reason
    return "-"


def _recipient_requests_requested_display(req, multiline=False):
    dt = getattr(req, 'time_of_request', None)
    if not dt:
        return "-"
    local_dt = timezone.localtime(dt)
    date_text = local_dt.strftime('%b %d, %Y')
    time_text = local_dt.strftime('%H:%M')
    fulfilled_at = getattr(req, 'fulfilled_at', None)
    fulfilled_text = None
    if fulfilled_at:
        fulfilled_text = f"Fulfilled {timezone.localtime(fulfilled_at).strftime('%b %d')}"
    if multiline:
        parts = [date_text, time_text]
        if fulfilled_text:
            parts.append(fulfilled_text)
        return "\n".join(parts)
    if fulfilled_text:
        return f"{date_text} {time_text} | {fulfilled_text}"
    return f"{date_text} {time_text}"


def export_recipient_requests_pdf(request, requests_data, recipient_profile):

    """Generate branded PDF report for recipient requests (table-matching columns/content)."""

    from reportlab.platypus import Paragraph
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle
    from .report_utils import (
        get_report_styles, build_report_header, get_branded_table_style,
        build_report_summary, build_pdf_document, collect_active_filters, make_full_width_table,
    )

    styles = get_report_styles()
    report_pagesize = landscape(A3)
    wrap = styles['wrap']
    wrap_small = ParagraphStyle('RecipientCellWrapSmall', parent=wrap, fontSize=7, leading=8.4)
    elements = []

    recipient_name = recipient_profile.full_name or recipient_profile.user.email

    active_filters = collect_active_filters(request, [
        ('status', 'Status'), ('type', 'Type'), ('delivery', 'Delivery'),
        ('anonymous', 'Anonymous'),
        ('acknowledgment', 'Acknowledgment'), ('date_from', 'Start Date'), ('date_to', 'End Date'),
        ('search', 'Search'),
    ])

    build_report_header(
        elements, "My Requests Report", recipient_name,
        len(requests_data), active_filters, styles,
    )

    if not requests_data:
        elements.append(Paragraph("No requests found matching the current filters.", styles['normal']))
    else:
        headers = [
            'S/No', 'Type', 'Category', 'Description', 'Foodbank',
            'Quantity Details', 'Quantity Breakdown', 'Amount (KES)',
            'Delivery', 'Location', 'Status', 'Decline Reason', 'Requested Date/Time', 'Notes'
        ]
        data = [headers]

        for idx, req in enumerate(requests_data, 1):
            recipient_note, donor_note = _recipient_requests_notes_columns(req)
            notes_text = f"<b>Recipient:</b> {html.escape(recipient_note)}<br/><b>Donor:</b> {html.escape(donor_note)}"
            quantity_text = html.escape(_recipient_requests_quantity_text(req, line_break='\n')).replace('\n', '<br/>')
            qty_details_text = html.escape(_recipient_requests_qty_details_breakdown(req, line_break='\n')).replace('\n', '<br/>')
            amount_value = _recipient_requests_money_amount(req)
            amount_text = "-" if amount_value == '' else f"{amount_value:,.2f}"
            decline_reason = _recipient_requests_decline_reason(req)
            requested_text = html.escape(_recipient_requests_requested_display(req, multiline=True)).replace('\n', '<br/>')

            data.append([
                str(idx),
                Paragraph(html.escape(_recipient_requests_category_display(req)), wrap_small),
                Paragraph(html.escape(_recipient_requests_type_display(req)), wrap_small),
                Paragraph(html.escape(req.description or "-"), wrap),
                Paragraph(html.escape(_recipient_requests_foodbank_display(req)), wrap_small),
                Paragraph(quantity_text, wrap_small),
                Paragraph(qty_details_text, wrap_small),
                Paragraph(html.escape(amount_text), wrap_small),
                Paragraph(html.escape(req.get_delivery_method_display() or "-"), wrap_small),
                Paragraph(html.escape(req.location or "-"), wrap_small),
                Paragraph(html.escape(_recipient_requests_status_display(req)), wrap_small),
                Paragraph(html.escape(decline_reason), wrap_small),
                Paragraph(requested_text, wrap_small),
                Paragraph(notes_text, wrap_small),
            ])

        col_weights = [0.28, 0.46, 0.46, 0.95, 0.72, 1.18, 1.28, 0.72, 0.58, 0.66, 0.78, 1.0, 0.76, 0.9]
        table = make_full_width_table(
            data,
            repeat_rows=1,
            col_weights=col_weights,
            pagesize=report_pagesize,
            left_margin=6,
            right_margin=6,
        )
        table.setStyle(get_branded_table_style(len(data)))
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ]))
        elements.append(table)

        build_report_summary(elements, [
            ("Total Requests", len(requests_data)),
        ], styles)

    return build_pdf_document(
        elements,
        "my_requests",
        recipient_name,
        pagesize=report_pagesize,
        left_margin=6,
        right_margin=6,
        top_margin=18,
        bottom_margin=24,
    )



def _get_request_status_display(req):

    """Backward-compatible wrapper for recipient status text in exports."""

    return _recipient_requests_status_display(req)



def export_recipient_requests_csv(request, requests_data, recipient_profile):

    """Generate CSV report for recipient requests with split quantity columns + amount."""

    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    recipient_name = recipient_profile.full_name or recipient_profile.user.email
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_requests_{timestamp}.csv"'

    writer = csv.writer(response, quoting=csv.QUOTE_ALL)

    writer.writerow(['Recipient', recipient_name])
    writer.writerow(['Report Generated', timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Total Records', len(requests_data)])
    writer.writerow([])

    writer.writerow([
        'S/No', 'Type', 'Category', 'Description', 'Foodbank',
        'Quantity Requested', 'Quantity Fulfilled', 'Quantity Remaining', 'Unit',
        'Quantity Details', 'Amount (KES)',
        'Delivery', 'Location', 'Status', 'Decline Reason', 'Requested Date/Time',
        'Recipient Note', 'Donor Note'
    ])

    for idx, req in enumerate(requests_data, 1):
        recipient_note, donor_note = _recipient_requests_notes_columns(req)
        qty_requested, qty_fulfilled, qty_remaining, qty_unit = _recipient_requests_quantity_values(req)
        quantity_details = _recipient_requests_qty_details_breakdown(req, line_break=' | ')
        amount_value = _recipient_requests_money_amount(req)
        decline_reason = _recipient_requests_decline_reason(req)

        writer.writerow([
            idx,
            _recipient_requests_category_display(req),
            _recipient_requests_type_display(req),
            req.description or '-',
            _recipient_requests_foodbank_display(req),
            qty_requested,
            qty_fulfilled,
            qty_remaining,
            qty_unit,
            quantity_details,
            amount_value if amount_value != '' else '',
            req.get_delivery_method_display() or '-',
            getattr(req, 'location', '') or '-',
            _recipient_requests_status_display(req),
            decline_reason,
            _recipient_requests_requested_display(req, multiline=True),
            recipient_note,
            donor_note,
        ])

    writer.writerow([])
    writer.writerow(['Summary'])
    writer.writerow(['Total Requests', len(requests_data)])

    return response



def export_recipient_requests_excel(request, requests_data, recipient_profile):

    """Generate Excel report for recipient requests with split fields + amount."""

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse("Excel export requires openpyxl library. Please install it.")

    from django.http import HttpResponse
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Requests"

    recipient_name = recipient_profile.full_name or recipient_profile.user.email

    ws['A1'] = "FOODBANKHUB"
    ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
    ws['A2'] = f"{recipient_name} - My Requests Report"
    ws['A2'].font = Font(size=13, bold=True, color="1F4E78")
    ws['A3'] = f"Generated on {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}  |  Total Records: {len(requests_data)}"
    ws['A3'].font = Font(size=10, italic=True)

    headers = [
        'S/No', 'Type', 'Category', 'Description', 'Foodbank',
        'Quantity Requested', 'Quantity Fulfilled', 'Quantity Remaining', 'Unit',
        'Quantity Details', 'Amount (KES)',
        'Delivery', 'Location', 'Status', 'Decline Reason', 'Requested Date/Time',
        'Recipient Note', 'Donor Note'
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Guard: ensure trailing note/date headers are always present and styled.
    trailing_headers = {
        'P': 'Requested Date/Time',
        'Q': 'Recipient Note',
        'R': 'Donor Note',
    }
    for col_letter, label in trailing_headers.items():
        cell = ws[f'{col_letter}5']
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    row_num = 6
    for idx, req in enumerate(requests_data, 1):
        recipient_note, donor_note = _recipient_requests_notes_columns(req)
        qty_requested, qty_fulfilled, qty_remaining, qty_unit = _recipient_requests_quantity_values(req)
        quantity_details = _recipient_requests_qty_details_breakdown(req)
        amount_value = _recipient_requests_money_amount(req)
        decline_reason = _recipient_requests_decline_reason(req)

        row_data = [
            idx,
            _recipient_requests_category_display(req),
            _recipient_requests_type_display(req),
            req.description or '-',
            _recipient_requests_foodbank_display(req),
            qty_requested,
            qty_fulfilled,
            qty_remaining,
            qty_unit,
            quantity_details,
            amount_value if amount_value != '' else None,
            req.get_delivery_method_display() or '-',
            req.location or '-',
            _recipient_requests_status_display(req),
            decline_reason,
            _recipient_requests_requested_display(req, multiline=True),
            recipient_note,
            donor_note,
        ]

        is_alt = idx % 2 == 0
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_alt:
                cell.fill = alt_fill
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            if col_num == 11 and value is not None:
                cell.number_format = '#,##0.00'

        row_num += 1

    summary_row = row_num + 2
    ws.cell(row=summary_row, column=1).value = "Summary"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="1F4E78")
    ws.cell(row=summary_row + 1, column=1).value = "Total Requests"
    ws.cell(row=summary_row + 1, column=2).value = len(requests_data)
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)

    column_widths = {
        'A': 6,
        'B': 12,
        'C': 12,
        'D': 30,
        'E': 20,
        'F': 10,
        'G': 10,
        'H': 10,
        'I': 10,
        'J': 32,
        'K': 14,
        'L': 14,
        'M': 18,
        'N': 20,
        'O': 28,
        'P': 18,
        'Q': 24,
        'R': 24,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[5].height = 35

    if len(requests_data) > 0:
        ws.auto_filter.ref = f"A5:R{row_num - 1}"

    ws.freeze_panes = "A6"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_requests_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


@login_required

def foodbank_inventory(request):

    """View for foodbanks to manage their inventory of available donations"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank privileges required.')

        return redirect('dashboard')



    foodbank_profile = request.user.foodbank_profile



    # Base: item, money, and subsidized donations accepted by foodbank (stock)

    base_stock = Donation.objects.filter(

        foodbank=foodbank_profile,

        donation_type__in=['item', 'money', 'subsidized'],

        status='accepted'

    ).select_related('donor', 'donor__donor_profile').annotate(

        allocated_amount=Coalesce(

            Sum('allocations__amount', filter=Q(allocations__declined_by_recipient=False)),

            Value(0)

        )

    ).annotate(

        remaining_amount=Case(

            When(

                donation_type='money',

                then=Coalesce(F('amount'), Value(0)) - F('allocated_amount')

            ),

            When(

                donation_type='subsidized',

                then=Coalesce(F('subsidized_price'), Value(0)) - F('allocated_amount')

            ),

            default=Value(0),

            output_field=DecimalField(max_digits=15, decimal_places=2)

        )

    )



    donations = base_stock.order_by('-donated_at')



    search_query = request.GET.get('search', '')

    category_filter = request.GET.get('category', 'all')

    status_filter = request.GET.get('status', 'all')

    type_filter = request.GET.get('type', 'all')



    if search_query:

        donations = donations.filter(

            Q(item_name__icontains=search_query) |

            Q(description__icontains=search_query) |

            Q(message__icontains=search_query) |

            Q(other_description__icontains=search_query) |

            Q(subsidized_product_type__icontains=search_query) |

            Q(donor__donor_profile__full_name__icontains=search_query) |

            Q(donor__donor_profile__organization_name__icontains=search_query) |

            Q(donor__email__icontains=search_query)

        )



    if category_filter != 'all':

        donations = donations.filter(donation_category=category_filter)



    if type_filter != 'all':

        donations = donations.filter(donation_type=type_filter)



    if status_filter == 'available':

        donations = donations.filter(

            Q(donation_type='item', is_allocated=False) |

            Q(donation_type='money', remaining_amount__gt=0) |

            Q(donation_type='subsidized', remaining_amount__gt=0)

        )

    elif status_filter == 'allocated':

        donations = donations.filter(

            Q(donation_type='item', is_allocated=True) |

            Q(donation_type='money', remaining_amount__lte=0) |

            Q(donation_type='subsidized', remaining_amount__lte=0)

        )



    paginator = Paginator(donations, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)



    stock_base = Donation.objects.filter(

        foodbank=foodbank_profile,

        donation_type__in=['item', 'money', 'subsidized'],

        status='accepted'

    ).annotate(

        allocated_amount=Coalesce(

            Sum('allocations__amount', filter=Q(allocations__declined_by_recipient=False)),

            Value(0)

        )

    ).annotate(

        remaining_amount=Case(

            When(donation_type='money', then=Coalesce(F('amount'), Value(0)) - F('allocated_amount')),

            When(donation_type='subsidized', then=Coalesce(F('subsidized_price'), Value(0)) - F('allocated_amount')),

            default=Value(0),

            output_field=DecimalField(max_digits=15, decimal_places=2)

        )

    )

    total_donations = stock_base.count()

    available_donations = stock_base.filter(

        Q(donation_type='item', is_allocated=False) |

        Q(donation_type='money', remaining_amount__gt=0) |

        Q(donation_type='subsidized', remaining_amount__gt=0)

    ).count()

    allocated_donations = stock_base.filter(

        Q(donation_type='item', is_allocated=True) |

        Q(donation_type='money', remaining_amount__lte=0) |

        Q(donation_type='subsidized', remaining_amount__lte=0)

    ).count()

    categories_count = stock_base.values('donation_category').distinct().count()



    context = {

        'donations': page_obj,

        'total_donations': total_donations,

        'available_donations': available_donations,

        'allocated_donations': allocated_donations,

        'categories_count': categories_count,

        'search_query': search_query,

        'category_filter': category_filter,

        'status_filter': status_filter,

        'type_filter': type_filter,

    }



    return render(request, 'authentication/foodbank_inventory.html', context)



@login_required

def foodbank_settings(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    context = {

        'foodbank_profile': foodbank_profile,

    }

    return render(request, 'authentication/foodbank_settings.html', context)



@login_required

def update_foodbank_profile(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    

    if request.method == 'POST':

        form = FoodBankProfileForm(request.POST, request.FILES, instance=foodbank_profile)

        if form.is_valid():

            form.save()

            messages.success(request, 'Profile updated successfully!')

            return redirect('foodbank_settings')

    else:

        form = FoodBankProfileForm(instance=foodbank_profile)

    

    context = {

        'form': form,

        'foodbank_profile': foodbank_profile,

    }

    return render(request, 'authentication/update_foodbank_profile.html', context)



@login_required

def change_foodbank_password(request):

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = FoodBankPasswordChangeForm(request.user, request.POST)

        if form.is_valid():

            user = form.save()

            update_session_auth_hash(request, user)

            messages.success(request, 'Password changed successfully!')

            return redirect('foodbank_settings')

    else:

        form = FoodBankPasswordChangeForm(request.user)

    

    context = {

        'form': form,

    }

    return render(request, 'authentication/change_foodbank_password.html', context)



@login_required

def foodbank_public_profile(request, foodbank_id):

    """Public profile page for a specific food bank"""

    try:

        foodbank = FoodBankProfile.objects.get(id=foodbank_id)

    except FoodBankProfile.DoesNotExist:

        messages.error(request, 'Food bank not found.')

        return redirect('dashboard')

    

    # Get gallery photos

    gallery_photos = FoodBankGalleryPhoto.objects.filter(foodbank=foodbank)

    featured_photos = gallery_photos.filter(is_featured=True)[:3]

    regular_photos = gallery_photos.filter(is_featured=False)[:6]

    

    # Get recent requests from this food bank

    recent_requests = FoodBankRequest.objects.filter(

        foodbank=foodbank,

        status='active'

    ).order_by('-created_at')[:5]

    

    # Calculate some stats

    total_requests = FoodBankRequest.objects.filter(foodbank=foodbank).count()

    active_requests = FoodBankRequest.objects.filter(foodbank=foodbank, status='active').count()

    donations_received = Donation.objects.filter(foodbank=foodbank).count()

    

    context = {

        'foodbank': foodbank,

        'featured_photos': featured_photos,

        'regular_photos': regular_photos,

        'recent_requests': recent_requests,

        'total_requests': total_requests,

        'active_requests': active_requests,

        'donations_received': donations_received,

        'is_profile_owner': request.user == foodbank.user,

    }

    

    return render(request, 'authentication/foodbank_public_profile.html', context)



@login_required

def manage_foodbank_profile(request):

    """Manage food bank public profile"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    profile = request.user.foodbank_profile

    gallery_photos = FoodBankGalleryPhoto.objects.filter(foodbank=profile)

    

    if request.method == 'POST':

        action = request.POST.get('action')

        

        if action == 'update_profile':

            # Handle profile update

            for field in ['about_text', 'mission_statement', 'contact_email', 'contact_phone', 'website_url', 'established_year']:

                if field in request.POST:

                    setattr(profile, field, request.POST[field] or None)

            

            # Handle header photo upload

            if 'header_photo' in request.FILES:

                profile.header_photo = request.FILES['header_photo']

            

            profile.save()

            messages.success(request, 'Profile updated successfully!')

            

        elif action == 'add_gallery_photo':

            # Handle gallery photo upload

            if 'gallery_photo' in request.FILES:

                caption = request.POST.get('photo_caption', '')

                is_featured = request.POST.get('is_featured') == 'on'

                

                FoodBankGalleryPhoto.objects.create(

                    foodbank=profile,

                    photo=request.FILES['gallery_photo'],

                    caption=caption,

                    is_featured=is_featured

                )

                messages.success(request, 'Photo added to gallery!')

            

        elif action == 'delete_photo':

            # Handle photo deletion

            photo_id = request.POST.get('photo_id')

            try:

                photo = FoodBankGalleryPhoto.objects.get(id=photo_id, foodbank=profile)

                photo.delete()

                messages.success(request, 'Photo deleted successfully!')

            except FoodBankGalleryPhoto.DoesNotExist:

                messages.error(request, 'Photo not found.')

        

        return redirect('manage_foodbank_profile')

    

    context = {

        'profile': profile,

        'gallery_photos': gallery_photos,

        'completion_percentage': profile.get_profile_completion_percentage(),

        'is_complete': profile.is_profile_complete(),

    }

    

    return render(request, 'authentication/manage_foodbank_profile.html', context)



@login_required

def allocate_donation(request, donation_id):

    """Allocate a donation to one or more recipients"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Only food banks can allocate donations.')

        return redirect('dashboard')

    

    try:

        donation = Donation.objects.get(id=donation_id, foodbank=request.user.foodbank_profile)

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found.')

        return redirect('dashboard')

    

    # Check if donation is already fully allocated

    if donation.is_fully_allocated():

        messages.warning(request, 'This donation has been fully allocated.')

        return redirect('dashboard')

    

    # Get remaining amount/quantity

    if donation.donation_type == 'item':

        remaining = donation.get_remaining_quantity()

    else:  # money or subsidized

        remaining = donation.get_remaining_amount()

    

    if request.method == 'POST':

        form = DonationAllocationForm(request.POST, donation=donation)

        if form.is_valid():

            allocation = form.save(commit=False)

            allocation.donation = donation

            allocation.save()

            

            # Create notification for recipient

            Notification.objects.create(

                user=allocation.recipient.user,

                notification_type='donation_received',

                message=f'You have been allocated {allocation.quantity or allocation.amount} from a donation by {donation.donor.email}.'

            )

            

            # Create notification for donor about allocation

            donation_display = f"{allocation.quantity} {donation.unit}" if allocation.quantity else f"KSH {allocation.amount}"

            Notification.objects.create(

                user=donation.donor,

                notification_type='donation_delivered',

                message=f'Your donation ({donation_display}) has been allocated to {allocation.recipient.full_name} by {donation.foodbank.foodbank_name}.'

            )

            

            messages.success(request, f'Donation allocated to {allocation.recipient.full_name} successfully!')

            return redirect('foodbank_donations')

    else:

        form = DonationAllocationForm(donation=donation)

    

    context = {

        'form': form,

        'donation': donation,

        'remaining': remaining,

        'allocations': donation.allocations.all(),

    }

    return render(request, 'authentication/allocate_donation.html', context)



@login_required

def view_donation_allocations(request, donation_id):

    """View all allocations for a specific donation"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Only food banks can view donation allocations.')

        return redirect('dashboard')

    

    try:

        donation = Donation.objects.get(id=donation_id, foodbank=request.user.foodbank_profile)

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found.')

        return redirect('dashboard')

    

    allocations = donation.allocations.all().order_by('-allocated_at')

    

    context = {

        'donation': donation,

        'allocations': allocations,

    }

    return render(request, 'authentication/view_donation_allocations.html', context)



# views.py

#from django.shortcuts import get_object_or_404, redirect

#from django.contrib import messages



@login_required

def accept_request(request, pk):

    req = get_object_or_404(RecipientRequest, pk=pk)



    # Get the current foodbank profile

    foodbank_profile = request.user.foodbank_profile  



    # Update status + assign foodbank if it was anonymous

    req.status = "accepted"

    if req.is_anonymous and req.foodbank is None:

        req.foodbank = foodbank_profile

    req.save()



    # Create notification for the recipient (not the foodbank)



    Notification.create_notification(

        user=req.recipient.user,  # Ensure this points to the recipient's user

        notification_type='acknowledgement',

        message=(

            f"Your request has been accepted by "

            f"{foodbank_profile.foodbank_name} and it is being worked on."

        ),

        related_object=req

    )

    return JsonResponse({"success": True})



@login_required

def decline_request(request, pk):

    foodbank_profile = request.user.foodbank_profile

    req = get_object_or_404(RecipientRequest, pk=pk)

    # Only decline if still pending

    if req.status != "pending":

        return JsonResponse({"success": False, "error": "Request already processed"})



    # If this request is directly assigned to the foodbank, mark as declined

    if req.foodbank == foodbank_profile:

        req.status = "declined"

        req.save()

    else:

        # If it's anonymous, just ignore it (don't update foodbank/status)

        # It remains

        # Anonymous request â†’ mark as declined only for this FB

        req.declined_by.add(foodbank_profile)

   



    Notification.objects.create(

        user=req.recipient.user,

        message=f"Your request '{req.title}' has been declined by {request.user.foodbank_profile.foodbank_name}."

    )



    return JsonResponse({"success": True})



import json

from django.http import HttpResponse

from datetime import datetime



@login_required

def download_recipient_data(request):

    """Allow recipient to download their data"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    # Gather all recipient data

    data = {

        'user_info': {

            'email': request.user.email,

            'user_type': request.user.user_type,

            'date_joined': request.user.date_joined.isoformat(),

        },

        'profile': {

            'full_name': recipient_profile.full_name,

            #'phone_number': recipient_profile.phone_number,

            #'address': recipient_profile.address,

            'is_organization': recipient_profile.is_organization,

            'organization_name': recipient_profile.organization_name if recipient_profile.is_organization else None,

            #'number_of_dependents': recipient_profile.number_of_dependents,

        },

        'requests': [],

        'allocations': [],

        'notifications': [],

    }

    



    # Get all requests

    requests = RecipientRequest.objects.filter(recipient=recipient_profile)

    for req in requests:

        data['requests'].append({

            'id': req.id,

            'title': req.title,

            'description': req.description,

            'status': req.status,

            'is_anonymous': req.is_anonymous,

            'created_at': req.created_at.isoformat(),

            'foodbank': req.foodbank.foodbank_name if req.foodbank else None,

        })

    

    # Get all donation allocations

    allocations = DonationAllocation.objects.filter(recipient=recipient_profile)

    for allocation in allocations:

        data['allocations'].append({

            'id': allocation.id,

            'quantity': str(allocation.quantity) if allocation.quantity else None,

            'amount': str(allocation.amount) if allocation.amount else None,

            'allocated_at': allocation.allocated_at.isoformat(),

            'is_acknowledged': allocation.is_acknowledged,

            'foodbank': allocation.donation.foodbank.foodbank_name if allocation.donation.foodbank else None,

        })

    

    # Get all notifications

    notifications = Notification.objects.filter(user=request.user)

    for notification in notifications:

        data['notifications'].append({

            'id': notification.id,

            'notification_type': notification.notification_type,

            'message': notification.message,

            'created_at': notification.created_at.isoformat(),

            'is_read': notification.is_read,

        })

    

    # Create JSON response

    response = HttpResponse(

        json.dumps(data, indent=2, ensure_ascii=False),

        content_type='application/json'

    )

    

    # Set filename with timestamp

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')

    filename = f'foodbankhub_data_{timestamp}.json'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    

    return response





from django.http import JsonResponse



@login_required

def dismiss_notification(request, pk):

    if request.method == "POST":

        try:

            note = Notification.objects.get(pk=pk, user=request.user)

            note.delete()   # delete is enough

            return JsonResponse({"success": True})

        except Notification.DoesNotExist:

            return JsonResponse({"success": False, "error": "Notification not found"})

#donations to acknowledge

@login_required

def all_donations_to_acknowledge(request):

    recipient_profile = request.user.recipient_profile

    unacknowledged_allocations = DonationAllocation.objects.filter(

        recipient=recipient_profile,

        is_acknowledged=False

    ).order_by('-allocated_at')



    context = {

        "unacknowledged_allocations": unacknowledged_allocations,

    }

    return render(request, "recipient/all_donations_to_acknowledge.html", context)            



# views.py

from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date



@login_required

def view_foodbank_requests(request):

    if request.user.user_type != 'DONOR':

        messages.error(request, "Access denied.")

        return redirect('dashboard')

    

    # Get filter parameters

    location_filter = request.GET.get('location', '').strip()
    search_filter = request.GET.get('search', '').strip()

    priority_filter = request.GET.get('priority', '')
    request_type_filter = (request.GET.get('request_type', '') or '').strip().lower()
    if request_type_filter in ('non-food', 'nonfood'):
        request_type_filter = 'non_food'

    delivery_filter = request.GET.get('delivery', '').strip().lower()

    service_type_filter = request.GET.get('service_type', '')

    sort_by_filter = request.GET.get('sort_by', '')

    due_date_from_filter = request.GET.get('due_date_from', '')

    due_date_to_filter = request.GET.get('due_date_to', '')

    

    

    # Build query for requests - start with active requests

    # Only show active requests where deadline hasn't passed (or no deadline set)

    # Exclude requests that this donor has already donated to

    requests = FoodBankRequest.objects.filter(

        status='active'

    ).filter(

        Q(deadline__isnull=True) | Q(deadline__gte=timezone.now())

    ).exclude(

        donations__donor=request.user

    )

    

    # Apply filters only if they have values
    if search_filter:
        requests = requests.filter(
            Q(title__icontains=search_filter) |
            Q(description__icontains=search_filter) |
            Q(foodbank__foodbank_name__icontains=search_filter) |
            Q(foodbank__address__icontains=search_filter)
        )

    if location_filter:

        requests = requests.filter(

            Q(foodbank__address__icontains=location_filter) |

            Q(foodbank__foodbank_name__icontains=location_filter)

        )

    

    if priority_filter:

        requests = requests.filter(priority=priority_filter)

    if request_type_filter in ('food', 'non_food'):
        requests = requests.filter(donation_type=request_type_filter)
    else:
        request_type_filter = ''

    

    if service_type_filter:

        requests = requests.filter(foodbank__service_type=service_type_filter)

    if delivery_filter:
        if delivery_filter == 'dropoff':
            delivery_filter = 'delivery'
        if delivery_filter == 'pickup':
            requests = requests.filter(Q(delivery_method='pickup') | Q(delivery_method='both'))
        elif delivery_filter == 'delivery':
            requests = requests.filter(
                Q(delivery_method='delivery') | Q(delivery_method='dropoff') | Q(delivery_method='both')
            )

    

    # Apply date filters

    if due_date_from_filter:

        try:

            from datetime import datetime

            due_date_from = datetime.strptime(due_date_from_filter, '%Y-%m-%d').date()

            requests = requests.filter(deadline__gte=due_date_from)

        except ValueError:

            pass  # Invalid date format, ignore filter

    

    if due_date_to_filter:

        try:

            from datetime import datetime

            due_date_to = datetime.strptime(due_date_to_filter, '%Y-%m-%d').date()

            requests = requests.filter(deadline__lte=due_date_to)

        except ValueError:

            pass  # Invalid date format, ignore filter

    

    # Apply sorting

    requests = requests.select_related('foodbank')

    

    # Import Django ORM functions for priority ordering

    from django.db.models import Case, When, Value, IntegerField

    

    if sort_by_filter == 'newest':

        requests = requests.order_by('-created_at')

    elif sort_by_filter == 'oldest':

        requests = requests.order_by('created_at')

    elif sort_by_filter == 'due_date_asc':

        requests = requests.order_by('deadline', '-created_at')

    elif sort_by_filter == 'due_date_desc':

        requests = requests.order_by('-deadline', '-created_at')

    elif sort_by_filter == 'priority':

        # Custom priority ordering: urgent, high, medium, low

        requests = requests.annotate(

            priority_order=Case(

                When(priority='urgent', then=Value(1)),

                When(priority='high', then=Value(2)),

                When(priority='medium', then=Value(3)),

                When(priority='low', then=Value(4)),

                default=Value(5),

                output_field=IntegerField()

            )

        ).order_by('priority_order', '-created_at')

    else:

        # Default ordering: newest requests first

        requests = requests.order_by('-created_at')

    

    # Filter out fully-fulfilled requests (100% or more fulfilled)

    requests = [req for req in requests if req.get_fulfillment_percentage() < 100]

    def _strip_recipient_note_text(text):
        if not text:
            return ''
        cleaned = str(text)
        cleaned = re.sub(
            r'---\s*Recipient Note\s*---[\s\S]*?(?=\n\s*---\s*[^-].*?---|$)',
            '',
            cleaned,
            flags=re.IGNORECASE
        )
        cleaned = re.sub(r'^\s*Recipient Note\s*:\s*.*$', '', cleaned, flags=re.IGNORECASE | re.MULTILINE)
        cleaned = re.sub(r'^\s*---\s*Recipient Note\s*---\s*$', '', cleaned, flags=re.IGNORECASE | re.MULTILINE)
        cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
        return cleaned

    # Avoid duplicate title/description display on remaining requests.
    for req in requests:
        raw_title = (getattr(req, 'title', None) or '').strip()
        raw_description = (getattr(req, 'description', None) or '').strip()
        original_request = getattr(req, 'original_request', None)
        original_note = (getattr(original_request, 'additional_notes', None) or '').strip() if original_request else ''
        fulfilled_qty = int(getattr(original_request, 'quantity_fulfilled', 0) or 0) if original_request else 0
        is_remainder_request = bool(original_request and fulfilled_qty > 0)
        clean_title = _strip_recipient_note_text(raw_title)
        clean_description = _strip_recipient_note_text(raw_description)
        clean_original_note = _strip_recipient_note_text(original_note)
        req.display_title = clean_title or clean_description or raw_title or 'Untitled Request'

        # Remainder requests should not show the recipient note line in donor list views.
        if is_remainder_request:
            req.display_description = ''
            continue

        if clean_description and clean_title and clean_description == clean_title:
            req.display_description = clean_original_note
        elif clean_description:
            req.display_description = clean_description
        elif clean_original_note:
            req.display_description = clean_original_note
        else:
            req.display_description = ''

    

    # Pagination

    paginator = Paginator(requests, 9)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    context = {

        'requests': page_obj,

        'priority_choices': FoodBankRequest.PRIORITY_CHOICES,

        'location_filter': location_filter,
        'search_filter': search_filter,

        'priority_filter': priority_filter,
        'request_type_filter': request_type_filter,
        'delivery_filter': delivery_filter,

        'service_type_filter': service_type_filter,

        'sort_by_filter': sort_by_filter,

        'due_date_from_filter': due_date_from_filter,

        'due_date_to_filter': due_date_to_filter,

    }

    return render(request, 'donor/foodbank_requests.html', context)







@login_required

def donation_analytics(request):

    """Analytics view showing donation impact and request fulfillment data"""

    if request.user.user_type not in ['DONOR', 'FOODBANK']:

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    # General statistics

    total_donations = Donation.objects.count()

    total_donors = Donation.objects.values('donor').distinct().count()

    total_foodbanks = FoodBankProfile.objects.count()

    

    # Request fulfillment statistics

    total_requests = FoodBankRequest.objects.count()

    fulfilled_requests = FoodBankRequest.objects.filter(status='fulfilled').count()

    active_requests = FoodBankRequest.objects.filter(status='active').count()

    

    # Calculate fulfillment rate

    fulfillment_rate = (fulfilled_requests / total_requests * 100) if total_requests > 0 else 0

    

    # Recent donation activity

    recent_donations = Donation.objects.select_related('donor', 'foodbank', 'foodbank_request').order_by('-donated_at')[:10]

    

    # Top performing food banks (by donations received)

    top_foodbanks = FoodBankProfile.objects.annotate(

        donation_count=Count('donation'),

        total_impact=Sum('donation__quantity', default=0)

    ).order_by('-donation_count')[:5]

    

    # Request analytics by priority

    request_stats = FoodBankRequest.objects.values('priority').annotate(

        total=Count('id'),

        fulfilled=Count('id', filter=Q(status='fulfilled')),

        active=Count('id', filter=Q(status='active'))

    ).order_by('priority')

    

    # Donation type breakdown

    donation_types = Donation.objects.values('donation_type').annotate(

        count=Count('id'),

        total_impact=Sum('quantity', default=0)

    ).order_by('-count')

    

    # Monthly trends (last 6 months)

    from datetime import datetime, timedelta

    six_months_ago = datetime.now() - timedelta(days=180)

    monthly_trends = Donation.objects.filter(

        donated_at__gte=six_months_ago

    ).annotate(

        month=TruncMonth('donated_at')

    ).values('month').annotate(

        donations=Count('id'),

        unique_donors=Count('donor', distinct=True)

    ).order_by('month')

    

    context = {

        'total_donations': total_donations,

        'total_donors': total_donors,

        'total_foodbanks': total_foodbanks,

        'total_requests': total_requests,

        'fulfilled_requests': fulfilled_requests,

        'active_requests': active_requests,

        'fulfillment_rate': fulfillment_rate,

        'recent_donations': recent_donations,

        'top_foodbanks': top_foodbanks,

        'request_stats': request_stats,

        'donation_types': donation_types,

        'monthly_trends': monthly_trends,

    }

    

    return render(request, 'analytics/donation_analytics.html', context)





# Configure Stripe

stripe.api_key = settings.STRIPE_SECRET_KEY



@login_required

def create_payment_intent(request):

    """Create a Stripe payment intent for credit card donations"""

    if request.user.user_type != 'DONOR':

        return JsonResponse({'error': 'Only donors can make payments'}, status=403)

    

    if request.method == 'POST':

        try:

            data = json.loads(request.body)

            amount = float(data.get('amount', 0))

            donation_id = data.get('donation_id')

            

            if amount <= 0:

                return JsonResponse({'error': 'Invalid amount'}, status=400)

            

            # Convert to cents for Stripe (KES to cents)

            amount_cents = int(amount * 100)

            

            # Create payment intent

            intent = stripe.PaymentIntent.create(

                amount=amount_cents,

                currency='kes',  # Kenyan Shillings

                metadata={

                    'donation_id': donation_id,

                    'donor_email': request.user.email,

                }

            )

            

            return JsonResponse({

                'client_secret': intent.client_secret,

                'payment_intent_id': intent.id

            })

            

        except stripe.error.StripeError as e:

            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:

            return JsonResponse({'error': 'An error occurred'}, status=500)

    

    return JsonResponse({'error': 'Method not allowed'}, status=405)



@login_required

def confirm_payment(request):

    """Confirm payment and create donation record"""

    if request.user.user_type != 'DONOR':

        return JsonResponse({'error': 'Only donors can confirm payments'}, status=403)

    

    if request.method == 'POST':

        try:

            data = json.loads(request.body)

            payment_intent_id = data.get('payment_intent_id')

            foodbank_id = data.get('foodbank_id')

            donation_type = data.get('donation_type', 'money')

            message = data.get('message', '')

            

            # Retrieve payment intent from Stripe

            intent = stripe.PaymentIntent.retrieve(payment_intent_id)

            

            if intent.status != 'succeeded':

                return JsonResponse({'error': 'Payment not completed'}, status=400)

            

            # Get food bank

            try:

                foodbank = FoodBankProfile.objects.get(id=foodbank_id)

            except FoodBankProfile.DoesNotExist:

                return JsonResponse({'error': 'Food bank not found'}, status=404)

            

            # Create donation record

            amount = intent.amount / 100  # Convert from cents

            donation = Donation.objects.create(

                donor=request.user,

                donation_type=donation_type,

                foodbank=foodbank,

                amount=amount,

                message=message,

                delivery_status='pending'

            )

            

            # Create payment transaction record

            payment_transaction = PaymentTransaction.objects.create(

                donation=donation,

                stripe_payment_intent_id=payment_intent_id,

                payment_method='credit_card',

                status='completed',

                amount=amount,

                currency='KES',

                stripe_response=intent

            )

            payment_transaction.mark_completed()

            

            # Create notifications

            Notification.objects.create(

                user=request.user,

                notification_type='acknowledgement',

                message=f'Thank you for your credit card donation of KES {amount:,.2f} to {foodbank.foodbank_name}.'

            )

            

            Notification.objects.create(

                user=foodbank.user,

                notification_type='donation_received',

                message=f'Credit card donation received: KES {amount:,.2f} from {request.user.email}'

            )

            

            return JsonResponse({

                'success': True,

                'donation_id': donation.id,

                'message': 'Payment confirmed successfully!'

            })

            

        except stripe.error.StripeError as e:

            return JsonResponse({'error': f'Stripe error: {str(e)}'}, status=400)

        except Exception as e:

            return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)

    

    return JsonResponse({'error': 'Method not allowed'}, status=405)



@csrf_exempt

def stripe_webhook(request):

    """Handle Stripe webhooks for payment confirmations"""

    payload = request.body

    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')

    

    try:

        event = stripe.Webhook.construct_event(

            payload, sig_header, settings.STRIPE_WEBHOOK_SECRET

        )

    except ValueError:

        return HttpResponse(status=400)

    except stripe.error.SignatureVerificationError:

        return HttpResponse(status=400)

    

    # Handle payment intent succeeded

    if event['type'] == 'payment_intent.succeeded':

        payment_intent = event['data']['object']

        

        # Find and update payment transaction

        try:

            payment_transaction = PaymentTransaction.objects.get(

                stripe_payment_intent_id=payment_intent['id']

            )

            payment_transaction.mark_completed()

        except PaymentTransaction.DoesNotExist:

            pass

    

    return HttpResponse(status=200)



@login_required

def payment_success(request):

    """Payment success page"""

    donation_id = request.GET.get('donation_id')

    context = {

        'donation_id': donation_id,

    }

    return render(request, 'donor/payment_success.html', context)



@login_required

def payment_cancelled(request):

    """Payment cancelled page"""

    return render(request, 'donor/payment_cancelled.html')



# Recipient Settings Views

@login_required

def recipient_settings(request):

    if request.user.user_type != 'RECIPIENT':

        return redirect('dashboard')

    

    try:

        recipient_profile = request.user.recipient_profile

    except:

        messages.error(request, 'Recipient profile not found.')

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = RecipientProfileForm(request.POST, instance=recipient_profile, user=request.user)

        if form.is_valid():

            form.save()

            messages.success(request, 'Profile updated successfully!')

            return redirect('recipient_settings')

    else:

        form = RecipientProfileForm(instance=recipient_profile, user=request.user)

    

    context = {

        'form': form,

        'recipient_profile': recipient_profile

    }

    return render(request, 'recipient/recipient_settings.html', context)



@login_required

def change_recipient_password(request):

    if request.user.user_type != 'RECIPIENT':

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = RecipientPasswordChangeForm(request.user, request.POST)

        if form.is_valid():

            form.save()

            # Update session to prevent logout

            update_session_auth_hash(request, form.user)

            messages.success(request, 'Password changed successfully!')

            return redirect('recipient_settings')

    else:

        form = RecipientPasswordChangeForm(request.user)

    

    context = {

        'form': form

    }

    return render(request, 'recipient/change_password.html', context)



from django.db.models import Q

from django.core.paginator import Paginator



@login_required

def recipient_foodbanks_list(request):

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Access denied.")

        return redirect('dashboard')

    

    # Get all approved and active food banks

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True)

    total_count = foodbanks.count()

    

    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    service_type_filter = request.GET.get('service_type', '')
    sort_filter = request.GET.get('sort', 'newest')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    

    # Apply search filter

    if search_query:

        foodbanks = foodbanks.filter(

            Q(foodbank_name__icontains=search_query) |

            Q(address__icontains=search_query) |

            Q(about_text__icontains=search_query)

        )

    

    # Apply service type filter
    if service_type_filter:
        foodbanks = foodbanks.filter(service_type=service_type_filter)

    # Apply account creation date filter (application_date)
    if date_from:
        parsed_from = parse_date(date_from)
        if parsed_from:
            foodbanks = foodbanks.filter(application_date__date__gte=parsed_from)

    if date_to:
        parsed_to = parse_date(date_to)
        if parsed_to:
            foodbanks = foodbanks.filter(application_date__date__lte=parsed_to)

    

    # Apply sorting

    if sort_filter == 'oldest':

        foodbanks = foodbanks.order_by('application_date')

    else:

        foodbanks = foodbanks.order_by('-application_date')

    

    # Pagination

    paginator = Paginator(foodbanks, 9)  # 9 items per page (3x3 grid)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    context = {

        'foodbanks': page_obj,

        'search_query': search_query,

        'service_type_filter': service_type_filter,

        'sort_filter': sort_filter,
        'date_from': date_from,
        'date_to': date_to,

        'total_count': total_count,

    }

    return render(request, 'recipient/foodbanks_list.html', context)

# Simple session timeout - no additional views needed



@login_required

def admin_pending_registrations(request):

    """Admin view to manage pending foodbank registrations"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    # Get all pending foodbank registrations

    pending_registrations = FoodBankProfile.objects.filter(

        is_approved='pending'

    ).select_related('user').order_by('-application_date')

    

    # Get recently approved/rejected for reference

    recent_decisions = FoodBankProfile.objects.filter(

        is_approved__in=['approved', 'rejected']

    ).select_related('user', 'approved_by').order_by('-approval_date')[:10]

    

    context = {

        'pending_registrations': pending_registrations,

        'recent_decisions': recent_decisions,

        'pending_count': pending_registrations.count(),

    }

    return render(request, 'authentication/admin_pending_registrations.html', context)



@login_required

def approve_foodbank_registration(request, foodbank_id):

    """Approve a pending foodbank registration"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    try:

        foodbank_profile = FoodBankProfile.objects.get(id=foodbank_id, is_approved='pending')

    except FoodBankProfile.DoesNotExist:

        messages.error(request, 'Food bank registration not found or already processed.')

        return redirect('admin_pending_registrations')

    

    if request.method == 'POST':

        # Approve the registration

        foodbank_profile.is_approved = 'approved'

        foodbank_profile.approval_date = timezone.now()

        foodbank_profile.approved_by = request.user

        foodbank_profile.save()

        

        # Send approval email to foodbank

        try:

            send_approval_email(foodbank_profile.user)

        except Exception as e:

            print(f"Failed to send approval email: {e}")

        

        # Create notification for the foodbank

        Notification.objects.create(

            user=foodbank_profile.user,

            notification_type='approval',

            message=f'Congratulations! Your food bank "{foodbank_profile.foodbank_name}" has been approved and is now active on FoodBank Hub.'

        )

        

        messages.success(

            request, 

            f'Food bank "{foodbank_profile.foodbank_name}" has been approved successfully!'

        )

        return redirect('admin_pending_registrations')

    

    context = {

        'foodbank_profile': foodbank_profile,

        'action': 'approve'

    }

    return render(request, 'authentication/admin_confirm_action.html', context)



@login_required

def reject_foodbank_registration(request, foodbank_id):

    """Reject a pending foodbank registration"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    try:

        foodbank_profile = FoodBankProfile.objects.get(id=foodbank_id, is_approved='pending')

    except FoodBankProfile.DoesNotExist:

        messages.error(request, 'Food bank registration not found or already processed.')

        return redirect('admin_pending_registrations')

    

    if request.method == 'POST':

        rejection_reason = request.POST.get('rejection_reason', '').strip()

        

        if not rejection_reason:

            messages.error(request, 'Please provide a reason for rejection.')

            context = {

                'foodbank_profile': foodbank_profile,

                'action': 'reject'

            }

            return render(request, 'authentication/admin_confirm_action.html', context)

        

        # Reject the registration

        foodbank_profile.is_approved = 'rejected'

        foodbank_profile.approval_date = timezone.now()

        foodbank_profile.approved_by = request.user

        foodbank_profile.rejection_reason = rejection_reason

        foodbank_profile.save()

        

        # Send rejection email to foodbank

        try:

            send_rejection_email(foodbank_profile.user, rejection_reason)

        except Exception as e:

            print(f"Failed to send rejection email: {e}")

        

        # Create notification for the foodbank

        Notification.objects.create(

            user=foodbank_profile.user,

            notification_type='rejection',

            message=f'Your food bank application has been reviewed. Please check your email for details.'

        )

        

        messages.success(

            request, 

            f'Food bank "{foodbank_profile.foodbank_name}" application has been rejected.'

        )

        return redirect('admin_pending_registrations')

    

    context = {

        'foodbank_profile': foodbank_profile,

        'action': 'reject'

    }

    return render(request, 'authentication/admin_confirm_action.html', context)



@login_required

def view_foodbank_application(request, foodbank_id):

    """View detailed foodbank application"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    try:

        foodbank_profile = FoodBankProfile.objects.select_related('user').get(id=foodbank_id)

    except FoodBankProfile.DoesNotExist:

        messages.error(request, 'Food bank application not found.')

        return redirect('admin_pending_registrations')

    

    context = {

        'foodbank_profile': foodbank_profile,

    }

    return render(request, 'authentication/admin_view_application.html', context)



@login_required

def select_foodbank_for_request(request):

    """Step 1: Allow recipient to select a foodbank before making a request"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Only recipients can make requests.')

        return redirect('dashboard')

    

    # Get search and filter parameters

    search_query = request.GET.get('search', '')

    location_filter = request.GET.get('location', '')

    service_type_filter = request.GET.get('service_type', '')

    sort_by = request.GET.get('sort_by', 'name')

    

    # Start with all active and approved foodbanks

    foodbanks = FoodBankProfile.objects.select_related('user').filter(

        user__is_active=True,

        is_approved='approved'

    )

    

    # Apply search filter (name or location)

    if search_query:

        foodbanks = foodbanks.filter(

            Q(foodbank_name__icontains=search_query) |

            Q(address__icontains=search_query)

        )

    

    # Apply location filter

    if location_filter:

        foodbanks = foodbanks.filter(address__icontains=location_filter)

    

    # Apply service type filter

    if service_type_filter and service_type_filter != 'all':

        # Show foodbanks that provide the selected service (including those that provide both)

        foodbanks = foodbanks.filter(

            Q(service_type=service_type_filter) | Q(service_type='both')

        )

    

    # Apply sorting

    if sort_by == 'newest':

        foodbanks = foodbanks.order_by('-application_date')

    elif sort_by == 'oldest':

        foodbanks = foodbanks.order_by('application_date')

    else:  # default to name

        foodbanks = foodbanks.order_by('foodbank_name')

    

    # Get unique locations for filter dropdown (only approved food banks)

    all_locations = FoodBankProfile.objects.filter(

        user__is_active=True,

        is_approved='approved',

        address__isnull=False

    ).exclude(address='').values_list('address', flat=True).distinct()

    

    # Extract cities from addresses (assuming format: "City, State")

    cities = set()

    for address in all_locations:

        if address and ',' in address:

            city = address.split(',')[0].strip()

            if city:

                cities.add(city)

    

    cities = sorted(list(cities))

    

    context = {

        'foodbanks': foodbanks,

        'search_query': search_query,

        'location_filter': location_filter,

        'service_type_filter': service_type_filter,

        'sort_by': sort_by,

        'cities': cities,

        'service_type_choices': FoodBankProfile.SERVICE_TYPE_CHOICES,

        'total_count': foodbanks.count(),

    }

    return render(request, 'recipient/select_foodbank.html', context)



@login_required

def create_recipient_request_with_foodbank(request, foodbank_id):

    """Step 2: Create request for selected foodbank"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Only recipients can make requests.')

        return redirect('dashboard')

    

    try:

        foodbank = FoodBankProfile.objects.get(id=foodbank_id)

    except FoodBankProfile.DoesNotExist:

        messages.error(request, 'Food bank not found.')

        return redirect('select_foodbank_for_request')

    

    if request.method == 'POST':

        form = RecipientRequestForm(request.POST)

        # Remove foodbank from form validation since it's pre-selected

        if 'foodbank' in form.fields:

            del form.fields['foodbank']

        

        if form.is_valid():

            req = form.save(commit=False)

            req.recipient = request.user.recipient_profile

            req.foodbank = foodbank  # Set the selected foodbank

            req.save()

            

            # Create notification for the foodbank

            Notification.objects.create(

                user=foodbank.user,

                notification_type='request',

                message=f'New request from {req.recipient.full_name}: "{req.title}"'

            )

            

            messages.success(request, f"Request submitted successfully to {foodbank.foodbank_name}!")

            return redirect('recipient_requests_view')

        else:

            messages.error(request, 'Please correct the errors below.')

    else:

        # Create form without foodbank field since it's pre-selected

        form = RecipientRequestForm()

        if 'foodbank' in form.fields:

            del form.fields['foodbank']

    

    context = {

        'form': form,

        'foodbank': foodbank,

    }

    return render(request, 'recipient/create_request_with_foodbank.html', context)





@login_required

def support_payment(request):

    """Support payment page for system development and hosting"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can access the support payment page.')

        return redirect('dashboard')

    

    if request.method == 'POST':

        form = SystemSupportDonationForm(request.POST, request.FILES)

        if form.is_valid():

            donation = form.save(commit=False)

            donation.donor = request.user

            donation.save()

            

            # Create notification for donor

            Notification.objects.create(

                user=request.user,

                notification_type='acknowledgement',

                message=f'Thank you for your KES {donation.amount} donation to support FoodBank Hub! Your payment is being verified.'

            )

            

            # Create notification for admins

            admin_users = CustomUser.objects.filter(is_staff=True, is_active=True)

            for admin in admin_users:

                Notification.objects.create(

                    user=admin,

                    notification_type='system',

                    message=f'New system support donation of KES {donation.amount} from {request.user.email} pending verification.'

                )

            

            messages.success(

                request,

                f'Thank you for your donation of KES {donation.amount}! Your payment proof has been submitted and will be verified by our team shortly. You will be notified once it\'s approved.'

            )

            return redirect('dashboard')

        else:

            messages.error(request, 'Please correct the errors below.')

    else:

        form = SystemSupportDonationForm()

    

    # Get donor's previous donations

    previous_donations = SystemSupportDonation.objects.filter(donor=request.user).order_by('-created_at')[:5]

    

    context = {

        'user': request.user,

        'form': form,

        'previous_donations': previous_donations,

    }

    return render(request, 'authentication/support_payment.html', context)





@login_required

def my_support_donations(request):

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can access support donations.')

        return redirect('dashboard')



    status_filter = (request.GET.get('status') or '').strip()

    search_query = (request.GET.get('search') or '').strip()

    date_from = (request.GET.get('date_from') or '').strip()

    date_to = (request.GET.get('date_to') or '').strip()



    qs = SystemSupportDonation.objects.filter(donor=request.user).order_by('-created_at')



    if status_filter and status_filter in dict(SystemSupportDonation.STATUS_CHOICES):

        qs = qs.filter(status=status_filter)



    if date_from:

        from django.utils.dateparse import parse_date

        parsed_from = parse_date(date_from)

        if parsed_from:

            qs = qs.filter(created_at__date__gte=parsed_from)



    if date_to:

        from django.utils.dateparse import parse_date

        parsed_to = parse_date(date_to)

        if parsed_to:

            qs = qs.filter(created_at__date__lte=parsed_to)



    if search_query:

        qs = qs.filter(

            Q(transaction_reference__icontains=search_query) |

            Q(amount__icontains=search_query) |

            Q(notes__icontains=search_query)

        )



    paginator = Paginator(qs, 10)

    page_obj = paginator.get_page(request.GET.get('page'))



    context = {

        'title': 'My Support Donations',

        'page_obj': page_obj,

        'status_filter': status_filter,

        'search_query': search_query,

        'date_from': date_from,

        'date_to': date_to,

    }

    return render(request, 'authentication/my_support_donations.html', context)





@login_required

def export_my_support_donations_excel(request):

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can access support donations.')

        return redirect('dashboard')



    status_filter = (request.GET.get('status') or '').strip()

    search_query = (request.GET.get('search') or '').strip()

    date_from = (request.GET.get('date_from') or '').strip()

    date_to = (request.GET.get('date_to') or '').strip()



    qs = SystemSupportDonation.objects.filter(donor=request.user).select_related('verified_by').order_by('-created_at')



    if status_filter and status_filter in dict(SystemSupportDonation.STATUS_CHOICES):

        qs = qs.filter(status=status_filter)



    if date_from:

        from django.utils.dateparse import parse_date

        parsed_from = parse_date(date_from)

        if parsed_from:

            qs = qs.filter(created_at__date__gte=parsed_from)



    if date_to:

        from django.utils.dateparse import parse_date

        parsed_to = parse_date(date_to)

        if parsed_to:

            qs = qs.filter(created_at__date__lte=parsed_to)



    if search_query:

        qs = qs.filter(

            Q(transaction_reference__icontains=search_query) |

            Q(amount__icontains=search_query) |

            Q(notes__icontains=search_query)

        )



    donations_data = list(qs)



    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.utils import get_column_letter



    wb = Workbook()

    ws = wb.active

    ws.title = 'Support Donations'



    header_font = Font(bold=True, color='FFFFFF', size=10)

    header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')

    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    border = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin'),

    )



    ws.merge_cells('A1:L1')

    title_cell = ws['A1']

    title_cell.value = f"{request.user.email} - System Support Donations"

    title_cell.font = Font(bold=True, size=14, color='10b981')

    title_cell.alignment = Alignment(horizontal='center')



    ws.merge_cells('A2:L2')

    info_cell = ws['A2']

    info_cell.value = f"Generated on: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')} | Total: {len(donations_data)}"

    info_cell.alignment = Alignment(horizontal='center')



    headers = [

        'S/No',

        'Donation ID',

        'Amount (KES)',

        'Transaction Reference',

        'Status',

        'Submitted At',

        'Donor Notes',

        'Rejection Reason',

        'Verified By',

        'Verified At',

        'Updated At',

        'Proof URL',

    ]



    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=4, column=col_num)

        cell.value = header

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = header_alignment

        cell.border = border



    for idx, donation in enumerate(donations_data, 1):

        row_num = idx + 4



        proof_url = ''

        if getattr(donation, 'payment_proof', None) and getattr(donation.payment_proof, 'url', None):

            try:

                proof_url = request.build_absolute_uri(donation.payment_proof.url)

            except Exception:

                proof_url = donation.payment_proof.url



        verified_by = donation.verified_by.email if getattr(donation, 'verified_by', None) else ''

        verified_at = donation.verified_at.strftime('%Y-%m-%d %H:%M') if donation.verified_at else ''



        values = [

            idx,

            donation.id,

            float(donation.amount) if donation.amount is not None else 0,

            donation.transaction_reference or '',

            donation.get_status_display() if hasattr(donation, 'get_status_display') else donation.status,

            donation.created_at.strftime('%Y-%m-%d %H:%M') if donation.created_at else '',

            donation.notes or '',

            donation.rejection_reason or '',

            verified_by,

            verified_at,

            donation.updated_at.strftime('%Y-%m-%d %H:%M') if donation.updated_at else '',

            proof_url,

        ]



        for col_num, value in enumerate(values, 1):

            c = ws.cell(row=row_num, column=col_num, value=value)

            c.border = border

            if col_num in [7, 8, 12]:

                c.alignment = Alignment(wrap_text=True, vertical='top')



    column_widths = [6, 12, 14, 22, 14, 18, 28, 28, 20, 18, 18, 40]

    for col_num, width in enumerate(column_widths, 1):

        ws.column_dimensions[get_column_letter(col_num)].width = width



    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    filename = f"support_donations_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def export_my_support_donations_pdf(request):
    if request.user.user_type != 'DONOR':
        messages.error(request, 'Only donors can access support donations.')
        return redirect('dashboard')

    status_filter = (request.GET.get('status') or '').strip()
    search_query = (request.GET.get('search') or '').strip()
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()

    qs = SystemSupportDonation.objects.filter(donor=request.user).select_related('verified_by').order_by('-created_at')

    if status_filter and status_filter in dict(SystemSupportDonation.STATUS_CHOICES):
        qs = qs.filter(status=status_filter)

    if date_from:
        from django.utils.dateparse import parse_date
        parsed_from = parse_date(date_from)
        if parsed_from:
            qs = qs.filter(created_at__date__gte=parsed_from)

    if date_to:
        from django.utils.dateparse import parse_date
        parsed_to = parse_date(date_to)
        if parsed_to:
            qs = qs.filter(created_at__date__lte=parsed_to)

    if search_query:
        qs = qs.filter(
            Q(transaction_reference__icontains=search_query) |
            Q(amount__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    donations_data = list(qs)

    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=18,
        rightMargin=18,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SupportDonationTitle',
        parent=styles['Heading2'],
        textColor=colors.HexColor('#10b981'),
        alignment=TA_CENTER,
        fontSize=16,
        spaceAfter=6,
    )
    info_style = ParagraphStyle(
        'SupportDonationInfo',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
    )
    wrap_style = ParagraphStyle(
        'SupportDonationWrap',
        parent=styles['Normal'],
        alignment=TA_LEFT,
        fontSize=7,
        leading=9,
        wordWrap='CJK',
    )

    def _safe_text(value, max_len=180):
        text = str(value or '').strip()
        if not text:
            return '-'
        if len(text) > max_len:
            text = f"{text[:max_len - 3]}..."
        return html.escape(text)

    elements = [
        Paragraph(f"{html.escape(request.user.email)} - System Support Donations", title_style),
        Paragraph(
            f"Generated on: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')} | Total: {len(donations_data)}",
            info_style,
        ),
    ]

    table_data = [[
        'S/No',
        'Donation ID',
        'Amount (KES)',
        'Transaction Reference',
        'Status',
        'Submitted At',
        'Donor Notes',
        'Rejection Reason',
        'Verified By',
        'Verified At',
        'Updated At',
        'Proof URL',
    ]]

    for idx, donation in enumerate(donations_data, 1):
        proof_url = ''
        if getattr(donation, 'payment_proof', None) and getattr(donation.payment_proof, 'url', None):
            try:
                proof_url = request.build_absolute_uri(donation.payment_proof.url)
            except Exception:
                proof_url = donation.payment_proof.url

        verified_by = donation.verified_by.email if getattr(donation, 'verified_by', None) else '-'
        verified_at = donation.verified_at.strftime('%Y-%m-%d %H:%M') if donation.verified_at else '-'
        updated_at = donation.updated_at.strftime('%Y-%m-%d %H:%M') if donation.updated_at else '-'
        submitted_at = donation.created_at.strftime('%Y-%m-%d %H:%M') if donation.created_at else '-'
        amount_display = f"{float(donation.amount):,.2f}" if donation.amount is not None else "0.00"
        status_display = donation.get_status_display() if hasattr(donation, 'get_status_display') else (donation.status or '-')

        table_data.append([
            str(idx),
            str(donation.id),
            amount_display,
            Paragraph(_safe_text(donation.transaction_reference, 70), wrap_style),
            Paragraph(_safe_text(status_display, 40), wrap_style),
            submitted_at,
            Paragraph(_safe_text(donation.notes, 220), wrap_style),
            Paragraph(_safe_text(donation.rejection_reason, 220), wrap_style),
            Paragraph(_safe_text(verified_by, 80), wrap_style),
            verified_at,
            updated_at,
            Paragraph(_safe_text(proof_url, 240), wrap_style),
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[32, 45, 62, 96, 55, 78, 160, 160, 85, 78, 78, 189],
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"support_donations_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# Discussion System Views for 'Other' and 'CSR' Type Donations



@login_required

def available_other_donations(request):

    """View available 'other' and 'csr' type donations for foodbanks to start discussions"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Only foodbanks can view available donations.')

        return redirect('dashboard')

    

    # Get all 'other' and 'csr' type donations for this foodbank that require discussion

    other_donations = Donation.objects.filter(

        donation_type__in=['other', 'csr'],

        requires_discussion=True,

        foodbank=request.user.foodbank_profile,

        discussion_status__in=['pending', None]

    ).order_by('-donated_at')

    

    context = {

        'other_donations': other_donations,

    }

    return render(request, 'foodbank/available_other_donations.html', context)



@login_required

def start_donation_discussion(request, donation_id):

    """Start a discussion with donor about an 'other' or 'csr' type donation"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Only foodbanks can start discussions.')

        return redirect('dashboard')

    

    try:

        donation = Donation.objects.get(

            id=donation_id,

            donation_type__in=['other', 'csr'],

            requires_discussion=True,

            foodbank=request.user.foodbank_profile,

            discussion_status__in=['pending', None]

        )

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found or no longer available for discussion.')

        return redirect('available_other_donations')

    

    # Check if discussion already exists

    existing_discussion = DonationDiscussion.objects.filter(

        donation=donation,

        foodbank=request.user.foodbank_profile

    ).first()

    

    if existing_discussion:

        return redirect('donation_discussion_detail', discussion_id=existing_discussion.id)

    

    if request.method == 'POST':

        initial_message = request.POST.get('message', '').strip()

        if not initial_message:

            messages.error(request, 'Please provide an initial message.')

            return redirect('start_donation_discussion', donation_id=donation_id)

        

        # Create discussion

        discussion = DonationDiscussion.objects.create(

            donation=donation,

            donor=donation.donor,

            foodbank=request.user.foodbank_profile,

            status='in_progress'

        )

        

        # Create initial message

        DonationDiscussionMessage.objects.create(

            discussion=discussion,

            sender=request.user,

            message=initial_message

        )

        

        # Update donation status

        donation.discussion_status = 'in_progress'

        donation.save()

        

        # Notify donor

        Notification.objects.create(

            user=donation.donor,

            notification_type='new_donor',

            message=f'{request.user.foodbank_profile.foodbank_name} wants to discuss your donation: {donation.other_description[:50]}...'

        )

        

        messages.success(request, 'Discussion started! The donor has been notified.')

        return redirect('donation_discussion_detail', discussion_id=discussion.id)

    

    context = {

        'donation': donation,

    }

    return render(request, 'foodbank/start_donation_discussion.html', context)



@login_required

def donation_discussion_detail(request, discussion_id):

    """View and participate in a donation discussion"""

    try:

        discussion = DonationDiscussion.objects.get(id=discussion_id)

        

        # Check if user is part of this discussion

        if request.user not in [discussion.donor, discussion.foodbank.user]:

            messages.error(request, 'You are not authorized to view this discussion.')

            return redirect('dashboard')

            

    except DonationDiscussion.DoesNotExist:

        messages.error(request, 'Discussion not found.')

        return redirect('dashboard')

    

    # Mark messages as read for current user

    discussion.messages.filter(is_read=False).exclude(sender=request.user).update(is_read=True)

    

    if request.method == 'POST':

        action = request.POST.get('action')

        

        if action == 'send_message':

            message_text = request.POST.get('message', '').strip()

            if message_text:

                DonationDiscussionMessage.objects.create(

                    discussion=discussion,

                    sender=request.user,

                    message=message_text

                )

                

                # Notify the other party

                other_user = discussion.foodbank.user if request.user == discussion.donor else discussion.donor

                Notification.objects.create(

                    user=other_user,

                    notification_type='new_donor',

                    message=f'New message in donation discussion from {request.user.email}'

                )

        

        elif action == 'agree_terms' and request.user == discussion.foodbank.user:

            agreed_clause = request.POST.get('agreed_clause', '').strip()

            if not agreed_clause:

                messages.error(request, 'Please enter the agreed terms/clause before accepting.')

                return redirect('donation_discussion_detail', discussion_id=discussion_id)

            

            discussion.status = 'agreed'

            discussion.save()

            

            donation = discussion.donation

            donation.discussion_status = 'agreed'

            donation.status = 'accepted'  # Mark donation as accepted by foodbank

            donation.agreed_clause = agreed_clause  # Save the agreed clause

            donation.save()

            

            # Notify donor

            Notification.objects.create(

                user=discussion.donor,

                notification_type='acknowledgement',

                message=f'{discussion.foodbank.foodbank_name} has agreed to accept your donation!'

            )

        

        elif action == 'decline_terms' and request.user == discussion.foodbank.user:

            discussion.status = 'declined'

            discussion.save()

            

            donation = discussion.donation

            donation.discussion_status = 'declined'

            donation.status = 'declined'

            donation.save()

            

            # Notify donor

            Notification.objects.create(

                user=discussion.donor,

                notification_type='system',

                message=f'{discussion.foodbank.foodbank_name} has declined your donation.'

            )

        

        return redirect('donation_discussion_detail', discussion_id=discussion_id)

    

    discussion_messages = discussion.messages.all().order_by('sent_at')

    

    context = {

        'discussion': discussion,

        'messages': discussion_messages,

        'is_donor': request.user == discussion.donor,

        'is_foodbank': request.user == discussion.foodbank.user,

    }

    return render(request, 'foodbank/donation_discussion_detail.html', context)



@login_required

def my_donation_discussions(request):

    """View all discussions for current user (donor or foodbank)"""

    if request.user.user_type == 'DONOR':

        discussions = DonationDiscussion.objects.filter(donor=request.user).order_by('-updated_at')

        template = 'donor/my_discussions.html'

    elif request.user.user_type == 'FOODBANK':

        discussions = DonationDiscussion.objects.filter(foodbank=request.user.foodbank_profile).order_by('-updated_at')

        template = 'foodbank/my_discussions.html'

    else:

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    # Add unread message flag to each discussion

    for discussion in discussions:

        discussion.has_unread = discussion.messages.filter(is_read=False).exclude(sender=request.user).exists()

    

    # Calculate statistics

    agreed_count = discussions.filter(status='agreed').count()

    declined_count = discussions.filter(status='declined').count()

    in_progress_count = discussions.filter(status='in_progress').count()

    

    context = {

        'discussions': discussions,

        'agreed_count': agreed_count,

        'declined_count': declined_count,

        'in_progress_count': in_progress_count,

    }

    return render(request, template, context)



@login_required

def accepted_csr_donations(request):

    """View all accepted CSR and Other donations with agreed clauses for foodbank"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Only foodbanks can view this page.')

        return redirect('dashboard')

    

    # Get all accepted CSR and Other donations for this foodbank

    accepted_donations = Donation.objects.filter(

        foodbank=request.user.foodbank_profile,

        donation_type__in=['csr', 'other'],

        discussion_status='agreed',

        status='accepted'

    ).select_related('donor', 'donor__donor_profile').order_by('-donated_at')

    

    # Filter by donation type if specified

    donation_type_filter = request.GET.get('type', 'all')

    if donation_type_filter == 'csr':

        accepted_donations = accepted_donations.filter(donation_type='csr')

    elif donation_type_filter == 'other':

        accepted_donations = accepted_donations.filter(donation_type='other')

    

    # Calculate statistics

    total_count = accepted_donations.count()

    csr_count = accepted_donations.filter(donation_type='csr').count()

    other_count = accepted_donations.filter(donation_type='other').count()

    

    context = {

        'donations': accepted_donations,

        'total_count': total_count,

        'csr_count': csr_count,

        'other_count': other_count,

        'current_filter': donation_type_filter,

    }

    return render(request, 'foodbank/accepted_csr_donations.html', context)



@csrf_exempt

def mpesa_callback(request):

    """Handle M-Pesa payment callbacks"""

    if request.method == 'POST':

        try:

            import json

            callback_data = json.loads(request.body)

            

            # Extract callback data

            stk_callback = callback_data.get('Body', {}).get('stkCallback', {})

            checkout_request_id = stk_callback.get('CheckoutRequestID')

            result_code = stk_callback.get('ResultCode')

            result_desc = stk_callback.get('ResultDesc')

            

            if checkout_request_id:

                try:

                    # Find the payment transaction

                    payment = PaymentTransaction.objects.get(

                        mpesa_checkout_request_id=checkout_request_id

                    )

                    

                    if result_code == 0:  # Success

                        # Extract M-Pesa receipt number from callback items

                        callback_metadata = stk_callback.get('CallbackMetadata', {})

                        items = callback_metadata.get('Item', [])

                        

                        mpesa_receipt = None

                        for item in items:

                            if item.get('Name') == 'MpesaReceiptNumber':

                                mpesa_receipt = item.get('Value')

                                break

                        

                        # Update payment transaction

                        payment.status = 'completed'

                        payment.mpesa_receipt_number = mpesa_receipt

                        payment.mark_completed()

                        

                        # Create success notification

                        Notification.objects.create(

                            user=payment.donation.donor,

                            notification_type='acknowledgement',

                            message=f'M-Pesa payment successful! Receipt: {mpesa_receipt}. Thank you for your donation of KES {payment.amount}.'

                        )

                        

                    else:  # Failed

                        payment.status = 'failed'

                        payment.mark_failed()

                        

                        # Create failure notification

                        Notification.objects.create(

                            user=payment.donation.donor,

                            notification_type='system',

                            message=f'M-Pesa payment failed: {result_desc}. Please try again or contact support.'

                        )

                

                except PaymentTransaction.DoesNotExist:

                    pass  # Transaction not found

            

            return HttpResponse('OK')

            

        except Exception as e:

            print(f"M-Pesa callback error: {e}")

            return HttpResponse('ERROR', status=500)

    

    return HttpResponse('Method not allowed', status=405)



@login_required

def donor_contact_support(request):

    """Contact support page for donors"""

    if request.user.user_type != 'DONOR':

        return redirect('dashboard')

    

    if request.method == 'POST':

        subject = request.POST.get('subject')

        message_text = request.POST.get('message')

        

        if subject and message_text and len(message_text.strip()) >= SUPPORT_MESSAGE_MIN_LENGTH:

            # Create support message in database

            from .models import SupportMessage

            SupportMessage.objects.create(

                user=request.user,

                subject=subject,

                message=message_text.strip(),

            )

            messages.success(request, 'Your message has been sent to our support team. We will get back to you within 24-48 hours!')

            return redirect('donor_contact_support')

        else:

            messages.error(
                request,
                f'Please fill in all fields and ensure your message is at least {SUPPORT_MESSAGE_MIN_LENGTH} characters long.'
            )

    

    context = {

        'donor_profile': request.user.donor_profile

    }

    return render(request, 'donor/contact_support.html', context)



@login_required

def foodbank_contact_support(request):

    """Contact support page for food banks"""

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    if request.method == 'POST':

        subject = request.POST.get('subject')

        message_text = request.POST.get('message')

        

        if subject and message_text and len(message_text.strip()) >= SUPPORT_MESSAGE_MIN_LENGTH:

            # Create support message in database

            from .models import SupportMessage

            SupportMessage.objects.create(

                user=request.user,

                subject=subject,

                message=message_text.strip(),

            )

            messages.success(request, 'Your message has been sent to our support team. We will get back to you within 12-24 hours!')

            return redirect('foodbank_contact_support')

        else:

            messages.error(
                request,
                f'Please fill in all fields and ensure your message is at least {SUPPORT_MESSAGE_MIN_LENGTH} characters long.'
            )

    

    context = {

        'foodbank_profile': request.user.foodbank_profile

    }

    return render(request, 'foodbank/contact_support.html', context)



@login_required

def fulfill_recipient_request(request, pk):

    """Fulfill an approved recipient request by allocating donations or creating donor request"""

    if request.user.user_type != 'FOODBANK':

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    recipient_request = get_object_or_404(

        RecipientRequest, 

        pk=pk, 

        foodbank=foodbank_profile,

        status="accepted"

    )

    

    if request.method == 'POST':

        fulfillment_type = request.POST.get('fulfillment_type')

        

        if fulfillment_type == 'allocate_donation':

            # Allocate from an existing donation

            donation_id = request.POST.get('donation_id')

            quantity = request.POST.get('quantity')

            notes = request.POST.get('notes', '')

            

            try:

                donation = Donation.objects.get(id=donation_id, foodbank=foodbank_profile)

                

                # Create allocation

                allocation = DonationAllocation.objects.create(

                    donation=donation,

                    recipient=recipient_request.recipient,

                    quantity=quantity if donation.donation_type == 'item' else None,

                    amount=quantity if donation.donation_type in ['money', 'subsidized'] else None,

                    notes=f"Allocated for request: {recipient_request.title}. {notes}"

                )

                

                # Mark request as completed

                recipient_request.status = 'completed'

                recipient_request.fulfillment_notes = f"Fulfilled using donation #{donation.id}"

                recipient_request.save()

                

                # Notify recipient

                Notification.objects.create(

                    user=recipient_request.recipient.user,

                    notification_type='request_fulfilled',

                    message=f'Your request "{recipient_request.title}" has been fulfilled by {foodbank_profile.foodbank_name}. Please acknowledge receipt when you receive the donation.'

                )

                

                messages.success(request, 'Request fulfilled successfully! The recipient has been notified.')

                return redirect('dashboard')

                

            except Donation.DoesNotExist:

                messages.error(request, 'Selected donation not found.')

        

        elif fulfillment_type == 'create_donor_request':

            # Create a donor request for items needed

            title = request.POST.get('donor_request_title')

            description = request.POST.get('donor_request_description')

            quantity = request.POST.get('donor_request_quantity')

            quantity_unit = request.POST.get('donor_request_quantity_unit')

            priority = request.POST.get('priority', 'high')

            

            # Create FoodBank Request to donors

            donor_request = FoodBankRequest.objects.create(

                foodbank=foodbank_profile,

                title=title,

                description=f"{description}\n\n[Created to fulfill recipient request: {recipient_request.title}]",

                quantity_needed=quantity,

                quantity_unit=quantity_unit,

                priority=priority,

                status='active',

                donation_type='food',

                deadline=timezone.now() + timedelta(days=30),

                linked_recipient_request=recipient_request

            )

            

            # Link the recipient request to this donor request (optional - you may need to add this field)

            recipient_request.fulfillment_notes = f"Waiting for donor request #{donor_request.id} to be fulfilled"

            recipient_request.status = 'in_progress'  

            recipient_request.linked_donor_request = donor_request

            recipient_request.save()

            

            # Notify donors about urgent request

            if priority in ['urgent', 'high']:

                # Create notifications for active donors

                from django.contrib.auth import get_user_model

                User = get_user_model()

                active_donors = User.objects.filter(user_type='DONOR', is_active=True)

                

                for donor in active_donors[:50]:  # Limit to prevent spam

                    Notification.objects.create(

                        user=donor,

                        notification_type='urgent_request',

                        message=f'Urgent: {foodbank_profile.foodbank_name} needs {title}. Help fulfill this request!'

                    )

            

            messages.success(request, f'Donor request created successfully! Your request is now visible to donors.')

            return redirect('dashboard')

    

    # Get available donations for allocation

    available_donations = Donation.objects.filter(

        foodbank=foodbank_profile,

        donated_at__gte=timezone.now() - timedelta(days=30)  # Recent donations

    ).order_by('-donated_at')

    

    context = {

        'recipient_request': recipient_request,

        'available_donations': available_donations,

        'foodbank_profile': foodbank_profile,

        'QUANTITY_UNITS': QUANTITY_UNITS,

    }

    return render(request, 'authentication/fulfill_recipient_request.html', context)



@login_required

def view_in_progress_requests(request):

    foodbank_profile = get_object_or_404(FoodBankProfile, user=request.user)

    in_progress_requests = RecipientRequest.objects.filter(foodbank=foodbank_profile, status='in_progress')

    return render(request, 'authentication/view_in_progress_requests.html', {'in_progress_requests': in_progress_requests})



from django.contrib import messages

from django.contrib.auth.decorators import login_required

from django.db import transaction

from django.shortcuts import get_object_or_404, redirect

from django.core.exceptions import FieldDoesNotExist



@login_required

def mark_received(request, recipient_request_id):

    # Only accept POST

    if request.method != "POST":

        return redirect('view_in_progress_requests')



    recipient_request = get_object_or_404(RecipientRequest, id=recipient_request_id)



    # Authorization: ensure the current user belongs to the foodbank handling this request

    foodbank_profile = getattr(request.user, 'foodbankprofile', None)

    if not foodbank_profile or recipient_request.foodbank != foodbank_profile:

        messages.error(request, "You are not authorized to mark this request as received.")

        return redirect('view_in_progress_requests')



    donor_request = recipient_request.linked_donor_request

    if not donor_request:

        messages.error(request, "No linked donor request found for this recipient request.")

        return redirect('view_in_progress_requests')



    # Try to find Donation(s) linked to the donor_request.

    # Different projects call this FK differently â€” try a few common names.

    donation_qs = Donation.objects.none()

    link_field_candidates = ['linked_foodbank_request', 'linked_request', 'foodbank_request', 'donor_request']

    for field in link_field_candidates:

        try:

            # This will raise FieldError if field not present â€” catch by broad except below

            qs = Donation.objects.filter(**{field: donor_request})

            if qs.exists():

                donation_qs = qs

                break

        except Exception:

            # ignore and try next candidate

            donation_qs = Donation.objects.none()



    # fallback: try to find donations for this foodbank made around the donor_request time

    if not donation_qs.exists():

        donation_qs = Donation.objects.filter(

            foodbank=foodbank_profile

        ).order_by('-donated_at')[:5]  # handful to inspect/choose from



    try:

        with transaction.atomic():

            if donation_qs.exists():

                # Prefer the most recent relevant donation

                donation = donation_qs.first()



                # Assign the donation to the recipient and mark delivered/allocated

                donation.recipient = recipient_request.recipient



                # Update donation status field name variations (use what you have)

                if hasattr(donation, 'status'):

                    donation.status = 'delivered'  # or 'allocated' depending on your status vocabulary



                donation.save()



                messages.success(request, "Existing donation allocated to recipient and marked as delivered.")

            else:

                # No donation found: create an allocation-like Donation record.

                # Ensure we don't violate the Donation.donor FK constraint.

                donor_field = Donation._meta.get_field('donor')

                if donor_field.null:

                    donor_value = None

                else:

                    # If donor cannot be null, assign the current user (foodbank user) as a fallback.

                    # Alternative: change model to allow null for donor.

                    donor_value = request.user



                donation = Donation.objects.create(

                    donor=donor_value,

                    foodbank=foodbank_profile,

                    recipient=recipient_request.recipient,

                    # choose quantity field depending on your model

                    **({

                        'quantity': recipient_request.quantity

                    } if hasattr(Donation, 'quantity') else {}),

                    **({

                        'quantity_unit': recipient_request.quantity_unit

                    } if hasattr(Donation, 'quantity_unit') else {}),

                    donation_type=getattr(donor_request, 'donation_type', 'item'),

                    status='delivered',

                )



                messages.success(request, "Allocation record created and assigned to the recipient.")

            

            # Mark donor request fulfilled and recipient request completed

            donor_request.status = 'fulfilled'

            donor_request.save()



            recipient_request.status = 'completed'

            recipient_request.fulfillment_notes = (recipient_request.fulfillment_notes or '') + f"\nReceived via donor request #{donor_request.id}"

            recipient_request.save()



    except Exception as e:

        messages.error(request, f"Could not mark received: {e}")

        return redirect('view_in_progress_requests')



# Support Messages Views for Recipients

@login_required

def my_support_messages(request):

    """View for users to see their support messages and admin responses"""

    from .models import SupportMessage

    from django.db.models import Q

    from django.utils.dateparse import parse_date

    

    # Get all support messages for the current user

    support_messages = SupportMessage.objects.filter(

        user=request.user

    ).order_by('-created_at')



    status_filter = request.GET.get('status', '').strip()

    subject_filter = request.GET.get('subject', '').strip()

    search_query = request.GET.get('search', '').strip()

    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()



    if status_filter and status_filter in dict(SupportMessage.STATUS_CHOICES):

        support_messages = support_messages.filter(status=status_filter)

    if subject_filter and subject_filter in dict(SupportMessage.SUBJECT_CHOICES):

        support_messages = support_messages.filter(subject=subject_filter)





    if date_from:

        parsed_from = parse_date(date_from)

        if parsed_from:

            support_messages = support_messages.filter(created_at__date__gte=parsed_from)



    if date_to:

        parsed_to = parse_date(date_to)

        if parsed_to:

            support_messages = support_messages.filter(created_at__date__lte=parsed_to)



    if search_query:
        normalized_search = search_query.lower().replace('_', ' ').strip()
        matched_subject_values = [
            value
            for value, label in SupportMessage.SUBJECT_CHOICES
            if normalized_search in value.lower().replace('_', ' ')
            or normalized_search in label.lower()
        ]

        search_q = (
            Q(subject__icontains=search_query)
            | Q(message__icontains=search_query)
            | Q(admin_response__icontains=search_query)
        )
        if matched_subject_values:
            search_q = search_q | Q(subject__in=matched_subject_values)

        if search_query.isdigit():

            search_q = search_q | Q(id=int(search_query))

        support_messages = support_messages.filter(search_q)

    

    # Count messages with responses that user might not have seen

    messages_with_responses = support_messages.filter(

        Q(admin_response__isnull=False) & ~Q(admin_response='')

    ).distinct().count()

    

    # Count resolved and pending messages for foodbank users

    resolved_count = support_messages.filter(status__in=['resolved', 'closed']).count()

    pending_count = support_messages.exclude(status__in=['resolved', 'closed']).count()

    

    context = {

        'support_messages': support_messages,

        'messages_with_responses': messages_with_responses,

        'resolved_count': resolved_count,

        'pending_count': pending_count,

        'status_choices': SupportMessage.STATUS_CHOICES,

        'subject_choices': SupportMessage.SUBJECT_CHOICES,

        'status_filter': status_filter,

        'subject_filter': subject_filter,

        'search_query': search_query,

        'date_from': date_from,

        'date_to': date_to,

    }

    

    return render(request, 'authentication/my_support_messages.html', context)



@login_required

def support_message_detail_user(request, message_id):

    """View for users to see details of a specific support message"""

    from .models import SupportMessage, SupportMessageReply

    

    support_message = get_object_or_404(

        SupportMessage,

        id=message_id,

        user=request.user  # Ensure user can only see their own messages

    )



    if request.method == 'POST':

        action = request.POST.get('action')



        if action == 'add_reply':

            if support_message.status in ['resolved', 'closed']:

                messages.error(request, 'This support message is already resolved. Please create a new support message if you still need help.')

                return redirect('support_message_detail_user', message_id=message_id)



            reply_message = request.POST.get('reply_message')

            if reply_message and reply_message.strip():

                SupportMessageReply.objects.create(

                    support_message=support_message,

                    author=request.user,

                    message=reply_message.strip(),

                    is_from_admin=False,

                )

                support_message.status = 'new'

                support_message.resolved_at = None

                support_message.save()

                messages.success(request, 'Your reply has been sent.')

                return redirect('support_message_detail_user', message_id=message_id)



    replies = support_message.replies.select_related('author').all()



    context = {

        'support_message': support_message,

        'replies': replies,

    }

    

    return render(request, 'authentication/support_message_detail_user.html', context)



# Foodbank Testimonial Views

@login_required

def create_foodbank_testimonial(request):

    """Foodbanks can create testimonials with impact photos"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Only food banks can submit testimonials.')

        return redirect('dashboard')

    

    if request.method == 'POST':

        from .forms import FoodbankTestimonialForm

        from .models import FoodbankTestimonial

        form = FoodbankTestimonialForm(request.POST, request.FILES)

        if form.is_valid():

            testimonial = form.save(commit=False)

            testimonial.foodbank = request.user.foodbank_profile

            testimonial.save()

            messages.success(

                request, 

                "Testimonial submitted successfully! It will be reviewed by an admin before being displayed publicly."

            )

            return redirect('foodbank_testimonials_list')

    else:

        from .forms import FoodbankTestimonialForm

        form = FoodbankTestimonialForm()

    return render(request, 'authentication/create_foodbank_testimonial.html', {'form': form})



@login_required

def foodbank_testimonials_list(request):

    """List all testimonials for the logged-in foodbank"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial



    pending_testimonials = FoodbankTestimonial.objects.filter(

        foodbank=request.user.foodbank_profile,

        approval_status='pending'

    ).order_by('-created_at')



    approved_qs = FoodbankTestimonial.objects.filter(

        foodbank=request.user.foodbank_profile,

        approval_status='approved'

    ).order_by('-created_at')



    displayed_testimonials = [t for t in approved_qs if t.is_currently_displayed()]

    archived_approved = [t for t in approved_qs if not t.is_currently_displayed()]



    rejected_qs = FoodbankTestimonial.objects.filter(

        foodbank=request.user.foodbank_profile,

        approval_status='rejected'

    ).order_by('-created_at')



    archived_testimonials = archived_approved

    rejected_testimonials = rejected_qs

    

    context = {

        'pending_testimonials': pending_testimonials,

        'displayed_testimonials': displayed_testimonials,

        'archived_testimonials': archived_testimonials,

        'rejected_testimonials': rejected_testimonials,

    }

    return render(request, 'authentication/foodbank_testimonials_list.html', context)



@login_required

def edit_foodbank_testimonial(request, testimonial_id):

    """Edit a foodbank testimonial"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    testimonial = get_object_or_404(

        FoodbankTestimonial,

        id=testimonial_id,

        foodbank=request.user.foodbank_profile

    )

    

    # Only allow editing if pending or rejected

    if testimonial.approval_status == 'approved':

        messages.error(request, 'Cannot edit an approved testimonial.')

        return redirect('foodbank_testimonials_list')

    

    if request.method == 'POST':

        from .forms import FoodbankTestimonialForm

        form = FoodbankTestimonialForm(request.POST, request.FILES, instance=testimonial)

        if form.is_valid():

            testimonial = form.save(commit=False)

            testimonial.approval_status = 'pending'  # Reset to pending after edit

            testimonial.save()

            messages.success(request, 'Testimonial updated and resubmitted for review.')

            return redirect('foodbank_testimonials_list')

    else:

        from .forms import FoodbankTestimonialForm

        form = FoodbankTestimonialForm(instance=testimonial)

    

    context = {'form': form, 'testimonial': testimonial}

    return render(request, 'authentication/create_foodbank_testimonial.html', context)



@login_required

def delete_foodbank_testimonial(request, testimonial_id):

    """Delete a foodbank testimonial"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    testimonial = get_object_or_404(

        FoodbankTestimonial,

        id=testimonial_id,

        foodbank=request.user.foodbank_profile

    )

    

    if request.method == 'POST':

        testimonial.delete()

        messages.success(request, 'Testimonial deleted successfully.')

        return redirect('foodbank_testimonials_list')

    

    context = {'testimonial': testimonial}

    return render(request, 'authentication/delete_foodbank_testimonial.html', context)



@login_required

def toggle_foodbank_testimonial_display(request, testimonial_id):

    """Toggle public display of approved foodbank testimonial"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    testimonial = get_object_or_404(

        FoodbankTestimonial,

        id=testimonial_id,

        foodbank=request.user.foodbank_profile

    )

    

    if testimonial.approval_status != 'approved':

        messages.error(request, 'Only approved testimonials can be toggled.')

        return redirect('foodbank_testimonials_list')

    

    testimonial.display_on_public = not testimonial.display_on_public

    testimonial.save()

    

    status = "enabled" if testimonial.display_on_public else "disabled"

    messages.success(request, f'Public display {status} for this testimonial.')

    return redirect('foodbank_testimonials_list')



# Admin foodbank testimonial management views

@login_required

def admin_foodbank_testimonials_pending(request):

    """Admin view to see all pending foodbank testimonials"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    pending_testimonials = FoodbankTestimonial.objects.filter(

        approval_status='pending'

    ).select_related('foodbank__user').order_by('-created_at')

    

    context = {

        'title': 'Pending Foodbank Testimonials',

        'pending_testimonials': pending_testimonials,

        'pending_count': pending_testimonials.count(),

    }

    return render(request, 'authentication/admin_foodbank_testimonials_pending.html', context)



@login_required

def admin_approve_foodbank_testimonial(request, testimonial_id):

    """Admin approves a foodbank testimonial"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    testimonial = get_object_or_404(FoodbankTestimonial, id=testimonial_id)

    

    if request.method == 'POST':

        testimonial.approval_status = 'approved'

        testimonial.reviewed_by = request.user

        testimonial.reviewed_at = timezone.now()

        testimonial.set_default_display_period()

        testimonial.save()

        

        messages.success(request, f'Foodbank testimonial approved and will be displayed for 1 week.')

        return redirect('admin_foodbank_testimonials_pending')

    

    context = {

        'title': 'Approve Foodbank Testimonial',

        'testimonial': testimonial

    }

    return render(request, 'authentication/admin_approve_foodbank_testimonial.html', context)



@login_required

def admin_reject_foodbank_testimonial(request, testimonial_id):

    """Admin rejects a foodbank testimonial"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    testimonial = get_object_or_404(FoodbankTestimonial, id=testimonial_id)

    

    if request.method == 'POST':

        rejection_reason = request.POST.get('rejection_reason', '')

        testimonial.approval_status = 'rejected'

        testimonial.reviewed_by = request.user

        testimonial.reviewed_at = timezone.now()

        testimonial.rejection_reason = rejection_reason

        testimonial.save()

        

        messages.success(request, 'Foodbank testimonial rejected.')

        return redirect('admin_foodbank_testimonials_pending')

    

    context = {

        'title': 'Reject Foodbank Testimonial',

        'testimonial': testimonial

    }

    return render(request, 'authentication/admin_reject_foodbank_testimonial.html', context)



@login_required

def admin_all_foodbank_testimonials(request):

    """Admin view to see all foodbank testimonials with filters"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import FoodbankTestimonial

    status_filter = request.GET.get('status', 'all')

    

    testimonials = FoodbankTestimonial.objects.select_related('foodbank__user', 'reviewed_by')

    

    if status_filter != 'all':

        testimonials = testimonials.filter(approval_status=status_filter)

    

    testimonials = testimonials.order_by('-created_at')

    

    context = {

        'title': 'All Foodbank Testimonials',

        'testimonials': testimonials,

        'status_filter': status_filter,

        'total_count': testimonials.count(),

    }

    return render(request, 'authentication/admin_all_foodbank_testimonials.html', context)





# ==================== REQUEST MANAGEMENT VIEWS ====================



def _get_request_fulfillment_breakdown(req):
    """Build a list of accepted donation lines for the Quantity cell (thread of each fulfillment).
    Includes donations with status accepted, fulfilled, or partial so the breakdown still shows
    after the recipient has acknowledged (donation.status is set to fulfilled/partial then)."""
    breakdown = []
    seen_ids = set()
    unit_display = req.get_unit_display() if hasattr(req, 'get_unit_display') else (getattr(req, 'unit', '') or '')
    fb_requests = list(req.foodbank_request_created.all()) + list(req.donor_requests.all())
    for fb_req in fb_requests:
        for d in fb_req.donations.filter(status__in=('accepted', 'fulfilled', 'partial')):
            if d.id in seen_ids:
                continue
            seen_ids.add(d.id)
            if d.donation_type == 'item':
                qty = (d.quantity or 0)
                u = (d.quantity_unit or unit_display)
                breakdown.append({'type': 'item', 'label': 'Free', 'qty': qty, 'unit': u, 'amount': None})
            elif d.donation_type == 'money':
                qty = (d.foodbank_request.quantity_needed if d.foodbank_request else None)
                breakdown.append({'type': 'money', 'label': 'Monetary', 'qty': qty, 'unit': unit_display, 'amount': d.amount})
            elif d.donation_type == 'subsidized':
                qty = (d.subsidized_quantity or (d.foodbank_request.quantity_needed if d.foodbank_request else None))
                u = (d.subsidized_quantity_unit or unit_display)
                breakdown.append({'type': 'subsidized', 'label': 'Subsidized', 'qty': qty, 'unit': u, 'amount': d.subsidized_price})
    return breakdown


def _iter_request_linked_donations(req):
    """Collect unique donations linked to a request across all supported relation paths."""
    donations = []
    seen_ids = set()

    def add_donation_obj(donation):
        if not donation:
            return
        donation_id = getattr(donation, 'id', None)
        if donation_id and donation_id in seen_ids:
            return
        if donation_id:
            seen_ids.add(donation_id)
        donations.append(donation)

    # Direct FK on RequestManagement
    add_donation_obj(getattr(req, 'donation', None))

    # FoodBankRequest FK on RequestManagement
    fb_request = getattr(req, 'foodbank_request', None)
    if fb_request:
        for donation in fb_request.donations.all():
            add_donation_obj(donation)

    # Reverse relations from request -> FoodBankRequest(s)
    related_requests = list(req.foodbank_request_created.all()) + list(req.donor_requests.all())
    for linked_request in related_requests:
        for donation in linked_request.donations.all():
            add_donation_obj(donation)

    # Direct reverse relation from Donation.request_management
    for donation in req.donations.all():
        add_donation_obj(donation)

    # Donations that reached this request via allocations
    for allocation in req.donation_allocations.all():
        add_donation_obj(getattr(allocation, 'donation', None))

    return donations


def _get_request_primary_donation(req):
    """Pick the most recent linked donation for request type/category display."""
    donations = _iter_request_linked_donations(req)
    if not donations:
        return None

    donations.sort(
        key=lambda donation: (
            1 if getattr(donation, 'donated_at', None) else 0,
            getattr(donation, 'donated_at', None) or datetime.min,
            getattr(donation, 'id', 0),
        ),
        reverse=True,
    )
    return donations[0]


def _build_request_quantity_timeline(req):
    """Return chronological quantity entries (donations + allocations) for display."""
    entries = []
    seen_donation_ids = set()
    unit_display = req.get_unit_display() if hasattr(req, 'get_unit_display') else (getattr(req, 'unit', '') or '')
    # Only show donation entries that were explicitly submitted for this request.
    # Donations linked only through allocations represent stock sources and should
    # appear via allocation lines (used quantity/amount), not as full donation rows.
    request_donation_ids = set()

    def add_request_donation_id(donation):
        if not donation:
            return
        donation_id = getattr(donation, 'id', None)
        if donation_id:
            request_donation_ids.add(donation_id)

    add_request_donation_id(getattr(req, 'donation', None))

    fb_request = getattr(req, 'foodbank_request', None)
    if fb_request:
        for donation in fb_request.donations.all():
            add_request_donation_id(donation)

    related_requests = list(req.foodbank_request_created.all()) + list(req.donor_requests.all())
    for linked_request in related_requests:
        for donation in linked_request.donations.all():
            add_request_donation_id(donation)

    for donation in req.donations.all():
        add_request_donation_id(donation)

    def format_money(value):
        if value is None:
            return "KES 0"
        try:
            return f"KES {value:,.0f}"
        except (TypeError, ValueError):
            return f"KES {value}"

    def format_quantity(qty, unit):
        if qty is None:
            return unit or ''
        label = unit or unit_display
        return f"{qty} {label}" if label else str(qty)

    def add_entry(entry):
        timestamp = entry.get('timestamp')
        if not timestamp:
            timestamp = req.time_of_request or timezone.now()
        entry['timestamp'] = timestamp
        entries.append(entry)

    def add_donation(donation):
        if donation.id in seen_donation_ids:
            return
        seen_donation_ids.add(donation.id)

        display = ''
        css_class = 'text-success fw-bold'
        label = None
        accepted_by_recipient = bool(getattr(donation, 'accepted_by_recipient_id', None))
        # Treat a declined donation with a decline_message as foodbank-declined
        # even if a stale recipient flag exists from older buggy data.
        foodbank_declined = bool(
            donation.status == 'declined' and (
                not getattr(donation, 'declined_by_recipient_id', None)
                or (getattr(donation, 'decline_message', '') or '').strip()
            )
        )

        if donation.donation_type == 'item':
            display = format_quantity(donation.quantity or 0, donation.quantity_unit or unit_display)
        elif donation.donation_type == 'money':
            display = format_money(donation.amount or 0)
            qty_context = None
            if getattr(donation, 'foodbank_request', None):
                qty_context = donation.foodbank_request.quantity_needed
            qty_context = qty_context or getattr(req, 'get_donation_quantity_context', lambda: None)()
            if qty_context:
                display += f" (for {qty_context} {unit_display})"
        elif donation.donation_type == 'subsidized':
            display = format_money(donation.subsidized_price or donation.amount or 0)
            qty = donation.subsidized_quantity
            if not qty and getattr(donation, 'foodbank_request', None):
                qty = donation.foodbank_request.quantity_needed
            display += f" (for {format_quantity(qty, donation.subsidized_quantity_unit or unit_display)})" if qty else ''
            css_class = 'text-warning fw-bold' if donation.status == 'pending' else css_class
        else:
            display = donation.get_donation_type_display()

        if donation.status == 'pending':
            css_class = 'text-info fw-bold'
            label = 'Donated (Awaiting Approval)'

        is_declined = bool(getattr(donation, 'declined_by_recipient_id', None) or donation.status == 'declined')
        if is_declined:
            entry_state = 'declined'
        elif donation.status == 'pending':
            entry_state = 'pending'
        else:
            entry_state = 'accepted'

        add_entry({
            'display': display,
            'label': label,
            'css_class': css_class,
            'is_declined': is_declined,
            'timestamp': getattr(donation, 'donated_at', None),
            'state': entry_state,
            'accepted_by_recipient': accepted_by_recipient,
            'hidden_from_recipient': foodbank_declined
        })

    def add_allocation(allocation):
        display = ''
        css_class = 'text-success fw-bold'
        entry_state = 'declined' if allocation.declined_by_recipient else 'accepted'
        allocation_acknowledged = bool(getattr(allocation, 'is_acknowledged', False)) and not allocation.declined_by_recipient
        donation_unit = getattr(allocation.donation, 'quantity_unit', '') if allocation.donation else ''
        qty_unit = donation_unit or unit_display
        has_qty = allocation.quantity is not None
        has_amt = allocation.amount is not None
        if has_qty and has_amt:
            # One line per allocation: "2 pieces · KES 200 (for 2 pieces)" (whole line strike-through if declined)
            qty_part = format_quantity(allocation.quantity, qty_unit)
            amt_part = format_money(allocation.amount)
            display = f"{qty_part} · {amt_part} (for {format_quantity(allocation.quantity, qty_unit)})"
        elif has_qty:
            display = format_quantity(allocation.quantity, qty_unit)
        elif has_amt:
            display = format_money(allocation.amount)

        add_entry({
            'display': display,
            'label': None,
            'css_class': css_class,
            'is_declined': allocation.declined_by_recipient,
            'timestamp': allocation.allocated_at,
            'state': entry_state,
            'accepted_by_recipient': allocation_acknowledged
        })

    for donation in _iter_request_linked_donations(req):
        donation_id = getattr(donation, 'id', None)
        if donation_id and donation_id not in request_donation_ids:
            continue
        add_donation(donation)

    for allocation in req.donation_allocations.all():
        add_allocation(allocation)

    entries.sort(key=lambda entry: entry['timestamp'])
    return entries


def _build_recipient_timeline(entries):
    """Filter timeline entries to those recipients should see."""
    visible_entries = []
    for entry in entries:
        if entry.get('hidden_from_recipient'):
            continue

        normalized_entry = dict(entry)
        # Recipient view icons should reflect recipient action:
        # - declined => X with strike-through
        # - accepted/acknowledged => check
        # - otherwise => pending dot
        if normalized_entry.get('is_declined') or normalized_entry.get('state') == 'declined':
            normalized_entry['state'] = 'declined'
        elif normalized_entry.get('accepted_by_recipient'):
            normalized_entry['state'] = 'accepted'
        else:
            normalized_entry['state'] = 'pending'

        visible_entries.append(normalized_entry)

    has_declines = any(entry.get('is_declined') for entry in visible_entries)
    return visible_entries, has_declines


def _build_recipient_type_filter(type_filter):
    """Return a Q() expression for filtering requests by donation type."""
    if type_filter == 'all':
        return None

    sources = ['foodbank_request_created', 'donor_requests']

    def _mode_clause(field, value):
        return Q(**{field: value})

    combined_q = None

    if type_filter == 'free_goods':
        for src in sources:
            type_field = f"{src}__donations__donation_type"
            mode_field = f"{src}__donations__donation_mode"
            clause = Q(**{type_field: 'item'}) & _mode_clause(mode_field, 'free')
            combined_q = clause if combined_q is None else (combined_q | clause)
        return combined_q
    if type_filter == 'subsidized':
        for src in sources:
            type_field = f"{src}__donations__donation_type"
            mode_field = f"{src}__donations__donation_mode"
            clause = (
                Q(**{type_field: 'subsidized'}) |
                Q(**{mode_field: 'subsidized'}) |
                (
                    Q(**{type_field: 'item'}) &
                    Q(**{mode_field: 'subsidized'})
                )
            )
            combined_q = clause if combined_q is None else (combined_q | clause)
        return combined_q
    if type_filter == 'money':
        for src in sources:
            clause = Q(**{f"{src}__donations__donation_type": 'money'})
            combined_q = clause if combined_q is None else (combined_q | clause)
        return combined_q
    if type_filter == 'csr':
        for src in sources:
            clause = (
                Q(**{f"{src}__donations__donation_type": 'csr'}) |
                Q(**{f"{src}__donations__donation_mode": 'csr'})
            )
            combined_q = clause if combined_q is None else (combined_q | clause)
        return combined_q
    if type_filter == 'other':
        for src in sources:
            clause = Q(**{f"{src}__donations__donation_type": 'other'})
            combined_q = clause if combined_q is None else (combined_q | clause)
        return combined_q
    return combined_q


def _build_recipient_amount_filter(amount_filter):
    """Return Q() expression for amount-band filtering on recipient requests."""
    if amount_filter == 'all':
        return None

    sources = [
        'donation',
        'donations',
        'foodbank_request_created__donations',
        'donor_requests__donations',
    ]
    combined_q = Q()

    for src in sources:
        amount_field = f'{src}__amount'
        subsidized_field = f'{src}__subsidized_price'

        if amount_filter == 'small':
            clause = Q(**{f'{amount_field}__lte': 5000}) | Q(**{f'{subsidized_field}__lte': 5000})
        elif amount_filter == 'medium':
            clause = (
                Q(**{f'{amount_field}__gt': 5000, f'{amount_field}__lte': 20000}) |
                Q(**{f'{subsidized_field}__gt': 5000, f'{subsidized_field}__lte': 20000})
            )
        elif amount_filter == 'large':
            clause = Q(**{f'{amount_field}__gt': 20000}) | Q(**{f'{subsidized_field}__gt': 20000})
        else:
            return None

        combined_q |= clause

    return combined_q


@login_required

def recipient_requests_view(request):

    """View for recipients to see all their requests"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient privileges required.')

        return redirect('dashboard')



    recipient_profile = get_object_or_404(RecipientProfile, user=request.user)



    # Base queryset: all requests created by this recipient (including anonymous)

    base_queryset = RequestManagement.objects.filter(

        recipient=recipient_profile

    )



    # Apply default ordering and select-related data

    requests = base_queryset.select_related(
        'foodbank',
        'updated_by',
        'foodbank_request',
        'donation',
    ).prefetch_related(
        'foodbank_request__donations',
        'foodbank_request_created__donations',
        'donor_requests__donations',
        'donations',
        'donation_allocations__donation'
    ).order_by('-time_of_request')



    # Get filter parameters

    status_filter = request.GET.get('status', 'all')

    type_filter = request.GET.get('type', 'all')

    category_filter = request.GET.get('category', 'all')

    request_type_choices = {'all', 'food', 'non_food'}
    donation_type_choices = {'all', 'free_goods', 'subsidized', 'money'}

    # Backward compatibility: older UI had type/category semantics swapped.
    if type_filter in donation_type_choices and category_filter in request_type_choices:
        type_filter, category_filter = category_filter, type_filter
    elif type_filter in donation_type_choices and category_filter == 'all':
        category_filter = type_filter
        type_filter = 'all'
    elif category_filter in request_type_choices and type_filter == 'all':
        type_filter = category_filter
        category_filter = 'all'

    if type_filter not in request_type_choices:
        type_filter = 'all'
    if category_filter not in donation_type_choices:
        category_filter = 'all'

    quantity_filter = (request.GET.get('quantity', 'all') or 'all').strip().lower()
    amount_filter = (request.GET.get('amount', 'all') or 'all').strip().lower()
    valid_range_filters = {'all', 'small', 'medium', 'large'}
    if quantity_filter not in valid_range_filters:
        quantity_filter = 'all'
    if amount_filter not in valid_range_filters:
        amount_filter = 'all'

    anonymous_filter = request.GET.get('anonymous', 'all')

    delivery_filter = request.GET.get('delivery', 'all')
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'

    acknowledgment_filter = request.GET.get('acknowledgment', 'all')

    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()

    search_query = request.GET.get('search', '').strip()

    sort_filter = request.GET.get('sort', 'newest')



    # Filter by status if provided
    if status_filter != 'all':
        requests = requests.filter(status=status_filter)



    # Filter by type (food / non-food)
    if type_filter != 'all':
        requests = requests.filter(
            Q(request_category=type_filter) | Q(request_type=type_filter)
        )

    # Filter by category (based on linked donations)
    donation_filter = _build_recipient_type_filter(category_filter)
    if donation_filter is not None:
        requests = requests.filter(donation_filter).distinct()

    # Filter by requested quantity band
    if quantity_filter == 'small':
        requests = requests.filter(quantity__gte=1, quantity__lte=50)
    elif quantity_filter == 'medium':
        requests = requests.filter(quantity__gt=50, quantity__lte=200)
    elif quantity_filter == 'large':
        requests = requests.filter(quantity__gt=200)

    # Filter by linked donation amount/subsidized price band
    amount_q = _build_recipient_amount_filter(amount_filter)
    if amount_q is not None:
        requests = requests.filter(amount_q).distinct()

    if anonymous_filter == 'anonymous':
        requests = requests.filter(is_anonymous=True)
    elif anonymous_filter == 'not_anonymous':
        requests = requests.filter(is_anonymous=False)



    # Filter by delivery method if provided

    if delivery_filter != 'all':

        requests = requests.filter(delivery_method=delivery_filter)



    # Filter by acknowledgment status if provided

    if acknowledgment_filter != 'all':

        if acknowledgment_filter == 'received':
            requests = requests.filter(
                Q(additional_notes__icontains='Receipt Confirmed') |
                Q(status='donation_received')
            ).exclude(status='declined')
        elif acknowledgment_filter == 'acknowledged':
            requests = requests.filter(
                Q(acknowledged_by_recipient=True) &

                ~Q(additional_notes__icontains='Receipt Confirmed')

            )

        elif acknowledgment_filter == 'not_acknowledged':

            requests = requests.filter(

                Q(status__in=['fulfilled', 'partial']) &

                Q(acknowledged_by_recipient=False)

            )



    # Filter by date range (start/end only)
    if date_from:

        try:

            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')

            from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))

            requests = requests.filter(time_of_request__gte=from_date)

        except ValueError:

            pass

    if date_to:

        try:

            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')

            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))

            requests = requests.filter(time_of_request__lte=to_date)

        except ValueError:

            pass



    # Filter by search query (description, location, notes)

    if search_query:

        requests = requests.filter(

            Q(description__icontains=search_query) |

            Q(location__icontains=search_query) |

            Q(foodbank__foodbank_name__icontains=search_query) |

            Q(additional_notes__icontains=search_query)

        )



    # Sort

    if sort_filter == 'oldest':

        requests = requests.order_by('time_of_request')

    else:

        requests = requests.order_by('-time_of_request')



    # Pagination

    paginator = Paginator(requests, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    for req in page_obj.object_list:
        req.fulfillment_breakdown = _get_request_fulfillment_breakdown(req)
        raw_timeline = _build_request_quantity_timeline(req)
        req.quantity_timeline, req.quantity_timeline_has_declines = _build_recipient_timeline(raw_timeline)
        req.primary_donation = _get_request_primary_donation(req)
        if req.primary_donation:
            req.primary_donor_note = (
                req.primary_donation.message
                or req.primary_donation.csr_description
                or req.primary_donation.other_description
                or ""
            )
        else:
            req.primary_donor_note = ""


    context = {

        'title': 'My Requests',

        'requests': page_obj,

        'status_filter': status_filter,

        'type_filter': type_filter,
        'category_filter': category_filter,
        'quantity_filter': quantity_filter,
        'amount_filter': amount_filter,
        'anonymous_filter': anonymous_filter,

        'delivery_filter': delivery_filter,

        'acknowledgment_filter': acknowledgment_filter,

        'date_from': date_from,

        'date_to': date_to,

        'search_query': search_query,

        'sort_filter': sort_filter,

        'total_count': base_queryset.count(),

        'pending_count': base_queryset.filter(status='pending').count(),

        'fulfilled_count': base_queryset.filter(status='fulfilled').count(),

        'denied_count': base_queryset.filter(status='declined').count(),

        'partial_count': base_queryset.filter(status='partial').count(),

    }

    return render(request, 'authentication/recipient_requests.html', context)



from django.db.models import Q



@login_required

def foodbank_requests_view(request):

    """View for foodbanks to manage requests directed to them"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank privileges required.')

        return redirect('dashboard')

    

    foodbank_profile = get_object_or_404(FoodBankProfile, user=request.user)

    

    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')
    anonymous_filter = request.GET.get('anonymous', 'all')
    delivery_filter = request.GET.get('delivery', 'all')
    if delivery_filter == 'both':
        delivery_filter = 'all'
    date_filter = request.GET.get('date_range', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')
    search_query = request.GET.get('search', '').strip()
    sort_filter = request.GET.get('sort', 'newest')

    

    # Include: Direct requests + Assigned anonymous + Unassigned anonymous

    # Exclude requests that have ACCEPTED subsidized donations (they go to the dedicated subsidized table)

    # Keep requests with pending subsidized donations here until they're accepted

    requests = RequestManagement.objects.filter(

        Q(foodbank=foodbank_profile, is_anonymous=False) |  # Direct requests

        Q(is_anonymous=True, assigned_foodbank=foodbank_profile) |  # Assigned anonymous

        Q(is_anonymous=True, assigned_foodbank__isnull=True)  # Unassigned anonymous

    ).exclude(

        donations__donation_type='subsidized',

        donations__status='accepted'

    ).select_related('recipient', 'updated_by', 'donation__donor', 'donation__donor__donor_profile').prefetch_related('foodbank_request_created__donations', 'donor_requests__donations')

    

    # Filter by status if provided (align filter options with status labels shown in table)
    if status_filter != 'all':
        if status_filter == 'recipient_received':
            requests = requests.filter(status='fulfilled', acknowledged_by_recipient=True)
        elif status_filter == 'fulfilled':
            requests = requests.filter(status='fulfilled', acknowledged_by_recipient=False)
        elif status_filter == 'declined_foodbank':
            requests = requests.filter(status='declined', updated_by__user_type='FOODBANK')
        elif status_filter == 'declined_recipient':
            requests = requests.filter(status='declined', updated_by__user_type='RECIPIENT')
        else:
            requests = requests.filter(status=status_filter)

    

    # Filter by type if provided

    if type_filter != 'all':

        requests = requests.filter(request_type=type_filter)

    if anonymous_filter == 'anonymous':

        requests = requests.filter(is_anonymous=True)

    elif anonymous_filter == 'not_anonymous':

        requests = requests.filter(is_anonymous=False)

    

    # Filter by delivery method if provided

    if delivery_filter != 'all':

        requests = requests.filter(delivery_method=delivery_filter)

    

    # Filter by date range if provided

    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()

    

    if (date_filter == 'custom') or (date_filter in ('', 'all') and (date_from or date_to)):

        # Custom date range

        if date_from:

            try:

                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')

                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))

                requests = requests.filter(time_of_request__gte=from_date)

            except ValueError:

                pass

        if date_to:

            try:

                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')

                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))

                requests = requests.filter(time_of_request__lte=to_date)

            except ValueError:

                pass

    elif date_filter != 'all':

        now = timezone.now()

        if date_filter == 'today':

            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)

            requests = requests.filter(time_of_request__gte=start_date)

        elif date_filter == 'week':

            start_date = now - timedelta(days=7)

            requests = requests.filter(time_of_request__gte=start_date)

        elif date_filter == 'month':

            start_date = now - timedelta(days=30)

            requests = requests.filter(time_of_request__gte=start_date)

        elif date_filter == '3months':

            start_date = now - timedelta(days=90)

            requests = requests.filter(time_of_request__gte=start_date)

    

    # Filter by quantity range if provided

    if quantity_filter != 'all':

        if quantity_filter == 'small':

            requests = requests.filter(quantity__lte=10)

        elif quantity_filter == 'medium':

            requests = requests.filter(quantity__gt=10, quantity__lte=50)

        elif quantity_filter == 'large':

            requests = requests.filter(quantity__gt=50)

    # Filter by amount range using all donation links that feed this table:
    # - RequestManagement.donations (Donation.request_management FK)
    # - RequestManagement.donation (direct FK)
    # - RequestManagement.foodbank_request_created -> FoodBankRequest.donations
    # - RequestManagement.donor_requests -> FoodBankRequest.donations
    if amount_filter != 'all':
        if amount_filter == 'small':
            amount_q = (
                Q(donations__amount__lte=5000) |
                Q(donations__subsidized_price__lte=5000) |
                Q(donation__amount__lte=5000) |
                Q(donation__subsidized_price__lte=5000) |
                Q(foodbank_request__donations__amount__lte=5000) |
                Q(foodbank_request__donations__subsidized_price__lte=5000) |
                Q(foodbank_request_created__donations__amount__lte=5000) |
                Q(foodbank_request_created__donations__subsidized_price__lte=5000) |
                Q(donor_requests__donations__amount__lte=5000) |
                Q(donor_requests__donations__subsidized_price__lte=5000) |
                Q(donation_allocations__amount__lte=5000)
            )
            requests = requests.filter(amount_q).distinct()
        elif amount_filter == 'medium':
            amount_q = (
                Q(donations__amount__gt=5000, donations__amount__lte=20000) |
                Q(donations__subsidized_price__gt=5000, donations__subsidized_price__lte=20000) |
                Q(donation__amount__gt=5000, donation__amount__lte=20000) |
                Q(donation__subsidized_price__gt=5000, donation__subsidized_price__lte=20000) |
                Q(foodbank_request__donations__amount__gt=5000, foodbank_request__donations__amount__lte=20000) |
                Q(foodbank_request__donations__subsidized_price__gt=5000, foodbank_request__donations__subsidized_price__lte=20000) |
                Q(foodbank_request_created__donations__amount__gt=5000, foodbank_request_created__donations__amount__lte=20000) |
                Q(foodbank_request_created__donations__subsidized_price__gt=5000, foodbank_request_created__donations__subsidized_price__lte=20000) |
                Q(donor_requests__donations__amount__gt=5000, donor_requests__donations__amount__lte=20000) |
                Q(donor_requests__donations__subsidized_price__gt=5000, donor_requests__donations__subsidized_price__lte=20000) |
                Q(donation_allocations__amount__gt=5000, donation_allocations__amount__lte=20000)
            )
            requests = requests.filter(amount_q).distinct()
        elif amount_filter == 'large':
            amount_q = (
                Q(donations__amount__gt=20000) |
                Q(donations__subsidized_price__gt=20000) |
                Q(donation__amount__gt=20000) |
                Q(donation__subsidized_price__gt=20000) |
                Q(foodbank_request__donations__amount__gt=20000) |
                Q(foodbank_request__donations__subsidized_price__gt=20000) |
                Q(foodbank_request_created__donations__amount__gt=20000) |
                Q(foodbank_request_created__donations__subsidized_price__gt=20000) |
                Q(donor_requests__donations__amount__gt=20000) |
                Q(donor_requests__donations__subsidized_price__gt=20000) |
                Q(donation_allocations__amount__gt=20000)
            )
            requests = requests.filter(amount_q).distinct()

    

    # Filter by search query if provided (Description, Location, Recipient, Notes)

    if search_query:

        requests = requests.filter(

            Q(description__icontains=search_query) |

            Q(location__icontains=search_query) |

            Q(recipient__full_name__icontains=search_query) |

            Q(additional_notes__icontains=search_query)

        )

    

    # Handle sorting

    if sort_filter == 'oldest':

        requests = requests.order_by('time_of_request')  # Oldest first

    else:

        requests = requests.order_by('-time_of_request')  # Newest first (default)

    

    # Pagination

    paginator = Paginator(requests, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    for req in page_obj.object_list:
        req.manage_recipient_display = _manage_request_recipient_display(req)
        req.fulfillment_breakdown = _get_request_fulfillment_breakdown(req)
        req.quantity_timeline = _build_request_quantity_timeline(req)
        req.quantity_timeline_has_declines = any(entry.get('is_declined') for entry in req.quantity_timeline)



    # Update counts (counts don't need sorting)

    base_queryset = RequestManagement.objects.filter(

        Q(foodbank=foodbank_profile, is_anonymous=False) |

        Q(is_anonymous=True, assigned_foodbank=foodbank_profile) |

        Q(is_anonymous=True, assigned_foodbank__isnull=True)

    )

    

    context = {

        'title': 'Manage Requests',

        'requests': page_obj,

        'current_full_path': request.get_full_path(),

        'status_filter': status_filter,

        'type_filter': type_filter,

        'anonymous_filter': anonymous_filter,

        'delivery_filter': delivery_filter,

        'date_filter': date_filter,

        'date_from': date_from,

        'date_to': date_to,

        'quantity_filter': quantity_filter,
        'amount_filter': amount_filter,

        'search_query': search_query,

        'sort_filter': sort_filter,

        'total_count': base_queryset.count(),

        'pending_count': base_queryset.filter(status='pending').count(),

        'fulfilled_count': base_queryset.filter(status='fulfilled').count(),

        'denied_count': base_queryset.filter(status='declined').count(),

        'partial_count': base_queryset.filter(status='partial').count(),

        'awaiting_recipient_count': base_queryset.filter(status='awaiting_recipient').count(),

    }

    return render(request, 'authentication/foodbank_manage_requests.html', context)

@login_required

def create_request(request):

    """Create a new request from recipient to foodbank"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient privileges required.')

        return redirect('dashboard')

    

    recipient_profile = get_object_or_404(RecipientProfile, user=request.user)

    

    if request.method == 'POST':

        # Get form data

        foodbank_id = request.POST.get('foodbank')

        request_type = request.POST.get('request_type')

        description = (request.POST.get('description') or '').strip()

        quantity = request.POST.get('quantity')

        unit = request.POST.get('unit')

        custom_unit = request.POST.get('custom_unit', '')

        delivery_method = request.POST.get('delivery_method')

        location = request.POST.get('location')

        is_anonymous = request.POST.get('is_anonymous') == 'true'  

        additional_notes = request.POST.get('additional_notes', '')

        

        # Validate required fields (note: foodbank_id not required if anonymous)

        if not all([request_type, description, quantity, unit, delivery_method, location]):

            messages.error(request, 'All fields except foodbank are required.')

            return redirect('create_request')

        if len(description) > 400:
            messages.error(request, 'Description must be 400 characters or fewer.')
            return redirect('create_request')

        

        # Validate custom unit if 'other' is selected

        if unit == 'other' and not custom_unit.strip():

            messages.error(request, 'Custom unit is required when "Other" is selected.')

            return redirect('create_request')

        

        try:

            quantity = int(quantity)

            

            if is_anonymous:

                # Anonymous request: no specific foodbank

                # Format initial note with label

                formatted_notes = f"--- Recipient Note ---\n{additional_notes}" if additional_notes.strip() else ""

                

                new_request = RequestManagement.objects.create(

                    recipient=recipient_profile,

                    foodbank=None,  # â† No specific foodbank

                    request_type=request_type,

                    request_category=request_type,  # Category matches type

                    description=description,

                    quantity=quantity,

                    unit=unit,

                    custom_unit=custom_unit if unit == 'other' else None,

                    delivery_method=delivery_method,

                    location=location,

                    is_anonymous=True,  # â† Mark as anonymous

                    additional_notes=formatted_notes

                )

                

                # Notify ALL approved and active foodbanks

                foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True)

                for foodbank in foodbanks:

                    Notification.create_notification(

                        user=foodbank.user,

                        notification_type='anonymous_request',

                        message=f'New anonymous request: {description}',

                        related_object=new_request

                    )

                

                messages.success(request, 'Anonymous request created successfully! It is now visible to all foodbanks.')

                

            else:

                # Direct request: requires foodbank selection

                if not foodbank_id:

                    messages.error(request, 'Please select a foodbank for direct requests.')

                    return redirect('create_request')

                

                foodbank = FoodBankProfile.objects.get(id=foodbank_id)

                # Format initial note with label

                formatted_notes = f"--- Recipient Note ---\n{additional_notes}" if additional_notes.strip() else ""

                

                new_request = RequestManagement.objects.create(

                    recipient=recipient_profile,

                    foodbank=foodbank,

                    request_type=request_type,

                    request_category=request_type,  # Category matches type

                    description=description,

                    quantity=quantity,

                    unit=unit,

                    custom_unit=custom_unit if unit == 'other' else None,

                    delivery_method=delivery_method,

                    location=location,

                    is_anonymous=False,

                    additional_notes=formatted_notes

                )

                

                # Notify only the selected foodbank

                Notification.create_notification(

                    user=foodbank.user,

                    notification_type='request',

                    message=f'New request from {recipient_profile.full_name}: {description}',

                    related_object=new_request

                )

                

                messages.success(request, f'Request sent to {foodbank.foodbank_name} successfully!')

            

            return redirect('recipient_requests_view')

            

        except FoodBankProfile.DoesNotExist:

            messages.error(request, 'Selected foodbank not found.')

        except ValueError:

            messages.error(request, 'Please enter a valid quantity.')

        except Exception as e:

            messages.error(request, f'Error creating request: {str(e)}')

    

    # Get approved and active foodbanks for the form

    foodbanks = FoodBankProfile.objects.filter(is_approved='approved', user__is_active=True)

    

    context = {

        'title': 'Create New Request',

        'foodbanks': foodbanks,

        'quantity_units': QUANTITY_UNITS,

    }

    return render(request, 'authentication/create_request.html', context)



@login_required

def update_request_status(request, request_id):

    """Update request status (for foodbanks)"""

    if request.user.user_type != 'FOODBANK':

        return JsonResponse({'success': False, 'error': 'Access denied'})

    

    try:

        foodbank_profile = get_object_or_404(FoodBankProfile, user=request.user)

        request_obj = get_object_or_404(RequestManagement, id=request_id, foodbank=foodbank_profile)

        

        if request.method == 'POST':

            new_status = request.POST.get('status')

            notes = request.POST.get('notes', '')

            

            if new_status not in ['pending', 'fulfilled', 'denied', 'partial']:

                return JsonResponse({'success': False, 'error': 'Invalid status'})

            

            # Update the request

            request_obj.status = new_status

            request_obj.additional_notes = notes

            request_obj.updated_by = request.user

            

            if new_status == 'fulfilled':

                request_obj.fulfilled_at = timezone.now()

            

            request_obj.save()

            

            # Create notification for recipient

            status_messages = {

                'fulfilled': f'Your request "{request_obj.description}" has been fulfilled!',

                'denied': f'Your request "{request_obj.description}" has been denied.',

                'partial': f'Your request "{request_obj.description}" has been partially fulfilled.',

                'pending': f'Your request "{request_obj.description}" status has been updated to pending.'

            }

            

            Notification.create_notification(

                user=request_obj.recipient.user,

                notification_type='request_fulfilled' if new_status == 'fulfilled' else 'system',

                message=status_messages.get(new_status, 'Your request status has been updated.'),

                related_object=request_obj

            )

            

            return JsonResponse({

                'success': True, 

                'message': f'Request status updated to {new_status}',

                'new_status': new_status,

                'badge_class': request_obj.get_status_badge_class()

            })

    

    except Exception as e:

        return JsonResponse({'success': False, 'error': str(e)})

    

    return JsonResponse({'success': False, 'error': 'Invalid request method'})



@login_required

def request_detail(request, request_id):

    """View detailed information about a specific request"""

    try:

        request_obj = get_object_or_404(RequestManagement, id=request_id)

        

        # Check permissions

        if request.user.user_type == 'RECIPIENT':

            if request_obj.recipient.user != request.user:

                messages.error(request, 'Access denied.')

                return redirect('recipient_requests_view')

        elif request.user.user_type == 'FOODBANK':

            # Foodbank can view:
            # 1) Direct requests addressed to their foodbank
            # 2) Anonymous requests assigned to their foodbank
            # 3) Anonymous requests not yet assigned (so they can inspect before assigning)
            if getattr(request_obj, 'is_anonymous', False):
                assigned_fb = getattr(request_obj, 'assigned_foodbank', None)
                if assigned_fb and assigned_fb.user != request.user:
                    messages.error(request, 'Access denied.')
                    return redirect('foodbank_requests')
            else:
                target_fb = getattr(request_obj, 'foodbank', None)
                if not target_fb or target_fb.user != request.user:
                    messages.error(request, 'Access denied.')
                    return redirect('foodbank_requests')

        else:

            messages.error(request, 'Access denied.')

            return redirect('dashboard')

        

        context = {

            'title': 'Request Details',

            'request': request_obj,

            'can_update': request.user.user_type == 'FOODBANK',

        }

        return render(request, 'authentication/request_detail.html', context)

        

    except Exception as e:

        messages.error(request, f'Error loading request: {str(e)}')

        return redirect('dashboard')

    





@login_required

def manage_subscription(request):

    user = request.user

    subscription, created = Subscription.objects.get_or_create(user=user, user_type=user.user_type)



    if request.method == 'POST':

        # Handle subscription payment

        subscription.is_active = True

        subscription.payment_status = 'paid'

        subscription.last_payment_date = now()

        subscription.end_date = now() + timedelta(days=30)  # 1-month subscription

        subscription.save()



        # Notify admin (example: send email or log notification)

        messages.success(request, "Your subscription has been activated successfully!")

        return redirect('dashboard')



    return render(request, 'subscriptions/manage_subscription.html', {'subscription': subscription})  



from django.core.mail import send_mail
from django.conf import settings



def notify_admin(subscription):

    subject = f"Subscription Update: {subscription.user.username}"

    message = f"""

    User: {subscription.user.username}

    Type: {subscription.user_type}

    Status: {'Paid' if subscription.payment_status == 'paid' else 'Unpaid'}

    Last Payment: {subscription.last_payment_date}

    """

    send_mail(
        subject,
        message,
        getattr(settings, 'DEFAULT_FROM_EMAIL', 'admin@foodbankhub.com'),
        ['admin@foodbankhub.com'],
    )

    

from django.utils import timezone

from django.db.models import Sum

from django.contrib import messages

from django.shortcuts import get_object_or_404, redirect



# ...existing code...

# ...existing code...

@login_required

def accept_donation(request, donation_id):

    # Ensure only the foodbank that owns the donation can accept it

    donation = get_object_or_404(

        Donation,

        id=donation_id,

        foodbank=request.user.foodbank_profile

    )



    # Prevent re-accepting

    if donation.status == 'accepted':

        messages.warning(request, "This donation has already been accepted.")

        return _redirect_back_or_default(request, 'foodbank_donations_list')



    # Mark donation accepted

    donation.status = 'accepted'

    update_fields = ['status']

    if donation.foodbank_request_id is not None and donation.accepted_by_recipient_id is not None:

        donation.accepted_by_recipient = None

        update_fields.append('accepted_by_recipient')

    donation.save(update_fields=update_fields)



    # If this donation is linked to a foodbank_request, try to update the original recipient request

    if donation.foodbank_request:

        # Try a few common attribute names that might point to the recipient's request

        original_req = (

            getattr(donation.foodbank_request, 'original_request', None)

            or getattr(donation.foodbank_request, 'linked_recipient_request', None)

            or getattr(donation.foodbank_request, 'recipient_request', None)

            or getattr(donation.foodbank_request, 'linked_request_management', None)

        )



        if not original_req and hasattr(donation.foodbank_request, 'recipient_requests'):

            original_req = donation.foodbank_request.recipient_requests.order_by('-time_of_request').first()



        if original_req:

            try:

                # Use atomic update to avoid partial writes

                from django.db import transaction

                with transaction.atomic():

                    if hasattr(original_req, 'acknowledged_by_recipient'):

                        original_req.acknowledged_by_recipient = False

                    if hasattr(original_req, 'acknowledged_at'):

                        original_req.acknowledged_at = None

                    if hasattr(original_req, 'acknowledged_by_recipient') or hasattr(original_req, 'acknowledged_at'):

                        ack_update_fields = []

                        if hasattr(original_req, 'acknowledged_by_recipient'):

                            ack_update_fields.append('acknowledged_by_recipient')

                        if hasattr(original_req, 'acknowledged_at'):

                            ack_update_fields.append('acknowledged_at')

                        if ack_update_fields:

                            original_req.save(update_fields=ack_update_fields)



                    # ITEM-based request handling (quantity) - only for item/subsidized donations, NOT monetary

                    donated_qty = None

                    if donation.donation_type == 'subsidized':

                        donated_qty = donation.subsidized_quantity or donation.quantity

                    elif donation.donation_type == 'item':

                        donated_qty = donation.quantity



                    if donated_qty is None and donation.foodbank_request and getattr(donation.foodbank_request, 'quantity_needed', None):

                        donated_qty = donation.foodbank_request.quantity_needed



                    if donation.donation_type in ['item', 'subsidized'] and donated_qty is not None and getattr(original_req, 'quantity', None) is not None:

                        donated_qty = int(donated_qty or 0)

                        fulfilled = int(getattr(original_req, 'quantity_fulfilled', 0) or 0)

                        needed = int(getattr(original_req, 'quantity', 0) or 0)

                        remaining = needed - fulfilled



                        if donated_qty > 0 and remaining > 0:

                            use = min(donated_qty, remaining)

                            original_req.quantity_fulfilled = fulfilled + use



                            if original_req.quantity_fulfilled >= needed:

                                original_req.status = 'awaiting_recipient'

                            else:

                                original_req.status = 'partial'



                            # ðŸ”‘ CRITICAL RESET

                            if hasattr(original_req, 'acknowledged_by_recipient'):

                                original_req.acknowledged_by_recipient = False

                            if hasattr(original_req, 'acknowledged_at'):

                                original_req.acknowledged_at = None



                            original_req.save()





                            # Notify recipient with clear message depending on partial vs full coverage

                            if original_req.status == 'partial':

                                Notification.objects.create(

                                    user=original_req.recipient.user,

                                    message=(

                                        f"Good news â€” {use} {getattr(original_req,'quantity_unit', '')} "

                                        f"has been allocated for your request '{getattr(original_req, 'description', original_req.id)}'. "

                                        f"Status: Partially fulfilled ({original_req.quantity_fulfilled}/{original_req.quantity})."

                                    )

                                )

                            else:

                                Notification.objects.create(

                                    user=original_req.recipient.user,

                                    notification_type='donation_received',

                                    message=(

                                        f"A donation covering your request '{getattr(original_req, 'description', original_req.id)}' "

                                        f"is ready. Please acknowledge receipt so we can mark the request as fulfilled."

                                    )

                                )



                    # MONEY-based request handling (amount)

                    elif getattr(donation, 'amount', None) is not None and getattr(original_req, 'amount_needed', None) is not None:

                        donated_amount = float(donation.amount or 0)

                        received = float(getattr(original_req, 'amount_received', 0) or 0)

                        needed_amount = float(getattr(original_req, 'amount_needed', 0) or 0)

                        remaining_amount = needed_amount - received



                        if donated_amount > 0 and remaining_amount > 0:

                            use_amount = min(donated_amount, remaining_amount)

                            original_req.amount_received = received + use_amount



                            if original_req.amount_received >= needed_amount:

                                original_req.status = 'awaiting_recipient'

                            else:

                                original_req.status = 'partial'



                            original_req.save()



                            if original_req.status == 'partial':

                                Notification.objects.create(

                                    user=original_req.recipient.user,

                                    message=(

                                        f"Partial monetary contribution received for your request '{getattr(original_req, 'description', original_req.id)}': "

                                        f"KSH {original_req.amount_received}/{original_req.amount_needed}."

                                    )

                                )

                            else:

                                Notification.objects.create(

                                    user=original_req.recipient.user,

                                    notification_type='donation_received',

                                    message=(

                                        f"A monetary donation covering your request '{getattr(original_req, 'description', original_req.id)}' is ready. "

                                        "Please acknowledge receipt to complete the process."

                                    )

                                )

                    # MONETARY donation to a quantity-based request: add the quantity this money is "for" to quantity_fulfilled

                    elif donation.donation_type == 'money' and getattr(donation, 'amount', None) is not None:

                        # Quantity this monetary donation is for (e.g. "KES 500 for 50 kg" -> 50)
                        qty_for_money = None
                        if donation.foodbank_request and getattr(donation.foodbank_request, 'quantity_needed', None):
                            qty_for_money = int(donation.foodbank_request.quantity_needed or 0)

                        if qty_for_money is not None and qty_for_money > 0 and getattr(original_req, 'quantity', None) is not None:
                            # Quantity-based request: add fulfilled amount so partial + monetary can become fulfilled
                            fulfilled = int(getattr(original_req, 'quantity_fulfilled', 0) or 0)
                            needed = int(getattr(original_req, 'quantity', 0) or 0)
                            remaining = needed - fulfilled
                            if remaining > 0:
                                use = min(qty_for_money, remaining)
                                original_req.quantity_fulfilled = fulfilled + use
                                if original_req.quantity_fulfilled >= needed:
                                    original_req.status = 'awaiting_recipient'
                                else:
                                    original_req.status = 'partial'
                                if hasattr(original_req, 'acknowledged_by_recipient'):
                                    original_req.acknowledged_by_recipient = False
                                if hasattr(original_req, 'acknowledged_at'):
                                    original_req.acknowledged_at = None
                                original_req.save()
                                if original_req.status == 'partial':
                                    Notification.objects.create(
                                        user=original_req.recipient.user,
                                        message=(
                                            f"Monetary donation received. Request partially fulfilled: "
                                            f"{original_req.quantity_fulfilled}/{original_req.quantity}."
                                        )
                                    )
                                else:
                                    Notification.objects.create(
                                        user=original_req.recipient.user,
                                        notification_type='donation_received',
                                        message=(
                                            f"A monetary donation covering your request '{getattr(original_req, 'description', original_req.id)}' "
                                            "is ready. Please acknowledge receipt to complete the process."
                                        )
                                    )
                            else:
                                if getattr(original_req, 'status', '') not in ['fulfilled', 'awaiting_recipient']:
                                    original_req.status = 'awaiting_recipient'
                                    original_req.save()
                                Notification.objects.create(
                                    user=original_req.recipient.user,
                                    notification_type='donation_received',
                                    message=f"A monetary donation of KES {donation.amount} has been received for your request '{getattr(original_req, 'description', original_req.id)}'. Please acknowledge receipt."
                                )
                        else:
                            # No quantity_needed on foodbank request or non-quantity request: just mark as received
                            if getattr(original_req, 'status', '') not in ['fulfilled', 'awaiting_recipient']:
                                original_req.status = 'awaiting_recipient'
                                original_req.save()
                            Notification.objects.create(
                                user=original_req.recipient.user,
                                notification_type='donation_received',
                                message=f"A monetary donation of KES {donation.amount} has been received for your request '{getattr(original_req, 'description', original_req.id)}'. Please acknowledge receipt."
                            )

                    else:

                        # Unknown request shape - mark as awaiting_recipient if not already fulfilled

                        if getattr(original_req, 'status', '') not in ['fulfilled', 'awaiting_recipient']:

                            original_req.status = 'awaiting_recipient'

                            original_req.save()

                            Notification.objects.create(

                                user=original_req.recipient.user,

                                notification_type='donation_received',

                                message=f"A donation for your request '{getattr(original_req, 'description', original_req.id)}' is ready. Please acknowledge receipt."

                            )

            except Exception as e:

                # Log but don't break flow

                print(f"Error updating original request on accept_donation: {e}")



    messages.success(request, "Donation accepted and request updated!")

    return _redirect_back_or_default(request, 'foodbank_requests_view')

# ...existing code...

@login_required

def decline_donation(request, donation_id):

    donation = get_object_or_404(

        Donation,

        id=donation_id,

        foodbank=request.user.foodbank_profile

    )



    if request.method == 'POST':

        donation.status = 'declined'

        donation.decline_message = request.POST.get('message', '').strip()

        donation.save(update_fields=['status', 'decline_message'])



        foodbank_request = donation.foodbank_request



        # ðŸ”— Find the original recipient request safely

        request_obj = (

            getattr(foodbank_request, 'original_request', None)

            or getattr(foodbank_request, 'linked_recipient_request', None)

            or getattr(foodbank_request, 'recipient_request', None)

        )



        if request_obj:

            # Check if any non-declined donations remain

            remaining_donations = foodbank_request.donations.exclude(status='declined')



            if not remaining_donations.exists():

                update_fields = ['status']

                if getattr(request_obj, 'quantity_fulfilled', 0) > 0:

                    request_obj.status = 'partial'

                else:

                    request_obj.status = 'declined'



                if hasattr(request_obj, 'decline_message'):

                    request_obj.decline_message = donation.decline_message

                    update_fields.append('decline_message')



                # Track that the foodbank performed this update so recipients can mask the status

                if hasattr(request_obj, 'updated_by'):

                    request_obj.updated_by = request.user

                    update_fields.append('updated_by')



                request_obj.save(update_fields=update_fields)



        messages.warning(request, "Donation declined.")



    return _redirect_back_or_default(request, 'foodbank_requests_view')



    



from django.shortcuts import redirect, get_object_or_404

from django.contrib import messages

from .models import RequestManagement



from django.utils import timezone

from django.contrib import messages

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required

from django.db.models import Q



# Helper functions used by fulfillment flows ---------------------------------


def _filter_donations_by_request_keywords(queryset, req):

    """Limit donations to those whose metadata matches request keywords."""

    text_parts = [
        getattr(req, 'title', ''),
        getattr(req, 'description', ''),
        getattr(req, 'additional_notes', ''),
    ]
    combined_text = ' '.join(filter(None, text_parts)).lower()
    keywords = [word for word in re.findall(r'\w+', combined_text) if len(word) > 2]

    if not keywords:
        return queryset

    keyword_filter = Q()
    for word in keywords:
        keyword_filter |= (
            Q(item_name__icontains=word) |
            Q(message__icontains=word) |
            Q(other_description__icontains=word) |
            Q(foodbank_request__title__icontains=word) |
            Q(foodbank_request__description__icontains=word)
        )

    return queryset.filter(keyword_filter)


def _get_request_quantity_value(request_obj):
    """Return the numeric quantity associated with a request-like object."""
    if not request_obj:
        return None
    qty = getattr(request_obj, 'quantity_needed', None)
    if qty:
        return qty
    qty = getattr(request_obj, 'quantity', None)
    if qty:
        return qty
    linked_req = getattr(request_obj, 'linked_request_management', None)
    if linked_req:
        return _get_request_quantity_value(linked_req)
    return None


def _get_request_unit_label(request_obj):
    """Resolve the human-readable unit label for FoodBankRequest/RequestManagement objects."""
    if not request_obj:
        return ''

    custom_unit = getattr(request_obj, 'custom_unit', None)

    quantity_unit = getattr(request_obj, 'quantity_unit', None)
    if quantity_unit:
        if quantity_unit == 'other' and custom_unit:
            return custom_unit
        try:
            return request_obj.get_quantity_unit_display()
        except Exception:
            return quantity_unit

    unit_field = getattr(request_obj, 'unit', None)
    if unit_field:
        if unit_field == 'other' and custom_unit:
            return custom_unit
        try:
            return request_obj.get_unit_display()
        except Exception:
            return unit_field

    if custom_unit:
        return custom_unit

    linked_req = getattr(request_obj, 'linked_request_management', None)
    if linked_req:
        return _get_request_unit_label(linked_req)

    return ''


def _resolve_donation_target_quantity_and_unit(donation):
    """Determine the total quantity/unit the donation was meant to cover."""
    source_req = getattr(donation, 'foodbank_request', None)
    target_qty = _get_request_quantity_value(source_req)
    target_unit = _get_request_unit_label(source_req)
    return target_qty, target_unit


def _get_donation_remaining_supported_units(donation):
    """Return how many units this donation can still cover (for money/subsidized)."""
    if donation.donation_type not in ('money', 'subsidized'):
        return None

    target_qty = _get_request_quantity_value(getattr(donation, 'foodbank_request', None))
    if not target_qty:
        return None

    from django.db.models import Sum as _Sum

    used_units = donation.allocations.filter(declined_by_recipient=False).aggregate(
        total=_Sum('quantity')
    )['total'] or 0

    remaining_units = target_qty - used_units
    return remaining_units if remaining_units > 0 else 0


def _prepare_donation_status_metadata(donations_queryset):
    """Attach status display/class metadata used by donor tables."""

    donations = list(donations_queryset)

    for donation in donations:

        status_display = get_display_status(donation)

        donation.status_display = status_display

        donation.status_class = STATUS_CLASS_MAP.get(status_display, 'pending')


    return donations, len(donations)


def _get_available_stock_donations(foodbank_to_search, req, request_get=None, is_full_fulfillment=False):
    """Return (list of donations, count) for use available stock: item + money + subsidized with remaining > 0."""
    from django.db.models import Sum, F, Value
    from django.db.models.functions import Coalesce

    request_get = request_get or {}
    search_query = (request_get.get('search') or '').strip()
    delivery_filter = (request_get.get('delivery') or '').strip().lower()
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    category_filter = (request_get.get('category') or '').strip().lower()
    type_filter = (request_get.get('type') or '').strip().lower().replace('-', '_')
    quantity_filter = (request_get.get('quantity') or '').strip().lower()
    amount_filter = (request_get.get('amount') or '').strip().lower()
    if quantity_filter == 'all':
        quantity_filter = ''
    if amount_filter == 'all':
        amount_filter = ''
    date_from = (request_get.get('date_from') or '').strip()
    date_to = (request_get.get('date_to') or '').strip()
    sort_filter = (request_get.get('sort') or 'most_qty').strip()

    base_item_filter = dict(
        foodbank=foodbank_to_search,
        donation_category=req.request_type,
        status='accepted',
        request_management__isnull=True,
        foodbank_request__isnull=False,
        foodbank_request__original_request__isnull=True,
    )

    # Item donations (existing logic)
    item_qs = (
        Donation.objects.filter(
            **base_item_filter,
            donation_type='item',
            quantity__isnull=False,
            quantity_unit=req.unit,
            is_allocated=False,
        )
        .annotate(
            allocated_qty=Coalesce(
                Sum('allocations__quantity', filter=Q(allocations__declined_by_recipient=False)),
                Value(0)
            )
        )
        .annotate(remaining_qty=F('quantity') - F('allocated_qty'))
        .filter(remaining_qty__gte=1)
        .select_related('donor', 'donor__donor_profile', 'foodbank_request')
    )
    item_qs = _filter_donations_by_request_keywords(item_qs, req)
    if search_query:
        item_qs = item_qs.filter(
            Q(item_name__icontains=search_query)
            | Q(message__icontains=search_query)
            | Q(donor__donor_profile__full_name__icontains=search_query)
            | Q(donor__email__icontains=search_query)
            | Q(foodbank_request__title__icontains=search_query)
            | Q(foodbank_request__description__icontains=search_query)
        )
    if delivery_filter in ('pickup', 'delivery'):
        if delivery_filter == 'delivery':
            item_qs = item_qs.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            item_qs = item_qs.filter(delivery_method='pickup')
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
            from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
            item_qs = item_qs.filter(donated_at__gte=from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
            item_qs = item_qs.filter(donated_at__lte=to_date)
        except ValueError:
            pass
    if sort_filter == 'oldest':
        item_qs = item_qs.order_by('donated_at')
    elif sort_filter == 'newest':
        item_qs = item_qs.order_by('-donated_at')
    elif sort_filter == 'least_qty':
        item_qs = item_qs.order_by('remaining_qty')
    else:
        item_qs = item_qs.order_by('-remaining_qty')

    item_list = list(item_qs)

    # Money donations: category must match request type (food/non_food), same as foodbank requests table
    monetary_categories = [req.request_type]
    if 'monetary' not in monetary_categories:
        monetary_categories.append('monetary')

    money_qs = (
        Donation.objects.filter(
            foodbank=foodbank_to_search,
            donation_category__in=[cat for cat in monetary_categories if cat],
            status='accepted',
            donation_type='money',
            amount__isnull=False,
        )
        .filter(request_management__isnull=True)
        .filter(
            Q(foodbank_request__isnull=False, foodbank_request__original_request__isnull=True)
            | Q(foodbank_request__isnull=True)
        )
        .select_related('donor', 'donor__donor_profile', 'foodbank_request')
    )
    if search_query:
        money_qs = money_qs.filter(
            Q(message__icontains=search_query)
            | Q(donor__donor_profile__full_name__icontains=search_query)
            | Q(donor__email__icontains=search_query)
            | Q(foodbank_request__title__icontains=search_query)
            | Q(foodbank_request__description__icontains=search_query)
        )
    if delivery_filter in ('pickup', 'delivery'):
        if delivery_filter == 'delivery':
            money_qs = money_qs.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            money_qs = money_qs.filter(delivery_method='pickup')
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
            from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
            money_qs = money_qs.filter(donated_at__gte=from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
            money_qs = money_qs.filter(donated_at__lte=to_date)
        except ValueError:
            pass
    money_list = []
    for d in money_qs:
        rem = d.get_remaining_amount()
        if rem > 0:
            d.remaining_amount = rem
            d.remaining_qty = _get_donation_remaining_supported_units(d)
            d.remaining_subsidized_qty = 0
            d.remaining_supported_units = d.remaining_qty
            money_list.append(d)
    if sort_filter == 'oldest':
        money_list.sort(key=lambda x: x.donated_at)
    elif sort_filter == 'newest':
        money_list.sort(key=lambda x: x.donated_at, reverse=True)
    else:
        money_list.sort(key=lambda x: x.remaining_amount, reverse=True)

    # Subsidized donations: linked to request OR direct stock
    subsidized_qs = (
        Donation.objects.filter(
            foodbank=foodbank_to_search,
            donation_category=req.request_type,
            status='accepted',
            donation_type='subsidized',
        )
        .filter(request_management__isnull=True)
        .filter(
            Q(foodbank_request__isnull=False, foodbank_request__original_request__isnull=True)
            | Q(foodbank_request__isnull=True)
        )
        .select_related('donor', 'donor__donor_profile', 'foodbank_request')
    )
    if search_query:
        subsidized_qs = subsidized_qs.filter(
            Q(message__icontains=search_query)
            | Q(subsidized_product_type__icontains=search_query)
            | Q(csr_description__icontains=search_query)
            | Q(donor__donor_profile__full_name__icontains=search_query)
            | Q(donor__email__icontains=search_query)
            | Q(foodbank_request__title__icontains=search_query)
            | Q(foodbank_request__description__icontains=search_query)
        )
    if delivery_filter in ('pickup', 'delivery'):
        if delivery_filter == 'delivery':
            subsidized_qs = subsidized_qs.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            subsidized_qs = subsidized_qs.filter(delivery_method='pickup')
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')
            from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))
            subsidized_qs = subsidized_qs.filter(donated_at__gte=from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')
            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))
            subsidized_qs = subsidized_qs.filter(donated_at__lte=to_date)
        except ValueError:
            pass
    subsidized_list = []
    for d in subsidized_qs:
        rem_qty = d.get_remaining_subsidized_quantity()
        rem_amt = d.get_remaining_amount()
        if rem_qty > 0 or rem_amt > 0:
            d.remaining_amount = rem_amt
            d.remaining_subsidized_qty = rem_qty
            d.remaining_qty = rem_qty
            d.remaining_supported_units = rem_qty
            subsidized_list.append(d)
    if sort_filter == 'oldest':
        subsidized_list.sort(key=lambda x: x.donated_at)
    elif sort_filter == 'newest':
        subsidized_list.sort(key=lambda x: x.donated_at, reverse=True)
    elif sort_filter == 'least_qty':
        subsidized_list.sort(key=lambda x: x.remaining_subsidized_qty or 0)
    else:
        subsidized_list.sort(key=lambda x: x.remaining_subsidized_qty or 0, reverse=True)

    combined = item_list + money_list + subsidized_list
    request_unit_label = (_get_request_unit_label(req) or '').strip().lower()

    filtered_donations = []
    for donation in combined:
        target_qty, target_unit = _resolve_donation_target_quantity_and_unit(donation)
        donation.target_quantity = target_qty
        donation.target_unit_label = target_unit

        if donation.donation_type in ('money', 'subsidized') and request_unit_label:
            donation_unit = (target_unit or '').strip().lower()
            if not donation_unit or donation_unit != request_unit_label:
                continue

        filtered_donations.append(donation)

    valid_category_filters = {'item', 'money', 'subsidized'}
    if category_filter in valid_category_filters:
        filtered_donations = [
            donation for donation in filtered_donations
            if (getattr(donation, 'donation_type', '') or '').lower() == category_filter
        ]

    valid_type_filters = {'food', 'non_food'}
    if type_filter in valid_type_filters:
        def _normalize_type(value):
            return (value or '').strip().lower().replace('-', '_').replace(' ', '_')

        def _donation_type_key(donation):
            # Match the "Type" column semantics in the table.
            if donation.donation_type == 'money':
                source_req = getattr(donation, 'foodbank_request', None)
                source_type = _normalize_type(getattr(source_req, 'donation_type', None))
                if source_type in valid_type_filters:
                    return source_type
                original_req = getattr(source_req, 'original_request', None)
                original_type = _normalize_type(getattr(original_req, 'request_type', None))
                if original_type in valid_type_filters:
                    return original_type

            donation_category = _normalize_type(getattr(donation, 'donation_category', None))
            if donation_category in valid_type_filters:
                return donation_category

            request_type = _normalize_type(getattr(req, 'request_type', None))
            return request_type if request_type in valid_type_filters else ''

        filtered_donations = [
            donation for donation in filtered_donations
            if _donation_type_key(donation) == type_filter
        ]

    def _bucket_match(value, bucket):
        if value is None:
            return False
        if bucket == 'small':
            return value <= 50
        if bucket == 'medium':
            return 50 < value <= 200
        if bucket == 'large':
            return value > 200
        return True

    def _amount_bucket_match(value, bucket):
        if value is None:
            return False
        if bucket == 'small':
            return value <= 5000
        if bucket == 'medium':
            return 5000 < value <= 20000
        if bucket == 'large':
            return value > 20000
        return True

    def _safe_number(value):
        try:
            if value is None:
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    if quantity_filter in {'small', 'medium', 'large'}:
        def _donation_quantity_value(donation):
            if donation.donation_type == 'money':
                return _safe_number(
                    getattr(donation, 'remaining_supported_units', None)
                    or getattr(donation, 'remaining_qty', None)
                    or getattr(donation, 'target_quantity', None)
                )
            if donation.donation_type == 'subsidized':
                return _safe_number(
                    getattr(donation, 'remaining_supported_units', None)
                    or getattr(donation, 'remaining_subsidized_qty', None)
                    or getattr(donation, 'remaining_qty', None)
                    or getattr(donation, 'target_quantity', None)
                )
            return _safe_number(getattr(donation, 'remaining_qty', None) or getattr(donation, 'quantity', None))

        filtered_donations = [
            donation for donation in filtered_donations
            if _bucket_match(_donation_quantity_value(donation), quantity_filter)
        ]

    if amount_filter in {'small', 'medium', 'large'}:
        def _donation_amount_value(donation):
            if donation.donation_type in ('money', 'subsidized'):
                return _safe_number(
                    getattr(donation, 'remaining_amount', None)
                    or getattr(donation, 'amount', None)
                    or getattr(donation, 'subsidized_price', None)
                )
            return _safe_number(getattr(donation, 'amount', None))

        filtered_donations = [
            donation for donation in filtered_donations
            if _amount_bucket_match(_donation_amount_value(donation), amount_filter)
        ]

    if sort_filter == 'oldest':
        filtered_donations.sort(key=lambda x: x.donated_at)
    elif sort_filter == 'newest':
        filtered_donations.sort(key=lambda x: x.donated_at, reverse=True)

    available_donations, available_donations_count = _prepare_donation_status_metadata(filtered_donations)
    return available_donations, available_donations_count


@login_required

def fulfill_request(request, request_id):

    req = get_object_or_404(RequestManagement, id=request_id)

    current_foodbank = request.user.foodbank_profile

    

    # Handle both direct and anonymous requests

    if req.is_anonymous:

        if not req.assigned_foodbank:

            messages.error(request, "This anonymous request hasn't been assigned to your foodbank yet.")

            return _redirect_back_or_default(request, 'foodbank_requests_view')

        if req.assigned_foodbank != current_foodbank:

            messages.error(request, "This anonymous request is assigned to a different foodbank.")

            return _redirect_back_or_default(request, 'foodbank_requests_view')

        foodbank_to_search = req.assigned_foodbank

    else:

        if req.foodbank != current_foodbank:

            messages.error(request, "This request doesn't belong to your foodbank.")

            return _redirect_back_or_default(request, 'foodbank_requests_view')

        foodbank_to_search = req.foodbank

    

    if request.method == 'POST':

        option = request.POST.get('fulfillment_option')

        

        if option == 'available':

            available_donations, available_donations_count = _get_available_stock_donations(

                foodbank_to_search, req, request.GET, is_full_fulfillment=True

            )

            if not available_donations_count:

                messages.error(request, "No available donations match this request.")

                return _redirect_back_or_default(request, 'foodbank_requests_view')

            return render(request, 'authentication/select_donation.html', {

                'request_obj': req,

                'remaining': req.quantity,

                'available_donations': available_donations,

                'available_donations_count': available_donations_count,

                'is_full_fulfillment': True

            })

        

        elif option == 'new_request':

            remaining_quantity = req.quantity - req.quantity_fulfilled

            if remaining_quantity <= 0:

                messages.error(request, "This request is already fully fulfilled.")

                return _redirect_back_or_default(request, 'foodbank_requests_view')



            # Create donor request (your existing logic - updated for anonymous requests)

            donor_request = FoodBankRequest.objects.create(

                foodbank=foodbank_to_search,        # Use correct foodbank

                original_request=req,

                title=req.description,

                description=req.description or "Request from recipient",

                donation_type=req.request_type,

                quantity_needed=remaining_quantity,

                quantity_unit=req.unit,

                priority='high',

                status='active',

                deadline=timezone.now() + timedelta(days=30),

            )

            

            req.status = 'submitted'

            req.save()

            

            # Notify donors

            active_donors = CustomUser.objects.filter(user_type='DONOR', is_active=True)

            for donor in active_donors:

                Notification.objects.create(

                    user=donor,

                    notification_type='urgent_request',

                    message=f"New urgent request: {donor_request.title} from {foodbank_to_search.foodbank_name}."

                )

            

            messages.success(request, "Donor request created successfully.")

            return _redirect_back_or_default(request, 'foodbank_requests_view')

    

    return _redirect_back_or_default(request, 'foodbank_requests_view')

# views.py

@login_required

def partial_fulfill_request(request, request_id):

    """Show available donations for partial fulfillment"""

    req = get_object_or_404(RequestManagement, id=request_id)

    current_foodbank = request.user.foodbank_profile

    

    # Determine the correct foodbank to search donations for

    if req.is_anonymous:

        if not req.assigned_foodbank:

            messages.error(request, "This anonymous request hasn't been assigned to any foodbank yet.")

            return redirect('foodbank_requests_view')

        if req.assigned_foodbank != current_foodbank:

            messages.error(request, "This anonymous request is assigned to a different foodbank.")

            return redirect('foodbank_requests_view')

        foodbank_to_search = req.assigned_foodbank

    else:

        if req.foodbank != current_foodbank:

            messages.error(request, "This request doesn't belong to your foodbank.")

            return redirect('foodbank_requests_view')

        foodbank_to_search = req.foodbank

    

    # Check if already fulfilled

    remaining = req.quantity - req.quantity_fulfilled

    if remaining <= 0:

        messages.error(request, "This request is already fully fulfilled.")

        return redirect('foodbank_requests_view')

    available_donations, available_donations_count = _get_available_stock_donations(
        foodbank_to_search, req, request.GET, is_full_fulfillment=False
    )

    page_number = request.GET.get('page')
    paginator = Paginator(available_donations, 10)
    available_donations_page = paginator.get_page(page_number)

    search_query = (request.GET.get('search') or '').strip()
    delivery_filter = (request.GET.get('delivery') or '').strip().lower()
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    category_filter = (request.GET.get('category') or '').strip()
    type_filter = (request.GET.get('type') or '').strip()
    quantity_filter = (request.GET.get('quantity') or '').strip().lower()
    amount_filter = (request.GET.get('amount') or '').strip().lower()
    if quantity_filter == 'all':
        quantity_filter = ''
    if amount_filter == 'all':
        amount_filter = ''
    date_from = (request.GET.get('date_from') or '').strip()

    date_to = (request.GET.get('date_to') or '').strip()

    sort_filter = (request.GET.get('sort') or 'most_qty').strip()



    return render(request, 'authentication/select_donation.html', {

        'request_obj': req,

        'remaining': remaining,

        'available_donations': available_donations,
        'available_donations_page': available_donations_page,

        'available_donations_count': available_donations_count,

        'show_filters': True,

        'search_query': search_query,

        'delivery_filter': delivery_filter,

        'category_filter': category_filter,

        'type_filter': type_filter,
        'quantity_filter': quantity_filter,
        'amount_filter': amount_filter,

        'date_from': date_from,

        'date_to': date_to,

        'sort_filter': sort_filter,

        'delivery_choices': [
            ('pickup', 'Pickup'),
            ('delivery', 'Delivery'),
        ],

    })

# views.py

@login_required

def allocate_donation_to_request(request, donation_id):

    """Allocate donation to a SPECIFIC request (for partial fulfillment)"""

    if request.method == 'POST':

        donation = get_object_or_404(Donation, id=donation_id)

        current_foodbank = request.user.foodbank_profile

        

        try:

            request_id = int(request.POST.get('request_id'))

        except (ValueError, TypeError):

            messages.error(request, "Invalid request data.")

            return redirect('foodbank_requests_view')

        quantity_to_use = None

        amount_to_use = None

        if request.POST.get('quantity_to_use') not in (None, ''):

            try:

                quantity_to_use = int(request.POST.get('quantity_to_use'))

            except (ValueError, TypeError):

                pass

        if request.POST.get('amount_to_use') not in (None, ''):

            try:

                from decimal import Decimal

                amount_to_use = Decimal(request.POST.get('amount_to_use'))

            except (ValueError, TypeError, Exception):

                pass

        req = get_object_or_404(RequestManagement, id=request_id)

        unit_label = (
            _get_request_unit_label(req)
            or getattr(req, 'get_unit_display', lambda: '')()
            or getattr(req, 'custom_unit', '')
            or getattr(req, 'quantity_unit', '')
            or getattr(req, 'unit', '')
            or 'units'
        )

        remaining = req.quantity - req.quantity_fulfilled

        # For monetary and subsidized donations, do not allow partial fulfillment.
        # These should only be allocated if they can cover the full remaining requirement.
        if donation.donation_type in ('money', 'subsidized') and remaining > 0:
            if quantity_to_use is not None and quantity_to_use > 0 and quantity_to_use != remaining:
                messages.error(request, "Monetary/subsidized donations must fulfill the full remaining quantity (no partial fulfillment).")
                return redirect('foodbank_requests_view')

        if donation.foodbank != current_foodbank:

            messages.error(request, "Donation doesn't belong to your foodbank.")

            return redirect('foodbank_requests_view')

        if donation.donation_type == 'money' and donation.donation_category == 'monetary':
            pass
        elif donation.donation_category != req.request_type:
            messages.error(request, "Donation category doesn't match request type.")
            return redirect('foodbank_requests_view')

        if donation.donation_type == 'item':

            if quantity_to_use is None or quantity_to_use <= 0:

                messages.error(request, "Quantity must be greater than zero.")

                return redirect('foodbank_requests_view')

            if quantity_to_use > remaining:

                messages.error(request, f"Cannot fulfill more than {remaining} {req.get_unit_display()} remaining.")

                return redirect('foodbank_requests_view')

            if donation.quantity_unit != req.unit:

                messages.error(request, "Donation unit doesn't match request unit.")

                return redirect('foodbank_requests_view')

        elif donation.donation_type == 'money':

            from decimal import Decimal, InvalidOperation

            total_amt = Decimal(str(donation.amount or 0))
            donation_target_units = _get_request_quantity_value(getattr(donation, 'foodbank_request', None))
            remaining_supported_units = _get_donation_remaining_supported_units(donation)

            per_unit_amount = None
            if donation_target_units:
                try:
                    per_unit_amount = total_amt / Decimal(donation_target_units)
                except (InvalidOperation, ZeroDivisionError):
                    per_unit_amount = None

            if quantity_to_use is not None and quantity_to_use > 0:

                if per_unit_amount is None:
                    messages.error(request, "Cannot convert quantity to amount for this monetary donation.")
                    return redirect('foodbank_requests_view')

                if remaining_supported_units is not None and quantity_to_use > remaining_supported_units:
                    messages.error(request, "Cannot use more than the donation covers.")
                    return redirect('foodbank_requests_view')

                # Enforce full fulfillment for monetary donations
                if quantity_to_use != remaining:
                    messages.error(request, "Monetary donations cannot be used for partial fulfillment. Please fulfill the entire remaining quantity.")
                    return redirect('foodbank_requests_view')

                if quantity_to_use > remaining:
                    messages.error(request, f"Cannot fulfill more than {remaining} {req.get_unit_display()} remaining.")
                    return redirect('foodbank_requests_view')

                amount_to_use = per_unit_amount * Decimal(quantity_to_use)

                if amount_to_use > donation.get_remaining_amount():
                    messages.error(request, "Cannot use more than available donation amount.")
                    return redirect('foodbank_requests_view')

            elif amount_to_use is not None and amount_to_use > 0:

                if amount_to_use > donation.get_remaining_amount():
                    messages.error(request, "Cannot use more than available donation amount.")
                    return redirect('foodbank_requests_view')

                if per_unit_amount:
                    potential_units = int(amount_to_use / per_unit_amount)
                    if potential_units <= 0:
                        messages.error(request, "Amount entered is too small to cover any units from this donation.")
                        return redirect('foodbank_requests_view')
                    quantity_to_use = min(potential_units, remaining)
                    if remaining_supported_units is not None:
                        quantity_to_use = min(quantity_to_use, remaining_supported_units)
                else:
                    quantity_to_use = 1

                # Enforce full fulfillment for monetary donations
                if quantity_to_use != remaining:
                    messages.error(request, "Monetary donations cannot be used for partial fulfillment. Please fulfill the entire remaining quantity.")
                    return redirect('foodbank_requests_view')

            else:

                messages.error(request, f"Amount or {unit_label} must be greater than zero.")

                return redirect('foodbank_requests_view')

        elif donation.donation_type == 'subsidized':

            if quantity_to_use is None or quantity_to_use <= 0:

                messages.error(request, f"Quantity ({unit_label}) must be greater than zero.")

                return redirect('foodbank_requests_view')

            if quantity_to_use > remaining:

                messages.error(request, f"Cannot fulfill more than {remaining} {req.get_unit_display()} remaining.")

                return redirect('foodbank_requests_view')

            # Enforce full fulfillment for subsidized donations
            if quantity_to_use != remaining:
                messages.error(request, "Subsidized donations cannot be used for partial fulfillment. Please fulfill the entire remaining quantity.")
                return redirect('foodbank_requests_view')

            if quantity_to_use > donation.get_remaining_subsidized_quantity():

                messages.error(request, "Cannot use more than available subsidized quantity.")

                return redirect('foodbank_requests_view')

            total_qty = donation.subsidized_quantity or donation.quantity or 1

            total_amt = float(donation.subsidized_price or 0)

            amount_to_use = (total_amt / total_qty) * quantity_to_use if total_qty else 0

        else:

            messages.error(request, "Unsupported donation type for allocation.")

            return redirect('foodbank_requests_view')

        try:

            with transaction.atomic():

                req = RequestManagement.objects.select_for_update().get(id=req.id)

                remaining = req.quantity - req.quantity_fulfilled

                if quantity_to_use > remaining:

                    messages.error(request, f"Cannot fulfill more than {remaining} {req.get_unit_display()} remaining.")

                    return redirect('foodbank_requests_view')

                donation = Donation.objects.select_for_update().get(id=donation.id)

                if donation.donation_type == 'item':

                    if quantity_to_use > donation.get_remaining_quantity():

                        messages.error(request, "Cannot use more than available donation quantity.")

                        return redirect('foodbank_requests_view')

                elif donation.donation_type == 'money':

                    if amount_to_use > donation.get_remaining_amount():

                        messages.error(request, "Cannot use more than available donation amount.")

                        return redirect('foodbank_requests_view')

                elif donation.donation_type == 'subsidized':

                    if quantity_to_use > donation.get_remaining_subsidized_quantity():

                        messages.error(request, "Cannot use more than available subsidized quantity.")

                        return redirect('foodbank_requests_view')

                alloc_quantity = quantity_to_use if donation.donation_type in ('item', 'subsidized', 'money') else None

                alloc_amount = None

                if donation.donation_type == 'money':

                    alloc_amount = amount_to_use

                elif donation.donation_type == 'subsidized':

                    total_qty = donation.subsidized_quantity or donation.quantity or 1

                    total_amt = float(donation.subsidized_price or 0)

                    alloc_amount = (total_amt / total_qty) * quantity_to_use if total_qty else None

                DonationAllocation.objects.create(

                    donation=donation,

                    recipient=req.recipient,

                    request_management=req,

                    quantity=alloc_quantity,

                    amount=alloc_amount,

                    is_acknowledged=False,

                )

                donation.is_allocated = donation.is_fully_allocated()

                donation.save(update_fields=['is_allocated'])

                req.quantity_fulfilled += quantity_to_use

                if req.quantity_fulfilled >= req.quantity:

                    req.acknowledged_by_recipient = False

                    req.status = 'awaiting_recipient'

                else:

                    req.acknowledged_by_recipient = False

                    if req.status != 'awaiting_recipient':

                        req.status = 'partial'

                req.save()

                if req.status == 'awaiting_recipient':

                    Notification.objects.create(

                        user=req.recipient.user,

                        notification_type='request',

                        message=f"Your request '{req.description[:50]}...' has been fulfilled from available stock and is awaiting your acceptance."

                    )

        except Donation.DoesNotExist:

            messages.error(request, "Donation not found.")

            return redirect('foodbank_requests_view')

        if donation.donation_type == 'money':

            if quantity_to_use:

                messages.success(request, f"Successfully allocated KES {amount_to_use:,.0f} (for {quantity_to_use} {req.get_unit_display()}) for {req.recipient_name}.")

            else:

                messages.success(request, f"Successfully allocated KES {amount_to_use:,.0f} for {req.recipient_name}.")

        elif donation.donation_type == 'subsidized':

            messages.success(
                request,
                f"Successfully allocated KES {amount_to_use:,.0f} (for {quantity_to_use} {unit_label}) for {req.recipient_name}."
            )

        else:

            messages.success(request, f"Successfully fulfilled {quantity_to_use} {req.get_unit_display()} for {req.recipient_name}.")

        return redirect('foodbank_requests_view')

    

    # Don't allow GET requests - redirect

    return redirect('foodbank_requests_view')



@login_required

def fulfill_request_rest(request, request_id):

    """Fulfill the remaining quantity of a partially fulfilled request"""

    req = get_object_or_404(RequestManagement, id=request_id)

    current_foodbank = request.user.foodbank_profile

    

    # Handle foodbank assignment (same as before)

    if req.is_anonymous:

        if not req.assigned_foodbank or req.assigned_foodbank != current_foodbank:

            messages.error(request, "This request is not assigned to your foodbank.")

            return _redirect_back_or_default(request, 'foodbank_requests_view')

        foodbank_to_search = req.assigned_foodbank

    else:

        if req.foodbank != current_foodbank:

            messages.error(request, "This request doesn't belong to your foodbank.")

            return _redirect_back_or_default(request, 'foodbank_requests_view')

        foodbank_to_search = req.foodbank

    

    remaining = req.quantity - req.quantity_fulfilled

    if remaining <= 0:

        messages.error(request, "This request is already fully fulfilled.")

        return redirect('foodbank_requests_view')

    available_donations, available_donations_count = _get_available_stock_donations(

        foodbank_to_search, req, request.GET, is_full_fulfillment=False

    )

    if not available_donations_count:

        messages.error(request, f"No donations available to fulfill the remaining {remaining} {req.get_unit_display()}.")

        return redirect('foodbank_requests_view')

    return render(request, 'authentication/select_donation.html', {

        'request_obj': req,

        'remaining': remaining,

        'available_donations': available_donations,

        'available_donations_count': available_donations_count,

        'is_full_fulfillment': True  # This will show "Fulfill Rest" button

    })



@login_required

def decline_request(request, request_id):

    req = get_object_or_404(RequestManagement, id=request_id)



    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank privileges required.')

        return redirect('dashboard')



    current_foodbank = request.user.foodbank_profile



    # Ensure the request belongs to this foodbank (direct or assigned anonymous)

    if req.is_anonymous:

        if not req.assigned_foodbank or req.assigned_foodbank != current_foodbank:

            messages.error(request, "This anonymous request is not assigned to your foodbank.")

            return redirect('foodbank_requests_view')

    else:

        if req.foodbank != current_foodbank:

            messages.error(request, "This request doesn't belong to your foodbank.")

            return redirect('foodbank_requests_view')



    if request.method == 'POST':

        decline_message = (request.POST.get('decline_message') or '').strip()



        if not decline_message:

            messages.error(request, 'Please provide a reason for declining.')

            return _redirect_back_or_default(request, 'foodbank_requests_view')



        # Update the request status

        req.status = 'declined'

        req.decline_message = decline_message

        req.updated_by = request.user

        req.save()



        # Notify the recipient

        Notification.objects.create(

            user=req.recipient.user,

            notification_type='request',

            message=f"Your request '{req.description}' has been declined. Reason: {decline_message}"

        )



        messages.warning(request, "Request declined with a message.")

        return _redirect_back_or_default(request, 'foodbank_requests_view')



    return _redirect_back_or_default(request, 'foodbank_requests_view')



from django.shortcuts import get_object_or_404, redirect

from django.contrib import messages

from .models import RequestManagement, Donation



@login_required

def respond_to_request(request, request_id):

    req = get_object_or_404(RequestManagement, id=request_id)



    if request.method == 'POST':

        donated_qty = int(request.POST.get('quantity'))

        remaining = req.quantity - req.quantity_fulfilled



        actual_applied_qty = min(donated_qty, remaining)

        extra_qty = max(0, donated_qty - remaining)

        

        donor_message = request.POST.get('message', '')



        donation = Donation.objects.create(

            donor=request.user,

            foodbank=req.foodbank,

            donation_type=request.POST.get('donation_type'),

            quantity=donated_qty,

            quantity_unit=request.POST.get('quantity_unit'),

            message=donor_message,

            delivery_status='pending',

            request_management=req,

        )



        req.quantity_fulfilled += actual_applied_qty



        if req.quantity_fulfilled >= req.quantity:

            req.status = 'fulfilled'

            req.awaiting_donors = False

            req.fulfilled_at = timezone.now()

        else:

            req.status = 'partial'

            req.awaiting_donors = True

        

        # Append donor note to request's additional_notes

        if donor_message.strip():

            donor_name = request.user.donor_profile.full_name if hasattr(request.user, 'donor_profile') else request.user.email

            timestamp = timezone.localtime().strftime("%Y-%m-%d %H:%M")

            current_notes = req.additional_notes or ""

            req.additional_notes = current_notes + f"\n\n--- Donor Note ({donor_name}, {timestamp}) ---\n{donor_message}"



        req.save()



        if extra_qty > 0:

            messages.info(

                request,

                f"Thank you! {extra_qty} extra {req.get_unit_display()} will be visible as additional donation."

            )

        else:

            messages.success(request, "Thank you for responding to the request!")



        return redirect('donor_dashboard')



    return redirect('donor_dashboard')



# Excel Export for Recipient Requests

import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from openpyxl.utils import get_column_letter



@login_required

def download_recipient_requests_excel(request):

    """Download all recipient requests as Excel file"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Only recipients can download their request data.')

        return redirect('dashboard')

    

    # Get all requests for this recipient

    requests = RecipientRequest.objects.filter(recipient=request.user.recipient_profile).order_by('-created_at')

    

    # Create workbook and worksheet

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "My Requests"

    

    # Define styles

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=12)

    border = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )

    

    # Headers

    headers = [

        'Request ID', 'Title', 'Description', 'Food Bank', 'Quantity', 'Unit',

        'Status', 'Delivery Method', 'Location', 'Created Date', 'Is Anonymous', 'Fulfillment Notes'

    ]

    

    # Write headers

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col_num)

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell.border = border

    

    # Write data

    for row_num, req in enumerate(requests, 2):

        data = [

            req.id,

            req.title,

            req.description,

            req.foodbank.foodbank_name if req.foodbank else 'Not Assigned',

            req.quantity,

            req.quantity_unit if req.quantity_unit else 'N/A',

            req.get_status_display(),

            req.delivery_method if req.delivery_method else 'N/A',

            req.location if req.location else 'N/A',

            req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else '',

            'Yes' if req.is_anonymous else 'No',

            req.fulfillment_notes if req.fulfillment_notes else ''

        ]

        

        for col_num, value in enumerate(data, 1):

            cell = ws.cell(row=row_num, column=col_num)

            cell.value = value

            cell.border = border

            cell.alignment = Alignment(vertical='center', wrap_text=True)

    

    # Auto-adjust column widths

    for col_num in range(1, len(headers) + 1):

        column_letter = get_column_letter(col_num)

        max_length = 0

        for cell in ws[column_letter]:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:

                pass

        adjusted_width = min(max_length + 2, 50)

        ws.column_dimensions[column_letter].width = adjusted_width

    

    # Prepare response

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    

    # Generate filename with timestamp

    filename = f"my_requests_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    

    # Save workbook to response

    wb.save(response)

    

    return response



@login_required

def assign_anonymous_request(request, request_id):

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank privileges required.')

        return redirect('dashboard')



    if request.method != 'POST':

        return _redirect_back_or_default(request, 'foodbank_requests_view')



    req = get_object_or_404(

        RequestManagement, 

        id=request_id, 

        is_anonymous=True,

        assigned_foodbank__isnull=True

    )

    req.assigned_foodbank = request.user.foodbank_profile

    req.status = 'assigned'  # Add this status to your STATUS_CHOICES

    req.save()

    

    # Check if it's an AJAX request

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

        return JsonResponse({

            'success': True,

            'message': 'Anonymous request assigned to you successfully!'

        })

    else:

        messages.success(request, "Anonymous request assigned to you.")

        return _redirect_back_or_default(request, 'foodbank_requests_view')

    

# In your view

# In your view

@login_required
def select_foodbank_for_donation(request):

    if request.user.user_type != 'DONOR':
        messages.error(request, 'Only donors can make donations.')
        return redirect('dashboard')

    foodbanks = FoodBankProfile.objects.select_related('user').filter(
        is_approved='approved',
        user__is_active=True
    )

    # Service type filter
    service_type = (request.GET.get('service_type') or '').strip()
    if service_type in {'food', 'non_food', 'both'}:
        foodbanks = foodbanks.filter(service_type=service_type)

    # Location search
    location = (request.GET.get('location') or '').strip()
    if location:
        foodbanks = foodbanks.filter(address__icontains=location)

    # Food bank name search
    foodbank_name = (request.GET.get('foodbank_name') or '').strip()
    if foodbank_name:
        foodbanks = foodbanks.filter(foodbank_name__icontains=foodbank_name)

    # Date joined range filter
    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()
    date_from_value = None
    date_to_value = None

    if date_from_raw:
        try:
            date_from_value = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
        except ValueError:
            messages.warning(request, 'Invalid "From" date. Please use YYYY-MM-DD.')

    if date_to_raw:
        try:
            date_to_value = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
        except ValueError:
            messages.warning(request, 'Invalid "To" date. Please use YYYY-MM-DD.')

    if date_from_value and date_to_value and date_from_value > date_to_value:
        messages.warning(request, '"From" date cannot be later than "To" date.')
    else:
        if date_from_value:
            foodbanks = foodbanks.filter(user__date_joined__date__gte=date_from_value)
        if date_to_value:
            foodbanks = foodbanks.filter(user__date_joined__date__lte=date_to_value)

    # Sorting
    sort = (request.GET.get('sort') or 'newest').strip()
    if sort == 'oldest':
        foodbanks = foodbanks.order_by('user__date_joined', 'foodbank_name')
    else:
        foodbanks = foodbanks.order_by('-user__date_joined', 'foodbank_name')

    context = {
        'foodbanks': foodbanks,
    }
    return render(request, 'donor/select_foodbank.html', context)

from django.contrib import messages

from django.shortcuts import render, redirect

from django.contrib.auth.decorators import login_required



@login_required

def request_account_deletion(request):

    """Request account deletion - requires admin approval"""

    if request.method == 'POST':

        # Check if user already has a pending request

        existing_request = AccountDeletionRequest.objects.filter(

            user=request.user,

            status='pending'

        ).first()

        

        if existing_request:

            messages.warning(request, "You already have a pending deletion request.")

        else:

            # Create deletion request

            AccountDeletionRequest.objects.create(user=request.user)

            

            # Create notification for admins

            admin_users = CustomUser.objects.filter(is_superuser=True)

            for admin in admin_users:

                Notification.objects.create(

                    user=admin,

                    notification_type='account_deletion_request',

                    message=f"Account deletion request from {request.user.email}. Please review in admin panel."

                )

            

            messages.success(request, "Your account deletion request has been submitted. An administrator will review your request.")

        

        return redirect('dashboard')



    deletion_request = AccountDeletionRequest.objects.filter(

        user=request.user

    ).order_by('-requested_at').first()



    return render(request, 'donor/request_deletion.html', {

        'deletion_request': deletion_request,

    })



# views.py

# authentication/views.py

# authentication/views.py

# views.py

@login_required

def recipient_subsidized_donations(request):

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Access denied.")

        return redirect('dashboard')



    recipient = request.user.recipient_profile



    # Get filter parameters

    delivery_filter = request.GET.get('delivery', 'all')

    status_filter = request.GET.get('status', 'all')

    date_filter = request.GET.get('date_range', 'all')

    legacy_category_param = request.GET.get('category', 'all').strip()
    type_filter = request.GET.get('type', 'all').strip()
    if type_filter == 'all' and legacy_category_param in ['food', 'non_food', 'monetary']:
        type_filter = legacy_category_param

    donation_type_filter = request.GET.get('donation_type', '').strip()
    if legacy_category_param in ['subsidized', 'csr', 'other']:
        donation_type_filter = legacy_category_param
    if donation_type_filter not in ['subsidized', 'csr', 'other']:
        donation_type_filter = 'subsidized'

    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')

    search_query = request.GET.get('search', '').strip()

    sort_filter = request.GET.get('sort', 'newest')



    # Base query for available donations

    donations_query = Q(status='accepted')



    # Filter by donation type

    if donation_type_filter == 'subsidized':

        donations_query &= Q(donation_type='subsidized')

        # Only show unspecified subsidized donations (donor-initiated, not linked to a request)

        donations_query &= Q(foodbank_request__isnull=True)

        # Subsidized donations are available to all recipients

        donations_query &= (

            Q(accepted_by_recipient__isnull=True) |  # Not accepted by anyone

            Q(accepted_by_recipient=recipient)       # Accepted by this recipient

        )

    elif donation_type_filter in ['csr', 'other']:

        # CSR and Other donations are only available to organizations

        if not getattr(recipient, 'is_organization', False):

            # If not an organization, show empty queryset

            donations_query &= Q(pk__in=[])

        else:

            donations_query &= Q(donation_type=donation_type_filter)

            # These donations are shown to organizations for their review/acceptance

            donations_query &= Q(discussion_status='agreed')  # Only show agreed discussions



    available_donations = Donation.objects.filter(donations_query).select_related(

        'foodbank',

        'donor',

        'donor__donor_profile',

        'accepted_by_recipient'

    ).order_by('-donated_at', '-id')



    # Apply filters

    if delivery_filter != 'all':
        if delivery_filter == 'delivery':
            available_donations = available_donations.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            available_donations = available_donations.filter(delivery_method='pickup')



    if status_filter != 'all':

        if status_filter == 'available':

            available_donations = available_donations.filter(accepted_by_recipient__isnull=True).exclude(
                declined_by_recipient=recipient
            )

        elif status_filter == 'accepted':

            available_donations = available_donations.filter(accepted_by_recipient=recipient).exclude(
                delivery_status='delivered'
            )

        elif status_filter == 'received':

            available_donations = available_donations.filter(accepted_by_recipient=recipient, delivery_status='delivered')
        elif status_filter == 'declined':
            available_donations = available_donations.filter(declined_by_recipient=recipient)



    # Apply date range filter

    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()

    

    if date_filter == 'custom' and (date_from or date_to):

        # Custom date range

        if date_from:

            try:

                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')

                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))

                available_donations = available_donations.filter(donated_at__gte=from_date)

            except ValueError:

                pass

        if date_to:

            try:

                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')

                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))

                available_donations = available_donations.filter(donated_at__lte=to_date)

            except ValueError:

                pass

    elif date_filter != 'all':

        now = timezone.now()

        if date_filter == 'today':

            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)

            available_donations = available_donations.filter(donated_at__gte=start_date)

        elif date_filter == 'week':

            start_date = now - timedelta(days=7)

            available_donations = available_donations.filter(donated_at__gte=start_date)

        elif date_filter == 'month':

            start_date = now - timedelta(days=30)

            available_donations = available_donations.filter(donated_at__gte=start_date)

        elif date_filter == '3months':

            start_date = now - timedelta(days=90)

            available_donations = available_donations.filter(donated_at__gte=start_date)



    # Apply type filter (matches Type column -> donation_category)
    if type_filter != 'all':

        available_donations = available_donations.filter(donation_category=type_filter)



    # Apply quantity filter (for subsidized donations)

    if quantity_filter != 'all' and donation_type_filter == 'subsidized':

        if quantity_filter == 'small':

            available_donations = available_donations.filter(subsidized_quantity__lte=100)

        elif quantity_filter == 'medium':

            available_donations = available_donations.filter(subsidized_quantity__gt=100, subsidized_quantity__lte=500)

        elif quantity_filter == 'large':

            available_donations = available_donations.filter(subsidized_quantity__gt=500)

    # Apply amount filter (for subsidized donations new price)
    if amount_filter != 'all' and donation_type_filter == 'subsidized':

        if amount_filter == 'small':

            available_donations = available_donations.filter(subsidized_price__lte=100)

        elif amount_filter == 'medium':

            available_donations = available_donations.filter(subsidized_price__gt=100, subsidized_price__lte=500)

        elif amount_filter == 'large':

            available_donations = available_donations.filter(subsidized_price__gt=500)



    if search_query:

        if donation_type_filter == 'subsidized':

            available_donations = available_donations.filter(

                Q(subsidized_product_type__icontains=search_query) |

                Q(csr_description__icontains=search_query) |

                Q(message__icontains=search_query) |
                Q(foodbank__foodbank_name__icontains=search_query) |
                Q(foodbank__address__icontains=search_query)

            )

        else:

            available_donations = available_donations.filter(

                Q(other_description__icontains=search_query) |

                Q(csr_description__icontains=search_query) |

                Q(message__icontains=search_query) |
                Q(foodbank__foodbank_name__icontains=search_query) |
                Q(foodbank__address__icontains=search_query)

            )



    # Sort

    if sort_filter == 'oldest':

        available_donations = available_donations.order_by('donated_at', 'id')

    else:

        available_donations = available_donations.order_by('-donated_at', '-id')



    # Get donation IDs this recipient has responded to

    responded_donation_ids = DonationResponse.objects.filter(

        recipient=recipient

    ).values_list('donation_id', flat=True)



    # Get counts for summary stats

    available_count = Donation.objects.filter(

        donation_type=donation_type_filter,

        status='accepted',

        accepted_by_recipient__isnull=True

    ).exclude(
        declined_by_recipient=recipient
    ).count()

    

    accepted_count = Donation.objects.filter(

        donation_type=donation_type_filter,

        status='accepted',

        accepted_by_recipient=recipient

    ).exclude(delivery_status='delivered').count()

    

    received_count = Donation.objects.filter(

        donation_type=donation_type_filter,

        status='accepted',

        accepted_by_recipient=recipient,

        delivery_status='delivered'

    ).count()



    # Pagination

    paginator = Paginator(available_donations, 10)  # 10 items per page

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)



    # Attach latest recipient notes for displayed donations

    donation_list = list(page_obj.object_list)

    donation_ids = [donation.id for donation in donation_list]

    latest_notes = {}



    if donation_ids:

        responses = DonationResponse.objects.filter(

            donation_id__in=donation_ids

        ).exclude(

            notes__isnull=True

        ).exclude(

            notes__exact=''

        ).order_by('-responded_at')



        for response in responses:

            if response.donation_id not in latest_notes:

                latest_notes[response.donation_id] = response.notes

    # Decline reasons for donations this recipient declined (for "View reason" button)
    decline_reasons = {}
    if donation_ids:
        declined_responses = DonationResponse.objects.filter(
            donation_id__in=donation_ids,
            recipient=recipient,
            response_type='declined'
        ).values_list('donation_id', 'notes')
        for did, notes in declined_responses:
            decline_reasons[did] = notes or ''

    for donation in donation_list:

        donation.latest_recipient_note = latest_notes.get(donation.id)
        donation.recipient_decline_reason = decline_reasons.get(donation.id, '')

    category_filter = donation_type_filter

    context = {

        'donations': page_obj,

        'page_obj': page_obj,

        'responded_donation_ids': list(responded_donation_ids),

        'current_recipient': recipient,

        'delivery_filter': delivery_filter,

        'status_filter': status_filter,

        'date_filter': date_filter,

        'date_from': date_from,

        'date_to': date_to,

        'category_filter': category_filter,

        'type_filter': type_filter,

        'quantity_filter': quantity_filter,

        'amount_filter': amount_filter,

        'search_query': search_query,

        'sort_filter': sort_filter,

        'donation_type_filter': donation_type_filter,

        'is_organization': getattr(recipient, 'is_organization', False),

        'available_count': available_count,

        'accepted_count': accepted_count,

        'received_count': received_count,

        'total_count': available_count + accepted_count + received_count,

    }

    return render(request, 'recipient/subsidized_donations.html', context)



@login_required

def export_subsidized_donations(request, format):

    """Unified export handler for subsidized donations - supports PDF, CSV, and Excel.

    Mirrors ALL filters from recipient_subsidized_donations view."""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Access denied.")

        return redirect('dashboard')



    recipient = request.user.recipient_profile



    # â”€â”€ Read all filter params (same as recipient_subsidized_donations) â”€â”€

    legacy_category_param = request.GET.get('category', 'all').strip()
    type_filter = request.GET.get('type', 'all').strip()
    if type_filter == 'all' and legacy_category_param in ['food', 'non_food', 'monetary']:
        type_filter = legacy_category_param

    donation_type_filter = request.GET.get('donation_type', '').strip()
    if legacy_category_param in ['subsidized', 'csr', 'other']:
        donation_type_filter = legacy_category_param
    if donation_type_filter not in ['subsidized', 'csr', 'other']:
        donation_type_filter = 'subsidized'

    delivery_filter = request.GET.get('delivery', 'all')

    status_filter = request.GET.get('status', 'all')

    date_filter = request.GET.get('date_range', 'all')

    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')

    search_query = request.GET.get('search', '').strip()



    # Base query

    donations_query = Q(status='accepted')



    if donation_type_filter == 'subsidized':

        donations_query &= Q(donation_type='subsidized')

        donations_query &= Q(foodbank_request__isnull=True)

        donations_query &= (

            Q(accepted_by_recipient__isnull=True) |

            Q(accepted_by_recipient=recipient)

        )

    elif donation_type_filter in ['csr', 'other']:

        if not getattr(recipient, 'is_organization', False):

            donations_query &= Q(pk__in=[])

        else:

            donations_query &= Q(donation_type=donation_type_filter)

            donations_query &= Q(discussion_status='agreed')



    donations = Donation.objects.filter(donations_query).select_related(

        'foodbank', 'donor', 'donor__donor_profile', 'accepted_by_recipient'

    ).order_by('-donated_at', '-id')



    # Apply simple filters

    if delivery_filter != 'all':
        if delivery_filter == 'delivery':
            donations = donations.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            donations = donations.filter(delivery_method='pickup')

    if status_filter != 'all':

        if status_filter == 'available':

            donations = donations.filter(accepted_by_recipient__isnull=True).exclude(
                declined_by_recipient=recipient
            )

        elif status_filter == 'accepted':

            donations = donations.filter(accepted_by_recipient=recipient).exclude(
                delivery_status='delivered'
            )

        elif status_filter == 'received':

            donations = donations.filter(accepted_by_recipient=recipient, delivery_status='delivered')
        elif status_filter == 'declined':
            donations = donations.filter(declined_by_recipient=recipient)



    # Date filter

    date_from = request.GET.get('date_from', '').strip()

    date_to = request.GET.get('date_to', '').strip()



    if date_filter == 'custom' and (date_from or date_to):

        if date_from:

            try:

                from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d')

                from_date = timezone.make_aware(from_date.replace(hour=0, minute=0, second=0, microsecond=0))

                donations = donations.filter(donated_at__gte=from_date)

            except ValueError:

                pass

        if date_to:

            try:

                to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d')

                to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59, microsecond=999999))

                donations = donations.filter(donated_at__lte=to_date)

            except ValueError:

                pass

    elif date_filter != 'all':

        now = timezone.now()

        if date_filter == 'today':

            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)

            donations = donations.filter(donated_at__gte=start_date)

        elif date_filter == 'week':

            start_date = now - timedelta(days=7)

            donations = donations.filter(donated_at__gte=start_date)

        elif date_filter == 'month':

            start_date = now - timedelta(days=30)

            donations = donations.filter(donated_at__gte=start_date)

        elif date_filter == '3months':

            start_date = now - timedelta(days=90)

            donations = donations.filter(donated_at__gte=start_date)



    # Type filter (matches Type column -> donation_category)
    if type_filter != 'all':

        donations = donations.filter(donation_category=type_filter)



    # Quantity filter (subsidized only)

    if quantity_filter != 'all' and donation_type_filter == 'subsidized':

        if quantity_filter == 'small':

            donations = donations.filter(subsidized_quantity__lte=100)

        elif quantity_filter == 'medium':

            donations = donations.filter(subsidized_quantity__gt=100, subsidized_quantity__lte=500)

        elif quantity_filter == 'large':

            donations = donations.filter(subsidized_quantity__gt=500)

    # Amount filter (subsidized only - new price)
    if amount_filter != 'all' and donation_type_filter == 'subsidized':

        if amount_filter == 'small':

            donations = donations.filter(subsidized_price__lte=100)

        elif amount_filter == 'medium':

            donations = donations.filter(subsidized_price__gt=100, subsidized_price__lte=500)

        elif amount_filter == 'large':

            donations = donations.filter(subsidized_price__gt=500)



    # Search

    if search_query:

        if donation_type_filter == 'subsidized':

            donations = donations.filter(

                Q(subsidized_product_type__icontains=search_query) |

                Q(csr_description__icontains=search_query) |

                Q(message__icontains=search_query) |
                Q(foodbank__foodbank_name__icontains=search_query) |
                Q(foodbank__address__icontains=search_query)

            )

        else:

            donations = donations.filter(

                Q(other_description__icontains=search_query) |

                Q(csr_description__icontains=search_query) |

                Q(message__icontains=search_query) |
                Q(foodbank__foodbank_name__icontains=search_query) |
                Q(foodbank__address__icontains=search_query)

            )



    sort_filter = request.GET.get('sort', 'newest')

    if sort_filter == 'oldest':

        donations = donations.order_by('donated_at', 'id')

    else:

        donations = donations.order_by('-donated_at', '-id')



    donations_data = list(donations)

    # Fetch latest recipient notes from DonationResponse
    donation_ids = [d.id for d in donations_data]
    latest_recipient_notes = {}
    if donation_ids:
        responses = (
            DonationResponse.objects
            .filter(donation_id__in=donation_ids)
            .exclude(notes__isnull=True)
            .exclude(notes__exact='')
            .order_by('-responded_at')
        )
        for resp in responses:
            if resp.donation_id not in latest_recipient_notes:
                latest_recipient_notes[resp.donation_id] = resp.notes

    if format.lower() == 'pdf':

        return export_subsidized_donations_pdf_helper(request, donations_data, recipient, donation_type_filter, latest_recipient_notes)

    elif format.lower() == 'csv':

        return export_subsidized_donations_csv_helper(request, donations_data, recipient, donation_type_filter, latest_recipient_notes)

    elif format.lower() == 'excel':

        return export_subsidized_donations_excel_helper(request, donations_data, recipient, donation_type_filter, latest_recipient_notes)

    else:

        messages.error(request, 'Invalid export format.')

        return redirect('recipient_subsidized_donations')





def _get_subsidized_status_display(donation, recipient):

    """Shared helper: human-readable status for a subsidized/CSR/Other donation."""

    if donation.accepted_by_recipient:

        if donation.accepted_by_recipient == recipient:

            return 'Received by recipient' if donation.delivery_status == 'delivered' else 'Accepted by recipient'

        else:

            return 'Delivered' if donation.delivery_status == 'delivered' else 'Accepted by recipient'

    elif donation.declined_by_recipient == recipient:

        return 'Declined by recipient'

    return 'Available for recipients'





def export_subsidized_donations_csv_helper(request, donations_data, recipient, donation_type_filter, latest_recipient_notes=None):

    """Generate CSV report for subsidized donations.

    Columns match the table: S/NO, Food Bank, Type, Category, Description,

    Qty, Unit, Market Price, Subsidy, New Price, Status, Delivery, Donor Note, Recipient Note, Decline Note

    """

    if latest_recipient_notes is None:
        latest_recipient_notes = {}

    import csv

    import datetime

    from django.http import HttpResponse



    recipient_name = recipient.full_name or recipient.user.email

    response = HttpResponse(content_type='text/csv')

    response['Content-Disposition'] = (

        f'attachment; filename="{recipient_name}_{donation_type_filter}_donations_'

        f'{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.csv"'

    )



    writer = csv.writer(response, quoting=csv.QUOTE_ALL, lineterminator='\r\n')

    donation_ids = [d.id for d in donations_data]
    decline_notes = {}
    if donation_ids:
        declined_responses = (
            DonationResponse.objects
            .filter(
                donation_id__in=donation_ids,
                recipient=recipient,
                response_type='declined',
            )
            .exclude(notes__isnull=True)
            .exclude(notes__exact='')
            .order_by('-responded_at')
            .values_list('donation_id', 'notes')
        )
        for did, note in declined_responses:
            if did not in decline_notes:
                decline_notes[did] = note or ''



    # Metadata

    writer.writerow(['Recipient', recipient_name])

    writer.writerow(['Report Generated', timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')])

    writer.writerow(['Total Records', len(donations_data)])

    writer.writerow([])



    # Headers â€“ matching table columns with separate price columns for calculations

    if donation_type_filter == "subsidized":

        writer.writerow([

            '#', 'Date', 'Food Bank', 'Donor', 'Type', 'Category', 'Product',

            'Quantity', 'Market Price (KES)', 'Subsidy (KES)', 'New Price (KES)',

            'Status', 'Delivery', 'Location', 'Donor Note', 'Recipient Note', 'Decline Note',

        ])

    else:

        writer.writerow([

            '#', 'Food Bank', 'Type', 'Category', 'Description',

            'Delivery', 'Status', 'Decline Note', 'Date Added',

        ])



    for idx, donation in enumerate(donations_data, 1):

        status = _get_subsidized_status_display(donation, recipient)



        if donation_type_filter == "subsidized":

            market_price = donation.subsidized_initial_amount or donation.subsidized_market_price or ''

            subsidy = donation.subsidized_subsidy_amount or ''

            new_price = donation.subsidized_price or ''

            type_label = donation.get_donation_category_display() or 'Food'
            category_label = donation.get_donation_type_display() or (donation.donation_type.title() if donation.donation_type else 'Subsidized')
            product = (
                donation.subsidized_product_type
                or donation.csr_description
                or donation.other_description
                or donation.message
                or 'No description'
            )
            qty_value = donation.subsidized_quantity if donation.subsidized_quantity is not None else donation.quantity
            qty_unit = donation.subsidized_quantity_unit or donation.quantity_unit or 'units'
            quantity_display = f"{qty_value} {qty_unit}" if qty_value is not None else '-'
            donor_profile = getattr(donation.donor, 'donor_profile', None) if donation.donor else None
            donor_name = '-'
            if donor_profile:
                donor_name = (
                    (donor_profile.organization_name or '').strip()
                    if getattr(donor_profile, 'is_organization', False) and (donor_profile.organization_name or '').strip()
                    else (donor_profile.full_name or '').strip()
                ) or (getattr(donation.donor, 'email', '') or '').strip() or '-'
            elif donation.donor:
                donor_name = (donation.donor.email or '').strip() or '-'
            location = (
                (getattr(donation.foodbank_request, 'location', None) or '').strip()
                or (getattr(donation.foodbank, 'address', None) or '').strip()
                or 'Not provided'
            )



            donor_note = (donation.message or donation.csr_description or donation.other_description or '').replace('\n', ' | ').replace('\r', '')
            recipient_note = (latest_recipient_notes.get(donation.id) or '').replace('\n', ' | ').replace('\r', '')
            decline_note = (
                decline_notes.get(donation.id)
                or (getattr(donation, 'decline_message', None) or '').strip()
                or 'No decline note'
            ).replace('\n', ' | ').replace('\r', '')

            writer.writerow([

                idx,
                donation.donated_at.strftime('%Y-%m-%d') if donation.donated_at else '',

                donation.foodbank.foodbank_name,
                donor_name,

                type_label,

                category_label,

                product,

                quantity_display,

                market_price,

                subsidy,

                new_price,

                status,

                donation.get_delivery_method_display() if donation.delivery_method else '',
                location,

                donor_note,

                recipient_note,

                decline_note,

            ])

        else:

            desc = (donation.csr_description or donation.other_description

                    or donation.message or 'No description')

            cat = 'CSR' if donation.donation_type == 'csr' else 'Other'

            writer.writerow([

                idx,

                donation.foodbank.foodbank_name,

                donation.donation_type.title(),

                cat,

                desc,

                donation.get_delivery_method_display() if donation.delivery_method else '',

                status,

                (
                    decline_notes.get(donation.id)
                    or (getattr(donation, 'decline_message', None) or '').strip()
                    or 'No decline note'
                ).replace('\n', ' | ').replace('\r', ''),

                donation.donated_at.strftime('%Y-%m-%d'),

            ])



    # Summary

    writer.writerow([])

    writer.writerow(['Summary'])

    writer.writerow(['Total Donations', len(donations_data)])

    writer.writerow(['Available', sum(1 for d in donations_data if not d.accepted_by_recipient)])

    writer.writerow(['Accepted', sum(1 for d in donations_data if d.accepted_by_recipient == recipient and d.delivery_status != 'delivered')])

    writer.writerow(['Received', sum(1 for d in donations_data if d.accepted_by_recipient == recipient and d.delivery_status == 'delivered')])



    return response





def export_subsidized_donations_pdf_helper(request, donations_data, recipient, donation_type_filter, latest_recipient_notes=None):

    """Generate branded PDF report for subsidized / CSR / Other donations.

    Columns match the table: S/NO, Food Bank, Type, Category, Description,

    Quantity (combined), Market Price, Subsidy, New Price, Status, Delivery,
    Donor Note, Recipient Note, Decline Comment

    """

    if latest_recipient_notes is None:
        latest_recipient_notes = {}

    from reportlab.lib.pagesizes import A3, landscape

    from reportlab.platypus import Paragraph

    from .report_utils import (

        get_report_styles, build_report_header, get_branded_table_style,

        build_report_summary, build_pdf_document, collect_active_filters, make_full_width_table,

    )



    styles = get_report_styles()
    report_pagesize = landscape(A3)

    wrap = styles['wrap']

    elements = []



    recipient_name = recipient.full_name or recipient.user.email



    active_filters = collect_active_filters(request, [

        ('donation_type', 'Type'),

        ('delivery', 'Delivery'), ('status', 'Status'),

        ('category', 'Category'), ('quantity', 'Quantity'),

        ('amount', 'New Price Range'),

        ('date_range', 'Date Range'), ('search', 'Search'),

    ])



    report_title = f"{donation_type_filter.title()} Donations Report"

    build_report_header(

        elements, report_title, recipient_name,

        len(donations_data), active_filters, styles,

    )



    if not donations_data:

        elements.append(Paragraph("No donations found matching the current filters.", styles['normal']))

    else:

        donation_ids = [d.id for d in donations_data]
        decline_comments = {}
        if donation_ids:
            declined_responses = (
                DonationResponse.objects
                .filter(
                    donation_id__in=donation_ids,
                    recipient=recipient,
                    response_type='declined',
                )
                .exclude(notes__isnull=True)
                .exclude(notes__exact='')
                .order_by('-responded_at')
                .values_list('donation_id', 'notes')
            )
            for did, note in declined_responses:
                if did not in decline_comments:
                    decline_comments[did] = note or ''

        if donation_type_filter == 'subsidized':

            data = [['S/NO', 'Date', 'Food Bank', 'Donor', 'Type', 'Category', 'Product',

                     'Quantity', 'Market Price', 'Subsidy', 'New Price',

                     'Status', 'Delivery', 'Location', 'Donor Note', 'Recipient Note', 'Decline Comment']]

        else:

            data = [['S/NO', 'Food Bank', 'Type', 'Category', 'Description',

                     'Delivery', 'Status', 'Decline Comment', 'Date']]



        for idx, donation in enumerate(donations_data, 1):

            status = _get_subsidized_status_display(donation, recipient)



            if donation_type_filter == 'subsidized':

                product = (
                    donation.subsidized_product_type
                    or donation.csr_description
                    or donation.other_description
                    or donation.message
                    or 'No description'
                )

                qty_str = ''

                if donation.subsidized_quantity:

                    unit = donation.subsidized_quantity_unit or 'units'

                    qty_str = f"{donation.subsidized_quantity} {unit}"

                elif donation.quantity:

                    qty_str = f"{donation.quantity} {donation.quantity_unit or 'units'}"



                # Price columns

                market = ''

                if donation.subsidized_initial_amount:

                    market = f"KES {donation.subsidized_initial_amount:,.2f}"

                elif donation.subsidized_market_price:

                    market = f"KES {donation.subsidized_market_price:,.2f}"



                subsidy = ''

                if donation.subsidized_subsidy_amount:

                    subsidy = f"KES {donation.subsidized_subsidy_amount:,.2f}"



                new_price = ''

                if donation.subsidized_price is not None:

                    new_price = f"<font color='#059669'><b>KES {donation.subsidized_price:,.2f}</b></font>"

                elif donation.amount:

                    new_price = f"KES {donation.amount:,.2f}"



                donor_profile = getattr(donation.donor, 'donor_profile', None) if donation.donor else None
                donor_name = '-'
                if donor_profile:
                    donor_name = (
                        (donor_profile.organization_name or '').strip()
                        if getattr(donor_profile, 'is_organization', False) and (donor_profile.organization_name or '').strip()
                        else (donor_profile.full_name or '').strip()
                    ) or (getattr(donation.donor, 'email', '') or '').strip() or '-'
                elif donation.donor:
                    donor_name = (donation.donor.email or '').strip() or '-'
                type_label = donation.get_donation_category_display() or 'Food'
                category = donation.get_donation_type_display() or (donation.donation_type.title() if donation.donation_type else 'Subsidized')
                location = (
                    (getattr(donation.foodbank_request, 'location', None) or '').strip()
                    or (getattr(donation.foodbank, 'address', None) or '').strip()
                    or 'Not provided'
                )
                donor_note_text = (donation.message or donation.csr_description or donation.other_description or '-')[:100]
                recipient_note_text = (latest_recipient_notes.get(donation.id) or '-')[:100]
                decline_comment_text = (
                    decline_comments.get(donation.id)
                    or (getattr(donation, 'decline_message', None) or '').strip()
                    or 'No decline comment'
                )[:100]

                data.append([

                    str(idx),
                    donation.donated_at.strftime('%b %d, %Y') if donation.donated_at else '-',

                    Paragraph(donation.foodbank.foodbank_name, wrap),
                    Paragraph(donor_name, wrap),

                    Paragraph(type_label, wrap),

                    Paragraph(category, wrap),

                    Paragraph(product[:120], wrap),

                    Paragraph(qty_str or '-', wrap),

                    Paragraph(market or '-', wrap),

                    Paragraph(subsidy or '-', wrap),

                    Paragraph(new_price or '-', wrap),

                    Paragraph(status, wrap),

                    Paragraph(donation.get_delivery_method_display() if donation.delivery_method else '-', wrap),
                    Paragraph(location[:120], wrap),

                    Paragraph(donor_note_text, wrap),

                    Paragraph(recipient_note_text, wrap),

                    Paragraph(decline_comment_text, wrap),

                ])

            else:

                desc = (donation.csr_description or donation.other_description

                        or donation.message or 'No description')

                cat = 'CSR' if donation.donation_type == 'csr' else 'Other'
                decline_comment_text = (
                    decline_comments.get(donation.id)
                    or (getattr(donation, 'decline_message', None) or '').strip()
                    or 'No decline comment'
                )[:100]
                data.append([

                    str(idx),

                    Paragraph(donation.foodbank.foodbank_name, wrap),

                    Paragraph(donation.donation_type.title(), wrap),

                    Paragraph(cat, wrap),

                    Paragraph(desc[:120], wrap),

                    Paragraph(donation.get_delivery_method_display() if donation.delivery_method else '-', wrap),

                    Paragraph(status, wrap),

                    Paragraph(decline_comment_text, wrap),

                    donation.donated_at.strftime('%b %d, %Y'),

                ])



        if donation_type_filter == 'subsidized':

            col_weights = [

                0.55, 0.95, 1.3, 1.1, 0.9, 0.9, 2.0,

                1.1, 1.0, 0.95, 1.0,

                1.1, 0.95, 1.5, 1.6, 1.6, 1.6,

            ]

        else:

            col_weights = [

                0.55, 1.8, 1.0, 1.0, 3.0,

                1.3, 1.3, 2.1, 1.2,

            ]



        table = make_full_width_table(
            data,
            repeat_rows=1,
            col_weights=col_weights,
            pagesize=report_pagesize,
        )

        table.setStyle(get_branded_table_style(len(data)))

        elements.append(table)



        build_report_summary(elements, [

            ("Total Donations", len(donations_data)),

            ("Available", sum(1 for d in donations_data if not d.accepted_by_recipient)),

            ("Accepted", sum(1 for d in donations_data if d.accepted_by_recipient == recipient and d.delivery_status != 'delivered')),

            ("Received", sum(1 for d in donations_data if d.accepted_by_recipient == recipient and d.delivery_status == 'delivered')),

        ], styles)



    return build_pdf_document(elements, f"{donation_type_filter}_donations", recipient_name, pagesize=report_pagesize)



def export_subsidized_donations_excel_helper(request, donations_data, recipient, donation_type_filter, latest_recipient_notes=None):

    """Generate Excel report for subsidized / CSR / Other donations.

    Columns match table with separate numeric price & qty columns for calculations,
    including Decline Note.

    """

    if latest_recipient_notes is None:
        latest_recipient_notes = {}

    try:

        import openpyxl

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    except ImportError:

        from django.http import HttpResponse

        return HttpResponse("Excel export requires openpyxl. Please install it.")



    from django.http import HttpResponse

    from io import BytesIO

    from datetime import datetime



    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Donations Report"



    recipient_name = recipient.full_name or recipient.user.email

    donation_ids = [d.id for d in donations_data]
    decline_notes = {}
    if donation_ids:
        declined_responses = (
            DonationResponse.objects
            .filter(
                donation_id__in=donation_ids,
                recipient=recipient,
                response_type='declined',
            )
            .exclude(notes__isnull=True)
            .exclude(notes__exact='')
            .order_by('-responded_at')
            .values_list('donation_id', 'notes')
        )
        for did, note in declined_responses:
            if did not in decline_notes:
                decline_notes[did] = note or ''



    # â”€â”€ Branded header â”€â”€

    ws["A1"] = "FOODBANKHUB"

    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")

    ws["A2"] = f"{recipient_name} - {donation_type_filter.title()} Donations Report"

    ws["A2"].font = Font(size=13, bold=True, color="1F4E78")

    ws["A3"] = f"Generated on {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}  |  Total Records: {len(donations_data)}"

    ws["A3"].font = Font(size=10, italic=True)



    # â”€â”€ Headers (row 5) â”€â”€

    if donation_type_filter == "subsidized":

        headers = [

            '#', 'Date', 'Food Bank', 'Donor', 'Type', 'Category', 'Product',

            'Quantity', 'Market Price (KES)', 'Subsidy (KES)', 'New Price (KES)',

            'Status', 'Delivery', 'Location', 'Donor Note', 'Recipient Note', 'Decline Note',

        ]

    else:

        headers = [

            '#', 'Food Bank', 'Type', 'Category', 'Description',

            'Delivery', 'Status', 'Decline Note', 'Date Added',

        ]



    header_fill = PatternFill("solid", fgColor="1F4E78")

    header_font = Font(color="FFFFFF", bold=True, size=10)

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(

        left=Side(style="thin"), right=Side(style="thin"),

        top=Side(style="thin"), bottom=Side(style="thin"),

    )

    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")



    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=5, column=col, value=header)

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = header_alignment

        cell.border = thin_border



    # â”€â”€ Data rows â”€â”€

    row_num = 6

    for idx, donation in enumerate(donations_data, start=1):

        status = _get_subsidized_status_display(donation, recipient)



        if donation_type_filter == "subsidized":

            market_price = donation.subsidized_initial_amount or donation.subsidized_market_price or ''

            subsidy = donation.subsidized_subsidy_amount or ''

            new_price = donation.subsidized_price or ''

            type_label = donation.get_donation_category_display() or "Food"
            category = donation.get_donation_type_display() or (donation.donation_type.title() if donation.donation_type else "Subsidized")
            product = (
                donation.subsidized_product_type
                or donation.csr_description
                or donation.other_description
                or donation.message
                or "No description"
            )
            qty_value = donation.subsidized_quantity if donation.subsidized_quantity is not None else donation.quantity
            qty_unit = donation.subsidized_quantity_unit or donation.quantity_unit or "units"
            quantity_display = f"{qty_value} {qty_unit}" if qty_value is not None else "-"
            donor_profile = getattr(donation.donor, "donor_profile", None) if donation.donor else None
            donor_name = "-"
            if donor_profile:
                donor_name = (
                    (donor_profile.organization_name or "").strip()
                    if getattr(donor_profile, "is_organization", False) and (donor_profile.organization_name or "").strip()
                    else (donor_profile.full_name or "").strip()
                ) or (getattr(donation.donor, "email", "") or "").strip() or "-"
            elif donation.donor:
                donor_name = (donation.donor.email or "").strip() or "-"
            location = (
                (getattr(donation.foodbank_request, "location", None) or "").strip()
                or (getattr(donation.foodbank, "address", None) or "").strip()
                or "Not provided"
            )



            donor_note = donation.message or donation.csr_description or donation.other_description or ''
            recipient_note = latest_recipient_notes.get(donation.id) or ''
            decline_note = (
                decline_notes.get(donation.id)
                or (getattr(donation, 'decline_message', None) or '').strip()
                or 'No decline note'
            )

            row_data = [

                idx,
                donation.donated_at.strftime("%Y-%m-%d") if donation.donated_at else "",

                donation.foodbank.foodbank_name,
                donor_name,

                type_label,

                category,

                product,

                quantity_display,

                market_price,

                subsidy,

                new_price,

                status,

                donation.get_delivery_method_display() if donation.delivery_method else '',
                location,

                donor_note,

                recipient_note,

                decline_note,

            ]

        else:

            desc = (donation.csr_description or donation.other_description

                    or donation.message or "No description")

            cat = "CSR" if donation.donation_type == "csr" else "Other"

            row_data = [

                idx,

                donation.foodbank.foodbank_name,

                donation.donation_type.title(),

                cat,

                desc,

                donation.get_delivery_method_display() if donation.delivery_method else '',

                status,

                decline_notes.get(donation.id) or (getattr(donation, 'decline_message', None) or '').strip() or 'No decline note',

                donation.donated_at.strftime("%Y-%m-%d"),

            ]



        is_alt = idx % 2 == 0

        for col, value in enumerate(row_data, start=1):

            cell = ws.cell(row=row_num, column=col, value=value)

            cell.border = thin_border

            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if is_alt:

                cell.fill = alt_fill



            # Right-align numeric price columns (9=Market, 10=Subsidy, 11=NewPrice)

            if donation_type_filter == "subsidized" and col in [9, 10, 11]:

                cell.alignment = Alignment(horizontal="right", vertical="top")



        row_num += 1



    # â”€â”€ Summary â”€â”€

    summary_row = row_num + 2

    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12, color="1F4E78")



    summary_data = [

        ("Total Donations", len(donations_data)),

        ("Available", sum(1 for d in donations_data if not d.accepted_by_recipient)),

        ("Accepted", sum(1 for d in donations_data if d.accepted_by_recipient == recipient and d.delivery_status != 'delivered')),

        ("Received", sum(1 for d in donations_data if d.accepted_by_recipient == recipient and d.delivery_status == 'delivered')),

    ]

    for i, (label, value) in enumerate(summary_data, start=1):

        ws.cell(row=summary_row + i, column=1, value=label).font = Font(bold=True)

        ws.cell(row=summary_row + i, column=2, value=value)



    # â”€â”€ Column widths â”€â”€

    if donation_type_filter == "subsidized":

        widths = {

            "A": 6, "B": 12, "C": 20, "D": 18, "E": 12,

            "F": 12, "G": 24, "H": 14, "I": 15,

            "J": 14, "K": 15, "L": 16, "M": 14, "N": 20, "O": 22, "P": 22, "Q": 22,

        }

    else:

        widths = {

            "A": 6, "B": 18, "C": 12, "D": 12, "E": 30,

            "F": 14, "G": 14, "H": 22, "I": 14,

        }



    for col, width in widths.items():

        ws.column_dimensions[col].width = width

    ws.row_dimensions[5].height = 30



    # â”€â”€ Add auto-filter to all columns â”€â”€

    if len(donations_data) > 0:

        if donation_type_filter == "subsidized":

            ws.auto_filter.ref = f"A5:Q{row_num - 1}"

        else:

            ws.auto_filter.ref = f"A5:I{row_num - 1}"



    # â”€â”€ Freeze top rows (header) â”€â”€

    ws.freeze_panes = "A6"



    # â”€â”€ Response â”€â”€

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)



    response = HttpResponse(

        buffer,

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

    )

    response["Content-Disposition"] = (

        f'attachment; filename="{recipient_name}_'

        f'{donation_type_filter}_donations_'

        f'{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    )

    return response





# authentication/views.py



# views.py

@login_required

def respond_to_subsidized(request, donation_id):

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Access denied.")

        return redirect('dashboard')



    donation = get_object_or_404(Donation, id=donation_id)

    recipient = request.user.recipient_profile

    

    # Check if already responded

    if DonationResponse.objects.filter(donation=donation, recipient=recipient).exists():
        messages.warning(request, "You have already responded to this donation.")
        return _redirect_back_or_default(request, 'recipient_subsidized_donations')

    

    if request.method == 'POST':

        response_type = request.POST.get('response_type')

        

        if response_type == 'accept':

            recipient_notes = request.POST.get('recipient_notes', '').strip()



            # âœ… Record who accepted it

            donation.accepted_by_recipient = recipient

            donation.save()

             # âœ… Mark as allocated immediately for subsidized donations

            donation.is_allocated = True

            donation.save()

            DonationResponse.objects.create(

                donation=donation,

                recipient=recipient,

                response_type='accepted',

                notes=recipient_notes if recipient_notes else None

            )

            messages.success(request, f"Your acceptance of the {donation.get_donation_type_display()} has been recorded.")

            

        elif response_type == 'partial':

            partial_quantity = int(request.POST.get('partial_quantity', 0))

            DonationResponse.objects.create(

                donation=donation,

                recipient=recipient,

                response_type='partial',

                partial_quantity=partial_quantity

            )

            messages.success(request, f"Your partial acceptance of the {donation.get_donation_type_display()} has been recorded.")

            

        elif response_type == 'decline':

            decline_message = request.POST.get('decline_reason', '').strip() or request.POST.get('decline_message', '').strip()

            

            # âœ… Record who accepted it

            donation.declined_by_recipient = recipient

            donation.save()

            DonationResponse.objects.create(

                donation=donation,

                recipient=recipient,

                response_type='declined',

                notes=decline_message

            )

            messages.info(request, "Your decline has been recorded.")

        

        # Notify food bank

        Notification.objects.create(

            user=donation.foodbank.user,

            notification_type='subsidized_response',

            message=f"Recipient {recipient.full_name} {response_type} your {donation.get_donation_type_display()} donation"

        )



        # Redirect back to caller when provided, else preserve donation-type fallback.
        if donation.donation_type == 'csr':
            fallback_url = f"{reverse('recipient_subsidized_donations')}?donation_type=csr"
        elif donation.donation_type == 'other':
            fallback_url = f"{reverse('recipient_subsidized_donations')}?donation_type=other"
        else:
            fallback_url = 'recipient_subsidized_donations'
        return _redirect_back_or_default(request, fallback_url)

    

    return render(request, 'recipient/respond_to_subsidized.html', {'donation': donation})



@login_required

def confirm_subsidized_received(request, donation_id):

    """Allow recipients to confirm they have received subsidized donations"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Access denied.")

        return redirect('dashboard')



    donation = get_object_or_404(Donation, id=donation_id)

    recipient = request.user.recipient_profile



    # Check if this recipient has accepted this donation

    if donation.accepted_by_recipient != recipient:
        messages.error(request, "You can only confirm receipt for donations you have accepted.")
        return _redirect_back_or_default(request, 'recipient_subsidized_donations')



    # Check if already confirmed

    if donation.delivery_status == 'delivered':
        messages.warning(request, "This donation has already been marked as received.")
        return _redirect_back_or_default(request, 'recipient_subsidized_donations')



    if request.method == 'POST':

        # Update delivery status

        donation.delivery_status = 'delivered'

        donation.save()



        # Create notification for food bank

        Notification.objects.create(

            user=donation.foodbank.user,

            notification_type='donation_received',

            message=f"Recipient {recipient.full_name} has confirmed receipt of your {donation.get_donation_type_display()} donation"

        )



        messages.success(request, "Thank you for confirming receipt. The food bank has been notified.")



        # Redirect back to caller when provided, else preserve donation-type fallback.
        if donation.donation_type == 'csr':
            fallback_url = f"{reverse('recipient_subsidized_donations')}?donation_type=csr"
        elif donation.donation_type == 'other':
            fallback_url = f"{reverse('recipient_subsidized_donations')}?donation_type=other"
        else:
            fallback_url = 'recipient_subsidized_donations'
        return _redirect_back_or_default(request, fallback_url)



    # If GET request, redirect to donations page
    return _redirect_back_or_default(request, 'recipient_subsidized_donations')



@login_required

def foodbank_subsidized_responses(request):

    """Show subsidized donation responses for this food bank"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, "Access denied.")

        return redirect('dashboard')

    

    foodbank = request.user.foodbank_profile

    

    # âœ… Get DonationResponse objects (not RequestManagement)

    responses = DonationResponse.objects.filter(

        donation__foodbank=foodbank,

        donation__donation_type='subsidized'

    ).select_related(

        'donation', 

        'recipient',

        'recipient__user'

    ).order_by('-responded_at')

    

    # Add pagination

    paginator = Paginator(responses, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    context = {

        'responses': page_obj,  # âœ… Changed from 'requests' to 'responses'

        'total_count': responses.count(),

    }

    return render(request, 'foodbank/subsidized_responses.html', context)



# Donor Testimonial Views

@login_required

def create_donor_testimonial(request):

    """Donors can create testimonials with impact photos"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Only donors can submit testimonials.')

        return redirect('dashboard')

    

    if request.method == 'POST':

        from .forms import DonorTestimonialForm

        from .models import DonorTestimonial

        form = DonorTestimonialForm(request.POST, request.FILES)

        if form.is_valid():

            testimonial = form.save(commit=False)

            testimonial.donor = request.user.donor_profile

            testimonial.save()

            messages.success(

                request, 

                "Testimonial submitted successfully! It will be reviewed by an admin before being displayed publicly."

            )

            return redirect('donor_testimonials_list')

    else:

        from .forms import DonorTestimonialForm

        form = DonorTestimonialForm()

    return render(request, 'authentication/create_donor_testimonial.html', {'form': form})



@login_required

def donor_testimonials_list(request):

    """List all testimonials for the logged-in donor"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial



    pending_testimonials = DonorTestimonial.objects.filter(

        donor=request.user.donor_profile,

        approval_status='pending'

    ).order_by('-created_at')



    approved_qs = DonorTestimonial.objects.filter(

        donor=request.user.donor_profile,

        approval_status='approved'

    ).order_by('-created_at')



    displayed_testimonials = [t for t in approved_qs if t.is_currently_displayed()]

    archived_approved = [t for t in approved_qs if not t.is_currently_displayed()]



    rejected_qs = DonorTestimonial.objects.filter(

        donor=request.user.donor_profile,

        approval_status='rejected'

    ).order_by('-created_at')



    archived_testimonials = archived_approved

    rejected_testimonials = rejected_qs



    context = {

        'pending_testimonials': pending_testimonials,

        'displayed_testimonials': displayed_testimonials,

        'archived_testimonials': archived_testimonials,

        'rejected_testimonials': rejected_testimonials,

    }

    return render(request, 'authentication/donor_testimonials_list.html', context)



@login_required

def edit_donor_testimonial(request, testimonial_id):

    """Edit a donor testimonial"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    testimonial = get_object_or_404(

        DonorTestimonial,

        id=testimonial_id,

        donor=request.user.donor_profile

    )

    

    # Only allow editing if pending or rejected

    if testimonial.approval_status == 'approved':

        messages.error(request, 'Cannot edit an approved testimonial.')

        return redirect('donor_testimonials_list')

    

    if request.method == 'POST':

        from .forms import DonorTestimonialForm

        form = DonorTestimonialForm(request.POST, request.FILES, instance=testimonial)

        if form.is_valid():

            testimonial = form.save(commit=False)

            testimonial.approval_status = 'pending'  # Reset to pending after edit

            testimonial.save()

            messages.success(request, 'Testimonial updated and resubmitted for review.')

            return redirect('donor_testimonials_list')

    else:

        from .forms import DonorTestimonialForm

        form = DonorTestimonialForm(instance=testimonial)

    

    context = {'form': form, 'testimonial': testimonial}

    return render(request, 'authentication/create_donor_testimonial.html', context)



@login_required

def delete_donor_testimonial(request, testimonial_id):

    """Delete a donor testimonial"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    testimonial = get_object_or_404(

        DonorTestimonial,

        id=testimonial_id,

        donor=request.user.donor_profile

    )

    

    if request.method == 'POST':

        testimonial.delete()

        messages.success(request, 'Testimonial deleted successfully.')

        return redirect('donor_testimonials_list')

    

    context = {'testimonial': testimonial}

    return render(request, 'authentication/delete_donor_testimonial.html', context)



@login_required

def toggle_donor_testimonial_display(request, testimonial_id):

    """Toggle public display of approved donor testimonial"""

    if request.user.user_type != 'DONOR':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    testimonial = get_object_or_404(

        DonorTestimonial,

        id=testimonial_id,

        donor=request.user.donor_profile

    )

    

    if testimonial.approval_status != 'approved':

        messages.error(request, 'Only approved testimonials can be toggled.')

        return redirect('donor_testimonials_list')

    

    testimonial.display_on_public = not testimonial.display_on_public

    testimonial.save()

    

    status = "enabled" if testimonial.display_on_public else "disabled"

    messages.success(request, f'Public display {status} for this testimonial.')

    return redirect('donor_testimonials_list')



# Admin donor testimonial management views

@login_required

def admin_donor_testimonials_pending(request):

    """Admin view to see all pending donor testimonials"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    pending_testimonials = DonorTestimonial.objects.filter(

        approval_status='pending'

    ).select_related('donor__user').order_by('-created_at')

    

    context = {

        'title': 'Pending Donor Testimonials',

        'pending_testimonials': pending_testimonials,

        'pending_count': pending_testimonials.count(),

    }

    return render(request, 'authentication/admin_donor_testimonials_pending.html', context)



@login_required

def admin_approve_donor_testimonial(request, testimonial_id):

    """Admin approves a donor testimonial"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    testimonial = get_object_or_404(DonorTestimonial, id=testimonial_id)

    

    if request.method == 'POST':

        testimonial.approval_status = 'approved'

        testimonial.reviewed_by = request.user

        testimonial.reviewed_at = timezone.now()

        testimonial.set_default_display_period()

        testimonial.save()

        

        messages.success(request, f'Donor testimonial approved and will be displayed for 1 week.')

        return redirect('admin_donor_testimonials_pending')

    

    context = {

        'title': 'Approve Donor Testimonial',

        'testimonial': testimonial

    }

    return render(request, 'authentication/admin_approve_donor_testimonial.html', context)



@login_required

def admin_reject_donor_testimonial(request, testimonial_id):

    """Admin rejects a donor testimonial"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    testimonial = get_object_or_404(DonorTestimonial, id=testimonial_id)

    

    if request.method == 'POST':

        rejection_reason = request.POST.get('rejection_reason', '')

        testimonial.approval_status = 'rejected'

        testimonial.reviewed_by = request.user

        testimonial.reviewed_at = timezone.now()

        testimonial.rejection_reason = rejection_reason

        testimonial.save()

        

        messages.success(request, 'Donor testimonial rejected.')

        return redirect('admin_donor_testimonials_pending')

    

    context = {

        'title': 'Reject Donor Testimonial',

        'testimonial': testimonial

    }

    return render(request, 'authentication/admin_reject_donor_testimonial.html', context)



@login_required

def admin_all_donor_testimonials(request):

    """Admin view to see all donor testimonials with filters"""

    if request.user.user_type != 'ADMIN':

        messages.error(request, 'Access denied. Admin privileges required.')

        return redirect('dashboard')

    

    from .models import DonorTestimonial

    status_filter = request.GET.get('status', 'all')

    

    testimonials = DonorTestimonial.objects.select_related('donor__user', 'reviewed_by')

    

    if status_filter != 'all':

        testimonials = testimonials.filter(approval_status=status_filter)

    

    testimonials = testimonials.order_by('-created_at')

    

    context = {

        'title': 'All Donor Testimonials',

        'testimonials': testimonials,

        'status_filter': status_filter,

        'total_count': testimonials.count(),

    }

    return render(request, 'authentication/admin_all_donor_testimonials.html', context)



# views.py

@login_required

def respond_to_foodbank_request(request):

    if request.method == 'POST':

        request_id = request.POST.get('request_id')

        foodbank_request = get_object_or_404(FoodBankRequest, id=request_id)

        

        # Get the quantity donor wants to contribute

        quantity_donated = int(request.POST.get('quantity', 0))

        donation_category = request.POST.get('donation_category')

        delivery_method = request.POST.get('delivery_method')

        pickup_time = request.POST.get('pickup_time')

        message = request.POST.get('message', '')

        

        # Validate quantity

        if quantity_donated <= 0:

            messages.error(request, "Invalid quantity.")

            return redirect('respond_to_request', request_id=request_id)

        

        if quantity_donated > foodbank_request.quantity_needed:

            messages.error(request, "Donation exceeds requested quantity.")

            return redirect('respond_to_request', request_id=request_id)

        

        # Create the donation

        donation = Donation.objects.create(

            donor=request.user,

            foodbank=foodbank_request.foodbank,

            foodbank_request=foodbank_request,

            donation_type='item',

            donation_category=donation_category,

            item_name=foodbank_request.title,

            quantity=quantity_donated,

            quantity_unit=foodbank_request.quantity_unit,

            delivery_method=delivery_method,

            pickup_time=pickup_time if pickup_time else None,

            message=message,

            status='pending'

        )

        

        # âœ… Find the corresponding recipient request

        # This is the key missing piece - link to the original request

        try:

            recipient_request = RequestManagement.objects.get(

                foodbank_request=foodbank_request

            )

            

            # âœ… Update the recipient request's fulfillment status

            recipient_request.quantity_fulfilled += quantity_donated

            

            # âœ… Set status based on fulfillment level

            if recipient_request.quantity_fulfilled >= recipient_request.quantity:

                recipient_request.status = 'fulfilled'

                recipient_request.fulfilled_at = timezone.now()

            else:

                recipient_request.status = 'partial'

            

            # Append donor note to request's additional_notes

            if message.strip():

                donor_name = request.user.donor_profile.full_name if hasattr(request.user, 'donor_profile') else request.user.email

                timestamp = timezone.localtime().strftime("%Y-%m-%d %H:%M")

                current_notes = recipient_request.additional_notes or ""

                recipient_request.additional_notes = current_notes + f"\n\n--- Donor Note ({donor_name}, {timestamp}) ---\n{message}"

            

            recipient_request.save()

            

        except RequestManagement.DoesNotExist:

            # Handle case where no recipient request exists

            messages.warning(request, "Donation recorded, but no recipient request found.")

        

        messages.success(request, f"Your donation of {quantity_donated} {foodbank_request.quantity_unit} has been submitted.")

        return redirect('dashboard')

    

    return redirect('dashboard')

# views.py

@login_required

def acknowledge_request(request, request_id):

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Only recipients can acknowledge requests.")

        return redirect('dashboard')



    req = get_object_or_404(

        RequestManagement,

        id=request_id,

        recipient=request.user.recipient_profile

    )



    if request.method == 'POST':

        if req.status not in ['awaiting_recipient', 'fulfilled', 'partial']:

            messages.error(

                request,

                "Can only acknowledge requests that are awaiting your response."

            )

            return redirect('recipient_requests_view')



        acknowledgment_note = request.POST.get('acknowledgment_note', '').strip()



        # Mark acknowledged

        req.acknowledged_by_recipient = True



        # Mark any linked stock allocations as acknowledged

        DonationAllocation.objects.filter(

            request_management=req,

            recipient=req.recipient,

            is_acknowledged=False,

        ).update(is_acknowledged=True)



        # Finalize request

        if req.status == 'awaiting_recipient':

            req.status = 'fulfilled'

            req.fulfilled_at = timezone.now()



        # Append acknowledgment note

        if acknowledgment_note:

            current_notes = req.additional_notes or ""

            timestamp = timezone.localtime().strftime("%Y-%m-%d %H:%M")

            req.additional_notes = (

                current_notes +

                f"\n\n--- Acknowledgment Note ({timestamp}) ---\n{acknowledgment_note}"

            )



        req.save()



        # ðŸ” SYNC ALL RELATED DONATIONS

        foodbank_requests = FoodBankRequest.objects.filter(

            original_request=req

        )



        donations = Donation.objects.filter(

            foodbank_request__in=foodbank_requests

        )


        for donation in donations:
            # Preserve original decline markers so recipients don't suddenly "accept" them.
            if donation.status == 'declined':
                continue

            donation.status = req.status  # fulfilled or partial
            donation.accepted_by_recipient = req.recipient


            # Don't overwrite the original donor message with acknowledgement note
            # Acknowledgement notes are stored in additional_notes with timestamps


            donation.save(update_fields=[
                'status',
                'accepted_by_recipient'
            ])


        messages.success(request, "Request acknowledged successfully!")


    return redirect('recipient_requests_view')


@login_required
def confirm_request_received(request, request_id):
    """Allow recipients to confirm they have received the products for their request"""
    if request.user.user_type != 'RECIPIENT':
        messages.error(request, "Access denied.")
        return redirect('dashboard')


    req = get_object_or_404(RequestManagement, id=request_id, recipient=request.user.recipient_profile)


    # Check if this recipient has acknowledged this request
    if not req.acknowledged_by_recipient:
        messages.error(request, "You can only confirm receipt for acknowledged requests.")
        return redirect('recipient_requests_view')


    # Check if already confirmed
    if getattr(req, 'delivery_status', None) == 'delivered':
        messages.warning(request, "This request has already been marked as received.")
        return redirect('recipient_requests_view')


    if request.method == 'POST':
        try:
            # Update delivery status if field exists
            if hasattr(req, 'delivery_status'):
                req.delivery_status = 'delivered'


            # Always add confirmation note for template checking
            current_notes = req.additional_notes or ""
            timestamp = timezone.localtime().strftime("%Y-%m-%d %H:%M")
            receipt_text = f"\n\n--- Receipt Confirmed ({timestamp}) ---\nRecipient has confirmed receiving the products."
            req.additional_notes = current_notes + receipt_text


            req.save()


            # Create notification for food bank
            if hasattr(req, 'foodbank') and req.foodbank:
                # Use description if available, otherwise use a generic message
                request_desc = getattr(req, 'description', 'their request')
                if len(request_desc) > 50:
                    request_desc = request_desc[:50] + "..."
                Notification.objects.create(
                    user=req.foodbank.user,
                    notification_type='request_received',
                    message=f"Recipient {req.recipient.full_name} has confirmed receipt of products for request: {request_desc}"
                )


            # Ensure all direct donations linked to this request are marked delivered
            linked_requests = FoodBankRequest.objects.filter(original_request=req)
            donations = Donation.objects.filter(foodbank_request__in=linked_requests)


            for donation in donations:
                # Skip previously declined donations so they keep their decline state/visibility.
                if donation.status == 'declined':
                    continue

                fields_to_update = []


                if donation.delivery_status != 'delivered':
                    donation.delivery_status = 'delivered'
                    fields_to_update.append('delivery_status')


                if donation.status != 'accepted':
                    donation.status = 'accepted'
                    fields_to_update.append('status')


                if donation.accepted_by_recipient is None:
                    donation.accepted_by_recipient = req.recipient
                    fields_to_update.append('accepted_by_recipient')


                if fields_to_update:
                    donation.save(update_fields=fields_to_update)


            messages.success(request, f"Thank you for confirming receipt! Note added: {receipt_text[:50]}...")

            return redirect('recipient_requests_view')



        except Exception as e:

            messages.error(request, f"Error confirming receipt: {str(e)}")

            return redirect('recipient_requests_view')



    # If GET request, redirect back

    return redirect('recipient_requests_view')





@login_required

def recipient_decline_request(request, request_id):

    """Allow recipients to decline a request they received from a foodbank"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, "Only recipients can decline requests.")

        return redirect('dashboard')



    req = get_object_or_404(

        RequestManagement,

        id=request_id,

        recipient=request.user.recipient_profile

    )



    if request.method == 'POST':

        # Only allow declining requests that are awaiting recipient response

        if req.status not in ['awaiting_recipient', 'fulfilled', 'partial']:

            messages.error(

                request,

                "Can only decline requests that are awaiting your response."

            )

            return redirect('recipient_requests_view')



        decline_note = request.POST.get('decline_note', '').strip()



        if not decline_note:

            messages.error(request, "Please provide a reason for declining.")

            return redirect('recipient_requests_view')



        linked_fb_requests = list(
            FoodBankRequest.objects.filter(
                Q(original_request=req) | Q(linked_request_management=req)
            ).prefetch_related('donations', 'donations__foodbank_request')
        )

        # Data guardrail: if a foodbank-declined donation was previously mis-flagged
        # as recipient-declined, clear that stale recipient flag.
        for fb_req in linked_fb_requests:
            for donation in fb_req.donations.all():
                if (
                    donation.status == 'declined'
                    and donation.declined_by_recipient_id
                    and (donation.decline_message or '').strip()
                ):
                    donation.declined_by_recipient = None
                    donation.save(update_fields=['declined_by_recipient'])

        def _get_qty_from_donation(donation_obj):
            fr = getattr(donation_obj, 'foodbank_request', None)
            if donation_obj.donation_type == 'money':
                if fr and getattr(fr, 'quantity_needed', None):
                    return int(fr.quantity_needed or 0)
                quantity_context = getattr(req, 'get_donation_quantity_context', lambda: req.quantity)()
                return int(quantity_context or 0)
            if donation_obj.donation_type == 'subsidized':
                base_qty = donation_obj.subsidized_quantity or donation_obj.quantity
                if base_qty:
                    return int(base_qty or 0)
                if fr and getattr(fr, 'quantity_needed', None):
                    return int(fr.quantity_needed or 0)
                return 0
            return int(donation_obj.quantity or 0)

        def _recalculate_quantity_fulfilled():
            total_qty = 0

            allocations_qs = DonationAllocation.objects.filter(
                request_management=req,
                recipient=req.recipient,
                declined_by_recipient=False,
            )
            for allocation in allocations_qs:
                total_qty += int(allocation.quantity or 0)

            seen_donation_ids = set()
            for fb_req in linked_fb_requests:
                for donation in fb_req.donations.all():
                    donation_id = donation.id
                    if donation_id in seen_donation_ids:
                        continue
                    seen_donation_ids.add(donation_id)

                    if donation.declined_by_recipient_id:
                        continue
                    if donation.status == 'declined':
                        continue

                    has_request_allocation = donation.allocations.filter(
                        request_management=req,
                        recipient=req.recipient,
                    ).exists()
                    if has_request_allocation:
                        continue

                    total_qty += _get_qty_from_donation(donation)

            return max(0, int(total_qty))

        pending_donations = []
        for fb_req in linked_fb_requests:
            for donation in fb_req.donations.all():
                if donation.declined_by_recipient_id:
                    continue
                if donation.accepted_by_recipient_id:
                    continue
                # Keep foodbank-declined donations untouched (do not relabel as recipient-declined).
                if donation.status == 'declined':
                    continue
                pending_donations.append(donation)

        pending_donations.sort(
            key=lambda d: (d.donated_at or timezone.now(), getattr(d, 'id', 0)),
            reverse=True,
        )
        latest_pending_donation = pending_donations[0] if pending_donations else None

        all_allocations = DonationAllocation.objects.select_related('donation').filter(
            request_management=req,
            recipient=req.recipient,
        ).order_by('-allocated_at', '-id')
        latest_pending_allocation = all_allocations.filter(
            declined_by_recipient=False,
            is_acknowledged=False,
        ).first()

        selected_target = None
        if latest_pending_donation and latest_pending_allocation:
            donation_ts = latest_pending_donation.donated_at or timezone.now()
            allocation_ts = latest_pending_allocation.allocated_at or timezone.now()
            selected_target = 'allocation' if allocation_ts >= donation_ts else 'donation'
        elif latest_pending_allocation:
            selected_target = 'allocation'
        elif latest_pending_donation:
            selected_target = 'donation'

        if selected_target == 'allocation':
            latest_allocation = latest_pending_allocation
            reversed_qty = int(latest_allocation.quantity or 0)
            req.quantity_fulfilled = max(0, (req.quantity_fulfilled or 0) - reversed_qty)

            latest_allocation.declined_by_recipient = True
            latest_allocation.save(update_fields=['declined_by_recipient'])

            donation = latest_allocation.donation
            if donation:
                remaining_qty = donation.get_remaining_quantity()
                donation.is_allocated = (remaining_qty == 0)
                donation.save(update_fields=['is_allocated'])

            decline_feedback = "Latest delivery declined. Quantity fulfilled has been reduced; earlier fulfillments remain."
        elif selected_target == 'donation':
            qty_to_reverse = _get_qty_from_donation(latest_pending_donation)
            req.quantity_fulfilled = max(0, (req.quantity_fulfilled or 0) - qty_to_reverse)

            latest_pending_donation.declined_by_recipient = req.recipient
            if latest_pending_donation.status != 'declined':
                latest_pending_donation.status = 'declined'
                latest_pending_donation.save(update_fields=['declined_by_recipient', 'status'])
            else:
                latest_pending_donation.save(update_fields=['declined_by_recipient'])

            decline_feedback = "Latest donation declined. Quantity fulfilled has been reduced; earlier fulfillments remain."
        else:
            messages.error(request, "No pending donation or stock allocation is available to decline.")
            return redirect('recipient_requests_view')

        # Recipient decline is terminal for this request:
        # auto-decline any remaining pending allocations/donations so it won't reopen.
        pending_allocations_qs = DonationAllocation.objects.filter(
            request_management=req,
            recipient=req.recipient,
            declined_by_recipient=False,
            is_acknowledged=False,
        )
        allocation_donation_ids = set(
            pending_allocations_qs.values_list('donation_id', flat=True)
        )
        pending_allocations_qs.update(declined_by_recipient=True)

        for donation in Donation.objects.filter(id__in=allocation_donation_ids):
            remaining_qty = donation.get_remaining_quantity()
            donation.is_allocated = (remaining_qty == 0)
            donation.save(update_fields=['is_allocated'])

        for fb_req in linked_fb_requests:
            fb_req.donations.filter(
                accepted_by_recipient__isnull=True,
                declined_by_recipient__isnull=True
            ).exclude(status='declined').update(
                declined_by_recipient=req.recipient,
                status='declined'
            )

        req.quantity_fulfilled = _recalculate_quantity_fulfilled()
        req.status = 'declined'

        req.decline_message = decline_note
        req.updated_by = request.user
        req.save(update_fields=['status', 'decline_message', 'updated_by', 'quantity_fulfilled'])

        if req.foodbank:
            notification_text = (
                f"Recipient {req.recipient.full_name} has declined the request "
                f"'{req.description[:50]}...'. Reason: {decline_note}"
            )
            Notification.objects.create(
                user=req.foodbank.user,
                notification_type='request_declined',
                message=notification_text
            )

        messages.warning(request, "Request declined. No further donations will be requested for this request.")

        return redirect('recipient_requests_view')



    return redirect('recipient_requests_view')





# ============================================================================

# UNSPECIFIED DONATION MANAGEMENT VIEWS

# ============================================================================



def _apply_foodbank_unspecified_status_filter(queryset, status_filter):
    """Apply status filter values used by foodbank unspecified donations table/export."""
    status_key = (status_filter or 'all').strip()
    if status_key in ('', 'all'):
        return queryset

    if status_key == 'pending_foodbank':
        return queryset.filter(foodbank_status='pending_foodbank')
    if status_key == 'declined_by_foodbank':
        return queryset.filter(foodbank_status='declined_by_foodbank')
    if status_key == 'accepted_by_foodbank':
        return queryset.filter(foodbank_status='accepted_by_foodbank')

    if status_key == 'sent_to_recipients':
        return queryset.filter(foodbank_status='accepted_by_foodbank').exclude(
            recipient_status__in=['accepted_by_recipient', 'received', 'declined_by_recipient']
        )
    if status_key == 'claimed_by_recipient':
        return queryset.filter(
            foodbank_status='accepted_by_foodbank',
            recipient_status='accepted_by_recipient',
        )
    if status_key == 'received_by_recipient':
        return queryset.filter(
            foodbank_status='accepted_by_foodbank',
            recipient_status='received',
        )
    if status_key == 'rejected_by_recipient_broadcasted':
        return queryset.filter(
            foodbank_status='accepted_by_foodbank',
            recipient_status='declined_by_recipient',
        )

    # Backward compatibility for legacy links.
    return queryset.filter(foodbank_status=status_key)


def _build_unspecified_response_note_maps(items):
    """Collect latest accepted/non-declined and declined recipient notes per donation."""
    latest_non_decline_notes = {}
    latest_decline_notes = {}
    accepted_recipient_notes = {}

    item_list = [item for item in items if getattr(item, 'donation_id', None)]
    if not item_list:
        return {
            'latest_non_decline_notes': latest_non_decline_notes,
            'latest_decline_notes': latest_decline_notes,
            'accepted_recipient_notes': accepted_recipient_notes,
        }

    donation_ids = [item.donation_id for item in item_list]
    accepted_recipient_by_donation = {
        item.donation_id: item.accepted_by_recipient_id
        for item in item_list
        if getattr(item, 'accepted_by_recipient_id', None)
    }

    responses = (
        DonationResponse.objects
        .filter(donation_id__in=donation_ids)
        .exclude(notes__isnull=True)
        .exclude(notes__exact='')
        .order_by('-responded_at')
    )

    for response in responses:
        note = (response.notes or '').strip()
        if not note:
            continue

        donation_id = response.donation_id
        response_type = (response.response_type or '').strip().lower()

        if response_type == 'declined':
            if donation_id not in latest_decline_notes:
                latest_decline_notes[donation_id] = note
            continue

        if donation_id not in latest_non_decline_notes:
            latest_non_decline_notes[donation_id] = note

        if (
            response_type == 'accepted'
            and accepted_recipient_by_donation.get(donation_id) == response.recipient_id
            and donation_id not in accepted_recipient_notes
        ):
            accepted_recipient_notes[donation_id] = note

    return {
        'latest_non_decline_notes': latest_non_decline_notes,
        'latest_decline_notes': latest_decline_notes,
        'accepted_recipient_notes': accepted_recipient_notes,
    }


def _resolve_unspecified_recipient_notes(item, response_note_maps=None):
    """Resolve recipient and recipient-decline notes for unspecified donations."""
    response_note_maps = response_note_maps or {}
    latest_non_decline_notes = response_note_maps.get('latest_non_decline_notes') or {}
    latest_decline_notes = response_note_maps.get('latest_decline_notes') or {}
    accepted_recipient_notes = response_note_maps.get('accepted_recipient_notes') or {}

    donation_id = getattr(item, 'donation_id', None)
    recipient_status = (getattr(item, 'recipient_status', None) or '').strip()
    has_accepted_recipient = bool(getattr(item, 'accepted_by_recipient_id', None)) or recipient_status in (
        'accepted_by_recipient',
        'received',
    )

    base_recipient_note = (getattr(item, 'recipient_notes', None) or '').strip()
    base_recipient_decline_note = (getattr(item, 'recipient_decline_reason', None) or '').strip()

    latest_non_decline_note = (latest_non_decline_notes.get(donation_id) or '').strip()
    latest_decline_note = (latest_decline_notes.get(donation_id) or '').strip()
    accepted_note = (accepted_recipient_notes.get(donation_id) or '').strip()

    if has_accepted_recipient:
        effective_recipient_note = accepted_note or latest_non_decline_note or base_recipient_note
        effective_recipient_decline_note = ''
    else:
        effective_recipient_note = latest_non_decline_note or base_recipient_note
        effective_recipient_decline_note = latest_decline_note or base_recipient_decline_note

    return effective_recipient_note, effective_recipient_decline_note


@login_required

def foodbank_unspecified_donations(request):

    """

    View for foodbanks to see and manage unspecified donations.

    Shows all donations in one table with filters.

    """

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank users only.')

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    

    # Get all unspecified donations for this foodbank

    from .models import UnspecifiedDonationManagement

    

    # Get filter parameters - default to 'all' to show everything

    status_filter = request.GET.get('status', 'all')

    category_filter = request.GET.get('category', 'all')
    type_filter = request.GET.get('type', 'all')
    quantity_range = (request.GET.get('quantity_range') or 'all').strip()
    amount_range = (request.GET.get('amount_range') or 'all').strip()
    # Backward compatibility for old min/max links.
    quantity_min = (request.GET.get('quantity_min') or '').strip()
    quantity_max = (request.GET.get('quantity_max') or '').strip()
    amount_min = (request.GET.get('amount_min') or '').strip()
    amount_max = (request.GET.get('amount_max') or '').strip()

    delivery_filter = request.GET.get('delivery', 'all')

    if delivery_filter == 'dropoff':

        delivery_filter = 'delivery'

    recipient_filter = request.GET.get('recipient', request.GET.get('claimed', 'all'))

    date_from = request.GET.get('date_from', '')

    date_to = request.GET.get('date_to', '')

    search_query = request.GET.get('search', '')

    apply_status_filter = globals().get('_apply_foodbank_unspecified_status_filter')
    if not callable(apply_status_filter):
        def apply_status_filter(queryset, status_filter_value):
            status_key = (status_filter_value or 'all').strip()
            if status_key in ('', 'all'):
                return queryset
            if status_key == 'pending_foodbank':
                return queryset.filter(foodbank_status='pending_foodbank')
            if status_key == 'declined_by_foodbank':
                return queryset.filter(foodbank_status='declined_by_foodbank')
            if status_key == 'accepted_by_foodbank':
                return queryset.filter(foodbank_status='accepted_by_foodbank')
            if status_key == 'sent_to_recipients':
                return queryset.filter(foodbank_status='accepted_by_foodbank').exclude(
                    recipient_status__in=['accepted_by_recipient', 'received', 'declined_by_recipient']
                )
            if status_key == 'claimed_by_recipient':
                return queryset.filter(foodbank_status='accepted_by_foodbank', recipient_status='accepted_by_recipient')
            if status_key == 'received_by_recipient':
                return queryset.filter(foodbank_status='accepted_by_foodbank', recipient_status='received')
            if status_key == 'rejected_by_recipient_broadcasted':
                return queryset.filter(foodbank_status='accepted_by_foodbank', recipient_status='declined_by_recipient')
            return queryset.filter(foodbank_status=status_key)

    

    unspecified_donations = UnspecifiedDonationManagement.objects.filter(

        donation__foodbank=foodbank_profile

    ).select_related(

        'donation', 'donation__donor', 'donation__donor__donor_profile', 'accepted_by_recipient'

    ).order_by('-created_at')

    

    # Apply status filter
    unspecified_donations = apply_status_filter(unspecified_donations, status_filter)

    

    # Apply type filter (Type column -> donation_category)
    normalized_type_filter = 'non_food' if type_filter == 'nonfood' else type_filter
    if normalized_type_filter and normalized_type_filter != 'all':
        unspecified_donations = unspecified_donations.filter(donation__donation_category=normalized_type_filter)

    # Apply category filter (Category column -> donation_type/donation_mode)
    if category_filter and category_filter != 'all':
        if category_filter == 'free_goods':
            unspecified_donations = unspecified_donations.filter(
                donation__donation_type='item',
                donation__donation_mode='free',
            )
        elif category_filter == 'subsidized':
            unspecified_donations = unspecified_donations.filter(
                Q(donation__donation_type='subsidized') |
                Q(donation__donation_type='item', donation__donation_mode='subsidized')
            )
        elif category_filter == 'monetary':
            unspecified_donations = unspecified_donations.filter(donation__donation_type='money')
        elif category_filter in ('csr', 'other'):
            unspecified_donations = unspecified_donations.filter(donation__donation_type=category_filter)



    # Apply delivery method filter

    if delivery_filter and delivery_filter != 'all':
        if delivery_filter == 'delivery':
            unspecified_donations = unspecified_donations.filter(
                donation__delivery_method__in=['dropoff', 'delivery']
            )
        else:
            unspecified_donations = unspecified_donations.filter(donation__delivery_method=delivery_filter)



    # Apply recipient filter

    if recipient_filter == 'claimed':

        unspecified_donations = unspecified_donations.filter(accepted_by_recipient__isnull=False)

    elif recipient_filter == 'unclaimed':

        unspecified_donations = unspecified_donations.filter(accepted_by_recipient__isnull=True)

    # Apply quantity range filter
    quantity_range_map = {
        '1-100': (1, 100),
        '101-500': (101, 500),
        '501-1000': (501, 1000),
        '1001-5000': (1001, 5000),
        '5001_or_more': (5001, None),
        '5001+': (5001, None),
    }
    quantity_min_value, quantity_max_value = quantity_range_map.get(quantity_range, (None, None))
    if quantity_min_value is None and quantity_max_value is None:
        try:
            quantity_min_value = int(quantity_min) if quantity_min else None
        except (TypeError, ValueError):
            quantity_min_value = None
        try:
            quantity_max_value = int(quantity_max) if quantity_max else None
        except (TypeError, ValueError):
            quantity_max_value = None

    if quantity_min_value is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__quantity__gte=quantity_min_value) |
            Q(donation__subsidized_quantity__gte=quantity_min_value)
        )
    if quantity_max_value is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__quantity__lte=quantity_max_value) |
            Q(donation__subsidized_quantity__lte=quantity_max_value)
        )

    # Apply amount range filter
    amount_range_map = {
        '1-10000': (1, 10000),
        '10001-50000': (10001, 50000),
        '50001-100000': (50001, 100000),
        '100001-500000': (100001, 500000),
        '500001_or_more': (500001, None),
        '500001+': (500001, None),
    }
    amount_min_value, amount_max_value = amount_range_map.get(amount_range, (None, None))
    if amount_min_value is None and amount_max_value is None:
        try:
            amount_min_value = float(amount_min) if amount_min else None
        except (TypeError, ValueError):
            amount_min_value = None
        try:
            amount_max_value = float(amount_max) if amount_max else None
        except (TypeError, ValueError):
            amount_max_value = None

    if amount_min_value is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__amount__gte=amount_min_value) |
            Q(donation__subsidized_price__gte=amount_min_value)
        )
    if amount_max_value is not None:
        unspecified_donations = unspecified_donations.filter(
            Q(donation__amount__lte=amount_max_value) |
            Q(donation__subsidized_price__lte=amount_max_value)
        )



    # Apply date filters

    if date_from:

        from django.utils.dateparse import parse_date

        parsed_from = parse_date(date_from)

        if parsed_from:

            unspecified_donations = unspecified_donations.filter(created_at__date__gte=parsed_from)



    if date_to:

        from django.utils.dateparse import parse_date

        parsed_to = parse_date(date_to)

        if parsed_to:

            unspecified_donations = unspecified_donations.filter(created_at__date__lte=parsed_to)

    

    # Apply search filter (match any visible table value)

    if search_query and search_query.strip():
        raw_search = search_query.strip()
        search_lower = raw_search.lower()

        unspecified_donations = unspecified_donations.annotate(
            search_quantity_text=Cast('donation__quantity', output_field=CharField()),
            search_subsidized_quantity_text=Cast('donation__subsidized_quantity', output_field=CharField()),
            search_amount_text=Cast('donation__amount', output_field=CharField()),
            search_subsidized_amount_text=Cast('donation__subsidized_price', output_field=CharField()),
        )

        search_filters = (
            Q(donation__donor__email__icontains=raw_search) |
            Q(donation__donor__donor_profile__full_name__icontains=raw_search) |
            Q(donation__donor__donor_profile__organization_name__icontains=raw_search) |
            Q(accepted_by_recipient__full_name__icontains=raw_search) |
            Q(accepted_by_recipient__user__email__icontains=raw_search) |
            Q(donation__item_name__icontains=raw_search) |
            Q(donation__message__icontains=raw_search) |
            Q(donation__csr_description__icontains=raw_search) |
            Q(donation__other_description__icontains=raw_search) |
            Q(recipient_notes__icontains=raw_search) |
            Q(foodbank_decline_reason__icontains=raw_search) |
            Q(recipient_decline_reason__icontains=raw_search) |
            Q(donation__foodbank__address__icontains=raw_search) |
            Q(donation__quantity_unit__icontains=raw_search) |
            Q(donation__delivery_method__icontains=raw_search) |
            Q(donation__donation_type__icontains=raw_search) |
            Q(donation__donation_category__icontains=raw_search) |
            Q(donation__donation_mode__icontains=raw_search) |
            Q(foodbank_status__icontains=raw_search) |
            Q(recipient_status__icontains=raw_search) |
            Q(search_quantity_text__icontains=raw_search) |
            Q(search_subsidized_quantity_text__icontains=raw_search) |
            Q(search_amount_text__icontains=raw_search) |
            Q(search_subsidized_amount_text__icontains=raw_search)
        )

        if search_lower in ('money', 'monetary'):
            search_filters |= Q(donation__donation_type='money') | Q(donation__donation_category='monetary')
        if search_lower in ('delivery', 'dropoff'):
            search_filters |= Q(donation__delivery_method__in=['delivery', 'dropoff'])
        if search_lower == 'pickup':
            search_filters |= Q(donation__delivery_method='pickup')
        if search_lower in ('free', 'free goods', 'free_goods'):
            search_filters |= Q(donation__donation_type='item', donation__donation_mode='free')
        if search_lower in ('subsidized', 'subsidised'):
            search_filters |= Q(donation__donation_type='subsidized') | Q(donation__donation_mode='subsidized')
        if search_lower == 'food':
            search_filters |= Q(donation__donation_category='food')
        if search_lower in ('non food', 'non-food', 'nonfood', 'non_food'):
            search_filters |= Q(donation__donation_category='non_food')
        if search_lower == 'csr':
            search_filters |= Q(donation__donation_type='csr') | Q(donation__donation_category='csr')
        if search_lower == 'other':
            search_filters |= Q(donation__donation_type='other') | Q(donation__donation_category='other')
        if search_lower in ('pending', 'pending review'):
            search_filters |= Q(foodbank_status='pending_foodbank')
        if search_lower in ('accepted', 'approved'):
            search_filters |= Q(foodbank_status='accepted_by_foodbank')
        if search_lower in ('declined', 'rejected'):
            search_filters |= Q(foodbank_status='declined_by_foodbank') | Q(recipient_status='declined_by_recipient')
        if search_lower == 'claimed':
            search_filters |= Q(recipient_status='accepted_by_recipient')
        if search_lower == 'received':
            search_filters |= Q(recipient_status='received')
        if search_lower in ('unclaimed', 'not allocated'):
            search_filters |= Q(accepted_by_recipient__isnull=True)

        from django.utils.dateparse import parse_date
        parsed_search_date = parse_date(raw_search)
        if parsed_search_date:
            search_filters |= Q(created_at__date=parsed_search_date)

        unspecified_donations = unspecified_donations.filter(search_filters).distinct()

    

    # Get counts for summary stats (always show total counts, not filtered)

    base_status_qs = UnspecifiedDonationManagement.objects.filter(donation__foodbank=foodbank_profile)
    status_counts = {
        'pending_foodbank': apply_status_filter(base_status_qs, 'pending_foodbank').count(),
        'sent_to_recipients': apply_status_filter(base_status_qs, 'sent_to_recipients').count(),
        'claimed_by_recipient': apply_status_filter(base_status_qs, 'claimed_by_recipient').count(),
        'received_by_recipient': apply_status_filter(base_status_qs, 'received_by_recipient').count(),
        'rejected_by_recipient_broadcasted': apply_status_filter(base_status_qs, 'rejected_by_recipient_broadcasted').count(),
        'accepted_by_foodbank': apply_status_filter(base_status_qs, 'accepted_by_foodbank').count(),
        'declined_by_foodbank': apply_status_filter(base_status_qs, 'declined_by_foodbank').count(),
    }

    

    from django.core.paginator import Paginator

    paginator = Paginator(unspecified_donations, 10)

    page_number = request.GET.get('page')

    donations_page = paginator.get_page(page_number)
    response_note_maps = _build_unspecified_response_note_maps(donations_page.object_list)

    # UI display fields: donor display priority + unified notes payload
    for item in donations_page:
        donation = item.donation
        donor = getattr(donation, 'donor', None)
        donor_profile = getattr(donor, 'donor_profile', None) if donor else None

        organization_name = (getattr(donor_profile, 'organization_name', None) or '').strip() if donor_profile else ''
        profile_full_name = (getattr(donor_profile, 'full_name', None) or '').strip() if donor_profile else ''
        donor_full_name = ''
        if donor:
            try:
                donor_full_name = (donor.get_full_name() or '').strip()
            except Exception:
                donor_full_name = ''

        item.donor_display = organization_name or profile_full_name or donor_full_name or getattr(donor, 'email', 'Unknown donor')

        donor_note = (
            (getattr(donation, 'message', None) or '').strip()
            or (getattr(donation, 'csr_description', None) or '').strip()
            or (getattr(donation, 'other_description', None) or '').strip()
        )
        recipient_note, recipient_decline_note = _resolve_unspecified_recipient_notes(item, response_note_maps)
        foodbank_decline_note = (getattr(item, 'foodbank_decline_reason', None) or '').strip()

        item.effective_recipient_note = recipient_note
        item.effective_recipient_decline_note = recipient_decline_note
        item.donor_note_display = donor_note or 'No donor note'
        item.recipient_note_display = recipient_note or 'No recipient note'
        item.foodbank_decline_note_display = foodbank_decline_note or 'No foodbank decline note'
        item.recipient_decline_note_display = recipient_decline_note or 'No recipient decline note'
        item.has_foodbank_decline_note = bool(foodbank_decline_note)
        item.has_recipient_decline_note = bool(recipient_decline_note)

    

    context = {

        'donations': donations_page,

        'results_count': paginator.count,

        'status': status_filter,

        'category_filter': category_filter,

        'type_filter': type_filter,

        'selected_category': category_filter,

        'delivery_filter': delivery_filter,

        'recipient_filter': recipient_filter,

        'claimed_filter': recipient_filter,

        'date_from': date_from,

        'date_to': date_to,

        'quantity_range': quantity_range,

        'amount_range': amount_range,

        'search_query': search_query,

        'status_counts': status_counts,

        # Keep backward-compatible count keys used elsewhere in template/UI.
        'pending_count': status_counts['pending_foodbank'],

        'accepted_count': status_counts['accepted_by_foodbank'],

        'declined_count': status_counts['declined_by_foodbank'],

        'total_count': (
            status_counts['pending_foodbank']
            + status_counts['accepted_by_foodbank']
            + status_counts['declined_by_foodbank']
        ),

    }

    

    return render(request, 'foodbank/unspecified_donations.html', context)





def _unspecified_donation_row_data(item, idx):
    """Build one row dict for export matching foodbank/unspecified_donations.html table columns."""
    d = item.donation
    type_display = d.get_donation_category_display() if getattr(d, 'donation_category', None) else 'Uncategorized'
    if getattr(d, 'donation_category', None) == 'csr':
        if getattr(d, 'csr_subcategory', None) == 'other' and getattr(d, 'csr_custom_subcategory', None):
            type_display = f"{type_display} / {d.csr_custom_subcategory}"
        elif getattr(d, 'csr_subcategory', None):
            type_display = f"{type_display} / {d.get_csr_subcategory_display()}"
        else:
            type_display = f"{type_display} / CSR Initiative"

    if d.donation_type == 'item':
        if d.donation_mode == 'free':
            category_display = 'Free Goods'
        elif d.donation_mode == 'subsidized':
            category_display = 'Subsidized'
        else:
            category_display = d.get_donation_mode_display() or d.get_donation_type_display()
    elif d.donation_type == 'money':
        category_display = 'Monetary'
    else:
        category_display = d.get_donation_type_display() or ''

    # Reports requirement: for CSR donations, swap Type and Category values.
    if d.donation_type == 'csr':
        type_display, category_display = category_display, type_display

    if d.donation_type == 'item':
        description = (d.item_name or 'General donation').strip()
    elif d.donation_type == 'csr':
        description = (d.csr_description or 'CSR initiative').strip()
    elif d.donation_type == 'other':
        description = (d.other_description or 'Other donation').strip()
    elif d.donation_type == 'money':
        description = (d.message or 'Monetary donation').strip() if getattr(d, 'message', None) else 'Monetary donation'
    else:
        description = 'No description'

    donor = getattr(d, 'donor', None)
    donor_email = getattr(donor, 'email', '') if donor else ''
    donor_name = ''
    donor_profile = getattr(donor, 'donor_profile', None) if donor else None
    if donor_profile:
        org_name = (getattr(donor_profile, 'organization_name', None) or '').strip()
        profile_name = (getattr(donor_profile, 'full_name', None) or '').strip()
        donor_name = org_name or profile_name
    if not donor_name and donor:
        try:
            donor_name = (donor.get_full_name() or '').strip()
        except Exception:
            donor_name = ''
    if not donor_name:
        donor_name = donor_email

    if item.accepted_by_recipient:
        rec = item.accepted_by_recipient
        recipient_display = getattr(rec, 'full_name', None) or 'Recipient'
        rec_email = getattr(getattr(rec, 'user', None), 'email', None) or getattr(getattr(rec, 'user', None), 'username', '')
        if rec_email:
            recipient_display = f"{recipient_display} ({rec_email})"
    else:
        recipient_display = 'Not allocated'

    if d.donation_type == 'item':
        qty = d.quantity
        unit = d.quantity_unit or 'units'
        quantity_display = f"{qty} {unit}" if qty is not None else '-'
        quantity_value = qty if qty is not None else ''
        unit_value = unit
        amount_value = ''
    elif d.donation_type == 'money':
        quantity_display = f"KES {d.amount:.0f}" if d.amount is not None else '-'
        quantity_value = ''
        unit_value = ''
        amount_value = d.amount if d.amount is not None else ''
    elif d.donation_type in ('csr', 'other'):
        parts = []
        if d.quantity is not None:
            parts.append(f"{d.quantity} {d.quantity_unit or 'units'}")
        if d.amount is not None:
            parts.append(f"{d.amount:.0f}")
        quantity_display = ' '.join(parts) if parts else '-'
        if d.quantity is not None:
            quantity_value = d.quantity
            unit_value = d.quantity_unit or 'units'
        else:
            quantity_value = ''
            unit_value = ''
        amount_value = d.amount if d.amount is not None else ''
    else:
        quantity_display = '-'
        quantity_value = unit_value = ''
        amount_value = ''

    if getattr(d, 'delivery_method', None) in ('dropoff', 'delivery'):
        delivery_display = 'Delivery'
    else:
        delivery_display = d.get_delivery_method_display() if getattr(d, 'delivery_method', None) else 'Not specified'
    location = getattr(getattr(d, 'foodbank', None), 'address', None) or 'Location not provided'

    donor_note = (
        (getattr(d, 'message', None) or '').strip()
        or (getattr(d, 'csr_description', None) or '').strip()
        or (getattr(d, 'other_description', None) or '').strip()
        or 'No donor note'
    )
    effective_recipient_note = (getattr(item, 'effective_recipient_note', None) or '').strip()
    if effective_recipient_note:
        recipient_note = effective_recipient_note
    else:
        recipient_note = (item.recipient_notes or '').strip() if getattr(item, 'recipient_notes', None) else ''
    if not recipient_note:
        recipient_note = 'No recipient note'

    effective_recipient_decline = (getattr(item, 'effective_recipient_decline_note', None) or '').strip()
    recipient_status = (getattr(item, 'recipient_status', None) or '').strip()
    has_accepted_recipient = bool(getattr(item, 'accepted_by_recipient_id', None)) or recipient_status in (
        'accepted_by_recipient',
        'received',
    )
    if effective_recipient_decline:
        recipient_decline_reason = effective_recipient_decline
    elif has_accepted_recipient:
        recipient_decline_reason = ''
    else:
        recipient_decline_reason = (getattr(item, 'recipient_decline_reason', None) or '').strip()

    foodbank_decline_reason = (getattr(item, 'foodbank_decline_reason', None) or '').strip()
    decline_reason = recipient_decline_reason or foodbank_decline_reason

    if item.foodbank_status == 'accepted_by_foodbank':
        if getattr(item, 'recipient_status', None) == 'received':
            status_display = 'Received by recipient'
        elif getattr(item, 'recipient_status', None) == 'accepted_by_recipient':
            status_display = 'Claimed by recipient'
        elif getattr(item, 'recipient_status', None) == 'declined_by_recipient':
            status_display = 'Declined by recipient-broadcasted to other recipients'
        else:
            status_display = 'Sent to recipients'
    elif item.foodbank_status == 'pending_foodbank':
        status_display = item.get_foodbank_status_display() or 'Pending Review'
    elif item.foodbank_status == 'declined_by_foodbank':
        status_display = item.get_foodbank_status_display() or 'Declined'
    else:
        status_display = item.get_foodbank_status_display() or 'â€”'

    created = item.created_at
    date_display = created.strftime('%b %d, %Y') if created else ''
    time_display = created.strftime('%H:%M') if created else ''

    return {
        'sno': idx, 'type_display': type_display, 'category_display': category_display, 'description': description,
        'donor_name': donor_name, 'donor_email': donor_email, 'recipient_display': recipient_display,
        'quantity_display': quantity_display, 'quantity_value': quantity_value, 'unit_value': unit_value,
        'amount_value': amount_value,
        'delivery_display': delivery_display, 'location': location,
        'donor_note': donor_note, 'recipient_note': recipient_note,
        'foodbank_rejection_note': foodbank_decline_reason,
        'recipient_rejection_note': recipient_decline_reason,
        'decline_reason': decline_reason,
        'status_display': status_display, 'date_display': date_display, 'time_display': time_display,
    }


@login_required
def foodbank_export_unspecified_donations(request):
    """Export unspecified donations to CSV, Excel, or PDF - columns match foodbank/unspecified_donations.html table."""
    import csv
    from django.http import HttpResponse
    from datetime import datetime

    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'Access denied. Foodbank users only.')
        return redirect('foodbank_unspecified_donations')

    foodbank_profile = request.user.foodbank_profile
    export_format = request.GET.get('format', 'csv')
    status_filter = request.GET.get('status', 'all')
    category_filter = (request.GET.get('category', 'all') or 'all').strip().lower()
    type_filter = (request.GET.get('type', 'all') or 'all').strip().lower()
    if type_filter in ('', 'all') and category_filter in ('food', 'non_food', 'nonfood'):
        # Backward compatibility for old links that used `category` for type values.
        type_filter = category_filter
    if type_filter not in ('all', 'food', 'non_food', 'nonfood'):
        type_filter = 'all'
    quantity_range = (request.GET.get('quantity_range') or 'all').strip()
    amount_range = (request.GET.get('amount_range') or 'all').strip()
    # Backward compatibility for old min/max links.
    quantity_min = (request.GET.get('quantity_min') or '').strip()
    quantity_max = (request.GET.get('quantity_max') or '').strip()
    amount_min = (request.GET.get('amount_min') or '').strip()
    amount_max = (request.GET.get('amount_max') or '').strip()
    delivery_filter = request.GET.get('delivery', 'all')
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    recipient_filter = request.GET.get('recipient', request.GET.get('claimed', 'all'))
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = (request.GET.get('search') or '').strip()
    apply_status_filter = globals().get('_apply_foodbank_unspecified_status_filter')
    if not callable(apply_status_filter):
        def apply_status_filter(queryset, status_filter_value):
            status_key = (status_filter_value or 'all').strip()
            if status_key in ('', 'all'):
                return queryset
            if status_key == 'pending_foodbank':
                return queryset.filter(foodbank_status='pending_foodbank')
            if status_key == 'declined_by_foodbank':
                return queryset.filter(foodbank_status='declined_by_foodbank')
            if status_key == 'accepted_by_foodbank':
                return queryset.filter(foodbank_status='accepted_by_foodbank')
            if status_key == 'sent_to_recipients':
                return queryset.filter(foodbank_status='accepted_by_foodbank').exclude(
                    recipient_status__in=['accepted_by_recipient', 'received', 'declined_by_recipient']
                )
            if status_key == 'claimed_by_recipient':
                return queryset.filter(foodbank_status='accepted_by_foodbank', recipient_status='accepted_by_recipient')
            if status_key == 'received_by_recipient':
                return queryset.filter(foodbank_status='accepted_by_foodbank', recipient_status='received')
            if status_key == 'rejected_by_recipient_broadcasted':
                return queryset.filter(foodbank_status='accepted_by_foodbank', recipient_status='declined_by_recipient')
            return queryset.filter(foodbank_status=status_key)

    from .models import UnspecifiedDonationManagement

    donations_qs = UnspecifiedDonationManagement.objects.filter(
        donation__foodbank=foodbank_profile
    ).select_related(
        'donation', 'donation__donor', 'donation__donor__donor_profile', 'donation__foodbank', 'accepted_by_recipient', 'accepted_by_recipient__user'
    ).order_by('-created_at')

    donations_qs = apply_status_filter(donations_qs, status_filter)
    normalized_type_filter = 'non_food' if type_filter == 'nonfood' else type_filter
    if normalized_type_filter and normalized_type_filter != 'all':
        donations_qs = donations_qs.filter(donation__donation_category=normalized_type_filter)
    if category_filter and category_filter != 'all':
        if category_filter == 'free_goods':
            donations_qs = donations_qs.filter(
                donation__donation_type='item',
                donation__donation_mode='free',
            )
        elif category_filter == 'subsidized':
            donations_qs = donations_qs.filter(
                Q(donation__donation_type='subsidized') |
                Q(donation__donation_type='item', donation__donation_mode='subsidized')
            )
        elif category_filter == 'monetary':
            donations_qs = donations_qs.filter(donation__donation_type='money')
        elif category_filter in ('csr', 'other'):
            donations_qs = donations_qs.filter(donation__donation_type=category_filter)
    if delivery_filter and delivery_filter != 'all':
        if delivery_filter == 'delivery':
            donations_qs = donations_qs.filter(donation__delivery_method__in=['dropoff', 'delivery'])
        else:
            donations_qs = donations_qs.filter(donation__delivery_method=delivery_filter)
    if recipient_filter == 'claimed':
        donations_qs = donations_qs.filter(accepted_by_recipient__isnull=False)
    elif recipient_filter == 'unclaimed':
        donations_qs = donations_qs.filter(accepted_by_recipient__isnull=True)

    # Apply quantity range filter
    quantity_range_map = {
        '1-100': (1, 100),
        '101-500': (101, 500),
        '501-1000': (501, 1000),
        '1001-5000': (1001, 5000),
        '5001_or_more': (5001, None),
        '5001+': (5001, None),
    }
    quantity_min_value, quantity_max_value = quantity_range_map.get(quantity_range, (None, None))
    if quantity_min_value is None and quantity_max_value is None:
        try:
            quantity_min_value = int(quantity_min) if quantity_min else None
        except (TypeError, ValueError):
            quantity_min_value = None
        try:
            quantity_max_value = int(quantity_max) if quantity_max else None
        except (TypeError, ValueError):
            quantity_max_value = None

    if quantity_min_value is not None:
        donations_qs = donations_qs.filter(
            Q(donation__quantity__gte=quantity_min_value) |
            Q(donation__subsidized_quantity__gte=quantity_min_value)
        )
    if quantity_max_value is not None:
        donations_qs = donations_qs.filter(
            Q(donation__quantity__lte=quantity_max_value) |
            Q(donation__subsidized_quantity__lte=quantity_max_value)
        )

    # Apply amount range filter
    amount_range_map = {
        '1-10000': (1, 10000),
        '10001-50000': (10001, 50000),
        '50001-100000': (50001, 100000),
        '100001-500000': (100001, 500000),
        '500001_or_more': (500001, None),
        '500001+': (500001, None),
    }
    amount_min_value, amount_max_value = amount_range_map.get(amount_range, (None, None))
    if amount_min_value is None and amount_max_value is None:
        try:
            amount_min_value = float(amount_min) if amount_min else None
        except (TypeError, ValueError):
            amount_min_value = None
        try:
            amount_max_value = float(amount_max) if amount_max else None
        except (TypeError, ValueError):
            amount_max_value = None

    if amount_min_value is not None:
        donations_qs = donations_qs.filter(
            Q(donation__amount__gte=amount_min_value) |
            Q(donation__subsidized_price__gte=amount_min_value)
        )
    if amount_max_value is not None:
        donations_qs = donations_qs.filter(
            Q(donation__amount__lte=amount_max_value) |
            Q(donation__subsidized_price__lte=amount_max_value)
        )
    if date_from:
        from django.utils.dateparse import parse_date
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations_qs = donations_qs.filter(created_at__date__gte=parsed_from)
    if date_to:
        from django.utils.dateparse import parse_date
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations_qs = donations_qs.filter(created_at__date__lte=parsed_to)
    if search_query and search_query.strip():
        raw_search = search_query.strip()
        search_lower = raw_search.lower()

        donations_qs = donations_qs.annotate(
            search_quantity_text=Cast('donation__quantity', output_field=CharField()),
            search_subsidized_quantity_text=Cast('donation__subsidized_quantity', output_field=CharField()),
            search_amount_text=Cast('donation__amount', output_field=CharField()),
            search_subsidized_amount_text=Cast('donation__subsidized_price', output_field=CharField()),
        )

        search_filters = (
            Q(donation__donor__email__icontains=raw_search) |
            Q(donation__donor__donor_profile__full_name__icontains=raw_search) |
            Q(donation__donor__donor_profile__organization_name__icontains=raw_search) |
            Q(accepted_by_recipient__full_name__icontains=raw_search) |
            Q(accepted_by_recipient__user__email__icontains=raw_search) |
            Q(donation__item_name__icontains=raw_search) |
            Q(donation__message__icontains=raw_search) |
            Q(donation__csr_description__icontains=raw_search) |
            Q(donation__other_description__icontains=raw_search) |
            Q(recipient_notes__icontains=raw_search) |
            Q(foodbank_decline_reason__icontains=raw_search) |
            Q(recipient_decline_reason__icontains=raw_search) |
            Q(donation__foodbank__address__icontains=raw_search) |
            Q(donation__quantity_unit__icontains=raw_search) |
            Q(donation__delivery_method__icontains=raw_search) |
            Q(donation__donation_type__icontains=raw_search) |
            Q(donation__donation_category__icontains=raw_search) |
            Q(donation__donation_mode__icontains=raw_search) |
            Q(foodbank_status__icontains=raw_search) |
            Q(recipient_status__icontains=raw_search) |
            Q(search_quantity_text__icontains=raw_search) |
            Q(search_subsidized_quantity_text__icontains=raw_search) |
            Q(search_amount_text__icontains=raw_search) |
            Q(search_subsidized_amount_text__icontains=raw_search)
        )

        if search_lower in ('money', 'monetary'):
            search_filters |= Q(donation__donation_type='money') | Q(donation__donation_category='monetary')
        if search_lower in ('delivery', 'dropoff'):
            search_filters |= Q(donation__delivery_method__in=['delivery', 'dropoff'])
        if search_lower == 'pickup':
            search_filters |= Q(donation__delivery_method='pickup')
        if search_lower in ('free', 'free goods', 'free_goods'):
            search_filters |= Q(donation__donation_type='item', donation__donation_mode='free')
        if search_lower in ('subsidized', 'subsidised'):
            search_filters |= Q(donation__donation_type='subsidized') | Q(donation__donation_mode='subsidized')
        if search_lower == 'food':
            search_filters |= Q(donation__donation_category='food')
        if search_lower in ('non food', 'non-food', 'nonfood', 'non_food'):
            search_filters |= Q(donation__donation_category='non_food')
        if search_lower == 'csr':
            search_filters |= Q(donation__donation_type='csr') | Q(donation__donation_category='csr')
        if search_lower == 'other':
            search_filters |= Q(donation__donation_type='other') | Q(donation__donation_category='other')
        if search_lower in ('pending', 'pending review'):
            search_filters |= Q(foodbank_status='pending_foodbank')
        if search_lower in ('accepted', 'approved'):
            search_filters |= Q(foodbank_status='accepted_by_foodbank')
        if search_lower in ('declined', 'rejected'):
            search_filters |= Q(foodbank_status='declined_by_foodbank') | Q(recipient_status='declined_by_recipient')
        if search_lower == 'claimed':
            search_filters |= Q(recipient_status='accepted_by_recipient')
        if search_lower == 'received':
            search_filters |= Q(recipient_status='received')
        if search_lower in ('unclaimed', 'not allocated'):
            search_filters |= Q(accepted_by_recipient__isnull=True)

        from django.utils.dateparse import parse_date
        parsed_search_date = parse_date(raw_search)
        if parsed_search_date:
            search_filters |= Q(created_at__date=parsed_search_date)

        donations_qs = donations_qs.filter(search_filters).distinct()

    donations_list = list(donations_qs)
    export_response_note_maps = _build_unspecified_response_note_maps(donations_list)
    for item in donations_list:
        recipient_note, recipient_decline_note = _resolve_unspecified_recipient_notes(item, export_response_note_maps)
        item.effective_recipient_note = recipient_note
        item.effective_recipient_decline_note = recipient_decline_note

    data_rows = [_unspecified_donation_row_data(item, i) for i, item in enumerate(donations_list, 1)]
    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')

    

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="unspecified_donations_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow([foodbank_profile.foodbank_name or 'Food Bank', 'Unspecified Donations'])
        writer.writerow(['Generated', timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        writer.writerow([
            'S/No', 'Type', 'Category', 'Description', 'Donor', 'Recipient',
            'Quantity', 'Unit', 'Amount (KES)', 'Delivery', 'Location',
            'Donor note', 'Recipient note', 'Foodbank rejection note', 'Recipient rejection note',
            'Status', 'Date', 'Time'
        ])
        for row in data_rows:
            writer.writerow([
                row['sno'], row['type_display'], row['category_display'], (row['description'] or '').replace('\n', ' '),
                row['donor_name'], row['recipient_display'],
                row.get('quantity_value', '') if row.get('quantity_value') is not None and row.get('quantity_value') != '' else '',
                row.get('unit_value', ''),
                row.get('amount_value', '') if row.get('amount_value') is not None and row.get('amount_value') != '' else '',
                row['delivery_display'], (row['location'] or '').replace('\n', ' '),
                (row['donor_note'] or '').replace('\n', ' '), (row['recipient_note'] or '').replace('\n', ' '),
                (row.get('foodbank_rejection_note') or '').replace('\n', ' '),
                (row.get('recipient_rejection_note') or '').replace('\n', ' '),
                row['status_display'], row['date_display'], row['time_display'],
            ])
        writer.writerow([])
        writer.writerow(['Total Donations', len(data_rows)])
        return response

    

    elif export_format == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unspecified Donations"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")

        headers = [
            'S/No', 'Type', 'Category', 'Description', 'Donor', 'Recipient',
            'Quantity', 'Unit', 'Amount (KES)', 'Delivery', 'Location',
            'Donor note', 'Recipient note', 'Foodbank rejection note', 'Recipient rejection note',
            'Status', 'Date', 'Time'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for row_idx, row in enumerate(data_rows, 2):
            ws.cell(row=row_idx, column=1, value=row['sno'])
            ws.cell(row=row_idx, column=2, value=row['type_display'])
            ws.cell(row=row_idx, column=3, value=row['category_display'])
            ws.cell(row=row_idx, column=4, value=(row['description'] or '').replace('\n', ' '))
            ws.cell(row=row_idx, column=5, value=row['donor_name'])
            ws.cell(row=row_idx, column=6, value=row['recipient_display'])
            qval = row.get('quantity_value')
            ws.cell(row=row_idx, column=7, value=qval if qval is not None and qval != '' else '')
            ws.cell(row=row_idx, column=8, value=row.get('unit_value', ''))
            aval = row.get('amount_value')
            ws.cell(row=row_idx, column=9, value=aval if aval is not None and aval != '' else '')
            ws.cell(row=row_idx, column=10, value=row['delivery_display'])
            ws.cell(row=row_idx, column=11, value=(row['location'] or '').replace('\n', ' '))
            ws.cell(row=row_idx, column=12, value=(row['donor_note'] or '').replace('\n', ' '))
            ws.cell(row=row_idx, column=13, value=(row['recipient_note'] or '').replace('\n', ' '))
            ws.cell(row=row_idx, column=14, value=(row.get('foodbank_rejection_note') or '').replace('\n', ' '))
            ws.cell(row=row_idx, column=15, value=(row.get('recipient_rejection_note') or '').replace('\n', ' '))
            ws.cell(row=row_idx, column=16, value=row['status_display'])
            ws.cell(row=row_idx, column=17, value=row['date_display'])
            ws.cell(row=row_idx, column=18, value=row['time_display'])

        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = min(22, max(10, len(headers[c - 1]) + 2))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="unspecified_donations_{timestamp}.xlsx"'
        wb.save(response)
        return response

    

    elif export_format == 'pdf':
        from reportlab.lib.units import inch
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.platypus import Table, Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from .report_utils import (
            get_report_styles, build_report_header, get_branded_table_style,
            build_report_summary, build_pdf_document, collect_active_filters, make_full_width_table,
        )
        styles = get_report_styles()
        wrap = styles['wrap']
        rejection_wrap = ParagraphStyle(
            'UnspecifiedRejectionWrap',
            parent=wrap,
            fontSize=6.3,
            leading=7.3,
            wordWrap='CJK',
        )
        report_pagesize = landscape(A3)
        elements = []
        name = foodbank_profile.foodbank_name or request.user.email
        active_filters = collect_active_filters(request, [
            ('status', 'Status'), ('type', 'Type'), ('category', 'Category'), ('delivery', 'Delivery'),
            ('recipient', 'Recipient'),
            ('quantity_range', 'Quantity Range'),
            ('amount_range', 'Amount Range'),
            ('date_from', 'From'), ('date_to', 'To'),
            ('search', 'Search'),
        ])
        build_report_header(elements, "Unspecified Donations Report", name, len(data_rows), active_filters, styles)

        if not data_rows:
            elements.append(Paragraph("No donations found matching the current filters.", styles['normal']))
        else:
            data = [[
                'S/No', 'Type', 'Category', 'Description', 'Donor', 'Recipient',
                'Quantity', 'Unit', 'Amount (KES)', 'Delivery', 'Location',
                'Donor note', 'Recipient note', 'Foodbank rejection note', 'Recipient rejection note',
                'Status', 'Date'
            ]]
            for row in data_rows:
                data.append([
                    str(row['sno']),
                    Paragraph((row['type_display'] or '-')[:30], wrap),
                    Paragraph((row['category_display'] or '-')[:40], wrap),
                    Paragraph((row['description'] or '-')[:80].replace('\n', '<br/>'), wrap),
                    Paragraph((row['donor_name'] or '-')[:50], wrap),
                    Paragraph((row['recipient_display'] or '-')[:40], wrap),
                    Paragraph(str(row.get('quantity_value', '') if row.get('quantity_value') is not None else ''), wrap),
                    Paragraph(str(row.get('unit_value', '') or ''), wrap),
                    Paragraph(str(row.get('amount_value', '') if row.get('amount_value') is not None else ''), wrap),
                    Paragraph((row['delivery_display'] or '-')[:15], wrap),
                    Paragraph((row['location'] or '-')[:50].replace('\n', ' '), wrap),
                    Paragraph((row['donor_note'] or '-')[:80].replace('\n', '<br/>'), wrap),
                    Paragraph((row['recipient_note'] or '-')[:80].replace('\n', '<br/>'), wrap),
                    Paragraph((row.get('foodbank_rejection_note') or '-')[:160].replace('\n', '<br/>'), rejection_wrap),
                    Paragraph((row.get('recipient_rejection_note') or '-')[:160].replace('\n', '<br/>'), rejection_wrap),
                    Paragraph((row['status_display'] or '-').replace('\n', ' '), wrap),
                    Paragraph(f"{row['date_display']} {row['time_display']}".strip(), wrap),
                ])
            col_weights = [
                0.40, 0.65, 0.85, 1.05, 0.90, 0.85, 0.60, 0.50, 0.75, 0.65, 0.90, 0.90, 0.90, 1.30, 1.30, 0.90, 0.80,
            ]
            table = make_full_width_table(data, repeat_rows=1, col_weights=col_weights, pagesize=report_pagesize)
            table.setStyle(get_branded_table_style(len(data)))
            elements.append(table)
            build_report_summary(elements, [("Total Donations", len(data_rows))], styles)
        safe_name = (foodbank_profile.foodbank_name or "foodbank").replace(" ", "_")[:30]
        return build_pdf_document(elements, "unspecified_donations", safe_name, pagesize=report_pagesize)

    

    return redirect('foodbank_unspecified_donations')





@login_required

def foodbank_accept_unspecified_donation(request, donation_id):

    """Foodbank accepts an unspecified donation - makes it available to org recipients"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank users only.')

        return redirect('dashboard')

    

    if request.method != 'POST':

        return redirect('foodbank_unspecified_donations')

    

    from .models import UnspecifiedDonationManagement

    

    try:

        unspecified = UnspecifiedDonationManagement.objects.get(

            id=donation_id,

            donation__foodbank=request.user.foodbank_profile,

            foodbank_status='pending_foodbank'

        )

    except UnspecifiedDonationManagement.DoesNotExist:

        messages.error(request, 'Donation not found or already processed.')

        return redirect('foodbank_unspecified_donations')

    

    unspecified.foodbank_accept()

    

    # Update the original donation status

    unspecified.donation.status = 'accepted'

    unspecified.donation.save()

    

    messages.success(request, 'Donation accepted! It is now available for organization recipients.')

    return redirect('foodbank_unspecified_donations')





@login_required

def foodbank_decline_unspecified_donation(request, donation_id):

    """Foodbank declines an unspecified donation"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank users only.')

        return redirect('dashboard')

    

    if request.method != 'POST':

        return redirect('foodbank_unspecified_donations')

    

    from .models import UnspecifiedDonationManagement

    

    try:

        unspecified = UnspecifiedDonationManagement.objects.get(

            id=donation_id,

            donation__foodbank=request.user.foodbank_profile,

            foodbank_status='pending_foodbank'

        )

    except UnspecifiedDonationManagement.DoesNotExist:

        messages.error(request, 'Donation not found or already processed.')

        return redirect('foodbank_unspecified_donations')

    

    reason = request.POST.get('decline_reason', '')

    unspecified.foodbank_decline(reason=reason)

    

    # Update the original donation status

    unspecified.donation.status = 'declined'

    unspecified.donation.decline_message = reason

    unspecified.donation.save()

    

    messages.success(request, 'Donation declined.')

    return redirect('foodbank_unspecified_donations')





@login_required

def foodbank_subsidized_donations(request):

    """

    View for foodbanks to see and manage subsidized donations.

    Shows all subsidized donations in one table with filters.

    """

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank users only.')

        return redirect('dashboard')

    

    foodbank_profile = request.user.foodbank_profile

    

    # Get filter parameters

    status_filter = request.GET.get('status', 'all')

    type_filter = (request.GET.get('type') or request.GET.get('category') or 'all').strip().lower()
    if type_filter not in ('all', 'food', 'non_food'):
        type_filter = 'all'

    delivery_filter = (request.GET.get('delivery', 'all') or 'all').strip().lower()
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    if delivery_filter not in ('all', 'pickup', 'delivery'):
        delivery_filter = 'all'

    recipient_filter = request.GET.get('recipient', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')

    date_from = request.GET.get('date_from', '')

    date_to = request.GET.get('date_to', '')

    search_query = request.GET.get('search', '')

    

    donations = Donation.objects.filter(

        foodbank=foodbank_profile,

        donation_type='subsidized',

        foodbank_request__isnull=True  # Only donor-initiated (unspecified) subsidized donations

    ).select_related('donor', 'donor__donor_profile', 'accepted_by_recipient', 'accepted_by_recipient__user').order_by('-donated_at')

    

    # Apply status filter

    if status_filter and status_filter != 'all':

        donations = donations.filter(status=status_filter)



    # Apply type filter (food / non-food)
    if type_filter != 'all':
        donations = donations.filter(donation_category=type_filter)



    # Apply delivery method filter

    if delivery_filter and delivery_filter != 'all':
        if delivery_filter == 'delivery':
            donations = donations.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            donations = donations.filter(delivery_method='pickup')



    # Apply recipient filter

    if recipient_filter == 'claimed':

        donations = donations.filter(accepted_by_recipient__isnull=False)

    elif recipient_filter == 'unclaimed':

        donations = donations.filter(accepted_by_recipient__isnull=True)

    # Apply quantity range filter
    if quantity_filter != 'all':
        if quantity_filter == 'small':
            donations = donations.filter(subsidized_quantity__lte=100)
        elif quantity_filter == 'medium':
            donations = donations.filter(subsidized_quantity__gt=100, subsidized_quantity__lte=500)
        elif quantity_filter == 'large':
            donations = donations.filter(subsidized_quantity__gt=500)

    # Apply amount range filter
    if amount_filter != 'all':
        if amount_filter == 'small':
            donations = donations.filter(subsidized_price__lte=100)
        elif amount_filter == 'medium':
            donations = donations.filter(subsidized_price__gt=100, subsidized_price__lte=500)
        elif amount_filter == 'large':
            donations = donations.filter(subsidized_price__gt=500)

    

    # Apply date filters

    if date_from:

        from django.utils.dateparse import parse_date

        parsed_from = parse_date(date_from)

        if parsed_from:

            donations = donations.filter(donated_at__date__gte=parsed_from)



    if date_to:

        from django.utils.dateparse import parse_date

        parsed_to = parse_date(date_to)

        if parsed_to:

            donations = donations.filter(donated_at__date__lte=parsed_to)



    # Apply search filter

    if search_query:

        donations = donations.filter(

            Q(donor__email__icontains=search_query) |

            Q(subsidized_product_type__icontains=search_query) |

            Q(message__icontains=search_query)

        )

    

    # Get counts for summary stats

    counts = {

        'pending': Donation.objects.filter(

            foodbank=foodbank_profile,

            donation_type='subsidized',

            status='pending',

            foodbank_request__isnull=True

        ).count(),

        'accepted': Donation.objects.filter(

            foodbank=foodbank_profile,

            donation_type='subsidized',

            status='accepted',

            foodbank_request__isnull=True

        ).count(),

        'declined': Donation.objects.filter(

            foodbank=foodbank_profile,

            donation_type='subsidized',

            status='declined',

            foodbank_request__isnull=True

        ).count(),

    }

    

    # Pagination

    paginator = Paginator(donations, 10)  # 10 items per page

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    # Attach recipient decline responses for "Show all rejections" (foodbank subsidized table only)
    donation_ids = [d.id for d in page_obj.object_list]
    decline_responses_by_donation = {}
    if donation_ids:
        for resp in DonationResponse.objects.filter(
            donation_id__in=donation_ids,
            response_type='declined'
        ).select_related('recipient').order_by('-responded_at'):
            decline_responses_by_donation.setdefault(resp.donation_id, []).append(resp)
    for donation in page_obj.object_list:
        donation.decline_responses = decline_responses_by_donation.get(donation.id, [])

    # Attach latest recipient note (acceptance or decline) for Notes column, same as recipient subsidized table
    latest_notes = {}
    if donation_ids:
        for response in DonationResponse.objects.filter(
            donation_id__in=donation_ids
        ).exclude(notes__isnull=True).exclude(notes__exact='').order_by('-responded_at'):
            if response.donation_id not in latest_notes:
                latest_notes[response.donation_id] = response.notes
    for donation in page_obj.object_list:
        donation.latest_recipient_note = latest_notes.get(donation.id)
        rec = getattr(donation, 'accepted_by_recipient', None)
        if rec:
            recipient_name = (
                (getattr(rec, 'full_name', None) or '').strip()
                or (getattr(rec, 'organization_name', None) or '').strip()
                or (getattr(getattr(rec, 'user', None), 'email', None) or '').strip()
                or 'Recipient'
            )
        else:
            recipient_name = ''
        donation.recipient_display_name = recipient_name

        donor_note = (
            (getattr(donation, 'message', None) or '').strip()
            or (getattr(donation, 'csr_description', None) or '').strip()
            or (getattr(donation, 'other_description', None) or '').strip()
        )
        recipient_note = (donation.latest_recipient_note or '').strip() if donation.latest_recipient_note else ''
        foodbank_decline_note = (getattr(donation, 'decline_message', None) or '').strip()

        recipient_rejection_entries = []
        for resp in getattr(donation, 'decline_responses', []):
            who = (
                getattr(getattr(resp, 'recipient', None), 'full_name', None)
                or getattr(getattr(getattr(resp, 'recipient', None), 'user', None), 'email', None)
                or 'Recipient'
            )
            when = resp.responded_at.strftime('%b %d, %Y %H:%M') if getattr(resp, 'responded_at', None) else ''
            note = (getattr(resp, 'notes', None) or '').strip() or 'No note provided.'
            if when:
                recipient_rejection_entries.append(f"{who} ({when}): {note}")
            else:
                recipient_rejection_entries.append(f"{who}: {note}")

        donation.donor_note_display = donor_note or 'No donor note'
        donation.recipient_note_display = recipient_note or 'No recipient note'
        donation.foodbank_decline_note_display = foodbank_decline_note or 'No foodbank rejection note'
        donation.recipient_rejection_notes_display = "\n".join(recipient_rejection_entries) if recipient_rejection_entries else 'No recipient rejection notes'
        donation.has_foodbank_decline_note = bool(foodbank_decline_note)
        donation.has_recipient_rejection_notes = bool(recipient_rejection_entries)

    context = {

        'donations': page_obj,

        'page_obj': page_obj,

        'status': status_filter,

        'type_filter': type_filter,
        'category_filter': type_filter,

        'delivery_filter': delivery_filter,

        'recipient_filter': recipient_filter,
        'quantity_filter': quantity_filter,
        'amount_filter': amount_filter,

        'date_from': date_from,

        'date_to': date_to,

        'search_query': search_query,

        'pending_count': counts['pending'],

        'accepted_count': counts['accepted'],

        'declined_count': counts['declined'],

        'total_count': counts['pending'] + counts['accepted'] + counts['declined'],

    }

    

    return render(request, 'foodbank/subsidized_donations.html', context)





@login_required

def foodbank_accept_subsidized_donation(request, donation_id):

    """Foodbank accepts a subsidized donation - makes it available to all recipients"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank users only.')

        return redirect('foodbank_subsidized_donations')

    

    if request.method != 'POST':

        return redirect('foodbank_subsidized_donations')

    

    try:

        donation = Donation.objects.get(

            id=donation_id,

            foodbank=request.user.foodbank_profile,

            donation_type='subsidized',

            status='pending'

        )

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found or already processed.')

        return redirect('foodbank_subsidized_donations')

    

    donation.status = 'accepted'

    donation.save()

    

    # Also update the related RequestManagement status if it exists

    if donation.request_management:

        request_mgmt = donation.request_management

        # Update quantity fulfilled if applicable

        if donation.subsidized_quantity and request_mgmt.quantity:

            donated_qty = int(donation.subsidized_quantity or 0)

            fulfilled = int(request_mgmt.quantity_fulfilled or 0)

            needed = int(request_mgmt.quantity or 0)

            remaining = needed - fulfilled

            

            if donated_qty > 0 and remaining > 0:

                use = min(donated_qty, remaining)

                request_mgmt.quantity_fulfilled = fulfilled + use

                

                # Update status based on fulfillment

                if request_mgmt.quantity_fulfilled >= needed:

                    request_mgmt.status = 'awaiting_recipient'

                else:

                    request_mgmt.status = 'partial'

                

                request_mgmt.save()

                

                # Notify recipient

                if request_mgmt.status == 'partial':

                    Notification.objects.create(

                        user=request_mgmt.recipient.user,

                        message=(

                            f"Good news â€” {use} {request_mgmt.get_unit_display()} "

                            f"has been allocated for your request '{request_mgmt.description}'. "

                            f"Status: Partially fulfilled ({request_mgmt.quantity_fulfilled}/{request_mgmt.quantity})."

                        )

                    )

                else:

                    Notification.objects.create(

                        user=request_mgmt.recipient.user,

                        notification_type='donation_received',

                        message=(

                            f"A subsidized donation covering your request '{request_mgmt.description}' "

                            f"is ready. Please check available donations to claim it."

                        )

                    )

    

    messages.success(request, 'Subsidized donation accepted! It is now available to all recipients.')

    return redirect(request.POST.get('next') or 'foodbank_subsidized_donations')





@login_required

def foodbank_decline_subsidized_donation(request, donation_id):

    """Foodbank declines a subsidized donation"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank users only.')

        return redirect('foodbank_subsidized_donations')

    

    if request.method != 'POST':

        return redirect('foodbank_subsidized_donations')

    

    try:

        donation = Donation.objects.get(

            id=donation_id,

            foodbank=request.user.foodbank_profile,

            donation_type='subsidized',

            status='pending'

        )

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found or already processed.')

        return redirect('foodbank_subsidized_donations')

    

    reason = request.POST.get('decline_reason', '')

    donation.status = 'declined'

    donation.decline_message = reason

    donation.save()

    

    messages.success(request, 'Subsidized donation declined.')

    return redirect(request.POST.get('next') or 'foodbank_subsidized_donations')





def _subsidized_donation_row_data(donation, idx, latest_recipient_note=None, recipient_decline_note=None):
    """Build one row dict for subsidized donations export - columns match foodbank/subsidized_donations.html."""
    dp = getattr(donation.donor, 'donor_profile', None)
    if dp:
        if getattr(dp, 'full_name', None):
            donor_name = (dp.full_name or '').strip()
        elif getattr(dp, 'is_organization', False) and getattr(dp, 'organization_name', None):
            donor_name = (dp.organization_name or '').strip()
        else:
            donor_name = (donation.donor.email or '').strip()
    else:
        donor_name = (donation.donor.email or '').strip()
    donor_name = donor_name or '-'

    type_display = 'Subsidized'
    category_display = donation.get_donation_category_display() or 'Food'
    product = (donation.subsidized_product_type or 'Subsidized Goods').strip()

    qty = donation.subsidized_quantity or donation.quantity
    unit = (donation.subsidized_quantity_unit or donation.quantity_unit or 'units').strip()
    quantity_value = qty if qty is not None else ''
    unit_value = unit

    market_price = ''
    if donation.subsidized_initial_amount is not None:
        market_price = donation.subsidized_initial_amount
    elif getattr(donation, 'subsidized_market_price', None) is not None:
        market_price = donation.subsidized_market_price
    subsidy_value = donation.subsidized_subsidy_amount if donation.subsidized_subsidy_amount is not None else ''
    new_price = donation.subsidized_price if donation.subsidized_price is not None else ''

    if donation.status == 'pending':
        status_display = 'Awaiting review'
    elif donation.status == 'accepted':
        if donation.accepted_by_recipient:
            status_display = 'Received by recipient' if getattr(donation, 'delivery_status', None) == 'delivered' else 'Accepted by recipient'
        elif getattr(donation, 'declined_by_recipient', None):
            status_display = 'Rejected - broadcasted to other recipients'
        else:
            status_display = 'Available to recipients'
    elif donation.status == 'fulfilled':
        status_display = 'Fulfilled & allocated'
    elif donation.status == 'declined':
        status_display = 'Rejected by foodbank'
    else:
        status_display = donation.get_status_display() or 'Status pending'

    delivery_display = donation.get_delivery_method_display() if donation.delivery_method else 'Not specified'

    if donation.status == 'accepted':
        if donation.accepted_by_recipient:
            recipient_display = (
                (getattr(donation.accepted_by_recipient, 'full_name', None) or '').strip()
                or (getattr(donation.accepted_by_recipient, 'organization_name', None) or '').strip()
                or (getattr(getattr(donation.accepted_by_recipient, 'user', None), 'email', None) or '').strip()
                or 'Claimed'
            )
        else:
            recipient_display = 'Not claimed'
    elif donation.status == 'declined':
        recipient_display = 'None'
    else:
        recipient_display = '-'

    donor_note = (donation.message or donation.csr_description or donation.other_description or '').strip() or 'No donor note'
    recipient_note = (latest_recipient_note or '').strip() if latest_recipient_note else ('Recipient hasn\'t added a note.' if donation.accepted_by_recipient else 'Not yet claimed.')

    recipient_decline_reason = (recipient_decline_note or '').strip()
    recipient_rejection_note = recipient_decline_reason
    foodbank_decline_reason = (getattr(donation, 'decline_message', None) or '').strip()
    decline_reason = recipient_decline_reason or foodbank_decline_reason

    quantity_display = f"{qty} {unit}" if qty is not None else '-'
    if donation.donated_at:
        try:
            date_display = timezone.localtime(donation.donated_at).strftime('%b %d, %Y')
        except Exception:
            date_display = donation.donated_at.strftime('%b %d, %Y')
    else:
        date_display = '-'

    return {
        'sno': idx, 'date_display': date_display, 'donor_name': donor_name, 'type_display': type_display, 'category_display': category_display,
        'product': product, 'quantity_value': quantity_value, 'unit_value': unit_value, 'quantity_display': quantity_display,
        'market_price': market_price, 'subsidy_value': subsidy_value, 'new_price': new_price,
        'status_display': status_display, 'delivery_display': delivery_display, 'recipient_display': recipient_display,
        'donor_note': donor_note, 'recipient_note': recipient_note,
        'recipient_rejection_note': recipient_rejection_note,
        'recipient_decline_reason': recipient_decline_reason,
        'foodbank_decline_reason': foodbank_decline_reason,
        'decline_reason': decline_reason,
    }


@login_required
def foodbank_subsidized_donations_export(request):
    """Export subsidized donations as CSV, PDF, or Excel - columns match foodbank/subsidized_donations.html table."""
    from django.http import HttpResponse
    from datetime import datetime

    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'Access denied. Foodbank users only.')
        return redirect('foodbank_subsidized_donations')

    foodbank_profile = request.user.foodbank_profile
    export_format = request.GET.get('format', 'csv')

    status_filter = request.GET.get('status', 'all')
    type_filter = (request.GET.get('type') or request.GET.get('category') or 'all').strip().lower()
    if type_filter not in ('all', 'food', 'non_food'):
        type_filter = 'all'
    delivery_filter = (request.GET.get('delivery', 'all') or 'all').strip().lower()
    if delivery_filter == 'dropoff':
        delivery_filter = 'delivery'
    if delivery_filter not in ('all', 'pickup', 'delivery'):
        delivery_filter = 'all'
    recipient_filter = request.GET.get('recipient', 'all')
    quantity_filter = request.GET.get('quantity', 'all')
    amount_filter = request.GET.get('amount', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')

    donations_qs = Donation.objects.filter(
        foodbank=foodbank_profile,
        donation_type='subsidized',
        foodbank_request__isnull=True
    ).select_related('donor', 'donor__donor_profile', 'accepted_by_recipient', 'accepted_by_recipient__user', 'declined_by_recipient').prefetch_related(
        'responses', 'responses__recipient'
    ).order_by('-donated_at')

    if status_filter and status_filter != 'all':
        donations_qs = donations_qs.filter(status=status_filter)
    if type_filter != 'all':
        donations_qs = donations_qs.filter(donation_category=type_filter)
    if delivery_filter and delivery_filter != 'all':
        if delivery_filter == 'delivery':
            donations_qs = donations_qs.filter(delivery_method__in=['delivery', 'dropoff'])
        else:
            donations_qs = donations_qs.filter(delivery_method='pickup')
    if recipient_filter == 'claimed':
        donations_qs = donations_qs.filter(accepted_by_recipient__isnull=False)
    elif recipient_filter == 'unclaimed':
        donations_qs = donations_qs.filter(accepted_by_recipient__isnull=True)
    if quantity_filter != 'all':
        if quantity_filter == 'small':
            donations_qs = donations_qs.filter(subsidized_quantity__lte=100)
        elif quantity_filter == 'medium':
            donations_qs = donations_qs.filter(subsidized_quantity__gt=100, subsidized_quantity__lte=500)
        elif quantity_filter == 'large':
            donations_qs = donations_qs.filter(subsidized_quantity__gt=500)
    if amount_filter != 'all':
        if amount_filter == 'small':
            donations_qs = donations_qs.filter(subsidized_price__lte=100)
        elif amount_filter == 'medium':
            donations_qs = donations_qs.filter(subsidized_price__gt=100, subsidized_price__lte=500)
        elif amount_filter == 'large':
            donations_qs = donations_qs.filter(subsidized_price__gt=500)
    if date_from:
        from django.utils.dateparse import parse_date
        parsed_from = parse_date(date_from)
        if parsed_from:
            donations_qs = donations_qs.filter(donated_at__date__gte=parsed_from)
    if date_to:
        from django.utils.dateparse import parse_date
        parsed_to = parse_date(date_to)
        if parsed_to:
            donations_qs = donations_qs.filter(donated_at__date__lte=parsed_to)
    if search_query:
        donations_qs = donations_qs.filter(
            Q(donor__email__icontains=search_query) |
            Q(subsidized_product_type__icontains=search_query) |
            Q(message__icontains=search_query)
        )

    donations_list = list(donations_qs)
    donation_ids = [d.id for d in donations_list]
    latest_notes = {}
    decline_notes = {}
    if donation_ids:
        for response in DonationResponse.objects.filter(
            donation_id__in=donation_ids
        ).exclude(notes__isnull=True).exclude(notes__exact='').order_by('-responded_at'):
            if response.donation_id not in latest_notes:
                latest_notes[response.donation_id] = response.notes

        for response in DonationResponse.objects.filter(
            donation_id__in=donation_ids, response_type='declined'
        ).exclude(notes__isnull=True).exclude(notes__exact='').order_by('-responded_at'):
            if response.donation_id not in decline_notes:
                decline_notes[response.donation_id] = response.notes

    data_rows = [_subsidized_donation_row_data(d, i, latest_notes.get(d.id), decline_notes.get(d.id)) for i, d in enumerate(donations_list, 1)]

    if export_format == 'csv':
        return export_subsidized_csv(data_rows, foodbank_profile)
    elif export_format == 'pdf':
        return export_subsidized_pdf(request, data_rows, foodbank_profile)
    elif export_format == 'excel':
        return export_subsidized_excel(data_rows, foodbank_profile)
    else:
        return HttpResponse("Invalid format", status=400)





def export_subsidized_csv(data_rows, foodbank_profile):
    """Export subsidized donations as CSV - columns match table; Quantity/Unit and Donor note/Recipient note separate; (KSH) in headers."""
    import csv
    from django.http import HttpResponse
    from datetime import datetime

    response = HttpResponse(content_type='text/csv')
    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f'{foodbank_profile.foodbank_name}_subsidized_donations_{timestamp}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow([f'Subsidized Donations Report - {foodbank_profile.foodbank_name}'])
    writer.writerow([f'Generated on: {timezone.localtime().strftime("%B %d, %Y at %I:%M %p")}'])
    writer.writerow([])
    writer.writerow([
        'S/No', 'Date', 'Donor', 'Category', 'Type', 'Product', 'Quantity', 'Unit',
        'Market price (KSH)', 'Subsidy (KSH)', 'New price (KSH)', 'Status', 'Delivery', 'Recipient',
        'Donor note', 'Recipient note', 'Recipient rejection note', 'Foodbank decline reason'
    ])
    for row in data_rows:
        writer.writerow([
            row['sno'], row.get('date_display', '-'), row['donor_name'], row['category_display'], row['type_display'], (row['product'] or '').replace('\n', ' '),
            row.get('quantity_value', ''), row.get('unit_value', ''),
            row.get('market_price', ''), row.get('subsidy_value', ''), row.get('new_price', ''),
            row['status_display'], row['delivery_display'], row['recipient_display'],
            (row['donor_note'] or '').replace('\n', ' '), (row['recipient_note'] or '').replace('\n', ' '),
            (row.get('recipient_rejection_note') or '').replace('\n', ' '),
            (row.get('foodbank_decline_reason') or '').replace('\n', ' '),
        ])
    writer.writerow([])
    writer.writerow(['Total', len(data_rows)])
    return response





def export_subsidized_pdf(request, data_rows, foodbank_profile):
    """Export subsidized donations as PDF - columns match table; Market price/Subsidy/New price (KSH); Donor note & Recipient note separate."""
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from .report_utils import (
        get_report_styles, build_report_header, get_branded_table_style,
        build_report_summary, build_pdf_document, collect_active_filters,
        make_full_width_table,
    )
    styles = get_report_styles()
    report_pagesize = landscape(A3)
    wrap = ParagraphStyle(
        'SubsidizedCellWrap',
        parent=styles['wrap'],
        fontSize=6,
        leading=7.5,
        alignment=TA_LEFT,
    )
    header_wrap = ParagraphStyle(
        'SubsidizedHeaderWrap',
        parent=styles['wrap'],
        fontSize=6,
        leading=8,
        alignment=TA_LEFT,
        textColor=colors.white,
        fontName='Helvetica-Bold',
    )
    elements = []
    name = foodbank_profile.foodbank_name or request.user.email
    active_filters = collect_active_filters(request, [
        ('status', 'Status'), ('type', 'Type'), ('delivery', 'Delivery'),
        ('recipient', 'Recipient'), ('date_from', 'From'), ('date_to', 'To'),
        ('search', 'Search'),
    ])
    build_report_header(elements, "Subsidized Donations Report", name, len(data_rows), active_filters, styles)

    if not data_rows:
        elements.append(Paragraph("No subsidized donations found matching the current filters.", styles['normal']))
    else:
        header_row = [
            Paragraph('S/No', header_wrap),
            Paragraph('Date', header_wrap),
            Paragraph('Donor', header_wrap),
            Paragraph('Category', header_wrap),
            Paragraph('Type', header_wrap),
            Paragraph('Product', header_wrap),
            Paragraph('Quantity', header_wrap),
            Paragraph('Market price (KSH)', header_wrap),
            Paragraph('Subsidy (KSH)', header_wrap),
            Paragraph('New price (KSH)', header_wrap),
            Paragraph('Status', header_wrap),
            Paragraph('Delivery', header_wrap),
            Paragraph('Recipient', header_wrap),
            Paragraph('Donor note', header_wrap),
            Paragraph('Recipient note', header_wrap),
            Paragraph('Recipient rejection note', header_wrap),
            Paragraph('Foodbank decline reason', header_wrap),
        ]
        data = [header_row]
        for row in data_rows:
            data.append([
                Paragraph(str(row['sno']), wrap),
                Paragraph(str(row.get('date_display', '-')), wrap),
                Paragraph((row['donor_name'] or '-').replace('\n', '<br/>'), wrap),
                Paragraph((row['category_display'] or '-'), wrap),
                Paragraph((row['type_display'] or '-'), wrap),
                Paragraph((row['product'] or '-').replace('\n', '<br/>'), wrap),
                Paragraph((row['quantity_display'] or '-'), wrap),
                Paragraph(str(row.get('market_price', '-')), wrap),
                Paragraph(str(row.get('subsidy_value', '-')), wrap),
                Paragraph(str(row.get('new_price', '-')), wrap),
                Paragraph((row['status_display'] or '-').replace('\n', '<br/>'), wrap),
                Paragraph((row['delivery_display'] or '-'), wrap),
                Paragraph((row['recipient_display'] or '-'), wrap),
                Paragraph((row['donor_note'] or '-').replace('\n', '<br/>'), wrap),
                Paragraph((row['recipient_note'] or '-').replace('\n', '<br/>'), wrap),
                Paragraph((row.get('recipient_rejection_note') or '-').replace('\n', '<br/>'), wrap),
                Paragraph((row.get('foodbank_decline_reason') or '-').replace('\n', '<br/>'), wrap),
            ])
        col_weights = [
            0.4, 0.8, 1.2, 0.5, 0.7, 1.0, 0.6,
            1.0, 0.8, 0.8, 1.2, 0.6, 0.8,
            1.3, 1.3, 1.3, 1.3,
        ]
        table = make_full_width_table(data, repeat_rows=1, col_weights=col_weights, pagesize=report_pagesize)
        table.setStyle(get_branded_table_style(len(data)))
        elements.append(table)
        build_report_summary(elements, [("Total Donations", len(data_rows))], styles)
    safe_name = (foodbank_profile.foodbank_name or "foodbank").replace(" ", "_")[:30]
    return build_pdf_document(elements, "subsidized_donations", safe_name, pagesize=report_pagesize)





def export_subsidized_excel(data_rows, foodbank_profile):
    """Export subsidized donations as Excel - columns match table; Quantity/Unit and Donor note/Recipient note separate; (KSH) in headers."""
    from django.http import HttpResponse
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Subsidized Donations"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        'S/No', 'Date', 'Donor', 'Category', 'Type', 'Product', 'Quantity', 'Unit',
        'Market price (KSH)', 'Subsidy (KSH)', 'New price (KSH)', 'Status', 'Delivery', 'Recipient',
        'Donor note', 'Recipient note', 'Recipient rejection note', 'Foodbank decline reason'
    ]
    num_cols = len(headers)
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'].value = f"Subsidized Donations Report - {foodbank_profile.foodbank_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    ws['A2'].value = f"Generated on: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')} | Total: {len(data_rows)}"
    ws['A2'].alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, row in enumerate(data_rows, 1):
        r = row_idx + 4
        ws.cell(row=r, column=1, value=row['sno']).border = border
        ws.cell(row=r, column=2, value=row.get('date_display', '-')).border = border
        ws.cell(row=r, column=3, value=(row['donor_name'] or '').replace('\n', ' ')).border = border
        ws.cell(row=r, column=4, value=row['category_display']).border = border
        ws.cell(row=r, column=5, value=row['type_display']).border = border
        ws.cell(row=r, column=6, value=(row['product'] or '').replace('\n', ' ')).border = border
        ws.cell(row=r, column=7, value=row.get('quantity_value', '')).border = border
        ws.cell(row=r, column=8, value=row.get('unit_value', '')).border = border
        ws.cell(row=r, column=9, value=row.get('market_price', '')).border = border
        ws.cell(row=r, column=10, value=row.get('subsidy_value', '')).border = border
        ws.cell(row=r, column=11, value=row.get('new_price', '')).border = border
        ws.cell(row=r, column=12, value=row['status_display']).border = border
        ws.cell(row=r, column=13, value=row['delivery_display']).border = border
        ws.cell(row=r, column=14, value=row['recipient_display']).border = border
        ws.cell(row=r, column=15, value=(row['donor_note'] or '').replace('\n', ' ')).border = border
        ws.cell(row=r, column=16, value=(row['recipient_note'] or '').replace('\n', ' ')).border = border
        ws.cell(row=r, column=17, value=(row.get('recipient_rejection_note') or '').replace('\n', ' ')).border = border
        ws.cell(row=r, column=18, value=(row.get('foodbank_decline_reason') or '').replace('\n', ' ')).border = border

    column_widths = [6, 12, 18, 10, 12, 18, 8, 8, 12, 10, 10, 22, 12, 16, 22, 22, 24, 24]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f'{foodbank_profile.foodbank_name}_subsidized_donations_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
def _annotate_recipient_unspecified_decline_state(donations, recipient_profile):
    """Annotate unspecified donations with whether this recipient has declined them."""
    declined_response_qs = DonationResponse.objects.filter(
        donation_id=OuterRef('donation_id'),
        recipient=recipient_profile,
        response_type='declined',
    )
    return donations.annotate(
        _declined_by_me=Exists(declined_response_qs),
        _declined_by_me_legacy=Case(
            When(donation__declined_by_recipient=recipient_profile, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        ),
    )


def _attach_recipient_unspecified_effective_status(items, recipient_profile):
    """
    Attach per-recipient effective status fields.
    Global status may be "declined_by_recipient" after rebroadcast; for recipients
    who did not decline, treat it as available with action buttons.
    """
    item_list = list(items)
    donation_ids = [item.donation_id for item in item_list if getattr(item, 'donation_id', None)]
    response_map = {
        donation_id: {'response_type': response_type, 'notes': (notes or '')}
        for donation_id, response_type, notes in DonationResponse.objects.filter(
            donation_id__in=donation_ids,
            recipient=recipient_profile,
        ).values_list('donation_id', 'response_type', 'notes')
    }

    status_labels = {
        'not_applicable': 'Not Claimed',
        'pending_recipient': 'Available for Recipients',
        'accepted_by_recipient': 'Accepted by Recipient',
        'declined_by_recipient': 'Declined by recipient-broadcasted',
        'received': 'Received by Recipient',
    }
    status_badges = {
        'not_applicable': 'bg-secondary',
        'pending_recipient': 'bg-info',
        'accepted_by_recipient': 'bg-primary',
        'declined_by_recipient': 'bg-danger',
        'received': 'bg-success',
    }

    for item in item_list:
        current_status = getattr(item, 'recipient_status', 'not_applicable')
        accepted_by_current_recipient = (
            getattr(item, 'accepted_by_recipient_id', None) == getattr(recipient_profile, 'id', None)
        )
        response_entry = response_map.get(item.donation_id, {})
        response_type = response_entry.get('response_type')
        response_note = (response_entry.get('notes') or '').strip()
        declined_by_me = (
            response_type == 'declined'
            or bool(getattr(item, '_declined_by_me', False))
            or bool(getattr(item, '_declined_by_me_legacy', False))
            or getattr(getattr(item, 'donation', None), 'declined_by_recipient_id', None) == getattr(recipient_profile, 'id', None)
        )

        effective_status = current_status
        if declined_by_me and not accepted_by_current_recipient:
            effective_status = 'declined_by_recipient'
        elif current_status == 'declined_by_recipient' and not declined_by_me:
            effective_status = 'pending_recipient'

        item.effective_recipient_status = effective_status
        item.effective_recipient_status_display = status_labels.get(effective_status, status_labels['not_applicable'])
        item.effective_recipient_status_badge_class = status_badges.get(effective_status, status_badges['not_applicable'])
        if effective_status == 'declined_by_recipient':
            item.effective_recipient_note = ''
        elif accepted_by_current_recipient or effective_status in ('accepted_by_recipient', 'received'):
            item.effective_recipient_note = response_note or (getattr(item, 'recipient_notes', None) or '').strip()
        else:
            item.effective_recipient_note = ''
        item.effective_decline_note = (
            (response_note if response_type == 'declined' else '').strip()
            or ((getattr(item, 'recipient_decline_reason', None) or '').strip() if declined_by_me else '')
        )
        item.effective_description = resolve_available_donation_description(getattr(item, 'donation', None))

    return item_list


def _recipient_unspecified_filters(request):
    status_choices = {'all', 'available', 'my_accepted', 'received', 'declined'}
    type_choices = {'all', 'food', 'non_food', 'monetary', 'csr', 'other'}
    category_choices = {'all', 'free_goods', 'subsidized', 'monetary', 'csr', 'other'}
    delivery_choices = {'all', 'pickup', 'delivery'}
    quantity_choices = {'all', '1-100', '101-500', '501-1000', '1001-5000', '5001_or_more'}
    amount_choices = {'all', '1-10000', '10001-50000', '50001-100000', '100001-500000', '500001_or_more'}
    sort_choices = {'newest', 'oldest'}

    legacy_category_map = {'free': 'free_goods', 'money': 'monetary'}
    legacy_amount_map = {'small': '1-10000', 'medium': '10001-50000', 'large': '50001_or_more'}

    status_filter = (request.GET.get('status') or 'all').strip()
    type_filter = (request.GET.get('type') or '').strip()
    category_filter = (request.GET.get('category') or '').strip()
    delivery_filter = (request.GET.get('delivery') or 'all').strip().lower()
    legacy_type_filter = (request.GET.get('donation_type') or '').strip()
    quantity_range = (request.GET.get('quantity_range') or 'all').strip()
    amount_range = (request.GET.get('amount_range') or 'all').strip()
    sort_order = (request.GET.get('sort') or 'newest').strip()
    search_query = (request.GET.get('search') or '').strip()

    # Backward compatibility for old query params.
    if not type_filter and category_filter in type_choices:
        type_filter = category_filter
        category_filter = ''
    if not category_filter and legacy_type_filter:
        category_filter = legacy_category_map.get(legacy_type_filter, legacy_type_filter)
    if type_filter in {'free', 'free_goods', 'subsidized', 'money'} and not category_filter:
        category_filter = legacy_category_map.get(type_filter, type_filter)
        type_filter = ''
    if amount_range == 'all':
        legacy_amount = (request.GET.get('amount') or '').strip()
        if legacy_amount in legacy_amount_map:
            amount_range = legacy_amount_map[legacy_amount]

    if status_filter not in status_choices:
        status_filter = 'all'
    if not type_filter:
        type_filter = 'all'
    if type_filter not in type_choices:
        type_filter = 'all'
    if not category_filter:
        category_filter = 'all'
    category_filter = legacy_category_map.get(category_filter, category_filter)
    if category_filter not in category_choices:
        category_filter = 'all'
    if delivery_filter in {'both', 'dropoff'}:
        delivery_filter = 'delivery'
    if delivery_filter not in delivery_choices:
        delivery_filter = 'all'
    if quantity_range not in quantity_choices:
        quantity_range = 'all'
    if amount_range not in amount_choices:
        amount_range = 'all'
    if sort_order not in sort_choices:
        sort_order = 'newest'

    return {
        'status_filter': status_filter,
        'type_filter': type_filter,
        'category_filter': category_filter,
        'delivery_filter': delivery_filter,
        'quantity_range': quantity_range,
        'amount_range': amount_range,
        'sort_order': sort_order,
        'search_query': search_query,
    }


def _recipient_unspecified_base_queryset(recipient_profile, status_filter):
    from .models import UnspecifiedDonationManagement

    donations = _annotate_recipient_unspecified_decline_state(
        UnspecifiedDonationManagement.objects.all(),
        recipient_profile,
    )

    if status_filter == 'available':
        donations = donations.filter(
            foodbank_status='accepted_by_foodbank'
        ).filter(
            Q(recipient_status='pending_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
            Q(recipient_status='declined_by_recipient', _declined_by_me=False, _declined_by_me_legacy=False)
        )
    elif status_filter == 'my_accepted':
        donations = donations.filter(
            accepted_by_recipient=recipient_profile,
            recipient_status='accepted_by_recipient',
        )
    elif status_filter == 'received':
        donations = donations.filter(
            accepted_by_recipient=recipient_profile,
            recipient_status='received',
        )
    elif status_filter == 'declined':
        donations = donations.filter(
            foodbank_status='accepted_by_foodbank'
        ).filter(
            Q(_declined_by_me=True) | Q(_declined_by_me_legacy=True)
        )
    else:
        donations = donations.filter(
            Q(
                foodbank_status='accepted_by_foodbank',
                recipient_status='pending_recipient',
                _declined_by_me=False,
                _declined_by_me_legacy=False,
            ) |
            Q(
                foodbank_status='accepted_by_foodbank',
                recipient_status='declined_by_recipient',
                _declined_by_me=False,
                _declined_by_me_legacy=False,
            ) |
            Q(
                foodbank_status='accepted_by_foodbank',
            ) & (Q(_declined_by_me=True) | Q(_declined_by_me_legacy=True)) |
            Q(accepted_by_recipient=recipient_profile, recipient_status='accepted_by_recipient') |
            Q(accepted_by_recipient=recipient_profile, recipient_status='received')
        )

    return donations.select_related('donation', 'donation__donor', 'donation__foodbank')


def _apply_recipient_unspecified_filters(donations, filters):
    type_filter = filters['type_filter']
    category_filter = filters['category_filter']
    delivery_filter = filters['delivery_filter']
    quantity_range = filters['quantity_range']
    amount_range = filters['amount_range']
    sort_order = filters['sort_order']
    search_query = filters['search_query']

    if type_filter != 'all':
        donations = donations.filter(donation__donation_category=type_filter)

    if category_filter != 'all':
        if category_filter == 'free_goods':
            donations = donations.filter(donation__donation_type='item', donation__donation_mode='free')
        elif category_filter == 'subsidized':
            donations = donations.filter(
                Q(donation__donation_type='subsidized') |
                Q(donation__donation_type='item', donation__donation_mode='subsidized')
            )
        elif category_filter == 'monetary':
            donations = donations.filter(
                Q(donation__donation_type='money') |
                Q(donation__donation_category='monetary')
            )
        elif category_filter == 'csr':
            donations = donations.filter(
                Q(donation__donation_type='csr') |
                Q(donation__donation_category='csr')
            )
        elif category_filter == 'other':
            donations = donations.filter(
                Q(donation__donation_type='other') |
                Q(donation__donation_category='other')
            )

    if delivery_filter in ('pickup', 'delivery'):
        if delivery_filter == 'delivery':
            donations = donations.filter(donation__delivery_method__in=['delivery', 'dropoff'])
        else:
            donations = donations.filter(donation__delivery_method='pickup')

    quantity_range_map = {
        '1-100': (1, 100),
        '101-500': (101, 500),
        '501-1000': (501, 1000),
        '1001-5000': (1001, 5000),
        '5001_or_more': (5001, None),
    }
    qty_min, qty_max = quantity_range_map.get(quantity_range, (None, None))
    if qty_min is not None:
        donations = donations.filter(
            Q(donation__quantity__gte=qty_min) |
            Q(donation__subsidized_quantity__gte=qty_min)
        )
    if qty_max is not None:
        donations = donations.filter(
            Q(donation__quantity__lte=qty_max) |
            Q(donation__subsidized_quantity__lte=qty_max)
        )

    amount_range_map = {
        '1-10000': (1, 10000),
        '10001-50000': (10001, 50000),
        '50001-100000': (50001, 100000),
        '100001-500000': (100001, 500000),
        '500001_or_more': (500001, None),
    }
    amt_min, amt_max = amount_range_map.get(amount_range, (None, None))
    if amt_min is not None:
        donations = donations.filter(
            Q(donation__amount__gte=amt_min) |
            Q(donation__subsidized_price__gte=amt_min)
        )
    if amt_max is not None:
        donations = donations.filter(
            Q(donation__amount__lte=amt_max) |
            Q(donation__subsidized_price__lte=amt_max)
        )

    if search_query:
        donations = donations.filter(
            Q(donation__foodbank__foodbank_name__icontains=search_query) |
            Q(donation__foodbank__address__icontains=search_query) |
            Q(donation__item_name__icontains=search_query) |
            Q(donation__message__icontains=search_query) |
            Q(donation__csr_description__icontains=search_query) |
            Q(donation__other_description__icontains=search_query)
        )

    if sort_order == 'oldest':
        return donations.order_by('foodbank_reviewed_at', 'created_at')
    return donations.order_by('-foodbank_reviewed_at', '-created_at')


@login_required
def recipient_available_donations(request):
    """View for organization recipients to see donations available for acceptance."""
    if request.user.user_type != 'RECIPIENT':
        messages.error(request, 'Access denied. Recipient users only.')
        return redirect('dashboard')

    recipient_profile = request.user.recipient_profile
    if not recipient_profile.is_organization:
        messages.error(request, 'This feature is only available for organization recipients.')
        return redirect('dashboard')

    filters = _recipient_unspecified_filters(request)
    status_filter = filters['status_filter']

    donations = _recipient_unspecified_base_queryset(recipient_profile, status_filter)
    donations = _apply_recipient_unspecified_filters(donations, filters)

    from .models import UnspecifiedDonationManagement
    available_count = _annotate_recipient_unspecified_decline_state(
        UnspecifiedDonationManagement.objects.filter(
            foodbank_status='accepted_by_foodbank',
        ),
        recipient_profile,
    ).filter(
        Q(recipient_status='pending_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
        Q(recipient_status='declined_by_recipient', _declined_by_me=False, _declined_by_me_legacy=False)
    ).count()
    my_accepted_count = UnspecifiedDonationManagement.objects.filter(
        accepted_by_recipient=recipient_profile,
        recipient_status='accepted_by_recipient'
    ).count()
    received_count = UnspecifiedDonationManagement.objects.filter(
        accepted_by_recipient=recipient_profile,
        recipient_status='received'
    ).count()

    paginator = Paginator(donations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_obj.object_list = _attach_recipient_unspecified_effective_status(page_obj.object_list, recipient_profile)

    context = {
        'donations': page_obj,
        'page_obj': page_obj,
        'status': status_filter,
        'type_filter': filters['type_filter'],
        'category_filter': filters['category_filter'],
        'delivery_filter': filters['delivery_filter'],
        'quantity_range': filters['quantity_range'],
        'amount_range': filters['amount_range'],
        'sort_order': filters['sort_order'],
        'search_query': filters['search_query'],
        # Backward-compatible context keys
        'selected_category': filters['type_filter'],
        'donation_type_filter': filters['category_filter'],
        'amount_filter': filters['amount_range'],
        'available_count': available_count,
        'my_accepted_count': my_accepted_count,
        'received_count': received_count,
        'total_count': available_count + my_accepted_count + received_count,
    }
    return render(request, 'recipient/available_donations.html', context)





@login_required

def export_available_donations(request, format):

    """Export available donations in PDF, CSV, or Excel format"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient users only.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if not recipient_profile.is_organization:

        messages.error(request, 'This feature is only available for organization recipients.')

        return redirect('dashboard')

    

    filters = _recipient_unspecified_filters(request)
    status_filter = filters['status_filter']

    donations = _recipient_unspecified_base_queryset(recipient_profile, status_filter)
    donations = _apply_recipient_unspecified_filters(donations, filters)

    # Get all records (no pagination for export)
    donations_data = _attach_recipient_unspecified_effective_status(list(donations), recipient_profile)

    

    if format.lower() == 'pdf':
        return available_donations_pdf_report(request, donations_data, recipient_profile)
    elif format.lower() == 'csv':
        return available_donations_csv_report(request, donations_data, recipient_profile)
    elif format.lower() == 'excel':
        return available_donations_excel_report(request, donations_data, recipient_profile)

    else:

        messages.error(request, 'Invalid export format.')

        return redirect('recipient_available_donations')





def export_available_donations_pdf(request, donations_data, recipient_profile):

    """Generate PDF report for available donations matching template structure"""

    from reportlab.lib.pagesizes import A4, landscape

    from reportlab.lib import colors

    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    from reportlab.lib.units import inch

    from reportlab.platypus import Table, TableStyle, Paragraph

    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    import datetime

    from django.http import HttpResponse

    from io import BytesIO

    from .report_utils import build_pdf_document



    # PDF setup

    response = HttpResponse(content_type='application/pdf')

    recipient_name = recipient_profile.full_name or recipient_profile.user.email
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_unspecified_donations_{timestamp}.pdf"'



    # Create elements list

    elements = []

    styles = getSampleStyleSheet()

    

    # Custom styles

    title_style = ParagraphStyle(

        'CustomTitle',

        parent=styles['Heading1'],

        fontSize=16,

        textColor=colors.HexColor('#1F4E78'),

        spaceAfter=12,

        alignment=TA_CENTER,

    )

    

    wrap = ParagraphStyle(

        'Wrap',

        parent=styles['BodyText'],

        fontSize=7,

        leading=9,

        wordWrap='CJK',

    )



    # Title

    elements.append(Paragraph("Unspecified Donations Report", title_style))

    elements.append(Paragraph(f"<b>Recipient:</b> {recipient_name}", styles['Normal']))

    elements.append(Paragraph(f"<b>Generated:</b> {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))

    elements.append(Paragraph("<br/><br/>", styles['Normal']))



    # Table data - 11 columns (excluding Actions)

    data = [['S/No', 'Type', 'Category', 'Description', 'Foodbank',

             'Quantity', 'Delivery', 'Location', 'Status', 'Date', 'Notes']]



    for idx, item in enumerate(donations_data, 1):

        don = item.donation

        

        # Type - donation type + item name for items

        if don.donation_type == 'item':

            if don.donation_mode == 'free':

                type_display = '<b>Free</b>'

            elif don.donation_mode == 'subsidized':

                type_display = '<b>Subsidized</b>'

            else:

                type_display = don.get_donation_mode_display()

            item_name = don.item_name or "General items"

            type_display += f"<br/><font size='6' color='#6b7280'>{item_name[:24]}</font>"

        else:

            type_display = don.get_donation_type_display()



        # Category with CSR subcategory

        category_map = {'food': 'Food', 'non_food': 'Non-Food', 'monetary': 'Monetary', 'csr': 'CSR'}

        category_display = category_map.get(don.donation_category, 'Other')

        if don.donation_category == 'csr':

            if don.csr_subcategory == 'other' and don.csr_custom_subcategory:

                category_display += f"<br/><font size='6'>{don.csr_custom_subcategory}</font>"

            elif don.csr_subcategory:

                category_display += f"<br/><font size='6'>{don.get_csr_subcategory_display()}</font>"



        # Description

        if don.message:

            description = don.message[:60]

        elif don.other_description:

            description = don.other_description[:60]

        elif don.csr_description:

            description = don.csr_description[:60]

        elif don.donation_type == 'item':

            description = (don.item_name or "General items")[:60]

        else:

            description = "No description"



        # Foodbank

        fb_name = don.foodbank.foodbank_name[:25] if don.foodbank else ""

        fb_contact = (don.foodbank.user.email or don.foodbank.contact_person or "")[:25] if don.foodbank else ""

        foodbank_display = f"<b>{fb_name}</b><br/><font size='6' color='#6b7280'>{fb_contact}</font>"



        # Quantity

        if don.donation_type == 'item':

            quantity_display = f"<b>{don.quantity}</b> {don.quantity_unit}"

        elif don.donation_type == 'money':

            quantity_display = f"<b>KES {don.amount:,.0f}</b>"

        elif don.donation_type in ['csr', 'other']:

            if don.quantity:

                quantity_display = f"<b>{don.quantity}</b> {don.quantity_unit or 'units'}"

                if don.amount:

                    quantity_display += f"<br/><font size='6' color='#6b7280'>(KES {don.amount:,.0f})</font>"

            elif don.amount:

                quantity_display = f"<b>KES {don.amount:,.0f}</b>"

            else:

                quantity_display = "Not specified"

        else:

            quantity_display = "-"



        # Delivery

        delivery_display = don.get_delivery_method_display() if don.delivery_method else "Not specified"



        # Location

        location_display = (don.foodbank.address or "Location not provided")[:30] if don.foodbank else "Not provided"



        # Status

        status_display = item.get_recipient_status_display()



        # Date

        date_ago = ""

        date_formatted = ""

        if item.foodbank_reviewed_at:

            date_formatted = item.foodbank_reviewed_at.strftime('%b %d, %Y')

            date_display = date_formatted

        else:

            date_display = "-"



        # Notes - combine all three notes

        donor_note = don.message or don.other_description or don.csr_description or "No donor notes provided."

        donor_note = donor_note[:60]

        

        fb_note = item.foodbank_decline_reason[:60] if item.foodbank_decline_reason else "No foodbank notes."

        

        recipient_note = (item.recipient_notes or "No notes yet.")[:60]

        

        notes_display = f"<font size='6' color='#6b7280'>DONOR:</font> {donor_note}<br/>"

        notes_display += f"<font size='6' color='#6b7280'>FOODBANK:</font> {fb_note}<br/>"

        notes_display += f"<font size='6' color='#6b7280'>MY NOTE:</font> {recipient_note}"



        data.append([

            str(idx),

            Paragraph(type_display, wrap),

            Paragraph(category_display, wrap),

            Paragraph(description, wrap),

            Paragraph(foodbank_display, wrap),

            Paragraph(quantity_display, wrap),

            Paragraph(delivery_display, wrap),

            Paragraph(location_display, wrap),

            Paragraph(status_display, wrap),

            Paragraph(date_display, wrap),

            Paragraph(notes_display, wrap),

        ])



    # Create table

    table = Table(data, colWidths=[0.4*inch, 0.9*inch, 0.8*inch, 1.2*inch, 1.1*inch,

                                    0.9*inch, 0.7*inch, 1.0*inch, 0.9*inch, 0.7*inch, 1.5*inch])



    # Table style

    table.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),

        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('FONTSIZE', (0, 0), (-1, 0), 8),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        ('TOPPADDING', (0, 0), (-1, 0), 8),

        ('BACKGROUND', (0, 1), (-1, -1), colors.white),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        ('FONTSIZE', (0, 1), (-1, -1), 7),

        ('VALIGN', (0, 0), (-1, -1), 'TOP'),

        ('LEFTPADDING', (0, 0), (-1, -1), 4),

        ('RIGHTPADDING', (0, 0), (-1, -1), 4),

    ]))



    elements.append(table)



    # Summary

    elements.append(Paragraph("<br/>", styles['Normal']))

    elements.append(Paragraph(f"<b>Total Records:</b> {len(donations_data)}", styles['Normal']))



    # Build PDF

    return build_pdf_document(elements, "available_donations", recipient_name)





def export_available_donations_csv(request, donations_data, recipient_profile):

    """Generate CSV report for available donations"""

    import csv

    import datetime

    from django.http import HttpResponse



    response = HttpResponse(content_type='text/csv')

    recipient_name = recipient_profile.full_name or recipient_profile.user.email

    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")

    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_available_donations_{timestamp}.csv"'



    writer = csv.writer(response, quoting=csv.QUOTE_ALL)



    # Metadata

    writer.writerow(['Recipient', recipient_name])

    writer.writerow(['Report Generated', timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')])

    writer.writerow(['Total Records', len(donations_data)])

    writer.writerow([])



    # Headers - detailed columns

    headers = [

        'S/No', 'Type', 'Item Name', 'Donation Mode', 'Category', 'CSR Subcategory',

        'Description', 'Foodbank Name', 'Foodbank Contact', 'Foodbank Address',

        'Quantity', 'Unit', 'Amount (KES)', 'Delivery Method', 'Status',

        'Date Reviewed', 'Donor Note', 'Foodbank Note', 'My Note'

    ]

    writer.writerow(headers)



    for idx, item in enumerate(donations_data, 1):

        don = item.donation



        # Type

        type_display = don.get_donation_type_display()



        # Item name

        item_name = don.item_name or "" if don.donation_type == 'item' else ""



        # Donation mode

        donation_mode = don.get_donation_mode_display() if don.donation_type == 'item' else ""



        # Category

        category_map = {'food': 'Food', 'non_food': 'Non-Food', 'monetary': 'Monetary', 'csr': 'CSR'}

        category_display = category_map.get(don.donation_category, 'Other')



        # CSR Subcategory

        csr_subcat = ""

        if don.donation_category == 'csr':

            if don.csr_subcategory == 'other' and don.csr_custom_subcategory:

                csr_subcat = don.csr_custom_subcategory

            elif don.csr_subcategory:

                csr_subcat = don.get_csr_subcategory_display()



        # Description

        if don.message:

            description = don.message

        elif don.other_description:

            description = don.other_description

        elif don.csr_description:

            description = don.csr_description

        elif don.donation_type == 'item':

            description = don.item_name or "General items"

        else:

            description = ""

        description = description.replace("\n", " ").replace("\r", "")



        # Foodbank info

        fb_name = don.foodbank.foodbank_name if don.foodbank else ""

        fb_contact = don.foodbank.user.email or don.foodbank.contact_person or "" if don.foodbank else ""

        fb_address = don.foodbank.address or "" if don.foodbank else ""



        # Quantity and amount

        quantity = don.quantity or "" if don.donation_type in ['item', 'csr', 'other'] else ""

        unit = don.quantity_unit or "" if don.donation_type in ['item', 'csr', 'other'] else ""

        amount = f"{don.amount:,.2f}" if don.amount else ""



        # Delivery

        delivery = don.get_delivery_method_display() if don.delivery_method else "Not specified"



        # Status

        status = item.get_recipient_status_display()



        # Date

        date_reviewed = item.foodbank_reviewed_at.strftime('%Y-%m-%d %H:%M') if item.foodbank_reviewed_at else ""



        # Notes

        donor_note = (don.message or don.other_description or don.csr_description or "").replace("\n", " ").replace("\r", "")

        fb_note = (item.foodbank_decline_reason or "").replace("\n", " ").replace("\r", "")

        recipient_note = (item.recipient_notes or "").replace("\n", " ").replace("\r", "")



        writer.writerow([

            idx,

            type_display,

            item_name,

            donation_mode,

            category_display,

            csr_subcat,

            description,

            fb_name,

            fb_contact,

            fb_address,

            quantity,

            unit,

            amount,

            delivery,

            status,

            date_reviewed,

            donor_note,

            fb_note,

            recipient_note,

        ])



    return response





def export_available_donations_excel(request, donations_data, recipient_profile):

    """Generate Excel report for available donations with enhanced formatting"""

    import datetime

    from django.http import HttpResponse

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



    wb = Workbook()

    ws = wb.active

    ws.title = "Available Donations"



    recipient_name = recipient_profile.full_name or recipient_profile.user.email



    # â”€â”€ Metadata section (rows 1-3) â”€â”€

    ws['A1'] = 'Recipient'

    ws['B1'] = recipient_name

    ws['A2'] = 'Report Generated'

    ws['B2'] = timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')

    ws['A3'] = 'Total Records'

    ws['B3'] = len(donations_data)



    # Style metadata

    for row in range(1, 4):

        ws[f'A{row}'].font = Font(bold=True)



    # â”€â”€ Headers (row 5) â”€â”€

    headers = [

        'S/No', 'Type', 'Item Name', 'Donation Mode', 'Category', 'CSR Subcategory',

        'Description', 'Foodbank Name', 'Foodbank Contact', 'Foodbank Address',

        'Quantity', 'Unit', 'Amount (KES)', 'Delivery Method', 'Status',

        'Date Reviewed', 'Donor Note', 'Foodbank Note', 'My Note'

    ]



    header_fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True, size=10)

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(

        left=Side(style='thin'), right=Side(style='thin'),

        top=Side(style='thin'), bottom=Side(style='thin'),

    )

    alt_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")



    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=5, column=col_num)

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = header_alignment

        cell.border = thin_border



    # â”€â”€ Data rows â”€â”€

    row_num = 6

    for idx, item in enumerate(donations_data, 1):

        don = item.donation



        # Type

        type_display = don.get_donation_type_display()



        # Item name

        item_name = don.item_name or "" if don.donation_type == 'item' else ""



        # Donation mode

        donation_mode = don.get_donation_mode_display() if don.donation_type == 'item' else ""



        # Category

        category_map = {'food': 'Food', 'non_food': 'Non-Food', 'monetary': 'Monetary', 'csr': 'CSR'}

        category_display = category_map.get(don.donation_category, 'Other')



        # CSR Subcategory

        csr_subcat = ""

        if don.donation_category == 'csr':

            if don.csr_subcategory == 'other' and don.csr_custom_subcategory:

                csr_subcat = don.csr_custom_subcategory

            elif don.csr_subcategory:

                csr_subcat = don.get_csr_subcategory_display()



        # Description

        if don.message:

            description = don.message

        elif don.other_description:

            description = don.other_description

        elif don.csr_description:

            description = don.csr_description

        elif don.donation_type == 'item':

            description = don.item_name or "General items"

        else:

            description = ""



        # Foodbank info

        fb_name = don.foodbank.foodbank_name if don.foodbank else ""

        fb_contact = don.foodbank.user.email or don.foodbank.contact_person or "" if don.foodbank else ""

        fb_address = don.foodbank.address or "" if don.foodbank else ""



        # Quantity and amount

        quantity = don.quantity or "" if don.donation_type in ['item', 'csr', 'other'] else ""

        unit = don.quantity_unit or "" if don.donation_type in ['item', 'csr', 'other'] else ""

        amount = don.amount if don.amount else ""



        # Delivery

        delivery = don.get_delivery_method_display() if don.delivery_method else "Not specified"



        # Status

        status = item.get_recipient_status_display()



        # Date

        date_reviewed = item.foodbank_reviewed_at.strftime('%Y-%m-%d %H:%M') if item.foodbank_reviewed_at else ""



        # Notes

        donor_note = don.message or don.other_description or don.csr_description or ""

        fb_note = item.foodbank_decline_reason or ""

        recipient_note = item.recipient_notes or ""



        row_data = [

            idx,

            type_display,

            item_name,

            donation_mode,

            category_display,

            csr_subcat,

            description,

            fb_name,

            fb_contact,

            fb_address,

            quantity,

            unit,

            amount,

            delivery,

            status,

            date_reviewed,

            donor_note,

            fb_note,

            recipient_note,

        ]



        is_alt = idx % 2 == 0

        for col_num, value in enumerate(row_data, 1):

            cell = ws.cell(row=row_num, column=col_num)

            cell.value = value

            cell.border = thin_border

            cell.alignment = Alignment(vertical='top', wrap_text=True)

            if is_alt:

                cell.fill = alt_fill



        row_num += 1



    # â”€â”€ Column widths â”€â”€

    column_widths = {

        'A': 6,   # S/No

        'B': 12,  # Type

        'C': 18,  # Item Name

        'D': 12,  # Donation Mode

        'E': 12,  # Category

        'F': 15,  # CSR Subcategory

        'G': 30,  # Description

        'H': 20,  # Foodbank Name

        'I': 20,  # Foodbank Contact

        'J': 25,  # Foodbank Address

        'K': 10,  # Quantity

        'L': 10,  # Unit

        'M': 15,  # Amount (KES)

        'N': 15,  # Delivery Method

        'O': 18,  # Status

        'P': 16,  # Date Reviewed

        'Q': 30,  # Donor Note

        'R': 25,  # Foodbank Note

        'S': 25,  # My Note

    }

    for col, width in column_widths.items():

        ws.column_dimensions[col].width = width

    ws.row_dimensions[5].height = 35



    # â”€â”€ Add auto-filter to all columns â”€â”€

    if len(donations_data) > 0:

        ws.auto_filter.ref = f"A5:S{row_num - 1}"



    # â”€â”€ Freeze top rows (header) â”€â”€

    ws.freeze_panes = "A6"



    # â”€â”€ Response â”€â”€

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)



    response = HttpResponse(

        buffer,

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

    )

    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")

    response['Content-Disposition'] = f'attachment; filename="{recipient_name}_available_donations_{timestamp}.xlsx"'

    return response





@login_required

def recipient_accept_unspecified_donation(request, donation_id):

    """Organization recipient accepts an available donation"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient users only.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if not recipient_profile.is_organization:

        messages.error(request, 'This feature is only available for organization recipients.')

        return redirect('dashboard')

    

    if request.method != 'POST':
        return _redirect_back_or_default(request, 'recipient_available_donations')

    

    from .models import UnspecifiedDonationManagement

    

    try:
        unspecified = _annotate_recipient_unspecified_decline_state(
            UnspecifiedDonationManagement.objects.filter(
                id=donation_id,
                foodbank_status='accepted_by_foodbank',
                recipient_status__in=['pending_recipient', 'declined_by_recipient'],
            ),
            recipient_profile,
        ).filter(
            Q(recipient_status='pending_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
            Q(recipient_status='declined_by_recipient', _declined_by_me=False, _declined_by_me_legacy=False)
        ).get()
    except UnspecifiedDonationManagement.DoesNotExist:

        messages.error(request, 'Donation not found or already claimed by another recipient.')
        return _redirect_back_or_default(request, 'recipient_available_donations')

    

    notes = request.POST.get('notes', '')

    unspecified.recipient_accept(recipient_profile, notes=notes)

    # Persist recipient-specific response state and clear declined marker when accepted.
    DonationResponse.objects.update_or_create(
        donation=unspecified.donation,
        recipient=recipient_profile,
        defaults={
            'response_type': 'accepted',
            'notes': notes or '',
        },
    )
    if unspecified.donation.declined_by_recipient_id:
        unspecified.donation.declined_by_recipient = None
        unspecified.donation.save(update_fields=['declined_by_recipient'])

    

    messages.success(request, 'You have accepted the donation! Please confirm when you receive it.')
    return _redirect_back_or_default(request, 'recipient_available_donations')





@login_required

def recipient_decline_unspecified_donation(request, donation_id):

    """Organization recipient declines an available donation"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient users only.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if not recipient_profile.is_organization:

        messages.error(request, 'This feature is only available for organization recipients.')

        return redirect('dashboard')

    

    if request.method != 'POST':
        return _redirect_back_or_default(request, 'recipient_available_donations')

    

    from .models import UnspecifiedDonationManagement

    

    try:
        unspecified = _annotate_recipient_unspecified_decline_state(
            UnspecifiedDonationManagement.objects.filter(
                id=donation_id,
                foodbank_status='accepted_by_foodbank',
                recipient_status__in=['pending_recipient', 'declined_by_recipient'],
            ),
            recipient_profile,
        ).filter(
            Q(recipient_status='pending_recipient', _declined_by_me=False, _declined_by_me_legacy=False) |
            Q(recipient_status='declined_by_recipient', _declined_by_me=False, _declined_by_me_legacy=False)
        ).get()
    except UnspecifiedDonationManagement.DoesNotExist:

        messages.error(request, 'Donation not found or already processed.')
        return _redirect_back_or_default(request, 'recipient_available_donations')

    

    decline_reason = request.POST.get('decline_reason', '').strip()

    unspecified.recipient_status = 'declined_by_recipient'

    unspecified.accepted_by_recipient = None

    unspecified.recipient_accepted_at = None

    if decline_reason:

        unspecified.recipient_decline_reason = decline_reason

    unspecified.save()

    unspecified.donation.declined_by_recipient = recipient_profile
    unspecified.donation.save(update_fields=['declined_by_recipient'])

    DonationResponse.objects.update_or_create(
        donation=unspecified.donation,
        recipient=recipient_profile,
        defaults={
            'response_type': 'declined',
            'notes': decline_reason,
        },
    )

    

    messages.info(request, 'Thank you for the update. The donation has been declined for your organization.')
    return _redirect_back_or_default(request, 'recipient_available_donations')





@login_required

def recipient_confirm_unspecified_received(request, donation_id):

    """Recipient confirms they have received the donation"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient users only.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if request.method != 'POST':
        return _redirect_back_or_default(request, 'recipient_available_donations')

    

    from .models import UnspecifiedDonationManagement

    

    try:

        unspecified = UnspecifiedDonationManagement.objects.get(

            id=donation_id,

            accepted_by_recipient=recipient_profile,

            recipient_status='accepted_by_recipient'

        )

    except UnspecifiedDonationManagement.DoesNotExist:

        messages.error(request, 'Donation not found or not accepted by you.')
        return _redirect_back_or_default(request, 'recipient_available_donations')

    

    unspecified.confirm_received()

    

    # Update delivery status on the original donation

    unspecified.donation.delivery_status = 'delivered'

    unspecified.donation.save()

    

    messages.success(request, 'Thank you for confirming receipt of the donation!')
    return _redirect_back_or_default(request, 'recipient_available_donations')





@login_required

def recipient_accept_subsidized_donation(request, donation_id):

    """Recipient accepts an available subsidized donation"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient users only.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if request.method != 'POST':
        return _redirect_back_or_default(request, 'dashboard')

    

    try:

        donation = Donation.objects.get(

            id=donation_id,

            donation_type='subsidized',

            status='accepted',

            accepted_by_recipient__isnull=True

        )

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found or already claimed by another recipient.')
        return _redirect_back_or_default(request, 'dashboard')

    

    notes = (request.POST.get('notes') or request.POST.get('recipient_notes') or '').strip()

    donation.accepted_by_recipient = recipient_profile
    donation.save(update_fields=['accepted_by_recipient'])

    response, created = DonationResponse.objects.get_or_create(
        donation=donation,
        recipient=recipient_profile,
        defaults={
            'response_type': 'accepted',
            'notes': notes,
        }
    )
    if not created:
        response.response_type = 'accepted'
        response.notes = notes
        response.responded_at = timezone.now()
        response.save(update_fields=['response_type', 'notes', 'responded_at'])

    

    messages.success(request, 'You have accepted the subsidized donation! Please confirm when you receive it.')
    return _redirect_back_or_default(request, 'dashboard')





@login_required

def recipient_confirm_subsidized_donation_received(request, donation_id):

    """Recipient confirms they have received the subsidized donation"""

    if request.user.user_type != 'RECIPIENT':

        messages.error(request, 'Access denied. Recipient users only.')

        return redirect('dashboard')

    

    recipient_profile = request.user.recipient_profile

    

    if request.method != 'POST':
        return _redirect_back_or_default(request, 'dashboard')

    

    try:

        donation = Donation.objects.get(

            id=donation_id,

            donation_type='subsidized',

            accepted_by_recipient=recipient_profile

        )

    except Donation.DoesNotExist:

        messages.error(request, 'Donation not found or not accepted by you.')
        return _redirect_back_or_default(request, 'dashboard')

    

    donation.delivery_status = 'delivered'

    donation.save()

    

    messages.success(request, 'Thank you for confirming receipt of the subsidized donation!')
    return _redirect_back_or_default(request, 'dashboard')





@login_required

def foodbank_inventory(request):

    """View for foodbanks to manage their inventory of available donations (item, money, subsidized)"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied. Foodbank privileges required.')

        return redirect('dashboard')



    foodbank_profile = request.user.foodbank_profile



    base_stock = Donation.objects.filter(

        foodbank=foodbank_profile,

        donation_type__in=['item', 'money', 'subsidized'],

        status='accepted'

    ).select_related('donor', 'donor__donor_profile').annotate(

        allocated_amount=Coalesce(

            Sum('allocations__amount', filter=Q(allocations__declined_by_recipient=False)),

            Value(0)

        )

    ).annotate(

        remaining_amount=Case(

            When(donation_type='money', then=Coalesce(F('amount'), Value(0)) - F('allocated_amount')),

            When(donation_type='subsidized', then=Coalesce(F('subsidized_price'), Value(0)) - F('allocated_amount')),

            default=Value(0),

            output_field=DecimalField(max_digits=15, decimal_places=2)

        )

    )

    donations = base_stock.order_by('-donated_at')

    search_query = request.GET.get('search', '')

    category_filter = request.GET.get('category', 'all')

    status_filter = request.GET.get('status', 'all')

    type_filter = request.GET.get('type', 'all')



    if search_query:

        donations = donations.filter(

            Q(item_name__icontains=search_query) |

            Q(description__icontains=search_query) |

            Q(message__icontains=search_query) |

            Q(other_description__icontains=search_query) |

            Q(subsidized_product_type__icontains=search_query) |

            Q(donor__donor_profile__full_name__icontains=search_query) |

            Q(donor__donor_profile__organization_name__icontains=search_query) |

            Q(donor__email__icontains=search_query)

        )

    if category_filter != 'all':

        donations = donations.filter(donation_category=category_filter)

    if type_filter != 'all':

        donations = donations.filter(donation_type=type_filter)

    if status_filter == 'available':

        donations = donations.filter(

            Q(donation_type='item', is_allocated=False) |

            Q(donation_type='money', remaining_amount__gt=0) |

            Q(donation_type='subsidized', remaining_amount__gt=0)

        )

    elif status_filter == 'allocated':

        donations = donations.filter(

            Q(donation_type='item', is_allocated=True) |

            Q(donation_type='money', remaining_amount__lte=0) |

            Q(donation_type='subsidized', remaining_amount__lte=0)

        )



    paginator = Paginator(donations, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)



    stock_base = Donation.objects.filter(

        foodbank=foodbank_profile,

        donation_type__in=['item', 'money', 'subsidized'],

        status='accepted'

    ).annotate(

        allocated_amount=Coalesce(

            Sum('allocations__amount', filter=Q(allocations__declined_by_recipient=False)),

            Value(0)

        )

    ).annotate(

        remaining_amount=Case(

            When(donation_type='money', then=Coalesce(F('amount'), Value(0)) - F('allocated_amount')),

            When(donation_type='subsidized', then=Coalesce(F('subsidized_price'), Value(0)) - F('allocated_amount')),

            default=Value(0),

            output_field=DecimalField(max_digits=15, decimal_places=2)

        )

    )

    total_donations = stock_base.count()

    available_donations = stock_base.filter(

        Q(donation_type='item', is_allocated=False) |

        Q(donation_type='money', remaining_amount__gt=0) |

        Q(donation_type='subsidized', remaining_amount__gt=0)

    ).count()

    allocated_donations = stock_base.filter(

        Q(donation_type='item', is_allocated=True) |

        Q(donation_type='money', remaining_amount__lte=0) |

        Q(donation_type='subsidized', remaining_amount__lte=0)

    ).count()

    categories_count = stock_base.values('donation_category').distinct().count()



    context = {

        'donations': page_obj,

        'total_donations': total_donations,

        'available_donations': available_donations,

        'allocated_donations': allocated_donations,

        'categories_count': categories_count,

        'search_query': search_query,

        'category_filter': category_filter,

        'status_filter': status_filter,

        'type_filter': type_filter,

    }

    return render(request, 'authentication/foodbank_inventory.html', context)





@login_required

def export_foodbank_inventory(request):

    """Export foodbank inventory as Excel, PDF, or CSV"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    from django.http import HttpResponse

    import csv

    import io

    from datetime import datetime

    from reportlab.lib.pagesizes import A4, landscape

    from reportlab.lib import colors

    from reportlab.lib.units import inch

    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.utils import get_column_letter

    

    foodbank_profile = request.user.foodbank_profile

    export_format = request.GET.get('format', 'csv')

    

    # Same base as foodbank_inventory: item, money, subsidized

    donations = Donation.objects.filter(

        foodbank=foodbank_profile,

        donation_type__in=['item', 'money', 'subsidized'],

        status='accepted'

    ).select_related('donor', 'donor__donor_profile').annotate(

        allocated_amount=Coalesce(

            Sum('allocations__amount', filter=Q(allocations__declined_by_recipient=False)),

            Value(0)

        )

    ).annotate(

        remaining_amount=Case(

            When(donation_type='money', then=Coalesce(F('amount'), Value(0)) - F('allocated_amount')),

            When(donation_type='subsidized', then=Coalesce(F('subsidized_price'), Value(0)) - F('allocated_amount')),

            default=Value(0),

            output_field=DecimalField(max_digits=15, decimal_places=2)

        )

    ).order_by('-donated_at')

    search_query = request.GET.get('search', '')

    category_filter = request.GET.get('category', 'all')

    status_filter = request.GET.get('status', 'all')

    type_filter = request.GET.get('type', 'all')

    if search_query:

        donations = donations.filter(

            Q(item_name__icontains=search_query) |

            Q(description__icontains=search_query) |

            Q(message__icontains=search_query) |

            Q(other_description__icontains=search_query) |

            Q(subsidized_product_type__icontains=search_query) |

            Q(donor__donor_profile__full_name__icontains=search_query) |

            Q(donor__donor_profile__organization_name__icontains=search_query) |

            Q(donor__email__icontains=search_query)

        )

    if category_filter != 'all':

        donations = donations.filter(donation_category=category_filter)

    if type_filter != 'all':

        donations = donations.filter(donation_type=type_filter)

    if status_filter == 'available':

        donations = donations.filter(

            Q(donation_type='item', is_allocated=False) |

            Q(donation_type='money', remaining_amount__gt=0) |

            Q(donation_type='subsidized', remaining_amount__gt=0)

        )

    elif status_filter == 'allocated':

        donations = donations.filter(

            Q(donation_type='item', is_allocated=True) |

            Q(donation_type='money', remaining_amount__lte=0) |

            Q(donation_type='subsidized', remaining_amount__lte=0)

        )

    if export_format == 'csv':

        return export_inventory_csv(donations, foodbank_profile)

    elif export_format == 'pdf':

        return export_inventory_pdf(donations, foodbank_profile)

    elif export_format == 'excel':

        return export_inventory_excel(donations, foodbank_profile)

    else:

        return HttpResponse("Invalid format", status=400)





def export_inventory_csv(donations, foodbank_profile):

    """Export inventory as CSV"""

    import csv

    from django.http import HttpResponse

    from datetime import datetime

    

    response = HttpResponse(content_type='text/csv')

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')

    filename = f'{foodbank_profile.foodbank_name}_inventory_{timestamp}.csv'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    

    writer = csv.writer(response)

    

    # Header

    writer.writerow([f'Inventory Report - {foodbank_profile.foodbank_name}'])

    writer.writerow([f'Generated on: {timezone.localtime().strftime("%B %d, %Y at %I:%M %p")}'])

    writer.writerow([])

    

    # Column headers

    writer.writerow([

        'S/No', 'Type', 'Donor', 'Description', 'Category', 'Quantity/Amount', 'Unit',

        'Status', 'Received Date', 'Notes'

    ])

    for idx, donation in enumerate(donations, 1):

        donor_name = donation.donor.donor_profile.full_name if getattr(donation.donor, 'donor_profile', None) and donation.donor.donor_profile.full_name else donation.donor.email

        if donation.donation_type == 'item':

            status = 'Out of Stock' if donation.is_allocated else 'In Stock'

            desc = donation.item_name or '-'

            qty_amt = donation.quantity or 0

            unit = donation.get_quantity_unit_display() if donation.quantity_unit else '-'

        elif donation.donation_type == 'money':

            status = 'In Stock' if getattr(donation, 'remaining_amount', 0) > 0 else 'Out of Stock'

            desc = 'Monetary donation'

            qty_amt = float(donation.amount or 0)

            unit = 'KES'

        else:

            status = 'In Stock' if getattr(donation, 'remaining_amount', 0) > 0 else 'Out of Stock'

            desc = donation.subsidized_product_type or 'Subsidized goods'

            qty_amt = float(donation.subsidized_price or 0)

            unit = 'KES'

        notes = (donation.description or donation.message or '-')[:50]

        writer.writerow([

            idx,

            donation.get_donation_type_display(),

            donor_name,

            desc,

            donation.get_donation_category_display() if donation.donation_category else '-',

            qty_amt,

            unit,

            status,

            donation.donated_at.strftime('%Y-%m-%d'),

            notes

        ])

    

    return response





def export_inventory_pdf(donations, foodbank_profile):

    """Export inventory as PDF"""

    from django.http import HttpResponse

    from datetime import datetime

    import io

    from reportlab.lib.pagesizes import A4, landscape

    from reportlab.lib import colors

    from reportlab.lib.units import inch

    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)

    elements = []

    

    # Styles

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(

        'CustomTitle',

        parent=styles['Heading1'],

        fontSize=18,

        textColor=colors.HexColor('#10b981'),

        spaceAfter=30,

        alignment=1

    )

    

    # Title

    title = Paragraph(f"Inventory Report - {foodbank_profile.foodbank_name}", title_style)

    elements.append(title)

    

    # Report info

    report_info = Paragraph(

        f"Generated on: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')}<br/>"

        f"Total Items: {donations.count()}",

        styles['Normal']

    )

    elements.append(report_info)

    elements.append(Spacer(1, 20))

    

    # Table data

    data = [['S/No', 'Type', 'Donor', 'Description', 'Category', 'Qty/Amount', 'Unit', 'Status', 'Date']]

    for idx, donation in enumerate(donations, 1):

        donor_name = donation.donor.donor_profile.full_name if getattr(donation.donor, 'donor_profile', None) and donation.donor.donor_profile.full_name else donation.donor.email

        if donation.donation_type == 'item':

            status = 'Out of Stock' if donation.is_allocated else 'In Stock'

            desc = (donation.item_name or '-')[:20]

            qty_amt = str(donation.quantity or 0)

            unit = (str(donation.quantity_unit) if donation.quantity_unit else '-')[:10]

        elif donation.donation_type == 'money':

            status = 'In Stock' if getattr(donation, 'remaining_amount', 0) > 0 else 'Out of Stock'

            desc = 'Monetary'[:20]

            qty_amt = str(int(donation.amount or 0))

            unit = 'KES'

        else:

            status = 'In Stock' if getattr(donation, 'remaining_amount', 0) > 0 else 'Out of Stock'

            desc = (donation.subsidized_product_type or 'Subsidized')[:20]

            qty_amt = str(int(donation.subsidized_price or 0))

            unit = 'KES'

        data.append([

            str(idx),

            donation.get_donation_type_display(),

            donor_name[:20],

            desc,

            (donation.get_donation_category_display() or '-')[:10],

            qty_amt,

            unit,

            status,

            donation.donated_at.strftime('%m/%d/%Y')

        ])

    

    # Create table

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),

        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('FONTSIZE', (0, 0), (-1, 0), 9),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        ('FONTSIZE', (0, 1), (-1, -1), 8),

        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

    ]))

    

    elements.append(table)

    

    # Build PDF

    doc.build(elements)

    

    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')

    filename = f'{foodbank_profile.foodbank_name}_inventory_{timestamp}.pdf'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    

    return response





def export_inventory_excel(donations, foodbank_profile):

    """Export inventory as Excel"""

    from django.http import HttpResponse

    from datetime import datetime

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.utils import get_column_letter

    

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventory"

    

    # Define styles

    header_font = Font(bold=True, color="FFFFFF", size=11)

    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )

    

    # Title

    ws.merge_cells('A1:J1')

    title_cell = ws['A1']

    title_cell.value = f"Inventory Report - {foodbank_profile.foodbank_name}"

    title_cell.font = Font(bold=True, size=14, color="10b981")

    title_cell.alignment = Alignment(horizontal="center")

    

    ws.merge_cells('A2:J2')

    info_cell = ws['A2']

    info_cell.value = f"Generated on: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')} | Total: {donations.count()} (Item, Money, Subsidized)"

    info_cell.alignment = Alignment(horizontal="center")

    

    # Headers

    headers = [

        'S/No', 'Type', 'Donor', 'Description', 'Category', 'Quantity/Amount', 'Unit',

        'Status', 'Received Date', 'Notes'

    ]

    

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=4, column=col_num)

        cell.value = header

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = header_alignment

        cell.border = border

    

    # Data rows

    for idx, donation in enumerate(donations, 1):

        row_num = idx + 4

        donor_name = donation.donor.donor_profile.full_name if getattr(donation.donor, 'donor_profile', None) and donation.donor.donor_profile.full_name else donation.donor.email

        if donation.donation_type == 'item':

            status = 'Out of Stock' if donation.is_allocated else 'In Stock'

            desc = donation.item_name or '-'

            qty_amt = donation.quantity or 0

            unit = donation.quantity_unit or '-'

        elif donation.donation_type == 'money':

            status = 'In Stock' if getattr(donation, 'remaining_amount', 0) > 0 else 'Out of Stock'

            desc = 'Monetary donation'

            qty_amt = float(donation.amount or 0)

            unit = 'KES'

        else:

            status = 'In Stock' if getattr(donation, 'remaining_amount', 0) > 0 else 'Out of Stock'

            desc = donation.subsidized_product_type or 'Subsidized goods'

            qty_amt = float(donation.subsidized_price or 0)

            unit = 'KES'

        notes = (donation.description or donation.message or '-')[:50]

        ws.cell(row=row_num, column=1, value=idx).border = border

        ws.cell(row=row_num, column=2, value=donation.get_donation_type_display()).border = border

        ws.cell(row=row_num, column=3, value=donor_name).border = border

        ws.cell(row=row_num, column=4, value=desc).border = border

        ws.cell(row=row_num, column=5, value=donation.get_donation_category_display() or '-').border = border

        ws.cell(row=row_num, column=6, value=qty_amt).border = border

        ws.cell(row=row_num, column=7, value=unit).border = border

        ws.cell(row=row_num, column=8, value=status).border = border

        ws.cell(row=row_num, column=9, value=donation.donated_at.strftime('%Y-%m-%d')).border = border

        ws.cell(row=row_num, column=10, value=notes).border = border

    

    # Adjust column widths

    column_widths = [8, 12, 25, 25, 12, 14, 10, 12, 12, 40]

    for col_num, width in enumerate(column_widths, 1):

        ws.column_dimensions[get_column_letter(col_num)].width = width

    

    # Create response

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')

    filename = f'{foodbank_profile.foodbank_name}_inventory_{timestamp}.xlsx'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    

    wb.save(response)

    return response





@login_required

def submit_remaining_to_donors(request, request_id):

    req = get_object_or_404(RequestManagement, id=request_id)

    if request.user.user_type != 'FOODBANK':
        messages.error(request, "Access denied.")
        return redirect('dashboard')

    current_foodbank = request.user.foodbank_profile
    if not current_foodbank:
        messages.error(request, "Foodbank profile not found.")
        return _redirect_back_or_default(request, 'foodbank_requests_view')

    if req.is_anonymous:
        if not req.assigned_foodbank or req.assigned_foodbank != current_foodbank:
            messages.error(request, "This anonymous request is not assigned to your foodbank.")
            return _redirect_back_or_default(request, 'foodbank_requests_view')
        foodbank_to_use = req.assigned_foodbank
    else:
        if req.foodbank and req.foodbank != current_foodbank:
            messages.error(request, "This request doesn't belong to your foodbank.")
            return _redirect_back_or_default(request, 'foodbank_requests_view')
        foodbank_to_use = req.foodbank or req.assigned_foodbank or current_foodbank

    if not foodbank_to_use:
        messages.error(request, "Unable to determine foodbank for this request.")
        return _redirect_back_or_default(request, 'foodbank_requests_view')



    if req.awaiting_donors:

        messages.error(request, "This request is already awaiting donor response.")

        return _redirect_back_or_default(request, 'foodbank_requests_view')



    if request.method == 'POST':

        option = request.POST.get('fulfillment_option')



        if option == 'donors':

            remaining_qty = req.quantity - req.quantity_fulfilled
            if remaining_qty <= 0:
                messages.error(request, "This request has no remaining quantity to submit.")
                return _redirect_back_or_default(request, 'foodbank_requests_view')



            # ðŸ”¹ CREATE donor-facing request

            donor_request = FoodBankRequest.objects.create(

                foodbank=foodbank_to_use,

                original_request=req,

                title=req.description,

                description=req.description or "Request from recipient",

                quantity_needed=remaining_qty,

                quantity_unit=req.unit,

                priority='high',

                status='active',

                donation_type=req.request_type,

                deadline=timezone.now() + timedelta(days=30),

                linked_request_management=req

            )



            # ðŸ”¹ Update original request

            req.status = 'submitted'

            req.awaiting_donors = True
            if req.foodbank_id is None:
                req.foodbank = foodbank_to_use

            req.save()



            messages.success(

                request,

                "Remaining quantity submitted to donors and is now visible on donor dashboard."

            )

            return _redirect_back_or_default(request, 'foodbank_requests_view')



        elif option == 'available':

            return redirect('partial_fulfill_request', request_id=req.id)



    return _redirect_back_or_default(request, 'foodbank_requests_view')





@login_required

def accept_request_donation(request, donation_id):

    """Accept a pending donation for a foodbank request"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    donation = get_object_or_404(

        Donation,

        id=donation_id,

        foodbank=request.user.foodbank_profile,

        status='pending',

        foodbank_request__isnull=False

    )

    

    if request.method == 'POST':

        comment = request.POST.get('comment', '').strip()

        

        # Accept the donation

        donation.status = 'accepted'

        if comment:

            donation.decline_message = f"Acceptance note: {comment}"

        donation.save()

        

        # Reference the foodbank request so templates know stock exists

        foodbank_request = donation.foodbank_request

        foodbank_request.save(update_fields=['updated_at'])

        

        # Notify donor

        Notification.objects.create(

            user=donation.donor,

            notification_type='donation_accepted',

            message=f'Your donation for "{foodbank_request.title}" has been accepted by {request.user.foodbank_profile.foodbank_name}. {comment if comment else ""}'

        )

        

        messages.success(request, f'Donation accepted successfully and added to available stock.')

        return _redirect_back_or_default(request, 'foodbank_requests')

    

    next_url = (request.GET.get('next') or request.POST.get('next') or '').strip()
    context = {

        'donation': donation,

        'action': 'accept',
        'next': next_url,

    }

    return render(request, 'authentication/confirm_donation_action.html', context)





@login_required

def reject_request_donation(request, donation_id):

    """Reject a pending donation for a foodbank request"""

    if request.user.user_type != 'FOODBANK':

        messages.error(request, 'Access denied.')

        return redirect('dashboard')

    

    donation = get_object_or_404(

        Donation,

        id=donation_id,

        foodbank=request.user.foodbank_profile,

        status='pending',

        foodbank_request__isnull=False

    )

    

    if request.method == 'POST':

        comment = request.POST.get('comment', '').strip()

        

        if not comment:

            messages.error(request, 'Please provide a reason for rejection.')
            next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
            if next_url:
                return redirect(f"{reverse('reject_request_donation', args=[donation_id])}?next={quote(next_url, safe='/:?=&%#')}")
            return redirect('reject_request_donation', donation_id=donation_id)

        

        # Reject the donation

        donation.status = 'declined'

        donation.decline_message = comment

        donation.save()

        

        # Notify donor

        foodbank_request = donation.foodbank_request

        Notification.objects.create(

            user=donation.donor,

            notification_type='donation_declined',

            message=f'Your donation for "{foodbank_request.title}" was declined by {request.user.foodbank_profile.foodbank_name}. Reason: {comment}'

        )

        

        messages.success(request, 'Donation rejected successfully.')

        return _redirect_back_or_default(request, 'foodbank_requests')

    

    next_url = (request.GET.get('next') or request.POST.get('next') or '').strip()
    context = {

        'donation': donation,

        'action': 'reject',
        'next': next_url,

    }

    return render(request, 'authentication/confirm_donation_action.html', context)








