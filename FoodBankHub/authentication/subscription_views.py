from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from urllib.parse import urlencode
from datetime import timedelta
from django.db.models import Q

from .models import FoodBankProfile, FoodBankSubscription, SubscriptionPayment
from .forms import SubscriptionPaymentForm


def _filtered_subscription_payments(request, foodbank_profile):
    base_payment_qs = SubscriptionPayment.objects.filter(
        foodbank=foodbank_profile
    ).select_related('verified_by').order_by('-submitted_at')

    status_filter = (request.GET.get('status') or '').strip()
    plan_filter = (request.GET.get('plan_type') or '').strip()
    method_filter = (request.GET.get('payment_method') or '').strip()
    q_filter = (request.GET.get('q') or '').strip()
    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()

    date_from = parse_date(date_from_raw) if date_from_raw else None
    date_to = parse_date(date_to_raw) if date_to_raw else None

    payment_history_filtered = base_payment_qs

    if status_filter:
        payment_history_filtered = payment_history_filtered.filter(status=status_filter)
    if plan_filter:
        payment_history_filtered = payment_history_filtered.filter(plan_type=plan_filter)
    if method_filter:
        payment_history_filtered = payment_history_filtered.filter(payment_method=method_filter)
    if date_from:
        payment_history_filtered = payment_history_filtered.filter(payment_date__gte=date_from)
    if date_to:
        payment_history_filtered = payment_history_filtered.filter(payment_date__lte=date_to)
    if q_filter:
        payment_history_filtered = payment_history_filtered.filter(
            Q(transaction_reference__icontains=q_filter) |
            Q(notes__icontains=q_filter)
        )

    current_filters = {
        'status': status_filter,
        'plan_type': plan_filter,
        'payment_method': method_filter,
        'q': q_filter,
        'date_from': date_from_raw,
        'date_to': date_to_raw,
    }

    return base_payment_qs, payment_history_filtered, current_filters


@login_required
def subscription_status(request):
    """Display subscription status and options for food banks"""
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'This page is only accessible to food banks.')
        return redirect('dashboard')
    
    try:
        foodbank_profile = request.user.foodbank_profile
        subscription = foodbank_profile.subscription
        
        # Fix missing trial_end_date for existing subscriptions
        if subscription.status == 'trial' and not subscription.trial_end_date and subscription.trial_start_date:
            subscription.trial_end_date = subscription.trial_start_date + timedelta(days=90)
            subscription.save()
            
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank profile not found.')
        return redirect('dashboard')
    except FoodBankSubscription.DoesNotExist:
        # Create subscription if it doesn't exist
        subscription = FoodBankSubscription.objects.create(
            foodbank=foodbank_profile,
            status='trial',
            plan='trial'
        )
    
    base_payment_qs, payment_history_filtered, current_filters = _filtered_subscription_payments(request, foodbank_profile)
    
    # Get pending payments
    pending_payments = base_payment_qs.filter(status='pending')
    pending_payments_count = pending_payments.count()

    status_filter = current_filters.get('status') or ''
    plan_filter = current_filters.get('plan_type') or ''
    method_filter = current_filters.get('payment_method') or ''
    q_filter = current_filters.get('q') or ''
    date_from_raw = current_filters.get('date_from') or ''
    date_to_raw = current_filters.get('date_to') or ''
    filter_query = urlencode({k: v for k, v in current_filters.items() if v})

    paginator = Paginator(payment_history_filtered, 10)
    payment_page_obj = paginator.get_page(request.GET.get('page'))
    
    context = {
        'subscription': subscription,
        'payment_history': base_payment_qs,
        'payment_history_filtered': payment_history_filtered,
        'payment_page_obj': payment_page_obj,
        'pending_payments': pending_payments,
        'pending_payments_count': pending_payments_count,
        'can_access': subscription.can_access_features(),
        'days_remaining': subscription.days_remaining(),
        'is_trial': subscription.status == 'trial',
        'is_expired': subscription.status == 'expired',
        'status_choices': SubscriptionPayment.STATUS_CHOICES,
        'plan_choices': SubscriptionPayment.PLAN_CHOICES,
        'payment_method_choices': SubscriptionPayment.PAYMENT_METHOD_CHOICES,
        'filter_query': filter_query,
        'filters': {
            'status': status_filter,
            'plan_type': plan_filter,
            'payment_method': method_filter,
            'q': q_filter,
            'date_from': date_from_raw,
            'date_to': date_to_raw,
        },
    }
    
    return render(request, 'authentication/foodbank_subscription.html', context)


@login_required
def subscription_requests(request):
    """Dedicated view for subscription payment requests/history"""
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'This page is only accessible to food banks.')
        return redirect('dashboard')

    try:
        foodbank_profile = request.user.foodbank_profile
        subscription = foodbank_profile.subscription
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank profile not found.')
        return redirect('dashboard')
    except FoodBankSubscription.DoesNotExist:
        subscription = FoodBankSubscription.objects.create(
            foodbank=foodbank_profile,
            status='trial',
            plan='trial'
        )

    base_payment_qs, payment_history_filtered, current_filters = _filtered_subscription_payments(request, foodbank_profile)
    filter_query = urlencode({k: v for k, v in current_filters.items() if v})

    paginator = Paginator(payment_history_filtered, 10)
    payment_page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'subscription': subscription,
        'payment_page_obj': payment_page_obj,
        'status_choices': SubscriptionPayment.STATUS_CHOICES,
        'plan_choices': SubscriptionPayment.PLAN_CHOICES,
        'payment_method_choices': SubscriptionPayment.PAYMENT_METHOD_CHOICES,
        'filter_query': filter_query,
        'filters': current_filters,
        'total_payments': base_payment_qs.count(),
        'pending_payments_count': base_payment_qs.filter(status='pending').count(),
    }

    return render(request, 'authentication/subscription_requests.html', context)


@login_required
def subscribe(request):
    """Subscription payment submission page"""
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'This page is only accessible to food banks.')
        return redirect('dashboard')
    
    try:
        foodbank_profile = request.user.foodbank_profile
        subscription = foodbank_profile.subscription
    except:
        messages.error(request, 'Subscription information not found.')
        return redirect('subscription_status')
    
    if request.method == 'POST':
        form = SubscriptionPaymentForm(request.POST, request.FILES)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.subscription = subscription
            payment.foodbank = foodbank_profile
            payment.status = 'pending'
            payment.save()
            
            messages.success(
                request,
                'Payment evidence submitted successfully! We will verify your payment within 24 hours.'
            )
            return redirect('subscription_status')
    else:
        form = SubscriptionPaymentForm()
    
    context = {
        'form': form,
        'subscription': subscription,
        'monthly_price': 2000,
        'yearly_price': 10000,
    }
    
    return render(request, 'authentication/subscribe.html', context)


@login_required
def payment_detail(request, payment_id):
    """View details of a specific payment submission"""
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    payment = get_object_or_404(
        SubscriptionPayment,
        id=payment_id,
        foodbank=request.user.foodbank_profile
    )
    
    context = {
        'payment': payment,
    }
    
    return render(request, 'authentication/payment_detail.html', context)


@login_required
def subscription_info(request):
    """Information page about subscription plans"""
    context = {
        'monthly_price': 2000,
        'yearly_price': 10000,
        'trial_days': 90,
    }
    
    return render(request, 'authentication/subscription_info.html', context)


@login_required
def export_subscription_payments_excel(request):
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'This page is only accessible to food banks.')
        return redirect('dashboard')

    try:
        foodbank_profile = request.user.foodbank_profile
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank profile not found.')
        return redirect('dashboard')

    _, payment_history_filtered, _ = _filtered_subscription_payments(request, foodbank_profile)
    payments_data = list(payment_history_filtered)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Subscription Payments'

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"{foodbank_profile.foodbank_name} - Subscription Payments"
    title_cell.font = Font(bold=True, size=14, color='10b981')
    title_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')} | Total: {len(payments_data)}"
    info_cell.alignment = Alignment(horizontal='center')

    headers = [
        'S/No',
        'Payment Date',
        'Plan',
        'Amount',
        'Payment Method',
        'Transaction Reference',
        'Status',
        'Submitted At',
        'Verified By',
        'Verified At',
        'Message',
        'Admin Rejection Notes',
        'Evidence URL',
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for idx, payment in enumerate(payments_data, 1):
        row_num = idx + 4

        evidence_url = ''
        if getattr(payment, 'payment_evidence', None) and getattr(payment.payment_evidence, 'url', None):
            try:
                evidence_url = request.build_absolute_uri(payment.payment_evidence.url)
            except Exception:
                evidence_url = payment.payment_evidence.url

        verified_by = payment.verified_by.email if getattr(payment, 'verified_by', None) else ''
        verified_at = payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else ''

        values = [
            idx,
            payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
            payment.get_plan_type_display() if hasattr(payment, 'get_plan_type_display') else payment.plan_type,
            float(payment.amount) if payment.amount is not None else 0,
            payment.get_payment_method_display() if hasattr(payment, 'get_payment_method_display') else payment.payment_method,
            payment.transaction_reference,
            payment.get_status_display() if hasattr(payment, 'get_status_display') else payment.status,
            payment.submitted_at.strftime('%Y-%m-%d %H:%M') if payment.submitted_at else '',
            verified_by,
            verified_at,
            payment.notes or '',
            payment.rejection_reason or '',
            evidence_url,
        ]

        for col_num, value in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col_num, value=value)
            c.border = border
            if col_num in [11, 12, 13]:
                c.alignment = Alignment(wrap_text=True, vertical='top')

    column_widths = [6, 14, 18, 12, 18, 22, 16, 18, 22, 18, 30, 30, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"subscription_payments_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_subscription_payments_pdf(request):
    if request.user.user_type != 'FOODBANK':
        messages.error(request, 'This page is only accessible to food banks.')
        return redirect('dashboard')

    try:
        foodbank_profile = request.user.foodbank_profile
    except FoodBankProfile.DoesNotExist:
        messages.error(request, 'Food bank profile not found.')
        return redirect('dashboard')

    _, payment_history_filtered, _ = _filtered_subscription_payments(request, foodbank_profile)
    payments_data = list(payment_history_filtered)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#10b981'),
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=6,
        leading=8,
        alignment=TA_LEFT,
        wordWrap='CJK',
    )

    elements.append(Paragraph(f"{foodbank_profile.foodbank_name} - Subscription Payments", title_style))
    elements.append(Paragraph(f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(payments_data)}", styles['Normal']))
    elements.append(Spacer(1, 0.15 * inch))

    data = [[
        'S/No',
        'Payment Date',
        'Plan',
        'Amount',
        'Method',
        'Transaction Ref',
        'Status',
        'Submitted',
        'Verified By',
        'Verified At',
        'Message',
        'Admin Rejection\nNotes',
        'Evidence URL',
    ]]

    for idx, payment in enumerate(payments_data, start=1):
        evidence_url = ''
        if getattr(payment, 'payment_evidence', None) and getattr(payment.payment_evidence, 'url', None):
            try:
                evidence_url = request.build_absolute_uri(payment.payment_evidence.url)
            except Exception:
                evidence_url = payment.payment_evidence.url

        verified_by = payment.verified_by.email if getattr(payment, 'verified_by', None) else ''
        verified_at = payment.verified_at.strftime('%Y-%m-%d %H:%M') if payment.verified_at else ''
        submitted_at = payment.submitted_at.strftime('%Y-%m-%d %H:%M') if payment.submitted_at else ''
        payment_date = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else ''
        plan_label = payment.get_plan_type_display() if hasattr(payment, 'get_plan_type_display') else payment.plan_type
        method_label = payment.get_payment_method_display() if hasattr(payment, 'get_payment_method_display') else payment.payment_method
        status_label = payment.get_status_display() if hasattr(payment, 'get_status_display') else payment.status

        data.append([
            str(idx),
            Paragraph(payment_date, cell_style),
            Paragraph(plan_label, cell_style),
            Paragraph(f"{payment.amount}", cell_style),
            Paragraph(method_label, cell_style),
            Paragraph(payment.transaction_reference or '', cell_style),
            Paragraph(status_label, cell_style),
            Paragraph(submitted_at, cell_style),
            Paragraph(verified_by, cell_style),
            Paragraph(verified_at, cell_style),
            Paragraph((payment.notes or '')[:250], cell_style),
            Paragraph((payment.rejection_reason or '')[:250], cell_style),
            Paragraph(evidence_url[:250], cell_style),
        ])

    col_widths = [
        0.35 * inch,  # S/No
        0.65 * inch,  # Payment Date
        0.75 * inch,  # Plan
        0.6 * inch,   # Amount
        0.7 * inch,   # Method
        0.95 * inch,  # Transaction Ref
        0.6 * inch,   # Status
        0.8 * inch,   # Submitted
        0.8 * inch,   # Verified By
        0.75 * inch,  # Verified At
        1.0 * inch,   # Message
        1.0 * inch,   # Admin Rejection Notes
        1.25 * inch,  # Evidence URL
    ]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="subscription_payments_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response
